import time
import traceback
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from bit.bit_api import closeBrowser
from bit.bit_clash import get_public_ip, switch_random_hongkong_node
from bit.bit_mysql import insert_task_record
from bit.bit_summary_delayfile import summary_delayFile
from bit.bit_utils import get_now_time
from playwright.bit_email_info import get_mail_info, read_email_info_all
from playwright.common import BitPlaywrightSession, select_country


def download_relay_mail(window_id, site):
    with BitPlaywrightSession(window_id) as session:
        page = session.page
        clicked = False
        for _ in range(3):
            try:
                clicked = click_download(page, site)
                if clicked:
                    break
            except Exception as exc:
                print("点击下载失败", exc)
                traceback.print_exc()
                switch_random_hongkong_node()
                get_public_ip()

        if not clicked:
            return "没有需要下载的文件"

        for _ in range(3):
            time.sleep(60)
            try:
                mail_item = scan_email(page, 1)
                if mail_item == "读取邮件失败":
                    break
                if mail_item and download_excel(page, mail_item):
                    return "下载文件成功"
            except Exception as exc:
                print("下载延误邮件失败", exc)
                traceback.print_exc()
        return "下载文件失败"


def click_download(page, site):
    page.goto("https://global-selling.mercadolibre.com/reputation", wait_until="domcontentloaded", timeout=60000)
    page.reload(wait_until="domcontentloaded", timeout=60000)
    time.sleep(8)
    select_country(page, site)
    time.sleep(3)
    link = page.locator(
        "xpath=//a[contains(text(), 'Download affected orders') and "
        "not(../descendant::*[contains(text(), 'Review in Metrics') or contains(text(), 'Review')])]"
    )
    try:
        link.first.click(timeout=10000)
        print("点击下载成功")
        return True
    except Exception:
        return False


def scan_email(page, isAll):
    email_infos = read_email_info_all(page) if isAll == 1 else get_mail_info(page, "普通邮件")
    if email_infos is None:
        print("读取邮件失败")
        return "读取邮件失败"
    email_infos_sorted = sorted(list(email_infos), key=lambda x: x[1], reverse=True)
    for subject, mail_time, element, text in email_infos_sorted:
        if subject != "Your orders that you shipped with delay report is ready":
            continue
        diff = datetime.now() - mail_time
        print("相差时间为", diff)
        if diff.total_seconds() <= 3600.0:
            return (subject, mail_time, element, text)
    return None


def download_excel(page, mail_item):
    subject, mail_time, element, text = mail_item
    if text == "垃圾邮件":
        element.click(timeout=10000)
    else:
        page.get_by_title("收件箱").click(timeout=30000)
        mail_item_2 = scan_email(page, 0)
        mail_item_2[2].click(timeout=10000)

    print("点击时间为的邮件:", str(mail_time))
    page.wait_for_timeout(10000)
    url = page.get_by_text("Go to download report").first.get_attribute("href", timeout=30000)
    print(url)
    page.goto(url, wait_until="domcontentloaded", timeout=60000)
    page.get_by_text("Download", exact=True).click(timeout=30000)
    print("已下载延误文件")
    return True


def download_relay_mail_all():
    root_path = Path(__file__).resolve().parent.parent / "bit"
    file_path = root_path / "比特配置文件.xlsx"
    start = int(time.time())
    wb = load_workbook(file_path)
    sheet = wb.active
    result = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        browser_id, name, remark, sites = row[:4]
        if "忽略" in str(remark or "").strip() or not browser_id or not sites:
            continue
        for site in str(sites).split("，"):
            site = site.strip()
            if not site:
                continue
            try:
                message = download_relay_mail(browser_id, site)
                print(get_now_time() + name + site + message)
                result.append(("下载延误表格", name, site, "成功", get_now_time()))
            except Exception as exc:
                print(get_now_time() + name + site + "执行失败", exc)
                result.append(("下载延误表格", name, site, "失败", get_now_time()))
        try:
            closeBrowser(str(browser_id))
        except Exception:
            pass
    print("总花费", int(time.time()) - start)
    insert_task_record(result)
    summary_delayFile()


if __name__ == "__main__":
    download_relay_mail_all()
