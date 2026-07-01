import time
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from bit.bit_api import closeBrowser
from bit.bit_send_mail import send_info
from bit.bit_utils import get_now_time
from playwright.common import BitPlaywrightSession, first_text, select_country


def get_reputation_info(window_id, site):
    with BitPlaywrightSession(window_id) as session:
        page = session.page
        page.goto("https://global-selling.mercadolibre.com/sales-summary", wait_until="domcontentloaded", timeout=60000)
        page.reload(wait_until="domcontentloaded", timeout=60000)
        time.sleep(8)
        select_country(page, site)

        data_delay = ""
        data_complain = ""
        data_cancel = ""
        try:
            data_delay = page.locator(
                "xpath=//p[text()='Non-compliant shipments']/ancestor::div[contains(@class,'metric')]"
                "//div[contains(@class, 'metric__title')]"
            ).inner_text(timeout=10000)
            data_complain = page.locator(
                "xpath=//p[text()='Complaints']/ancestor::div[contains(@class,'metric')]"
                "//div[contains(@class, 'metric__title')]"
            ).inner_text(timeout=10000)
            data_cancel = page.locator(
                "xpath=//p[text()='Canceled by you']/ancestor::div[contains(@class,'metric')]"
                "//div[contains(@class, 'metric__title')]"
            ).inner_text(timeout=10000)
        except Exception as exc:
            print(get_now_time() + str(site) + "获取声誉数据失败", exc)

        data_color = first_text(page, ".panel-segment__focus-item-title-container")
        data_orders = first_text(page, ".value__sales")

        color_map = {
            "Green": "绿色",
            "Yellow": "黄色",
            "Orange": "橙色",
            "Red": "红色",
            "You still have no color": "无色",
        }
        for marker, value in color_map.items():
            if marker in data_color:
                data_color = value
                break

        return [data_color, data_orders, data_complain, data_delay]


def get_reputation_info_all():
    start = int(time.time())
    root_path = Path(__file__).resolve().parent.parent / "bit"
    file_path = root_path / "比特配置文件.xlsx"
    wb = load_workbook(file_path)
    sheet = wb.active
    reputation_info_sum = []
    result = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        browser_id, name, remark, sites = row[:4]
        if remark == "忽略" or not browser_id or not sites:
            continue
        for site in str(sites).split("，"):
            site = site.strip()
            if not site:
                continue
            try:
                reputation_info = get_reputation_info(browser_id, site)
                reputation_info.extend([name, site])
                reputation_info_sum.append(reputation_info)
                result.append(name + site + "获取声誉信息执行成功")
            except Exception as exc:
                print(get_now_time() + name + site + "执行失败", exc)
                result.append(name + site + "获取声誉信息执行失败")
                time.sleep(180)
            time.sleep(10)
        try:
            closeBrowser(browser_id)
        except Exception:
            pass

    body = "\n".join(map(str, reputation_info_sum))
    print(body)
    print(get_now_time() + "总花费", int(time.time()) - start)

    df = pd.DataFrame(
        reputation_info_sum,
        columns=["声誉颜色", "总单量", "投诉率", "延误率", "店铺名", "站点"],
    )
    date_str = datetime.now().strftime("%Y-%m-%d-%H")
    output_path = root_path / f"美客多-武汉泽顺店铺声誉信息汇总-{date_str}.xlsx"
    df.to_excel(output_path, index=False)
    send_info("美客多所有店铺声誉汇总", body, output_path, output_path.name)


if __name__ == "__main__":
    get_reputation_info_all()
