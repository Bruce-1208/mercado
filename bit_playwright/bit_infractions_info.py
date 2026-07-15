import importlib
import re
import sys
import time
from concurrent.futures import ProcessPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from bit.bit_api import closeBrowser, openBrowser
from bit.bit_runtime_lock import create_window_lease
from bit.bit_clash import get_public_ip, switch_random_hongkong_node
from bit.bit_db_api import insert_task_record, inset_infraction_info
from bit.bit_send_mail import send_info
from bit.bit_utils import get_now_time


INFRACTIONS_URL = "https://global-selling.mercadolibre.com/noindex/pppi/infractions?tab=detections&offset=0"
INFRACTIONS_TAB_URLS = {
    "侵权": "https://global-selling.mercadolibre.com/noindex/pppi/infractions?tab=detections&offset=0",
    "权利人": "https://global-selling.mercadolibre.com/noindex/pppi/infractions?tab=denounces&offset=0",
}

SITE_PREFIX_MAP = {
    "墨西哥": "MLM",
    "MX": "MLM",
    "MLM": "MLM",
    "巴西": "MLB",
    "BR": "MLB",
    "MLB": "MLB",
    "哥伦比亚": "MCO",
    "CO": "MCO",
    "MCO": "MCO",
    "智利": "MLC",
    "CL": "MLC",
    "MLC": "MLC",
    "阿根廷": "MLA",
    "AR": "MLA",
    "MLA": "MLA",
    "乌拉圭": "MLU",
    "UY": "MLU",
    "MLU": "MLU",
}

SITE_SWITCH_SELECTOR_MAP = {
    "墨西哥": 'div[data-value="MLM-remote"]',
    "MX": 'div[data-value="MLM-remote"]',
    "MLM": 'div[data-value="MLM-remote"]',
    "巴西": 'div[data-value="MLB-remote"]',
    "BR": 'div[data-value="MLB-remote"]',
    "MLB": 'div[data-value="MLB-remote"]',
    "哥伦比亚": 'div[data-value="MCO-remote"]',
    "CO": 'div[data-value="MCO-remote"]',
    "MCO": 'div[data-value="MCO-remote"]',
    "智利": 'div[data-value="MLC-remote"]',
    "CL": 'div[data-value="MLC-remote"]',
    "MLC": 'div[data-value="MLC-remote"]',
    "阿根廷": 'div[data-value="MLA-remote"]',
    "AR": 'div[data-value="MLA-remote"]',
    "MLA": 'div[data-value="MLA-remote"]',
    "乌拉圭": 'div[data-value="MLU-remote"]',
    "UY": 'div[data-value="MLU-remote"]',
    "MLU": 'div[data-value="MLU-remote"]',
}

_PLAYWRIGHT_API = None


class MercadoRateLimitError(RuntimeError):
    """Mercado Libre 或浏览器接口返回访问限频。"""


class MercadoAuthenticationError(RuntimeError):
    """Mercado Libre 登录态已经失效。"""


class MercadoPageStructureError(RuntimeError):
    """Mercado Libre 页面已打开，但预期的侵权页结构不存在。"""


class BitBrowserWindowError(RuntimeError):
    """比特浏览器窗口配置无效或无法打开。"""


RATE_LIMIT_MARKERS = (
    "429",
    "too many requests",
    "rate limit",
    "request limit",
    "请求太过频繁",
    "请求过于频繁",
    "每秒最多可以发起",
    "访问过于频繁",
    "demasiadas solicitudes",
    "muitas solicitações",
    "hubo un error accediendo a esta página",
    "hubo un error accediendo a esta pagina",
)
SPANISH_IP_SWITCH_MARKERS = (
    "hubo un error accediendo a esta página",
    "hubo un error accediendo a esta pagina",
)
LOGIN_URL_MARKERS = ("/login/", "/lgz/", "/legacy-user")
LOGIN_TEXT_MARKERS = (
    "fill out your e-mail address to log in",
    "fill out your email address to log in",
    "iniciar sesión",
    "iniciar sesion",
    "iniciar sessão",
    "iniciar sessao",
    "登录您的账户",
    "登录你的账户",
    "登录账号",
    "请登录",
    "填写您的电子邮件地址以登录",
)
RIGHTS_HOLDER_LABELS = (
    "Reported by rights holders",
    "Reported by rights holder",
    "权利人举报",
    "由权利人举报",
    "权利持有人举报",
    "权利人报告",
)
DETECTED_LABELS = (
    "Detected by Mercado Libre",
    "Detected by MercadoLibre",
    "Detected",
    "Mercado Libre 检测到",
    "由 Mercado Libre 检测到",
    "美客多检测到",
    "平台检测",
    "系统检测",
)


def _is_rate_limited_text(value):
    text = str(value or "").lower()
    return any(marker in text for marker in RATE_LIMIT_MARKERS)


def _is_spanish_ip_switch_text(value):
    text = str(value or "").lower()
    return any(marker in text for marker in SPANISH_IP_SWITCH_MARKERS)


def _get_page_text(page):
    try:
        return page.locator("body").inner_text(timeout=3000)
    except Exception:
        return ""


def _is_login_page(current_url, body_text):
    url = str(current_url or "").casefold()
    text = str(body_text or "").casefold()
    return any(marker in url for marker in LOGIN_URL_MARKERS) or any(
        marker in text for marker in LOGIN_TEXT_MARKERS
    )


def _raise_if_page_unavailable(page, context="侵权页面"):
    """识别限频和登录跳转，避免把登录页误判为“0 条侵权”。"""
    body_text = _get_page_text(page)
    current_url = page.url or ""
    if _is_rate_limited_text(f"{current_url}\n{body_text}"):
        raise MercadoRateLimitError(f"Mercado 页面限频：{current_url} {body_text[:300]}")
    if _is_login_page(current_url, body_text):
        raise MercadoAuthenticationError(f"{context}登录态失效，已跳转登录页：{current_url}")
    return body_text


def _raise_if_page_rate_limited(page):
    """兼容旧调用；现在同时校验登录状态。"""
    return _raise_if_page_unavailable(page)


def _validate_infractions_page(page):
    body_text = _raise_if_page_unavailable(page)
    current_url = page.url or ""
    if "/noindex/pppi/infractions" not in current_url:
        raise MercadoPageStructureError(f"页面未停留在侵权中心：{current_url}")

    has_structure = page.locator(
        ".infractions__tab, .infraction-item__id, .infraction-item"
    ).count() > 0
    semantic_markers = (
        "intellectual property",
        "infringement",
        "propiedad intelectual",
        "propriedade intelectual",
        "知识产权",
        "侵权",
        "违规",
    )
    if not has_structure and not any(
        marker in body_text.casefold() for marker in semantic_markers
    ):
        raise MercadoPageStructureError(
            f"侵权页面结构不匹配：{current_url} {body_text[:200]}"
        )
    return True


def _load_playwright_sync_api():
    """
    This project has local Playwright helper modules. Temporarily hide the project
    root so Python imports the official ``playwright.sync_api`` package.
    """
    global _PLAYWRIGHT_API
    if _PLAYWRIGHT_API is not None:
        return _PLAYWRIGHT_API

    project_root = Path(__file__).resolve().parent.parent
    original_path = list(sys.path)
    local_parent = sys.modules.get("bit_playwright")
    current_module = sys.modules.get(__name__)
    removed_modules = {}

    for name, module in list(sys.modules.items()):
        if name == "bit_playwright" or (name.startswith("bit_playwright.") and name != __name__):
            removed_modules[name] = module
            sys.modules.pop(name, None)

    sys.path = [
        path
        for path in sys.path
        if Path(path or ".").resolve() != project_root.resolve()
    ]

    try:
        sync_api = importlib.import_module("playwright.sync_api")
        _PLAYWRIGHT_API = (sync_api.sync_playwright, sync_api.TimeoutError)
        return _PLAYWRIGHT_API
    finally:
        sys.path = original_path
        if local_parent is not None:
            sys.modules["bit_playwright"] = local_parent
        if current_module is not None:
            sys.modules[__name__] = current_module


def _site_key(site):
    return str(site or "").strip()


def _site_selector(site):
    key = _site_key(site)
    return SITE_SWITCH_SELECTOR_MAP.get(key) or SITE_SWITCH_SELECTOR_MAP.get(key.upper())


def _site_prefix(site):
    key = _site_key(site)
    return SITE_PREFIX_MAP.get(key) or SITE_PREFIX_MAP.get(key.upper(), "")


def _is_ignored_config_value(value):
    return "忽略" in str(value or "").strip()


def _split_sites(value):
    return [site.strip() for site in re.split(r"[，,、;；\n]+", str(value or "")) if site.strip()]


def _failure_status(exc):
    text = re.sub(r"\s+", " ", str(exc or "")).strip()
    lower = text.casefold()
    if isinstance(exc, MercadoAuthenticationError) or _is_login_page("", text):
        reason = "登录失效"
    elif isinstance(exc, MercadoRateLimitError) or _is_rate_limited_text(text):
        reason = "访问限频"
    elif isinstance(exc, BitBrowserWindowError) and (
        "没有找到相应数据" in text or "missing" in lower or "不存在" in text
    ):
        reason = "窗口ID不存在"
    elif isinstance(exc, BitBrowserWindowError) and "timeout" in lower:
        reason = "窗口打开超时"
    elif "窗口正在被其他任务占用" in text:
        reason = "窗口被其他任务占用"
    elif "站点切换失败" in text:
        reason = "站点切换失败"
    elif isinstance(exc, MercadoPageStructureError):
        reason = "页面结构不匹配"
    elif "timeout" in lower:
        reason = "页面元素等待超时"
    else:
        reason = text or (exc.__class__.__name__ if exc is not None else "未知异常")
    return f"失败：{reason}"[:180]


def _deduplicate_config_rows(rows):
    unique_rows = []
    seen = set()
    for row in rows:
        key = (str(row[0]).strip(), str(row[1]).strip(), str(row[3] or "").strip())
        if key in seen:
            print(f"{get_now_time()}跳过重复配置：{row[1]} {row[3]}")
            continue
        seen.add(key)
        unique_rows.append(row)
    return unique_rows


def _open_bitbrowser(window_id, max_retries=3, retry_delay=10):
    last_result = None
    for attempt in range(1, max_retries + 1):
        result = openBrowser(window_id)
        last_result = result
        print(result)
        data = result.get("data") if isinstance(result, dict) else None
        if data and (data.get("ws") or data.get("http")):
            return result

        message = result.get("msg", "") if isinstance(result, dict) else str(result)
        if "没有找到相应数据" in message or "不存在" in message:
            raise BitBrowserWindowError(f"比特浏览器窗口无效或不存在：{result}")
        if attempt < max_retries:
            print(
                f"{get_now_time()}比特浏览器打开窗口失败，第{attempt}/{max_retries}次："
                f"{message}；{retry_delay}秒后重试"
            )
            time.sleep(retry_delay)

    if _is_rate_limited_text(last_result):
        raise MercadoRateLimitError(
            f"比特浏览器打开窗口被限频，已重试{max_retries}次：{last_result}"
        )
    raise BitBrowserWindowError(
        f"比特浏览器打开窗口失败，已重试{max_retries}次：{last_result}"
    )


def _text_list(page, selector, timeout=30000):
    try:
        page.wait_for_selector(selector, timeout=timeout)
    except Exception:
        pass
    values = page.locator(selector).all_text_contents()
    return [value.strip() for value in values if value and value.strip()]


def _get_page_signature(page):
    try:
        ids = _text_list(page, ".infraction-item__id", timeout=5000)
        return tuple(ids)
    except Exception:
        return tuple()


def _wait_infractions_ready(page, timeout=30000):
    try:
        page.wait_for_load_state("domcontentloaded", timeout=timeout)
    except Exception:
        pass
    try:
        page.wait_for_function(
            """
            () => location.href.includes('/noindex/pppi/infractions') &&
              (document.readyState === 'complete' || document.readyState === 'interactive')
            """,
            timeout=timeout,
        )
    except Exception:
        pass

 
def _safe_goto_infractions(page, url, timeout=60000):
    try:
        response = page.goto(url, wait_until="domcontentloaded", timeout=timeout)
        if response is not None and response.status == 429:
            raise MercadoRateLimitError(f"Mercado 页面返回 HTTP 429：{url}")
        if response is not None and response.status in (401, 403):
            raise MercadoAuthenticationError(
                f"Mercado 页面返回 HTTP {response.status}：{url}"
            )
    except Exception as exc:
        _raise_if_page_unavailable(page)
        if isinstance(exc, (MercadoAuthenticationError, MercadoRateLimitError)):
            raise
        if _is_rate_limited_text(exc):
            raise MercadoRateLimitError(str(exc)) from exc
        current_url = page.url or ""
        if "/noindex/pppi/infractions" not in current_url:
            raise
        print(f"页面自动跳转，继续使用当前侵权页: {current_url}")
    _raise_if_page_unavailable(page)
    _wait_infractions_ready(page)
    _validate_infractions_page(page)
    return page.url


def _current_infraction_type(page):
    current_url = page.url or ""
    match = re.search(r"[?&]tab=([^&]+)", current_url)
    current_tab = match.group(1) if match else ""
    if current_tab == "denounces":
        return "权利人"
    if current_tab == "detections":
        return "侵权"
    return "侵权"


def _goto_infractions_type(page, infraction_type):
    target_url = INFRACTIONS_TAB_URLS[infraction_type]
    _safe_goto_infractions(page, target_url)
    _reset_current_offset(page)
    actual_type = _current_infraction_type(page)
    if actual_type != infraction_type:
        print(f"请求打开{infraction_type}标签，但页面实际停留在{actual_type}标签: {page.url}")
    return actual_type


def _extract_last_submit_times(page):
    return page.evaluate(
        """
        () => {
          const cards = [...document.querySelectorAll('.infraction-item__id')]
            .map((idNode) =>
              idNode.closest('.infraction-item') ||
              idNode.closest('[class*="infraction-item"]') ||
              idNode.closest('li') ||
              idNode.parentElement
            )
            .filter(Boolean);

          const labelPattern = /(last\\s+submitted|last\\s+submission|submitted|submission|最后提交|提交时间|提交日期|已提交)/i;
          const datePattern = /(?:\\d{4}年\\d{1,2}月\\d{1,2}日(?:\\s+\\d{1,2}:\\d{2}(?::\\d{2})?)?|\\d{4}[-/]\\d{1,2}[-/]\\d{1,2}(?:\\s+\\d{1,2}:\\d{2}(?::\\d{2})?)?|\\d{1,2}[-/]\\d{1,2}[-/]\\d{4}(?:\\s+\\d{1,2}:\\d{2}(?::\\d{2})?)?|[A-Z][a-z]{2,8}\\.?\\s+\\d{1,2},?\\s*\\d{4}(?:\\s+\\d{1,2}:\\d{2}(?:\\s*[AP]M)?)?|\\d{1,2}\\s+[A-Z][a-z]{2,8}\\.?\\s+\\d{4}(?:\\s+\\d{1,2}:\\d{2})?|\\d{1,2}\\s+de\\s+[a-záéíóúñç]+\\s+de\\s+\\d{4})/i;

          function clean(text) {
            return (text || '').replace(/\\s+/g, ' ').trim();
          }

          function normalizeCandidate(text) {
            const value = clean(text);
            if (!value) return '';
            const match = value.match(datePattern);
            if (match) return match[0].trim();
            return value.replace(labelPattern, '').replace(/^[:：\\s-]+/, '').trim();
          }

          function fromLines(text) {
            const lines = (text || '')
              .split(/\\n+/)
              .map((line) => clean(line))
              .filter(Boolean);
            for (let index = 0; index < lines.length; index += 1) {
              const line = lines[index];
              if (!labelPattern.test(line)) continue;

              const sameLine = normalizeCandidate(line);
              if (sameLine && !labelPattern.test(sameLine)) return sameLine;

              for (let offset = 1; offset <= 4 && index + offset < lines.length; offset += 1) {
                const nextLine = normalizeCandidate(lines[index + offset]);
                if (nextLine && !labelPattern.test(nextLine)) return nextLine;
              }
            }
            return '';
          }

          return cards.map((card) => {
            const fromCardText = fromLines(card.innerText || card.textContent || '');
            if (fromCardText) return fromCardText;

            const candidates = [
              ...card.querySelectorAll(
                '[class*="submit"], [class*="submission"], [class*="submitted"], [class*="date"], li, p, span, div'
              ),
            ]
              .map((el) => clean(el.textContent || ''))
              .filter(Boolean);

            const labeled = candidates.find((text) => labelPattern.test(text));
            if (labeled) {
              return normalizeCandidate(labeled);
            }
            return '';
          });
        }
        """
    )


def _read_current_infractions_page(page, name, site, infraction_type="侵权"):
    ids = _text_list(page, ".infraction-item__id", timeout=30000)
    titles = _text_list(page, ".infraction-item__title", timeout=5000)
    dates = _text_list(page, ".infraction-denounce__date", timeout=5000)
    submit_times = _extract_last_submit_times(page)

    rows = []
    submit_count = sum(1 for value in submit_times if value)
    if ids:
        print(f"{name}{site}{infraction_type}提交时间解析成功 {submit_count}/{len(ids)} 条")
    prefix = _site_prefix(site)
    for index, id_text in enumerate(ids):
        title = titles[index] if index < len(titles) else ""
        date = dates[index] if index < len(dates) else ""
        submit_time = submit_times[index] if index < len(submit_times) else ""
        rows.append(
            [
                name,
                site,
                id_text.replace("#", prefix),
                title,
                date,
                submit_time,
                get_now_time(),
                infraction_type,
            ]
        )
    return rows


def _is_next_button_disabled(element):
    try:
        return element.evaluate(
            """
            el => {
              const cls = String(el.className || '').toLowerCase();
              const parentCls = String(el.closest('li')?.className || '').toLowerCase();
              return !!el.disabled || el.getAttribute('aria-disabled') === 'true' ||
                cls.includes('disabled') || parentCls.includes('disabled');
            }
            """
        )
    except Exception:
        return True


def _find_next_button(page):
    selectors = [
        ".andes-pagination__button--next a",
        ".andes-pagination__button--next button",
        'a:has(span.andes-pagination__arrow-title:has-text("Next"))',
        'button:has(span.andes-pagination__arrow-title:has-text("Next"))',
        'a:has(span.andes-pagination__arrow-title:has-text("下一页"))',
        'button:has(span.andes-pagination__arrow-title:has-text("下一页"))',
        "a[aria-label*='Next']",
        "button[aria-label*='Next']",
        "a[aria-label*='下一']",
        "button[aria-label*='下一']",
        "a[title*='Next']",
        "button[title*='Next']",
        "a[title*='下一']",
        "button[title*='下一']",
        "text=Next",
        "text=下一页",
    ]
    for selector in selectors:
        try:
            locator = page.locator(selector)
            count = locator.count()
            for index in range(count):
                element = locator.nth(index)
                if element.is_visible():
                    return element
        except Exception:
            continue
    return None


def _offset_url(current_url, previous_signature):
    page_size = max(1, len(previous_signature or ()))
    match = re.search(r"([?&]offset=)(\d+)", current_url)
    if match:
        next_offset = int(match.group(2)) + page_size
        return current_url[: match.start(2)] + str(next_offset) + current_url[match.end(2) :]

    separator = "&" if "?" in current_url else "?"
    return f"{current_url}{separator}offset={page_size}"


def _goto_next_offset(page, previous_signature):
    next_url = _offset_url(page.url, previous_signature)
    _safe_goto_infractions(page, next_url)
    page.wait_for_function(
        """
        previous => {
          const ids = [...document.querySelectorAll('.infraction-item__id')]
            .map(el => (el.textContent || '').trim())
            .filter(Boolean);
          return ids.length && ids.join('|') !== previous.join('|');
        }
        """,
        arg=list(previous_signature),
        timeout=30000,
    )
    return True


def _reset_current_offset(page):
    next_url = re.sub(r"([?&]offset=)\d+", r"\g<1>0", page.url)
    if next_url != page.url:
        _safe_goto_infractions(page, next_url)
        time.sleep(2)


def _collect_current_infractions_tab(page, name, site, infraction_type):
    infractions_list = []
    seen_ids = set()
    page_no = 1
    while True:
        _validate_infractions_page(page)
        page_rows = _read_current_infractions_page(page, name, site, infraction_type)
        new_count = 0
        for row in page_rows:
            row_key = (row[2], row[7])
            if row_key in seen_ids:
                continue
            seen_ids.add(row_key)
            infractions_list.append(row)
            new_count += 1

        print(
            f"{get_now_time()}{name}{site}{infraction_type}第{page_no}页抓取{len(page_rows)}条，新增{new_count}条"
        )
        previous_signature = _get_page_signature(page)
        if not previous_signature:
            print(f"当前页面没有{infraction_type}数据，结束当前标签抓取")
            break

        if not _click_next_page(page, previous_signature, page_no):
            break
        page_no += 1

    return infractions_list


def _open_rights_holder_report_tab(page):
    clicked = page.evaluate(
        """
        (targetTexts) => {
          const normalizedTargets = targetTexts.map((text) => text.toLocaleLowerCase());
          const isVisible = (el) => {
            const style = window.getComputedStyle(el);
            const rect = el.getBoundingClientRect();
            return style && style.visibility !== 'hidden' && style.display !== 'none' &&
              rect.width > 0 && rect.height > 0;
          };
          const nodes = [...document.querySelectorAll('a, button, [role="tab"], li, span, div')]
            .filter((el) => {
              const text = (el.textContent || '').toLocaleLowerCase();
              return isVisible(el) && normalizedTargets.some((target) => text.includes(target));
            });
          if (!nodes.length) return false;
          const node = nodes.find((el) => ['A', 'BUTTON'].includes(el.tagName)) || nodes[0];
          const clickable = node.closest('a, button, [role="tab"]') || node;
          clickable.scrollIntoView({block: 'center', inline: 'center'});
          clickable.click();
          return true;
        }
        """,
        list(RIGHTS_HOLDER_LABELS),
    )
    if not clicked:
        return False

    try:
        page.wait_for_load_state("domcontentloaded", timeout=30000)
    except Exception:
        pass
    time.sleep(3)
    _reset_current_offset(page)
    return True


def _open_detected_report_tab(page):
    clicked = page.evaluate(
        """
        (texts) => {
          const normalizedTexts = texts.map((text) => text.toLocaleLowerCase());
          const isVisible = (el) => {
            const style = window.getComputedStyle(el);
            const rect = el.getBoundingClientRect();
            return style && style.visibility !== 'hidden' && style.display !== 'none' &&
              rect.width > 0 && rect.height > 0;
          };
          const nodes = [...document.querySelectorAll('a, button, [role="tab"], li, span, div')]
            .filter((el) => {
              const text = (el.textContent || '').toLocaleLowerCase();
              return isVisible(el) && normalizedTexts.some((target) => text.includes(target));
            });
          if (!nodes.length) return false;
          const node = nodes.find((el) => ['A', 'BUTTON'].includes(el.tagName)) || nodes[0];
          const clickable = node.closest('a, button, [role="tab"]') || node;
          clickable.scrollIntoView({block: 'center', inline: 'center'});
          clickable.click();
          return true;
        }
        """,
        list(DETECTED_LABELS),
    )
    if not clicked:
        return False

    try:
        page.wait_for_load_state("domcontentloaded", timeout=30000)
    except Exception:
        pass
    time.sleep(3)
    _reset_current_offset(page)
    return True


def _collect_type_once(page, name, site, infraction_type, collected_types):
    actual_type = _current_infraction_type(page)
    if actual_type != infraction_type:
        print(f"当前实际标签是{actual_type}，跳过按{infraction_type}采集")
        return []
    if actual_type in collected_types:
        print(f"{name}{site}{actual_type}已采集，跳过重复采集")
        return []
    collected_types.add(actual_type)
    return _collect_current_infractions_tab(page, name, site, actual_type)


def _click_next_page(page, previous_signature, page_no):
    for attempt in range(3):
        previous_url = page.url
        try:
            next_button = _find_next_button(page)
            if next_button is None or _is_next_button_disabled(next_button):
                print("当前已经是最后一页，翻页结束")
                return False

            next_button.scroll_into_view_if_needed(timeout=10000)
            next_button.click(timeout=10000)
            page.wait_for_function(
                """
                ([previous, previousUrl]) => {
                  const ids = [...document.querySelectorAll('.infraction-item__id')]
                    .map(el => (el.textContent || '').trim())
                    .filter(Boolean);
                  return (ids.length && ids.join('|') !== previous.join('|')) ||
                    location.href !== previousUrl;
                }
                """,
                arg=[list(previous_signature), previous_url],
                timeout=30000,
            )
            page.wait_for_function(
                """
                previous => {
                  const ids = [...document.querySelectorAll('.infraction-item__id')]
                    .map(el => (el.textContent || '').trim())
                    .filter(Boolean);
                  return ids.length && ids.join('|') !== previous.join('|');
                }
                """,
                arg=list(previous_signature),
                timeout=30000,
            )
            _validate_infractions_page(page)
            print(f"成功点击下一页，当前第{page_no + 1}页")
            time.sleep(1)
            return True
        except Exception as exc:
            print(f"点击下一页不稳定，尝试 offset 兜底，第{attempt + 1}次: {exc}")
            try:
                if _goto_next_offset(page, previous_signature):
                    print(f"offset兜底翻页成功，当前第{page_no + 1}页")
                    return True
            except Exception as fallback_error:
                print("offset兜底翻页失败", fallback_error)
            time.sleep(2)

    print("翻页多次失败，结束当前站点抓取")
    return False


def _switch_site_if_needed(page, name, site, retries=3):
    selector = _site_selector(site)
    if not selector:
        return

    for attempt in range(1, retries + 1):
        try:
            page.locator(".nav-header-cbt__site-switcher").click(timeout=10000)
            print(f"{name}打开站点选择器")
            page.locator(selector).click(timeout=30000)
            try:
                page.reload(wait_until="domcontentloaded", timeout=60000)
            except Exception:
                current_url = page.url or ""
                if "/noindex/pppi/infractions" not in current_url:
                    _raise_if_page_unavailable(page, context=f"{name}{site}切换站点后的页面")
                    raise
                print(f"{name}{site}切换站点后页面自动跳转，继续使用当前页: {current_url}")
            _raise_if_page_unavailable(page, context=f"{name}{site}切换站点后的页面")
            _wait_infractions_ready(page)
            time.sleep(3)
            _validate_infractions_page(page)
            print(get_now_time() + name + site + "选择站点成功")
            return
        except (MercadoAuthenticationError, MercadoRateLimitError):
            raise
        except Exception as exc:
            print(get_now_time() + name + site + f"重新执行选择站点，第{attempt}次: {exc}")
            time.sleep(3)
    raise MercadoPageStructureError(f"{name}{site} 站点切换失败")


def _connect_bitbrowser_with_playwright(playwright, open_result):
    data = open_result.get("data") or {}
    endpoint = data.get("ws")
    if not endpoint and data.get("http"):
        endpoint = f"http://{data['http']}"
    if not endpoint:
        raise BitBrowserWindowError(f"BitBrowser open result missing ws/http: {open_result}")

    browser = playwright.chromium.connect_over_cdp(endpoint)
    context = browser.contexts[0] if browser.contexts else browser.new_context()
    page = context.new_page()
    return browser, page


def get_infractions_info(window_id, name, site, isSwitch=1):
    sync_playwright, _ = _load_playwright_sync_api()
    res = _open_bitbrowser(window_id)

    with sync_playwright() as playwright:
        _browser, page = _connect_bitbrowser_with_playwright(playwright, res)
        try:
            _safe_goto_infractions(page, INFRACTIONS_URL)
            time.sleep(5)
            _validate_infractions_page(page)
            if isSwitch == 1:
                _switch_site_if_needed(page, name, site)
                _validate_infractions_page(page)

            infractions_list = []
            collected_types = set()

            current_type = _current_infraction_type(page)
            print(get_now_time() + name + site + f"当前实际侵权标签: {current_type}")
            infractions_list.extend(_collect_type_once(page, name, site, current_type, collected_types))

            if "侵权" not in collected_types:
                print(get_now_time() + name + site + "尝试切换到普通侵权报告")
                opened = _open_detected_report_tab(page)
                if not opened or _current_infraction_type(page) != "侵权":
                    _goto_infractions_type(page, "侵权")
                infractions_list.extend(_collect_type_once(page, name, site, "侵权", collected_types))

            if "权利人" not in collected_types:
                opened = _open_rights_holder_report_tab(page)
                if not opened or _current_infraction_type(page) != "权利人":
                    _goto_infractions_type(page, "权利人")
                print(get_now_time() + name + site + "开始抓取权利人侵权报告")
                infractions_list.extend(_collect_type_once(page, name, site, "权利人", collected_types))

            return infractions_list
        finally:
            try:
                page.close()
            except Exception:
                pass


def _run_infractions_for_browser_locked(row):
    browser_id = row[0]
    name = row[1]
    remark = row[2]
    if _is_ignored_config_value(remark):
        return [], []

    print(get_now_time() + "开始打开窗口:" + name)
    if not row[3]:
        return [], [("获取侵权信息", name, "", "失败：未配置站点", get_now_time())]

    site_list = _split_sites(row[3])
    infraction_info_sum = []
    result = []
    fatal_profile_error = None

    for site in site_list:
        if fatal_profile_error is not None:
            result.append(
                (
                    "获取侵权信息",
                    name,
                    site,
                    _failure_status(fatal_profile_error),
                    get_now_time(),
                )
            )
            continue

        succeeded = False
        last_error = None
        for attempt in range(1, 4):
            try:
                infraction_info = get_infractions_info(browser_id, name, site, 1)
                infraction_info_sum.extend(infraction_info)
                print(get_now_time() + name + site + "成功")
                result.append(("获取侵权信息", name, site, "成功", get_now_time()))
                succeeded = True
                break
            except Exception as exc:
                last_error = exc
                print(get_now_time() + name + site + "执行失败", exc)
                if isinstance(exc, (MercadoAuthenticationError, BitBrowserWindowError)):
                    fatal_profile_error = exc
                    break
                if attempt < 3:
                    if _is_spanish_ip_switch_text(exc):
                        print(
                            f"{get_now_time()}{name}{site}检测到指定西语错误页，"
                            f"第{attempt}次切换随机香港节点后重试"
                        )
                        switch_random_hongkong_node()
                        get_public_ip()
                    time.sleep(5)
        if not succeeded:
            result.append(
                ("获取侵权信息", name, site, _failure_status(last_error), get_now_time())
            )

    print(get_now_time() + "结束，正在关闭窗口")
    try:
        closeBrowser(browser_id)
    except Exception as exc:
        print(get_now_time() + name + "关闭窗口失败", exc)
    print(get_now_time() + "已经关闭窗口")
    return infraction_info_sum, result


def _run_infractions_for_browser(row):
    browser_id = row[0]
    name = row[1]
    lease = create_window_lease(
        browser_id,
        owner=f"infraction_collection:{name}",
        shop_name=name,
        task_type="infraction_collection",
    )
    if not lease.acquire(timeout=0):
        print(get_now_time() + name + "窗口已被其他任务占用，跳过本次侵权采集")
        return [], [("获取侵权信息", name, "", "跳过：窗口被其他任务占用", get_now_time())]
    try:
        return _run_infractions_for_browser_locked(row)
    finally:
        lease.release()


def get_infractions_info_all(max_workers=20):
    """使用独立进程并发采集各店铺侵权，主进程负责汇总、导出和入库。"""
    start = int(time.time())
    print(start)
    bit_dir = Path(__file__).resolve().parent.parent / "bit"
    file_path = bit_dir / "比特配置文件.xlsx"

    wb = load_workbook(file_path)
    sheet = wb.active
    infraction_info_sum = []
    result = []
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    rows = [row for row in rows if row and row[0] and not _is_ignored_config_value(row[2])]
    rows = _deduplicate_config_rows(rows)

    worker_count = max(1, min(int(max_workers), len(rows))) if rows else 1
    with ProcessPoolExecutor(max_workers=worker_count) as executor:
        future_map = {
            executor.submit(_run_infractions_for_browser, row): row for row in rows
        }
        for future in as_completed(future_map):
            row = future_map[future]
            name = row[1]
            try:
                browser_infractions, browser_result = future.result()
                infraction_info_sum.extend(browser_infractions)
                result.extend(browser_result)
                print(get_now_time() + name + "窗口任务完成")
            except Exception as exc:
                print(get_now_time() + name + "窗口任务异常", exc)
                result.append(
                    ("获取侵权信息", name, "", _failure_status(exc), get_now_time())
                )

    infraction_info_sum_str = "\n".join(map(str, infraction_info_sum))
    print(infraction_info_sum_str)

    end = int(time.time())
    print(get_now_time() + "总花费", end - start)

    df = pd.DataFrame(
        infraction_info_sum,
        columns=["店铺名", "站点", "编号", "标题", "侵权时间", "提交时间", "执行时间", "类型"],
    )

    date_str = datetime.now().strftime("%Y-%m-%d-%H")
    output_path = bit_dir / f"美客多-武汉泽顺店铺侵权信息汇总-{date_str}.xlsx"
    df.to_excel(output_path, index=False)

    send_info(
        "美客多所有店铺侵权汇总",
        infraction_info_sum_str,
        output_path,
        output_path.name,
    )
    print(get_now_time() + "发送邮件成功")

    insert_task_record(result)
    inset_infraction_info(infraction_info_sum)


if __name__ == "__main__":
    get_infractions_info_all()
