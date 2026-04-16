import time

from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.wait import WebDriverWait

from bit_api import *
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import pyautogui
from switch_country import *
from openpyxl import load_workbook
from send_mail import *
import pandas as pd

from datetime import datetime
from pathlib import Path


def print_orders(window_id, site):
    # /browser/open 接口会返回 selenium使用的http地址，以及webdriver的path，直接使用即可
    # 龙
    # window_id='9812f185f7ab49d98f3988994d9e8ebf'
    res = openBrowser(window_id)  # 窗口ID从窗口配置界面中复制，或者api创建后返回

    print(res)

    driverPath = res['data']['driver']
    debuggerAddress = res['data']['http']

    # selenium 连接代码
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_experimental_option("debuggerAddress", debuggerAddress)

    chrome_service = Service(driverPath)
    driver = webdriver.Chrome(service=chrome_service, options=chrome_options)

    driver.implicitly_wait(10)
    # 设置最长等待时间为 10 秒
    wait = WebDriverWait(driver, 10)

    driver.get(
        "https://global-selling.mercadolibre.com/orders/omni/list?filters=&subFilters=&search=&limit=50&offset=0&startPeriod=WITH_DATE_CLOSED_2M_OLD&selectedTab=TAB_TODAY_CBT")
    driver.refresh()
    time.sleep(5)

    # 这段 JS 脚本会自动寻找页面上所有隐藏的 Shadow DOM 并在其中搜索目标
    deep_click_script = """
    function findAndClick(root, selector) {
        // 1. 在当前层级寻找
        const node = root.querySelector(selector);
        if (node) {
            node.click();
            return true;
        }
    
        // 2. 递归寻找所有子节点的 Shadow DOM
        const allNodes = root.querySelectorAll('*');
        for (let i = 0; i < allNodes.length; i++) {
            if (allNodes[i].shadowRoot) {
                if (findAndClick(allNodes[i].shadowRoot, selector)) {
                    return true;
                }
            }
        }
        return false;
    }
    return findAndClick(document, 'button[aria-label="Select country"]');
    """
    # 打开选择器
    success = driver.execute_script(deep_click_script)
    # 选择站点
    name = ''
    if site == '墨西哥':
        name = 'Mexico'
    if site == '巴西':
        name = 'Brazil'
    if site == '哥伦比亚':
        name = 'Colombia'
    if site == '智利':
        name = 'Chile'
    if site == '阿根廷':
        name = 'Argentina'
    if site == '乌拉圭':
        name = 'Uruguay'
    #
    force_select_country(driver, name)
    print('成功选择站点')
    driver.get(
        "https://global-selling.mercadolibre.com/orders/omni/list?filters=&subFilters=&search=&limit=50&offset=0&startPeriod=WITH_DATE_CLOSED_2M_OLD&selectedTab=TAB_TODAY_CBT")
    driver.refresh()
    time.sleep(5)
    try:
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH,
                                        "/html/body/main/div/div[3]/div/div/div[3]/div/div[2]/div/div/section/div/div[1]/div/div/div[1]/div[1]/div/div/span/input"))).click()
    except Exception as e:
        print("无法勾选打印", e)
    try:
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH,
                                        "/html/body/main/div/div[3]/div/div/div[3]/div/div[2]/div/div/section/div/div[1]/div/div/div[2]/div/button"))).click()
    except Exception as e:
        print("没有可以打印的订单", e)
    return True



def print_orders_all():
    start = int(time.time())

    file_path = Path(__file__).resolve().parent / "比特配置文件.xlsx"
    print(start)

    wb = load_workbook(file_path)

    sheet = wb.active
    reputation_info_sum = []
    # 使用 min_row=2 跳过第一行
    result=[]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        print(row)  # row 是一个元组，包含该行所有数据
        id = row[0]
        name = row[1]
        remark = row[2]
        if remark == '忽略':
            continue
        print("开始打开窗口:", name)
        site_list = row[3].split("，")
        for site in site_list:
            i = 0
            while (i < 3):
                try:
                    success = print_orders(id, site)
                    if (success == True):
                        result.append(name+site+"打印订单任务执行成功")
                        break
                except Exception as e:
                    print("窗口" + name + site + "执行失败", e)
                    if(i==2):
                        result.append(name + site + "打印订单任务执行失败")
                time.sleep(300)

            time.sleep(5)

        print("结束，正在å关闭窗口")
        try:
            closeBrowser(str(id))
        except Exception as e:
            print("关闭窗口失败", e)
        print("已经关闭窗口")
        time.sleep(5)
    end = int(time.time())
    print("总花费", end - start)

if __name__ == '__main__':
    while(1==1):
        print_orders_all()
        time.sleep(3600*6)