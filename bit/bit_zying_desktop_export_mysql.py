"""Export ZYing desktop product pages and upsert every Excel row into MySQL.

The desktop exporter reproduces the verified 3440x1440 workflow:

1. Open 产品 -> 产品列表 and search for the requested category by name.
2. Visit each page, enable 多选, and select every product on that page.
3. Enable 只导出父项忽略变体 and save each workbook by page number.
4. Import all exported parent-product rows into MySQL.
5. In ``all`` mode, directly visit each source and checkpoint its main image.

The MySQL connection comes from ``bit.bit_mysql.config`` and therefore honors
the existing MYSQL_* / DB_* environment variables.  No credentials are stored
in this file.

Examples::

    # Import the already-exported 12 workbooks only.
    python bit/bit_zying_desktop_export_mysql.py --mode import \
        --input-dir outputs/zying_yimaxianshan_20260811

    # Run the desktop export and then import the new files.
    python bit/bit_zying_desktop_export_mysql.py --mode all \
        --category 一马当先 --pages 12
"""

from __future__ import annotations

import argparse
import ctypes
import hashlib
import json
import math
import os
import re
import subprocess
import sys
import time
from dataclasses import dataclass
from ctypes import wintypes
from datetime import date, datetime, time as datetime_time
from decimal import Decimal, InvalidOperation
from pathlib import Path


if __package__ in (None, ""):
    project_root = str(Path(__file__).resolve().parent.parent)
    if project_root not in sys.path:
        sys.path.insert(0, project_root)

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
    sys.stderr.reconfigure(encoding="utf-8", errors="backslashreplace")
except (AttributeError, ValueError):
    pass

import pymysql
from openpyxl import load_workbook

from bit.bit_mysql import config as mysql_config


DEFAULT_CATEGORY = "一马当先"
ALL_CATEGORIES = "全部分类"
DEFAULT_TABLE = "zying_desktop_products"
CURRENT_EXPORT_DIR = (
    Path(__file__).resolve().parent.parent
    / "outputs"
    / "zying_yimaxianshan_20260811"
)


def _enable_windows_dpi_awareness():
    """Use physical pixels consistently at 100%-300% Windows scaling."""
    if os.name != "nt":
        return
    try:
        # DPI_AWARENESS_CONTEXT_PER_MONITOR_AWARE_V2
        ctypes.windll.user32.SetProcessDpiAwarenessContext(ctypes.c_void_p(-4))
        return
    except Exception:
        pass
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(2)
    except Exception:
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass


# Excel header -> database column.  Order must match the ZYing export exactly.
HEADER_COLUMNS = (
    ("id", "product_id"),
    ("key", "variant_key"),
    ("theme", "variant_theme"),
    ("child", "variant_child"),
    ("SKU", "sku"),
    ("品牌", "brand"),
    ("币种", "currency"),
    ("分销价", "distribution_price"),
    ("成本价", "cost_price"),
    ("运费", "shipping_fee"),
    ("毛重", "gross_weight"),
    ("尺寸", "dimensions"),
    ("库存", "inventory"),
    ("材料", "material"),
    ("包装", "packaging"),
    ("中文简称", "short_name_zh"),
    ("英文简称", "short_name_en"),
    ("来源", "source_url"),
    ("采购链接", "purchase_url"),
    ("标题", "title"),
    ("英语-标题", "title_en"),
    ("西班牙-标题", "title_es"),
    ("葡萄牙-标题", "title_pt"),
    ("中文-标题", "title_zh"),
    ("短标题", "short_title"),
    ("英语-短标题", "short_title_en"),
    ("西班牙-短标题", "short_title_es"),
    ("葡萄牙-短标题", "short_title_pt"),
    ("中文-短标题", "short_title_zh"),
    ("关键字", "keywords"),
    ("英语-关键字", "keywords_en"),
    ("西班牙-关键字", "keywords_es"),
    ("葡萄牙-关键字", "keywords_pt"),
    ("中文-关键字", "keywords_zh"),
    ("要点", "bullet_points"),
    ("英语-要点", "bullet_points_en"),
    ("西班牙-要点", "bullet_points_es"),
    ("葡萄牙-要点", "bullet_points_pt"),
    ("中文-要点", "bullet_points_zh"),
    ("描述", "description"),
    ("英语-描述", "description_en"),
    ("西班牙-描述", "description_es"),
    ("葡萄牙-描述", "description_pt"),
    ("中文-描述", "description_zh"),
    ("分类", "category"),
    ("禁止平台", "forbidden_platforms"),
    ("禁止地区", "forbidden_regions"),
)

EXPECTED_HEADERS = tuple(header for header, _ in HEADER_COLUMNS)
DATA_COLUMNS = tuple(column for _, column in HEADER_COLUMNS)
NUMERIC_COLUMNS = {
    "distribution_price",
    "cost_price",
    "shipping_fee",
    "gross_weight",
    "inventory",
}
LONGTEXT_COLUMNS = {
    "variant_key",
    "variant_theme",
    "variant_child",
    "source_url",
    "purchase_url",
    "title",
    "title_en",
    "title_es",
    "title_pt",
    "title_zh",
    "short_title",
    "short_title_en",
    "short_title_es",
    "short_title_pt",
    "short_title_zh",
    "keywords",
    "keywords_en",
    "keywords_es",
    "keywords_pt",
    "keywords_zh",
    "bullet_points",
    "bullet_points_en",
    "bullet_points_es",
    "bullet_points_pt",
    "bullet_points_zh",
    "description",
    "description_en",
    "description_es",
    "description_pt",
    "description_zh",
    "forbidden_platforms",
    "forbidden_regions",
}

METADATA_COLUMNS = (
    "record_key",
    "content_hash",
    "export_category",
    "export_department",
    "export_salesperson",
    "export_collection_site",
    "export_source_region",
    "export_start_date",
    "export_end_date",
    "export_currency",
    "export_filters_json",
    "source_batch",
    "source_file",
    "source_page",
    "source_row",
    "raw_json",
)
INSERT_COLUMNS = DATA_COLUMNS + METADATA_COLUMNS


def _validate_table_name(table_name: str) -> str:
    table_name = str(table_name or "").strip()
    if not re.fullmatch(r"[A-Za-z][A-Za-z0-9_]{0,63}", table_name):
        raise ValueError(f"不安全的 MySQL 表名：{table_name!r}")
    return table_name


def _parse_category_result_count(texts) -> int | None:
    for text in texts:
        match = re.search(r"共有\s*(\d+)\s*条记录", str(text or ""))
        if match:
            return int(match.group(1))
    return None


def _parse_product_count(texts) -> int | None:
    for text in texts:
        match = re.search(r"共找到\s*(\d+)\s*个产品", str(text or ""))
        if match:
            return int(match.group(1))
    return None


def _safe_filename_component(value: str) -> str:
    value = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", str(value or "").strip())
    value = value.rstrip(" .")
    return value or "未命名分类"


def _cell_text(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, (datetime, date, datetime_time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        if value.is_integer():
            return str(int(value))
        return format(value, ".15g")
    text = str(value).strip()
    return text or None


def _decimal_value(value):
    text = _cell_text(value)
    if text is None:
        return None
    text = text.replace(",", "")
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def _hash_parts(parts) -> str:
    serialized = json.dumps(
        [_cell_text(value) for value in parts],
        ensure_ascii=False,
        separators=(",", ":"),
    )
    return hashlib.sha256(serialized.encode("utf-8")).hexdigest()


def _page_from_file(path: Path) -> int | None:
    match = re.search(r"第(\d+)页", path.stem)
    return int(match.group(1)) if match else None


def _normalize_headers(values) -> tuple[str, ...]:
    return tuple(str(value or "").strip() for value in values)


def _header_positions(headers) -> dict[str, int]:
    positions = {header: index for index, header in enumerate(headers)}
    # Parent-only exports intentionally omit the three variant identity fields.
    # Keep those values NULL while still accepting the complete legacy export.
    optional_headers = {"key", "theme", "child"}
    missing = [
        header
        for header in EXPECTED_HEADERS
        if header not in positions and header not in optional_headers
    ]
    if missing:
        raise ValueError(f"Excel 缺少字段：{', '.join(missing)}")
    return positions


def build_record(
    headers,
    row,
    *,
    category: str,
    department: str | None = None,
    salesperson: str | None = None,
    collection_site: str | None = None,
    source_region: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    currency: str | None = None,
    export_filters: dict | None = None,
    source_batch: str,
    source_file: str,
    source_page: int | None,
    source_row: int,
):
    """Convert one Excel row to a typed, idempotent MySQL record."""
    positions = _header_positions(headers)
    raw = {
        header: _cell_text(row[positions[header]])
        if header in positions and positions[header] < len(row)
        else None
        for header in EXPECTED_HEADERS
    }
    if not any(value is not None for value in raw.values()):
        return None
    product_id = raw["id"]
    if not product_id:
        raise ValueError(f"{source_file} 第 {source_row} 行没有产品 id")

    record = {}
    for header, column in HEADER_COLUMNS:
        value = raw[header]
        record[column] = _decimal_value(value) if column in NUMERIC_COLUMNS else value

    identity_parts = (
        category,
        record["product_id"],
        record["variant_key"],
        record["variant_theme"],
        record["variant_child"],
        record["sku"],
    )
    record["record_key"] = _hash_parts(identity_parts)
    record["content_hash"] = _hash_parts(raw.values())
    record["export_category"] = category
    record["export_department"] = _cell_text(department)
    record["export_salesperson"] = _cell_text(salesperson)
    record["export_collection_site"] = _cell_text(collection_site)
    record["export_source_region"] = _cell_text(source_region)
    record["export_start_date"] = _cell_text(start_date)
    record["export_end_date"] = _cell_text(end_date)
    record["export_currency"] = _cell_text(currency)
    record["export_filters_json"] = json.dumps(
        export_filters or {}, ensure_ascii=False, separators=(",", ":")
    )
    record["source_batch"] = source_batch
    record["source_file"] = source_file
    record["source_page"] = source_page
    record["source_row"] = source_row
    record["raw_json"] = json.dumps(raw, ensure_ascii=False, separators=(",", ":"))
    return record


def discover_export_files(input_dir: str | Path) -> list[Path]:
    input_dir = Path(input_dir).expanduser().resolve()
    if not input_dir.is_dir():
        raise FileNotFoundError(f"Excel 目录不存在：{input_dir}")
    files = [
        path for path in input_dir.glob("*.xlsx")
        if not path.name.startswith("~$")
    ]
    files.sort(key=lambda path: (_page_from_file(path) or 10**9, path.name))
    if not files:
        raise FileNotFoundError(f"目录内没有 .xlsx 文件：{input_dir}")
    return files


def iter_workbook_records(
    paths,
    *,
    category: str,
    source_batch: str,
    department: str | None = None,
    salesperson: str | None = None,
    collection_site: str | None = None,
    source_region: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    currency: str | None = None,
    export_filters: dict | None = None,
):
    for path in paths:
        page = _page_from_file(path)
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.worksheets[0]
            rows = sheet.iter_rows(values_only=True)
            try:
                headers = _normalize_headers(next(rows))
            except StopIteration as exc:
                raise ValueError(f"Excel 是空表：{path}") from exc
            _header_positions(headers)
            for row_number, row in enumerate(rows, start=2):
                record = build_record(
                    headers,
                    row,
                    category=category,
                    department=department,
                    salesperson=salesperson,
                    collection_site=collection_site,
                    source_region=source_region,
                    start_date=start_date,
                    end_date=end_date,
                    currency=currency,
                    export_filters=export_filters,
                    source_batch=source_batch,
                    source_file=path.name,
                    source_page=page,
                    source_row=row_number,
                )
                if record is not None:
                    yield record
        finally:
            workbook.close()


def _column_definition(column: str) -> str:
    if column == "product_id":
        return "VARCHAR(64) NOT NULL"
    if column in NUMERIC_COLUMNS:
        return "DECIMAL(24,6) NULL"
    if column in LONGTEXT_COLUMNS:
        return "LONGTEXT NULL"
    if column in {"sku", "brand", "dimensions", "material", "packaging", "category"}:
        return "VARCHAR(2048) NULL"
    if column == "currency":
        return "VARCHAR(32) NULL"
    return "LONGTEXT NULL"


def ensure_table(cursor, table_name: str):
    table_name = _validate_table_name(table_name)
    data_definitions = ",\n".join(
        f"`{column}` {_column_definition(column)}" for column in DATA_COLUMNS
    )
    cursor.execute(
        f"""
        CREATE TABLE IF NOT EXISTS `{table_name}` (
            {data_definitions},
            `record_key` CHAR(64) NOT NULL,
            `content_hash` CHAR(64) NOT NULL,
            `export_category` VARCHAR(255) NOT NULL,
            `export_department` VARCHAR(255) NULL,
            `export_salesperson` VARCHAR(255) NULL,
            `export_collection_site` VARCHAR(255) NULL,
            `export_source_region` VARCHAR(255) NULL,
            `export_start_date` DATE NULL,
            `export_end_date` DATE NULL,
            `export_currency` VARCHAR(32) NULL,
            `export_filters_json` LONGTEXT NULL,
            `source_batch` VARCHAR(255) NOT NULL,
            `source_file` VARCHAR(512) NOT NULL,
            `source_page` INT NULL,
            `source_row` INT NOT NULL,
            `raw_json` LONGTEXT NOT NULL,
            `imported_at` DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
            `updated_at` DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
                ON UPDATE CURRENT_TIMESTAMP,
            PRIMARY KEY (`record_key`),
            KEY `idx_zying_desktop_product` (`product_id`),
            KEY `idx_zying_desktop_category` (`export_category`),
            KEY `idx_zying_desktop_batch` (`source_batch`),
            KEY `idx_zying_desktop_page` (`source_page`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """
    )
    metadata_definitions = {
        "export_department": "VARCHAR(255) NULL",
        "export_salesperson": "VARCHAR(255) NULL",
        "export_collection_site": "VARCHAR(255) NULL",
        "export_source_region": "VARCHAR(255) NULL",
        "export_start_date": "DATE NULL",
        "export_end_date": "DATE NULL",
        "export_currency": "VARCHAR(32) NULL",
        "export_filters_json": "LONGTEXT NULL",
    }
    cursor.execute(f"SHOW COLUMNS FROM `{table_name}`")
    existing = {row["Field"] for row in cursor.fetchall()}
    for column, definition in metadata_definitions.items():
        if column not in existing:
            cursor.execute(
                f"ALTER TABLE `{table_name}` ADD COLUMN `{column}` {definition}"
            )


def _upsert_sql(table_name: str) -> str:
    table_name = _validate_table_name(table_name)
    quoted = ", ".join(f"`{column}`" for column in INSERT_COLUMNS)
    placeholders = ", ".join(["%s"] * len(INSERT_COLUMNS))
    updates = ", ".join(
        f"`{column}`=VALUES(`{column}`)"
        for column in INSERT_COLUMNS
        if column != "record_key"
    )
    return (
        f"INSERT INTO `{table_name}` ({quoted}) VALUES ({placeholders}) "
        f"ON DUPLICATE KEY UPDATE {updates}"
    )


def import_workbooks(
    input_dir: str | Path,
    *,
    table_name: str = DEFAULT_TABLE,
    category: str = DEFAULT_CATEGORY,
    department: str | None = None,
    salesperson: str | None = None,
    collection_site: str | None = None,
    source_region: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    currency: str | None = None,
    export_filters: dict | None = None,
    batch_size: int = 500,
    connection_factory=None,
):
    """Read every exported row and upsert it into MySQL in bounded batches."""
    paths = discover_export_files(input_dir)
    source_batch = Path(input_dir).resolve().name
    connection_factory = connection_factory or (lambda: pymysql.connect(**mysql_config))
    connection = connection_factory()
    processed = 0
    affected = 0
    batch = []
    sql = _upsert_sql(table_name)
    try:
        with connection.cursor() as cursor:
            ensure_table(cursor, table_name)
            for record in iter_workbook_records(
                paths,
                category=category,
                source_batch=source_batch,
                department=department,
                salesperson=salesperson,
                collection_site=collection_site,
                source_region=source_region,
                start_date=start_date,
                end_date=end_date,
                currency=currency,
                export_filters=export_filters,
            ):
                batch.append(tuple(record[column] for column in INSERT_COLUMNS))
                if len(batch) >= batch_size:
                    affected += cursor.executemany(sql, batch)
                    processed += len(batch)
                    print(f"MySQL 已处理 {processed} 行", flush=True)
                    batch.clear()
            if batch:
                affected += cursor.executemany(sql, batch)
                processed += len(batch)
                print(f"MySQL 已处理 {processed} 行", flush=True)
            connection.commit()

            cursor.execute(
                f"""
                SELECT COUNT(*) AS row_count,
                       COUNT(DISTINCT `product_id`) AS unique_products,
                       COUNT(DISTINCT `source_file`) AS source_files,
                       COUNT(DISTINCT `source_page`) AS source_pages
                FROM `{_validate_table_name(table_name)}`
                WHERE `export_category`=%s AND `source_batch`=%s
                """,
                (category, source_batch),
            )
            summary = cursor.fetchone() or {}
            cursor.execute(
                f"""
                SELECT `source_page`, COUNT(*) AS row_count,
                       COUNT(DISTINCT `product_id`) AS unique_products
                FROM `{_validate_table_name(table_name)}`
                WHERE `export_category`=%s AND `source_batch`=%s
                GROUP BY `source_page`
                ORDER BY `source_page`
                """,
                (category, source_batch),
            )
            page_summary = cursor.fetchall()
        return {
            "table": table_name,
            "category": category,
            "department": department,
            "salesperson": salesperson,
            "collection_site": collection_site,
            "source_region": source_region,
            "start_date": start_date,
            "end_date": end_date,
            "currency": currency,
            "export_filters": export_filters or {},
            "source_batch": source_batch,
            "excel_files": len(paths),
            "processed_rows": processed,
            "affected_rows": affected,
            "database_rows": int(summary.get("row_count") or 0),
            "unique_products": int(summary.get("unique_products") or 0),
            "source_files": int(summary.get("source_files") or 0),
            "source_pages": int(summary.get("source_pages") or 0),
            "pages": page_summary,
        }
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()


@dataclass(frozen=True)
class DesktopCoordinates:
    """DPI-relative fallbacks for the verified 1920x1080 window.

    Normal WinForms controls are located by automation id.  These coordinates
    are only used for the top product tab and custom-drawn category picker.
    """

    base_width: int = 1920
    base_height: int = 1080
    product_tab: tuple[int, int] = (295, 38)
    category_dropdown: tuple[int, int] = (895, 77)
    category_search: tuple[int, int] = (1120, 215)
    category_result: tuple[int, int] = (900, 327)
    salesperson: tuple[int, int] = (1135, 77)
    start_date: tuple[int, int] = (1530, 77)
    product_detail_close: tuple[int, int] = (1308, 77)
    page_y: int = 956
    page_x: tuple[int, ...] = tuple(156 + 31 * index for index in range(35))


class ZyingDesktopExporter:
    def __init__(
        self,
        *,
        category=ALL_CATEGORIES,
        product_ids=None,
        sku=None,
        keyword=None,
        department=None,
        salesperson=None,
        collection_site=None,
        source_region=None,
        start_date=None,
        end_date=None,
        currency=None,
        expected_count=None,
        parent_only=True,
        filter_wait=5.0,
        coordinates=None,
    ):
        if os.name != "nt":
            raise RuntimeError("智赢桌面端自动导出仅支持 Windows")
        _enable_windows_dpi_awareness()
        import pyautogui

        self.pyautogui = pyautogui
        self.pyautogui.FAILSAFE = True
        self.pyautogui.PAUSE = 0.12
        self.category = str(category or ALL_CATEGORIES).strip() or ALL_CATEGORIES
        self.product_ids = str(product_ids or "").strip() or None
        self.sku = str(sku or "").strip() or None
        self.keyword = str(keyword or "").strip() or None
        self.department = str(department or "").strip() or None
        self.salesperson = str(salesperson or "").strip() or None
        self.collection_site = str(collection_site or "").strip() or None
        self.source_region = str(source_region or "").strip() or None
        self.start_date = str(start_date or "").strip() or None
        self.end_date = str(end_date or "").strip() or None
        self.currency = str(currency or "").strip() or None
        self.expected_count = expected_count
        self.parent_only = bool(parent_only)
        self.filter_wait = max(0.0, float(filter_wait))
        self.coordinates = coordinates or DesktopCoordinates()
        self.hwnd = self._find_window()
        self._filters_prepared = False
        self.exported_product_count = None

    @staticmethod
    def _find_window():
        user32 = ctypes.windll.user32
        matches = []
        callback_type = ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.c_void_p, ctypes.c_void_p)

        def callback(hwnd, _):
            if not user32.IsWindowVisible(hwnd):
                return True
            length = user32.GetWindowTextLengthW(hwnd)
            if not length:
                return True
            buffer = ctypes.create_unicode_buffer(length + 1)
            user32.GetWindowTextW(hwnd, buffer, length + 1)
            if "分销系统" in buffer.value:
                matches.append(hwnd)
            return True

        user32.EnumWindows(callback_type(callback), 0)
        if not matches:
            raise RuntimeError("没有找到标题含“分销系统”的智赢桌面端窗口")
        return matches[0]

    def activate(self):
        user32 = ctypes.windll.user32
        kernel32 = ctypes.windll.kernel32
        foreground = user32.GetForegroundWindow()
        current_thread = kernel32.GetCurrentThreadId()
        foreground_thread = user32.GetWindowThreadProcessId(foreground, None)
        target_thread = user32.GetWindowThreadProcessId(self.hwnd, None)
        attached = []
        for thread_id in {foreground_thread, target_thread}:
            if thread_id and thread_id != current_thread:
                if user32.AttachThreadInput(current_thread, thread_id, True):
                    attached.append(thread_id)
        user32.ShowWindow(self.hwnd, 3)  # SW_MAXIMIZE
        try:
            user32.BringWindowToTop(self.hwnd)
            user32.SetForegroundWindow(self.hwnd)
            user32.SetActiveWindow(self.hwnd)
            user32.SetFocus(self.hwnd)
        finally:
            for thread_id in attached:
                user32.AttachThreadInput(current_thread, thread_id, False)
        time.sleep(0.45)

    def _scaled(self, point):
        rect = wintypes.RECT()
        ctypes.windll.user32.GetWindowRect(self.hwnd, ctypes.byref(rect))
        width = rect.right - rect.left
        height = rect.bottom - rect.top
        x = rect.left + round(point[0] * width / self.coordinates.base_width)
        y = rect.top + round(point[1] * height / self.coordinates.base_height)
        return x, y

    def _main_tab_point(self, reference_x: int, reference_y: int):
        """Map a top-tab reference point while preserving the 54px sidebar."""
        rect = wintypes.RECT()
        ctypes.windll.user32.GetWindowRect(self.hwnd, ctypes.byref(rect))
        width = rect.right - rect.left
        height = rect.bottom - rect.top
        sidebar_reference = 132
        if reference_x >= sidebar_reference:
            content_width = max(1, width - sidebar_reference)
            x = rect.left + sidebar_reference + round(
                (reference_x - sidebar_reference)
                * content_width
                / (self.coordinates.base_width - sidebar_reference)
            )
        else:
            x = rect.left + round(reference_x * width / self.coordinates.base_width)
        y = rect.top + round(reference_y * height / self.coordinates.base_height)
        return x, y

    def click(self, point, *, clicks=1, interval=0.0):
        self.pyautogui.click(*self._scaled(point), clicks=clicks, interval=interval)

    def _win32_control(self, automation_id: str):
        """Return a visible child control by stable WinForms automation id."""
        try:
            from pywinauto import Desktop

            window = Desktop(backend="win32").window(handle=self.hwnd)
            matches = []
            for child in window.descendants():
                try:
                    if (
                        getattr(child.element_info, "automation_id", "")
                        == automation_id
                        and child.is_visible()
                    ):
                        matches.append(child)
                except Exception:
                    continue
            return matches[-1] if matches else None
        except Exception:
            return None

    def _control_center(self, automation_id: str):
        control = self._win32_control(automation_id)
        if control is None:
            return None
        rect = control.rectangle()
        return ((rect.left + rect.right) // 2, (rect.top + rect.bottom) // 2)

    def _click_control(self, automation_id: str, fallback=None):
        center = self._control_center(automation_id)
        if center is not None:
            self.pyautogui.click(*center)
            return
        if fallback is None:
            raise RuntimeError(f"智赢界面没有找到控件：{automation_id}")
        self.click(fallback)

    def _fill_control(self, automation_id: str, value, fallback=None):
        self._click_control(automation_id, fallback)
        self.pyautogui.hotkey("ctrl", "a")
        self.paste(value)
        self.pyautogui.press("enter")
        if self.filter_wait:
            time.sleep(self.filter_wait)

    def _hide_category_popup(self):
        """Hide only the category popup without sending Escape to filters."""
        user32 = ctypes.windll.user32
        process_id = ctypes.c_ulong()
        user32.GetWindowThreadProcessId(self.hwnd, ctypes.byref(process_id))
        callback_type = ctypes.WINFUNCTYPE(
            ctypes.c_bool,
            ctypes.c_void_p,
            ctypes.c_void_p,
        )

        def callback(hwnd, _):
            window_process_id = ctypes.c_ulong()
            user32.GetWindowThreadProcessId(hwnd, ctypes.byref(window_process_id))
            if (
                hwnd != self.hwnd
                and window_process_id.value == process_id.value
                and user32.IsWindowVisible(hwnd)
            ):
                class_name = ctypes.create_unicode_buffer(256)
                user32.GetClassNameW(hwnd, class_name, len(class_name))
                if "WindowsForms10.Window.20808" in class_name.value:
                    user32.ShowWindow(hwnd, 0)
            return True

        user32.EnumWindows(callback_type(callback), 0)

    def _category_popup_rect(self):
        user32 = ctypes.windll.user32
        process_id = ctypes.c_ulong()
        user32.GetWindowThreadProcessId(self.hwnd, ctypes.byref(process_id))
        matches = []
        callback_type = ctypes.WINFUNCTYPE(
            ctypes.c_bool,
            ctypes.c_void_p,
            ctypes.c_void_p,
        )

        def callback(hwnd, _):
            window_process_id = ctypes.c_ulong()
            user32.GetWindowThreadProcessId(hwnd, ctypes.byref(window_process_id))
            if (
                hwnd != self.hwnd
                and window_process_id.value == process_id.value
                and user32.IsWindowVisible(hwnd)
            ):
                class_name = ctypes.create_unicode_buffer(256)
                user32.GetClassNameW(hwnd, class_name, len(class_name))
                if "WindowsForms10.Window.20808" in class_name.value:
                    rect = wintypes.RECT()
                    user32.GetWindowRect(hwnd, ctypes.byref(rect))
                    matches.append(rect)
            return True

        user32.EnumWindows(callback_type(callback), 0)
        return matches[-1] if matches else None

    def _show_export_dialog(self):
        """Restore ZYing's reusable export dialog when it was hidden."""
        user32 = ctypes.windll.user32
        process_id = ctypes.c_ulong()
        user32.GetWindowThreadProcessId(self.hwnd, ctypes.byref(process_id))
        matches = []
        callback_type = ctypes.WINFUNCTYPE(
            ctypes.c_bool,
            ctypes.c_void_p,
            ctypes.c_void_p,
        )

        def callback(hwnd, _):
            window_process_id = ctypes.c_ulong()
            user32.GetWindowThreadProcessId(hwnd, ctypes.byref(window_process_id))
            if window_process_id.value == process_id.value:
                length = user32.GetWindowTextLengthW(hwnd)
                buffer = ctypes.create_unicode_buffer(length + 1)
                user32.GetWindowTextW(hwnd, buffer, length + 1)
                if buffer.value == "DFulfill":
                    matches.append(hwnd)
            return True

        user32.EnumWindows(callback_type(callback), 0)
        if matches:
            user32.ShowWindow(matches[-1], 5)  # SW_SHOW
            user32.BringWindowToTop(matches[-1])
            return matches[-1]
        return None

    @staticmethod
    def _child_texts(hwnd):
        user32 = ctypes.windll.user32
        texts = []
        callback_type = ctypes.WINFUNCTYPE(
            ctypes.c_bool,
            ctypes.c_void_p,
            ctypes.c_void_p,
        )

        def callback(child_hwnd, _):
            if not user32.IsWindowVisible(child_hwnd):
                return True
            length = user32.GetWindowTextLengthW(child_hwnd)
            if length:
                buffer = ctypes.create_unicode_buffer(length + 1)
                user32.GetWindowTextW(child_hwnd, buffer, length + 1)
                texts.append(buffer.value.strip())
            return True

        user32.EnumChildWindows(hwnd, callback_type(callback), 0)
        return texts

    def _category_result_count(self):
        user32 = ctypes.windll.user32
        process_id = ctypes.c_ulong()
        user32.GetWindowThreadProcessId(self.hwnd, ctypes.byref(process_id))
        texts = []
        callback_type = ctypes.WINFUNCTYPE(
            ctypes.c_bool,
            ctypes.c_void_p,
            ctypes.c_void_p,
        )

        def callback(hwnd, _):
            window_process_id = ctypes.c_ulong()
            user32.GetWindowThreadProcessId(hwnd, ctypes.byref(window_process_id))
            if (
                hwnd != self.hwnd
                and window_process_id.value == process_id.value
                and user32.IsWindowVisible(hwnd)
            ):
                texts.extend(self._child_texts(hwnd))
            return True

        user32.EnumWindows(callback_type(callback), 0)
        return _parse_category_result_count(texts)

    def _current_product_count(self):
        return _parse_product_count(self._child_texts(self.hwnd))

    def _assert_filter_state(self, context=""):
        texts = self._child_texts(self.hwnd)
        if self.salesperson:
            salesperson_values = []
            # Page changes briefly rebuild the custom filter controls. UIA can
            # return an empty value during that redraw even though the selected
            # salesperson is still present. Retry reads only; never reapply the
            # filter while an export is in progress.
            for attempt in range(4):
                try:
                    from pywinauto import Desktop

                    uia_window = Desktop(backend="uia").window(handle=self.hwnd)
                    salesperson_values = [
                        child.window_text().strip()
                        for child in uia_window.descendants()
                        if 1050 <= child.rectangle().left <= 1300
                        and 50 <= child.rectangle().top <= 100
                    ]
                except Exception:
                    salesperson_values = []
                if self.salesperson in salesperson_values:
                    break
                if attempt < 3:
                    time.sleep(0.5)
            if self.salesperson not in salesperson_values:
                prefix = f"{context}：" if context else ""
                raise RuntimeError(
                    f"{prefix}业务员筛选未显示“{self.salesperson}”，已停止导出"
                )
        current_count = _parse_product_count(texts)
        if self.expected_count is not None and current_count != self.expected_count:
            prefix = f"{context}：" if context else ""
            raise RuntimeError(
                f"{prefix}产品数为 {current_count}，预期 {self.expected_count}，"
                "筛选条件可能已丢失"
            )
        return current_count

    @staticmethod
    def _checkbox_looks_checked(image, center) -> bool:
        center_x, center_y = center
        blue_pixels = 0
        for y in range(center_y - 9, center_y + 10):
            for x in range(center_x - 9, center_x + 10):
                red, green, blue = image.getpixel((x, y))[:3]
                if blue > 130 and blue - red > 40 and blue - green > 25:
                    blue_pixels += 1
        return blue_pixels >= 4

    def _set_checkbox_state(self, point, checked: bool):
        center = self._scaled(point)
        current = self._checkbox_looks_checked(self.pyautogui.screenshot(), center)
        if current != checked:
            self.click(point)
            time.sleep(0.5)
        actual = self._checkbox_looks_checked(self.pyautogui.screenshot(), center)
        if actual != checked:
            state = "选中" if checked else "未选中"
            raise RuntimeError(f"无法把复选框设置为{state}")

    @staticmethod
    def _set_clipboard(text: str):
        user32 = ctypes.windll.user32
        kernel32 = ctypes.windll.kernel32
        cf_unicode_text = 13
        gmem_moveable = 0x0002
        kernel32.GlobalAlloc.argtypes = (ctypes.c_uint, ctypes.c_size_t)
        kernel32.GlobalAlloc.restype = ctypes.c_void_p
        kernel32.GlobalLock.argtypes = (ctypes.c_void_p,)
        kernel32.GlobalLock.restype = ctypes.c_void_p
        kernel32.GlobalUnlock.argtypes = (ctypes.c_void_p,)
        kernel32.GlobalFree.argtypes = (ctypes.c_void_p,)
        user32.SetClipboardData.argtypes = (ctypes.c_uint, ctypes.c_void_p)
        user32.SetClipboardData.restype = ctypes.c_void_p
        data = (str(text) + "\0").encode("utf-16-le")
        handle = kernel32.GlobalAlloc(gmem_moveable, len(data))
        if not handle:
            raise OSError("GlobalAlloc failed")
        pointer = kernel32.GlobalLock(handle)
        ctypes.memmove(pointer, data, len(data))
        kernel32.GlobalUnlock(handle)
        if not user32.OpenClipboard(None):
            raise OSError("OpenClipboard failed")
        try:
            user32.EmptyClipboard()
            if not user32.SetClipboardData(cf_unicode_text, handle):
                raise OSError("SetClipboardData failed")
            handle = None
        finally:
            user32.CloseClipboard()
            if handle:
                kernel32.GlobalFree(handle)

    def paste(self, text: str):
        self._set_clipboard(text)
        self.pyautogui.hotkey("ctrl", "v")

    def _visible_child_handles(self, text=None):
        user32 = ctypes.windll.user32
        matches = []
        seen = set()
        process_id = ctypes.c_ulong()
        user32.GetWindowThreadProcessId(self.hwnd, ctypes.byref(process_id))
        callback_type = ctypes.WINFUNCTYPE(
            ctypes.c_bool,
            ctypes.c_void_p,
            ctypes.c_void_p,
        )

        def callback(hwnd, _):
            if hwnd in seen:
                return True
            seen.add(hwnd)
            if not user32.IsWindowVisible(hwnd):
                return True
            length = user32.GetWindowTextLengthW(hwnd)
            value = ""
            if length:
                buffer = ctypes.create_unicode_buffer(length + 1)
                user32.GetWindowTextW(hwnd, buffer, length + 1)
                value = buffer.value.strip()
            if text is None or value == text:
                rect = wintypes.RECT()
                user32.GetWindowRect(hwnd, ctypes.byref(rect))
                matches.append((hwnd, rect))
            return True

        user32.EnumChildWindows(self.hwnd, callback_type(callback), 0)

        def top_level_callback(hwnd, _):
            window_process_id = ctypes.c_ulong()
            user32.GetWindowThreadProcessId(hwnd, ctypes.byref(window_process_id))
            if window_process_id.value == process_id.value:
                callback(hwnd, None)
                user32.EnumChildWindows(hwnd, callback_type(callback), 0)
            return True

        user32.EnumWindows(callback_type(top_level_callback), 0)
        matches.sort(key=lambda item: (item[1].top, item[1].left))
        return matches

    @staticmethod
    def _message_click(hwnd, *, x=None, y=None):
        try:
            from pywinauto import Desktop
        except ImportError as exc:
            raise RuntimeError("请先安装 pywinauto：pip install pywinauto") from exc
        wrapper = Desktop(backend="win32").window(handle=hwnd)
        rect = wrapper.rectangle()
        wrapper.click(
            coords=(
                (rect.width() // 2) if x is None else x,
                (rect.height() // 2) if y is None else y,
            )
        )

    def _click_child_text(self, text, *, last=True, x=None, y=None):
        matches = self._visible_child_handles(text)
        if not matches:
            raise RuntimeError(f"没有找到智赢控件：{text}")
        hwnd = matches[-1 if last else 0][0]
        self._message_click(hwnd, x=x, y=y)
        return hwnd

    def prepare_filters(self):
        if self._filters_prepared:
            return
        self.activate()
        self.pyautogui.click(*self._main_tab_point(*self.coordinates.product_tab))
        time.sleep(3)
        # Prefer stable WinForms automation ids. This keeps the filter row
        # usable across screen resolutions, DPI scaling and window sizes.
        simple_filters = (
            ("PID", self.product_ids, None),
            ("PKey", self.sku, None),
            ("PWord", self.keyword, None),
            ("PSection", self.department, None),
            ("PLogin", self.salesperson, self.coordinates.salesperson),
            ("PSource", self.collection_site, None),
            ("PArea", self.source_region, None),
            ("PCur", self.currency, None),
        )
        for automation_id, value, fallback in simple_filters:
            if value:
                self._fill_control(automation_id, value, fallback)

        if self.category != ALL_CATEGORIES:
            self._click_control("PKind", self.coordinates.category_dropdown)
            time.sleep(1)
            category = self._win32_control("PKind")
            popup_rect = self._category_popup_rect()
            if popup_rect is not None:
                search_point = (
                    popup_rect.left + round((popup_rect.right - popup_rect.left) * 0.91),
                    popup_rect.top + round((popup_rect.bottom - popup_rect.top) * 0.32),
                )
                self.pyautogui.click(*search_point)
            else:
                self.click(self.coordinates.category_search)
            self.pyautogui.hotkey("ctrl", "a")
            self.paste(self.category)
            self.pyautogui.press("enter")
            time.sleep(1.5)
            result_count = self._category_result_count()
            if result_count != 1:
                self._hide_category_popup()
                if result_count == 0:
                    raise RuntimeError(f"智赢中没有找到产品分类：{self.category}")
                if result_count is None:
                    raise RuntimeError("无法读取智赢分类搜索结果，请检查分类弹窗")
                raise RuntimeError(
                    f"分类“{self.category}”匹配到 {result_count} 条，请输入更精确的名称"
                )
            # The category results are custom drawn. Select the single verified
            # result at the first-row center relative to the actual PKind box.
            popup_rect = self._category_popup_rect()
            if popup_rect is not None:
                result_point = (
                    popup_rect.left + round((popup_rect.right - popup_rect.left) * 0.50),
                    popup_rect.top + round((popup_rect.bottom - popup_rect.top) * 0.60),
                )
                self.pyautogui.click(*result_point, clicks=2, interval=0.15)
            else:
                self.click(self.coordinates.category_result, clicks=2, interval=0.15)
            time.sleep(max(1.0, self.filter_wait))

        if self.start_date:
            date_control = self._win32_control("PDate")
            if date_control is not None:
                rect = date_control.rectangle()
                self.pyautogui.click(
                    rect.left + round(rect.width() * 0.25),
                    (rect.top + rect.bottom) // 2,
                )
                self.pyautogui.hotkey("ctrl", "a")
                self.paste(self.start_date)
                self.pyautogui.press("enter")
                if self.filter_wait:
                    time.sleep(self.filter_wait)
            else:
                self._fill_control("PDate", self.start_date, self.coordinates.start_date)
        if self.end_date:
            date_control = self._win32_control("PDate")
            if date_control is None:
                raise RuntimeError("智赢界面没有找到截止日期控件 PDate")
            rect = date_control.rectangle()
            self.pyautogui.click(
                rect.left + round(rect.width() * 0.75),
                (rect.top + rect.bottom) // 2,
            )
            self.pyautogui.hotkey("ctrl", "a")
            self.paste(self.end_date)
            self.pyautogui.press("enter")
            if self.filter_wait:
                time.sleep(self.filter_wait)
        self._filters_prepared = True

    def prepare_category(self):
        return self.prepare_filters()

    def _ensure_multi_select(self):
        texts = set(self._child_texts(self.hwnd))
        if {"全选", "不选"} & texts:
            return
        self._click_child_text("多选 |", x=28, y=14)
        time.sleep(2)
        if not ({"全选", "不选"} & set(self._child_texts(self.hwnd))):
            raise RuntimeError("智赢未能进入多选模式")

    def _deselect_current_page(self):
        if "不选" in self._child_texts(self.hwnd):
            self._click_child_text("不选", x=14, y=14)
            time.sleep(1)

    def _select_page(self, page: int):
        if not 1 <= page <= len(self.coordinates.page_x):
            raise ValueError(f"当前坐标配置仅支持 1-{len(self.coordinates.page_x)} 页")
        pagers = self._visible_child_handles("eTurn1")
        if not pagers:
            raise RuntimeError("没有找到智赢分页控件")
        # Pages 1-9 use 27 px buttons with a 4 px gap. Two-digit page
        # numbers widen the button to 29 px, so pages 10+ advance by 33 px.
        relative_x = (
            13 + 31 * (page - 1)
            if page <= 9
            else 293 + 33 * (page - 10)
        )
        pager_width = pagers[-1][1].right - pagers[-1][1].left
        if relative_x >= pager_width:
            self.activate()
            close_control = self._control_center("close")
            if close_control is not None:
                self.pyautogui.click(*close_control)
            else:
                self.click(self.coordinates.product_detail_close)
            time.sleep(2)
            pagers = self._visible_child_handles("eTurn1")
            pager_width = pagers[-1][1].right - pagers[-1][1].left
        if relative_x >= pager_width:
            raise RuntimeError(f"第 {page} 页当前不可见，请先关闭右侧产品详情")
        # eTurn is a custom-drawn WinForms pager. A background/control click
        # does not change pages even though it reports success; use a real
        # mouse click and verify the selected-page background colour.
        pager_rect = pagers[-1][1]
        for attempt in range(3):
            self.activate()
            self.pyautogui.click(
                pager_rect.left + relative_x,
                pager_rect.top + 14,
            )
            time.sleep(5)
            screenshot = self.pyautogui.screenshot()
            active_pixel = screenshot.getpixel(
                (pager_rect.left + relative_x, pager_rect.top + 4)
            )[:3]
            if not (max(active_pixel) > 235 and min(active_pixel) > 225):
                return
        raise RuntimeError(f"智赢未能切换到第 {page} 页")

    def _select_all_current_page(self):
        self._ensure_multi_select()
        texts = self._child_texts(self.hwnd)
        if "不选" in texts:
            return
        if "全选" not in texts:
            raise RuntimeError("没有找到智赢产品列表的“全选”按钮")
        self._click_child_text("全选", x=14, y=14)
        time.sleep(1.8)
        if "不选" not in self._child_texts(self.hwnd):
            raise RuntimeError("当前页产品未能全部选中")

    def _select_product_export_action(self):
        """Select 产品导出 in the bottom-right batch action dropdown."""
        if self._visible_child_handles("产品导出"):
            return
        current = None
        for text in (
            "批量翻译",
            "检查侵权",
            "批量替换",
            "业务合并",
            "Meli分类属性",
        ):
            matches = self._visible_child_handles(text)
            if matches:
                current = matches[-1]
                break
        if current is None:
            raise RuntimeError("没有找到智赢右下角的批处理下拉框")
        hwnd, rect = current
        self._message_click(hwnd)
        time.sleep(0.5)
        # The WinForms popup is drawn rather than exposed as a child window.
        # 产品导出 is the fifth menu row and its centre is 72 px above the
        # collapsed control on the verified desktop layout.
        self.pyautogui.click(rect.left + 45, rect.top - 72)
        time.sleep(0.8)
        if not self._visible_child_handles("产品导出"):
            raise RuntimeError("智赢批处理动作未能切换到“产品导出”")

    @staticmethod
    def _window_children(hwnd):
        user32 = ctypes.windll.user32
        rows = []
        callback_type = ctypes.WINFUNCTYPE(
            ctypes.c_bool,
            ctypes.c_void_p,
            ctypes.c_void_p,
        )

        def callback(child, _):
            class_name = ctypes.create_unicode_buffer(256)
            user32.GetClassNameW(child, class_name, 256)
            length = user32.GetWindowTextLengthW(child)
            value = ""
            if length:
                buffer = ctypes.create_unicode_buffer(length + 1)
                user32.GetWindowTextW(child, buffer, length + 1)
                value = buffer.value
            rect = wintypes.RECT()
            user32.GetWindowRect(child, ctypes.byref(rect))
            rows.append((child, class_name.value, value, rect))
            return True

        user32.EnumChildWindows(hwnd, callback_type(callback), 0)
        return rows

    def _find_save_dialog(self, timeout=30):
        user32 = ctypes.windll.user32
        deadline = time.time() + timeout
        while time.time() < deadline:
            windows = []
            callback_type = ctypes.WINFUNCTYPE(
                ctypes.c_bool,
                ctypes.c_void_p,
                ctypes.c_void_p,
            )

            def callback(hwnd, _):
                if user32.IsWindowVisible(hwnd):
                    class_name = ctypes.create_unicode_buffer(256)
                    user32.GetClassNameW(hwnd, class_name, 256)
                    if class_name.value == "#32770":
                        windows.append(hwnd)
                return True

            user32.EnumWindows(callback_type(callback), 0)
            for hwnd in windows:
                children = self._window_children(hwnd)
                dialog_rect = wintypes.RECT()
                user32.GetWindowRect(hwnd, ctypes.byref(dialog_rect))
                # Current Windows save dialogs render the selected filename
                # through DirectUI, so GetWindowTextW may return an empty
                # string for its Edit child. The filename edit is the wide,
                # visible Edit in the lower half of the dialog; the other
                # wide Edit near the top belongs to the address bar.
                edits = [
                    row for row in children
                    if row[1] == "Edit"
                    and user32.IsWindowVisible(row[0])
                    and (row[3].right - row[3].left) > 200
                    and row[3].top > dialog_rect.top + (
                        dialog_rect.bottom - dialog_rect.top
                    ) // 2
                ]
                buttons = [
                    row for row in children
                    if row[1] == "Button"
                    and user32.IsWindowVisible(row[0])
                    and user32.IsWindowEnabled(row[0])
                    and (row[3].right - row[3].left) > 0
                ]
                if edits and buttons:
                    edits.sort(key=lambda row: row[3].top, reverse=True)
                    save_buttons = [
                        row for row in buttons
                        if row[2].startswith("保存")
                    ]
                    if not save_buttons:
                        bottom = max(row[3].top for row in buttons)
                        save_buttons = [
                            row for row in buttons if row[3].top >= bottom - 5
                        ]
                        save_buttons.sort(key=lambda row: row[3].left)
                    return edits[0][0], save_buttons[0][0]
            time.sleep(0.25)
        raise TimeoutError("没有找到智赢的 Excel 保存对话框")

    def _save_dialog(self, output_path: Path):
        output_path.parent.mkdir(parents=True, exist_ok=True)
        filename_edit, save_button = self._find_save_dialog()
        ctypes.windll.user32.SendMessageW(
            filename_edit,
            0x000C,
            0,
            str(output_path),
        )
        ctypes.windll.user32.SendMessageW(save_button, 0x00F5, 0, 0)

    @staticmethod
    def _close_wps_workbook(path: Path, attempts=20):
        quoted = str(path).replace("'", "''")
        command = (
            "$app=[Runtime.InteropServices.Marshal]::GetActiveObject('Ket.Application');"
            "$found=$false;"
            "foreach($book in @($app.Workbooks)){"
            f"if($book.FullName -eq '{quoted}')"
            "{$book.Close($false);$found=$true;break}};"
            "if($found){exit 0}else{exit 2}"
        )
        for _ in range(attempts):
            result = subprocess.run(
                ["powershell", "-NoProfile", "-Command", command],
                capture_output=True,
                text=True,
                check=False,
            )
            if result.returncode == 0:
                return True
            time.sleep(0.5)
        return False

    @staticmethod
    def _wait_for_file(path: Path, timeout=90):
        deadline = time.time() + timeout
        last_size = -1
        stable = 0
        while time.time() < deadline:
            if path.exists():
                size = path.stat().st_size
                stable = stable + 1 if size > 0 and size == last_size else 0
                last_size = size
                if stable >= 2:
                    return
            time.sleep(0.5)
        raise TimeoutError(f"智赢导出文件未生成：{path}")

    def export_page(self, page: int, output_path: Path):
        self.activate()
        # Do not press Escape here.  In the ZYing product list the salesperson
        # selector is an editable drop-down, and Escape clears its selected
        # value even after the drop-down has already closed.  That silently
        # turns a salesperson-scoped export back into an all-products export.
        # Stale export dialogs are handled explicitly below instead.
        self._assert_filter_state(f"导出第 {page} 页前")
        self._ensure_multi_select()
        self._deselect_current_page()
        self._select_page(page)
        self._assert_filter_state(f"切换到第 {page} 页后")
        self._select_all_current_page()
        self._select_product_export_action()
        self._click_child_text("GO", x=17, y=12)
        time.sleep(3)
        field_all = self._visible_child_handles("全选")
        parent_only = self._visible_child_handles("只导出父项忽略变体")
        if not field_all or not parent_only:
            self._show_export_dialog()
            time.sleep(1)
            field_all = self._visible_child_handles("全选")
            parent_only = self._visible_child_handles("只导出父项忽略变体")
        if not field_all or not parent_only:
            raise RuntimeError("智赢没有打开产品批量导出设置窗口")
        screenshot = self.pyautogui.screenshot()
        all_rect = field_all[-1][1]
        if not self._checkbox_looks_checked(
            screenshot,
            (all_rect.left + 8, all_rect.top + 14),
        ):
            self._message_click(field_all[-1][0], x=8, y=14)
        screenshot = self.pyautogui.screenshot()
        parent_rect = parent_only[-1][1]
        parent_checked = self._checkbox_looks_checked(
            screenshot,
            (parent_rect.left + 8, parent_rect.top + 14),
        )
        if parent_checked != self.parent_only:
            self._message_click(parent_only[-1][0], x=8, y=14)
        if output_path.exists():
            output_path.unlink()
        self._click_child_text("确认导出")
        time.sleep(3)
        self._save_dialog(output_path)
        self._wait_for_file(output_path)
        time.sleep(1)
        self._close_wps_workbook(output_path)
        # Hide the modeless settings window between pages. Its own 中止 button
        # stops generation but leaves the window visible, while WM_CLOSE
        # disposes the singleton and breaks all later exports in this session.
        confirm = self._visible_child_handles("确认导出")
        if confirm:
            dialog = ctypes.windll.user32.GetParent(confirm[-1][0])
            ctypes.windll.user32.ShowWindow(dialog, 0)  # SW_HIDE
        time.sleep(1)
        self._assert_filter_state(f"导出第 {page} 页后")

    @staticmethod
    def _inspect_export_file(path: Path):
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            rows = workbook.worksheets[0].iter_rows(values_only=True)
            headers = _normalize_headers(next(rows))
            positions = _header_positions(headers)
            ids = []
            for row in rows:
                value = row[positions["id"]] if positions["id"] < len(row) else None
                if value is not None:
                    ids.append(str(value))
            return headers, ids
        finally:
            workbook.close()

    def export_pages(
        self,
        output_dir: str | Path,
        *,
        pages=None,
        allow_count_drift=False,
    ):
        output_dir = Path(output_dir).expanduser().resolve()
        output_dir.mkdir(parents=True, exist_ok=True)
        self.prepare_filters()
        product_count = self._assert_filter_state("筛选初始化后")
        if product_count is None:
            raise RuntimeError("无法读取当前分类的产品总数")
        if (
            self.expected_count is not None
            and product_count != self.expected_count
            and not allow_count_drift
        ):
            raise RuntimeError(
                f"智赢当前有 {product_count} 个产品，与预期 "
                f"{self.expected_count} 不一致，已停止导出"
            )
        inferred_pages = max(1, math.ceil(product_count / 500))
        pages = inferred_pages if pages is None else pages
        if pages > len(self.coordinates.page_x):
            raise RuntimeError(
                f"当前分类需要 {pages} 页，坐标配置最多支持 "
                f"{len(self.coordinates.page_x)} 页"
            )
        print(
            f"分类“{self.category}”共有 {product_count} 个产品，导出 {pages} 页",
            flush=True,
        )
        exported = []
        safe_category = (
            "全部产品"
            if self.category == ALL_CATEGORIES
            else _safe_filename_component(self.category)
        )
        seen_ids = set()
        for page in range(1, pages + 1):
            output_path = output_dir / f"{safe_category}_第{page:02d}页.xlsx"
            if output_path.exists() and output_path.stat().st_size > 0:
                print(f"第 {page} 页文件已存在，跳过：{output_path.name}", flush=True)
            else:
                print(f"开始导出第 {page}/{pages} 页", flush=True)
                self.export_page(page, output_path)
                print(
                    f"第 {page}/{pages} 页完成：{output_path.name} "
                    f"({output_path.stat().st_size} bytes)",
                    flush=True,
                )
            headers, ids = self._inspect_export_file(output_path)
            expected_rows = min(500, product_count - (page - 1) * 500)
            last_page_drift = (
                allow_count_drift
                and page == pages
                and 0 < len(ids) <= expected_rows
                and len(ids) == len(set(ids))
            )
            if (
                len(ids) != expected_rows or len(ids) != len(set(ids))
            ) and not last_page_drift:
                raise RuntimeError(
                    f"第 {page} 页校验失败：应有 {expected_rows} 个唯一产品，"
                    f"实际 {len(ids)} 行/{len(set(ids))} 个唯一产品"
                )
            overlap = seen_ids.intersection(ids)
            if overlap:
                raise RuntimeError(
                    f"第 {page} 页与前序页面重复 {len(overlap)} 个产品，已停止"
                )
            seen_ids.update(ids)
            print(
                f"第 {page}/{pages} 页校验通过：{len(headers)} 列，"
                f"{len(ids)} 个唯一产品",
                flush=True,
            )
            exported.append(output_path)
        if len(seen_ids) != product_count and not allow_count_drift:
            raise RuntimeError(
                f"全量校验失败：智赢显示 {product_count} 个产品，"
                f"Excel 合计 {len(seen_ids)} 个唯一产品"
            )
        self.exported_product_count = len(seen_ids)
        if len(seen_ids) != product_count:
            print(
                f"注意：智赢界面计数为 {product_count}，Excel 实际为 "
                f"{len(seen_ids)} 个唯一产品；已按实际导出结果继续",
                flush=True,
            )
        return exported


def _default_new_export_dir():
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return Path(__file__).resolve().parent.parent / "outputs" / f"zying_desktop_{stamp}"


def main():
    config_probe = argparse.ArgumentParser(add_help=False)
    config_probe.add_argument("--config")
    config_args, _ = config_probe.parse_known_args()
    config_defaults = {}
    if config_args.config:
        config_path = Path(config_args.config).expanduser().resolve()
        with config_path.open("r", encoding="utf-8-sig") as stream:
            config_defaults = json.load(stream)
        if not isinstance(config_defaults, dict):
            raise ValueError("--config 必须是一个 JSON 对象")
        config_defaults = {
            str(key).strip().replace("-", "_"): value
            for key, value in config_defaults.items()
        }

    parser = argparse.ArgumentParser(
        description="智赢桌面端分页导出并写入 MySQL",
    )
    parser.add_argument("--config", help="JSON 配置文件；命令行参数可覆盖配置值")
    parser.add_argument(
        "--mode",
        choices=("export", "import", "images", "all"),
        default="all",
        help="export=仅导出，import=仅入库，images=仅抓主图，all=全部流程",
    )
    parser.add_argument(
        "--category",
        help="产品分类；使用“全部分类”表示不限制分类",
    )
    parser.add_argument("--product-ids", help="产品编号；多个编号按智赢支持的分隔符填写")
    parser.add_argument("--sku", help="SKU 筛选")
    parser.add_argument("--keyword", help="关键词筛选")
    parser.add_argument("--department", help="部门名称筛选")
    parser.add_argument("--salesperson", help="可选业务员筛选")
    parser.add_argument("--collection-site", help="采集网站筛选")
    parser.add_argument("--source-region", help="来源地区筛选")
    parser.add_argument("--start-date", help="可选起始日期，格式 YYYY-MM-DD")
    parser.add_argument("--end-date", help="可选截止日期，格式 YYYY-MM-DD")
    parser.add_argument("--currency", help="币种筛选")
    parser.add_argument(
        "--expected-count",
        type=int,
        help="导出前必须匹配的产品总数，不一致则停止",
    )
    parser.add_argument(
        "--pages",
        type=int,
        help="导出页数；未提供时按当前分类产品总数自动计算",
    )
    parser.add_argument("--table", default=DEFAULT_TABLE)
    parser.add_argument("--batch-size", type=int, default=500)
    parser.add_argument(
        "--source-images",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="all 模式导入完成后直连来源获取主图（默认启用）",
    )
    parser.add_argument(
        "--image-limit",
        type=int,
        help="本次最多获取多少个产品的主图；未提供时处理全部待处理产品",
    )
    parser.add_argument("--image-delay", type=float, default=0.4)
    parser.add_argument("--image-timeout", type=float, default=30)
    parser.add_argument("--image-token-env", default="MELI_ACCESS_TOKEN")
    parser.add_argument(
        "--source-batch",
        help="images 模式要处理的 MySQL 批次名（通常是 Excel 目录名）",
    )
    parser.add_argument(
        "--force-images",
        action=argparse.BooleanOptionalAction,
        default=False,
        help="即使已有成功记录也重新抓取主图",
    )
    parser.add_argument(
        "--parent-only",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="只导出父项、忽略变体（默认启用；用 --no-parent-only 关闭）",
    )
    parser.add_argument(
        "--reuse-current-filters",
        action="store_true",
        help="复用智赢界面当前筛选，只做校验，不重新选择筛选条件",
    )
    parser.add_argument(
        "--allow-count-drift",
        action=argparse.BooleanOptionalAction,
        default=False,
        help="允许智赢显示总数有缓存偏差，以 Excel 实际唯一产品数为准",
    )
    parser.add_argument(
        "--filter-wait",
        type=float,
        default=5.0,
        help="每个筛选条件提交后的等待秒数",
    )
    parser.add_argument(
        "--output-dir",
        help="桌面端导出的目标目录；未指定时创建带时间戳的新目录",
    )
    parser.add_argument(
        "--input-dir",
        help="import 模式的 Excel 目录；默认使用本次已导出的12个文件",
    )
    valid_config_keys = {action.dest for action in parser._actions}
    unknown_config_keys = sorted(set(config_defaults) - valid_config_keys)
    if unknown_config_keys:
        parser.error(
            "JSON 配置含未知参数：" + ", ".join(unknown_config_keys)
        )
    parser.set_defaults(**config_defaults)
    args = parser.parse_args()

    if args.category is None:
        if args.mode in {"import", "images"}:
            args.category = DEFAULT_CATEGORY
        else:
            try:
                args.category = input(
                    "请输入产品分类（直接回车表示全部分类）："
                ).strip() or ALL_CATEGORIES
            except EOFError:
                args.category = ALL_CATEGORIES
    else:
        args.category = args.category.strip()
    if not args.category:
        parser.error("产品分类不能为空")
    for date_argument in ("start_date", "end_date"):
        value = getattr(args, date_argument)
        if value:
            try:
                setattr(args, date_argument, date.fromisoformat(value).isoformat())
            except ValueError:
                parser.error(f"--{date_argument.replace('_', '-')} 必须是 YYYY-MM-DD 格式")
    if args.start_date and args.end_date and args.start_date > args.end_date:
        parser.error("--start-date 不能晚于 --end-date")
    if args.expected_count is not None and args.expected_count < 1:
        parser.error("--expected-count 必须大于 0")

    _validate_table_name(args.table)
    if args.pages is not None and not (
        1 <= args.pages <= len(DesktopCoordinates().page_x)
    ):
        parser.error(
            f"--pages 必须在 1 到 {len(DesktopCoordinates().page_x)} 之间"
        )
    if args.batch_size < 1:
        parser.error("--batch-size 必须大于 0")
    if args.image_limit is not None and args.image_limit < 1:
        parser.error("--image-limit 必须大于 0")
    if args.image_delay < 0:
        parser.error("--image-delay 不能小于 0")
    if args.image_timeout <= 0:
        parser.error("--image-timeout 必须大于 0")
    if args.filter_wait < 0:
        parser.error("--filter-wait 不能小于 0")

    output_dir = (
        Path(args.output_dir).expanduser().resolve()
        if args.output_dir
        else _default_new_export_dir()
    )
    input_dir = None
    if args.mode == "import":
        input_dir = (
            Path(args.input_dir).expanduser().resolve()
            if args.input_dir
            else CURRENT_EXPORT_DIR
        )
    elif args.mode in {"export", "all"}:
        exporter = ZyingDesktopExporter(
            category=args.category,
            product_ids=args.product_ids,
            sku=args.sku,
            keyword=args.keyword,
            department=args.department,
            salesperson=args.salesperson,
            collection_site=args.collection_site,
            source_region=args.source_region,
            start_date=args.start_date,
            end_date=args.end_date,
            currency=args.currency,
            expected_count=args.expected_count,
            parent_only=args.parent_only,
            filter_wait=args.filter_wait,
        )
        if args.reuse_current_filters:
            exporter._filters_prepared = True
        exporter.export_pages(
            output_dir,
            pages=args.pages,
            allow_count_drift=args.allow_count_drift,
        )
        input_dir = output_dir

    export_filters = {
        key: value
        for key, value in {
            "product_ids": args.product_ids,
            "sku": args.sku,
            "keyword": args.keyword,
            "category": args.category,
            "department": args.department,
            "salesperson": args.salesperson,
            "collection_site": args.collection_site,
            "source_region": args.source_region,
            "start_date": args.start_date,
            "end_date": args.end_date,
            "currency": args.currency,
        }.items()
        if value not in (None, "", ALL_CATEGORIES)
    }

    result = {
        "mode": args.mode,
        "category": args.category,
        "department": args.department,
        "salesperson": args.salesperson,
        "collection_site": args.collection_site,
        "source_region": args.source_region,
        "start_date": args.start_date,
        "end_date": args.end_date,
        "currency": args.currency,
        "filters": export_filters,
    }
    if args.mode == "import":
        result["input_dir"] = str(input_dir)
    elif args.mode in {"export", "all"}:
        result["output_dir"] = str(output_dir)
        result["parent_only"] = args.parent_only
        result["exported_product_count"] = exporter.exported_product_count
    if args.mode in {"import", "all"}:
        result["mysql"] = import_workbooks(
            input_dir,
            table_name=args.table,
            category=args.category,
            department=args.department,
            salesperson=args.salesperson,
            collection_site=args.collection_site,
            source_region=args.source_region,
            start_date=args.start_date,
            end_date=args.end_date,
            currency=args.currency,
            export_filters=export_filters,
            batch_size=args.batch_size,
        )
    if args.mode == "all" and args.source_images:
        from bit.bit_source_main_images import fetch_images_to_mysql

        result["source_images"] = fetch_images_to_mysql(
            table_name=args.table,
            category=args.category,
            source_batch=result["mysql"]["source_batch"],
            limit=args.image_limit,
            delay=args.image_delay,
            timeout=args.image_timeout,
            force=args.force_images,
            token_env=args.image_token_env,
        )
    elif args.mode == "images":
        if not args.source_batch:
            parser.error("--mode images 必须提供 --source-batch")
        from bit.bit_source_main_images import fetch_images_to_mysql

        result["source_images"] = fetch_images_to_mysql(
            table_name=args.table,
            category=args.category,
            source_batch=args.source_batch,
            limit=args.image_limit,
            delay=args.image_delay,
            timeout=args.image_timeout,
            force=args.force_images,
            token_env=args.image_token_env,
        )
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
