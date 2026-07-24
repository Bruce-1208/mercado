"""Mercado 店铺登录态检测，供申诉、声誉和侵权任务共用。"""

import json
import sys
import threading
import time
from collections import OrderedDict
from concurrent.futures import ProcessPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

from bit import bit_db_api
from bit.bit_clash import switch_random_hongkong_node
from bit.bit_config import list_shop_configs, require_shop_config
from bit.bit_send_mail import send_info
from bit.bit_runtime_lock import InterProcessLock, get_lock_owner
from bit.bit_utils import get_bit_path, get_now_time


MERCADO_LOGIN_JOB_LOCK_KEY = "mercado_login_job"


LOGIN_URL_MARKERS = ("/login", "/lgz/", "/legacy-user")
LOGIN_TEXT_MARKERS = (
    "fill out your e-mail address to log in",
    "fill out your email address to log in",
    "log in to your account",
    "sign in to your account",
    "enter your e-mail address",
    "enter your email address",
    "ingresa tu e-mail",
    "ingrese su e-mail",
    "iniciar sesión",
    "iniciar sesion",
    "填写您的电子邮件地址以登录",
)
VERIFICATION_TEXT_MARKERS = (
    "enter the verification code",
    "enter verification code",
    "enter the code we sent",
    "enter the code sent",
    "enter the security code",
    "we sent you a code",
    "we've sent you a code",
    "we'll send you a code",
    "code sent to your email",
    "check your email for a code",
    "verify your identity",
    "confirm your identity",
    "verify it's you",
    "we will send you a code",
    "ingresa el código",
    "ingrese el código",
    "código de verificación",
    "codigo de verificacion",
    "código de seguridad",
    "verifica tu identidad",
    "confirma tu identidad",
    "digite o código",
    "código de verificação",
    "verifique sua identidade",
    "验证码",
    "输入代码",
)
CAPTCHA_TEXT_MARKERS = (
    "i'm not a robot",
    "i am not a robot",
    "im not a robot",
    "no soy un robot",
    "não sou um robô",
    "nao sou um robo",
    "人机验证",
    "我不是机器人",
)
MERCADO_RATE_LIMIT_MARKERS = (
    "429 too many",
    "http 429",
    "too many requests",
    "rate limit",
    "rate-limit",
    "request limit exceeded",
    "request limit",
    "access denied",
    "请求太过频繁",
    "请求过于频繁",
    "访问过于频繁",
    "操作太频繁",
    "每秒最多可以发起",
    "demasiadas solicitudes",
    "muitas solicitações",
    "hubo un error accediendo a esta página",
    "hubo un error accediendo a esta pagina",
)

EMAIL_INPUT_SELECTORS = (
    "input[type='email']",
    "#user_id",
    "input[name='email']:not([type='radio']):not([type='checkbox'])",
    "input[name='user_id']:not([type='radio']):not([type='checkbox'])",
    "input[autocomplete='username']",
    "input[placeholder*='mail' i]",
    "input[id*='email']:not([type='radio']):not([type='checkbox'])",
    "input[id*='user']:not([type='radio']):not([type='checkbox'])",
)
PASSWORD_INPUT_SELECTORS = (
    "input[type='password']",
    "input[name='password']",
    "input[autocomplete='current-password']",
    "#password",
)
PASSWORD_OPTION_SELECTORS = (
    "button[aria-labelledby='password_validation-content']",
    "button[aria-labelledby^='password_validation']",
    "li#password_validation > button",
    "#password_validation button",
)
CODE_INPUT_SELECTORS = (
    "input[autocomplete='one-time-code']",
    "input[inputmode='numeric'][maxlength='1']",
    "input[inputmode='numeric'][maxlength='6']",
    "input[data-testid*='code' i]",
    "input[aria-label*='code' i]",
    "input[aria-label*='digit' i]",
    "input[name*='code']",
    "input[id*='code']",
    "input[name*='otp']",
    "input[id*='otp']",
)
CAPTCHA_SELECTORS = (
    "iframe[src*='recaptcha']",
    "iframe[title*='recaptcha' i]",
    "iframe[src*='hcaptcha']",
    "iframe[title*='hcaptcha' i]",
    ".g-recaptcha",
    ".h-captcha",
    "[data-sitekey][class*='captcha' i]",
    "[data-sitekey][id*='captcha' i]",
)
CLICKABLE_SELECTORS = (
    "button",
    "a",
    "[role='button']",
    "input[type='button']",
    "input[type='submit']",
)
LOGIN_ENTRY_TEXT_MARKERS = (
    "log in",
    "sign in",
    "iniciar sesión",
    "iniciar sesion",
    "ingresar",
    "entrar",
    "登录",
)
LOGIN_ENTRY_NEGATIVE_MARKERS = (
    "sign up",
    "create account",
    "register",
    "registrar",
    "cadastre",
    "google",
    "facebook",
    "apple",
    "注册",
)
AUTH_METHOD_TEXT_MARKERS = (
    "choose a verification method",
    "choose how to log in",
    "select a verification method",
    "elige un método de verificación",
    "elige un metodo de verificacion",
    "escolha um método de verificação",
    "escolha um metodo de verificacao",
    "选择验证方式",
)
CONTINUE_BUTTON_EXACT_TEXTS = (
    "continue",
    "next",
    "continuar",
    "siguiente",
    "avançar",
    "avancar",
    "继续",
    "下一步",
)
CONFIRM_BUTTON_EXACT_TEXTS = (
    "confirm",
    "continue",
    "log in",
    "sign in",
    "iniciar sesión",
    "iniciar sesion",
    "ingresar",
    "entrar",
    "continuar",
    "确认",
    "登录",
)
PASSWORD_OPTION_TEXT_MARKERS = (
    "password",
    "use password",
    "log in with password",
    "sign in with password",
    "enter with password",
    "ingresar con contraseña",
    "iniciar sesión con contraseña",
    "usar contraseña",
    "entrar com senha",
    "usar senha",
    "使用密码",
    "密码登录",
    "用密码",
)
PASSWORD_OPTION_EXACT_TEXTS = (
    "password",
    "contraseña",
    "senha",
    "密码",
)
PASSWORD_OPTION_NEGATIVE_MARKERS = (
    "forgot",
    "reset",
    "change",
    "olvid",
    "recuper",
    "esquec",
    "忘记",
    "重置",
    "修改",
)
PASSWORD_INCORRECT_TEXT_MARKERS = (
    "incorrect password",
    "invalid password",
    "wrong password",
    "password is incorrect",
    "password isn't correct",
    "contraseña incorrecta",
    "contrasena incorrecta",
    "senha incorreta",
    "密码错误",
    "密码不正确",
    "密码不对",
)
PASSWORD_REQUIRED_TEXT_MARKERS = (
    "enter your password",
    "password is required",
    "please enter your password",
    "ingresa tu contraseña",
    "ingresa tu contrasena",
    "digite sua senha",
    "请输入密码",
    "密码不能为空",
)
EMAIL_REJECTED_TEXT_MARKERS = (
    "enter a valid email",
    "enter a valid e-mail",
    "email is not valid",
    "e-mail is not valid",
    "account not found",
    "we couldn't find",
    "we could not find",
    "correo no válido",
    "correo no valido",
    "cuenta no encontrada",
    "e-mail inválido",
    "e-mail invalido",
    "邮箱格式不正确",
    "找不到该账号",
    "账号不存在",
)
LOGIN_ALREADY_ACTIVE = "已登录"
LOGIN_SUCCESS = "登录成功"
LOGIN_NOT_LOGGED_IN = "未登录"
LOGIN_VERIFICATION_REQUIRED = "需要验证码"
LOGIN_CAPTCHA_REQUIRED = "需要人机验证"
LOGIN_FAILED = "登录状态检测失败"
LOGIN_EMAIL_MISSING = "数据库未配置邮箱"
LOGIN_EMAIL_REJECTED = "数据库邮箱未通过登录页校验"
LOGIN_SAVED_PASSWORD_MISSING = "浏览器未保存默认密码"
LOGIN_SAVED_PASSWORD_INCORRECT = "浏览器默认密码错误"
LOGIN_WINDOW_BUSY = "窗口正在被其他任务占用"
INITIAL_LOGIN_ACTIVE = "已登录"
INITIAL_LOGIN_INACTIVE = "未登录"
INITIAL_LOGIN_UNKNOWN = "未确认"
PROGRAM_LOGIN_NOT_REQUIRED = "无需登录"
PROGRAM_LOGIN_SUCCESS = "登录成功"
PROGRAM_LOGIN_FAILED = "登录失败"
PROGRAM_LOGIN_VERIFICATION_REQUIRED = "遇到验证码"
PROGRAM_LOGIN_CAPTCHA_REQUIRED = "遇到人机验证"
PROGRAM_LOGIN_NOT_RUN = "未执行"
LOGIN_OUTCOME_ALREADY_ACTIVE = "原本已登录"
LOGIN_OUTCOME_AUTO_LOGIN_SUCCESS = "未登录，程序登录成功"
LOGIN_OUTCOME_AUTO_LOGIN_FAILED = "未登录，程序登录失败"
LOGIN_OUTCOME_VERIFICATION_REQUIRED = "未登录，程序登录遇到验证码"
LOGIN_OUTCOME_CAPTCHA_REQUIRED = "未登录，程序登录遇到人机验证"
LOGIN_OUTCOME_NOT_DETERMINED = "未完成登录判断"
MERCADO_HOME_URL = "https://global-selling.mercadolibre.com/"
RATE_LIMIT_RETRY_WAIT_SECONDS = 30
LOGIN_EVENT_LOG_PATH = get_bit_path() / "logs" / "mercado_unlogged_shops.jsonl"
LOGIN_REPORT_DIR = get_bit_path() / "登录状态汇总"
_LOGIN_EVENT_LOG_GUARD = threading.Lock()
SAVED_PASSWORD_SUBMITTED_DETAIL = "已提交浏览器保存的默认密码"
SAVED_PASSWORD_SELECTION_ATTEMPTED_DETAIL = "已尝试选择浏览器保存的默认密码并提交"


def load_shop_login_config(shop_name, window_id="", config_path=None):
    """从数据库或数据库接口读取店铺窗口及登录邮箱。"""
    if config_path is not None:
        raise RuntimeError(
            "运行时不再读取比特配置 Excel；请先用 bit.bit_config 导入数据库"
        )
    record = require_shop_config(shop_name=shop_name, window_id=window_id)
    return {
        "shop_name": record["shop_name"] or str(shop_name or "").strip(),
        "window_id": record["window_id"] or str(window_id or "").strip(),
        "email": record["email"],
        "email_column_exists": True,
        "config_source": "database",
    }


def _page_snapshot(driver):
    try:
        driver.switch_to.default_content()
    except Exception:
        pass
    try:
        page_text = driver.execute_script(
            "return document.body ? document.body.innerText : '';"
        ) or ""
    except Exception:
        page_text = ""
    try:
        current_url = driver.current_url or ""
    except Exception:
        current_url = ""
    try:
        title = driver.title or ""
    except Exception:
        title = ""
    return {
        "page_text": str(page_text),
        "current_url": str(current_url),
        "title": str(title),
    }


def is_mercado_rate_limited_page(driver=None, state=None):
    """识别 Mercado 可见限频页；空白错误页才回退检查源码。"""
    state = dict(state or _page_snapshot(driver))
    visible_state = "\n".join(
        (
            str(state.get("page_text") or ""),
            str(state.get("title") or ""),
            str(state.get("current_url") or ""),
            str(state.get("navigation_error") or ""),
        )
    ).casefold()
    # 明确的人机验证仍按验证码流程处理，不能误当成限频后自动换 IP。
    if any(marker in visible_state for marker in CAPTCHA_TEXT_MARKERS):
        return False
    if any(marker in visible_state for marker in MERCADO_RATE_LIMIT_MARKERS):
        return True

    if str(state.get("page_text") or "").strip() or str(
        state.get("title") or ""
    ).strip():
        return False

    page_source = state.get("page_source")
    if page_source is None and driver is not None:
        try:
            page_source = driver.page_source or ""
        except Exception:
            page_source = ""
    source_text = str(page_source or "").casefold()
    return any(marker in source_text for marker in MERCADO_RATE_LIMIT_MARKERS)


def _is_visible_enabled(element):
    try:
        return element.is_displayed() and element.is_enabled()
    except Exception:
        return False


def _visible_elements_in_current_context(driver, selectors):
    elements = []
    seen = set()
    for selector in selectors:
        try:
            found = driver.find_elements(By.CSS_SELECTOR, selector)
        except Exception:
            continue
        for element in found:
            marker = getattr(element, "id", None) or id(element)
            if marker in seen:
                continue
            seen.add(marker)
            if _is_visible_enabled(element):
                elements.append(element)

    if elements:
        return elements

    # Selenium 的普通 find_elements 不会穿透 Web Component 的 Shadow DOM。
    try:
        shadow_elements = driver.execute_script(
            """
            const selectors = arguments[0];
            const results = [];
            const seen = new Set();
            function addMatches(root) {
                if (!root || !root.querySelectorAll) return;
                for (const selector of selectors) {
                    let nodes = [];
                    try { nodes = root.querySelectorAll(selector); } catch (_) {}
                    for (const node of nodes) {
                        if (!seen.has(node)) {
                            seen.add(node);
                            results.push(node);
                        }
                    }
                }
                let descendants = [];
                try { descendants = root.querySelectorAll('*'); } catch (_) {}
                for (const node of descendants) {
                    if (node.shadowRoot) addMatches(node.shadowRoot);
                }
            }
            addMatches(document);
            return results;
            """,
            list(selectors),
        ) or []
    except Exception:
        shadow_elements = []
    for element in shadow_elements:
        marker = getattr(element, "id", None) or id(element)
        if marker in seen:
            continue
        seen.add(marker)
        if _is_visible_enabled(element):
            elements.append(element)
    return elements


def _search_visible_elements_in_frames(driver, selectors, depth=0, max_depth=4):
    elements = _visible_elements_in_current_context(driver, selectors)
    if elements or depth >= max_depth:
        return elements

    try:
        frames = driver.find_elements(By.CSS_SELECTOR, "iframe, frame")
    except Exception:
        frames = []
    for frame in frames:
        try:
            if not frame.is_displayed():
                continue
            driver.switch_to.frame(frame)
        except Exception:
            continue
        elements = _search_visible_elements_in_frames(
            driver,
            selectors,
            depth=depth + 1,
            max_depth=max_depth,
        )
        if elements:
            # 保持在找到控件的 frame 中，后续输入、点击必须使用同一上下文。
            return elements
        try:
            driver.switch_to.parent_frame()
        except Exception:
            try:
                driver.switch_to.default_content()
            except Exception:
                pass
    return []


def _visible_elements(driver, selectors):
    # 登录探测需要测试许多候选选择器。若继承业务代码的 10 秒隐式等待，
    # 每个未命中的选择器都会单独等待，最终可能耗时数分钟。
    implicit_wait_changed = False
    try:
        driver.implicitly_wait(0)
        implicit_wait_changed = True
    except Exception:
        pass
    try:
        # 优先搜索当前 frame，用于判断登录控件所处的页面上下文。
        elements = _visible_elements_in_current_context(driver, selectors)
        if elements:
            return elements
        try:
            driver.switch_to.default_content()
        except Exception:
            pass
        elements = _search_visible_elements_in_frames(driver, selectors)
        if elements:
            return elements
        try:
            driver.switch_to.default_content()
        except Exception:
            pass
        return []
    finally:
        if implicit_wait_changed:
            try:
                driver.implicitly_wait(10)
            except Exception:
                pass


def _has_visible_captcha_widget(elements):
    """过滤页面里不可交互的小型反爬占位 iframe，避免误报人机验证。"""
    for element in elements or []:
        try:
            size = element.size or {}
            width = float(size.get("width") or 0)
            height = float(size.get("height") or 0)
        except Exception:
            width = 0
            height = 0
        if width >= 100 and height >= 40:
            return True
    return False


def is_mercado_login_page(driver):
    if not driver:
        return False
    state = _page_snapshot(driver)
    combined = "\n".join(state.values()).casefold()
    if _visible_elements(
        driver,
        EMAIL_INPUT_SELECTORS + PASSWORD_INPUT_SELECTORS + CODE_INPUT_SELECTORS,
    ):
        return True
    if _has_visible_captcha_widget(_visible_elements(driver, CAPTCHA_SELECTORS)):
        return True
    return any(
        marker in combined
        for marker in (
            LOGIN_URL_MARKERS
            + LOGIN_TEXT_MARKERS
            + VERIFICATION_TEXT_MARKERS
            + CAPTCHA_TEXT_MARKERS
        )
    )


def detect_login_stage(driver):
    """返回 rate_limited/captcha/email/password/verification/login/logged_in。"""
    # 先看页面级文案，避免“验证方式选择页”里的 Email 选项/隐藏单选框
    # 被误认成需要填写的 Email 输入框。
    state = _page_snapshot(driver)
    combined = "\n".join(state.values()).casefold()
    if is_mercado_rate_limited_page(driver=driver, state=state):
        return "rate_limited"
    if any(marker in combined for marker in CAPTCHA_TEXT_MARKERS):
        return "captcha"
    if _has_visible_captcha_widget(_visible_elements(driver, CAPTCHA_SELECTORS)):
        return "captcha"
    if any(marker in combined for marker in AUTH_METHOD_TEXT_MARKERS):
        return "login"

    if _visible_elements(driver, CODE_INPUT_SELECTORS):
        return "verification"
    if _visible_elements(driver, PASSWORD_INPUT_SELECTORS):
        return "password"
    if _visible_elements(driver, EMAIL_INPUT_SELECTORS):
        return "email"

    if any(marker in combined for marker in VERIFICATION_TEXT_MARKERS):
        return "verification"
    if any(marker in combined for marker in LOGIN_URL_MARKERS + LOGIN_TEXT_MARKERS):
        return "login"
    return "logged_in"


def _result(ok, status, message, **extra):
    return {"ok": bool(ok), "status": status, "message": message, **extra}


def _unlogged_program_login_result(ok, status, message, **extra):
    """记录已经确认未登录后的程序登录结果。"""
    if status == LOGIN_SUCCESS:
        program_login_result = PROGRAM_LOGIN_SUCCESS
        result_category = LOGIN_OUTCOME_AUTO_LOGIN_SUCCESS
    elif status == LOGIN_VERIFICATION_REQUIRED:
        program_login_result = PROGRAM_LOGIN_VERIFICATION_REQUIRED
        result_category = LOGIN_OUTCOME_VERIFICATION_REQUIRED
    elif status == LOGIN_CAPTCHA_REQUIRED:
        program_login_result = PROGRAM_LOGIN_CAPTCHA_REQUIRED
        result_category = LOGIN_OUTCOME_CAPTCHA_REQUIRED
    else:
        program_login_result = PROGRAM_LOGIN_FAILED
        result_category = LOGIN_OUTCOME_AUTO_LOGIN_FAILED
    return _result(
        ok,
        status,
        message,
        initial_login_status=INITIAL_LOGIN_INACTIVE,
        program_login_result=program_login_result,
        result_category=result_category,
        **extra,
    )


def _normalize_login_judgement(result):
    """为旧结果、异常结果和测试替身补齐统一的登录判断字段。"""
    normalized = dict(result or {})
    if all(
        normalized.get(field)
        for field in (
            "initial_login_status",
            "program_login_result",
            "result_category",
        )
    ):
        return normalized

    status = str(normalized.get("status") or LOGIN_FAILED)
    action = str(normalized.get("action") or "")
    if status == LOGIN_ALREADY_ACTIVE:
        judgement = (
            INITIAL_LOGIN_ACTIVE,
            PROGRAM_LOGIN_NOT_REQUIRED,
            LOGIN_OUTCOME_ALREADY_ACTIVE,
        )
    elif status == LOGIN_SUCCESS:
        judgement = (
            INITIAL_LOGIN_INACTIVE,
            PROGRAM_LOGIN_SUCCESS,
            LOGIN_OUTCOME_AUTO_LOGIN_SUCCESS,
        )
    elif status == LOGIN_VERIFICATION_REQUIRED:
        judgement = (
            INITIAL_LOGIN_INACTIVE,
            PROGRAM_LOGIN_VERIFICATION_REQUIRED,
            LOGIN_OUTCOME_VERIFICATION_REQUIRED,
        )
    elif status == LOGIN_CAPTCHA_REQUIRED:
        judgement = (
            INITIAL_LOGIN_INACTIVE,
            PROGRAM_LOGIN_CAPTCHA_REQUIRED,
            LOGIN_OUTCOME_CAPTCHA_REQUIRED,
        )
    elif action in ("未执行", "执行异常"):
        judgement = (
            INITIAL_LOGIN_UNKNOWN,
            PROGRAM_LOGIN_NOT_RUN,
            LOGIN_OUTCOME_NOT_DETERMINED,
        )
    else:
        judgement = (
            INITIAL_LOGIN_INACTIVE,
            PROGRAM_LOGIN_FAILED,
            LOGIN_OUTCOME_AUTO_LOGIN_FAILED,
        )

    normalized.setdefault("initial_login_status", judgement[0])
    normalized.setdefault("program_login_result", judgement[1])
    normalized.setdefault("result_category", judgement[2])
    return normalized


def _count_login_outcomes(results):
    categories = (
        LOGIN_OUTCOME_ALREADY_ACTIVE,
        LOGIN_OUTCOME_AUTO_LOGIN_SUCCESS,
        LOGIN_OUTCOME_AUTO_LOGIN_FAILED,
        LOGIN_OUTCOME_VERIFICATION_REQUIRED,
        LOGIN_OUTCOME_CAPTCHA_REQUIRED,
        LOGIN_OUTCOME_NOT_DETERMINED,
    )
    counts = {category: 0 for category in categories}
    for result in results or []:
        category = _normalize_login_judgement(result).get("result_category")
        if category not in counts:
            counts[category] = 0
        counts[category] += 1
    return counts


def sync_login_results_to_window_anomalies(results):
    """把登录结果同步到店铺状态；成功店铺自动解除待登录状态。"""
    summary = {
        "anomaly_count": 0,
        "resolved_count": 0,
        "skipped_count": 0,
        "error_count": 0,
        "errors": [],
    }
    for raw_result in results or []:
        result = _normalize_login_judgement(raw_result)
        window_id = str(result.get("window_id") or "").strip()
        shop_name = str(result.get("shop_name") or "").strip()
        if not window_id:
            summary["skipped_count"] += 1
            continue
        try:
            if result.get("result_category") in (
                LOGIN_OUTCOME_ALREADY_ACTIVE,
                LOGIN_OUTCOME_AUTO_LOGIN_SUCCESS,
            ) or result.get("ok"):
                bit_db_api.resolve_window_anomaly(window_id)
                summary["resolved_count"] += 1
                continue

            status = str(result.get("status") or LOGIN_FAILED).strip()
            stage = str(result.get("login_stage") or "").strip()
            anomaly_type = "美客多限频" if stage == "rate_limited" else status
            reason = str(result.get("message") or "登录任务未完成").strip()
            bit_db_api.upsert_window_anomaly(
                window_id,
                shop_name,
                str(result.get("sites") or result.get("site") or "").strip(),
                anomaly_type=anomaly_type,
                reason=reason,
                source="bit_mercado_login",
            )
            summary["anomaly_count"] += 1
        except Exception as exc:
            summary["error_count"] += 1
            summary["errors"].append(
                {
                    "shop_name": shop_name,
                    "window_id": window_id,
                    "error": str(exc),
                }
            )
            print(
                f"{get_now_time()} {shop_name or window_id} 同步店铺状态失败：{exc}",
                flush=True,
            )
    print(
        f"{get_now_time()} bit_mercado_login 店铺状态同步完成："
        f"异常 {summary['anomaly_count']} 家，解除 {summary['resolved_count']} 家，"
        f"跳过 {summary['skipped_count']} 家，失败 {summary['error_count']} 家",
        flush=True,
    )
    return summary


def ensure_mercado_login(
    driver,
    shop_name,
    window_id="",
    config_path=None,
    default_password=None,
    wait_seconds=20,
    alert_sender=None,
):
    """兼容旧调用名：仅检测登录状态，绝不输入账号、密码或提交表单。"""
    del window_id, config_path, default_password, wait_seconds, alert_sender
    if not is_mercado_login_page(driver):
        return _result(True, LOGIN_ALREADY_ACTIVE, LOGIN_ALREADY_ACTIVE)

    stage = detect_login_stage(driver)
    print(
        f"{get_now_time()} {shop_name} 登录页面识别阶段：{stage}",
        flush=True,
    )
    stage_labels = {
        "captcha": LOGIN_CAPTCHA_REQUIRED,
        "verification": LOGIN_VERIFICATION_REQUIRED,
        "email": "邮箱登录页",
        "password": "密码登录页",
        "login": "登录页",
    }
    detail = stage_labels.get(stage, stage)
    return _result(
        False,
        LOGIN_NOT_LOGGED_IN,
        f"{shop_name} 未登录（{detail}），不执行任何自动登录操作",
        login_stage=stage,
    )


def ensure_mercado_login_from_home(
    driver,
    shop_name,
    window_id="",
    config_path=None,
    default_password=None,
    wait_seconds=20,
    navigation_wait_seconds=5,
    alert_sender=None,
    max_rate_limit_retries=2,
    rate_limit_retry_wait_seconds=RATE_LIMIT_RETRY_WAIT_SECONDS,
):
    """访问首页；每次遇到限频都切换香港节点，等待后最多重试两次。"""
    max_rate_limit_retries = max(0, int(max_rate_limit_retries))
    rate_limit_retry_count = 0
    rate_limit_detected = False
    node_switch_result = {}
    node_switch_results = []

    while True:
        print(
            f"{get_now_time()} {shop_name} 正在访问美客多首页检测登录状态："
            f"{MERCADO_HOME_URL}<br>",
            flush=True,
        )
        navigation_error = ""
        navigation_method = ""
        try:
            driver.switch_to.default_content()
        except Exception:
            pass

        # 附着到 BitBrowser 的 ChromeDriver 在 driver.get() 上可能长期等待页面资源。
        # CDP Page.navigate 只发出导航命令，不等待 load 事件，因此适合登录态探测。
        try:
            driver.execute_cdp_cmd("Page.navigate", {"url": MERCADO_HOME_URL})
            navigation_method = "cdp"
        except Exception as cdp_error:
            try:
                driver.execute_script(
                    "window.location.replace(arguments[0]);",
                    MERCADO_HOME_URL,
                )
                navigation_method = "javascript"
            except Exception as js_error:
                navigation_error = (
                    f"CDP导航失败：{cdp_error}；JS导航失败：{js_error}"
                )

        settle_seconds = max(0, float(navigation_wait_seconds))
        settle_deadline = time.monotonic() + settle_seconds
        while time.monotonic() < settle_deadline:
            remaining = max(0, int(settle_deadline - time.monotonic()))
            print(
                f"{get_now_time()} {shop_name} 首页导航已发出"
                f"（{navigation_method or '失败'}），等待页面稳定，剩余 {remaining} 秒",
                flush=True,
            )
            time.sleep(min(1, max(0, settle_deadline - time.monotonic())))

        # 不调用 Page.stopLoading：部分 BitBrowser/ChromeDriver 组合会在该 CDP
        # 命令上无限等待。登录探测使用 0 秒隐式等待，无需停止页面加载。
        print(
            f"{get_now_time()} {shop_name} 首页稳定等待结束，开始扫描登录控件",
            flush=True,
        )

        state = _page_snapshot(driver)
        current_url = str(state.get("current_url") or "").casefold()
        reached_mercado = "mercadolibre.com" in current_url
        rate_limited = is_mercado_rate_limited_page(
            driver=driver,
            state={**state, "navigation_error": navigation_error},
        )
        login_detected = False if rate_limited else is_mercado_login_page(driver)

        if not rate_limited:
            break

        rate_limit_detected = True
        if rate_limit_retry_count >= max_rate_limit_retries:
            switch_reason = str(node_switch_result.get("reason") or "未执行")
            result = _result(
                False,
                LOGIN_FAILED,
                f"{shop_name} 检测到美客多限频，切换节点后重试 "
                f"{rate_limit_retry_count} 次仍未恢复（节点切换：{switch_reason}）",
                login_stage="rate_limited",
                action="限频重试失败",
                rate_limited=True,
                rate_limit_retry_count=rate_limit_retry_count,
                node_switch_result=node_switch_result,
                node_switch_results=node_switch_results,
            )
            return {
                **result,
                "login_detected_before": login_detected,
                "navigation_error": navigation_error,
                "navigation_method": navigation_method,
                "login_check_url": MERCADO_HOME_URL,
            }

        print(
            f"{get_now_time()} {shop_name} 检测到美客多限频，正在切换香港节点",
            flush=True,
        )
        try:
            node_switch_result = switch_random_hongkong_node() or {}
        except Exception as exc:
            node_switch_result = {
                "switched": False,
                "reason": "exception",
                "error": str(exc),
            }
        node_switch_results.append(dict(node_switch_result))

        rate_limit_retry_count += 1
        print(
            f"{get_now_time()} {shop_name} 限频后准备第 "
            f"{rate_limit_retry_count}/{max_rate_limit_retries} 次重试，"
            f"等待 {max(0, int(rate_limit_retry_wait_seconds))} 秒",
            flush=True,
        )
        time.sleep(max(0, int(rate_limit_retry_wait_seconds)))

    if navigation_error and not reached_mercado:
        result = _result(
            False,
            LOGIN_FAILED,
            f"无法打开美客多首页进行登录检测：{navigation_error}",
        )
    elif not reached_mercado:
        result = _result(
            False,
            LOGIN_FAILED,
            f"首页导航后未进入 Mercado 页面：{state.get('current_url') or '空地址'}",
        )
    elif login_detected:
        result = ensure_mercado_login(
            driver,
            shop_name,
            window_id=window_id,
            config_path=config_path,
            default_password=default_password,
            wait_seconds=wait_seconds,
            alert_sender=alert_sender,
        )
    else:
        result = _result(True, LOGIN_ALREADY_ACTIVE, LOGIN_ALREADY_ACTIVE)

    if rate_limit_retry_count:
        result = dict(result)
        result["message"] = (
            f"{result.get('message') or ''}；限频后已重试 "
            f"{rate_limit_retry_count} 次"
        ).strip("；")

    return {
        **result,
        "login_detected_before": login_detected,
        "navigation_error": navigation_error,
        "navigation_method": navigation_method,
        "login_check_url": MERCADO_HOME_URL,
        "rate_limited": rate_limit_detected,
        "rate_limit_retry_count": rate_limit_retry_count,
        "node_switch_result": node_switch_result,
        "node_switch_results": node_switch_results,
    }


def _element_search_text(element):
    values = []
    try:
        values.append(element.text or "")
    except Exception:
        pass
    # Mercado 的认证方式卡片是外层 button + 内层文字节点，WebElement.text
    # 在部分 BitBrowser 内核上为空，因此同时读取后代 innerText/textContent。
    for attribute in ("aria-label", "title", "value", "innerText", "textContent"):
        try:
            values.append(element.get_attribute(attribute) or "")
        except Exception:
            pass
    normalized_values = []
    seen = set()
    for value in values:
        normalized = " ".join(str(value).split()).casefold()
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        normalized_values.append(normalized)
    return " ".join(normalized_values)


def _first_visible_element(driver, selectors):
    elements = _visible_elements(driver, selectors)
    return elements[0] if elements else None


def _click_matching_control(
    driver,
    *,
    exact_texts=(),
    text_markers=(),
    negative_markers=(),
):
    """点击文案明确的按钮；普通 click 失败时回退到页面内 click。"""
    exact_texts = {str(value).casefold() for value in exact_texts}
    text_markers = tuple(str(value).casefold() for value in text_markers)
    negative_markers = tuple(str(value).casefold() for value in negative_markers)
    for element in _visible_elements(driver, CLICKABLE_SELECTORS):
        text = " ".join(_element_search_text(element).split())
        if not text or any(marker in text for marker in negative_markers):
            continue
        if text not in exact_texts and not any(marker in text for marker in text_markers):
            continue
        try:
            element.click()
            return True
        except Exception:
            try:
                driver.execute_script("arguments[0].click();", element)
                return True
            except Exception:
                continue
    return False


def _click_login_entry_button(driver, shop_name=""):
    clicked = _click_matching_control(
        driver,
        text_markers=LOGIN_ENTRY_TEXT_MARKERS,
        negative_markers=LOGIN_ENTRY_NEGATIVE_MARKERS
        + PASSWORD_OPTION_NEGATIVE_MARKERS,
    )
    if clicked:
        print(f"{get_now_time()} {shop_name} 已点击登录入口", flush=True)
    return clicked


def _click_continue_button(driver, shop_name=""):
    clicked = _click_matching_control(
        driver,
        exact_texts=CONTINUE_BUTTON_EXACT_TEXTS,
        negative_markers=LOGIN_ENTRY_NEGATIVE_MARKERS,
    )
    if clicked:
        print(f"{get_now_time()} {shop_name} 已点击 Continue", flush=True)
    return clicked


def _click_confirm_button(driver, shop_name=""):
    clicked = _click_matching_control(
        driver,
        exact_texts=CONFIRM_BUTTON_EXACT_TEXTS,
        negative_markers=PASSWORD_OPTION_NEGATIVE_MARKERS
        + LOGIN_ENTRY_NEGATIVE_MARKERS,
    )
    if clicked:
        print(f"{get_now_time()} {shop_name} 已点击 Confirm 登录", flush=True)
    return clicked


def _click_password_login_option(driver, shop_name=""):
    """只点击明确标为“密码”的登录方式，不点击发送验证码入口。"""
    # Mercado 当前验证方式页的卡片按钮本身无文字，通过 aria-labelledby
    # 指向相邻内容；先使用稳定的 password_validation 语义标识定位。
    for element in _visible_elements(driver, PASSWORD_OPTION_SELECTORS):
        try:
            element.click()
        except Exception:
            try:
                driver.execute_script("arguments[0].click();", element)
            except Exception:
                continue
        print(
            f"{get_now_time()} {shop_name} 已选择密码登录方式",
            flush=True,
        )
        return True

    clicked = _click_matching_control(
        driver,
        exact_texts=PASSWORD_OPTION_EXACT_TEXTS,
        text_markers=PASSWORD_OPTION_TEXT_MARKERS,
        negative_markers=PASSWORD_OPTION_NEGATIVE_MARKERS,
    )
    if clicked:
        print(
            f"{get_now_time()} {shop_name} 已选择密码登录方式",
            flush=True,
        )
    return clicked


def _wait_for_stage_transition(driver, previous_stage, timeout=30):
    """等待页面离开上一阶段；logged_in 连续出现两次才认为页面稳定。"""
    deadline = time.monotonic() + max(0, float(timeout))
    last_stage = previous_stage
    logged_in_hits = 0
    while time.monotonic() <= deadline:
        try:
            stage = detect_login_stage(driver)
        except Exception:
            stage = last_stage
        last_stage = stage
        if stage == "logged_in":
            logged_in_hits += 1
            if logged_in_hits >= 2:
                return stage
        else:
            logged_in_hits = 0
            if stage != previous_stage:
                return stage
        if time.monotonic() >= deadline:
            break
        time.sleep(min(1, max(0, deadline - time.monotonic())))
    return last_stage


def _fill_email_and_continue(driver, email, shop_name=""):
    email_input = _first_visible_element(driver, EMAIL_INPUT_SELECTORS)
    if email_input is None:
        return False
    try:
        email_input.click()
    except Exception:
        pass
    accelerators = (
        [Keys.COMMAND, Keys.CONTROL]
        if sys.platform == "darwin"
        else [Keys.CONTROL]
    )
    value_readable = False
    actual_value = ""
    for accelerator in accelerators:
        try:
            email_input.clear()
        except Exception:
            pass
        try:
            email_input.send_keys(accelerator, "a")
            email_input.send_keys(Keys.BACKSPACE)
        except Exception:
            pass
        email_input.send_keys(email)
        try:
            actual_value = str(email_input.get_attribute("value") or "").strip()
            value_readable = True
        except Exception:
            value_readable = False
            break
        if actual_value == email:
            break

    if value_readable and actual_value != email:
        # React 受控输入或浏览器自动填充可能拦截键盘清空；使用原生 value
        # setter 并派发 input/change 事件，使 React 状态与 DOM 保持一致。
        try:
            actual_value = str(
                driver.execute_script(
                    """
                    const input = arguments[0];
                    const value = arguments[1];
                    const setter = Object.getOwnPropertyDescriptor(
                        window.HTMLInputElement.prototype, 'value'
                    ).set;
                    setter.call(input, value);
                    input.dispatchEvent(new Event('input', {bubbles: true}));
                    input.dispatchEvent(new Event('change', {bubbles: true}));
                    return input.value;
                    """,
                    email_input,
                    email,
                )
                or ""
            ).strip()
        except Exception:
            actual_value = ""
        if actual_value != email:
            raise RuntimeError(
                "Email 输入框未能清除默认账号，已停止提交，避免登录错误账号"
            )

    if not _click_continue_button(driver, shop_name=shop_name):
        email_input.send_keys(Keys.ENTER)
    return True


def _classify_email_page_failure(driver):
    state = _page_snapshot(driver)
    combined = "\n".join(state.values()).casefold()
    if any(marker in combined for marker in EMAIL_REJECTED_TEXT_MARKERS):
        return LOGIN_EMAIL_REJECTED
    return LOGIN_FAILED


def _password_input_has_saved_value(driver, password_input):
    """只让浏览器返回是否有值，绝不把保存的密码内容读入 Python。"""
    try:
        return bool(
            driver.execute_script(
                "return Boolean(arguments[0] && arguments[0].value);",
                password_input,
            )
        )
    except Exception:
        return False


def _classify_password_page_failure(driver, saved_password_detected=False):
    """区分浏览器无保存密码与已保存密码错误，不读取密码内容。"""
    state = _page_snapshot(driver)
    combined = "\n".join(state.values()).casefold()
    if any(marker in combined for marker in PASSWORD_INCORRECT_TEXT_MARKERS):
        return LOGIN_SAVED_PASSWORD_INCORRECT
    if not saved_password_detected and any(
        marker in combined for marker in PASSWORD_REQUIRED_TEXT_MARKERS
    ):
        return LOGIN_SAVED_PASSWORD_MISSING

    password_input = _first_visible_element(driver, PASSWORD_INPUT_SELECTORS)
    has_value = bool(
        password_input is not None
        and _password_input_has_saved_value(driver, password_input)
    )
    if not saved_password_detected and not has_value:
        return LOGIN_SAVED_PASSWORD_MISSING
    return LOGIN_FAILED


def _submit_browser_saved_password(driver, wait_seconds=15, shop_name=""):
    """选择浏览器密码建议并提交；不读取、不保存、不生成密码。"""
    if is_mercado_rate_limited_page(driver):
        return False, "密码页面遇到限频", "rate_limited"

    password_input = _first_visible_element(driver, PASSWORD_INPUT_SELECTORS)
    if password_input is None:
        if is_mercado_rate_limited_page(driver):
            return False, "密码页面遇到限频", "rate_limited"
        return False, "未找到密码输入框", "password"

    has_saved_password = _password_input_has_saved_value(driver, password_input)
    if not has_saved_password:
        try:
            password_input.click()
            # Chrome/BitBrowser 的密码建议为浏览器原生弹层，DOM 无法直接访问。
            # 用键盘选择第一条（默认）保存密码。浏览器可能不允许 JS 读取自动
            # 填充值，因此不能只按 value 是否可读来断定“没有保存密码”。
            time.sleep(0.5)
            password_input.send_keys(Keys.ARROW_DOWN)
            password_input.send_keys(Keys.ENTER)
        except Exception:
            pass

        fill_deadline = time.monotonic() + min(max(1, float(wait_seconds)), 5)
        while time.monotonic() <= fill_deadline:
            current_stage = detect_login_stage(driver)
            if current_stage == "rate_limited":
                return False, "密码页面遇到限频", current_stage
            if current_stage != "password":
                return True, "密码页面已提交", current_stage
            password_input = _first_visible_element(driver, PASSWORD_INPUT_SELECTORS)
            if password_input is not None and _password_input_has_saved_value(
                driver, password_input
            ):
                has_saved_password = True
                break
            if time.monotonic() >= fill_deadline:
                break
            time.sleep(min(0.5, max(0, fill_deadline - time.monotonic())))

    if password_input is None:
        return False, "未找到密码输入框", "password"

    # 即使 JS 看不到密码值，也尝试点击 Confirm；最终以页面是否离开密码页判断。
    if not _click_confirm_button(driver, shop_name=shop_name):
        try:
            password_input.send_keys(Keys.ENTER)
        except Exception as exc:
            return False, f"提交浏览器保存密码失败：{exc}", "password"
    detail = (
        SAVED_PASSWORD_SUBMITTED_DETAIL
        if has_saved_password
        else SAVED_PASSWORD_SELECTION_ATTEMPTED_DETAIL
    )
    return True, detail, "password"


def login_mercado_with_saved_password(
    driver,
    shop_name,
    window_id="",
    email=None,
    wait_seconds=60,
    navigation_wait_seconds=5,
):
    """独立自动登录流程；普通申诉/声誉任务仍只调用检测函数。"""
    shop_name = str(shop_name or "").strip()
    window_id = str(window_id or "").strip()
    if email is None:
        config = load_shop_login_config(shop_name, window_id=window_id)
        window_id = window_id or config["window_id"]
        email = config.get("email")
    email = str(email or "").strip()

    initial = ensure_mercado_login_from_home(
        driver,
        shop_name,
        window_id=window_id,
        wait_seconds=wait_seconds,
        navigation_wait_seconds=navigation_wait_seconds,
    )
    if initial.get("ok"):
        return _result(
            True,
            LOGIN_ALREADY_ACTIVE,
            f"{shop_name} 已处于登录状态，无需登录",
            initial_login_status=INITIAL_LOGIN_ACTIVE,
            program_login_result=PROGRAM_LOGIN_NOT_REQUIRED,
            result_category=LOGIN_OUTCOME_ALREADY_ACTIVE,
            login_stage="logged_in",
            action="无需登录",
        )
    if initial.get("status") == LOGIN_FAILED:
        initial = dict(initial)
        initial.update(
            initial_login_status=INITIAL_LOGIN_UNKNOWN,
            program_login_result=PROGRAM_LOGIN_NOT_RUN,
            result_category=LOGIN_OUTCOME_NOT_DETERMINED,
        )
        return initial

    stage = str(initial.get("login_stage") or detect_login_stage(driver))
    if stage == "captcha":
        return _unlogged_program_login_result(
            False,
            LOGIN_CAPTCHA_REQUIRED,
            f"{shop_name} 出现人机验证，需要人工处理",
            login_stage=stage,
            action="未登录",
        )

    # 部分账号先落在只有“Log in / Iniciar sesión”的入口页，必须先点击入口，
    # 才会渲染 Email 输入框。旧逻辑在这里直接停留为 login 阶段。
    if stage == "login" and _click_login_entry_button(driver, shop_name):
        stage = _wait_for_stage_transition(
            driver,
            "login",
            timeout=min(max(1, int(wait_seconds)), 20),
        )

    email_submitted = False
    if stage == "email":
        if not email:
            return _unlogged_program_login_result(
                False,
                LOGIN_EMAIL_MISSING,
                f"{shop_name} 数据库配置没有邮箱，无法自动登录",
                login_stage=stage,
                action="未登录",
            )
        try:
            submitted = _fill_email_and_continue(
                driver,
                email,
                shop_name=shop_name,
            )
        except Exception as exc:
            return _unlogged_program_login_result(
                False,
                LOGIN_FAILED,
                f"{shop_name} 输入数据库邮箱失败：{exc}",
                login_stage=stage,
                action="未登录",
            )
        if not submitted:
            return _unlogged_program_login_result(
                False,
                LOGIN_FAILED,
                f"{shop_name} 未找到可输入的 Email 控件",
                login_stage=stage,
                action="未登录",
            )
        email_submitted = True
        print(f"{get_now_time()} {shop_name} 已输入数据库邮箱并继续", flush=True)
        stage = _wait_for_stage_transition(driver, "email", timeout=wait_seconds)
        if stage == "email":
            print(
                f"{get_now_time()} {shop_name} 邮箱首次提交后页面未继续，正在重试一次",
                flush=True,
            )
            try:
                _fill_email_and_continue(driver, email, shop_name=shop_name)
            except Exception as exc:
                return _unlogged_program_login_result(
                    False,
                    LOGIN_FAILED,
                    f"{shop_name} 重试提交数据库邮箱失败：{exc}",
                    login_stage=stage,
                    action="自动登录失败",
                )
            stage = _wait_for_stage_transition(
                driver,
                "email",
                timeout=min(max(1, int(wait_seconds)), 20),
            )

    if stage == "email":
        email_status = _classify_email_page_failure(driver)
        message = (
            f"{shop_name} 数据库邮箱未通过登录页校验"
            if email_status == LOGIN_EMAIL_REJECTED
            else f"{shop_name} 邮箱重复提交后页面仍未继续"
        )
        return _unlogged_program_login_result(
            False,
            email_status,
            message,
            login_stage=stage,
            action="自动登录失败",
        )

    if stage == "logged_in":
        return _unlogged_program_login_result(
            True,
            LOGIN_SUCCESS,
            (
                f"{shop_name} 邮箱提交后已登录"
                if email_submitted
                else f"{shop_name} 点击登录入口后已登录"
            ),
            login_stage=stage,
            action="自动登录",
        )

    # 认证方式选择页可能会被通用检测识别为 login 或 verification。
    # 优先选择明确的密码方式；没有密码入口时，验证码页面留给人工处理。
    if stage in ("login", "verification") and _click_password_login_option(
        driver, shop_name
    ):
        stage = _wait_for_stage_transition(driver, stage, timeout=wait_seconds)

    if stage == "captcha":
        return _unlogged_program_login_result(
            False,
            LOGIN_CAPTCHA_REQUIRED,
            f"{shop_name} 出现人机验证，需要人工处理",
            login_stage=stage,
            action="未登录",
        )
    if stage == "verification":
        return _unlogged_program_login_result(
            False,
            LOGIN_VERIFICATION_REQUIRED,
            f"{shop_name} 需要验证码，未继续自动操作",
            login_stage=stage,
            action="未登录",
        )
    if stage != "password":
        return _unlogged_program_login_result(
            False,
            LOGIN_FAILED,
            f"{shop_name} 登录页面停留在无法处理的阶段：{stage}",
            login_stage=stage,
            action="未登录",
        )

    submitted, detail, observed_stage = _submit_browser_saved_password(
        driver,
        wait_seconds=min(max(1, int(wait_seconds)), 15),
        shop_name=shop_name,
    )
    if not submitted:
        status = (
            LOGIN_SAVED_PASSWORD_MISSING
            if detail == LOGIN_SAVED_PASSWORD_MISSING
            else LOGIN_FAILED
        )
        return _unlogged_program_login_result(
            False,
            status,
            f"{shop_name} {detail}",
            login_stage=observed_stage,
            action="未登录",
        )

    final_stage = observed_stage
    if final_stage == "password":
        final_stage = _wait_for_stage_transition(
            driver,
            "password",
            timeout=min(max(1, int(wait_seconds)), 20),
        )
    if final_stage == "captcha":
        return _unlogged_program_login_result(
            False,
            LOGIN_CAPTCHA_REQUIRED,
            f"{shop_name} 提交密码后出现人机验证，需要人工处理",
            login_stage=final_stage,
            action="自动登录未完成",
        )
    if final_stage == "verification":
        return _unlogged_program_login_result(
            False,
            LOGIN_VERIFICATION_REQUIRED,
            f"{shop_name} 提交密码后需要验证码",
            login_stage=final_stage,
            action="自动登录未完成",
        )
    if final_stage == "password":
        password_status = _classify_password_page_failure(
            driver,
            saved_password_detected=(detail == SAVED_PASSWORD_SUBMITTED_DETAIL),
        )
        if password_status == LOGIN_SAVED_PASSWORD_INCORRECT:
            message = f"{shop_name} 已提交浏览器默认密码，但页面提示密码错误"
        elif password_status == LOGIN_SAVED_PASSWORD_MISSING:
            message = f"{shop_name} 浏览器未保存可用的默认密码"
        else:
            message = f"{shop_name} 提交默认密码后仍停留在密码页面，原因未识别"
        return _unlogged_program_login_result(
            False,
            password_status,
            message,
            login_stage=final_stage,
            action="自动登录失败",
        )

    verification = ensure_mercado_login_from_home(
        driver,
        shop_name,
        window_id=window_id,
        wait_seconds=wait_seconds,
        navigation_wait_seconds=min(3, max(0, navigation_wait_seconds)),
    )
    if verification.get("ok"):
        return _unlogged_program_login_result(
            True,
            LOGIN_SUCCESS,
            f"{shop_name} 已使用浏览器保存的默认密码登录成功",
            login_stage="logged_in",
            action="自动登录",
        )

    verified_stage = str(
        verification.get("login_stage") or detect_login_stage(driver)
    )
    if verified_stage == "captcha":
        status = LOGIN_CAPTCHA_REQUIRED
    elif verified_stage == "verification":
        status = LOGIN_VERIFICATION_REQUIRED
    else:
        status = LOGIN_FAILED
    return _unlogged_program_login_result(
        False,
        status,
        f"{shop_name} 提交浏览器保存密码后仍未登录："
        f"{verification.get('message') or verified_stage}",
        login_stage=verified_stage,
        action="自动登录未完成",
    )


def is_login_blocking_result(value):
    text = str(value or "")
    return any(
        marker in text
        for marker in (
            "未登录",
            LOGIN_VERIFICATION_REQUIRED,
            LOGIN_CAPTCHA_REQUIRED,
            LOGIN_FAILED,
        )
    )


def unlogged_entries_from_task_records(records):
    """从任务记录五元组中提取未登录店铺，按店铺去重并合并站点。"""
    entries = {}
    for record in records or []:
        if not isinstance(record, (list, tuple)) or len(record) < 4:
            continue
        status = str(record[3] or "")
        if "未登录" not in status and "登录失效" not in status:
            continue
        shop_name = str(record[1] or "").strip()
        site = str(record[2] or "").strip()
        if not shop_name:
            continue
        entry = entries.setdefault(
            shop_name,
            {"shop_name": shop_name, "sites": set(), "details": set()},
        )
        if site:
            entry["sites"].add(site)
        if status:
            entry["details"].add(status)
    return [
        {
            "shop_name": entry["shop_name"],
            "sites": sorted(entry["sites"]),
            "details": sorted(entry["details"]),
        }
        for entry in sorted(entries.values(), key=lambda item: item["shop_name"])
    ]


def _normalize_unlogged_entries(entries):
    normalized = {}
    for raw in entries or []:
        if isinstance(raw, str):
            raw = {"shop_name": raw}
        if not isinstance(raw, dict):
            continue
        shop_name = str(raw.get("shop_name") or raw.get("name") or "").strip()
        if not shop_name:
            continue
        entry = normalized.setdefault(
            shop_name,
            {"shop_name": shop_name, "sites": set(), "details": set()},
        )
        sites = raw.get("sites") or ([raw.get("site")] if raw.get("site") else [])
        details = raw.get("details") or ([raw.get("detail")] if raw.get("detail") else [])
        for site in sites:
            if str(site or "").strip():
                entry["sites"].add(str(site).strip())
        for detail in details:
            if str(detail or "").strip():
                entry["details"].add(str(detail).strip())
    return [
        {
            "shop_name": entry["shop_name"],
            "sites": sorted(entry["sites"]),
            "details": sorted(entry["details"]),
        }
        for entry in sorted(normalized.values(), key=lambda item: item["shop_name"])
    ]


def record_unlogged_shop_summary(task_name, entries, log_path=None):
    """把本批次未登录店铺写入 JSONL 日志，一批一行。"""
    normalized = _normalize_unlogged_entries(entries)
    if not normalized:
        return ""
    path = Path(log_path or LOGIN_EVENT_LOG_PATH)
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "task_name": str(task_name or "美客多任务"),
        "shop_count": len(normalized),
        "shops": normalized,
    }
    line = json.dumps(payload, ensure_ascii=False) + "\n"
    with _LOGIN_EVENT_LOG_GUARD:
        with path.open("a", encoding="utf-8") as file:
            file.write(line)
    print(f"{get_now_time()} 未登录店铺日志已写入：{path}<br>")
    return str(path)


def send_unlogged_shop_summary(task_name, entries, log_path=None):
    """所有店铺任务结束后发送一封未登录店铺汇总邮件。"""
    normalized = _normalize_unlogged_entries(entries)
    if not normalized:
        print(f"{get_now_time()} {task_name} 本批次没有未登录店铺，无需发送提醒<br>")
        return False

    path = record_unlogged_shop_summary(task_name, normalized, log_path=log_path)
    lines = [
        f"任务：{task_name}",
        f"结束时间：{get_now_time()}",
        f"未登录店铺数量：{len(normalized)}",
        "",
    ]
    for index, entry in enumerate(normalized, start=1):
        sites = "、".join(entry["sites"]) or "全部/未指定站点"
        details = "；".join(entry["details"]) or "检测到未登录状态"
        lines.append(f"{index}. {entry['shop_name']}｜站点：{sites}｜{details}")
    body = "\n".join(lines)
    sent = send_info(
        f"美客多未登录店铺提醒：{task_name}（{len(normalized)}家）",
        body,
        path,
        Path(path).name if path else "",
    )
    if sent is False:
        print(f"{get_now_time()} {task_name} 未登录店铺汇总邮件发送失败<br>")
        return False
    print(f"{get_now_time()} {task_name} 未登录店铺汇总邮件已发送<br>")
    return True


def _connect_to_open_bit_browser(window_id, page_load_timeout=20):
    import requests
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service

    from bit.bit_api import closeBrowser, openBrowser

    last_error = ""
    max_browser_attempts = 3
    for browser_attempt in range(1, max_browser_attempts + 1):
        try:
            response = openBrowser(window_id)
        except Exception as exc:
            response = None
            last_error = f"BitBrowser 打开接口异常：{exc}"
        data = response.get("data") if isinstance(response, dict) else None
        if response is None:
            pass
        elif isinstance(response, dict) and response.get("success") is False:
            last_error = str(
                response.get("msg") or f"打开比特浏览器失败：{response}"
            )
        elif not (
            isinstance(data, dict) and data.get("driver") and data.get("http")
        ):
            last_error = f"打开比特浏览器失败：{response}"
        else:
            debugger_address = str(data["http"]).strip()
            debugger_url = (
                debugger_address
                if debugger_address.startswith(("http://", "https://"))
                else f"http://{debugger_address}"
            )
            debugger_error = ""
            debugger_ready = False
            for attempt in range(1, 7):
                try:
                    debugger_response = requests.get(
                        f"{debugger_url.rstrip('/')}/json/version",
                        timeout=3,
                    )
                    debugger_response.raise_for_status()
                    debugger_ready = True
                    break
                except Exception as exc:
                    debugger_error = str(exc)
                    if attempt < 6:
                        time.sleep(2)
            if debugger_ready:
                try:
                    options = webdriver.ChromeOptions()
                    options.add_experimental_option(
                        "debuggerAddress", debugger_address
                    )
                    driver = webdriver.Chrome(
                        service=Service(data["driver"]),
                        options=options,
                    )
                    driver.implicitly_wait(10)
                    try:
                        driver.set_page_load_timeout(
                            max(1, int(page_load_timeout))
                        )
                    except Exception:
                        pass
                    return driver
                except Exception as exc:
                    last_error = f"WebDriver 附着失败：{exc}"
            else:
                last_error = (
                    f"BitBrowser 调试端口未就绪："
                    f"{debugger_error or debugger_address}"
                )

        if browser_attempt >= max_browser_attempts:
            break
        print(
            f"{get_now_time()} 窗口 {window_id} 第 "
            f"{browser_attempt}/{max_browser_attempts} 次打开未就绪：{last_error}；"
            "正在关闭残留启动后重试",
            flush=True,
        )
        for close_attempt in range(1, 7):
            try:
                close_response = closeBrowser(window_id)
            except Exception as exc:
                last_error = f"{last_error}；重试前关闭失败：{exc}"
                break
            if not (
                isinstance(close_response, dict)
                and close_response.get("success") is False
            ):
                break
            close_message = str(close_response.get("msg") or close_response)
            if "正在打开" not in close_message or close_attempt >= 6:
                last_error = f"{last_error}；重试前关闭失败：{close_message}"
                break
            time.sleep(5)
        time.sleep(3)

    raise RuntimeError(last_error or "BitBrowser 打开后无法附着 WebDriver")


def _finalize_shop_login_result(
    config,
    result,
    started_at,
    browser_closed,
    close_error,
    browser_opened=False,
):
    ended_at = datetime.now()
    result = _normalize_login_judgement(result)
    return {
        "config_index": int(config.get("config_index") or 0),
        "shop_name": str(config.get("shop_name") or "").strip(),
        "window_id": str(config.get("window_id") or "").strip(),
        "email": str(config.get("email") or "").strip(),
        "ok": bool(result.get("ok")),
        "status": str(result.get("status") or LOGIN_FAILED),
        "initial_login_status": str(
            result.get("initial_login_status") or INITIAL_LOGIN_UNKNOWN
        ),
        "program_login_result": str(
            result.get("program_login_result") or PROGRAM_LOGIN_NOT_RUN
        ),
        "result_category": str(
            result.get("result_category") or LOGIN_OUTCOME_NOT_DETERMINED
        ),
        "login_stage": str(result.get("login_stage") or ""),
        "action": str(result.get("action") or ""),
        "message": str(result.get("message") or ""),
        "started_at": started_at.strftime("%Y-%m-%d %H:%M:%S"),
        "ended_at": ended_at.strftime("%Y-%m-%d %H:%M:%S"),
        "duration_seconds": round((ended_at - started_at).total_seconds(), 1),
        "browser_opened": bool(browser_opened),
        "browser_closed": bool(browser_closed),
        "close_error": str(close_error or ""),
    }


def login_one_database_shop(config, wait_seconds=60, page_load_timeout=20):
    """打开一个数据库店铺、按需登录，最后关闭浏览器并返回可序列化结果。"""
    from bit.bit_api import closeBrowser
    from bit.bit_runtime_lock import create_window_lease

    config = dict(config or {})
    shop_name = str(config.get("shop_name") or "").strip()
    window_id = str(config.get("window_id") or "").strip()
    email = str(config.get("email") or "").strip()
    started_at = datetime.now()
    browser_opened = False
    browser_closed = False
    close_error = ""
    driver = None

    if not window_id:
        return _finalize_shop_login_result(
            config,
            _result(
                False,
                LOGIN_FAILED,
                f"{shop_name or '未命名店铺'} 数据库配置缺少窗口ID",
                action="未执行",
            ),
            started_at,
            browser_closed,
            close_error,
        )

    lease = create_window_lease(
        window_id,
        owner=f"mercado_batch_login:{shop_name}",
        shop_name=shop_name,
        task_type="mercado_batch_login",
    )
    if not lease.acquire(timeout=0):
        return _finalize_shop_login_result(
            config,
            _result(
                False,
                LOGIN_WINDOW_BUSY,
                f"{shop_name} 的比特浏览器窗口正在被其他任务占用",
                action="未执行",
            ),
            started_at,
            browser_closed,
            close_error,
        )

    try:
        print(f"{get_now_time()} {shop_name} 正在打开比特浏览器", flush=True)
        # 即使 WebDriver 附着失败，BitBrowser API 也可能已经打开窗口，必须尝试关闭。
        browser_opened = True
        driver = _connect_to_open_bit_browser(
            window_id,
            page_load_timeout=page_load_timeout,
        )
        result = login_mercado_with_saved_password(
            driver,
            shop_name,
            window_id=window_id,
            email=email,
            wait_seconds=wait_seconds,
        )
    except Exception as exc:
        print(
            f"{get_now_time()} {shop_name} 登录任务异常：{exc}",
            flush=True,
        )
        result = _result(
            False,
            LOGIN_FAILED,
            f"{shop_name} 登录任务异常：{exc}",
            action="执行异常",
        )
    finally:
        if browser_opened:
            try:
                for close_attempt in range(1, 7):
                    close_response = closeBrowser(window_id, lease=lease)
                    if not (
                        isinstance(close_response, dict)
                        and close_response.get("success") is False
                    ):
                        browser_closed = True
                        close_error = ""
                        break
                    close_error = str(close_response.get("msg") or close_response)
                    if "正在打开" not in close_error or close_attempt >= 6:
                        break
                    print(
                        f"{get_now_time()} {shop_name} 浏览器仍在打开中，"
                        f"5 秒后重试关闭（{close_attempt}/6）",
                        flush=True,
                    )
                    time.sleep(5)
            except Exception as exc:
                close_error = str(exc)
            print(
                f"{get_now_time()} {shop_name} "
                f"{'已关闭比特浏览器' if browser_closed else '关闭比特浏览器失败'}",
                flush=True,
            )
        service = getattr(driver, "service", None)
        if service is not None:
            try:
                service.stop()
            except Exception:
                pass
        lease.release()

    if close_error:
        result = dict(result)
        result["message"] = (
            f"{result.get('message') or ''}；关闭浏览器失败：{close_error}"
        ).strip("；")
    return _finalize_shop_login_result(
        config,
        result,
        started_at,
        browser_closed,
        close_error,
        browser_opened=browser_opened,
    )


def _login_config_group_worker(configs, wait_seconds, page_load_timeout):
    # 相同窗口 ID 的数据库记录在同一进程内串行执行，避免互抢同一个浏览器配置。
    results = []
    for index, config in enumerate(configs):
        if index:
            # 数据库中可能有多个店铺记录复用同一个窗口；关闭后立即重开会让
            # BitBrowser 返回“浏览器正在打开中”，留出短暂的状态稳定时间。
            time.sleep(3)
        results.append(
            login_one_database_shop(
                config,
                wait_seconds=wait_seconds,
                page_load_timeout=page_load_timeout,
            )
        )
    return results


def _config_order_key(config):
    sequence = str(config.get("sequence_no") or "").strip()
    try:
        sequence_key = (0, int(float(sequence)))
    except (TypeError, ValueError):
        sequence_key = (1, sequence.casefold())
    return sequence_key, str(config.get("shop_name") or "").casefold()


def _excel_safe_text(value):
    text = str(value or "")
    return f"'{text}" if text.startswith(("=", "+", "-", "@")) else text


def write_login_status_report(results, output_path=None):
    """生成包含汇总公式和登录明细的 Excel。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    rows = [_normalize_login_judgement(row) for row in (results or [])]
    if output_path is None:
        LOGIN_REPORT_DIR.mkdir(parents=True, exist_ok=True)
        output_path = LOGIN_REPORT_DIR / (
            f"美客多店铺登录状态汇总-{datetime.now().strftime('%Y-%m-%d-%H-%M-%S')}.xlsx"
        )
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    summary = workbook.active
    summary.title = "汇总"
    details = workbook.create_sheet("登录明细")

    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    success_fill = PatternFill("solid", fgColor="E2F0D9")
    warning_fill = PatternFill("solid", fgColor="FFF2CC")
    failure_fill = PatternFill("solid", fgColor="FCE4D6")
    thin_gray = Side(style="thin", color="D9E2F3")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    summary.merge_cells("A1:B1")
    summary["A1"] = "美客多店铺登录状态汇总"
    summary["A1"].fill = title_fill
    summary["A1"].font = Font(name="微软雅黑", color="FFFFFF", bold=True, size=15)
    summary["A1"].alignment = Alignment(horizontal="center", vertical="center")
    summary.row_dimensions[1].height = 28
    summary_rows = (
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("店铺总数", "=COUNTA('登录明细'!B:B)-1"),
        ("原本已登录", f'=COUNTIF(\'登录明细\'!G:G,"{LOGIN_OUTCOME_ALREADY_ACTIVE}")'),
        (
            "未登录，程序登录成功",
            f'=COUNTIF(\'登录明细\'!G:G,"{LOGIN_OUTCOME_AUTO_LOGIN_SUCCESS}")',
        ),
        (
            "未登录，程序登录失败",
            f'=COUNTIF(\'登录明细\'!G:G,"{LOGIN_OUTCOME_AUTO_LOGIN_FAILED}")',
        ),
        (
            "其中：浏览器未保存默认密码",
            f'=COUNTIF(\'登录明细\'!H:H,"{LOGIN_SAVED_PASSWORD_MISSING}")',
        ),
        (
            "其中：浏览器默认密码错误",
            f'=COUNTIF(\'登录明细\'!H:H,"{LOGIN_SAVED_PASSWORD_INCORRECT}")',
        ),
        (
            "未登录，程序登录遇到验证码",
            f'=COUNTIF(\'登录明细\'!G:G,"{LOGIN_OUTCOME_VERIFICATION_REQUIRED}")',
        ),
        (
            "未登录，程序登录遇到人机验证",
            f'=COUNTIF(\'登录明细\'!G:G,"{LOGIN_OUTCOME_CAPTCHA_REQUIRED}")',
        ),
        (
            "未完成登录判断",
            f'=COUNTIF(\'登录明细\'!G:G,"{LOGIN_OUTCOME_NOT_DETERMINED}")',
        ),
        ("浏览器关闭失败", '=COUNTIF(\'登录明细\'!N:N,"失败")'),
    )
    for row_index, (label, value) in enumerate(summary_rows, start=3):
        summary.cell(row=row_index, column=1, value=label)
        summary.cell(row=row_index, column=2, value=value)
        for cell in summary[row_index]:
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            cell.font = Font(name="微软雅黑")
        summary.cell(row=row_index, column=1).fill = header_fill
        summary.cell(row=row_index, column=1).font = Font(name="微软雅黑", bold=True)
    summary.column_dimensions["A"].width = 20
    summary.column_dimensions["B"].width = 30

    headers = (
        "序号",
        "店铺名",
        "窗口ID",
        "邮箱",
        "初始登录状态",
        "程序登录结果",
        "最终判断",
        "技术状态",
        "登录阶段",
        "是否成功",
        "执行动作",
        "说明",
        "开始时间",
        "浏览器关闭情况",
        "结束时间",
        "耗时（秒）",
    )
    details.append(headers)
    for cell in details[1]:
        cell.fill = title_fill
        cell.font = Font(name="微软雅黑", color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    details.row_dimensions[1].height = 25

    for index, result in enumerate(rows, start=1):
        ok = bool(result.get("ok"))
        status = str(result.get("status") or LOGIN_FAILED)
        browser_opened = bool(result.get("browser_opened"))
        browser_closed = bool(result.get("browser_closed"))
        close_status = (
            "是" if browser_closed else ("失败" if browser_opened else "未打开")
        )
        values = (
            index,
            _excel_safe_text(result.get("shop_name")),
            _excel_safe_text(result.get("window_id")),
            _excel_safe_text(result.get("email")),
            result.get("initial_login_status"),
            result.get("program_login_result"),
            result.get("result_category"),
            status,
            _excel_safe_text(result.get("login_stage")),
            "是" if ok else "否",
            _excel_safe_text(result.get("action")),
            _excel_safe_text(result.get("message")),
            _excel_safe_text(result.get("started_at")),
            close_status,
            _excel_safe_text(result.get("ended_at")),
            float(result.get("duration_seconds") or 0),
        )
        details.append(values)
        row_index = details.max_row
        if result.get("result_category") in (
            LOGIN_OUTCOME_ALREADY_ACTIVE,
            LOGIN_OUTCOME_AUTO_LOGIN_SUCCESS,
        ):
            row_fill = success_fill
        elif result.get("result_category") in (
            LOGIN_OUTCOME_VERIFICATION_REQUIRED,
            LOGIN_OUTCOME_CAPTCHA_REQUIRED,
        ):
            row_fill = warning_fill
        else:
            row_fill = failure_fill
        for cell in details[row_index]:
            cell.border = border
            cell.font = Font(name="微软雅黑")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill = row_fill

    details.freeze_panes = "A2"
    details.auto_filter.ref = f"A1:P{max(1, details.max_row)}"
    details.sheet_view.showGridLines = False
    widths = (8, 20, 36, 30, 14, 20, 32, 20, 14, 12, 18, 55, 21, 16, 21, 14)
    for column_index, width in enumerate(widths, start=1):
        details.column_dimensions[get_column_letter(column_index)].width = width

    workbook.save(output_path)
    workbook.close()
    return str(output_path)


def send_login_status_report(results, report_path):
    rows = [_normalize_login_judgement(row) for row in (results or [])]
    outcome_counts = _count_login_outcomes(rows)
    already_active_count = outcome_counts[LOGIN_OUTCOME_ALREADY_ACTIVE]
    auto_success_count = outcome_counts[LOGIN_OUTCOME_AUTO_LOGIN_SUCCESS]
    success_count = already_active_count + auto_success_count
    missing_password_count = sum(
        1 for row in rows if row.get("status") == LOGIN_SAVED_PASSWORD_MISSING
    )
    incorrect_password_count = sum(
        1 for row in rows if row.get("status") == LOGIN_SAVED_PASSWORD_INCORRECT
    )
    body = "\n".join(
        (
            "美客多数据库店铺登录检查已完成。",
            f"结束时间：{get_now_time()}",
            f"店铺总数：{len(rows)}",
            f"原本已登录：{already_active_count}",
            f"未登录，程序登录成功：{auto_success_count}",
            f"未登录，程序登录失败：{outcome_counts[LOGIN_OUTCOME_AUTO_LOGIN_FAILED]}",
            f"其中浏览器未保存默认密码：{missing_password_count}",
            f"其中浏览器默认密码错误：{incorrect_password_count}",
            "未登录，程序登录遇到验证码："
            f"{outcome_counts[LOGIN_OUTCOME_VERIFICATION_REQUIRED]}",
            "未登录，程序登录遇到人机验证："
            f"{outcome_counts[LOGIN_OUTCOME_CAPTCHA_REQUIRED]}",
            f"未完成登录判断：{outcome_counts[LOGIN_OUTCOME_NOT_DETERMINED]}",
            "",
            "详细情况见附件 Excel。",
        )
    )
    return send_info(
        f"美客多店铺登录状态汇总（{success_count}/{len(rows)}）",
        body,
        report_path,
        Path(report_path).name,
    )


def run_all_database_shop_logins(
    max_workers=3,
    wait_seconds=60,
    page_load_timeout=20,
    output_path=None,
    send_email=True,
):
    """并发处理数据库中所有未忽略店铺，生成 Excel，并在全部结束后发邮件。"""
    configs = sorted(
        (dict(row) for row in list_shop_configs(include_ignored=False)),
        key=_config_order_key,
    )
    for index, config in enumerate(configs, start=1):
        config["config_index"] = index

    grouped = OrderedDict()
    for config in configs:
        # 缺少窗口 ID 的记录使用独立键，仍会进入报告。
        group_key = str(config.get("window_id") or f"__missing_{config['config_index']}")
        grouped.setdefault(group_key, []).append(config)

    results = []
    worker_count = max(1, min(int(max_workers or 3), len(grouped))) if grouped else 1
    print(
        f"{get_now_time()} 开始检查 {len(configs)} 个未忽略店铺，"
        f"并发进程数：{worker_count}",
        flush=True,
    )
    if grouped:
        with ProcessPoolExecutor(max_workers=worker_count) as executor:
            future_map = {
                executor.submit(
                    _login_config_group_worker,
                    group,
                    max(1, int(wait_seconds)),
                    max(1, int(page_load_timeout)),
                ): group
                for group in grouped.values()
            }
            for future in as_completed(future_map):
                group = future_map[future]
                try:
                    results.extend(future.result())
                except Exception as exc:
                    for config in group:
                        results.append(
                            _finalize_shop_login_result(
                                config,
                                _result(
                                    False,
                                    LOGIN_FAILED,
                                    f"{config.get('shop_name') or ''} 子进程异常：{exc}",
                                    action="执行异常",
                                ),
                                datetime.now(),
                                False,
                                "",
                            )
                        )

    results = [_normalize_login_judgement(row) for row in results]
    results.sort(key=lambda row: int(row.get("config_index") or 0))
    window_anomaly_sync = sync_login_results_to_window_anomalies(results)
    outcome_counts = _count_login_outcomes(results)
    status_counts = {
        LOGIN_SAVED_PASSWORD_MISSING: sum(
            1 for row in results if row.get("status") == LOGIN_SAVED_PASSWORD_MISSING
        ),
        LOGIN_SAVED_PASSWORD_INCORRECT: sum(
            1 for row in results if row.get("status") == LOGIN_SAVED_PASSWORD_INCORRECT
        ),
    }
    success_count = (
        outcome_counts[LOGIN_OUTCOME_ALREADY_ACTIVE]
        + outcome_counts[LOGIN_OUTCOME_AUTO_LOGIN_SUCCESS]
    )
    report_path = write_login_status_report(results, output_path=output_path)
    email_sent = bool(send_login_status_report(results, report_path)) if send_email else False
    print(
        f"{get_now_time()} 登录任务完成：{len(results)} 家，"
        f"原本已登录 {outcome_counts[LOGIN_OUTCOME_ALREADY_ACTIVE]} 家，"
        f"程序登录成功 {outcome_counts[LOGIN_OUTCOME_AUTO_LOGIN_SUCCESS]} 家，"
        f"程序登录失败 {outcome_counts[LOGIN_OUTCOME_AUTO_LOGIN_FAILED]} 家，"
        f"无默认密码 {status_counts[LOGIN_SAVED_PASSWORD_MISSING]} 家，"
        f"默认密码错误 {status_counts[LOGIN_SAVED_PASSWORD_INCORRECT]} 家，"
        f"遇到验证码 {outcome_counts[LOGIN_OUTCOME_VERIFICATION_REQUIRED]} 家，"
        f"Excel：{report_path}，邮件：{'已发送' if email_sent else '未发送'}",
        flush=True,
    )
    return {
        "shop_count": len(results),
        "success_count": success_count,
        "outcome_counts": outcome_counts,
        "status_counts": status_counts,
        "results": results,
        "report_path": report_path,
        "email_sent": email_sent,
        "max_workers": worker_count,
        "window_anomaly_sync": window_anomaly_sync,
    }


def build_command_line_parser():
    """构建单店检测及全店自动登录的命令行参数。"""
    import argparse

    parser = argparse.ArgumentParser(
        description=(
            "指定 --shop 时检测单店登录状态，追加 --auto-login 可自动登录；"
            "指定 --all-active-login 时，并发处理数据库全部未忽略店铺、"
            "按需登录并发送 Excel 汇总。"
        )
    )
    target = parser.add_mutually_exclusive_group(required=True)
    target.add_argument(
        "--shop",
        help="数据库比特浏览器配置中的账号名，例如：龙凤呈祥",
    )
    target.add_argument(
        "--all-active-login",
        "--all",
        dest="all_active_login",
        action="store_true",
        help="处理数据库中全部未忽略店铺，并按需使用邮箱和浏览器保存密码登录",
    )
    parser.add_argument(
        "--wait-seconds",
        type=int,
        default=30,
        help="页面状态变化的最长等待秒数，默认 30",
    )
    parser.add_argument(
        "--page-load-timeout",
        type=int,
        default=20,
        help="打开美客多 Global Selling 首页的超时秒数，默认 20",
    )
    parser.add_argument(
        "--no-navigate",
        action="store_true",
        help="不打开 Global Selling 首页，只检测比特窗口当前页面",
    )
    parser.add_argument(
        "--auto-login",
        action="store_true",
        help="与 --shop 配合：检测到未登录时使用数据库邮箱和浏览器保存密码自动登录，并关闭窗口",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=3,
        help="全店自动登录的并发进程数，默认 3",
    )
    parser.add_argument(
        "--output",
        default="",
        help="全店登录汇总 Excel 路径；默认写入 bit/登录状态汇总",
    )
    parser.add_argument(
        "--no-email",
        action="store_true",
        help="只生成 Excel，不发送汇总邮件（测试时使用）",
    )
    return parser


def run_login_test_from_command_line(args):
    """执行命令行登录状态检测；不自动登录，也不关闭浏览器窗口。"""
    import json
    from urllib.parse import urlparse

    import requests
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service

    from bit.bit_api import openBrowser
    from bit.bit_runtime_lock import create_window_lease

    shop_name = str(args.shop or "").strip()
    print(
        f"[{get_now_time()}] 读取店铺配置：{shop_name}",
        flush=True,
    )
    config = load_shop_login_config(shop_name)
    window_id = config["window_id"]
    lease = create_window_lease(
        window_id,
        owner=f"mercado_login_cli:{shop_name}",
        shop_name=shop_name,
        task_type="mercado_login_test",
    )
    if not lease.acquire(timeout=0):
        result = {
            "ok": False,
            "status": "窗口正在被其他任务占用",
            "shop": shop_name,
        }
        print(json.dumps(result, ensure_ascii=False))
        return 2

    try:
        print(
            f"[{get_now_time()}] 已获取窗口任务锁，正在打开比特浏览器...",
            flush=True,
        )
        response = openBrowser(window_id)
        data = response.get("data") if isinstance(response, dict) else None
        if not isinstance(data, dict) or not data.get("driver") or not data.get("http"):
            raise RuntimeError(f"打开比特浏览器失败：{response}")

        debugger_address = str(data["http"]).strip()
        debugger_url = (
            debugger_address
            if debugger_address.startswith(("http://", "https://"))
            else f"http://{debugger_address}"
        )
        print(
            f"[{get_now_time()}] BitBrowser 已返回调试地址，等待浏览器调试端口就绪...",
            flush=True,
        )
        debugger_error = ""
        debugger_ready = False
        for attempt in range(1, 7):
            try:
                debugger_response = requests.get(
                    f"{debugger_url.rstrip('/')}/json/version",
                    timeout=3,
                )
                debugger_response.raise_for_status()
                debugger_ready = True
                break
            except Exception as exc:
                debugger_error = str(exc)
                print(
                    f"[{get_now_time()}] 调试端口尚未就绪，"
                    f"第 {attempt}/6 次：{debugger_error[:160]}",
                    flush=True,
                )
                if attempt < 6:
                    time.sleep(2)
        if not debugger_ready:
            raise RuntimeError(
                f"BitBrowser 调试端口未就绪：{debugger_error or debugger_address}"
            )

        options = webdriver.ChromeOptions()
        options.add_experimental_option("debuggerAddress", debugger_address)
        driver = webdriver.Chrome(
            service=Service(data["driver"]),
            options=options,
        )
        driver.implicitly_wait(10)
        print(
            f"[{get_now_time()}] 已连接比特浏览器 Selenium 会话",
            flush=True,
        )
        try:
            driver.set_page_load_timeout(max(1, int(args.page_load_timeout)))
        except Exception:
            pass

        navigate_error = ""
        if not args.no_navigate:
            print(
                f"[{get_now_time()}] 正在打开美客多首页检测登录状态，页面超时 "
                f"{max(1, int(args.page_load_timeout))} 秒...",
                flush=True,
            )
            result = ensure_mercado_login_from_home(
                driver,
                shop_name,
                window_id=window_id,
                wait_seconds=max(0, int(args.wait_seconds)),
                navigation_wait_seconds=5,
            )
            navigate_error = str(result.get("navigation_error") or "")[:240]
            login_detected = bool(result.get("login_detected_before"))
        else:
            login_detected = is_mercado_login_page(driver)
            print(
                f"[{get_now_time()}] 当前页面检测完成："
                f"{'检测到登录页' if login_detected else '当前不是登录页'}",
                flush=True,
            )
            result = ensure_mercado_login(
                driver,
                shop_name,
                window_id=window_id,
                wait_seconds=max(0, int(args.wait_seconds)),
            )
        try:
            page_host = urlparse(driver.current_url or "").netloc
        except Exception:
            page_host = ""
        safe_result = {
            "ok": bool(result.get("ok")),
            "status": result.get("status", ""),
            "message": result.get("message", ""),
            "shop": shop_name,
            "login_detected_before": login_detected,
            "page_host": page_host,
            "navigate_error": navigate_error,
            "browser_kept_open": True,
        }
        print(json.dumps(safe_result, ensure_ascii=False), flush=True)
        return 0 if safe_result["ok"] else 3
    finally:
        lease.release()
        print(
            f"[{get_now_time()}] 已释放窗口任务锁；比特浏览器保持打开",
            flush=True,
        )


def run_single_auto_login_from_command_line(args):
    """执行单店自动登录，供控制台“店铺状态”的重新自动登录操作调用。"""
    shop_name = str(args.shop or "").strip()
    config = load_shop_login_config(shop_name)
    result = login_one_database_shop(
        config,
        wait_seconds=max(1, int(args.wait_seconds)),
        page_load_timeout=max(1, int(args.page_load_timeout)),
    )
    sync_login_results_to_window_anomalies([result])
    safe_result = {
        key: value
        for key, value in result.items()
        if key not in ("email",)
    }
    print(json.dumps(safe_result, ensure_ascii=False), flush=True)
    return 0 if result.get("ok") else 3


def main(argv=None):
    parser = build_command_line_parser()
    args = parser.parse_args(argv)
    target = str(getattr(args, "shop", "") or "").strip() or "全部未忽略店铺"
    job_lock = InterProcessLock(
        MERCADO_LOGIN_JOB_LOCK_KEY,
        owner=f"bit_mercado_login:{target}",
        metadata={"target": target, "task_type": "mercado_login"},
    )
    if not job_lock.acquire(timeout=0):
        owner = get_lock_owner(MERCADO_LOGIN_JOB_LOCK_KEY)
        print(
            json.dumps(
                {
                    "ok": False,
                    "status": "已有登录检测任务运行中",
                    "shop": str(getattr(args, "shop", "") or ""),
                    "message": f"请等待现有任务结束：{owner.get('owner') or 'mercado_login'}",
                    "lock_owner": owner,
                },
                ensure_ascii=False,
            ),
            flush=True,
        )
        return 5
    try:
        if args.all_active_login:
            batch_result = run_all_database_shop_logins(
                max_workers=max(1, int(args.workers)),
                wait_seconds=max(1, int(args.wait_seconds)),
                page_load_timeout=max(1, int(args.page_load_timeout)),
                output_path=args.output or None,
                send_email=not args.no_email,
            )
            print(
                json.dumps(
                    {
                        "ok": bool(args.no_email or batch_result["email_sent"]),
                        "shop_count": batch_result["shop_count"],
                        "success_count": batch_result["success_count"],
                        "outcome_counts": batch_result["outcome_counts"],
                        "status_counts": batch_result.get("status_counts", {}),
                        "report_path": batch_result["report_path"],
                        "email_sent": batch_result["email_sent"],
                        "max_workers": batch_result["max_workers"],
                        "window_anomaly_sync": batch_result.get(
                            "window_anomaly_sync", {}
                        ),
                    },
                    ensure_ascii=False,
                ),
                flush=True,
            )
            return 0 if args.no_email or batch_result["email_sent"] else 4
        if args.auto_login:
            return run_single_auto_login_from_command_line(args)
        return run_login_test_from_command_line(args)
    except Exception as exc:
        import traceback

        print(
            json.dumps(
                {
                    "ok": False,
                    "status": "测试异常",
                    "shop": str(getattr(args, "shop", "") or ""),
                    "message": str(exc),
                    "browser_kept_open": True,
                },
                ensure_ascii=False,
            )
        )
        traceback.print_exc()
        return 1
    finally:
        job_lock.release()


if __name__ == "__main__":
    raise SystemExit(main())
