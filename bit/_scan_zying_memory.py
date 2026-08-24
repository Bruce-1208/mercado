from __future__ import annotations

import ctypes
import re
from ctypes import wintypes
from pathlib import Path

from openpyxl import load_workbook


PROCESS_VM_READ = 0x0010
PROCESS_QUERY_INFORMATION = 0x0400
MEM_COMMIT = 0x1000
PAGE_GUARD = 0x100
PAGE_NOACCESS = 0x01


class MEMORY_BASIC_INFORMATION(ctypes.Structure):
    _fields_ = [
        ("BaseAddress", ctypes.c_void_p),
        ("AllocationBase", ctypes.c_void_p),
        ("AllocationProtect", wintypes.DWORD),
        ("PartitionId", wintypes.WORD),
        ("RegionSize", ctypes.c_size_t),
        ("State", wintypes.DWORD),
        ("Protect", wintypes.DWORD),
        ("Type", wintypes.DWORD),
    ]


def first_product_id() -> str:
    path = Path("outputs/zying_boutique_17074_20260812_run3/全部产品_第35页.xlsx")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return str(next(workbook.active.iter_rows(min_row=2, values_only=True))[0])
    finally:
        workbook.close()


kernel32 = ctypes.windll.kernel32
pid = None
snapshot = kernel32.CreateToolhelp32Snapshot(0x00000002, 0)


class PROCESSENTRY32W(ctypes.Structure):
    _fields_ = [
        ("dwSize", wintypes.DWORD),
        ("cntUsage", wintypes.DWORD),
        ("th32ProcessID", wintypes.DWORD),
        ("th32DefaultHeapID", ctypes.POINTER(ctypes.c_ulong)),
        ("th32ModuleID", wintypes.DWORD),
        ("cntThreads", wintypes.DWORD),
        ("th32ParentProcessID", wintypes.DWORD),
        ("pcPriClassBase", ctypes.c_long),
        ("dwFlags", wintypes.DWORD),
        ("szExeFile", wintypes.WCHAR * 260),
    ]


entry = PROCESSENTRY32W()
entry.dwSize = ctypes.sizeof(entry)
if kernel32.Process32FirstW(snapshot, ctypes.byref(entry)):
    while True:
        if entry.szExeFile.casefold() == "zying.exe":
            pid = entry.th32ProcessID
            break
        if not kernel32.Process32NextW(snapshot, ctypes.byref(entry)):
            break
kernel32.CloseHandle(snapshot)
if not pid:
    raise RuntimeError("ZYing.exe is not running")

handle = kernel32.OpenProcess(PROCESS_VM_READ | PROCESS_QUERY_INFORMATION, False, pid)
if not handle:
    raise ctypes.WinError()

needle_text = first_product_id()
needles = (needle_text.encode(), needle_text.encode("utf-16-le"))
address = 0
matches = []
mbi = MEMORY_BASIC_INFORMATION()
try:
    while kernel32.VirtualQueryEx(
        handle, ctypes.c_void_p(address), ctypes.byref(mbi), ctypes.sizeof(mbi)
    ):
        base = int(mbi.BaseAddress or 0)
        size = int(mbi.RegionSize)
        readable = (
            mbi.State == MEM_COMMIT
            and not (mbi.Protect & PAGE_GUARD)
            and not (mbi.Protect & PAGE_NOACCESS)
            and size <= 128 * 1024 * 1024
        )
        if readable:
            buffer = ctypes.create_string_buffer(size)
            read = ctypes.c_size_t()
            if kernel32.ReadProcessMemory(
                handle,
                ctypes.c_void_p(base),
                buffer,
                size,
                ctypes.byref(read),
            ):
                data = buffer.raw[: read.value]
                for needle in needles:
                    start = 0
                    while True:
                        index = data.find(needle, start)
                        if index < 0:
                            break
                        chunk = data[max(0, index - 262144): min(len(data), index + 262144)]
                        ascii_text = chunk.decode("utf-8", errors="ignore")
                        utf16_text = chunk.decode("utf-16-le", errors="ignore")
                        utf16_text_odd = chunk[1:].decode("utf-16-le", errors="ignore")
                        urls = sorted(set(re.findall(r"https?://[^\s\x00\"'<>]+", ascii_text)))
                        urls += sorted(set(re.findall(r"https?://[^\s\x00\"'<>]+", utf16_text)))
                        urls += sorted(set(re.findall(r"https?://[^\s\x00\"'<>]+", utf16_text_odd)))
                        cache_names = sorted(set(re.findall(
                            r"[^\s\x00\"'<>]{0,100}(?:mlstatic|alicdn|hzzying)[^\s\x00\"'<>]{0,200}",
                            ascii_text + "\n" + utf16_text + "\n" + utf16_text_odd,
                            flags=re.IGNORECASE,
                        )))
                        printable = sorted(set(re.findall(
                            r"[A-Za-z0-9_./:\-\u4e00-\u9fff ]{20,300}",
                            ascii_text + "\n" + utf16_text + "\n" + utf16_text_odd,
                        )))
                        matches.append((hex(base + index), urls[:100], cache_names[:100], printable[:100]))
                        start = index + len(needle)
        next_address = base + size
        if next_address <= address:
            break
        address = next_address
finally:
    kernel32.CloseHandle(handle)

print("PID", pid, "PRODUCT", needle_text, "MATCHES", len(matches))
for match in matches:
    print(match)
