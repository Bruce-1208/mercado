import time
import re
from concurrent.futures import ProcessPoolExecutor, as_completed

from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.wait import WebDriverWait

from bit.bit_utils import get_now_time
from bit.bit_api import *
from bit.bit_runtime_lock import create_window_lease
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import pyautogui
from bit.bit_switch_country import *
from openpyxl import load_workbook
from bit.bit_send_mail import *
import pandas as pd

from datetime import datetime
from pathlib import Path
from bit.bit_db_api import insert_task_record, inset_reputation_info
from bit.bit_clash import *


REPUTATION_URL = "https://global-selling.mercadolibre.com/reputation"
SALES_SUMMARY_URL = "https://global-selling.mercadolibre.com/sales-summary"
METRICS_URL = "https://global-selling.mercadolibre.com/metrics#sc-menu"
RATE_LIMIT_MARKERS = (
    "429",
    "too many requests",
    "rate limit",
    "request limit",
    "access denied",
    "请求太过频繁",
    "请求过于频繁",
    "每秒最多可以发起",
    "访问过于频繁",
    "demasiadas solicitudes",
    "muitas solicitações",
    "hubo un error accediendo a esta página",
    "hubo un error accediendo a esta pagina",
)
SPANISH_IP_SWITCH_MARKERS = (
    "hubo un error accediendo a esta página",
    "hubo un error accediendo a esta pagina",
)
LOGIN_URL_MARKERS = ("/login/", "/lgz/", "/legacy-user")
LOGIN_TEXT_MARKERS = (
    "fill out your e-mail address to log in",
    "fill out your email address to log in",
    "iniciar sesión",
    "iniciar sesion",
    "iniciar sessão",
    "iniciar sessao",
    "登录您的账户",
    "登录你的账户",
    "登录账号",
    "请登录",
    "填写您的电子邮件地址以登录",
)
SITE_CODE_MAP = {
    "墨西哥": "MLM",
    "MX": "MLM",
    "MLM": "MLM",
    "巴西": "MLB",
    "BR": "MLB",
    "MLB": "MLB",
    "哥伦比亚": "MCO",
    "CO": "MCO",
    "MCO": "MCO",
    "智利": "MLC",
    "CL": "MLC",
    "MLC": "MLC",
    "阿根廷": "MLA",
    "AR": "MLA",
    "MLA": "MLA",
    "乌拉圭": "MLU",
    "UY": "MLU",
    "MLU": "MLU",
}
METRIC_LABEL_ALIASES = {
    "complaints": (
        "complaints",
        "complaint",
        "claims",
        "投诉",
        "买家投诉",
        "客诉",
    ),
    "shipments": (
        "non-compliant shipments",
        "non compliant shipments",
        "late shipments",
        "delayed shipments",
        "shipping delays",
        "handling time",
        "不合规发货",
        "不合规的发货",
        "未按时发货",
        "发货延迟",
        "延迟发货",
        "延误发货",
        "派送时间",
    ),
    "cancellations": (
        "cancellations",
        "cancellation",
        "cancelled orders",
        "canceled orders",
        "取消",
        "订单取消",
        "取消订单",
        "卖家取消",
    ),
}
VISITS_LABEL_ALIASES = ("visits", "visit", "访问量", "访问次数", "访问", "访客量")


class MercadoRateLimitError(RuntimeError):
    """Mercado Libre 或浏览器接口返回访问限频。"""


class MercadoAuthenticationError(RuntimeError):
    """Mercado Libre 登录态已经失效。"""


class MercadoPageStructureError(RuntimeError):
    """Mercado Libre 页面已打开，但预期的业务结构不存在。"""


class BitBrowserWindowError(RuntimeError):
    """比特浏览器窗口配置无效或无法打开。"""


def _is_rate_limited_text(value):
    text = str(value or "").lower()
    return any(marker in text for marker in RATE_LIMIT_MARKERS)


def _is_spanish_ip_switch_text(value):
    text = str(value or "").lower()
    return any(marker in text for marker in SPANISH_IP_SWITCH_MARKERS)


def _is_bit_api_rate_limited(res):
    return _is_rate_limited_text(res)


def _connect_browser(window_id, max_retries=3, retry_delay=30):
    last_res = None
    for attempt in range(1, max_retries + 1):
        res = openBrowser(window_id)  # 窗口ID从窗口配置界面中复制，或者api创建后返回
        last_res = res
        print(res)

        data = res.get("data") if isinstance(res, dict) else None
        if data and data.get("driver") and data.get("http"):
            driverPath = data["driver"]
            debuggerAddress = data["http"]
            break

        msg = res.get("msg", "") if isinstance(res, dict) else str(res)
        if _is_bit_api_rate_limited(res):
            print(
                f"{get_now_time()} 比特浏览器打开窗口被限频，等待 {retry_delay} 秒后重试："
                f"{window_id}，第 {attempt}/{max_retries} 次，原因：{msg}"
            )
        else:
            print(
                f"{get_now_time()} 比特浏览器打开窗口返回异常，等待 {retry_delay} 秒后重试："
                f"{window_id}，第 {attempt}/{max_retries} 次，返回：{res}"
            )
        time.sleep(retry_delay)
    else:
        if _is_bit_api_rate_limited(last_res):
            raise MercadoRateLimitError(
                f"比特浏览器打开窗口被限频，已重试 {max_retries} 次：{last_res}"
            )
        raise BitBrowserWindowError(
            f"打开比特浏览器窗口失败，已重试 {max_retries} 次，最后返回：{last_res}"
        )

    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_experimental_option("debuggerAddress", debuggerAddress)

    chrome_service = Service(driverPath)
    driver = webdriver.Chrome(service=chrome_service, options=chrome_options)
    driver.implicitly_wait(10)
    return driver


def _get_mercado_page_state(driver):
    try:
        driver.switch_to.default_content()
    except Exception:
        pass
    try:
        page_text = driver.execute_script("return document.body ? document.body.innerText : '';") or ""
    except Exception:
        page_text = ""
    try:
        current_url = driver.current_url or ""
    except Exception:
        current_url = ""
    try:
        title = driver.title or ""
    except Exception:
        title = ""
    try:
        page_source = driver.page_source or ""
    except Exception:
        page_source = ""
    return {
        "page_text": str(page_text),
        "current_url": str(current_url),
        "title": str(title),
        "page_source": str(page_source),
    }


def _is_mercado_rate_limited_page(driver=None, state=None):
    state = state or _get_mercado_page_state(driver)
    visible_state = "\n".join(
        (
            state.get("page_text", ""),
            state.get("title", ""),
            state.get("current_url", ""),
        )
    )
    if _is_rate_limited_text(visible_state):
        return True
    if not state.get("page_text", "").strip() and not state.get("title", "").strip():
        return _is_rate_limited_text(state.get("page_source", ""))
    return False


def _is_spanish_ip_switch_page(driver=None, state=None):
    state = state or _get_mercado_page_state(driver)
    visible_state = "\n".join(
        (
            state.get("page_text", ""),
            state.get("title", ""),
            state.get("current_url", ""),
        )
    )
    if _is_spanish_ip_switch_text(visible_state):
        return True
    if not state.get("page_text", "").strip() and not state.get("title", "").strip():
        return _is_spanish_ip_switch_text(state.get("page_source", ""))
    return False


def _is_mercado_login_state(state):
    current_url = str(state.get("current_url", "") or "").lower()
    visible_text = "\n".join(
        (str(state.get("title", "") or ""), str(state.get("page_text", "") or ""))
    ).lower()
    return any(marker in current_url for marker in LOGIN_URL_MARKERS) or any(
        marker in visible_text for marker in LOGIN_TEXT_MARKERS
    )


def _raise_if_mercado_unavailable(driver=None, state=None, context="页面"):
    state = state or _get_mercado_page_state(driver)
    if _is_mercado_rate_limited_page(state=state):
        raise MercadoRateLimitError(
            f"{context}访问受限：{state.get('current_url', '')} "
            f"{state.get('page_text', '')[:160]}"
        )
    if _is_mercado_login_state(state):
        raise MercadoAuthenticationError(
            f"{context}登录态失效，已跳转登录页：{state.get('current_url', '')}"
        )
    return state


def _open_reputation_page_with_validation(
    driver,
    name="",
    site="",
    max_hongkong_switches=3,
    switch_wait_seconds=8,
):
    """打开声誉页并验证；命中 Mercado 限频页时切换香港节点后重试。"""
    max_hongkong_switches = max(0, int(max_hongkong_switches))
    for attempt in range(1, max_hongkong_switches + 2):
        navigate_error = ""
        try:
            driver.get(REPUTATION_URL)
        except Exception as exc:
            navigate_error = str(exc)

        time.sleep(10)
        state = _get_mercado_page_state(driver)
        if _is_spanish_ip_switch_page(state=state):
            if attempt > max_hongkong_switches:
                raise MercadoRateLimitError(
                    f"{name}{site}声誉页面持续显示指定西语错误，已切换香港 IP "
                    f"{max_hongkong_switches} 次仍未恢复"
                )
            print(
                f"{get_now_time()}{name}{site}声誉页面显示限频，"
                f"第{attempt}/{max_hongkong_switches}次切换香港 IP 后重试"
            )
            switch_random_hongkong_node()
            get_public_ip()
            time.sleep(max(0, int(switch_wait_seconds)))
            continue

        _raise_if_mercado_unavailable(
            state=state,
            context=f"{name}{site}声誉页面",
        )

        if navigate_error:
            raise RuntimeError(f"{name}{site}声誉页面打开失败：{navigate_error}")

        try:
            WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.CLASS_NAME, "title__page--cbt"))
            )
        except Exception as exc:
            raise MercadoPageStructureError(
                f"{name}{site}声誉页面结构不匹配：{state.get('current_url', '')}"
            ) from exc

        print(f"{get_now_time()}{name}{site}声誉页面打开验证通过")
        return True

    raise RuntimeError(f"{name}{site}声誉页面打开验证失败")


def _get_country_name(site):
    country_map = {
        "墨西哥": "Mexico",
        "巴西": "Brazil",
        "哥伦比亚": "Colombia",
        "智利": "Chile",
        "阿根廷": "Argentina",
        "乌拉圭": "Uruguay",
    }
    return country_map.get(site, site)


def _deep_shadow_click(driver, selectors):
    return bool(
        driver.execute_script(
            """
            const selectors = arguments[0];
            function findAndClick(root) {
                for (const selector of selectors) {
                    let node = null;
                    try { node = root.querySelector(selector); } catch (_) {}
                    if (node) {
                        node.scrollIntoView({block: 'center', inline: 'center'});
                        node.click();
                        return true;
                    }
                }
                for (const node of root.querySelectorAll('*')) {
                    if (node.shadowRoot && findAndClick(node.shadowRoot)) return true;
                }
                return false;
            }
            return findAndClick(document);
            """,
            list(selectors),
        )
    )


def _select_country(driver, site, shop_name=""):
    if not site:
        return True

    site_key = str(site).strip()
    site_code = SITE_CODE_MAP.get(site_key) or SITE_CODE_MAP.get(site_key.upper())
    country = _get_country_name(site)
    for attempt in range(1, 4):
        try:
            opened = _deep_shadow_click(
                driver,
                (
                    'button[aria-label="Select country"]',
                    'button[aria-label*="country" i]',
                    'button[aria-label*="国家"]',
                    'button[aria-label*="站点"]',
                    ".nav-header-cbt__site-switcher",
                    '[data-testid*="site-switcher"]',
                ),
            )
            if not opened:
                opened = bool(oepn_country_switch(driver))
            if not opened:
                raise MercadoPageStructureError("没有找到站点选择器")

            time.sleep(1)
            success = False
            if site_code:
                success = _deep_shadow_click(
                    driver,
                    (
                        f'[data-value="{site_code}-remote"]',
                        f'[data-value^="{site_code}-"]',
                    ),
                )
            if not success:
                success = force_select_country(driver, country)
            if not success and site_key != country:
                success = force_select_country(driver, site_key)
            if success:
                print(get_now_time() + shop_name + "成功选择站点:", site)
                time.sleep(3)
                state = _raise_if_mercado_unavailable(
                    driver=driver,
                    context=f"{shop_name}{site}切换站点后的页面",
                )
                try:
                    WebDriverWait(driver, 10).until(
                        EC.visibility_of_element_located(
                            (By.CLASS_NAME, "title__page--cbt")
                        )
                    )
                except Exception as exc:
                    raise MercadoPageStructureError(
                        f"{shop_name}{site}切换站点后声誉页结构不匹配："
                        f"{state.get('current_url', '')}"
                    ) from exc
                return True
            raise MercadoPageStructureError(f"没有找到站点选项：{site}")
        except (MercadoAuthenticationError, MercadoRateLimitError):
            raise
        except Exception as exc:
            print(
                get_now_time()
                + shop_name
                + f"选择站点失败，第{attempt}/3次:"
                + site,
                exc,
            )
            if attempt < 3:
                try:
                    driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
                except Exception:
                    pass
                time.sleep(3)
    raise MercadoPageStructureError(f"{shop_name}{site}站点切换失败")


def _click_visits_metric(driver):
    selectors = [
        (By.XPATH, "//*[self::button or @role='button' or @role='tab'][contains(., 'Visits')]"),
        (By.XPATH, "//*[normalize-space()='Visits']"),
        (By.XPATH, "//*[self::button or @role='button' or @role='tab'][contains(., '访问')]"),
        (By.XPATH, "//*[normalize-space()='访问量' or normalize-space()='访问次数']"),
        (
            By.XPATH,
            "/html/body/main/div/div/div[3]/div/div/div[3]/section/div[2]/div[2]/div[2]/div[1]/div/div/div/div[4]",
        ),
    ]
    for by, selector in selectors:
        try:
            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((by, selector))
            ).click()
            time.sleep(2)
            return
        except Exception:
            continue

    clicked = driver.execute_script(
        """
        function allNodes(root) {
            const nodes = [...root.querySelectorAll('*')];
            for (const node of [...nodes]) {
                if (node.shadowRoot) {
                    nodes.push(...allNodes(node.shadowRoot));
                }
            }
            return nodes;
        }

        const labels = arguments[0].map((value) => value.toLocaleLowerCase());
        const visitNode = allNodes(document).find((node) => {
            const text = (node.innerText || '').replace(/\\s+/g, ' ').trim();
            if (!text || text.length > 80) return false;
            const normalized = text.toLocaleLowerCase();
            return labels.some((label) => normalized === label || normalized.includes(label));
        });
        if (!visitNode) {
            return false;
        }
        const clickable = visitNode.closest('button, a, [role="button"], [role="tab"]') || visitNode;
        clickable.scrollIntoView({block: 'center', inline: 'center'});
        clickable.click();
        return true;
        """,
        list(VISITS_LABEL_ALIASES),
    )
    if not clicked:
        raise MercadoPageStructureError("没有找到 Visits/访问量 入口")
    time.sleep(2)


def _parse_visits_tooltip(text):
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if not lines:
        return None

    value = None
    for line in reversed(lines):
        match = re.search(r"[\d,.]+", line)
        if match:
            value = match.group(0).replace(",", "")
            break
    if value is None:
        return None

    date_text = ""
    for line in lines:
        if re.search(r"[A-Za-z]{3,}|月|日|/|-", line) and not re.fullmatch(
            r"[\d,.]+", line
        ):
            date_text = line
            break
    if not date_text:
        date_text = lines[0]

    return {"date": date_text, "visits": value, "raw": text}


def _parse_visit_records_from_json(data, days):
    records = []

    def walk(node, parent=None):
        if isinstance(node, dict):
            lower_keys = {str(key).lower(): key for key in node.keys()}
            value_key = None
            for key in node.keys():
                key_text = str(key).lower()
                if "visit" in key_text and isinstance(node.get(key), (int, float, str)):
                    value_key = key
                    break

            date_key = None
            for key_text, raw_key in lower_keys.items():
                if any(token in key_text for token in ["date", "day", "period", "label"]):
                    date_key = raw_key
                    break

            if value_key is not None:
                records.append(
                    {
                        "date": str(node.get(date_key, "")),
                        "visits": str(node.get(value_key, "")).replace(",", ""),
                        "raw": node,
                    }
                )

            for value in node.values():
                walk(value, node)
        elif isinstance(node, list):
            numeric_items = [item for item in node if isinstance(item, (int, float))]
            if parent and len(numeric_items) >= days:
                parent_text = str(parent).lower()
                if "visit" in parent_text:
                    for value in numeric_items[-days:]:
                        records.append({"date": "", "visits": str(value), "raw": parent})
            for item in node:
                walk(item, parent)

    walk(data)

    cleaned = []
    seen = set()
    for record in records:
        visits = record.get("visits", "")
        if not re.search(r"\d", visits):
            continue
        key = (record.get("date", ""), visits)
        if key in seen:
            continue
        seen.add(key)
        cleaned.append(record)
    return cleaned[-days:]


def _extract_visits_from_network(driver, days):
    entries = driver.execute_script(
        """
        return performance.getEntriesByType('resource')
            .map((entry) => entry.name)
            .filter((url) => /metric|visit|traffic|analytics|sales-summary/i.test(url))
            .slice(-30);
        """
    )

    for url in reversed(entries):
        try:
            data = driver.execute_async_script(
                """
                const url = arguments[0];
                const done = arguments[arguments.length - 1];
                fetch(url, {credentials: 'include'})
                    .then((response) => {
                        const contentType = response.headers.get('content-type') || '';
                        if (!contentType.includes('json')) {
                            done(null);
                            return;
                        }
                        return response.json().then((json) => done(json));
                    })
                    .catch(() => done(null));
                """,
                url,
            )
            if not data:
                continue
            visits = _parse_visit_records_from_json(data, days)
            if visits:
                return visits
        except Exception:
            continue
    return []


def _get_tooltip_text(driver):
    tooltip_selectors = [
        ".andes-tooltip",
        ".recharts-tooltip-wrapper",
        ".highcharts-tooltip",
        "[role='tooltip']",
        "[class*='tooltip']",
    ]
    for selector in tooltip_selectors:
        elements = driver.find_elements(By.CSS_SELECTOR, selector)
        for element in elements:
            text = element.text.strip()
            if text:
                return text
    return ""


def _get_chart_rect(driver):
    rect = driver.execute_script(
        """
        const candidates = [...document.querySelectorAll('svg, canvas')]
            .map((node) => {
                const rect = node.getBoundingClientRect();
                return {
                    x: rect.left,
                    y: rect.top,
                    width: rect.width,
                    height: rect.height,
                    area: rect.width * rect.height,
                    visible: rect.width > 160 && rect.height > 100 && rect.bottom > 0 && rect.right > 0,
                    tag: node.tagName,
                    text: node.closest('section, main, div')?.innerText || ''
                };
            })
            .filter((item) => item.visible)
            .sort((a, b) => {
                const aVisit = /Visits|访问/i.test(a.text) ? 1 : 0;
                const bVisit = /Visits|访问/i.test(b.text) ? 1 : 0;
                return (bVisit - aVisit) || (b.area - a.area);
            });
        return candidates[0] || null;
        """
    )
    return rect


def _extract_visits_from_dom(driver, days):
    data = driver.execute_script(
        """
        const results = [];
        const candidates = [
            ...document.querySelectorAll('svg circle, svg [class*="dot"], svg [class*="point"]')
        ];
        for (const node of candidates) {
            const aria = node.getAttribute('aria-label') || node.getAttribute('data-testid') || '';
            const title = node.querySelector('title')?.textContent || '';
            const text = `${aria} ${title}`.trim();
            if (/visit|访问/i.test(text) && /\\d/.test(text)) {
                results.push(text);
            }
        }
        return [...new Set(results)].slice(-arguments[0]);
        """,
        days,
    )
    visits = []
    for item in data:
        parsed = _parse_visits_tooltip(item)
        if parsed:
            visits.append(parsed)
    return visits


def _extract_visits_by_hover(driver, days):
    rect = _get_chart_rect(driver)
    if not rect:
        return []

    left = rect["x"] + rect["width"] * 0.08
    right = rect["x"] + rect["width"] * 0.96
    y_list = [
        rect["y"] + rect["height"] * 0.35,
        rect["y"] + rect["height"] * 0.5,
        rect["y"] + rect["height"] * 0.65,
    ]
    visits = []

    for i in range(days):
        x = right - ((days - 1 - i) * (right - left) / max(days - 1, 1))
        for y in y_list:
            try:
                driver.execute_script(
                    """
                    const x = arguments[0];
                    const y = arguments[1];
                    const target = document.elementFromPoint(x, y) || document.body;
                    for (const type of ['pointerover', 'pointermove']) {
                        const eventClass = window.PointerEvent || window.MouseEvent;
                        target.dispatchEvent(new eventClass(type, {
                            bubbles: true,
                            cancelable: true,
                            clientX: x,
                            clientY: y,
                            view: window,
                            pointerType: 'mouse'
                        }));
                    }
                    for (const type of ['mouseover', 'mousemove']) {
                        target.dispatchEvent(new MouseEvent(type, {
                            bubbles: true,
                            cancelable: true,
                            clientX: x,
                            clientY: y,
                            view: window
                        }));
                    }
                    """,
                    x,
                    y,
                )
                time.sleep(0.35)
                tooltip = _get_tooltip_text(driver)
                parsed = _parse_visits_tooltip(tooltip)
                if parsed and parsed not in visits:
                    visits.append(parsed)
                    break
            except Exception:
                continue

    return visits[-days:]


def _to_visit_number_list(visits, days):
    numbers = []
    for item in visits[-days:]:
        value = item.get("visits", item) if isinstance(item, dict) else item
        match = re.search(r"\d+(?:[,.]\d+)*", str(value))
        if not match:
            continue
        numbers.append(int(match.group(0).replace(",", "").replace(".", "")))
    return numbers[-days:]


def get_recent_visits_info(driver,window_id, name, site, days=8):
    # driver = _connect_browser(window_id)
    driver.get(METRICS_URL)
    driver.refresh()
    time.sleep(5)
    _raise_if_mercado_unavailable(
        driver=driver,
        context=f"{name}{site}流量页面",
    )

    # _select_country(driver, site, name)
    _click_visits_metric(driver)
    time.sleep(3)

    visits = _extract_visits_from_network(driver, days)
    if len(visits) < days:
        visits = _extract_visits_from_dom(driver, days)
    if len(visits) < days:
        visits = _extract_visits_by_hover(driver, days)

    if not visits:
        print("没有读取到Visits/访问量流量数据，请确认页面已加载并且折线图可见")
        debug_path = Path(__file__).resolve().parent / "visits_debug.png"
        driver.save_screenshot(str(debug_path))
        print("已保存调试截图:", debug_path)

    result = _to_visit_number_list(visits, days)
    print(f"最近{days}天Visits/访问量流量数据:", result)
    return result


def get_visits_info(driver,window_id, name="", site="", days=8):
    return get_recent_visits_info(driver,window_id, name, site, days)


def _normalize_label(value):
    return re.sub(r"[\W_]+", "", str(value or "").casefold(), flags=re.UNICODE)


def _metric_kind_from_title(title):
    normalized = _normalize_label(title)
    for kind, aliases in METRIC_LABEL_ALIASES.items():
        if any(_normalize_label(alias) in normalized for alias in aliases):
            return kind
    return ""


def _extract_reputation_metrics(driver):
    cards = driver.execute_script(
        """
        return [...document.querySelectorAll('.variable__title')]
          .map((titleNode) => {
            const card = titleNode.closest('.andes-card') ||
              titleNode.closest('[class*="variable"]') || titleNode.parentElement;
            const percentage = card?.querySelector('.variable__percentage');
            return {
              title: (titleNode.innerText || titleNode.textContent || '').trim(),
              percentage: (percentage?.innerText || percentage?.textContent || '').trim()
            };
          })
          .filter((item) => item.title && item.percentage);
        """
    ) or []

    metrics = {}
    for card in cards:
        kind = _metric_kind_from_title(card.get("title", ""))
        if kind and kind not in metrics:
            metrics[kind] = card.get("percentage", "")

    # 中英文翻译版本都保留相同的变量卡片顺序。若平台再次调整译文，
    # 使用结构顺序兜底：投诉在首项、取消在倒数第二项、发货延误在末项。
    if len(cards) >= 3:
        first_kind = _metric_kind_from_title(cards[0].get("title", ""))
        cancel_kind = _metric_kind_from_title(cards[-2].get("title", ""))
        shipment_kind = _metric_kind_from_title(cards[-1].get("title", ""))
        if not first_kind:
            metrics.setdefault("complaints", cards[0].get("percentage", ""))
        if not cancel_kind:
            metrics.setdefault("cancellations", cards[-2].get("percentage", ""))
        if not shipment_kind:
            metrics.setdefault("shipments", cards[-1].get("percentage", ""))

    missing = [
        kind
        for kind in ("complaints", "shipments", "cancellations")
        if not metrics.get(kind)
    ]
    if missing:
        discovered = ", ".join(card.get("title", "") for card in cards) or "无"
        raise MercadoPageStructureError(
            f"声誉指标卡片解析失败，缺少 {','.join(missing)}；页面卡片：{discovered}"
        )
    return metrics


def _normalize_reputation_color(text, class_name=""):
    value = f"{text or ''} {class_name or ''}".casefold()
    color_aliases = (
        ("无色", ("you still have no color", "no color", "sin color", "sem cor", "无色", "暂无颜色")),
        ("红色", ("red", "rojo", "vermelho", "红色", "红")),
        ("橘色", ("orange", "naranja", "laranja", "橘色", "橙色")),
        ("黄色", ("yellow", "amarillo", "amarelo", "黄色", "黄")),
        ("绿色", ("green", "verde", "绿色", "绿", "mercadoleader", "mercado leader")),
    )
    for translated, aliases in color_aliases:
        if any(alias in value for alias in aliases):
            return translated
    return str(text or "").strip()


def _parse_gradient(value):
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    replacements = (
        (r"decreased?|下降|减少|下滑", "下滑"),
        (r"increased?|上升|增加|增长", "增长"),
        (r"unchanged|no\s+change|持平|未变化|无变化|不变", "持平"),
    )
    normalized = text
    for pattern, replacement in replacements:
        normalized = re.sub(pattern, replacement, normalized, flags=re.IGNORECASE)

    direction_match = re.search(r"下滑|增长|持平", normalized)
    rate_match = re.search(r"(?:\d+(?:[.,]\d+)?|—|-)\s*%", normalized)
    direction = direction_match.group(0) if direction_match else normalized
    rate = rate_match.group(0).replace(" ", "") if rate_match else normalized
    return direction, rate


def get_reputation_info(window_id, name, site, driver=None):
    if driver is None:
        driver = _connect_browser(window_id)

    _open_reputation_page_with_validation(driver, name, site)
    _select_country(driver, site, name)

    metrics = _extract_reputation_metrics(driver)
    data_complain = metrics["complaints"]
    data_delay = metrics["shipments"]
    data_cancel = metrics["cancellations"]
    print("提取到的投诉率为:", data_complain)
    print("提取到的延误率为:", data_delay)
    print("提取到的取消率为:", data_cancel)

    color_element = driver.find_element(By.CLASS_NAME, "thermometer__level")
    data_color = _normalize_reputation_color(
        color_element.text,
        color_element.get_attribute("class"),
    )
    print("账号的声誉为:", data_color)

    data_orders = driver.find_element(By.CLASS_NAME, "value__sales").text
    print("总单数为：", data_orders)

    reputation = [
        name,
        site,
        data_color,
        data_orders,
        data_complain,
        data_delay,
        data_cancel,
    ]

    data_warn = "正常"
    direction = ""
    gradient_rate = ""
    auxiliary_errors = []
    try:
        driver.get(SALES_SUMMARY_URL)
        time.sleep(3)
        _raise_if_mercado_unavailable(
            driver=driver,
            context=f"{name}{site}销售汇总页面",
        )
        try:
            data_warn = (
                WebDriverWait(driver, 10)
                .until(
                    EC.visibility_of_element_located(
                        (By.CLASS_NAME, "andes-message__content")
                    )
                )
                .text
            )
        except Exception:
            data_warn = "正常"

        try:
            data_gradient = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, ".andes-badge .andes-visually-hidden")
                )
            ).text
        except Exception:
            data_gradient = "持平"
        print("近七天变化情况为:", data_gradient)
        direction, gradient_rate = _parse_gradient(data_gradient)
    except Exception as exc:
        auxiliary_errors.append(f"销售汇总{_failure_status(exc).removeprefix('失败：')}")

    print("系统提示为:", data_warn)

    try:
        visits = str(get_visits_info(driver, window_id, name, site, 8))
    except Exception as exc:
        visits = "[]"
        auxiliary_errors.append(f"流量{_failure_status(exc).removeprefix('失败：')}")

    if auxiliary_errors:
        auxiliary_message = "辅助采集失败：" + "；".join(auxiliary_errors)
        data_warn = auxiliary_message if data_warn == "正常" else f"{data_warn}\n{auxiliary_message}"

    reputation.extend([direction, gradient_rate, data_warn, get_now_time()])
    reputation.append(visits)
    return reputation


def _failure_status(exc):
    text = re.sub(r"\s+", " ", str(exc or "")).strip()
    lower = text.casefold()
    if isinstance(exc, MercadoAuthenticationError) or any(
        marker in lower for marker in LOGIN_TEXT_MARKERS
    ):
        reason = "登录失效"
    elif isinstance(exc, MercadoRateLimitError) or _is_rate_limited_text(text):
        reason = "访问限频"
    elif isinstance(exc, BitBrowserWindowError) and (
        "没有找到相应数据" in text or "missing" in lower or "不存在" in text
    ):
        reason = "窗口ID不存在"
    elif isinstance(exc, BitBrowserWindowError) and "timeout" in lower:
        reason = "窗口打开超时"
    elif "窗口正在被其他任务占用" in text:
        reason = "窗口被其他任务占用"
    elif "站点切换失败" in text or "没有找到站点" in text:
        reason = "站点切换失败"
    elif isinstance(exc, MercadoPageStructureError):
        reason = "页面结构不匹配"
    elif isinstance(exc, TimeoutException) or "timeout" in lower:
        reason = "页面元素等待超时"
    else:
        reason = text or exc.__class__.__name__ if exc is not None else "未知异常"
    return f"失败：{reason}"[:180]


def _split_sites(value):
    return [site.strip() for site in re.split(r"[，,、;；\n]+", str(value or "")) if site.strip()]


def _build_reputation_failure_row(name, site, failure_status="失败：未知异常"):
    return [
        name,
        site,
        "执行失败",
        "",
        "",
        "",
        "",
        "",
        "",
        failure_status,
        get_now_time(),
        "",
    ]


def _is_ignored_config_value(value):
    return "忽略" in str(value or "").strip()


def _deduplicate_config_rows(rows):
    unique_rows = []
    seen = set()
    for row in rows:
        key = (str(row[0]).strip(), str(row[1]).strip(), str(row[3] or "").strip())
        if key in seen:
            print(f"{get_now_time()}跳过重复配置：{row[1]} {row[3]}")
            continue
        seen.add(key)
        unique_rows.append(row)
    return unique_rows


def _run_reputation_for_browser(row):
    id = row[0]
    name = row[1]
    remark = row[2]
    if _is_ignored_config_value(remark):
        return [], []

    if not row[3]:
        return [], [("获取声誉信息", name, "", "失败：未配置站点", get_now_time())]

    lease = create_window_lease(
        id,
        owner=f"reputation_collection:{name}",
        shop_name=name,
        task_type="reputation_collection",
    )
    if not lease.acquire(timeout=0):
        print(get_now_time() + name + "窗口已被其他任务占用，跳过本次声誉采集")
        return [], [("获取声誉信息", name, "", "跳过：窗口被其他任务占用", get_now_time())]

    print(get_now_time() + "开始打开窗口:" + name)
    reputation_info_sum = []
    result = []
    sites = _split_sites(row[3])

    try:
        try:
            driver = _connect_browser(id)
        except Exception as exc:
            status = _failure_status(exc)
            print(get_now_time() + name + "打开窗口失败：" + status, exc)
            for site in sites:
                result.append(("获取声誉信息", name, site, status, get_now_time()))
                reputation_info_sum.append(
                    _build_reputation_failure_row(name, site, status)
                )
            return reputation_info_sum, result

        fatal_profile_error = None
        for site in sites:
            if fatal_profile_error is not None:
                status = _failure_status(fatal_profile_error)
                result.append(("获取声誉信息", name, site, status, get_now_time()))
                reputation_info_sum.append(
                    _build_reputation_failure_row(name, site, status)
                )
                continue

            succeeded = False
            last_error = None
            for attempt in range(1, 4):
                try:
                    reputation_info = get_reputation_info(id, name, site, driver=driver)
                    reputation_info_sum.append(reputation_info)
                    print(get_now_time() + name + site + "获取声誉信息成功")
                    result.append(("获取声誉信息", name, site, "成功", get_now_time()))
                    succeeded = True
                    break
                except Exception as e:
                    last_error = e
                    print(get_now_time() + name + site + "执行失败", e)
                    if isinstance(e, MercadoAuthenticationError):
                        fatal_profile_error = e
                        break
                    if attempt < 3:
                        should_switch_ip = (
                            _is_spanish_ip_switch_text(e)
                            or _is_spanish_ip_switch_page(driver=driver)
                        )
                        if should_switch_ip:
                            print(get_now_time() + name + site + "检测到指定西语错误页，切换香港 IP 后重试")
                            switch_random_hongkong_node()
                            get_public_ip()
                        time.sleep(5)

            if not succeeded:
                status = _failure_status(last_error)
                result.append(("获取声誉信息", name, site, status, get_now_time()))
                reputation_info_sum.append(
                    _build_reputation_failure_row(name, site, status)
                )
            time.sleep(10)
    finally:
        print(get_now_time() + "结束，正在关闭窗口")
        try:
            closeBrowser(id, lease=lease)
        except Exception as e:
            print(get_now_time() + name + "关闭窗口失败", e)
        lease.release()
        print(get_now_time() + "已经关闭窗口")

    return reputation_info_sum, result


def get_reputation_info_all(max_workers=20):
    """使用独立进程并发采集各店铺声誉，主进程负责汇总、导出和入库。"""
    start = int(time.time())
    print(start)
    root_path = Path(__file__).resolve().parent
    file_path = root_path / "比特配置文件.xlsx"
    # file_path = root_path / "比特配置文件测试.xlsx"

    wb = load_workbook(file_path)
    sheet = wb.active
    reputation_info_sum = []
    result = []
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    rows = [row for row in rows if row and row[0] and not _is_ignored_config_value(row[2])]
    rows = _deduplicate_config_rows(rows)

    worker_count = max(1, min(int(max_workers), len(rows))) if rows else 1
    with ProcessPoolExecutor(max_workers=worker_count) as executor:
        future_map = {
            executor.submit(_run_reputation_for_browser, row): row for row in rows
        }
        for future in as_completed(future_map):
            row = future_map[future]
            name = row[1]
            try:
                browser_reputations, browser_result = future.result()
                reputation_info_sum.extend(browser_reputations)
                result.extend(browser_result)
                print(get_now_time() + name + "窗口任务完成")
            except Exception as e:
                print(get_now_time() + name + "窗口任务异常", e)
                result.append(
                    ("获取声誉信息", name, "", _failure_status(e), get_now_time())
                )

    reputation_info_sum_str = "\n".join(map(str, reputation_info_sum))
    print(reputation_info_sum_str)

    end = int(time.time())
    print(get_now_time() + "总花费", end - start)
    df = pd.DataFrame(
        reputation_info_sum,
        columns=[
            "店铺名",
            "站点",
            "声誉颜色",
            "总单量",
            "投诉率",
            "延误率",
            "取消率",
            "增加或减少",
            "近七天变化率",
            "系统告警",
            "更新时间",
            "一周流量趋势"
        ],
    )

    now = datetime.now()
    date_str = datetime.now().strftime("%Y-%m-%d-%H")

    df.to_excel(
        root_path / ("美客多声誉/武汉泽顺店铺声誉信息汇总" + date_str + ".xlsx"),
        index=False,
    )

    send_info(
        "美客多所有店铺声誉汇总",
        "",
        root_path / ("美客多声誉/武汉泽顺店铺声誉信息汇总" + date_str + ".xlsx"),
        r"武汉泽顺店铺声誉信息汇总" + date_str + ".xlsx",
    )
    print(get_now_time() + "发送邮件成功")

    inset_reputation_info(reputation_info_sum)
    insert_task_record(result)


def main():
    return get_reputation_info_all()


if __name__ == "__main__":
    # get_reputation_info('22139511815a4bf588fe96d5fdafded6','四季如春','墨西哥')
    main()
