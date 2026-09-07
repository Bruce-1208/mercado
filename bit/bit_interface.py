
import queue
import json
import ipaddress
import re
from collections import deque
import functools
import html
import hashlib
import hmac
import multiprocessing
import os
import secrets
import signal
import subprocess
import sys
import threading
import time
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed
from io import BytesIO
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.parse import quote, urlsplit
from urllib.request import Request, urlopen

from flask import Flask, Response, request, render_template, jsonify, send_file, session, redirect, url_for, g, has_request_context
from itsdangerous import BadSignature, SignatureExpired, URLSafeTimedSerializer
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

CURRENT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = CURRENT_DIR.parent
for path in (str(CURRENT_DIR), str(PROJECT_ROOT)):
    if path not in sys.path:
        sys.path.insert(0, path)

# 必须在导入 bit_db_api、bit_mysql 及业务模块前确定角色；这些模块会在
# 导入时固定本进程使用 MySQL 还是数据库 HTTP 接口。
from bit.workbench_runtime import bootstrap_runtime

RUNTIME_SETTINGS = bootstrap_runtime()


def resolve_template_dir():
    bundle_root = Path(getattr(sys, "_MEIPASS", CURRENT_DIR))
    candidates = (
        CURRENT_DIR / "templates",
        CURRENT_DIR / "bit" / "templates",
        bundle_root / "bit" / "templates",
        bundle_root / "templates",
        PROJECT_ROOT / "bit" / "templates",
    )
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return CURRENT_DIR / "templates"

import bit.bit_appeal_ai as bit_appeal_ai
import bit.bit_check_risk as bit_check_risk
import bit.bit_db_api as bit_db_api
import bit.bit_daily_task as bit_daily_task
import bit.bit_infractions_info as bit_infractions_info
import bit.bit_infringement_knowledge_analysis as bit_infringement_knowledge_analysis
import bit.bit_inventory as bit_inventory
import bit.bit_pago_info as bit_pago_info
try:
    import bit.bit_print as bit_print
except ModuleNotFoundError as exc:
    if exc.name != "bit.bit_print":
        raise
    import bit_playwright.bit_print as bit_print
import bit.bit_reputation_info as bit_reputation_info
import bit.bit_order_sync as bit_order_sync
import bit.bit_prohibited_listing_sync as bit_prohibited_listing_sync
import bit.bit_store_link_remote_update as bit_store_link_remote_update
import bit.bit_store_link_sync as bit_store_link_sync
import bit.bit_update_orders as bit_update_orders
import bit.bit_zying_caiji as bit_zying_caiji
import bit.mercado_communications as mercado_communications
import bit.mercado_infraction_sync as mercado_infraction_sync
import bit.mercado_reputation as mercado_reputation
import bit.mercado_tokens as mercado_tokens
from bit.local_agent_bundle import build_business_bundle
from bit.local_agent_distribution import build_agent_distribution
from bit.local_agent_hub import (
    LocalAgentStore,
    TERMINAL_JOB_STATUSES,
    normalize_agent_id,
)
from erp.mercadolibre_infraction_store import (
    current_infraction_counts_by_token_site,
    list_infraction_dashboard,
)
from bit.bit_appeal import *
from bit.bit_collection_control import DEFAULT_COLLECTION_MAX_WORKERS
from bit.bit_config import list_shop_configs, split_config_sites
from bit.bit_runtime_lock import (
    InterProcessLock,
    RUNTIME_LOCK_DIR,
    create_window_lease,
    get_lock_owner,
)
from bit.bit_mercado_login import (
    MERCADO_LOGIN_JOB_LOCK_KEY,
    is_login_blocking_result,
    is_shop_status_anomaly,
)
from bit.bit_utils import *
from bit.bit_api import *

# 引入数据库入库需要的模块
import logging
from decimal import Decimal
from datetime import datetime, timedelta, timezone
# from db_pool import get_db_connection  # 确保你的连接池文件在这个目录下


WORKBENCH_REMEMBER_HOURS = 6
WORKBENCH_SECRET_KEY_FILE = CURRENT_DIR / "runtime_locks" / "workbench_secret.key"


def resolve_workbench_secret_key(secret_file=None):
    configured = str(os.environ.get("WORKBENCH_SECRET_KEY", "")).strip()
    if configured:
        return configured

    configured_file = str(os.environ.get("WORKBENCH_SECRET_KEY_FILE", "")).strip()
    secret_path = Path(secret_file or configured_file or WORKBENCH_SECRET_KEY_FILE)
    try:
        if secret_path.exists():
            persisted = secret_path.read_text(encoding="utf-8").strip()
            if len(persisted) >= 32:
                return persisted
            logging.warning("工作台会话密钥文件内容无效，将使用本次进程临时密钥：%s", secret_path)
            return secrets.token_hex(32)

        secret_path.parent.mkdir(parents=True, exist_ok=True)
        generated = secrets.token_hex(32)
        try:
            file_descriptor = os.open(
                str(secret_path),
                os.O_WRONLY | os.O_CREAT | os.O_EXCL,
                0o600,
            )
        except FileExistsError:
            persisted = secret_path.read_text(encoding="utf-8").strip()
            return persisted if len(persisted) >= 32 else secrets.token_hex(32)
        with os.fdopen(file_descriptor, "w", encoding="utf-8") as secret_handle:
            secret_handle.write(generated)
        return generated
    except OSError as exc:
        logging.warning("无法持久化工作台会话密钥，将使用本次进程临时密钥：%s", exc)
        return secrets.token_hex(32)


app = Flask(__name__, template_folder=str(resolve_template_dir()))
app.secret_key = resolve_workbench_secret_key()
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    PERMANENT_SESSION_LIFETIME=timedelta(hours=WORKBENCH_REMEMBER_HOURS),
    SESSION_REFRESH_EACH_REQUEST=False,
    TEMPLATES_AUTO_RELOAD=True,
    SEND_FILE_MAX_AGE_DEFAULT=0,
)

LOCAL_EXECUTOR_TOKEN_SALT = "zeshun-local-executor-v1"
LOCAL_EXECUTOR_SESSION_TOKEN_SALT = "zeshun-local-executor-session-v2"
LOCAL_EXECUTOR_SESSION_TOKEN_PREFIX = "session:"
LOCAL_EXECUTOR_TOKEN_MAX_AGE_SECONDS = 5 * 60
LOCAL_EXECUTOR_BROWSER_URL = str(
    os.environ.get("BIT_LOCAL_EXECUTOR_URL") or "http://127.0.0.1:5000"
).strip().rstrip("/")
LOCAL_EXECUTOR_PERMISSIONS = frozenset(
    ("appeal.execute", "tasks.view", "tasks.execute")
)
LOCAL_AGENT_ENROLLMENT_TOKEN_SALT = "zeshun-local-agent-enrollment-v1"
LOCAL_AGENT_CREDENTIAL_TOKEN_SALT = "zeshun-local-agent-credential-v1"
LOCAL_AGENT_ENROLLMENT_MAX_AGE_SECONDS = 24 * 60 * 60
LOCAL_AGENT_CREDENTIAL_MAX_AGE_SECONDS = 180 * 24 * 60 * 60
LOCAL_AGENT_ONLINE_SECONDS = max(
    15,
    int(os.environ.get("BIT_LOCAL_AGENT_ONLINE_SECONDS", "45")),
)
LOCAL_AGENT_HUB_PATH = Path(
    os.environ.get("BIT_LOCAL_AGENT_HUB_PATH")
    or (PROJECT_ROOT / ".data" / "local-agent-hub.sqlite3")
)
_local_agent_store_instance = None
_local_agent_store_lock = threading.Lock()
_local_agent_bundle_snapshot = None
_local_agent_bundle_snapshot_at = 0.0
_local_agent_bundle_lock = threading.Lock()


def get_local_agent_store():
    global _local_agent_store_instance
    if _local_agent_store_instance is None:
        with _local_agent_store_lock:
            if _local_agent_store_instance is None:
                _local_agent_store_instance = LocalAgentStore(LOCAL_AGENT_HUB_PATH)
    return _local_agent_store_instance


def current_local_agent_bundle(force=False):
    global _local_agent_bundle_snapshot, _local_agent_bundle_snapshot_at
    now = time.monotonic()
    if (
        not force
        and _local_agent_bundle_snapshot is not None
        and now - _local_agent_bundle_snapshot_at < 10
    ):
        return _local_agent_bundle_snapshot
    with _local_agent_bundle_lock:
        now = time.monotonic()
        if (
            force
            or _local_agent_bundle_snapshot is None
            or now - _local_agent_bundle_snapshot_at >= 10
        ):
            _local_agent_bundle_snapshot = build_business_bundle(PROJECT_ROOT)
            _local_agent_bundle_snapshot_at = now
        return _local_agent_bundle_snapshot


def local_executor_target_address_space(base_url=None):
    """Describe the browser-visible executor address for Chromium LNA."""

    try:
        hostname = urlsplit(
            str(base_url or LOCAL_EXECUTOR_BROWSER_URL).strip()
        ).hostname
    except ValueError:
        hostname = None
    normalized = str(hostname or "").strip().casefold()
    if normalized == "localhost":
        return "loopback"
    try:
        if ipaddress.ip_address(normalized).is_loopback:
            return "loopback"
    except ValueError:
        pass
    return "local"


def _configured_local_executor_origins():
    configured = str(
        os.environ.get("BIT_LOCAL_EXECUTOR_ALLOWED_ORIGINS") or ""
    ).strip()
    origins = {
        "https://zeshun.nat100.top",
        "http://zeshun.nat100.top",
    }
    if configured:
        origins.update(
            value.strip().rstrip("/")
            for value in configured.split(",")
            if value.strip()
        )
    try:
        parsed = urlsplit(RUNTIME_SETTINGS.api_base_url)
        if parsed.scheme and parsed.netloc:
            origins.add(f"{parsed.scheme}://{parsed.netloc}")
    except ValueError:
        pass
    return frozenset(origins)


LOCAL_EXECUTOR_ALLOWED_ORIGINS = _configured_local_executor_origins()


@app.after_request
def disable_workbench_html_cache(response):
    """Keep browser-rendered pages in sync with template and style edits."""

    if response.mimetype == "text/html":
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
        response.headers["Pragma"] = "no-cache"
    return response


@app.after_request
def allow_local_executor_browser_requests(response):
    """Allow the authenticated public workbench to reach this loopback bridge."""

    if not request.path.startswith("/api/local-executor/"):
        return response
    origin = str(request.headers.get("Origin") or "").strip().rstrip("/")
    if origin and origin in LOCAL_EXECUTOR_ALLOWED_ORIGINS:
        response.headers["Access-Control-Allow-Origin"] = origin
        response.headers["Vary"] = "Origin"
        response.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
        response.headers["Access-Control-Allow-Headers"] = (
            "Authorization, Content-Type"
        )
        response.headers["Access-Control-Max-Age"] = "300"
        if str(
            request.headers.get("Access-Control-Request-Private-Network") or ""
        ).strip().lower() == "true":
            response.headers["Access-Control-Allow-Private-Network"] = "true"
    return response


# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

PASSWORD_ITERATIONS = 260000

WORKBENCH_PERMISSION_GROUPS = (
    ("appeal", "自动化 AI 申诉", (("appeal.view", "查看"), ("appeal.execute", "执行/终止"))),
    ("tasks", "任务模块", (("tasks.view", "查看"), ("tasks.execute", "启动任务"))),
    ("order_print", "订单打印", (("order_print.view", "查看"), ("order_print.execute", "执行/终止"))),
    ("order_analysis", "订单分析", (("order_analysis.view", "查看"), ("order_analysis.execute", "导入订单"))),
    ("inventory", "库存管理", (("inventory.view", "查看"), ("inventory.execute", "出入库"), ("inventory.manage", "管理货架"))),
    ("shop_status", "店铺状态", (("shop_status.view", "查看"), ("shop_status.execute", "检测/处理"))),
    ("funds", "资金管理", (("funds.view", "查看"), ("funds.execute", "采集/终止"))),
    ("zying_collection", "智赢产品采集", (("zying_collection.view", "查看"), ("zying_collection.execute", "执行采集"))),
    ("ai_weight_price", "AI核重核价", (("ai_weight_price.view", "查看/导出"), ("ai_weight_price.execute", "采集/咨询/回写/配置"))),
    ("risk_check", "侵权检测", (("risk_check.view", "查看"), ("risk_check.execute", "执行检测"))),
    ("infringement_knowledge", "侵权知识库", (("infringement_knowledge.view", "查看"), ("infringement_knowledge.manage", "新增/修改/删除"))),
    ("infractions", "违规商品总览", (("infractions.view", "查看/导出"), ("infractions.execute", "采集"))),
    ("reputation", "声誉数据", (("reputation.view", "查看/导出"), ("reputation.execute", "采集/更新"))),
    ("customer_service", "客户消息", (("customer_service.view", "查看"), ("customer_service.manage", "回复/删除"))),
    ("ai_appeals", "AI 申诉记录", (("ai_appeals.view", "查看"),)),
    ("access", "人员与权限", (("access.view", "查看"), ("access.manage", "管理账号、角色和店铺配置"))),
)
WORKBENCH_PERMISSION_KEYS = tuple(
    permission_key
    for _, _, permissions in WORKBENCH_PERMISSION_GROUPS
    for permission_key, _ in permissions
)
WORKBENCH_DEFAULT_ROLES = (
    {
        "role_key": "super_admin",
        "role_name": "超级管理员",
        "description": "拥有控制台全部权限",
        "permissions": ("*",),
        "is_system": True,
    },
    {
        "role_key": "operator",
        "role_name": "运营人员",
        "description": "可查看并执行各业务模块，不可管理账号与权限",
        "permissions": tuple(
            key for key in WORKBENCH_PERMISSION_KEYS if not key.startswith("access.")
        ),
        "is_system": True,
    },
    {
        "role_key": "viewer",
        "role_name": "只读人员",
        "description": "只能查看业务数据，不能执行任务或修改配置",
        "permissions": tuple(
            key for key in WORKBENCH_PERMISSION_KEYS if key.endswith(".view")
        ),
        "is_system": True,
    },
)

try:
    YANDEX_CONSOLE_PORT = int(os.environ.get("WORKBENCH_YANDEX_PORT", "8011"))
except ValueError:
    YANDEX_CONSOLE_PORT = 8011
YANDEX_CONSOLE_HOST = "127.0.0.1"
YANDEX_CONSOLE_BASE_URL = f"http://{YANDEX_CONSOLE_HOST}:{YANDEX_CONSOLE_PORT}"
YANDEX_CONSOLE_PROXY_PATH = "/yandex-console"
YANDEX_PACKAGE_ROOT = PROJECT_ROOT / "yandex"
_yandex_console_lock = threading.Lock()
_yandex_console_process = None


def _yandex_console_health():
    try:
        with urlopen(f"{YANDEX_CONSOLE_BASE_URL}/api/health", timeout=1.0) as response:
            payload = json.loads(response.read().decode("utf-8"))
        return bool(
            response.status == 200
            and payload.get("status") == "ok"
            and payload.get("service") == "yandex-console"
        )
    except Exception:
        return False


def _yandex_console_python():
    configured = str(os.environ.get("WORKBENCH_YANDEX_PYTHON", "")).strip()
    if configured:
        return Path(configured)
    if os.name == "nt":
        return YANDEX_PACKAGE_ROOT / ".venv" / "Scripts" / "python.exe"
    return YANDEX_PACKAGE_ROOT / ".venv" / "bin" / "python"


def _yandex_console_setup_command(platform_name=None):
    if (platform_name or os.name) == "nt":
        return r".\yandex\run.ps1"
    return "./yandex/run.sh"


def ensure_yandex_console():
    global _yandex_console_process
    if _yandex_console_health():
        return True, "Yandex 控制台已运行"

    with _yandex_console_lock:
        if _yandex_console_health():
            return True, "Yandex 控制台已运行"
        python_executable = _yandex_console_python()
        if not python_executable.exists():
            setup_command = _yandex_console_setup_command()
            return False, f"Yandex 运行环境不存在，请先执行 {setup_command} 完成安装"
        if not (YANDEX_PACKAGE_ROOT / "__main__.py").exists():
            return False, "mercado/yandex 包不完整，缺少 __main__.py"

        data_dir = YANDEX_PACKAGE_ROOT / ".data"
        data_dir.mkdir(parents=True, exist_ok=True)
        log_path = data_dir / "workbench-server.log"
        environment = os.environ.copy()
        environment["YANDEX_HOST"] = YANDEX_CONSOLE_HOST
        environment["YANDEX_PORT"] = str(YANDEX_CONSOLE_PORT)
        creation_flags = (
            getattr(subprocess, "CREATE_NO_WINDOW", 0) if os.name == "nt" else 0
        )
        try:
            with log_path.open("ab", buffering=0) as log_file:
                _yandex_console_process = subprocess.Popen(
                    [str(python_executable), "-m", "yandex"],
                    cwd=str(PROJECT_ROOT),
                    env=environment,
                    stdout=log_file,
                    stderr=subprocess.STDOUT,
                    creationflags=creation_flags,
                    start_new_session=os.name != "nt",
                )
        except Exception as exc:
            logging.exception("启动 Yandex 控制台失败")
            return False, f"启动 Yandex 控制台失败：{exc}"

        for _ in range(40):
            if _yandex_console_health():
                return True, "Yandex 控制台已启动"
            if _yandex_console_process.poll() is not None:
                break
            time.sleep(0.25)
        return False, f"Yandex 控制台启动失败，请检查日志：{log_path}"


def _yandex_console_public_urls():
    return {
        "url": f"{YANDEX_CONSOLE_PROXY_PATH}/?embedded=1",
        "external_url": f"{YANDEX_CONSOLE_PROXY_PATH}/",
    }


def _yandex_console_upstream_url(proxy_path):
    safe_path = quote(str(proxy_path or ""), safe="/-._~")
    target = f"{YANDEX_CONSOLE_BASE_URL}/{safe_path}"
    if request.query_string:
        target = f"{target}?{request.query_string.decode('latin-1')}"
    return target


def _proxy_yandex_console_request(proxy_path):
    request_headers = {
        "Accept": request.headers.get("Accept", "*/*"),
        "X-Forwarded-Prefix": YANDEX_CONSOLE_PROXY_PATH,
    }
    for header_name in (
        "Content-Type",
        "If-Modified-Since",
        "If-None-Match",
        "Range",
    ):
        if request.headers.get(header_name):
            request_headers[header_name] = request.headers[header_name]

    body = (
        request.get_data()
        if request.method in {"POST", "PUT", "PATCH", "DELETE"}
        else None
    )
    upstream_request = Request(
        _yandex_console_upstream_url(proxy_path),
        data=body,
        headers=request_headers,
        method=request.method,
    )
    try:
        upstream = urlopen(upstream_request, timeout=300)
    except HTTPError as exc:
        upstream = exc
    except (URLError, TimeoutError):
        return jsonify({"detail": "Yandex 控制台暂时不可用，请重新加载后重试"}), 502

    try:
        payload = b"" if request.method == "HEAD" else upstream.read()
        content_type = str(upstream.headers.get("Content-Type") or "")
        if content_type.lower().startswith("text/html"):
            prefix = YANDEX_CONSOLE_PROXY_PATH.encode("utf-8")
            payload = payload.replace(b'href="/static/', b'href="' + prefix + b'/static/')
            payload = payload.replace(b'src="/static/', b'src="' + prefix + b'/static/')
            payload = payload.replace(
                b'window.YANDEX_BASE_PATH = "";',
                b"window.YANDEX_BASE_PATH = " + json.dumps(
                    YANDEX_CONSOLE_PROXY_PATH
                ).encode("utf-8") + b";",
            )
        response = Response(payload, status=upstream.status)
        for header_name in (
            "Content-Type",
            "Cache-Control",
            "Content-Disposition",
            "Content-Range",
            "Accept-Ranges",
            "ETag",
            "Last-Modified",
        ):
            value = upstream.headers.get(header_name)
            if value:
                response.headers[header_name] = value
        return response
    finally:
        upstream.close()


def _truthy_env(value):
    return str(value or "").strip().lower() in ("1", "true", "yes", "on")


def _resolve_use_db_api():
    return RUNTIME_SETTINGS.is_client


USE_DB_API = _resolve_use_db_api()

if USE_DB_API:
    db_get_latest_infraction_info = bit_db_api.get_latest_infraction_info
    db_get_latest_order_print_records = bit_db_api.get_latest_order_print_records
    db_get_latest_pago_info = bit_db_api.get_latest_pago_info
    db_get_latest_reputation_info = bit_db_api.get_latest_reputation_info
    db_get_ai_appeal_records = bit_db_api.get_ai_appeal_records
    db_list_appeal_phrases = bit_db_api.list_appeal_phrases
    db_get_random_appeal_phrase = bit_db_api.get_random_appeal_phrase
    db_create_appeal_phrase = bit_db_api.create_appeal_phrase
    db_update_appeal_phrase = bit_db_api.update_appeal_phrase
    db_delete_appeal_phrase = bit_db_api.delete_appeal_phrase
    db_list_infringement_knowledge = bit_db_api.list_infringement_knowledge
    db_create_infringement_knowledge = bit_db_api.create_infringement_knowledge
    db_update_infringement_knowledge = bit_db_api.update_infringement_knowledge
    db_delete_infringement_knowledge = bit_db_api.delete_infringement_knowledge
    db_bulk_create_infringement_knowledge = bit_db_api.bulk_create_infringement_knowledge
    db_get_infringement_knowledge_analysis_sources = bit_db_api.get_infringement_knowledge_analysis_sources
    db_upsert_analyzed_infringement_knowledge = bit_db_api.upsert_analyzed_infringement_knowledge
    db_get_window_anomalies = bit_db_api.get_window_anomalies
    db_insert_chat_info = bit_db_api.insert_chat_info
    db_insert_appeal_chat_record = bit_db_api.insert_appeal_chat_record
    db_insert_ai_appeal_record = bit_db_api.insert_ai_appeal_record
    db_insert_orders = bit_db_api.insert_orders
    db_get_high_after_sale_alerts = bit_db_api.get_high_after_sale_alerts
    db_get_high_profit_products = bit_db_api.get_high_profit_products
    db_list_orders = bit_db_api.list_orders
    db_insert_task_record = bit_db_api.insert_task_record
    db_insert_zying_product_info = bit_db_api.insert_zying_product_info
    db_upsert_zying_products_to_products = bit_db_api.upsert_zying_products_to_products
    db_get_existing_zying_product_ids = bit_db_api.get_existing_zying_product_ids
    db_get_zying_risk_candidates = bit_db_api.get_zying_risk_candidates
    db_update_zying_product_risks = bit_db_api.update_zying_product_risks
    db_list_zying_risk_categories = bit_db_api.list_zying_risk_categories
    db_get_zying_risk_results = bit_db_api.get_zying_risk_results
    db_resolve_window_anomaly = bit_db_api.resolve_window_anomaly
    db_upsert_window_anomaly = bit_db_api.upsert_window_anomaly
    db_inset_delay_info = bit_db_api.inset_delay_info
    db_inset_infraction_info = bit_db_api.inset_infraction_info
    db_inset_pago_info = bit_db_api.inset_pago_info
    db_inset_reputation_info = bit_db_api.inset_reputation_info
    db_create_mercado_collection_task = bit_db_api.create_mercado_collection_task
    db_update_mercado_collection_task = bit_db_api.update_mercado_collection_task
    db_get_mercado_collection_task = bit_db_api.get_mercado_collection_task
    db_upsert_mercado_collection_items = bit_db_api.upsert_mercado_collection_items
    db_list_mercado_collection_items = bit_db_api.list_mercado_collection_items
    db_list_mercado_product_items = bit_db_api.list_mercado_product_items
    db_add_mercado_collection_items_to_products = bit_db_api.add_mercado_collection_items_to_products
    db_delete_mercado_collection_items = bit_db_api.delete_mercado_collection_items
    db_delete_mercado_product_items = bit_db_api.delete_mercado_product_items
    db_move_mercado_product_items_to_collection = bit_db_api.move_mercado_product_items_to_collection
    db_get_mercado_product_items_by_ids = bit_db_api.get_mercado_product_items_by_ids
    db_update_mercado_product_publish_state = bit_db_api.update_mercado_product_publish_state
    db_create_mercado_product_publish_records = bit_db_api.create_mercado_product_publish_records
    db_get_mercado_product_publish_records_by_ids = bit_db_api.get_mercado_product_publish_records_by_ids
    db_get_published_mercado_product_item_ids = bit_db_api.get_published_mercado_product_item_ids
    db_update_mercado_product_publish_record = bit_db_api.update_mercado_product_publish_record
    db_list_mercado_product_publish_records = bit_db_api.list_mercado_product_publish_records
    db_update_mercado_product_review_status = bit_db_api.update_mercado_product_review_status
    db_update_mercado_product_item = bit_db_api.update_mercado_product_item
    db_list_mercado_management_categories = bit_db_api.list_mercado_management_categories
    db_create_mercado_management_category = bit_db_api.create_mercado_management_category
    db_update_mercado_management_category = bit_db_api.update_mercado_management_category
    db_delete_mercado_management_category = bit_db_api.delete_mercado_management_category
    db_assign_mercado_management_category = bit_db_api.assign_mercado_management_category
    db_list_mercado_store_links = bit_db_api.list_mercado_store_links
    db_bulk_update_mercado_store_links = bit_db_api.bulk_update_mercado_store_links
else:
    import pymysql
    from bit.bit_mysql import config as mysql_config
    from bit.bit_mysql import (
        get_latest_infraction_info,
        get_latest_order_print_records,
        get_latest_pago_info,
        get_latest_reputation_info,
        get_ai_appeal_records,
        list_appeal_phrases,
        get_random_appeal_phrase,
        create_appeal_phrase,
        update_appeal_phrase,
        delete_appeal_phrase,
        list_infringement_knowledge,
        create_infringement_knowledge,
        update_infringement_knowledge,
        delete_infringement_knowledge,
        bulk_create_infringement_knowledge,
        get_infringement_knowledge_analysis_sources,
        upsert_analyzed_infringement_knowledge,
        get_window_anomalies,
        insert_chat_info,
        insert_appeal_chat_record,
        insert_ai_appeal_record,
        insert_orders,
        get_high_after_sale_alerts,
        get_high_profit_products,
        list_orders,
        insert_task_record,
        insert_zying_product_info,
        get_existing_zying_product_ids,
        get_zying_risk_candidates,
        update_zying_product_risks,
        list_zying_risk_categories,
        get_zying_risk_results,
        resolve_window_anomaly,
        inset_delay_info,
        inset_infraction_info,
        inset_pago_info,
        inset_reputation_info,
        upsert_window_anomaly,
    )

    db_get_latest_infraction_info = get_latest_infraction_info
    db_get_latest_order_print_records = get_latest_order_print_records
    db_get_latest_pago_info = get_latest_pago_info
    db_get_latest_reputation_info = get_latest_reputation_info
    db_get_ai_appeal_records = get_ai_appeal_records
    db_list_appeal_phrases = list_appeal_phrases
    db_get_random_appeal_phrase = get_random_appeal_phrase
    db_create_appeal_phrase = create_appeal_phrase
    db_update_appeal_phrase = update_appeal_phrase
    db_delete_appeal_phrase = delete_appeal_phrase
    db_list_infringement_knowledge = list_infringement_knowledge
    db_create_infringement_knowledge = create_infringement_knowledge
    db_update_infringement_knowledge = update_infringement_knowledge
    db_delete_infringement_knowledge = delete_infringement_knowledge
    db_bulk_create_infringement_knowledge = bulk_create_infringement_knowledge
    db_get_infringement_knowledge_analysis_sources = get_infringement_knowledge_analysis_sources
    db_upsert_analyzed_infringement_knowledge = upsert_analyzed_infringement_knowledge
    db_get_window_anomalies = get_window_anomalies
    db_insert_chat_info = insert_chat_info
    db_insert_appeal_chat_record = insert_appeal_chat_record
    db_insert_ai_appeal_record = insert_ai_appeal_record
    db_insert_orders = insert_orders
    db_get_high_after_sale_alerts = get_high_after_sale_alerts
    db_get_high_profit_products = get_high_profit_products
    db_list_orders = list_orders
    db_insert_task_record = insert_task_record
    db_insert_zying_product_info = insert_zying_product_info
    db_get_existing_zying_product_ids = get_existing_zying_product_ids
    db_get_zying_risk_candidates = get_zying_risk_candidates
    db_update_zying_product_risks = update_zying_product_risks
    db_list_zying_risk_categories = list_zying_risk_categories
    db_get_zying_risk_results = get_zying_risk_results
    db_resolve_window_anomaly = resolve_window_anomaly
    db_upsert_window_anomaly = upsert_window_anomaly
    db_inset_delay_info = inset_delay_info
    db_inset_infraction_info = inset_infraction_info
    db_inset_pago_info = inset_pago_info
    db_inset_reputation_info = inset_reputation_info
    from erp.mercadolibre_collection_store import (
        add_collection_items_to_products as db_add_mercado_collection_items_to_products,
        create_collection_task as db_create_mercado_collection_task,
        delete_collection_items as db_delete_mercado_collection_items,
        delete_product_items as db_delete_mercado_product_items,
        get_collection_task as db_get_mercado_collection_task,
        get_product_items_by_ids as db_get_mercado_product_items_by_ids,
        move_product_items_to_collection as db_move_mercado_product_items_to_collection,
        create_product_publish_records as db_create_mercado_product_publish_records,
        get_product_publish_records_by_ids as db_get_mercado_product_publish_records_by_ids,
        get_published_product_item_ids as db_get_published_mercado_product_item_ids,
        list_product_publish_records as db_list_mercado_product_publish_records,
        list_collection_items as db_list_mercado_collection_items,
        list_product_items as db_list_mercado_product_items,
        list_management_categories as db_list_mercado_management_categories,
        create_management_category as db_create_mercado_management_category,
        update_management_category as db_update_mercado_management_category,
        delete_management_category as db_delete_mercado_management_category,
        assign_management_category as db_assign_mercado_management_category,
        update_collection_task as db_update_mercado_collection_task,
        update_product_publish_state as db_update_mercado_product_publish_state,
        update_product_publish_record as db_update_mercado_product_publish_record,
        update_product_item as db_update_mercado_product_item,
        update_product_items as db_update_mercado_product_items,
        update_product_review_status as db_update_mercado_product_review_status,
        upsert_zying_products_to_products as db_upsert_zying_products_to_products,
        upsert_collection_items as db_upsert_mercado_collection_items,
    )
    from erp.mercadolibre_store_link_store import (
        bulk_update_store_links as db_bulk_update_mercado_store_links,
        list_store_links as db_list_mercado_store_links,
    )


def make_password_hash(password):
    salt = secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac(
        "sha256",
        str(password).encode("utf-8"),
        salt.encode("utf-8"),
        PASSWORD_ITERATIONS,
    ).hex()
    return f"pbkdf2_sha256${PASSWORD_ITERATIONS}${salt}${digest}"


def verify_password(password, password_hash):
    try:
        method, iterations_text, salt, expected = str(password_hash or "").split("$", 3)
        if method != "pbkdf2_sha256":
            return False
        digest = hashlib.pbkdf2_hmac(
            "sha256",
            str(password).encode("utf-8"),
            salt.encode("utf-8"),
            int(iterations_text),
        ).hex()
        return hmac.compare_digest(digest, expected)
    except Exception:
        return False


def _normalize_workbench_permissions(value):
    if isinstance(value, str):
        try:
            value = json.loads(value)
        except (TypeError, ValueError):
            value = [value]
    if not isinstance(value, (list, tuple, set)):
        return []
    permissions = []
    for item in value:
        permission = str(item or "").strip()
        if permission and permission not in permissions:
            permissions.append(permission)
    return permissions


def workbench_permission_catalog():
    return [
        {
            "module": module,
            "label": label,
            "permissions": [
                {"key": permission_key, "label": permission_label}
                for permission_key, permission_label in permissions
            ],
        }
        for module, label, permissions in WORKBENCH_PERMISSION_GROUPS
    ]


def _validate_workbench_permissions(permissions):
    normalized = _normalize_workbench_permissions(permissions)
    invalid = [
        permission
        for permission in normalized
        if permission not in WORKBENCH_PERMISSION_KEYS
    ]
    if invalid:
        raise ValueError("包含不支持的权限：" + "、".join(invalid))
    for permission in tuple(normalized):
        if permission.endswith(".execute") or permission.endswith(".manage"):
            view_permission = permission.rsplit(".", 1)[0] + ".view"
            if view_permission in WORKBENCH_PERMISSION_KEYS and view_permission not in normalized:
                normalized.append(view_permission)
    return normalized


def _validate_workbench_password(password, required=True):
    password = str(password or "")
    if not password and not required:
        return ""
    if len(password) < 8:
        raise ValueError("密码至少需要 8 位")
    return password


def _validate_workbench_username(username):
    username = str(username or "").strip()
    allowed = set("abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-_.")
    if not 3 <= len(username) <= 64:
        raise ValueError("账号长度需要为 3–64 位")
    if any(character not in allowed for character in username):
        raise ValueError("账号只支持字母、数字、点、下划线和短横线")
    return username


_WORKBENCH_USER_REQUIRED_COLUMNS = {
    "id",
    "username",
    "password_hash",
    "display_name",
    "email",
    "department",
    "role_key",
    "is_active",
    "created_at",
    "updated_at",
}


def _workbench_schema_state(cursor):
    cursor.execute(
        """
        SELECT `TABLE_NAME` AS `table_name`
        FROM `information_schema`.`TABLES`
        WHERE `TABLE_SCHEMA` = DATABASE()
          AND `TABLE_NAME` IN ('workbench_roles', 'workbench_users')
        """
    )
    table_names = {
        str((row or {}).get("table_name") or "")
        for row in (cursor.fetchall() or [])
    }
    user_columns = set()
    if "workbench_users" in table_names:
        cursor.execute(
            """
            SELECT `COLUMN_NAME` AS `column_name`
            FROM `information_schema`.`COLUMNS`
            WHERE `TABLE_SCHEMA` = DATABASE()
              AND `TABLE_NAME` = 'workbench_users'
            """
        )
        user_columns = {
            str((row or {}).get("column_name") or "")
            for row in (cursor.fetchall() or [])
        }
    return table_names, user_columns


def _workbench_default_roles_are_current(cursor):
    cursor.execute(
        """
        SELECT `role_key`, `role_name`, `description`, `permissions_json`, `is_system`
        FROM `workbench_roles`
        WHERE `role_key` IN ('super_admin', 'operator', 'viewer')
        """
    )
    current_roles = {
        str(row.get("role_key") or ""): row
        for row in (cursor.fetchall() or [])
    }
    for role in WORKBENCH_DEFAULT_ROLES:
        current = current_roles.get(role["role_key"])
        if not current:
            return False
        current_permissions = set(
            _normalize_workbench_permissions(current.get("permissions_json"))
        )
        expected_permissions = set(role["permissions"])
        if (
            str(current.get("role_name") or "") != role["role_name"]
            or str(current.get("description") or "") != role["description"]
            or current_permissions != expected_permissions
            or bool(current.get("is_system")) != bool(role["is_system"])
        ):
            return False
    return True


def _workbench_users_are_current(cursor):
    cursor.execute(
        """
        SELECT COUNT(*) AS `total`,
               SUM(CASE WHEN `role_key` IS NULL OR `role_key` = '' THEN 1 ELSE 0 END)
                   AS `missing_role_count`
        FROM `workbench_users`
        """
    )
    row = cursor.fetchone() or {}
    return int(row.get("total") or 0) > 0 and int(row.get("missing_role_count") or 0) == 0


def _rollback_workbench_connection(connection):
    try:
        connection.rollback()
    except Exception as rollback_error:
        # 连接超时后 PyMySQL 会在 rollback() 上抛出 InterfaceError(0, "")。
        # 这里只记录清理失败，不能覆盖最初且更有价值的数据库异常。
        logging.warning(
            "工作台登录表初始化失败后的回滚未执行: %s: %s",
            type(rollback_error).__name__,
            rollback_error,
        )


def ensure_workbench_user_table():
    connection_config = dict(mysql_config)
    # 登录表初始化不能无限阻塞整个控制台启动；数据库暂时被锁定时先启动服务，
    # 后续登录请求仍会返回明确的数据库错误。
    connection_config.setdefault("connect_timeout", 5)
    connection_config.setdefault("read_timeout", 8)
    connection_config.setdefault("write_timeout", 8)
    connection = pymysql.connect(**connection_config)
    try:
        with connection.cursor() as cursor:
            table_names, user_columns = _workbench_schema_state(cursor)
            roles_exist = "workbench_roles" in table_names
            users_exist = "workbench_users" in table_names
            user_schema_ready = (
                users_exist
                and _WORKBENCH_USER_REQUIRED_COLUMNS.issubset(user_columns)
            )
            if (
                roles_exist
                and user_schema_ready
                and _workbench_default_roles_are_current(cursor)
                and _workbench_users_are_current(cursor)
            ):
                return False

            if not roles_exist:
                cursor.execute(
                    """
                    CREATE TABLE IF NOT EXISTS `workbench_roles` (
                        `role_key` VARCHAR(64) NOT NULL,
                        `role_name` VARCHAR(64) NOT NULL,
                        `description` VARCHAR(255) NULL,
                        `permissions_json` TEXT NOT NULL,
                        `is_system` TINYINT(1) NOT NULL DEFAULT 0,
                        `created_at` DATETIME NOT NULL,
                        `updated_at` DATETIME NOT NULL,
                        PRIMARY KEY (`role_key`),
                        UNIQUE KEY `uniq_workbench_role_name` (`role_name`)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                    """
                )
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            for role in WORKBENCH_DEFAULT_ROLES:
                cursor.execute(
                    """
                    INSERT INTO `workbench_roles`
                        (`role_key`, `role_name`, `description`, `permissions_json`,
                         `is_system`, `created_at`, `updated_at`)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    ON DUPLICATE KEY UPDATE
                        `role_name` = VALUES(`role_name`),
                        `description` = VALUES(`description`),
                        `permissions_json` = VALUES(`permissions_json`),
                        `is_system` = VALUES(`is_system`),
                        `updated_at` = VALUES(`updated_at`)
                    """,
                    (
                        role["role_key"],
                        role["role_name"],
                        role["description"],
                        json.dumps(role["permissions"], ensure_ascii=False),
                        1 if role["is_system"] else 0,
                        now,
                        now,
                    ),
                )
            if not users_exist:
                cursor.execute(
                    """
                    CREATE TABLE IF NOT EXISTS `workbench_users` (
                        `id` INT NOT NULL AUTO_INCREMENT,
                        `username` VARCHAR(64) NOT NULL,
                        `password_hash` VARCHAR(255) NOT NULL,
                        `display_name` VARCHAR(64) NULL,
                        `email` VARCHAR(128) NULL,
                        `department` VARCHAR(64) NULL,
                        `role_key` VARCHAR(64) NOT NULL DEFAULT 'viewer',
                        `is_active` TINYINT(1) NOT NULL DEFAULT 1,
                        `created_at` DATETIME NOT NULL,
                        `updated_at` DATETIME NOT NULL,
                        PRIMARY KEY (`id`),
                        UNIQUE KEY `uniq_workbench_username` (`username`)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                    """
                )
            for column_name, column_definition in (
                ("email", "VARCHAR(128) NULL"),
                ("department", "VARCHAR(64) NULL"),
                ("role_key", "VARCHAR(64) NULL"),
            ):
                if users_exist and column_name not in user_columns:
                    cursor.execute(
                        f"ALTER TABLE `workbench_users` ADD COLUMN `{column_name}` "
                        f"{column_definition}"
                    )
            cursor.execute(
                """
                UPDATE `workbench_users`
                SET `role_key` = 'super_admin'
                WHERE `role_key` IS NULL OR `role_key` = ''
                """
            )
            cursor.execute("SELECT COUNT(*) AS total FROM `workbench_users`")
            total = (cursor.fetchone() or {}).get("total") or 0
            if total == 0:
                username = os.environ.get("WORKBENCH_DEFAULT_USER", "admin")
                password = os.environ.get("WORKBENCH_DEFAULT_PASSWORD", "admin123456")
                cursor.execute(
                    """
                    INSERT INTO `workbench_users`
                        (`username`, `password_hash`, `display_name`, `role_key`,
                         `is_active`, `created_at`, `updated_at`)
                    VALUES (%s, %s, %s, 'super_admin', 1, %s, %s)
                    """,
                    (username, make_password_hash(password), "管理员", now, now),
                )
                logging.warning("workbench_users 为空，已创建默认账号 %s", username)
        connection.commit()
        return True
    except Exception:
        _rollback_workbench_connection(connection)
        raise
    finally:
        try:
            connection.close()
        except Exception as close_error:
            logging.warning(
                "关闭工作台登录表初始化连接失败: %s: %s",
                type(close_error).__name__,
                close_error,
            )


def get_workbench_user(username="", user_id=None):
    connection = pymysql.connect(**mysql_config)
    try:
        with connection.cursor() as cursor:
            where_clause = "u.`id` = %s" if user_id is not None else "u.`username` = %s"
            lookup_value = user_id if user_id is not None else username
            cursor.execute(
                f"""
                SELECT u.`id`, u.`username`, u.`password_hash`, u.`display_name`,
                       u.`email`, u.`department`, u.`role_key`, u.`is_active`,
                       r.`role_name`, r.`permissions_json`
                FROM `workbench_users` AS u
                LEFT JOIN `workbench_roles` AS r ON r.`role_key` = u.`role_key`
                WHERE {where_clause}
                LIMIT 1
                """,
                (lookup_value,),
            )
            return cursor.fetchone()
    finally:
        connection.close()


def build_workbench_session_user(user):
    role_key = str(user.get("role_key") or "viewer")
    permissions = (
        ["*"]
        if role_key == "super_admin"
        else _normalize_workbench_permissions(user.get("permissions_json"))
    )
    return {
        "id": user["id"],
        "username": user["username"],
        "display_name": user.get("display_name") or user["username"],
        "email": user.get("email") or "",
        "department": user.get("department") or "",
        "role_key": role_key,
        "role_name": user.get("role_name") or role_key,
        "permissions": permissions,
        "access_version": 1,
    }


def authenticate_workbench_user(username, password):
    if USE_DB_API:
        return bit_db_api.login_workbench_user(username, password)

    user = get_workbench_user(username)
    if not user or not user.get("is_active") or not verify_password(password, user.get("password_hash")):
        return None
    return build_workbench_session_user(user)


def _role_row_to_dict(row):
    row = dict(row or {})
    row["permissions"] = _normalize_workbench_permissions(
        row.pop("permissions_json", [])
    )
    row["is_system"] = bool(row.get("is_system"))
    row["user_count"] = int(row.get("user_count") or 0)
    return row


def list_workbench_roles_local():
    connection = pymysql.connect(**mysql_config)
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT r.`role_key`, r.`role_name`, r.`description`,
                       r.`permissions_json`, r.`is_system`, r.`created_at`,
                       r.`updated_at`, COUNT(u.`id`) AS `user_count`
                FROM `workbench_roles` AS r
                LEFT JOIN `workbench_users` AS u ON u.`role_key` = r.`role_key`
                GROUP BY r.`role_key`, r.`role_name`, r.`description`,
                         r.`permissions_json`, r.`is_system`, r.`created_at`,
                         r.`updated_at`
                ORDER BY FIELD(r.`role_key`, 'super_admin', 'operator', 'viewer'),
                         r.`role_name`
                """
            )
            return [_role_row_to_dict(row) for row in (cursor.fetchall() or [])]
    finally:
        connection.close()


def create_workbench_role_local(data):
    role_name = str((data or {}).get("role_name") or "").strip()
    if not role_name or len(role_name) > 64:
        raise ValueError("角色名称不能为空且最多 64 个字符")
    description = str((data or {}).get("description") or "").strip()[:255]
    permissions = _validate_workbench_permissions((data or {}).get("permissions"))
    role_key = "role_" + secrets.token_hex(8)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    connection = pymysql.connect(**mysql_config)
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO `workbench_roles`
                    (`role_key`, `role_name`, `description`, `permissions_json`,
                     `is_system`, `created_at`, `updated_at`)
                VALUES (%s, %s, %s, %s, 0, %s, %s)
                """,
                (
                    role_key,
                    role_name,
                    description,
                    json.dumps(permissions, ensure_ascii=False),
                    now,
                    now,
                ),
            )
        connection.commit()
    except pymysql.err.IntegrityError as exc:
        connection.rollback()
        raise ValueError("角色名称已存在") from exc
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()
    return role_key


def update_workbench_role_local(role_key, data):
    role_key = str(role_key or "").strip()
    role_name = str((data or {}).get("role_name") or "").strip()
    if not role_name or len(role_name) > 64:
        raise ValueError("角色名称不能为空且最多 64 个字符")
    description = str((data or {}).get("description") or "").strip()[:255]
    permissions = _validate_workbench_permissions((data or {}).get("permissions"))
    if role_key == "super_admin":
        permissions = ["*"]
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    connection = pymysql.connect(**mysql_config)
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT `role_key` FROM `workbench_roles` WHERE `role_key` = %s",
                (role_key,),
            )
            if not cursor.fetchone():
                raise ValueError("角色不存在")
            cursor.execute(
                """
                UPDATE `workbench_roles`
                SET `role_name` = %s, `description` = %s,
                    `permissions_json` = %s, `updated_at` = %s
                WHERE `role_key` = %s
                """,
                (
                    role_name,
                    description,
                    json.dumps(permissions, ensure_ascii=False),
                    now,
                    role_key,
                ),
            )
        connection.commit()
    except pymysql.err.IntegrityError as exc:
        connection.rollback()
        raise ValueError("角色名称已存在") from exc
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()


def delete_workbench_role_local(role_key):
    role_key = str(role_key or "").strip()
    connection = pymysql.connect(**mysql_config)
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT r.`is_system`, COUNT(u.`id`) AS `user_count`
                FROM `workbench_roles` AS r
                LEFT JOIN `workbench_users` AS u ON u.`role_key` = r.`role_key`
                WHERE r.`role_key` = %s
                GROUP BY r.`role_key`, r.`is_system`
                """,
                (role_key,),
            )
            role = cursor.fetchone()
            if not role:
                raise ValueError("角色不存在")
            if role.get("is_system"):
                raise ValueError("系统角色不能删除")
            if int(role.get("user_count") or 0) > 0:
                raise ValueError("该角色仍有关联账号，不能删除")
            cursor.execute(
                "DELETE FROM `workbench_roles` WHERE `role_key` = %s",
                (role_key,),
            )
        connection.commit()
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()


def list_workbench_users_local():
    connection = pymysql.connect(**mysql_config)
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT u.`id`, u.`username`, u.`display_name`, u.`email`,
                       u.`department`, u.`role_key`, u.`is_active`,
                       u.`created_at`, u.`updated_at`, r.`role_name`
                FROM `workbench_users` AS u
                LEFT JOIN `workbench_roles` AS r ON r.`role_key` = u.`role_key`
                ORDER BY u.`is_active` DESC, u.`id`
                """
            )
            rows = [dict(row) for row in (cursor.fetchall() or [])]
            for row in rows:
                row["is_active"] = bool(row.get("is_active"))
            return rows
    finally:
        connection.close()


def _require_workbench_role(cursor, role_key):
    cursor.execute(
        "SELECT `role_key` FROM `workbench_roles` WHERE `role_key` = %s",
        (role_key,),
    )
    if not cursor.fetchone():
        raise ValueError("所选角色不存在")


def create_workbench_user_local(data):
    data = data or {}
    username = _validate_workbench_username(data.get("username"))
    password = _validate_workbench_password(data.get("password"))
    display_name = str(data.get("display_name") or username).strip()[:64]
    email = str(data.get("email") or "").strip()[:128]
    department = str(data.get("department") or "").strip()[:64]
    role_key = str(data.get("role_key") or "viewer").strip()
    is_active = 1 if data.get("is_active", True) else 0
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    connection = pymysql.connect(**mysql_config)
    try:
        with connection.cursor() as cursor:
            _require_workbench_role(cursor, role_key)
            cursor.execute(
                """
                INSERT INTO `workbench_users`
                    (`username`, `password_hash`, `display_name`, `email`,
                     `department`, `role_key`, `is_active`, `created_at`, `updated_at`)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    username,
                    make_password_hash(password),
                    display_name,
                    email,
                    department,
                    role_key,
                    is_active,
                    now,
                    now,
                ),
            )
            user_id = cursor.lastrowid
        connection.commit()
        return user_id
    except pymysql.err.IntegrityError as exc:
        connection.rollback()
        raise ValueError("账号已存在") from exc
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()


def update_workbench_user_local(user_id, data):
    data = data or {}
    display_name = str(data.get("display_name") or "").strip()[:64]
    email = str(data.get("email") or "").strip()[:128]
    department = str(data.get("department") or "").strip()[:64]
    role_key = str(data.get("role_key") or "viewer").strip()
    is_active = 1 if data.get("is_active", True) else 0
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    connection = pymysql.connect(**mysql_config)
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT `id`, `role_key`, `is_active`
                FROM `workbench_users` WHERE `id` = %s
                """,
                (user_id,),
            )
            current = cursor.fetchone()
            if not current:
                raise ValueError("账号不存在")
            _require_workbench_role(cursor, role_key)
            removing_active_admin = (
                current.get("role_key") == "super_admin"
                and bool(current.get("is_active"))
                and (role_key != "super_admin" or not is_active)
            )
            if removing_active_admin:
                cursor.execute(
                    """
                    SELECT COUNT(*) AS total FROM `workbench_users`
                    WHERE `role_key` = 'super_admin' AND `is_active` = 1
                    """
                )
                if int((cursor.fetchone() or {}).get("total") or 0) <= 1:
                    raise ValueError("至少需要保留一个启用状态的超级管理员")
            cursor.execute(
                """
                UPDATE `workbench_users`
                SET `display_name` = %s, `email` = %s, `department` = %s,
                    `role_key` = %s, `is_active` = %s, `updated_at` = %s
                WHERE `id` = %s
                """,
                (
                    display_name,
                    email,
                    department,
                    role_key,
                    is_active,
                    now,
                    user_id,
                ),
            )
        connection.commit()
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()


def reset_workbench_user_password_local(user_id, password):
    password = _validate_workbench_password(password)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    connection = pymysql.connect(**mysql_config)
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                UPDATE `workbench_users`
                SET `password_hash` = %s, `updated_at` = %s
                WHERE `id` = %s
                """,
                (make_password_hash(password), now, user_id),
            )
            if cursor.rowcount == 0:
                raise ValueError("账号不存在")
        connection.commit()
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()


def _workbench_backend(function_name, *args):
    if USE_DB_API:
        return getattr(bit_db_api, function_name)(*args)
    return globals()[f"{function_name}_local"](*args)


def get_current_workbench_user():
    cached = getattr(g, "workbench_user", None)
    if cached is not None:
        return cached
    session_user = session.get("workbench_user")
    if not session_user:
        return None
    # 兼容升级前的会话和测试注入会话；新登录会话会实时读取账号状态和角色权限。
    if session_user.get("access_version") != 1:
        g.workbench_user = dict(session_user)
        return g.workbench_user
    if USE_DB_API:
        user = bit_db_api.get_workbench_session_user(session_user.get("id"))
    else:
        row = get_workbench_user(user_id=session_user.get("id"))
        user = (
            build_workbench_session_user(row)
            if row and row.get("is_active")
            else None
        )
    if not user:
        session.clear()
        return None
    session["workbench_user"] = user
    g.workbench_user = user
    return user


def workbench_user_has_permission(user, permission):
    if not user:
        return False
    if user.get("access_version") != 1:
        return True
    permissions = set(_normalize_workbench_permissions(user.get("permissions")))
    return "*" in permissions or permission in permissions


def login_required(view_func):
    @functools.wraps(view_func)
    def wrapper(*args, **kwargs):
        if get_current_workbench_user():
            return view_func(*args, **kwargs)
        if request.path.startswith(("/api/", f"{YANDEX_CONSOLE_PROXY_PATH}/api/")):
            return jsonify({"status": "error", "message": "请先登录"}), 401
        return redirect(url_for("login_page", next=request.full_path))

    return wrapper


BROWSER_EXTENSION_TOKEN_SALT = "zeshun-browser-extension-v1"


def create_browser_extension_token(user):
    """Create a signed, short-lived token for the Chrome/Edge collector."""
    payload = {
        "id": int(user.get("id") or 0),
        "username": str(user.get("username") or ""),
        "access_version": int(user.get("access_version") or 0),
    }
    return URLSafeTimedSerializer(
        app.secret_key, salt=BROWSER_EXTENSION_TOKEN_SALT
    ).dumps(payload)


def _browser_extension_user_from_token(token):
    try:
        payload = URLSafeTimedSerializer(
            app.secret_key, salt=BROWSER_EXTENSION_TOKEN_SALT
        ).loads(
            str(token or ""),
            max_age=WORKBENCH_REMEMBER_HOURS * 60 * 60,
        )
    except (BadSignature, SignatureExpired):
        return None
    if not isinstance(payload, dict) or not payload.get("id"):
        return None
    if payload.get("access_version") != 1:
        return payload
    if USE_DB_API:
        user = bit_db_api.get_workbench_session_user(payload.get("id"))
    else:
        row = get_workbench_user(user_id=payload.get("id"))
        user = (
            build_workbench_session_user(row)
            if row and row.get("is_active")
            else None
        )
    if not user or user.get("username") != payload.get("username"):
        return None
    return user


def browser_extension_login_required(view_func):
    @functools.wraps(view_func)
    def wrapper(*args, **kwargs):
        authorization = str(request.headers.get("Authorization") or "")
        token = authorization[7:].strip() if authorization.lower().startswith("bearer ") else ""
        try:
            user = _browser_extension_user_from_token(token)
        except Exception:
            logging.exception("校验泽顺商品采集助手登录状态失败")
            user = None
        if not user:
            return jsonify({"status": "error", "message": "插件登录已失效，请重新登录"}), 401
        g.browser_extension_user = user
        return view_func(*args, **kwargs)

    return wrapper


def create_local_executor_token(user, permission):
    permission = str(permission or "").strip()
    if permission not in LOCAL_EXECUTOR_PERMISSIONS:
        raise ValueError("不支持的本机执行权限")
    shared_secret = str(os.environ.get("BIT_DB_API_TOKEN") or "").strip()
    payload = {
        "id": int(user.get("id") or 0),
        "username": str(user.get("username") or ""),
        "permission": permission,
    }
    if not shared_secret:
        # The public workbench already has a persisted login signing key. Keep
        # it on the server; clients verify this token with that server over TLS.
        return LOCAL_EXECUTOR_SESSION_TOKEN_PREFIX + URLSafeTimedSerializer(
            app.secret_key,
            salt=LOCAL_EXECUTOR_SESSION_TOKEN_SALT,
        ).dumps(payload)
    return URLSafeTimedSerializer(
        shared_secret,
        salt=LOCAL_EXECUTOR_TOKEN_SALT,
    ).dumps(payload)


def _valid_local_executor_identity(payload):
    return (
        isinstance(payload, dict)
        and type(payload.get("id")) is int
        and payload["id"] > 0
        and isinstance(payload.get("username"), str)
        and bool(payload["username"].strip())
        and isinstance(payload.get("permission"), str)
        and payload["permission"] in LOCAL_EXECUTOR_PERMISSIONS
    )


def _verify_local_executor_session_token(token):
    """Server-only validation, including current account status and permissions."""
    if USE_DB_API or not str(token or "").startswith(LOCAL_EXECUTOR_SESSION_TOKEN_PREFIX):
        return None
    try:
        payload = URLSafeTimedSerializer(
            app.secret_key,
            salt=LOCAL_EXECUTOR_SESSION_TOKEN_SALT,
        ).loads(
            token[len(LOCAL_EXECUTOR_SESSION_TOKEN_PREFIX) :],
            max_age=LOCAL_EXECUTOR_TOKEN_MAX_AGE_SECONDS,
        )
    except (BadSignature, SignatureExpired):
        return None
    if not _valid_local_executor_identity(payload):
        return None
    row = get_workbench_user(user_id=payload["id"])
    if not row or not row.get("is_active") or row.get("username") != payload["username"]:
        return None
    user = build_workbench_session_user(row)
    if not workbench_user_has_permission(user, payload["permission"]):
        return None
    return {key: payload[key] for key in ("id", "username", "permission")}


def _verify_local_executor_token_with_server(token):
    # Never use an address supplied by the page/token or follow a redirect with
    # a bearer credential. The client's configured data server is the authority.
    base_url = str(RUNTIME_SETTINGS.api_base_url or "").strip().rstrip("/")
    try:
        parsed = urlsplit(base_url)
        if (
            not parsed.hostname
            or parsed.username
            or parsed.password
            or parsed.query
            or parsed.fragment
        ):
            raise ValueError("invalid server address")
        if parsed.scheme == "http" and parsed.netloc == "zeshun.nat100.top":
            base_url = "https://" + base_url[len("http://") :]
        elif parsed.scheme != "https" and not (
            parsed.scheme == "http" and parsed.hostname in ("127.0.0.1", "::1", "localhost")
        ):
            raise ValueError("HTTPS required")
    except ValueError as exc:
        raise RuntimeError("请将本机客户端的服务端地址配置为 HTTPS，以验证登录凭证") from exc
    try:
        response = bit_db_api.DB_API_SESSION.post(
            f"{base_url}/api/execution-targets/local-token/verify",
            headers={"Authorization": f"Bearer {token}"},
            timeout=10,
            allow_redirects=False,
        )
        if response.status_code in (401, 403):
            return None
        if response.status_code != 200:
            raise RuntimeError("请确认服务端已更新并重启，本机客户端指向同一工作台")
        result = response.json()
    except (bit_db_api.requests.RequestException, ValueError) as exc:
        raise RuntimeError("本机无法向服务端验证登录凭证，请检查客户端的服务端地址和连接") from exc
    payload = result.get("data") if isinstance(result, dict) else None
    if (
        not isinstance(result, dict)
        or result.get("status") != "success"
        or not _valid_local_executor_identity(payload)
    ):
        raise RuntimeError("服务端返回的本机执行凭证校验结果无效，请确认两端版本一致")
    return {key: payload[key] for key in ("id", "username", "permission")}


def _local_executor_user_from_token(token):
    token = str(token or "")
    if token.startswith(LOCAL_EXECUTOR_SESSION_TOKEN_PREFIX):
        if USE_DB_API:
            return _verify_local_executor_token_with_server(token)
        return _verify_local_executor_session_token(token)
    shared_secret = str(os.environ.get("BIT_DB_API_TOKEN") or "").strip()
    if not shared_secret:
        return None
    try:
        payload = URLSafeTimedSerializer(
            shared_secret,
            salt=LOCAL_EXECUTOR_TOKEN_SALT,
        ).loads(
            str(token or ""),
            max_age=LOCAL_EXECUTOR_TOKEN_MAX_AGE_SECONDS,
        )
    except (BadSignature, SignatureExpired):
        return None
    if (
        not isinstance(payload, dict)
        or not payload.get("id")
        or payload.get("permission") not in LOCAL_EXECUTOR_PERMISSIONS
    ):
        return None
    return payload


def create_local_agent_enrollment_token(user):
    permission = next(
        (value for value in ("appeal.execute", "tasks.execute")
         if workbench_user_has_permission(user, value)),
        "",
    )
    payload = {
        "id": int((user or {}).get("id") or 0),
        "username": str((user or {}).get("username") or ""),
        "permission": permission,
    }
    if not _valid_local_executor_identity(payload):
        raise ValueError("无法为当前账号创建 Agent 注册凭证")
    return URLSafeTimedSerializer(
        app.secret_key,
        salt=LOCAL_AGENT_ENROLLMENT_TOKEN_SALT,
    ).dumps(payload)


def _local_agent_enrollment_user(token):
    if USE_DB_API:
        return None
    try:
        payload = URLSafeTimedSerializer(
            app.secret_key,
            salt=LOCAL_AGENT_ENROLLMENT_TOKEN_SALT,
        ).loads(
            str(token or ""),
            max_age=LOCAL_AGENT_ENROLLMENT_MAX_AGE_SECONDS,
        )
    except (BadSignature, SignatureExpired):
        return None
    if not _valid_local_executor_identity(payload):
        return None
    row = get_workbench_user(user_id=payload["id"])
    if not row or not row.get("is_active") or row.get("username") != payload["username"]:
        return None
    user = build_workbench_session_user(row)
    if not workbench_user_has_permission(user, payload["permission"]):
        return None
    return payload


def create_local_agent_credential(agent_id, user_id=0):
    agent_id = normalize_agent_id(agent_id)
    signed = URLSafeTimedSerializer(
        app.secret_key,
        salt=LOCAL_AGENT_CREDENTIAL_TOKEN_SALT,
    ).dumps({"agent_id": agent_id, "user_id": int(user_id or 0)})
    return "agent:" + signed


def _verify_local_agent_credential(token):
    token = str(token or "")
    if not token.startswith("agent:"):
        return None
    try:
        payload, issued_at = URLSafeTimedSerializer(
            app.secret_key,
            salt=LOCAL_AGENT_CREDENTIAL_TOKEN_SALT,
        ).loads(
            token[len("agent:") :],
            max_age=LOCAL_AGENT_CREDENTIAL_MAX_AGE_SECONDS,
            return_timestamp=True,
        )
    except (BadSignature, SignatureExpired):
        return None
    try:
        agent = get_local_agent_store().get_agent(payload.get("agent_id"))
    except (KeyError, ValueError):
        return None
    if not agent:
        return None
    return {
        "agent_id": agent["agent_id"],
        "user_id": int(payload.get("user_id") or 0),
        "issued_at": issued_at.timestamp(),
    }


def local_agent_required(view_func):
    @functools.wraps(view_func)
    def wrapper(*args, **kwargs):
        if USE_DB_API:
            return jsonify({
                "status": "error",
                "message": "Agent 必须连接服务端工作台",
            }), 503
        request_token = str(request.headers.get("X-Local-Agent-Token") or "")
        shared_token = str(
            os.environ.get("BIT_LOCAL_AGENT_TOKEN")
            or os.environ.get("BIT_DB_API_TOKEN")
            or ""
        )
        claims = _verify_local_agent_credential(request_token)
        if not claims and shared_token and hmac.compare_digest(shared_token, request_token):
            claims = {"agent_id": "*", "user_id": 0}
        if not claims:
            return jsonify({"status": "error", "message": "Agent 凭证无效"}), 401
        g.local_agent_claims = claims
        return view_func(*args, **kwargs)

    return wrapper


def local_executor_required(*accepted_permissions):
    accepted = frozenset(
        str(permission or "").strip()
        for permission in accepted_permissions
        if str(permission or "").strip()
    )

    def decorator(view_func):
        @functools.wraps(view_func)
        def wrapper(*args, **kwargs):
            if request.method == "OPTIONS":
                return Response(status=204)
            if not USE_DB_API:
                return jsonify({
                    "status": "error",
                    "message": "该地址不是本机客户端执行端",
                }), 409
            if request.remote_addr not in ("127.0.0.1", "::1", "localhost"):
                return jsonify({"status": "error", "message": "Forbidden"}), 403
            origin = str(request.headers.get("Origin") or "").strip().rstrip("/")
            if origin and origin not in LOCAL_EXECUTOR_ALLOWED_ORIGINS:
                return jsonify({"status": "error", "message": "Forbidden"}), 403
            authorization = str(request.headers.get("Authorization") or "")
            token = (
                authorization[7:].strip()
                if authorization.lower().startswith("bearer ")
                else ""
            )
            try:
                user = _local_executor_user_from_token(token)
            except RuntimeError as exc:
                return jsonify({"status": "error", "message": str(exc)}), 503
            if not user:
                return jsonify({
                    "status": "error",
                    "message": "本机执行凭证无效或已过期，请重试",
                }), 401
            if accepted and user.get("permission") not in accepted:
                return jsonify({
                    "status": "error",
                    "message": "本机执行凭证没有该操作权限",
                }), 403
            g.local_executor_user = user
            return view_func(*args, **kwargs)

        return wrapper

    return decorator


def internal_api_required(view_func):
    @functools.wraps(view_func)
    def wrapper(*args, **kwargs):
        token = os.environ.get("BIT_DB_API_TOKEN", "")
        request_token = request.headers.get("X-Internal-Token", "")
        if token and hmac.compare_digest(token, request_token):
            return view_func(*args, **kwargs)
        if _verify_local_agent_credential(request_token):
            return view_func(*args, **kwargs)
        if request.remote_addr in ("127.0.0.1", "::1", "localhost"):
            return view_func(*args, **kwargs)
        return jsonify({"status": "error", "message": "Forbidden"}), 403

    return wrapper


def _required_workbench_permissions(path, method):
    method = str(method or "GET").upper()
    if path.startswith("/api/ai-weight-price/"):
        return ("ai_weight_price.view",) if method == "GET" else ("ai_weight_price.execute",)
    if path.startswith("/api/access/"):
        return ("access.view",) if method == "GET" else ("access.manage",)
    if path.startswith("/api/appeal-phrases"):
        return ("appeal.view",) if method == "GET" else ("appeal.execute",)
    if path.startswith("/api/infringement-knowledge"):
        return (
            ("infringement_knowledge.view",)
            if method == "GET"
            else ("infringement_knowledge.manage",)
        )
    if path.startswith("/api/run_shensu"):
        return ("appeal.execute",)
    if path == "/api/execution-agents":
        return ("appeal.view", "tasks.view", "tasks.execute")
    if path == "/api/local-agents/download":
        return ("appeal.execute", "tasks.execute")
    if path == "/api/collections/options":
        return (
            "appeal.view",
            "order_print.view",
            "infractions.view",
            "reputation.view",
        )
    if path.startswith("/api/infractions/"):
        return (
            ("infractions.execute",)
            if path == "/api/infractions/collect" and method == "POST"
            else ("infractions.view",)
        )
    if path.startswith("/api/official-infractions/"):
        return (
            ("infractions.execute",)
            if path == "/api/official-infractions/sync" and method == "POST"
            else ("infractions.view",)
        )
    if path.startswith("/api/reputation/"):
        return (
            ("reputation.execute",)
            if method == "POST"
            else ("reputation.view",)
        )
    if path.startswith("/api/funds/"):
        return (
            ("funds.execute",)
            if method == "POST"
            else ("funds.view",)
        )
    if path.startswith("/api/order-print/"):
        return (
            ("order_print.execute",)
            if method == "POST"
            else ("order_print.view",)
        )
    if path.startswith("/api/order-analysis/"):
        return (
            ("order_analysis.execute",)
            if method == "POST"
            else ("order_analysis.view",)
        )
    if path.startswith("/api/inventory/"):
        if path.startswith("/api/inventory/shelves") and method != "GET":
            return ("inventory.manage",)
        if path == "/api/inventory/movements" and method == "POST":
            return ("inventory.execute",)
        return ("inventory.view",)
    if path.startswith("/api/tasks/daily/"):
        return (
            ("tasks.execute",)
            if path.endswith(("/start", "/stop")) and method == "POST"
            else ("tasks.view",)
        )
    if path.startswith("/api/risk-check/"):
        return (
            ("risk_check.execute",)
            if path.endswith("/start") and method == "POST"
            else ("risk_check.view",)
        )
    if path.startswith("/api/zying-collection/"):
        return (
            ("zying_collection.execute",)
            if method == "POST"
            else ("zying_collection.view",)
        )
    if path == "/api/ai-appeal-records":
        return ("ai_appeals.view",)
    if path.startswith("/api/window-anomalies"):
        return (
            ("shop_status.execute",)
            if method == "POST"
            else ("shop_status.view",)
        )
    if path.startswith("/api/mercado-communications/"):
        action = path.rstrip("/").rsplit("/", 1)[-1].strip().lower()
        view_post_actions = {
            "customer-service-aggregate",
            "pre-sale-aggregate",
            "pre-sale-translate",
        }
        return (
            ("customer_service.manage",)
            if method == "POST" and action not in view_post_actions
            else ("customer_service.view",)
        )
    if path.startswith("/api/mercado-claims/"):
        return ("customer_service.view",)
    return ()


@app.before_request
def enforce_workbench_permissions():
    path = request.path
    if not path.startswith("/api/") or path.startswith("/api/db/"):
        return None
    if path in ("/api/login",):
        return None
    required_permissions = _required_workbench_permissions(path, request.method)
    if not required_permissions:
        return None
    user = get_current_workbench_user()
    if not user:
        return jsonify({"status": "error", "message": "请先登录"}), 401
    if not any(
        workbench_user_has_permission(user, permission)
        for permission in required_permissions
    ):
        return jsonify(
            {
                "status": "error",
                "message": "当前账号没有执行该操作的权限",
                "required_permissions": list(required_permissions),
            }
        ), 403
    return None


def reject_db_api_client_mode():
    if USE_DB_API:
        return jsonify({
            "status": "error",
            "message": "当前工作台是客户端模式，不提供数据库服务端接口；请在任一可访问数据库的电脑上以 --role server 启动。"
        }), 503
    return None


if USE_DB_API:
    logging.info(
        "工作台运行角色=client，数据库只通过接口访问：%s（配置来源：%s）",
        bit_db_api.DB_API_BASE_URL,
        RUNTIME_SETTINGS.source,
    )
else:
    logging.info(
        "工作台运行角色=server，MySQL 直连地址：%s（配置来源：%s）",
        mysql_config.get("host"),
        RUNTIME_SETTINGS.source,
    )
    try:
        ensure_workbench_user_table()
    except Exception as e:
        logging.error("初始化工作台登录表失败: %s", e)


def _authorize_ai_weight_price(permission):
    user = get_current_workbench_user()
    if not user:
        return jsonify({"message": "请先登录泽顺控制台"}), 401
    if not workbench_user_has_permission(user, permission):
        return jsonify({"message": "当前账号没有AI核重核价操作权限"}), 403
    return None


from erp.ai_weight_price.config import data_dir as ai_weight_price_data_dir
from erp.ai_weight_price.service import Service as AIWeightPriceService
from erp.ai_weight_price.web import create_blueprint as create_ai_weight_price_blueprint

ai_weight_price_service = AIWeightPriceService(ai_weight_price_data_dir())
app.register_blueprint(create_ai_weight_price_blueprint(ai_weight_price_service, _authorize_ai_weight_price))


# 1. 核心逻辑方法：改造成生成器
_original_stdout = sys.stdout
_original_stderr = sys.stderr
_thread_log_queues = {}
_thread_log_lock = threading.Lock()
_infraction_collect_lock = threading.Lock()
_infraction_collect_state = {
    "running": False,
    "started_at": "",
    "finished_at": "",
    "status": "idle",
    "message": "等待启动",
    "params": {},
}
_reputation_collect_lock = threading.Lock()
_reputation_collect_state = {
    "running": False,
    "started_at": "",
    "finished_at": "",
    "status": "idle",
    "message": "等待启动",
    "operation": "",
    "params": {},
}
API_REPUTATION_STATE_PATH = Path(
    os.environ.get("BIT_API_REPUTATION_STATE_PATH")
    or (RUNTIME_LOCK_DIR / "api_reputation_last_snapshot.json")
)
API_REPUTATION_AUTO_REFRESH_HOURS = (0, 12)
API_REPUTATION_MAX_WORKERS = 10


def _api_reputation_default_state():
    return {
        "running": False,
        "status": "idle",
        "message": "等待全量更新",
        "started_at": "",
        "finished_at": "",
        "elapsed_seconds": 0,
        "total_stores": 0,
        "completed_stores": 0,
        "success_stores": 0,
        "failed_stores": 0,
        "total_sites": 0,
        "rows": [],
        "failures": [],
    }


def _load_api_reputation_snapshot(state_path=None):
    path = Path(state_path or API_REPUTATION_STATE_PATH)
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, TypeError, ValueError, json.JSONDecodeError):
        return {**_api_reputation_default_state(), "logs": []}
    if not isinstance(payload, dict):
        return {**_api_reputation_default_state(), "logs": []}

    state = _api_reputation_default_state()
    state.update({key: payload.get(key, value) for key, value in state.items()})
    state["running"] = False
    state["rows"] = [
        dict(row) for row in (state.get("rows") or []) if isinstance(row, dict)
    ]
    state["failures"] = [
        dict(row) for row in (state.get("failures") or []) if isinstance(row, dict)
    ]
    if state["rows"] and state.get("status") not in {"success", "partial"}:
        state["status"] = "success"
    if state["rows"]:
        state["message"] = str(state.get("message") or "已加载上一次 API 声誉数据")
        try:
            state["total_sites"] = int(state.get("total_sites") or len(state["rows"]))
        except (TypeError, ValueError):
            state["total_sites"] = len(state["rows"])
    logs = payload.get("logs")
    return {
        **state,
        "logs": [str(line) for line in logs[-1000:]] if isinstance(logs, list) else [],
    }


_api_reputation_lock = threading.Lock()
_api_reputation_logs = deque(maxlen=1000)
_persisted_api_reputation = _load_api_reputation_snapshot()
_api_reputation_state = {
    key: value
    for key, value in _persisted_api_reputation.items()
    if key != "logs"
}
_api_reputation_logs.extend(_persisted_api_reputation.get("logs") or [])
_api_reputation_database_hydration_attempted = False
_api_reputation_scheduler_guard = threading.Lock()
_api_reputation_scheduler_thread = None
_api_reputation_scheduler_stop_event = threading.Event()
_risk_check_lock = threading.Lock()
_risk_check_state_lock = threading.RLock()
_risk_check_logs = deque(maxlen=500)
_risk_check_state = {
    "running": False,
    "started_at": "",
    "finished_at": "",
    "status": "idle",
    "message": "等待启动",
    "params": {},
    "summary": {},
}
_infringement_knowledge_analysis_lock = threading.Lock()
_infringement_knowledge_analysis_state_lock = threading.RLock()
_infringement_knowledge_analysis_logs = deque(maxlen=600)
_infringement_knowledge_analysis_state = {
    "running": False,
    "started_at": "",
    "finished_at": "",
    "status": "idle",
    "message": "等待启动",
    "params": {},
    "summary": {},
}
_zying_collection_lock = threading.Lock()
_zying_collection_stop_event = threading.Event()
_zying_collection_state_lock = threading.RLock()
_zying_collection_logs = deque(maxlen=800)
_zying_collection_state = {
    "running": False,
    "started_at": "",
    "finished_at": "",
    "status": "idle",
    "message": "等待启动",
    "params": {},
    "summary": {},
    "requires_login": False,
}
_fund_collect_lock = threading.Lock()
_fund_collect_state = {
    "running": False,
    "started_at": "",
    "finished_at": "",
    "status": "idle",
    "message": "等待启动",
    "scope": "",
    "target": "",
    "collected_count": 0,
}
_mercado_collection_lock = threading.RLock()
_mercado_collection_stop_event = threading.Event()
_mercado_collection_state = {
    "running": False,
    "task_id": None,
    "status": "idle",
    "message": "等待启动",
    "source_url": "",
    "keyword": "",
    "source_site_id": "MLM",
    "source_site_name": "墨西哥",
    "collection_scope": "all",
    "requested_count": 0,
    "worker_count": 4,
    "candidate_count": 0,
    "processed_count": 0,
    "completed_count": 0,
    "failed_count": 0,
    "elapsed_seconds": 0,
    "current_page": 0,
    "current_item_id": "",
    "started_at": "",
    "finished_at": "",
}
_mercado_playwright_setup_state = {
    "running": False,
    "status": "idle",
    "message": "采集浏览器尚未打开",
}
_mercado_profit_refresh_lock = threading.Lock()
_mercado_profit_refresh_started = False
_mercado_profit_refresh_stop_event = threading.Event()
_mercado_shipping_rate_refresh_lock = threading.RLock()
_mercado_shipping_rate_refresh_state = {
    "running": False,
    "status": "idle",
    "message": "等待从官方更新",
    "started_at": "",
    "finished_at": "",
    "elapsed_seconds": 0,
    "success_sites": 0,
    "failed_sites": 0,
    "unavailable_site_count": 0,
    "recalculation_count": 0,
    "sites": [],
    "errors": [],
    "unavailable_sites": [],
}
_mercado_publish_lock = threading.RLock()
_mercado_publish_state = {
    "running": False,
    "batch_id": "",
    "status": "idle",
    "message": "等待选择产品上架",
    "selection_mode": "accounts",
    "token_id": None,
    "token_ids": [],
    "group_names": [],
    "store_name": "",
    "site_id": "MLM",
    "site_ids": ["MLM"],
    "site_name": "墨西哥",
    "target_count": 0,
    "completed_target_count": 0,
    "skipped_target_count": 0,
    "quantity": 500,
    "worker_count": 10,
    "requested_count": 0,
    "processed_count": 0,
    "published_count": 0,
    "failed_count": 0,
    "moved_to_collection_count": 0,
    "skipped_published_count": 0,
    "elapsed_seconds": 0,
    "average_seconds_per_item": 0,
    "items_per_minute": 0,
    "estimated_remaining_seconds": 0,
    "average_stage_seconds": {},
    "current_item_id": "",
    "started_at": "",
    "finished_at": "",
    "results": [],
}
_fund_collect_stop_event = None
_order_print_lock = threading.RLock()
_order_print_stop_event = None
_order_print_logs = deque(maxlen=1000)
_order_print_state = {
    "running": False,
    "started_at": "",
    "finished_at": "",
    "status": "idle",
    "message": "等待启动",
    "params": {},
    "printed": 0,
    "partial": 0,
    "no_orders": 0,
    "failed": 0,
    "skipped": 0,
    "printed_order_count": 0,
    "shipment_count": 0,
    "fallback_store_count": 0,
    "task_id": "",
    "download_path": "",
    "download_name": "",
    "results": [],
    "site_last_runs": [],
}
_order_analysis_import_lock = threading.Lock()
# 每次点击“启动”都会创建独立任务实例。RLock 允许状态辅助函数在 API
# 已持锁时安全复用；真正的同一店铺窗口冲突仍由 window lease 控制。
_daily_task_lock = threading.RLock()
_daily_task_stop_event = None
_daily_task_stop_manager = None
_daily_task_log_lock = threading.Lock()
_daily_task_log_path = Path(
    os.environ.get("BIT_DAILY_TASK_LOG_PATH")
    or (Path(CURRENT_DIR) / "logs" / "bit_daily_task_console.log")
)
_daily_task_state = {
    "running": False,
    "started_at": "",
    "finished_at": "",
    "status": "idle",
    "message": "等待启动",
    "stop_requested": False,
    "params": {},
}
_daily_tasks = {}
_daily_task_controls = {}
DAILY_TASK_HISTORY_LIMIT = 50
DEFAULT_DAILY_TASK_MAX_CONCURRENT = 8


def _daily_task_max_concurrent():
    try:
        return max(
            1,
            int(
                os.environ.get(
                    "BIT_DAILY_TASK_MAX_CONCURRENT",
                    DEFAULT_DAILY_TASK_MAX_CONCURRENT,
                )
            ),
        )
    except (TypeError, ValueError):
        return DEFAULT_DAILY_TASK_MAX_CONCURRENT


def _daily_task_log_file(task_id=""):
    task_id = re.sub(r"[^0-9A-Za-z_.-]+", "_", str(task_id or "").strip())
    if not task_id:
        return Path(_daily_task_log_path)
    base_path = Path(_daily_task_log_path)
    suffix = base_path.suffix or ".log"
    return base_path.with_name(f"{base_path.stem}_{task_id}{suffix}")


def _reset_daily_task_log(log_path=None):
    target_path = Path(log_path or _daily_task_log_path)
    with _daily_task_log_lock:
        target_path.parent.mkdir(parents=True, exist_ok=True)
        target_path.write_text("", encoding="utf-8")


def _append_daily_task_log(text, log_path=None):
    text = str(text or "")
    if not text:
        return
    target_path = Path(log_path or _daily_task_log_path)
    with _daily_task_log_lock:
        target_path.parent.mkdir(parents=True, exist_ok=True)
        with target_path.open("a", encoding="utf-8") as log_file:
            log_file.write(text)
            log_file.flush()


def _read_daily_task_log(
    max_bytes=512 * 1024,
    max_lines=2000,
    log_path=None,
):
    target_path = Path(log_path or _daily_task_log_path)
    try:
        with target_path.open("rb") as log_file:
            log_file.seek(0, os.SEEK_END)
            size = log_file.tell()
            start = max(0, size - max_bytes)
            log_file.seek(start, os.SEEK_SET)
            content = log_file.read().decode("utf-8", errors="replace")
        if start:
            content = content.partition("\n")[2]
        lines = content.splitlines()[-max_lines:]
        content = "\n".join(lines)
        for marker in ("<br>", "<br/>", "<br />"):
            content = content.replace(f"{marker}\r\n", "\n")
            content = content.replace(f"{marker}\n", "\n")
        return format_log_text(content).strip()
    except OSError:
        return ""


class DailyTaskLogSink:
    def __init__(self, log_path=None):
        self.log_path = Path(log_path or _daily_task_log_path)

    def put(self, text):
        _append_daily_task_log(text, self.log_path)


def _daily_task_display_name(params):
    params = dict(params or {})
    group_names = [str(value).strip() for value in params.get("group_names") or () if str(value).strip()]
    salespeople = [str(value).strip() for value in params.get("salespeople") or () if str(value).strip()]
    appeal_types = [str(value).strip() for value in params.get("appeal_types") or () if str(value).strip()]
    scope = (
        "店铺组：" + "、".join(group_names)
        if group_names
        else "业务员：" + "、".join(salespeople)
        if salespeople
        else "全部授权店铺"
    )
    task_type = "、".join(appeal_types) or str(params.get("appeal_type") or "自动申诉")
    mode = "循环" if params.get("mode") == "loop" else "单轮"
    execution_label = (
        str(params.get("agent_name") or "本机 Agent")
        if params.get("execution_target") == "agent"
        else "本机比特浏览器"
        if params.get("execution_target") == "local"
        else "服务器比特浏览器"
    )
    return f"{execution_label}｜{scope}｜{task_type}｜{mode}"


def _update_daily_task_state(task_id, **updates):
    """更新一个任务实例；无 task_id 时保留旧调用的兼容行为。"""
    task_id = str(task_id or "").strip()
    with _daily_task_lock:
        if not task_id:
            _daily_task_state.update(updates)
            return dict(_daily_task_state)
        state = _daily_tasks.get(task_id)
        if state is None:
            return {}
        state.update(updates)
        if str(_daily_task_state.get("task_id") or "") == task_id:
            _daily_task_state.update(updates)
        return dict(state)


def _prune_daily_task_history():
    """只裁剪已结束的旧记录，运行中的任务绝不移除。"""
    removed = []
    with _daily_task_lock:
        if len(_daily_tasks) <= DAILY_TASK_HISTORY_LIMIT:
            return
        for task_id in list(_daily_tasks):
            if len(_daily_tasks) <= DAILY_TASK_HISTORY_LIMIT:
                break
            if not _daily_tasks[task_id].get("running"):
                state = _daily_tasks.pop(task_id, None) or {}
                control = _daily_task_controls.pop(task_id, None) or {}
                removed.append((state.get("log_path"), control.get("stop_manager")))

    # 文件删除和 Manager 关闭都可能阻塞，不能占着状态锁执行。
    base_log_path = Path(_daily_task_log_path)
    for log_path, stop_manager in removed:
        if stop_manager is not None:
            try:
                stop_manager.shutdown()
            except Exception:
                pass
        try:
            target_path = Path(log_path) if log_path else None
            if target_path and target_path != base_log_path:
                target_path.unlink(missing_ok=True)
        except OSError:
            pass


def _daily_task_snapshot(task_id, include_log=True):
    task_id = str(task_id or "").strip()
    with _daily_task_lock:
        state = dict(_daily_tasks.get(task_id) or {})
        if (
            not state
            and task_id == "legacy-daily-task"
            and not _daily_tasks
            and (_daily_task_state.get("running") or _daily_task_state.get("started_at"))
        ):
            state = dict(_daily_task_state)
            state.setdefault("task_id", "legacy-daily-task")
            state.setdefault("name", "daily_task")
            state.setdefault("can_stop", _daily_task_stop_event is not None)
    if not state:
        return {}
    state.setdefault(
        "execution_target",
        "local" if RUNTIME_SETTINGS.is_client else "server",
    )
    if include_log:
        state["log"] = _read_daily_task_log(log_path=state.get("log_path"))
    else:
        state.pop("log", None)
    return state


def _daily_tasks_snapshot():
    with _daily_task_lock:
        task_ids = list(_daily_tasks)
        legacy_state = dict(_daily_task_state)
    # 最新任务在页面顶部；列表只返回摘要，单任务详情才读取独立日志。
    tasks = [
        snapshot
        for task_id in reversed(task_ids)
        if (snapshot := _daily_task_snapshot(task_id, include_log=False))
    ]
    if not tasks and (legacy_state.get("running") or legacy_state.get("started_at")):
        legacy_state.setdefault("task_id", "legacy-daily-task")
        legacy_state.setdefault("name", "daily_task")
        legacy_state.setdefault("can_stop", _daily_task_stop_event is not None)
        legacy_state.setdefault(
            "execution_target",
            "local" if RUNTIME_SETTINGS.is_client else "server",
        )
        legacy_state.pop("log", None)
        tasks.append(legacy_state)

    result = {
        "tasks": tasks,
        "running": any(task.get("running") for task in tasks),
        "running_count": sum(1 for task in tasks if task.get("running")),
        "total_count": len(tasks),
    }
    if tasks:
        # 兼容仍按旧单任务结构读取 status 的客户端。
        compatibility_task = next(
            (task for task in tasks if task.get("running")),
            tasks[0],
        )
        result.update(compatibility_task)
        result["tasks"] = tasks
        result["running"] = any(task.get("running") for task in tasks)
        result["running_count"] = sum(1 for task in tasks if task.get("running"))
        result["total_count"] = len(tasks)
    return result


def _resolve_daily_task_id_for_stop(requested_task_id=""):
    requested_task_id = str(requested_task_id or "").strip()
    with _daily_task_lock:
        if requested_task_id:
            return requested_task_id if requested_task_id in _daily_tasks else ""
        running_ids = [
            task_id
            for task_id, state in _daily_tasks.items()
            if state.get("running")
        ]
    # 兼容旧客户端：只有一个任务运行时允许省略 task_id。
    return running_ids[0] if len(running_ids) == 1 else ""


_mercado_login_task_lock = threading.RLock()
_mercado_login_task_process = None
_mercado_login_task_processes = {}
_mercado_login_tasks = {}
MERCADO_LOGIN_TASK_HISTORY_LIMIT = 100
MERCADO_LOGIN_STOP_GRACE_SECONDS = 8
DEFAULT_MERCADO_LOGIN_WORKERS = 3
MAX_MERCADO_LOGIN_WORKERS = 10
_mercado_login_log_path = CURRENT_DIR / "logs" / "bit_mercado_login_console.log"
MERCADO_LOGIN_SINGLE_MANUAL_WAIT_SECONDS = 20 * 60
MERCADO_LOGIN_SELECTED_MANUAL_WAIT_SECONDS = 20 * 60


def _read_recent_mercado_login_logs(max_bytes=256 * 1024, max_lines=800):
    try:
        with _mercado_login_log_path.open("rb") as log_file:
            log_file.seek(0, os.SEEK_END)
            size = log_file.tell()
            log_file.seek(max(0, size - max_bytes), os.SEEK_SET)
            content = log_file.read().decode("utf-8", errors="replace")
        lines = content.splitlines(keepends=True)[-max_lines:]
        for index in range(len(lines) - 1, -1, -1):
            if "===== bit_mercado_login 启动：" in lines[index]:
                return lines[index:]
        return lines
    except OSError:
        return []


_mercado_login_task_logs = deque(
    _read_recent_mercado_login_logs(),
    maxlen=800,
)
_mercado_login_task_state = {
    "running": False,
    "started_at": "",
    "finished_at": "",
    "status": "idle",
    "message": "等待启动",
    "target": "全部未忽略店铺",
    "window_id": "",
    "window_ids": [],
    "pid": None,
    "returncode": None,
    "log_path": str(_mercado_login_log_path),
}

APPEAL_SITES = ("墨西哥", "巴西", "哥伦比亚", "智利", "阿根廷", "乌拉圭")
MERCADO_AUTH_SITE_NAMES = {
    "MLM": "墨西哥",
    "MX": "墨西哥",
    "MLB": "巴西",
    "BR": "巴西",
    "MCO": "哥伦比亚",
    "CO": "哥伦比亚",
    "MLC": "智利",
    "CL": "智利",
    "MLA": "阿根廷",
    "AR": "阿根廷",
    "MLU": "乌拉圭",
    "UY": "乌拉圭",
}
APPEAL_FORMS = ("延误", "侵权", "禁限售", "取消率", "投诉")
APPEAL_LOOP_COUNTS = (10, 20, 50)
DEFAULT_APPEAL_LOOP_COUNT = 10
PERMANENT_APPEAL_LOOP_COUNT = 0
AI_APPEAL_ROUND_INTERVAL_SECONDS = 60
MANUAL_APPEAL_ROUND_INTERVAL_SECONDS = 600
# 保留旧常量供其他模块调用，默认值等同人工客服的轮次间隔。
APPEAL_ROUND_INTERVAL_SECONDS = MANUAL_APPEAL_ROUND_INTERVAL_SECONDS
APPEAL_STREAM_HEARTBEAT_SECONDS = 15
_appeal_task_lock = threading.Lock()
_appeal_tasks = {}


def normalize_appeal_task_id(task_id):
    task_id = str(task_id or "").strip()[:96]
    if not task_id:
        return ""
    allowed = set("abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-_")
    return task_id if all(char in allowed for char in task_id) else ""


def register_appeal_task(task_id, metadata=None):
    task_id = normalize_appeal_task_id(task_id)
    if not task_id:
        raise ValueError("任务编号格式无效")
    with _appeal_task_lock:
        if task_id in _appeal_tasks:
            return None
        stop_event = threading.Event()
        _appeal_tasks[task_id] = {
            "stop_event": stop_event,
            "status": "running",
            "started_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            **(metadata or {}),
        }
        return stop_event


def request_appeal_task_stop(task_id):
    task_id = normalize_appeal_task_id(task_id)
    with _appeal_task_lock:
        task = _appeal_tasks.get(task_id)
        if task is None:
            return False
        task["status"] = "stopping"
        task["stop_event"].set()
        return True


def finish_appeal_task(task_id):
    task_id = normalize_appeal_task_id(task_id)
    with _appeal_task_lock:
        _appeal_tasks.pop(task_id, None)


class ThreadLogStream:
    def __init__(self, original_stream):
        self.original_stream = original_stream

    def write(self, text):
        if text:
            if isinstance(text, bytes):
                encoding = getattr(self.original_stream, "encoding", None) or "utf-8"
                text = text.decode(encoding, errors="replace")
            with _thread_log_lock:
                output_queue = _thread_log_queues.get(threading.get_ident())
            if output_queue:
                output_queue.put(text)
            else:
                self.original_stream.write(text)

    def flush(self):
        self.original_stream.flush()

    def isatty(self):
        return self.original_stream.isatty()

    @property
    def encoding(self):
        return getattr(self.original_stream, "encoding", "utf-8")

    def __getattr__(self, name):
        return getattr(self.original_stream, name)


sys.stdout = ThreadLogStream(_original_stdout)
sys.stderr = ThreadLogStream(_original_stderr)


def register_thread_log_queue(output_queue):
    with _thread_log_lock:
        _thread_log_queues[threading.get_ident()] = output_queue


def unregister_thread_log_queue():
    with _thread_log_lock:
        _thread_log_queues.pop(threading.get_ident(), None)


def _public_mercado_login_task(task):
    return {
        key: value
        for key, value in dict(task or {}).items()
        if key not in ("scope_keys", "log_chunks", "stop_worker_started")
    }


def _append_mercado_login_task_log(text, task_id=""):
    text = format_log_text(text)
    if not text:
        return
    if not text.endswith("\n"):
        text += "\n"
    with _mercado_login_task_lock:
        _mercado_login_task_logs.append(text)
        task = _mercado_login_tasks.get(str(task_id or ""))
        if task is not None:
            task.setdefault("log_chunks", deque(maxlen=800)).append(text)
        _mercado_login_log_path.parent.mkdir(parents=True, exist_ok=True)
        with _mercado_login_log_path.open("a", encoding="utf-8") as log_file:
            log_file.write(text)


def _terminate_mercado_login_process_tree(process):
    """终止登录控制进程及它创建的并发 worker。"""
    if process is None or process.poll() is not None:
        return

    used_process_group = False
    try:
        if os.name == "nt":
            subprocess.run(
                ["taskkill", "/PID", str(process.pid), "/T"],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                check=False,
            )
        else:
            process_group_id = os.getpgid(process.pid)
            if process_group_id == process.pid:
                os.killpg(process_group_id, signal.SIGTERM)
                used_process_group = True
            else:
                process.terminate()
    except (OSError, ProcessLookupError):
        return

    deadline = time.monotonic() + MERCADO_LOGIN_STOP_GRACE_SECONDS
    while process.poll() is None and time.monotonic() < deadline:
        time.sleep(0.1)
    if process.poll() is not None:
        return

    try:
        if os.name == "nt":
            subprocess.run(
                ["taskkill", "/PID", str(process.pid), "/T", "/F"],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                check=False,
            )
        elif used_process_group:
            os.killpg(process.pid, signal.SIGKILL)
        else:
            process.kill()
    except (OSError, ProcessLookupError):
        pass


def _stop_mercado_login_process_worker(task_id, process):
    _terminate_mercado_login_process_tree(process)
    with _mercado_login_task_lock:
        task = dict(_mercado_login_tasks.get(task_id) or {})
    for window_id in task.get("window_ids", ()):
        try:
            close_result = closeBrowser(window_id)
            if isinstance(close_result, dict) and close_result.get("success") is False:
                _append_mercado_login_task_log(
                    f"{get_now_time()} 停止任务后关闭窗口 {window_id} 失败："
                    f"{close_result.get('msg') or close_result}",
                    task_id=task_id,
                )
        except Exception as exc:
            _append_mercado_login_task_log(
                f"{get_now_time()} 停止任务后关闭窗口 {window_id} 失败：{exc}",
                task_id=task_id,
            )


def _start_mercado_login_stop_worker(task_id, process):
    if process is None:
        return
    with _mercado_login_task_lock:
        task = _mercado_login_tasks.get(task_id)
        if task is None or task.get("stop_worker_started"):
            return
        task["stop_worker_started"] = True
    threading.Thread(
        target=_stop_mercado_login_process_worker,
        args=(task_id, process),
        daemon=True,
    ).start()


def request_mercado_login_task_stop(task_id):
    task_id = normalize_appeal_task_id(task_id)
    with _mercado_login_task_lock:
        task = _mercado_login_tasks.get(task_id)
        if task is None or not task.get("running"):
            return False, _mercado_login_task_snapshot()
        task.update(
            {
                "status": "stopping",
                "message": f"{task.get('target') or '当前'} 登录任务正在停止",
                "stop_requested": True,
            }
        )
        process = _mercado_login_task_processes.get(task_id)
    _append_mercado_login_task_log(
        f"{get_now_time()} 已收到停止请求，正在终止登录任务及子进程",
        task_id=task_id,
    )
    _start_mercado_login_stop_worker(task_id, process)
    return True, _mercado_login_task_snapshot()


def request_mercado_login_window_tasks_stop(window_id):
    """停止指定窗口对应的独立登录任务，不影响其他窗口。"""
    window_id = str(window_id or "").strip()
    if not window_id:
        return []
    with _mercado_login_task_lock:
        task_ids = [
            task_id
            for task_id, task in _mercado_login_tasks.items()
            if task.get("running")
            and window_id in {
                str(item or "").strip()
                for item in task.get("window_ids", ())
            }
            and len(task.get("window_ids", ())) == 1
        ]
    stopped_task_ids = []
    for task_id in task_ids:
        stopped, _ = request_mercado_login_task_stop(task_id)
        if stopped:
            stopped_task_ids.append(task_id)
    return stopped_task_ids


def _mercado_login_task_snapshot():
    with _mercado_login_task_lock:
        completed_task_ids = [
            task_id
            for task_id, task in _mercado_login_tasks.items()
            if not task.get("running")
        ]
        for completed_task_id in completed_task_ids[:-MERCADO_LOGIN_TASK_HISTORY_LIMIT]:
            _mercado_login_tasks.pop(completed_task_id, None)
        active_tasks = [
            task for task in _mercado_login_tasks.values() if task.get("running")
        ]
        latest_task = next(reversed(_mercado_login_tasks.values()), None) if _mercado_login_tasks else None
        current_task = active_tasks[-1] if active_tasks else latest_task
        if active_tasks:
            primary = active_tasks[-1]
            if len(active_tasks) == 1:
                _mercado_login_task_state.update(_public_mercado_login_task(primary))
            else:
                active_window_ids = list(
                    dict.fromkeys(
                        window_id
                        for task in active_tasks
                        for window_id in task.get("window_ids", ())
                    )
                )
                _mercado_login_task_state.update(
                    {
                        "running": True,
                        "started_at": min(
                            str(task.get("started_at") or "") for task in active_tasks
                        ),
                        "finished_at": "",
                        "status": "running",
                        "message": f"{len(active_tasks)} 个登录任务正在异步执行",
                        "target": f"{len(active_tasks)} 个登录任务",
                        "window_id": "",
                        "window_ids": active_window_ids,
                        "pid": None,
                        "returncode": None,
                    }
                )
        elif latest_task:
            _mercado_login_task_state.update(_public_mercado_login_task(latest_task))

        public_tasks = [
            _public_mercado_login_task(task)
            for task in reversed(tuple(_mercado_login_tasks.values()))
        ]
        active_window_ids = list(
            dict.fromkeys(
                window_id
                for task in active_tasks
                for window_id in task.get("window_ids", ())
            )
        )
        all_running = any(task.get("scope") == "all" for task in active_tasks)
        latest_log = (
            "".join(current_task.get("log_chunks") or ())
            if current_task is not None
            else "".join(_mercado_login_task_logs)
        )
        snapshot = {
            **dict(_mercado_login_task_state),
            "log": latest_log,
            "tasks": public_tasks,
            "running_count": len(active_tasks),
            "active_window_ids": active_window_ids,
            "all_running": all_running,
            "current_task_id": (
                str(current_task.get("task_id") or "")
                if current_task is not None and current_task.get("running")
                else ""
            ),
            "current_task_target": (
                str(current_task.get("target") or "")
                if current_task is not None
                else ""
            ),
            "current_task_stopping": bool(
                current_task is not None and current_task.get("stop_requested")
            ),
            "can_stop": bool(current_task is not None and current_task.get("running")),
        }
    process_owner = get_lock_owner(MERCADO_LOGIN_JOB_LOCK_KEY)
    if process_owner:
        local_pids = {
            task.get("pid") for task in active_tasks if task.get("pid") is not None
        }
        owner_is_local_task = process_owner.get("pid") in local_pids
        snapshot["all_running"] = True
        if not owner_is_local_task:
            target = str(
                (process_owner.get("metadata") or {}).get("target")
                or process_owner.get("owner")
                or "现有登录检测任务"
            )
            snapshot.update(
                {
                    "running": True,
                    "status": "running",
                    "message": f"{target} 正在另一个进程中运行",
                    "target": target,
                    "pid": process_owner.get("pid"),
                    "running_count": int(snapshot.get("running_count") or 0) + 1,
                }
            )
    return snapshot


def _mercado_login_task_scope(shop_name="", window_id="", selected_shops=None):
    selected_shops = tuple(selected_shops or ())
    window_ids = list(
        dict.fromkeys(
            str(value or "").strip()
            for value in (
                [shop.get("window_id") for shop in selected_shops]
                if selected_shops
                else [window_id]
            )
            if str(value or "").strip()
        )
    )
    if window_ids:
        return "windows", window_ids, tuple(f"window:{value}" for value in window_ids)
    shop_name = str(shop_name or "").strip()
    if shop_name:
        return "shop", [], (f"shop:{shop_name}",)
    return "all", [], ("all",)


def _build_mercado_login_command(
    shop_name="",
    workers=DEFAULT_MERCADO_LOGIN_WORKERS,
    window_ids=None,
):
    command = [
        sys.executable,
        "-u",
        "-m",
        "bit.bit_mercado_login",
    ]
    shop_name = str(shop_name or "").strip()
    selected_window_ids = tuple(
        dict.fromkeys(
            str(window_id or "").strip()
            for window_id in (window_ids or ())
            if str(window_id or "").strip()
        )
    )
    if shop_name:
        command.extend(
            (
                "--shop",
                shop_name,
                "--auto-login",
                "--keep-browser-open",
                "--manual-login-wait-seconds",
                str(MERCADO_LOGIN_SINGLE_MANUAL_WAIT_SECONDS),
            )
        )
    elif selected_window_ids:
        worker_count = max(
            1,
            min(
                int(workers or DEFAULT_MERCADO_LOGIN_WORKERS),
                MAX_MERCADO_LOGIN_WORKERS,
                len(selected_window_ids),
            ),
        )
        for window_id in selected_window_ids:
            command.extend(("--window-id", window_id))
        command.extend(
            (
                "--workers",
                str(worker_count),
                "--no-email",
                "--keep-browser-open",
                "--manual-login-wait-seconds",
                str(MERCADO_LOGIN_SELECTED_MANUAL_WAIT_SECONDS),
            )
        )
    else:
        command.extend(
            (
                "--all-active-login",
                "--workers",
                str(
                    max(
                        1,
                        min(
                            int(workers or DEFAULT_MERCADO_LOGIN_WORKERS),
                            MAX_MERCADO_LOGIN_WORKERS,
                        ),
                    )
                ),
            )
        )
    command.extend(("--wait-seconds", "60", "--page-load-timeout", "20"))
    return command


def _mercado_login_selected_target(selected_shops):
    shops = tuple(selected_shops or ())
    if not shops:
        return ""
    names = [
        str(shop.get("window_name") or shop.get("shop_name") or "").strip()
        for shop in shops
    ]
    names = [name for name in names if name]
    preview = "、".join(names[:3])
    if len(names) > 3:
        preview += "等"
    return f"所选 {len(shops)} 家待处理登录异常店铺" + (f"（{preview}）" if preview else "")


def run_mercado_login_console_job(
    shop_name="",
    window_id="",
    workers=DEFAULT_MERCADO_LOGIN_WORKERS,
    selected_shops=None,
    task_id="",
):
    """后台运行登录任务，并把子进程及其工作进程输出持久化给控制台。"""
    global _mercado_login_task_process
    selected_shops = tuple(dict(shop) for shop in (selected_shops or ()))
    selected_window_ids = [shop.get("window_id") for shop in selected_shops]
    target = (
        _mercado_login_selected_target(selected_shops)
        or str(shop_name or "").strip()
        or "全部未忽略店铺"
    )
    command = _build_mercado_login_command(
        shop_name=shop_name,
        workers=workers,
        window_ids=selected_window_ids,
    )
    _append_mercado_login_task_log(
        f"{get_now_time()} ===== bit_mercado_login 启动：{target} =====",
        task_id=task_id,
    )
    try:
        child_env = os.environ.copy()
        child_env["PYTHONIOENCODING"] = "utf-8"
        child_env["PYTHONUNBUFFERED"] = "1"
        creationflags = (
            getattr(subprocess, "CREATE_NO_WINDOW", 0) if os.name == "nt" else 0
        )
        process = subprocess.Popen(
            command,
            cwd=str(PROJECT_ROOT),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
            errors="replace",
            bufsize=1,
            env=child_env,
            creationflags=creationflags,
            start_new_session=(os.name != "nt"),
        )
        with _mercado_login_task_lock:
            _mercado_login_task_process = process
            _mercado_login_task_processes[task_id] = process
            if task_id in _mercado_login_tasks:
                _mercado_login_tasks[task_id]["pid"] = process.pid
                stop_requested = bool(
                    _mercado_login_tasks[task_id].get("stop_requested")
                )
            else:
                stop_requested = False
        if stop_requested:
            _start_mercado_login_stop_worker(task_id, process)
        if process.stdout is not None:
            for line in process.stdout:
                _append_mercado_login_task_log(
                    f"[{target}] {line}",
                    task_id=task_id,
                )
        returncode = process.wait()
        with _mercado_login_task_lock:
            stop_requested = bool(
                (_mercado_login_tasks.get(task_id) or {}).get("stop_requested")
            )
        succeeded = returncode == 0 and not stop_requested
        resolve_message = ""
        if succeeded and window_id:
            try:
                db_resolve_window_anomaly(window_id)
                resolve_message = "，店铺待登录状态已自动解除"
            except Exception as exc:
                resolve_message = f"，但更新店铺状态失败：{exc}"
        if stop_requested:
            message = f"{target} 登录任务已停止"
        elif succeeded:
            message = f"{target} 登录任务完成{resolve_message}"
        else:
            message = f"{target} 登录任务失败，退出码：{returncode}"
        with _mercado_login_task_lock:
            if task_id in _mercado_login_tasks:
                _mercado_login_tasks[task_id].update({
                    "running": False,
                    "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "status": (
                        "stopped"
                        if stop_requested
                        else "success" if succeeded else "error"
                    ),
                    "message": message,
                    "returncode": returncode,
                })
        _append_mercado_login_task_log(
            f"{get_now_time()} {message}",
            task_id=task_id,
        )
    except Exception as exc:
        logging.error("bit_mercado_login console job failed: %s", exc)
        traceback.print_exc()
        with _mercado_login_task_lock:
            stop_requested = bool(
                (_mercado_login_tasks.get(task_id) or {}).get("stop_requested")
            )
            if task_id in _mercado_login_tasks:
                _mercado_login_tasks[task_id].update({
                    "running": False,
                    "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "status": "stopped" if stop_requested else "error",
                    "message": (
                        f"{target} 登录任务已停止" if stop_requested else str(exc)
                    ),
                    "returncode": None,
                })
        _append_mercado_login_task_log(
            (
                f"{get_now_time()} {target} 登录任务已停止"
                if stop_requested
                else f"{get_now_time()} bit_mercado_login 启动失败：{exc}"
            ),
            task_id=task_id,
        )
    finally:
        with _mercado_login_task_lock:
            _mercado_login_task_processes.pop(task_id, None)
            _mercado_login_task_process = next(
                iter(_mercado_login_task_processes.values()),
                None,
            )


def start_mercado_login_console_job(
    shop_name="",
    window_id="",
    workers=DEFAULT_MERCADO_LOGIN_WORKERS,
    selected_shops=None,
):
    selected_shops = tuple(dict(shop) for shop in (selected_shops or ()))
    workers = max(
        1,
        min(
            int(workers or DEFAULT_MERCADO_LOGIN_WORKERS),
            MAX_MERCADO_LOGIN_WORKERS,
        ),
    )
    scope, selected_window_ids, scope_keys = _mercado_login_task_scope(
        shop_name=shop_name,
        window_id=window_id,
        selected_shops=selected_shops,
    )
    with _mercado_login_task_lock:
        active_tasks = [
            task for task in _mercado_login_tasks.values() if task.get("running")
        ]
        active_all_task = any(task.get("scope") == "all" for task in active_tasks)
        if active_all_task or (scope == "all" and active_tasks):
            snapshot = _mercado_login_task_snapshot()
            snapshot["message"] = "全部店铺自动登录任务与单店任务不能同时运行"
            return False, snapshot
        if get_lock_owner(MERCADO_LOGIN_JOB_LOCK_KEY):
            snapshot = _mercado_login_task_snapshot()
            return False, snapshot
        occupied_scope_keys = {
            key for task in active_tasks for key in task.get("scope_keys", ())
        }
        overlapping_keys = occupied_scope_keys.intersection(scope_keys)
        if overlapping_keys:
            snapshot = _mercado_login_task_snapshot()
            snapshot["message"] = "所选店铺已有自动登录任务正在运行"
            return False, snapshot
        target = (
            _mercado_login_selected_target(selected_shops)
            or str(shop_name or "").strip()
            or "全部未忽略店铺"
        )
        task_id = secrets.token_hex(8)
        task_state = {
            "task_id": task_id,
            "running": True,
            "started_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "finished_at": "",
            "status": "running",
            "message": f"{target} 登录任务已启动",
            "target": target,
            "window_id": str(window_id or "").strip(),
            "window_ids": selected_window_ids,
            "workers": (
                min(workers, len(selected_window_ids))
                if selected_window_ids
                else workers
            ),
            "pid": None,
            "returncode": None,
            "stop_requested": False,
            "stop_worker_started": False,
            "scope": scope,
            "scope_keys": scope_keys,
            "log_chunks": deque(maxlen=800),
        }
        _mercado_login_tasks[task_id] = task_state
        task_thread = threading.Thread(
            target=run_mercado_login_console_job,
            args=(shop_name, window_id, workers, selected_shops, task_id),
            daemon=True,
        )
        try:
            task_thread.start()
        except Exception:
            _mercado_login_tasks.pop(task_id, None)
            raise
        snapshot = _mercado_login_task_snapshot()
        snapshot["started_task_id"] = task_id
        return True, snapshot


def _collection_status_failed(value):
    text = str(value or "").strip()
    return bool(text) and text not in ("成功", "无数据", "未知")


def _failed_collection_shop_options(shop_options, status_rows):
    """按店铺聚合失败站点；任一站点失败时店铺只出现一次。"""

    configured = {
        str(shop.get("shop_name") or "").strip(): shop
        for shop in (shop_options or [])
        if str(shop.get("shop_name") or "").strip()
    }
    failures_by_shop = {}
    for row in status_rows or []:
        shop_name = str(row.get("店铺名") or row.get("shop_name") or "").strip()
        site = str(row.get("站点") or row.get("site") or "").strip()
        status = str(row.get("状态") or row.get("status") or "").strip()
        shop = configured.get(shop_name)
        if (
            shop is None
            or not site
            or site not in (shop.get("sites") or [])
            or not _collection_status_failed(status)
        ):
            continue
        failed_shop = failures_by_shop.setdefault(
            shop_name,
            {
                "shop_name": shop_name,
                "salesperson": str(shop.get("salesperson") or "").strip(),
                "sites": [],
                "failures": [],
            },
        )
        if site not in failed_shop["sites"]:
            failed_shop["sites"].append(site)
        failure = {
            "site": site,
            "status": status,
            "status_time": str(
                row.get("状态时间") or row.get("status_time") or ""
            ),
        }
        if failure not in failed_shop["failures"]:
            failed_shop["failures"].append(failure)

    return [
        failures_by_shop[shop["shop_name"]]
        for shop in (shop_options or [])
        if shop.get("shop_name") in failures_by_shop
    ]


def _authorization_flag_enabled(value):
    if isinstance(value, str):
        return value.strip().casefold() not in ("", "0", "false", "no", "off")
    return bool(value)


def _authorized_task_shop_options(flag_name, token_data=None):
    """直接从店铺授权开关生成任务店铺和站点选项。"""
    if token_data is None:
        token_data = bit_db_api.list_mercado_store_tokens() or {}

    shops_by_name = {}
    for token in token_data.get("rows") or ():
        shop_name = str(
            token.get("display_name") or token.get("nickname") or ""
        ).strip()
        if not shop_name:
            continue
        enabled_sites = []
        relevant_settings = []
        for raw_setting in token.get("site_settings") or ():
            setting = dict(raw_setting or {})
            if not _authorization_flag_enabled(setting.get(flag_name)):
                continue
            site_name = MERCADO_AUTH_SITE_NAMES.get(
                str(setting.get("site_id") or "").strip().upper()
            )
            if not site_name:
                continue
            relevant_settings.append(setting)
            if site_name not in enabled_sites:
                enabled_sites.append(site_name)
        if not enabled_sites:
            continue

        shop = shops_by_name.setdefault(
            shop_name,
            {
                "shop_name": shop_name,
                "salesperson": next(
                    (
                        str(setting.get("salesperson") or "").strip()
                        for setting in relevant_settings
                        if str(setting.get("salesperson") or "").strip()
                    ),
                    "",
                ),
                "sites": [],
            },
        )
        for site_name in enabled_sites:
            if site_name not in shop["sites"]:
                shop["sites"].append(site_name)
    return list(shops_by_name.values())


def _collection_config_options(include_failures=False):
    token_data = bit_db_api.list_mercado_store_tokens() or {}
    infraction_shop_options = _authorized_task_shop_options(
        "visit_stats_enabled",
        token_data=token_data,
    )
    reputation_shop_options = _authorized_task_shop_options(
        "reputation_update_enabled",
        token_data=token_data,
    )
    appeal_shop_options = _authorized_task_shop_options(
        "appeal_enabled",
        token_data=token_data,
    )
    site_order = []
    for shop in reputation_shop_options:
        for site in shop.get("sites") or ():
            if site not in site_order:
                site_order.append(site)
    result = {
        "shops": reputation_shop_options,
        "sites": site_order,
        "infraction_shops": infraction_shop_options,
        "appeal_shops": appeal_shop_options,
    }
    if not include_failures:
        return result

    try:
        infraction_data = db_get_latest_infraction_info(30) or {}
        infraction_status_rows = infraction_data.get("summary") or []
    except Exception as exc:
        logging.warning("读取侵权失败店铺失败：%s", exc)
        infraction_status_rows = []

    try:
        reputation_data = db_get_latest_reputation_info() or {}
        reputation_status_rows = reputation_data.get("summary") or []
        if not reputation_status_rows:
            reputation_status_rows = [
                {
                    "店铺名": row.get("店铺名"),
                    "站点": row.get("站点"),
                    "状态": "失败",
                    "状态时间": row.get("更新时间") or row.get("提交时间"),
                }
                for row in (reputation_data.get("rows") or [])
                if "失败" in str(row.get("声誉颜色") or "")
                or str(row.get("系统告警") or "").strip().startswith("失败")
            ]
    except Exception as exc:
        logging.warning("读取声誉失败店铺失败：%s", exc)
        reputation_status_rows = []

    result["failed_shops"] = {
        "infraction": _failed_collection_shop_options(
            infraction_shop_options,
            infraction_status_rows,
        ),
        "reputation": _failed_collection_shop_options(
            reputation_shop_options,
            reputation_status_rows,
        ),
    }
    return result


def _normalized_collection_list(data, key):
    if key not in data:
        return ()
    raw_values = data.get(key)
    if not isinstance(raw_values, list):
        raise ValueError(f"{key} 必须是数组")
    values = tuple(
        dict.fromkeys(
            str(value or "").strip()
            for value in raw_values
            if str(value or "").strip()
        )
    )
    if not values:
        label = "店铺" if key == "shops" else "站点"
        raise ValueError(f"请至少选择一个{label}")
    return values


def _parse_collection_request(data, authorization_flag="visit_stats_enabled"):
    data = data if isinstance(data, dict) else {}
    shops = _normalized_collection_list(data, "shops")
    sites = _normalized_collection_list(data, "sites")
    max_workers = _parse_int_param(
        data,
        "max_workers",
        DEFAULT_COLLECTION_MAX_WORKERS,
        min_value=1,
        max_value=10,
    )
    configured_options = _authorized_task_shop_options(authorization_flag)
    configured = {shop["shop_name"]: shop for shop in configured_options}
    unknown_shops = [shop for shop in shops if shop not in configured]
    if unknown_shops:
        raise ValueError("店铺不存在或已被忽略：" + "、".join(unknown_shops))
    configured_sites = {
        site
        for shop in configured_options
        for site in (shop.get("sites") or ())
    }
    unknown_sites = [site for site in sites if site not in configured_sites]
    if unknown_sites:
        raise ValueError("站点不存在：" + "、".join(unknown_sites))

    target_shops = shops or tuple(configured)
    matching_shop_names = [
        shop_name
        for shop_name in target_shops
        if any(
            not sites or site in sites
            for site in configured[shop_name]["sites"]
        )
    ]
    matching_sites = {
        site
        for shop_name in matching_shop_names
        for site in configured[shop_name]["sites"]
        if not sites or site in sites
    }
    if not matching_sites:
        raise ValueError("所选店铺没有配置所选站点")
    return {
        "selected_shops": shops,
        "selected_sites": sites,
        "max_workers": max_workers,
        "target": (
            f"{len(matching_shop_names)} 家店铺"
        ) + " / " + (
            f"{len(sites)} 个站点" if sites else "全部站点"
        ),
    }


def run_infraction_collect_job(
    selected_shops=None,
    selected_sites=None,
    max_workers=DEFAULT_COLLECTION_MAX_WORKERS,
):
    try:
        print(
            f"{get_now_time()} 开始执行侵权数据采集："
            f"{len(selected_shops or ()) or '全部'} 家店铺，"
            f"{len(selected_sites or ()) or '全部'} 个站点，并发 {max_workers}<br>"
        )
        target = getattr(bit_infractions_info, "main", None) or bit_infractions_info.get_infractions_info_all
        result = target(
            max_workers=max_workers,
            selected_shops=selected_shops,
            selected_sites=selected_sites,
        )
        failed_count = sum(
            1
            for row in (result or {}).get("results", [])
            if len(row) >= 4 and str(row[3] or "") != "成功"
        )
        with _infraction_collect_lock:
            _infraction_collect_state.update({
                "running": False,
                "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "status": "success",
                "message": f"侵权数据采集完成，异常站点 {failed_count} 个",
            })
        print(f"{get_now_time()} 侵权数据采集完成<br>")
    except Exception as e:
        logging.error("Infraction collect failed: %s", e)
        traceback.print_exc()
        with _infraction_collect_lock:
            _infraction_collect_state.update({
                "running": False,
                "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "status": "error",
                "message": str(e),
            })


def run_reputation_collect_job(
    selected_shops=None,
    selected_sites=None,
    max_workers=DEFAULT_COLLECTION_MAX_WORKERS,
):
    try:
        with _reputation_collect_lock:
            operation = _reputation_collect_state.get("operation")
        action_label = (
            "所选店铺声誉更新"
            if operation == "selected_update"
            else "声誉数据补跑"
        )
        print(
            f"{get_now_time()} 开始执行{action_label}："
            f"{len(selected_shops or ()) or '全部'} 家店铺，"
            f"{len(selected_sites or ()) or '全部'} 个站点，并发 {max_workers}<br>"
        )
        target = getattr(bit_reputation_info, "main", None) or bit_reputation_info.get_reputation_info_all
        result = target(
            max_workers=max_workers,
            selected_shops=selected_shops,
            selected_sites=selected_sites,
        )
        failed_count = sum(
            1
            for row in (result or {}).get("results", [])
            if len(row) >= 4 and str(row[3] or "") != "成功"
        )
        with _reputation_collect_lock:
            _reputation_collect_state.update({
                "running": False,
                "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "status": "success",
                "message": f"{action_label}完成，异常站点 {failed_count} 个",
            })
        print(f"{get_now_time()} {action_label}完成<br>")
    except Exception as e:
        logging.error("Reputation collect failed: %s", e)
        traceback.print_exc()
        with _reputation_collect_lock:
            _reputation_collect_state.update({
                "running": False,
                "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "status": "error",
                "message": str(e),
            })


def run_fund_collect_job(
    all_shops=True,
    selected_window_ids=None,
    salesperson="",
    max_workers=DEFAULT_COLLECTION_MAX_WORKERS,
    stop_event=None,
):
    global _fund_collect_stop_event
    try:
        rows = bit_pago_info.get_pago_info_all(
            max_workers=max_workers,
            apply_shop_limit=False,
            selected_window_ids=None if all_shops else selected_window_ids,
            salesperson=salesperson,
            stop_event=stop_event,
        )
        collected_count = len(rows or [])
        with _fund_collect_lock:
            stopped = bool(stop_event is not None and stop_event.is_set())
            _fund_collect_state.update(
                {
                    "running": False,
                    "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "status": "stopped" if stopped else "success",
                    "message": (
                        f"资金数据采集已终止，已保留 {collected_count} 个店铺站点结果"
                        if stopped
                        else f"资金数据采集完成，共更新 {collected_count} 个店铺站点"
                    ),
                    "collected_count": collected_count,
                }
            )
    except Exception as e:
        logging.error("Fund collect failed: %s", e)
        traceback.print_exc()
        with _fund_collect_lock:
            stopped = bool(stop_event is not None and stop_event.is_set())
            _fund_collect_state.update({
                "running": False,
                "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "status": "stopped" if stopped else "error",
                "message": "资金数据采集已终止" if stopped else str(e),
            })
    finally:
        with _fund_collect_lock:
            if _fund_collect_stop_event is stop_event:
                _fund_collect_stop_event = None


def _parse_int_param(data, name, default, min_value=0, max_value=None):
    try:
        value = int(data.get(name, default))
    except (TypeError, ValueError):
        value = default
    value = max(min_value, value)
    if max_value is not None:
        value = min(max_value, value)
    return value


def _parse_bool_param(data, name, default=True):
    value = data.get(name, default)
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in ("1", "true", "yes", "on", "是", "启用")


def build_zying_collection_params(data):
    """校验智赢产品采集页面提交的参数。"""
    data = data if isinstance(data, dict) else {}
    start_page = _parse_int_param(
        data,
        "start_page",
        bit_zying_caiji.DEFAULT_ZYING_START_PAGE,
        min_value=1,
        max_value=10000,
    )
    end_page = _parse_int_param(
        data,
        "end_page",
        max(start_page, bit_zying_caiji.DEFAULT_ZYING_PAGE_COUNT),
        min_value=1,
        max_value=10000,
    )
    if start_page > end_page:
        raise ValueError(f"起始页 {start_page} 不能大于结束页 {end_page}")
    window_id = str(data.get("window_id") or "").strip()[:128]
    params = {
        "number": end_page,
        "window_id": window_id or bit_zying_caiji.DEFAULT_ZYING_WINDOW_ID,
        "start_page": start_page,
        "category": str(data.get("category") or "").strip()[:1024] or None,
    }
    if "category_name" in data:
        params["category_name"] = str(data.get("category_name") or "").strip()[:1024]
    # browser_type/window_name 是工作台的新参数；未提交时维持旧接口返回结构，
    # 兼容仍按 BitBrowser 窗口 ID 调用的脚本和客户端。
    if "browser_type" in data or "window_name" in data:
        params.update(
            {
                "browser_type": bit_zying_caiji.normalize_zying_browser_type(
                    data.get("browser_type")
                ),
                "window_name": str(data.get("window_name") or "").strip()[:256],
            }
        )
    return params


def build_zying_login_params(data):
    data = data if isinstance(data, dict) else {}
    window_id = str(data.get("window_id") or "").strip()[:128]
    return {
        "browser_type": bit_zying_caiji.normalize_zying_browser_type(
            data.get("browser_type")
        ),
        "window_id": window_id or bit_zying_caiji.DEFAULT_ZYING_WINDOW_ID,
        "window_name": str(data.get("window_name") or "").strip()[:256],
    }


def _append_zying_collection_log(message):
    text = format_log_text(message).strip()
    if not text:
        return
    timestamp = datetime.now().strftime("%H:%M:%S")
    with _zying_collection_state_lock:
        for line in text.splitlines():
            line = line.strip()
            if not line:
                continue
            _zying_collection_logs.append(f"[{timestamp}] {line}")
            if _zying_collection_state.get("running"):
                _zying_collection_state["message"] = line


class _ZyingCollectionLogSink:
    def __init__(self):
        self.buffer = ""

    def put(self, text):
        self.buffer += format_log_text(text)
        while "\n" in self.buffer:
            line, self.buffer = self.buffer.split("\n", 1)
            _append_zying_collection_log(line)

    def flush(self):
        if self.buffer.strip():
            _append_zying_collection_log(self.buffer)
        self.buffer = ""


def run_zying_collection_job(params, task_lock):
    """在后台执行智赢产品采集，并保留控制台实时日志。"""
    log_sink = _ZyingCollectionLogSink()
    register_thread_log_queue(log_sink)
    try:
        result = bit_zying_caiji.collect_zying_products(
            **params,
            stop_event=_zying_collection_stop_event,
            product_writer=db_insert_zying_product_info,
            existing_product_id_reader=db_get_existing_zying_product_ids,
            product_mirror_writer=db_upsert_zying_products_to_products,
            return_summary=True,
        )
        summary = {
            key: int((result or {}).get(key) or 0)
            for key in (
                "collected_count",
                "inserted_count",
                "skipped_existing_count",
                "duplicate_count",
                "detail_failed_count",
            )
        }
        completion_message = (
            f"智赢产品采集完成：入库 {summary['inserted_count']} 条，"
            f"已有产品跳过 {summary['skipped_existing_count']} 条，"
            f"页面重复跳过 {summary['duplicate_count']} 条"
        )
        _append_zying_collection_log(completion_message)
        with _zying_collection_state_lock:
            _zying_collection_state.update(
                {
                    "running": False,
                    "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "status": "success",
                    "message": completion_message,
                    "summary": summary,
                    "requires_login": False,
                }
            )
    except bit_zying_caiji.ZyingCollectionStopped as exc:
        stopped_message = str(exc) or "智赢产品采集已由用户结束"
        _append_zying_collection_log(stopped_message)
        with _zying_collection_state_lock:
            _zying_collection_state.update(
                {
                    "running": False,
                    "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "status": "stopped",
                    "message": stopped_message,
                    "summary": {},
                    "requires_login": False,
                }
            )
    except Exception as exc:
        logging.error("智赢产品采集失败：%s", exc)
        traceback.print_exc()
        _append_zying_collection_log(f"采集失败：{exc}")
        requires_login = isinstance(exc, bit_zying_caiji.ZyingAuthenticationError)
        with _zying_collection_state_lock:
            _zying_collection_state.update(
                {
                    "running": False,
                    "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "status": "error",
                    "message": str(exc),
                    "summary": {},
                    "requires_login": requires_login,
                }
            )
    finally:
        log_sink.flush()
        unregister_thread_log_queue()
        task_lock.release()


def build_risk_check_params(data):
    """校验泽顺控制台提交的侵权检测参数。"""
    data = data if isinstance(data, dict) else {}
    category = str(data.get("category") or "").strip()[:1024]
    model = str(data.get("model") or "").strip()[:128] or None
    return {
        "zying_category": category or None,
        "hours": _parse_int_param(data, "hours", 0, min_value=0, max_value=87600),
        "limit": _parse_int_param(data, "limit", 0, min_value=0, max_value=50000),
        "batch_size": _parse_int_param(
            data,
            "batch_size",
            bit_check_risk.DEFAULT_BATCH_SIZE,
            min_value=1,
            max_value=50,
        ),
        "model": model,
        "retries": _parse_int_param(
            data,
            "ai_retries",
            bit_check_risk.DEFAULT_AI_RETRIES,
            min_value=0,
            max_value=5,
        ),
        "recheck": _parse_bool_param(data, "recheck", False),
        "dry_run": _parse_bool_param(data, "dry_run", False),
    }


def _append_risk_check_log(message):
    text = str(message or "").strip()
    if not text:
        return
    timestamp = datetime.now().strftime("%H:%M:%S")
    with _risk_check_state_lock:
        for line in text.splitlines():
            line = line.strip()
            if line:
                _risk_check_logs.append(f"[{timestamp}] {line}")
                if _risk_check_state.get("running"):
                    _risk_check_state["message"] = line


def run_risk_check_job(params, task_lock):
    """在后台线程执行侵权检测，并更新控制台任务状态。"""
    try:
        summary = bit_check_risk.scan_products(
            **params,
            candidate_reader=db_get_zying_risk_candidates,
            risk_writer=db_update_zying_product_risks,
            log_callback=_append_risk_check_log,
        )
        public_summary = {
            key: int(summary.get(key) or 0)
            for key in ("checked", "risk_0", "risk_1", "risk_2", "updated")
        }
        completion_message = (
            f"检测完成：{public_summary['checked']} 条，"
            f"疑似 {public_summary['risk_1']} 条，"
            f"侵权 {public_summary['risk_2']} 条"
        )
        _append_risk_check_log(completion_message)
        with _risk_check_state_lock:
            _risk_check_state.update(
                {
                    "running": False,
                    "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "status": "success",
                    "message": completion_message,
                    "summary": public_summary,
                }
            )
    except Exception as exc:
        logging.error("侵权检测任务失败：%s", exc)
        traceback.print_exc()
        _append_risk_check_log(f"任务失败：{exc}")
        with _risk_check_state_lock:
            _risk_check_state.update(
                {
                    "running": False,
                    "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "status": "error",
                    "message": str(exc),
                    "summary": {},
                }
            )
    finally:
        task_lock.release()


def build_infringement_knowledge_analysis_params(data):
    data = data if isinstance(data, dict) else {}
    return {
        "infraction_limit": _parse_int_param(
            data, "infraction_limit", 10000, min_value=1, max_value=20000
        ),
        "active_limit": _parse_int_param(
            data, "active_limit", 5000, min_value=1, max_value=10000
        ),
        "batch_size": _parse_int_param(
            data, "batch_size", 300, min_value=20, max_value=300
        ),
    }


def _append_infringement_knowledge_analysis_log(message):
    text = str(message or "").strip()
    if not text:
        return
    timestamp = datetime.now().strftime("%H:%M:%S")
    with _infringement_knowledge_analysis_state_lock:
        for line in text.splitlines():
            line = line.strip()
            if not line:
                continue
            _infringement_knowledge_analysis_logs.append(f"[{timestamp}] {line}")
            if _infringement_knowledge_analysis_state.get("running"):
                _infringement_knowledge_analysis_state["message"] = line


def run_infringement_knowledge_analysis_job(params, task_lock):
    try:
        sources = db_get_infringement_knowledge_analysis_sources(
            infraction_limit=params["infraction_limit"],
            active_limit=params["active_limit"],
        )
        _append_infringement_knowledge_analysis_log(
            f"读取侵权商品 {len(sources.get('infraction_rows') or [])}/"
            f"{int(sources.get('infraction_total') or 0)} 个；"
            f"活跃成交链接 {len(sources.get('active_rows') or [])}/"
            f"{int(sources.get('active_total') or 0)} 条"
        )
        summary = bit_infringement_knowledge_analysis.analyze_knowledge_sources(
            sources,
            writer=db_upsert_analyzed_infringement_knowledge,
            batch_size=params["batch_size"],
            log_callback=_append_infringement_knowledge_analysis_log,
        )
        write_result = dict(summary.get("write_result") or {})
        completion_message = (
            f"自动分析完成：候选黑名单 {int(summary.get('blacklist_candidates') or 0)} 个，"
            f"候选白名单 {int(summary.get('whitelist_candidates') or 0)} 个；"
            f"新增 {int(write_result.get('inserted') or 0)} 个，"
            f"更新 {int(write_result.get('updated') or 0)} 个，"
            f"保护人工记录 {int(write_result.get('skipped_manual') or 0)} 个"
        )
        _append_infringement_knowledge_analysis_log(completion_message)
        with _infringement_knowledge_analysis_state_lock:
            _infringement_knowledge_analysis_state.update(
                {
                    "running": False,
                    "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "status": "success",
                    "message": completion_message,
                    "summary": summary,
                }
            )
    except Exception as exc:
        logging.error("侵权知识库自动分析失败：%s", exc)
        traceback.print_exc()
        _append_infringement_knowledge_analysis_log(f"自动分析失败：{exc}")
        with _infringement_knowledge_analysis_state_lock:
            _infringement_knowledge_analysis_state.update(
                {
                    "running": False,
                    "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "status": "error",
                    "message": str(exc),
                    "summary": {},
                }
            )
    finally:
        task_lock.release()


def _risk_result_query_params(args, export=False):
    risk_level = str(args.get("risk_level") or "").strip().lower()
    if risk_level not in {"", "0", "1", "2", "unchecked"}:
        raise ValueError("风险级别只能是 0、1、2 或未检测")
    sort_by = str(args.get("sort_by") or "risk_level").strip()
    if sort_by not in {
        "row_id",
        "product_id",
        "title",
        "zying_category",
        "risk_level",
        "keywords",
        "submitted_at",
    }:
        sort_by = "risk_level"
    sort_dir = "asc" if str(args.get("sort_dir") or "").lower() == "asc" else "desc"
    return {
        "zying_category": str(args.get("category") or "").strip()[:1024] or None,
        "risk_level": risk_level or None,
        "search": str(args.get("search") or "").strip()[:200],
        "sort_by": sort_by,
        "sort_dir": sort_dir,
        "limit": 0 if export else _parse_int_param(args, "limit", 1000, 1, 5000),
    }


def _append_order_print_log(message):
    text = format_log_text(message).rstrip()
    if not text:
        return
    with _order_print_lock:
        _order_print_logs.append(text + "\n")


def _order_print_snapshot():
    with _order_print_lock:
        runtime_log = "".join(_order_print_logs).rstrip()
        automatic_history = [
            row
            for row in (_order_print_state.get("site_last_runs") or [])
            if str(row.get("source") or "") == "系统自动打印"
        ]
        history_log = "\n".join(
            f"{row.get('finished_at') or '-'} "
            f"{row.get('shop_name') or '-'} / {row.get('site') or '-'}："
            f"{row.get('outcome') or '系统自动打印'}"
            for row in automatic_history
        )
        combined_log = runtime_log
        if history_log:
            automatic_section = "===== 系统自动打印最近记录 =====\n" + history_log
            combined_log = "\n\n".join(
                value for value in (runtime_log, automatic_section) if value
            )
        snapshot = {
            **dict(_order_print_state),
            "params": dict(_order_print_state.get("params") or {}),
            "results": [dict(row) for row in (_order_print_state.get("results") or [])],
            "site_last_runs": [
                dict(row) for row in (_order_print_state.get("site_last_runs") or [])
            ],
            "log": combined_log,
            "can_stop": bool(
                _order_print_state.get("running")
                and _order_print_stop_event is not None
            ),
        }
        snapshot.pop("download_path", None)
        snapshot["download_url"] = (
            "/api/order-print/download"
            if _order_print_state.get("download_path")
            else ""
        )
        return snapshot


def _order_print_history_status(outcome):
    text = str(outcome or "").strip()
    if "无待打印订单" in text:
        return "no_orders"
    if text.startswith("成功"):
        return "printed"
    if text.startswith("部分成功"):
        return "partial"
    if text.startswith("跳过"):
        return "skipped"
    if text.startswith("失败"):
        return "failed"
    return "unknown"


def _load_order_print_site_last_runs(current_results=None):
    """汇总全部 API 授权店铺站点及其最近一次订单打印时间。"""

    current_results = [dict(row) for row in (current_results or [])]
    configs_loaded = True
    try:
        configs = _order_print_config_options()["shops"]
    except Exception as exc:
        logging.warning("读取订单打印站点配置失败：%s", exc)
        configs = []
        configs_loaded = False
    try:
        history = db_get_latest_order_print_records() or []
    except Exception as exc:
        logging.warning("读取订单打印历史失败：%s", exc)
        history = []

    latest_by_key = {}
    for record in history:
        shop_name = str(record.get("shop_name") or "").strip()
        site = str(record.get("site") or "").strip()
        if not shop_name or not site:
            continue
        outcome = str(record.get("outcome") or "")
        latest_by_key[(shop_name, site)] = {
            "shop_name": shop_name,
            "site": site,
            "status": _order_print_history_status(outcome),
            "finished_at": str(record.get("finished_at") or ""),
            "outcome": outcome,
            "source": "系统自动打印" if "系统自动打印" in outcome else "",
        }
    for result in current_results:
        shop_name = str(result.get("shop_name") or "").strip()
        site = str(result.get("site") or "").strip()
        if not shop_name or not site:
            continue
        key = (shop_name, site)
        finished_at = str(result.get("finished_at") or "")
        existing = latest_by_key.get(key)
        if existing and str(existing.get("finished_at") or "") > finished_at:
            continue
        latest_by_key[key] = {
            "shop_name": shop_name,
            "site": site,
            "status": str(result.get("status") or "unknown"),
            "finished_at": finished_at,
            "outcome": str(result.get("message") or ""),
            "source": str(result.get("source") or ""),
        }

    rows = []
    seen = set()
    for config in configs:
        shop_name = str(config.get("shop_name") or "").strip()
        if not shop_name:
            continue
        for site in config.get("sites") or []:
            key = (shop_name, site)
            if key in seen:
                continue
            seen.add(key)
            rows.append(
                latest_by_key.get(
                    key,
                    {
                        "shop_name": shop_name,
                        "site": site,
                        "status": "not_run",
                        "finished_at": "",
                        "outcome": "",
                        "source": "",
                    },
                )
            )
    if not configs_loaded:
        for key, record in latest_by_key.items():
            if key not in seen:
                rows.append(record)
    return rows


def _order_print_config_options():
    """Return selectable stores from Mercado API authorizations only."""

    token_rows = list((bit_db_api.list_mercado_store_tokens() or {}).get("rows") or [])
    shops = []
    site_order = []
    for token in token_rows:
        token_id = int(token.get("id") or 0)
        shop_name = str(
            token.get("display_name") or token.get("nickname") or token_id or ""
        ).strip()
        if not token_id or not shop_name:
            continue
        sites = []
        salespeople = []
        for setting in token.get("site_settings") or []:
            site_id = str(setting.get("site_id") or "").strip().upper()
            site_name = str(
                setting.get("site_name")
                or bit_print.SITE_NAMES.get(site_id)
                or ""
            ).strip()
            if site_name and site_name not in sites:
                sites.append(site_name)
            salesperson = str(setting.get("salesperson") or "").strip()
            if salesperson and salesperson not in salespeople:
                salespeople.append(salesperson)
        default_site = bit_print.SITE_NAMES.get(
            str(token.get("site_id") or "").strip().upper()
        )
        if default_site and default_site not in sites:
            sites.append(default_site)
        if not sites:
            sites = list(bit_print.SITE_IDS)
        for site in sites:
            if site not in site_order:
                site_order.append(site)
        shops.append(
            {
                "token_id": token_id,
                "shop_name": shop_name,
                "salesperson": "、".join(salespeople),
                "sites": sites,
                "token_status": str(token.get("status") or "unknown"),
                "token_status_text": str(token.get("status_text") or ""),
            }
        )
    return {"shops": shops, "sites": site_order}


def _refresh_order_print_site_last_runs(current_results=None):
    if current_results is None:
        with _order_print_lock:
            current_results = [
                dict(row) for row in (_order_print_state.get("results") or [])
            ]
    rows = _load_order_print_site_last_runs(current_results)
    with _order_print_lock:
        _order_print_state["site_last_runs"] = rows
    return rows


def build_order_print_params(data):
    data = data if isinstance(data, dict) else {}
    local_tz = datetime.now().astimezone().tzinfo
    local_now = datetime.now(local_tz)

    def parse_range_value(name, label, default):
        raw = str(data.get(name) or "").strip()
        if not raw:
            return default
        try:
            parsed = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        except ValueError as exc:
            raise ValueError(f"{label}格式无效，请重新选择") from exc
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=local_tz)
        return parsed.astimezone(local_tz)

    range_end = parse_range_value("date_to", "结束时间", local_now)
    range_start = parse_range_value(
        "date_from",
        "开始时间",
        range_end - timedelta(hours=bit_print.DEFAULT_FALLBACK_HOURS),
    )
    if range_end > local_now + timedelta(minutes=5):
        range_end = local_now
    if range_start >= range_end:
        raise ValueError("订单打印开始时间必须早于结束时间")
    if range_end - range_start > timedelta(days=31):
        raise ValueError("单次订单打印时间段不能超过 31 天")

    raw_targets = data.get("targets")
    selected_targets = []
    if raw_targets is not None:
        if not isinstance(raw_targets, list):
            raise ValueError("targets 必须是数组")
        if not raw_targets:
            raise ValueError("请至少选择一个店铺站点")
        if len(raw_targets) > 1000:
            raise ValueError("单次最多选择 1000 个店铺站点")

        options = _order_print_config_options()
        configured_pairs = {
            (shop["shop_name"], site)
            for shop in options["shops"]
            for site in shop["sites"]
        }
        seen_targets = set()
        for raw_target in raw_targets:
            if not isinstance(raw_target, dict):
                raise ValueError("每个店铺站点必须包含 shop_name 和 site")
            shop_name = str(raw_target.get("shop_name") or "").strip()
            site = str(raw_target.get("site") or "").strip()
            if not shop_name or not site:
                raise ValueError("每个店铺站点必须包含 shop_name 和 site")
            pair = (shop_name, site)
            if pair not in configured_pairs:
                raise ValueError(f"店铺站点不存在或已被忽略：{shop_name} / {site}")
            if pair in seen_targets:
                continue
            seen_targets.add(pair)
            selected_targets.append({"shop_name": shop_name, "site": site})

        selected_shops = tuple(
            dict.fromkeys(target["shop_name"] for target in selected_targets)
        )
        selected_sites = tuple(
            dict.fromkeys(target["site"] for target in selected_targets)
        )
        target_label = (
            f"{len(selected_shops)} 家店铺 / "
            f"{len(selected_targets)} 个店铺站点"
        )
    else:
        selected_shops = _normalized_collection_list(data, "shops")
        selected_sites = _normalized_collection_list(data, "sites")
        options = _order_print_config_options()
        configured = {shop["shop_name"]: shop for shop in options["shops"]}
        unknown_shops = [shop for shop in selected_shops if shop not in configured]
        if unknown_shops:
            raise ValueError("API 授权店铺不存在：" + "、".join(unknown_shops))
        unknown_sites = [site for site in selected_sites if site not in options["sites"]]
        if unknown_sites:
            raise ValueError("站点不存在：" + "、".join(unknown_sites))
        matching_shops = [
            shop
            for shop in selected_shops
            if any(site in selected_sites for site in configured[shop]["sites"])
        ]
        if not matching_shops:
            raise ValueError("所选 API 授权店铺没有匹配站点")
        target_label = f"{len(matching_shops)} 家 API 授权店铺 / {len(selected_sites)} 个站点"
    return {
        "task_id": secrets.token_hex(16),
        "mode": "once",
        "max_retries": _parse_int_param(
            data, "max_retries", 3, min_value=1, max_value=3
        ),
        "retry_delay_seconds": _parse_int_param(
            data, "retry_delay_seconds", 3, min_value=0, max_value=60
        ),
        "fallback_hours": bit_print.DEFAULT_FALLBACK_HOURS,
        "start_at": range_start.astimezone(timezone.utc).isoformat(),
        "end_at": range_end.astimezone(timezone.utc).isoformat(),
        "date_from": range_start.strftime("%Y-%m-%dT%H:%M"),
        "date_to": range_end.strftime("%Y-%m-%dT%H:%M"),
        "range_label": (
            f"{range_start.strftime('%Y-%m-%d %H:%M')} 至 "
            f"{range_end.strftime('%Y-%m-%d %H:%M')}"
        ),
        "selected_shops": selected_shops,
        "selected_sites": selected_sites,
        "selected_targets": selected_targets,
        "target": target_label,
    }


def run_order_print_job(params, task_lock, stop_event):
    global _order_print_stop_event
    try:
        with _order_print_lock:
            _order_print_state["message"] = "正在通过美客多 API 生成面单"
        _append_order_print_log(f"{get_now_time()} ===== 美客多 API 订单打印开始 =====")
        summary = bit_print.print_orders_all(
            selected_shops=params["selected_shops"],
            selected_sites=params["selected_sites"],
            selected_targets=params.get("selected_targets"),
            max_retries=params["max_retries"],
            retry_delay_seconds=params["retry_delay_seconds"],
            fallback_hours=params.get(
                "fallback_hours", bit_print.DEFAULT_FALLBACK_HOURS
            ),
            start_at=params.get("start_at"),
            end_at=params.get("end_at"),
            stop_event=stop_event,
            logger=_append_order_print_log,
            task_id=params.get("task_id"),
        )
        site_last_runs = _load_order_print_site_last_runs(summary.get("results", []))
        with _order_print_lock:
            _order_print_state.update(
                {
                    "printed": summary.get("printed", 0),
                    "partial": summary.get("partial", 0),
                    "no_orders": summary.get("no_orders", 0),
                    "failed": summary.get("failed", 0),
                    "skipped": summary.get("skipped", 0),
                    "printed_order_count": summary.get("printed_order_count", 0),
                    "shipment_count": summary.get("shipment_count", 0),
                    "fallback_store_count": summary.get("fallback_store_count", 0),
                    "download_path": summary.get("download_path", ""),
                    "download_name": summary.get("download_name", ""),
                    "results": summary.get("results", []),
                    "site_last_runs": site_last_runs,
                }
            )

        stopped = stop_event.is_set()
        failed_sites = int(summary.get("failed") or 0)
        partial_sites = int(summary.get("partial") or 0)
        has_output = bool(summary.get("download_path"))
        if stopped:
            final_status = "stopped"
            final_message = "API 订单打印已停止"
        elif failed_sites and not has_output:
            final_status = "error"
            final_message = f"没有生成面单，{failed_sites} 个站点执行失败"
        elif failed_sites or partial_sites:
            final_status = "partial"
            final_message = (
                f"已生成 {summary.get('shipment_count', 0)} 个面单；"
                f"{partial_sites} 个站点部分成功，{failed_sites} 个站点失败"
            )
        else:
            final_status = "success"
            final_message = (
                f"API 面单已生成：{summary.get('shipment_count', 0)} 个"
                if has_output
                else "没有需要打印的订单"
            )
        with _order_print_lock:
            _order_print_state.update(
                {
                    "running": False,
                    "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "status": final_status,
                    "message": final_message,
                }
            )
    except Exception as exc:
        logging.error("Order print task failed: %s", exc)
        traceback.print_exc()
        _append_order_print_log(f"{get_now_time()} 订单打印异常：{exc}")
        with _order_print_lock:
            _order_print_state.update(
                {
                    "running": False,
                    "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "status": "error",
                    "message": str(exc),
                }
            )
    finally:
        if task_lock is not None:
            task_lock.release()
        with _order_print_lock:
            if _order_print_stop_event is stop_event:
                _order_print_stop_event = None


def _parse_rate_param(data, name="min_rate", default=0):
    value = data.get(name, default)
    text = str(value if value is not None else default).strip().replace("％", "%")
    if not text:
        return float(default)
    is_percent = "%" in text
    number_text = text.replace("%", "").replace("，", ".").replace(",", ".").strip()
    try:
        number = float(number_text)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{name} 必须是 0% 到 100% 之间的比率") from exc
    rate = number / 100 if is_percent or number > 1 else number
    if rate < 0 or rate > 1:
        raise ValueError(f"{name} 必须是 0% 到 100% 之间的比率")
    return rate


def _request_execution_target(data=None):
    if (
        has_request_context()
        and getattr(g, "local_executor_user", None)
    ) or RUNTIME_SETTINGS.is_client:
        return "local"
    if (data or {}).get("execution_target") == "agent":
        return "agent"
    return "server"


def build_daily_task_params(data):
    mode = str(data.get("mode", "loop")).strip().lower()
    if mode not in ("once", "loop"):
        mode = "loop"
    raw_appeal_types = (
        data.get("appeal_types")
        if "appeal_types" in data
        else data.get("appeal_type") or bit_daily_task.APPEAL_TYPE_INFRACTION
    )
    normalized_appeal_types = bit_daily_task.normalize_appeal_types(raw_appeal_types)
    appeal_types = [
        "延误率" if value == bit_daily_task.APPEAL_TYPE_DELAY else value
        for value in normalized_appeal_types
    ]
    raw_salespeople = data.get("salespeople", data.get("salesperson", []))
    if isinstance(raw_salespeople, str):
        raw_salespeople = [raw_salespeople]
    salespeople = []
    for value in raw_salespeople or ():
        salesperson = str(value or "").strip()
        if salesperson in ("", "全部业务员", "所有业务员", "all", "*"):
            continue
        if salesperson not in salespeople:
            salespeople.append(salesperson)
    raw_group_names = data.get("group_names", data.get("group_name", []))
    if isinstance(raw_group_names, str):
        raw_group_names = [raw_group_names]
    group_names = []
    for value in raw_group_names or ():
        group_name = str(value or "").strip()
        if group_name in ("", "全部店铺组", "所有店铺组", "all", "*"):
            continue
        if group_name not in group_names:
            group_names.append(group_name)
    legacy_min_rate = _parse_rate_param(data)
    return {
        "execution_target": _request_execution_target(data),
        "mode": mode,
        "appeal_types": appeal_types,
        "appeal_type": appeal_types[0] if len(appeal_types) == 1 else "多任务",
        "top_n": 0,
        "max_workers": _parse_int_param(
            data,
            "max_workers",
            bit_daily_task.DEFAULT_DAILY_MAX_WORKERS,
            1,
            bit_daily_task.MAX_DAILY_TASK_WORKERS,
        ),
        "recent_days": _parse_int_param(
            data,
            "recent_days",
            bit_daily_task.DEFAULT_DAILY_RECENT_DAYS,
            1,
            365,
        ),
        "round_interval": _parse_int_param(data, "round_interval", 600, 10, 86400),
        "site_pause": _parse_int_param(data, "site_pause", 30, 0, 3600),
        "stop_after_minutes": _parse_int_param(data, "stop_after_minutes", 360, 0, 24 * 60),
        "salespeople": salespeople,
        "group_names": group_names,
        "min_rate": legacy_min_rate,
        "infraction_min_count": _parse_int_param(
            data, "infraction_min_count", 0, 0, 1000000
        ),
        "delay_min_rate": _parse_rate_param(
            data, "delay_min_rate", legacy_min_rate
        ),
        "cancellation_min_rate": _parse_rate_param(
            data, "cancellation_min_rate", legacy_min_rate
        ),
        "complaint_min_rate": _parse_rate_param(
            data, "complaint_min_rate", legacy_min_rate
        ),
        "message": str(data.get("message", "") or ""),
    }


def execute_daily_task(params, task_lock, stop_event, task_id, effective_log_path, owned_window_ids):
    """Run daily_task business logic in either a workbench or an Agent worker."""
    appeal_types = params.get("appeal_types") or [
        params.get("appeal_type", bit_daily_task.APPEAL_TYPE_INFRACTION)
    ]
    appeal_task = appeal_types[0] if len(appeal_types) == 1 else appeal_types
    appeal_label = "、".join(appeal_types)
    min_rate = params.get("min_rate", 0)
    execution_standards = {
        "min_infraction_count": params.get("infraction_min_count", 0),
        "min_delay_rate": params.get("delay_min_rate", min_rate),
        "min_cancellation_rate": params.get("cancellation_min_rate", min_rate),
        "min_complaint_rate": params.get("complaint_min_rate", min_rate),
    }
    if params["mode"] == "loop":
        stop_at = None
        if params["stop_after_minutes"] > 0:
            stop_at = datetime.now() + timedelta(minutes=params["stop_after_minutes"])
        execution_result = bit_daily_task.loop_ai_appeal(
            appeal_task,
            top_n=params["top_n"],
            max_workers=params["max_workers"],
            recent_days=params["recent_days"],
            round_interval=params["round_interval"],
            site_pause=params["site_pause"],
            message=params["message"],
            min_rate=min_rate,
            **execution_standards,
            salespeople=params["salespeople"],
            group_names=params.get("group_names", []),
            stop_at=stop_at,
            stop_event=stop_event,
            log_path=str(effective_log_path),
            _task_lock=task_lock,
            task_id=task_id,
            owned_window_ids=owned_window_ids,
        )
        result_message = (
            f"daily_task {appeal_label}任务已停止"
            if stop_event.is_set()
            else f"daily_task {appeal_label}任务循环执行完成"
        )
    else:
        execution_result = bit_daily_task.run_ai_appeal_once(
            appeal_task,
            top_n=params["top_n"],
            max_workers=params["max_workers"],
            recent_days=params["recent_days"],
            site_pause=params["site_pause"],
            message=params["message"],
            min_rate=min_rate,
            **execution_standards,
            salespeople=params["salespeople"],
            group_names=params.get("group_names", []),
            stop_event=stop_event,
            log_path=str(effective_log_path),
            _task_lock=task_lock,
            task_id=task_id,
            owned_window_ids=owned_window_ids,
        )
        result_message = (
            f"daily_task {appeal_label}任务已停止"
            if stop_event.is_set()
            else f"daily_task {appeal_label}任务单轮执行完成"
        )

    execution_counts = bit_daily_task.task_execution_counts(execution_result)
    needs_attention = any(count and key not in {"sent", "replied", "no_data"}
                          for key, count in execution_counts.items())
    result_message += f"；执行统计：{execution_counts}；话术发送成功不代表申诉已批准"
    print(f"{get_now_time()} {result_message}<br>")
    return {
        "status": "stopped" if stop_event.is_set() else ("partial" if needs_attention else "success"),
        "execution_counts": execution_counts,
        "message": result_message,
    }


def run_daily_task_job(
    params,
    task_lock,
    stop_event=None,
    task_id="",
    log_path=None,
    owned_window_ids=None,
):
    global _daily_task_stop_event, _daily_task_stop_manager
    task_id = str(task_id or "").strip()
    effective_log_path = Path(log_path or _daily_task_log_file(task_id))
    stop_event = stop_event or threading.Event()
    register_thread_log_queue(DailyTaskLogSink(effective_log_path))
    try:
        print(f"{get_now_time()} 开始执行 daily_task：{params}<br>")
        result = execute_daily_task(
            params, task_lock, stop_event, task_id, effective_log_path, owned_window_ids
        )
        _update_daily_task_state(
            task_id,
            running=False,
            finished_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            status=result["status"],
            execution_counts=result["execution_counts"],
            message=result["message"],
            stop_requested=stop_event.is_set(),
            can_stop=False,
        )
    except Exception as e:
        logging.error("daily_task failed: %s", e)
        _append_daily_task_log(
            f"\ndaily_task 运行失败：{e}\n",
            effective_log_path,
        )
        traceback.print_exc()
        stopped = stop_event.is_set()
        _update_daily_task_state(
            task_id,
            running=False,
            finished_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            status="stopped" if stopped else "error",
            message="daily_task 已停止" if stopped else str(e),
            stop_requested=stopped,
            can_stop=False,
        )
    finally:
        if task_lock is not None:
            task_lock.release()
        stop_manager = None
        if task_id:
            with _daily_task_lock:
                control = _daily_task_controls.get(task_id) or {}
                if control.get("stop_event") is stop_event:
                    _daily_task_controls.pop(task_id, None)
                    stop_manager = control.get("stop_manager")
        else:
            with _daily_task_lock:
                if _daily_task_stop_event is stop_event:
                    _daily_task_stop_event = None
                    stop_manager = _daily_task_stop_manager
                    _daily_task_stop_manager = None
        if stop_manager is not None:
            try:
                stop_manager.shutdown()
            except Exception:
                pass
        unregister_thread_log_queue()
        _prune_daily_task_history()


def format_log_text(text):
    return str(text).replace("<br>", "\n").replace("<br/>", "\n").replace("<br />", "\n")


def shensu_logic_old(name, site, form, message):
    i = 0
    while i < 10:
        i = i + 1
        try:
            yield f"{get_now_time()}--- 任务启动第{i}次：{name}{site} ---<br>"
            shensu(name, site, form, message)
            # 模拟自动化操作步骤
            yield f"{get_now_time()}✅ {name}{site}申诉执行完毕,！<br>"
        except Exception as e:
            yield f"发生错误: {str(e)}<br>"
        finally:
            yield f"{get_now_time()}{name}{site}关闭浏览器等待十分钟，进行下一次申诉<br>"
            window_id = getWindowidByName(name)
            time.sleep(600)


# 2. 接口路由
def shensu_logic_previous(name, site, form, message):
    for i in range(1, 11):
        output_queue = queue.Queue()

        def run_task():
            writer = StreamWriter(output_queue)
            with contextlib.redirect_stdout(writer), contextlib.redirect_stderr(writer):
                try:
                    print(f"{get_now_time()} --- 任务启动第 {i} 次：{name} {site}")
                    shensu(name, site, form, message)
                    print(f"{get_now_time()} {name} {site} 申诉执行完毕")
                except Exception as e:
                    print(f"{get_now_time()} 发生错误: {str(e)}")
                    traceback.print_exc()
                finally:
                    output_queue.put(None)

        task_thread = threading.Thread(target=run_task, daemon=True)
        task_thread.start()

        while True:
            text = output_queue.get()
            if text is None:
                break
            yield format_log_text(text)
            sys.stdout.flush()

        yield f"{get_now_time()} {name} {site} 本轮结束，等待十分钟后进入下一轮\n"
        getWindowidByName(name)
        time.sleep(600)


def resolve_appeal_sites(sites):
    """规范化控制台选中的站点，兼容旧版单个 site 字符串调用。"""
    raw_sites = [sites] if isinstance(sites, str) else list(sites or [])
    selected_sites = []
    invalid_sites = []
    for raw_site in raw_sites:
        site = str(raw_site or "").strip()
        if not site:
            continue
        if site not in APPEAL_SITES:
            invalid_sites.append(site)
            continue
        if site not in selected_sites:
            selected_sites.append(site)
    if invalid_sites:
        raise ValueError(f"不支持的站点：{'、'.join(invalid_sites)}")
    if not selected_sites:
        raise ValueError("请至少选择一个站点")
    return tuple(selected_sites)


def validate_authorized_appeal_sites(shop_name, sites):
    """拒绝未在授权店铺信息中显式勾选“进行申诉”的店铺站点。"""
    name = str(shop_name or "").strip()
    scope = bit_daily_task.load_authorized_appeal_shop_site_config()
    allowed_sites = scope.get(name.casefold(), set())
    if not allowed_sites:
        raise ValueError(f"{name or '该店铺'}未在授权店铺信息中开启进行申诉")
    unauthorized = [
        site
        for site in sites
        if bit_appeal_ai.normalize_site_code(site) not in allowed_sites
    ]
    if unauthorized:
        raise ValueError(
            f"{name}以下站点未开启进行申诉：{'、'.join(unauthorized)}"
        )
    return tuple(sites)


def resolve_appeal_forms(forms):
    """规范化申诉类型，并按固定业务顺序执行，兼容旧版单个 form 字符串。"""
    raw_forms = [forms] if isinstance(forms, str) else list(forms or [])
    selected_forms = set()
    invalid_forms = []
    for raw_form in raw_forms:
        appeal_form = str(raw_form or "").strip()
        if not appeal_form:
            continue
        if appeal_form not in APPEAL_FORMS:
            invalid_forms.append(appeal_form)
            continue
        selected_forms.add(appeal_form)
    if invalid_forms:
        raise ValueError(f"不支持的任务类型：{'、'.join(dict.fromkeys(invalid_forms))}")
    if not selected_forms:
        raise ValueError("请至少选择一个任务类型")
    return tuple(
        appeal_form for appeal_form in APPEAL_FORMS if appeal_form in selected_forms
    )


def normalize_appeal_loop_count(value):
    """返回申诉轮数；0 表示永久循环，其他值只允许 10、20、50。"""
    text = str(value if value is not None else DEFAULT_APPEAL_LOOP_COUNT).strip()
    if text.casefold() in ("0", "permanent", "forever", "永久"):
        return PERMANENT_APPEAL_LOOP_COUNT
    try:
        loop_count = int(text)
    except (TypeError, ValueError) as exc:
        raise ValueError("循环次数只支持 10、20、50 或永久") from exc
    if loop_count not in APPEAL_LOOP_COUNTS:
        raise ValueError("循环次数只支持 10、20、50 或永久")
    return loop_count


def get_appeal_round_interval(mode):
    """AI 客服每轮间隔 1 分钟，人工客服每轮间隔 10 分钟。"""
    if mode == "AI客服":
        return AI_APPEAL_ROUND_INTERVAL_SECONDS
    return MANUAL_APPEAL_ROUND_INTERVAL_SECONDS


def stream_task_output(output_queue, stop_event=None):
    """持续转发任务日志，并在 AI 长时间无输出时发送心跳，避免请求连接超时。"""
    heartbeat_seconds = max(1, APPEAL_STREAM_HEARTBEAT_SECONDS)
    stop_notice_sent = False
    while True:
        if stop_event is not None and stop_event.is_set() and not stop_notice_sent:
            yield (
                f"{get_now_time()} 已收到终结请求，等待当前站点安全结束并释放窗口\n"
            )
            stop_notice_sent = True
        try:
            text = output_queue.get(timeout=heartbeat_seconds)
        except queue.Empty:
            yield f"{get_now_time()} 申诉任务仍在运行（保持连接）\n"
            continue
        if text is None:
            return
        yield format_log_text(text)
        sys.stdout.flush()


def stream_appeal_round_wait(seconds, stop_event=None):
    """分段等待下一轮，并持续向前端发送状态，避免静默导致连接断开。"""
    remaining = max(0, int(seconds))
    heartbeat_seconds = max(1, APPEAL_STREAM_HEARTBEAT_SECONDS)
    while remaining > 0:
        wait_seconds = min(heartbeat_seconds, remaining)
        if stop_event is not None:
            if stop_event.wait(wait_seconds):
                yield f"{get_now_time()} 已终结本次申诉任务，不再进入下一轮\n"
                return False
        else:
            time.sleep(wait_seconds)
        remaining -= wait_seconds
        if remaining > 0:
            yield f"{get_now_time()} 等待下一轮，剩余 {remaining} 秒（保持连接）\n"
    return True


def shensu_logic(
    name,
    sites,
    form,
    message,
    mode,
    loop_count=DEFAULT_APPEAL_LOOP_COUNT,
    stop_event=None,
):
    target_sites = resolve_appeal_sites(sites)
    target_forms = resolve_appeal_forms(form)
    round_limit = normalize_appeal_loop_count(loop_count)
    multiple_tasks_selected = len(target_sites) > 1 or len(target_forms) > 1
    round_interval_seconds = get_appeal_round_interval(mode)
    round_interval_minutes = round_interval_seconds // 60
    round_number = 0
    cancellation_enabled = stop_event is not None
    stop_event = stop_event or threading.Event()

    while round_limit == PERMANENT_APPEAL_LOOP_COUNT or round_number < round_limit:
        if stop_event.is_set():
            yield f"{get_now_time()} 已终结本次申诉任务\n"
            return
        round_number += 1
        if multiple_tasks_selected:
            yield (
                f"{get_now_time()} 第 {round_number} 轮开始，任务类型将按 "
                f"{' → '.join(target_forms)} 的顺序执行；"
                f"选中站点：{'、'.join(target_sites)}\n"
            )

        for current_form in target_forms:
            if len(target_forms) > 1:
                yield (
                    f"{get_now_time()} 第 {round_number} 轮开始执行任务类型："
                    f"{current_form}\n"
                )
            for current_site in target_sites:
                if stop_event.is_set():
                    yield (
                        f"{get_now_time()} 已终结本次申诉任务，"
                        "不再执行后续任务类型和站点\n"
                    )
                    return
                output_queue = queue.Queue()
                task_result = {"value": None}

                def run_task(
                    run_site=current_site,
                    run_form=current_form,
                    run_round=round_number,
                ):
                    register_thread_log_queue(output_queue)
                    window_lease = None
                    try:
                        print(
                            f"{get_now_time()} --- 第 {run_round} 轮任务启动："
                            f"{name} {run_site}，任务类型：{run_form}，客服模式：{mode}"
                        )
                        window_id = getWindowidByName(name)
                        window_lease = create_window_lease(
                            window_id,
                            owner=f"interface_appeal:{name}",
                            shop_name=name,
                            task_type="interface_appeal",
                        )
                        if not window_lease.acquire(timeout=0):
                            task_result["value"] = "窗口正在被其他任务占用"
                            print(
                                f"{get_now_time()} {name} {run_site} {run_form} "
                                f"{task_result['value']}，本轮已跳过"
                            )
                            return
                        if mode == "AI客服":
                            with bit_appeal_ai.appeal_controls(stop_event):
                                task_result["value"] = bit_appeal_ai.shensu(
                                    name, run_site, run_form, message
                                )
                        else:
                            task_result["value"] = shensu(
                                name, run_site, run_form, message, "人工客服"
                            )
                        print(
                            f"{get_now_time()} {name} {run_site} {run_form} "
                            f"申诉执行完毕：{task_result['value']}"
                        )
                    except Exception as e:
                        print(
                            f"{get_now_time()} {name} {run_site} {run_form} "
                            f"发生错误: {str(e)}"
                        )
                        traceback.print_exc()
                    finally:
                        if window_lease is not None:
                            window_lease.release()
                        unregister_thread_log_queue()
                        output_queue.put(None)

                task_thread = threading.Thread(target=run_task, daemon=True)
                task_thread.start()

                yield from stream_task_output(output_queue, stop_event=stop_event)

                if stop_event.is_set():
                    yield (
                        f"{get_now_time()} {name} {current_site} {current_form} "
                        "当前操作已安全结束，本次任务已终结\n"
                    )
                    return

                if bit_daily_task._is_login_required_result(task_result.get("value")):
                    yield (
                        f"{get_now_time()} {name} {current_site} "
                        f"{task_result.get('value')}，已停止该店铺后续站点、"
                        "任务类型和申诉循环\n"
                    )
                    return

        has_next_round = (
            round_limit == PERMANENT_APPEAL_LOOP_COUNT
            or round_number < round_limit
        )
        if multiple_tasks_selected:
            yield (
                f"{get_now_time()} 第 {round_number} 轮任务类型 "
                f"{'、'.join(target_forms)} 和选中站点执行完成，"
                + (
                    f"等待 {round_interval_minutes} 分钟后开始下一轮\n"
                    if has_next_round
                    else "已达到规定循环次数\n"
                )
            )
        else:
            yield (
                f"{get_now_time()} {name} {target_sites[0]} 第 {round_number} 轮结束，"
                + (
                    f"等待 {round_interval_minutes} 分钟后进入下一轮\n"
                    if has_next_round
                    else "已达到规定循环次数\n"
                )
            )
        if (
            round_limit != PERMANENT_APPEAL_LOOP_COUNT
            and round_number >= round_limit
        ):
            yield (
                f"{get_now_time()} 已完成规定的 {round_limit} 轮申诉，任务结束\n"
            )
            return
        wait_completed = yield from stream_appeal_round_wait(
            round_interval_seconds,
            stop_event=stop_event if cancellation_enabled else None,
        )
        if not wait_completed:
            return


@app.route('/api/run_shensu', methods=['GET'])
@login_required
def api_run_shensu():
    # 获取前端传入的参数
    name = request.args.get("name", "")
    requested_execution_target = str(
        request.args.get("execution_target") or ""
    ).strip().lower()
    execution_target = (
        "local"
        if getattr(g, "local_executor_user", None) or RUNTIME_SETTINGS.is_client
        else "agent"
        if requested_execution_target == "agent"
        else "server"
    )
    try:
        sites = resolve_appeal_sites(request.args.getlist("site"))
        sites = validate_authorized_appeal_sites(name, sites)
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    try:
        forms = resolve_appeal_forms(request.args.getlist("form"))
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    try:
        loop_count = normalize_appeal_loop_count(
            request.args.get("loop_count", DEFAULT_APPEAL_LOOP_COUNT)
        )
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    message = request.args.get("message", "")
    mode = request.args.get("mode", "人工客服")
    task_id = normalize_appeal_task_id(request.args.get("task_id", ""))
    if not task_id:
        task_id = secrets.token_hex(16)
    if execution_target == "agent":
        return enqueue_local_agent_appeal(
            task_id=task_id,
            agent_id=request.args.get("agent_id", ""),
            name=name,
            sites=sites,
            forms=forms,
            message=message,
            mode=mode,
            loop_count=loop_count,
        )
    stop_event = register_appeal_task(
        task_id,
        {
            "name": name,
            "sites": list(sites),
            "loop_count": "永久" if loop_count == 0 else loop_count,
            "form": forms[0] if len(forms) == 1 else "、".join(forms),
            "forms": list(forms),
            "mode": mode,
            "execution_target": execution_target,
        },
    )
    if stop_event is None:
        return jsonify({"status": "error", "message": "该任务编号正在运行"}), 409

    def generate():
        try:
            yield f"{get_now_time()} 申诉任务编号：{task_id}\n"
            yield (
                f"{get_now_time()} 执行端："
                + (
                    "本机比特浏览器\n"
                    if execution_target == "local"
                    else "服务器比特浏览器\n"
                )
            )
            yield from shensu_logic(
                name,
                sites,
                forms,
                message,
                mode,
                loop_count=loop_count,
                stop_event=stop_event,
            )
        finally:
            stop_event.set()
            finish_appeal_task(task_id)

    # 返回流式响应，mimetype 设为 text/html 或 text/event-stream
    response = Response(generate(), mimetype='text/plain; charset=utf-8')
    response.headers["Cache-Control"] = "no-cache"
    response.headers["X-Accel-Buffering"] = "no"
    response.headers["X-Appeal-Task-ID"] = task_id
    response.headers["X-Execution-Target"] = execution_target
    return response


@app.route('/api/run_shensu/stop', methods=['POST'])
@login_required
def api_stop_shensu():
    data = request.get_json(silent=True) or {}
    task_id = normalize_appeal_task_id(data.get("task_id"))
    if not task_id:
        return jsonify({"status": "error", "message": "缺少有效任务编号"}), 400
    try:
        agent_job = get_local_agent_store().get_job(task_id)
    except ValueError:
        agent_job = None
    if agent_job and agent_job.get("job_type") == "appeal":
        if get_local_agent_store().request_cancel(task_id):
            return jsonify({
                "status": "success",
                "message": "本机 Agent 申诉停止请求已提交",
            })
        return jsonify({"status": "error", "message": "任务已结束"}), 409
    if not request_appeal_task_stop(task_id):
        return jsonify({"status": "error", "message": "任务已结束或不存在"}), 404
    return jsonify(
        {
            "status": "success",
            "message": "已提交终结请求，当前站点结束后将释放窗口并停止后续任务",
        }
    )


@app.route('/api/appeal-phrases', methods=['GET', 'POST'])
@login_required
def api_appeal_phrases():
    try:
        if request.method == "POST":
            result = db_create_appeal_phrase(request.get_json(silent=True) or {})
        else:
            result = db_list_appeal_phrases()
        return jsonify({"status": "success", "data": result})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400


@app.route('/api/appeal-phrases/<int:phrase_id>', methods=['PUT', 'DELETE'])
@login_required
def api_appeal_phrase_detail(phrase_id):
    try:
        if request.method == "PUT":
            result = db_update_appeal_phrase(
                phrase_id,
                request.get_json(silent=True) or {},
            )
        else:
            result = db_delete_appeal_phrase(phrase_id)
        return jsonify({"status": "success", "data": result})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400


@app.route('/api/infringement-knowledge', methods=['GET', 'POST'])
@login_required
def api_infringement_knowledge():
    try:
        if request.method == "POST":
            result = db_create_infringement_knowledge(
                request.get_json(silent=True) or {}
            )
        else:
            result = db_list_infringement_knowledge(
                list_type=request.args.get("list_type", ""),
                search=request.args.get("search", ""),
                limit=request.args.get("limit", 2000),
            )
        return jsonify({"status": "success", "data": result})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400


@app.route('/api/infringement-knowledge/bulk', methods=['POST'])
@login_required
def api_bulk_infringement_knowledge():
    from bit.bit_infringement_knowledge import parse_bulk_brand_lines

    try:
        data = request.get_json(silent=True) or {}
        records = parse_bulk_brand_lines(
            data.get("brands_text"),
            data.get("list_type"),
            data.get("notes", ""),
        )
        result = db_bulk_create_infringement_knowledge(records)
        return jsonify({"status": "success", "data": result})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400


@app.route('/api/infringement-knowledge/analysis/start', methods=['POST'])
@login_required
def api_start_infringement_knowledge_analysis():
    params = build_infringement_knowledge_analysis_params(
        request.get_json(silent=True) or {}
    )
    if not _infringement_knowledge_analysis_lock.acquire(blocking=False):
        with _infringement_knowledge_analysis_state_lock:
            data = dict(_infringement_knowledge_analysis_state)
        return jsonify(
            {"status": "error", "message": "侵权知识库自动分析正在运行", "data": data}
        ), 409

    with _infringement_knowledge_analysis_state_lock:
        _infringement_knowledge_analysis_logs.clear()
        _infringement_knowledge_analysis_state.update(
            {
                "running": True,
                "started_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "finished_at": "",
                "status": "running",
                "message": "正在读取侵权记录和当前活跃成交链接",
                "params": dict(params),
                "summary": {},
            }
        )
        _append_infringement_knowledge_analysis_log(
            "自动分析已启动：侵权记录生成黑名单；active 且已有销量的链接生成白名单候选；黑名单优先"
        )
        data = {
            **dict(_infringement_knowledge_analysis_state),
            "logs": list(_infringement_knowledge_analysis_logs),
        }
    try:
        threading.Thread(
            target=run_infringement_knowledge_analysis_job,
            args=(params, _infringement_knowledge_analysis_lock),
            daemon=True,
            name="infringement-knowledge-analysis",
        ).start()
    except Exception:
        _infringement_knowledge_analysis_lock.release()
        raise
    return jsonify({"status": "success", "data": data})


@app.route('/api/infringement-knowledge/analysis/status', methods=['GET'])
@login_required
def api_infringement_knowledge_analysis_status():
    with _infringement_knowledge_analysis_state_lock:
        data = {
            **dict(_infringement_knowledge_analysis_state),
            "params": dict(_infringement_knowledge_analysis_state.get("params") or {}),
            "summary": dict(_infringement_knowledge_analysis_state.get("summary") or {}),
            "logs": list(_infringement_knowledge_analysis_logs),
        }
    return jsonify({"status": "success", "data": data})


@app.route('/api/infringement-knowledge/<int:record_id>', methods=['PUT', 'DELETE'])
@login_required
def api_infringement_knowledge_detail(record_id):
    try:
        if request.method == "PUT":
            result = db_update_infringement_knowledge(
                record_id,
                request.get_json(silent=True) or {},
            )
        else:
            result = db_delete_infringement_knowledge(record_id)
        return jsonify({"status": "success", "data": result})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400


@app.route('/api/infractions/latest', methods=['GET'])
@login_required
def api_latest_infractions():
    try:
        recent_days = request.args.get("days", 30)
        return jsonify({
            "status": "success",
            "data": db_get_latest_infraction_info(recent_days)
        })
    except Exception as e:
        logging.error(f"Latest infraction query failed: {str(e)}")
        return jsonify({
            "status": "error",
            "message": f"Database error: {str(e)}"
        }), 500


@app.route('/api/infractions/latest/export', methods=['GET'])
@login_required
def api_export_latest_infractions():
    try:
        recent_days = request.args.get("days", 30)
        data = db_get_latest_infraction_info(recent_days)
        rows = data.get("rows") or []
        recent_days = data.get("recent_days") or 30

        wb = Workbook()
        detail_ws = wb.active
        detail_ws.title = "侵权明细"

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        header_font = Font(bold=True, color="1F2937")

        detail_columns = ["店铺名", "站点", "类型", "编号", "标题", "侵权时间", "执行时间", "提交时间"]
        detail_ws.append(detail_columns)
        for row in rows:
            detail_ws.append([row.get(column, "") for column in detail_columns])

        for cell in detail_ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for column_cells in detail_ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            detail_ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 42)
        detail_ws.freeze_panes = "A2"

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        submit_time = str(data.get("latest_submit_time") or datetime.now().strftime("%Y%m%d%H%M%S"))
        safe_time = "".join(ch if ch.isdigit() else "" for ch in submit_time) or datetime.now().strftime("%Y%m%d%H%M%S")
        filename = f"最新侵权明细_最近{recent_days}天_{safe_time}.xlsx"
        encoded_filename = quote(filename)
        response = send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )
        response.headers["Content-Disposition"] = f"attachment; filename*=UTF-8''{encoded_filename}"
        return response
    except Exception as e:
        logging.error(f"Latest infraction export failed: {str(e)}")
        return jsonify({
            "status": "error",
            "message": f"Export error: {str(e)}"
        }), 500


@app.route('/api/infractions/collect', methods=['POST'])
@login_required
def api_collect_infractions():
    try:
        params = _parse_collection_request(request.get_json(silent=True) or {})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.error("读取侵权采集范围失败：%s", exc)
        return jsonify({"status": "error", "message": f"读取采集范围失败：{exc}"}), 500

    with _infraction_collect_lock:
        if _infraction_collect_state.get("running"):
            return jsonify({
                "status": "running",
                "data": dict(_infraction_collect_state),
                "message": "侵权数据采集正在运行中"
            }), 409

        started_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        _infraction_collect_state.update({
            "running": True,
            "started_at": started_at,
            "finished_at": "",
            "status": "running",
            "message": f"侵权数据采集已启动：{params['target']}，并发 {params['max_workers']}",
            "params": {
                "shops": list(params["selected_shops"]),
                "sites": list(params["selected_sites"]),
                "max_workers": params["max_workers"],
                "target": params["target"],
            },
        })

    task_thread = threading.Thread(
        target=run_infraction_collect_job,
        args=(
            params["selected_shops"],
            params["selected_sites"],
            params["max_workers"],
        ),
        daemon=True,
    )
    task_thread.start()
    return jsonify({
        "status": "success",
        "data": dict(_infraction_collect_state),
        "message": "侵权数据采集已在后台启动"
    })


@app.route('/api/infractions/collect/status', methods=['GET'])
@login_required
def api_collect_infractions_status():
    with _infraction_collect_lock:
        return jsonify({
            "status": "success",
            "data": dict(_infraction_collect_state),
        })


@app.route('/infringement-dashboard', methods=['GET'])
@login_required
def official_infraction_dashboard_page():
    user = get_current_workbench_user()
    if not workbench_user_has_permission(user, "infractions.view"):
        return Response(
            "当前账号没有查看违规商品数据的权限",
            status=403,
            content_type="text/plain; charset=utf-8",
        )
    return render_template(
        'infraction_dashboard.html',
        current_user=user or {},
        embedded=str(request.args.get("embedded") or "").strip().lower()
        in {"1", "true", "yes", "on"},
    )


@app.route('/api/official-infractions/dashboard', methods=['GET'])
@login_required
def api_official_infraction_dashboard():
    try:
        filters = {
            "days": request.args.get("days", 30),
            "view_mode": request.args.get("view_mode", "current"),
            "group_name": request.args.get("group_name", ""),
            "salesperson": request.args.get("salesperson", ""),
            "source_type": request.args.get("source_type", ""),
            "category": request.args.get("category", ""),
            "search": request.args.get("search", ""),
            "detail_token_id": request.args.get("detail_token_id", 0),
            "page": request.args.get("page", 1),
            "page_size": request.args.get("page_size", 100),
        }
        data = (
            bit_db_api.list_official_infraction_dashboard(**filters)
            if USE_DB_API
            else list_infraction_dashboard(**filters)
        )
        return jsonify({"status": "success", "data": data})
    except (TypeError, ValueError) as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("读取官方违规商品分组看板失败")
        return jsonify({"status": "error", "message": f"读取违规商品数据失败：{exc}"}), 500


@app.route('/api/official-infractions/sync', methods=['POST'])
@login_required
def api_start_official_infraction_sync():
    payload = request.get_json(silent=True) or {}
    token_ids = payload.get("token_ids") or []
    if not isinstance(token_ids, list):
        return jsonify({"status": "error", "message": "token_ids 必须是数组"}), 400
    try:
        start_operation = (
            bit_db_api.start_official_infraction_sync
            if USE_DB_API
            else mercado_infraction_sync.start_official_infraction_sync
        )
        started, state = start_operation(token_ids)
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("启动官方侵权同步失败")
        return jsonify({"status": "error", "message": f"启动同步失败：{exc}"}), 500
    if not started:
        return jsonify({
            "status": "running",
            "message": state.get("message") or "官方违规商品数据正在同步",
            "data": state,
        }), 409
    return jsonify({
        "status": "success",
        "message": "官方违规商品数据同步已在后台启动",
        "data": state,
    }), 202


@app.route('/api/official-infractions/sync/status', methods=['GET'])
@login_required
def api_official_infraction_sync_status():
    return jsonify({
        "status": "success",
        "data": (
            bit_db_api.get_official_infraction_sync_status()
            if USE_DB_API
            else mercado_infraction_sync.official_infraction_sync_status()
        ),
    })


@app.route('/api/reputation/latest', methods=['GET'])
@login_required
def api_latest_reputation():
    try:
        data = db_get_latest_reputation_info()
        _attach_reputation_token_ids(data)
        _attach_latest_reputation_infraction_counts(data)
        return jsonify({
            "status": "success",
            "data": data,
        })
    except Exception as e:
        logging.error(f"Latest reputation query failed: {str(e)}")
        return jsonify({
            "status": "error",
            "message": f"Database error: {str(e)}"
        }), 500


def _attach_reputation_token_ids(data):
    """按授权显示名/昵称给旧声誉表补充打开浏览器所需的 token_id。"""
    rows = (data or {}).get("rows") or []
    try:
        tokens = (bit_db_api.list_mercado_store_tokens() or {}).get("rows") or []
    except Exception as exc:
        logging.warning("声誉数据匹配店铺授权失败：%s", exc)
        return data
    candidates = {}
    settings_by_token_site = {}
    duplicate_aliases = set()
    for token in tokens:
        token_id = int(token.get("id") or 0)
        if not token_id:
            continue
        for value in (token.get("display_name"), token.get("nickname")):
            alias = str(value or "").strip().casefold()
            if not alias:
                continue
            if alias in candidates and candidates[alias] != token_id:
                duplicate_aliases.add(alias)
            else:
                candidates[alias] = token_id
        for setting in token.get("site_settings") or []:
            site_id = str(setting.get("site_id") or "").strip().upper()
            if site_id:
                settings_by_token_site[(token_id, site_id)] = dict(setting)
    for alias in duplicate_aliases:
        candidates.pop(alias, None)
    for row in rows:
        alias = str(row.get("店铺名") or "").strip().casefold()
        token_id = int(candidates.get(alias) or 0)
        site_id = bit_reputation_info._normalize_api_site_code(row.get("站点"))
        setting = settings_by_token_site.get((token_id, site_id)) or {}
        row["token_id"] = token_id
        row["业务员"] = str(setting.get("salesperson") or "").strip() or "未分配"
        row["账户组"] = str(setting.get("group_name") or "").strip() or "未分组"
    return data


def _attach_latest_reputation_infraction_counts(data, recent_days=100):
    """让声誉表数量实时对齐最新 API 侵权列表的当前视图。"""
    rows = (data or {}).get("rows") or []
    try:
        snapshot = (
            bit_db_api.get_current_infraction_counts_by_token_site(recent_days)
            if USE_DB_API
            else current_infraction_counts_by_token_site(recent_days)
        ) or {}
    except Exception as exc:
        logging.warning("声誉数据读取最新 API 侵权汇总失败：%s", exc)
        return data
    counts = snapshot.get("counts") or {}
    for row in rows:
        token_id = int(row.get("token_id") or 0)
        if token_id <= 0:
            continue
        key = (
            token_id,
            bit_reputation_info._normalize_api_site_code(row.get("站点")),
        )
        current = counts.get(key) or {}
        row["侵权数量"] = int(current.get("infraction_count") or 0)
        row["权利人数量"] = int(current.get("rights_holder_count") or 0)
        row["侵权统计天数"] = int(snapshot.get("days") or recent_days)
        row["侵权数据来源"] = "official_infraction_dashboard"
        row["侵权列表同步时间"] = str(snapshot.get("last_synced_at") or "")
    return data


@app.route('/api/reputation/latest/export', methods=['GET'])
@login_required
def api_export_latest_reputation():
    try:
        data = db_get_latest_reputation_info()
        _attach_reputation_token_ids(data)
        _attach_latest_reputation_infraction_counts(data)
        rows = data.get("rows") or []
        wb = Workbook()
        ws = wb.active
        ws.title = "最新声誉数据"

        columns = [
            "店铺名", "站点", "站点状态", "侵权数量", "权利人数量", "声誉颜色", "总单量", "投诉率", "延误率", "取消率",
            "增加或减少", "近七天变化率", "一周流量趋势", "系统告警",
            "更新时间", "提交时间"
        ]
        ws.append(columns)

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        header_font = Font(bold=True, color="1F2937")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in rows:
            ws.append([row.get(column, "") for column in columns])

        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 36)

        ws.freeze_panes = "A2"
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        submit_time = str(data.get("latest_submit_time") or datetime.now().strftime("%Y%m%d%H%M%S"))
        safe_time = "".join(ch if ch.isdigit() else "" for ch in submit_time) or datetime.now().strftime("%Y%m%d%H%M%S")
        filename = f"最新声誉数据_{safe_time}.xlsx"
        encoded_filename = quote(filename)
        response = send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )
        response.headers["Content-Disposition"] = f"attachment; filename*=UTF-8''{encoded_filename}"
        return response
    except Exception as e:
        logging.error(f"Latest reputation export failed: {str(e)}")
        return jsonify({
            "status": "error",
            "message": f"Export error: {str(e)}"
        }), 500


@app.route('/api/reputation/collect', methods=['POST'])
@login_required
def api_collect_reputation():
    try:
        params = _parse_collection_request(
            request.get_json(silent=True) or {},
            authorization_flag="reputation_update_enabled",
        )
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.error("读取声誉采集范围失败：%s", exc)
        return jsonify({"status": "error", "message": f"读取采集范围失败：{exc}"}), 500

    return _start_reputation_collect_task(params, operation="rerun")


@app.route('/api/reputation/update-selected', methods=['POST'])
@login_required
def api_update_selected_reputation():
    try:
        params = _parse_collection_request(
            request.get_json(silent=True) or {},
            authorization_flag="reputation_update_enabled",
        )
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.error("读取所选店铺声誉更新范围失败：%s", exc)
        return jsonify({"status": "error", "message": f"读取更新范围失败：{exc}"}), 500

    return _start_reputation_collect_task(params, operation="selected_update")


def _start_reputation_collect_task(params, *, operation):
    action_label = (
        "所选店铺声誉更新"
        if operation == "selected_update"
        else "声誉数据补跑"
    )

    with _reputation_collect_lock:
        if _reputation_collect_state.get("running"):
            return jsonify({
                "status": "running",
                "data": dict(_reputation_collect_state),
                "message": "另一个声誉数据任务正在运行中"
            }), 409

        started_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        _reputation_collect_state.update({
            "running": True,
            "started_at": started_at,
            "finished_at": "",
            "status": "running",
            "message": f"{action_label}已启动：{params['target']}，并发 {params['max_workers']}",
            "operation": operation,
            "params": {
                "shops": list(params["selected_shops"]),
                "sites": list(params["selected_sites"]),
                "max_workers": params["max_workers"],
                "target": params["target"],
            },
        })

    task_thread = threading.Thread(
        target=run_reputation_collect_job,
        args=(
            params["selected_shops"],
            params["selected_sites"],
            params["max_workers"],
        ),
        daemon=True,
    )
    task_thread.start()
    return jsonify({
        "status": "success",
        "data": dict(_reputation_collect_state),
        "message": f"{action_label}已在后台启动"
    })


@app.route('/api/reputation/collect/status', methods=['GET'])
@login_required
def api_collect_reputation_status():
    with _reputation_collect_lock:
        return jsonify({
            "status": "success",
            "data": dict(_reputation_collect_state),
        })


@app.route('/api/collections/options', methods=['GET'])
@login_required
def api_collection_options():
    try:
        response = jsonify({
            "status": "success",
            "data": _collection_config_options(include_failures=True),
        })
        response.headers["Cache-Control"] = "no-store"
        return response
    except Exception as exc:
        logging.error("读取采集店铺和站点失败：%s", exc)
        return jsonify({
            "status": "error",
            "message": f"读取采集店铺和站点失败：{exc}",
        }), 500


@app.route('/api/funds/latest', methods=['GET'])
@login_required
def api_latest_funds():
    try:
        salesperson = str(request.args.get("salesperson") or "").strip()
        response = jsonify({
            "status": "success",
            "data": db_get_latest_pago_info(salesperson),
        })
        response.headers["Cache-Control"] = "no-store"
        return response
    except Exception as e:
        logging.error("Latest funds query failed: %s", e)
        return jsonify({
            "status": "error",
            "message": f"Database error: {str(e)}",
        }), 500


@app.route('/api/funds/collect', methods=['POST'])
@login_required
def api_collect_funds():
    global _fund_collect_stop_event
    data = request.get_json(silent=True) or {}
    all_shops = _parse_bool_param(data, "all_shops", False)
    salesperson = str(data.get("salesperson") or "").strip()
    max_workers = _parse_int_param(
        data,
        "max_workers",
        DEFAULT_COLLECTION_MAX_WORKERS,
        1,
        10,
    )
    raw_window_ids = data.get("window_ids", [])
    if not isinstance(raw_window_ids, list):
        return jsonify({"status": "error", "message": "window_ids 必须是数组"}), 400
    selected_window_ids = tuple(
        dict.fromkeys(
            str(value or "").strip()
            for value in raw_window_ids
            if str(value or "").strip()
        )
    )
    legacy_window_id = str(data.get("window_id") or "").strip()
    legacy_shop_name = str(data.get("shop_name") or "").strip()
    if legacy_window_id and legacy_window_id not in selected_window_ids:
        selected_window_ids += (legacy_window_id,)
    if len(selected_window_ids) > 500:
        return jsonify({"status": "error", "message": "单次最多选择 500 家店铺"}), 400

    try:
        configs = list_shop_configs(include_ignored=False) or []
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"读取店铺配置失败：{str(e)}",
        }), 500
    valid_configs = [
        config
        for config in configs
        if str(config.get("window_id") or "").strip()
        and str(config.get("shop_name") or "").strip()
    ]
    if salesperson:
        valid_configs = [
            config
            for config in valid_configs
            if (str(config.get("salesperson") or "").strip() or "未分配") == salesperson
        ]
    config_by_window = {
        str(config.get("window_id") or "").strip(): config
        for config in valid_configs
    }
    if not selected_window_ids and legacy_shop_name:
        legacy_config = next(
            (
                config
                for config in valid_configs
                if str(config.get("shop_name") or "").strip() == legacy_shop_name
            ),
            None,
        )
        if legacy_config is not None:
            selected_window_ids = (
                str(legacy_config.get("window_id") or "").strip(),
            )

    if all_shops:
        target_configs = valid_configs
        selected_window_ids = ()
    else:
        if not selected_window_ids:
            return jsonify({"status": "error", "message": "请至少勾选一家店铺"}), 400
        unknown_ids = [window_id for window_id in selected_window_ids if window_id not in config_by_window]
        if unknown_ids:
            return jsonify({
                "status": "error",
                "message": "所选店铺不存在、已被忽略或不属于当前归属人",
            }), 404
        target_configs = [config_by_window[window_id] for window_id in selected_window_ids]

    if not target_configs:
        return jsonify({
            "status": "error",
            "message": "当前归属人没有可执行的有效店铺" if salesperson else "没有可执行的有效店铺",
        }), 400

    owner_label = f"归属人 {salesperson}的" if salesperson else ""
    target = (
        f"{owner_label}全部 {len(target_configs)} 家有效店铺"
        if all_shops
        else f"所选 {len(target_configs)} 家店铺"
    )
    with _fund_collect_lock:
        if _fund_collect_state.get("running"):
            return jsonify({
                "status": "running",
                "data": dict(_fund_collect_state),
                "message": "资金数据采集正在运行中",
            }), 409
        stop_event = threading.Event()
        _fund_collect_stop_event = stop_event
        _fund_collect_state.update({
            "running": True,
            "started_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "finished_at": "",
            "status": "running",
            "message": f"{target}资金数据采集已启动",
            "scope": "all" if all_shops else "selected",
            "target": target,
            "salesperson": salesperson,
            "target_count": len(target_configs),
            "window_ids": list(selected_window_ids),
            "collected_count": 0,
        })
        task_state = dict(_fund_collect_state)

    task_thread = threading.Thread(
        target=run_fund_collect_job,
        kwargs={
            "all_shops": all_shops,
            "selected_window_ids": selected_window_ids,
            "salesperson": salesperson,
            "max_workers": max_workers,
            "stop_event": stop_event,
        },
        daemon=True,
    )
    try:
        task_thread.start()
    except Exception:
        with _fund_collect_lock:
            if _fund_collect_stop_event is stop_event:
                _fund_collect_stop_event = None
            _fund_collect_state.update({
                "running": False,
                "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "status": "error",
                "message": "资金数据采集启动失败",
            })
        raise
    return jsonify({
        "status": "success",
        "data": task_state,
        "message": f"{target}资金数据采集已在后台启动",
    })


@app.route('/api/funds/collect/stop', methods=['POST'])
@login_required
def api_stop_funds_collection():
    with _fund_collect_lock:
        if not _fund_collect_state.get("running") or _fund_collect_stop_event is None:
            return jsonify({
                "status": "error",
                "data": dict(_fund_collect_state),
                "message": "当前没有正在运行的资金采集任务",
            }), 409
        _fund_collect_stop_event.set()
        _fund_collect_state.update({
            "status": "stopping",
            "message": "正在终止资金数据采集，已打开的窗口将在安全边界关闭",
        })
        state = dict(_fund_collect_state)
    return jsonify({
        "status": "success",
        "data": state,
        "message": "已发送终止指令",
    })


@app.route('/api/funds/collect/status', methods=['GET'])
@login_required
def api_collect_funds_status():
    with _fund_collect_lock:
        response = jsonify({
            "status": "success",
            "data": dict(_fund_collect_state),
        })
    response.headers["Cache-Control"] = "no-store"
    return response


@app.route('/api/order-print/options', methods=['GET'])
@login_required
def api_order_print_options():
    try:
        response = jsonify({"status": "success", "data": _order_print_config_options()})
        response.headers["Cache-Control"] = "no-store"
        return response
    except Exception as exc:
        logging.error("读取 API 打印店铺失败：%s", exc)
        return jsonify({
            "status": "error",
            "message": f"读取美客多授权店铺失败：{exc}",
        }), 500


@app.route('/api/order-print/start', methods=['POST'])
@login_required
def api_start_order_print():
    global _order_print_stop_event
    data = request.get_json(silent=True) or {}
    try:
        params = build_order_print_params(data)
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.error("读取订单打印范围失败：%s", exc)
        return jsonify({"status": "error", "message": f"读取店铺配置失败：{exc}"}), 500

    with _order_print_lock:
        if _order_print_state.get("running"):
            return jsonify({
                "status": "running",
                "data": _order_print_snapshot(),
                "message": "订单打印任务正在运行",
            }), 409
        task_lock = bit_print.acquire_order_print_lock(
            owner="bit_interface.py",
            mode="once",
        )
        if task_lock is None:
            owner = bit_print.get_order_print_lock_owner()
            return jsonify({
                "status": "running",
                "data": {**_order_print_snapshot(), "lock_owner": owner},
                "message": "订单打印已在其他进程中运行",
            }), 409

        stop_event = threading.Event()
        _order_print_stop_event = stop_event
        _order_print_logs.clear()
        _order_print_state.update({
            "running": True,
            "started_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "finished_at": "",
            "status": "running",
            "message": f"{params['target']} API 订单打印已启动",
            "params": params,
            "task_id": params["task_id"],
            "printed": 0,
            "partial": 0,
            "no_orders": 0,
            "failed": 0,
            "skipped": 0,
            "printed_order_count": 0,
            "shipment_count": 0,
            "fallback_store_count": 0,
            "download_path": "",
            "download_name": "",
            "results": [],
        })
        task_state = _order_print_snapshot()

    task_thread = threading.Thread(
        target=run_order_print_job,
        args=(params, task_lock, stop_event),
        daemon=True,
    )
    try:
        task_thread.start()
    except Exception:
        task_lock.release()
        with _order_print_lock:
            if _order_print_stop_event is stop_event:
                _order_print_stop_event = None
            _order_print_state.update({
                "running": False,
                "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "status": "error",
                "message": "订单打印任务启动失败",
            })
        raise
    return jsonify({
        "status": "success",
        "data": task_state,
        "message": "美客多 API 订单打印已在后台启动",
    })


@app.route('/api/order-print/stop', methods=['POST'])
@login_required
def api_stop_order_print():
    with _order_print_lock:
        if not _order_print_state.get("running") or _order_print_stop_event is None:
            return jsonify({
                "status": "error",
                "data": _order_print_snapshot(),
                "message": "当前没有正在运行的订单打印任务",
            }), 409
        _order_print_stop_event.set()
        _order_print_state.update({
            "status": "stopping",
            "message": "正在停止，当前 API 请求完成后会安全结束",
        })
        state = _order_print_snapshot()
    _append_order_print_log(f"{get_now_time()} 已收到停止 API 订单打印请求")
    return jsonify({
        "status": "success",
        "data": state,
        "message": "已发送停止指令",
    })


@app.route('/api/order-print/status', methods=['GET'])
@login_required
def api_order_print_status():
    state = _order_print_snapshot()
    # Polling happens every two seconds while a task runs.  Re-querying token
    # and history tables on every poll can make the status endpoint itself
    # unavailable when the database or Mercado sync is busy.
    if not state.get("site_last_runs") or not state.get("running"):
        _refresh_order_print_site_last_runs()
        state = _order_print_snapshot()
    external_owner = bit_print.get_order_print_lock_owner()
    if external_owner and not state.get("running"):
        state.update({
            "running": True,
            "status": "running",
            "message": "订单打印正在其他进程中运行",
            "started_at": external_owner.get("acquired_at", ""),
            "lock_owner": external_owner,
        })
    response = jsonify({"status": "success", "data": state})
    response.headers["Cache-Control"] = "no-store"
    return response


@app.route('/api/order-print/download', methods=['GET'])
@login_required
def api_order_print_download():
    with _order_print_lock:
        raw_path = str(_order_print_state.get("download_path") or "")
        download_name = str(_order_print_state.get("download_name") or "")
    if not raw_path:
        return jsonify({"status": "error", "message": "当前没有可下载的 API 面单"}), 404
    path = Path(raw_path).resolve()
    output_root = bit_print.ORDER_PRINT_OUTPUT_DIR.resolve()
    if not path.is_relative_to(output_root) or not path.is_file():
        return jsonify({"status": "error", "message": "面单文件不存在或已失效"}), 404
    response = send_file(
        path,
        mimetype="application/pdf",
        as_attachment=False,
        download_name=download_name or path.name,
        max_age=0,
    )
    response.headers["Cache-Control"] = "no-store"
    return response


@app.route('/api/order-analysis/import', methods=['POST'])
@login_required
def api_import_order_analysis():
    uploads = []
    skipped_files = []
    for upload in request.files.getlist("files"):
        filename = str(upload.filename or "").strip()
        if bit_update_orders.is_order_excel_file(filename):
            uploads.append((filename, upload.stream))
        elif filename:
            skipped_files.append(filename)
    if not uploads:
        return jsonify({
            "status": "error",
            "message": "所选文件夹中没有可导入的 .xlsx 或 .xlsm 文件",
        }), 400
    if not _order_analysis_import_lock.acquire(blocking=False):
        return jsonify({
            "status": "running",
            "message": "已有订单文件夹正在导入，请等待当前任务完成",
        }), 409
    try:
        result = bit_update_orders.update_order_sources(
            uploads,
            insert_func=db_insert_orders,
        )
        result["skipped_files"] = skipped_files
        result["completed_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        return jsonify({
            "status": "success",
            "data": result,
            "message": f"已更新 {result['imported_orders']} 个唯一订单",
        })
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("订单分析文件夹导入失败")
        return jsonify({
            "status": "error",
            "message": f"订单导入失败：{exc}",
        }), 500
    finally:
        _order_analysis_import_lock.release()


def _high_after_sale_query_params(values):
    try:
        limit = int(values.get("limit") or 100)
    except (TypeError, ValueError) as exc:
        raise ValueError("展示数量必须是整数") from exc
    return {
        "sort_by": str(values.get("sort_by") or "after_sale_quantity").strip(),
        "sort_dir": str(values.get("sort_dir") or "desc").strip(),
        "search": str(values.get("search") or "").strip(),
        "date_from": str(values.get("date_from") or "").strip(),
        "date_to": str(values.get("date_to") or "").strip(),
        "limit": max(1, min(limit, 500)),
    }


@app.route('/api/order-analysis/high-after-sales', methods=['GET'])
@login_required
def api_high_after_sale_alerts():
    try:
        data = db_get_high_after_sale_alerts(
            **_high_after_sale_query_params(request.args)
        )
        return jsonify({"status": "success", "data": data})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("读取高售后告警失败")
        return jsonify({"status": "error", "message": str(exc)}), 500


def _high_profit_query_params(values):
    try:
        limit = int(values.get("limit") or 100)
    except (TypeError, ValueError) as exc:
        raise ValueError("展示数量必须是整数") from exc
    return {
        "sort_by": str(values.get("sort_by") or "total_profit").strip(),
        "sort_dir": str(values.get("sort_dir") or "desc").strip(),
        "search": str(values.get("search") or "").strip(),
        "date_from": str(values.get("date_from") or "").strip(),
        "date_to": str(values.get("date_to") or "").strip(),
        "limit": max(1, min(limit, 500)),
    }


@app.route('/api/order-analysis/high-profits', methods=['GET'])
@login_required
def api_high_profit_products():
    try:
        data = db_get_high_profit_products(
            **_high_profit_query_params(request.args)
        )
        return jsonify({"status": "success", "data": data})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("读取高利润产品失败")
        return jsonify({"status": "error", "message": str(exc)}), 500


def _order_list_query_params(args):
    salespeople = []
    for value in args.getlist("salesperson"):
        name = str(value or "").strip()
        if name and name not in salespeople:
            salespeople.append(name)
    store_ids = []
    for value in args.getlist("store_id"):
        text = str(value or "").strip()
        if not text:
            continue
        try:
            store_id = int(text)
        except ValueError as exc:
            raise ValueError("店铺筛选参数无效") from exc
        if store_id <= 0:
            raise ValueError("店铺筛选参数无效")
        if store_id not in store_ids:
            store_ids.append(store_id)
    params = {
        "country": str(args.get("country") or "").strip(),
        "status": str(args.get("status") or "").strip(),
        "salesperson": salespeople[0] if len(salespeople) == 1 else "",
        "group_name": str(args.get("group_name") or "").strip(),
        "search": str(args.get("search") or "").strip(),
        "start_date": str(args.get("start_date") or "").strip(),
        "end_date": str(args.get("end_date") or "").strip(),
        "origin": str(args.get("origin") or "").strip(),
        "freight_variance": str(args.get("freight_variance") or "").strip(),
        "page": _parse_int_param(args, "page", 1, 1, 1000000),
        "page_size": _parse_int_param(args, "page_size", 200, 10, 200),
    }
    if len(salespeople) > 1:
        params["salespeople"] = salespeople
    if store_ids:
        params["store_ids"] = store_ids
    return params


@app.route('/api/orders', methods=['GET'])
@login_required
def api_orders():
    try:
        data = db_list_orders(**_order_list_query_params(request.args))
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("订单列表加载失败")
        return jsonify({
            "status": "error",
            "message": f"订单列表加载失败：{exc}",
        }), 502
    response = jsonify({"status": "success", "data": data})
    response.headers["Cache-Control"] = "no-store"
    return response


@app.route('/api/orders/weight-quote', methods=['POST'])
@login_required
def api_order_weight_quote():
    data = request.get_json(silent=True) or {}
    order_ids = data.get("order_ids") or []
    if not isinstance(order_ids, list):
        return jsonify({"status": "error", "message": "order_ids 必须是数组"}), 422
    try:
        quote = bit_db_api.get_order_weight_quote(order_ids)
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except KeyError as exc:
        return jsonify({"status": "error", "message": str(exc).strip("'")}), 404
    except Exception as exc:
        logging.exception("订单重量和重量表运费计算失败")
        return jsonify({
            "status": "error",
            "message": f"订单重量和重量表运费计算失败：{exc}",
        }), 502
    response = jsonify({"status": "success", "data": quote})
    response.headers["Cache-Control"] = "no-store"
    return response


@app.route('/api/order-sync/options', methods=['GET'])
@login_required
def api_order_sync_options():
    try:
        data = bit_db_api.list_mercado_store_tokens() or {}
        return jsonify({"status": "success", "data": data})
    except Exception as exc:
        logging.exception("读取订单同步店铺失败")
        return jsonify({"status": "error", "message": f"读取授权店铺失败：{exc}"}), 502


@app.route('/api/orders/bulk-update', methods=['POST'])
@login_required
def api_bulk_update_orders():
    data = request.get_json(silent=True) or {}
    order_ids = data.get("order_ids") or []
    if not isinstance(order_ids, list):
        return jsonify({"status": "error", "message": "order_ids 必须是数组"}), 422
    changes = {}
    for field in (
        "workflow_status", "purchase_order", "purchase_tracking",
        "logistics_company", "purchase_cost", "purchase_remark",
    ):
        if field in data:
            changes[field] = data.get(field)
    user = session.get("workbench_user") or {}
    try:
        result = bit_db_api.bulk_update_orders(
            order_ids,
            operator_id=user.get("id"),
            operator_name=user.get("display_name") or user.get("username") or "",
            **changes,
        )
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("批量更新订单失败")
        return jsonify({"status": "error", "message": f"批量更新订单失败：{exc}"}), 502
    return jsonify({
        "status": "success",
        "message": f"已更新 {int(result.get('matched') or 0)} 个订单",
        "data": result,
    })


@app.route('/api/orders/print', methods=['POST'])
@login_required
def api_print_orders_pdf():
    data = request.get_json(silent=True) or {}
    order_ids = data.get("order_ids") or []
    if not isinstance(order_ids, list):
        return jsonify({"status": "error", "message": "order_ids 必须是数组"}), 422
    try:
        result = bit_db_api.download_order_labels(order_ids)
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("美客多面单 PDF 下载失败")
        return jsonify({"status": "error", "message": f"美客多面单下载失败：{exc}"}), 502

    print_log_error = ""
    try:
        user = session.get("workbench_user") or {}
        bit_db_api.record_order_print_logs(
            result.get("order_ids") or [],
            operator_id=user.get("id"),
            operator_name=user.get("display_name") or user.get("username") or "",
        )
    except Exception as exc:
        # The PDF is already valid at this point.  Do not discard it just
        # because the audit write failed; report the condition to the UI so
        # operators can avoid immediately printing the same orders again.
        print_log_error = str(exc) or exc.__class__.__name__
        logging.exception("美客多面单已生成，但打印记录写入失败")
    response = send_file(
        BytesIO(result["content"]),
        mimetype="application/pdf",
        as_attachment=False,
        download_name=result.get("filename") or "mercado-labels.pdf",
        max_age=0,
    )
    successful_order_ids = [str(value) for value in result.get("order_ids") or []]
    skipped_order_count = int(
        result.get("skipped_order_count")
        or len(result.get("skipped_order_ids") or [])
    )
    failed_order_count = int(
        result.get("failed_order_count")
        or len(result.get("failed_order_ids") or [])
    )
    response.headers["X-Mercado-Shipment-Count"] = str(result.get("shipment_count") or 0)
    response.headers["X-Mercado-Printed-Order-Count"] = str(len(successful_order_ids))
    response.headers["X-Mercado-Printed-Order-Ids"] = ",".join(successful_order_ids)
    response.headers["X-Mercado-Skipped-Order-Count"] = str(skipped_order_count)
    response.headers["X-Mercado-Failed-Order-Count"] = str(failed_order_count)
    response.headers["X-Mercado-Result"] = (
        "partial" if skipped_order_count or failed_order_count else "success"
    )
    if print_log_error:
        response.headers["X-Mercado-Print-Log"] = "failed"
    response.headers["Cache-Control"] = "no-store"
    return response


@app.route('/api/orders/<order_id>/logs', methods=['GET'])
@login_required
def api_order_operation_logs(order_id):
    try:
        rows = bit_db_api.list_order_operation_logs(
            order_id,
            limit=_parse_int_param(request.args, "limit", 100, 1, 200),
        )
    except Exception as exc:
        logging.exception("订单操作日志加载失败")
        return jsonify({"status": "error", "message": f"订单操作日志加载失败：{exc}"}), 502
    response = jsonify({"status": "success", "data": {"rows": rows}})
    response.headers["Cache-Control"] = "no-store"
    return response


@app.route('/api/orders/<order_id>/tracking', methods=['GET'])
@login_required
def api_order_tracking(order_id):
    try:
        data = bit_db_api.get_order_tracking(order_id)
    except (KeyError, ValueError) as exc:
        return jsonify({"status": "error", "message": str(exc)}), 404
    except Exception as exc:
        logging.exception("物流轨迹查询失败")
        return jsonify({"status": "error", "message": f"物流轨迹查询失败：{exc}"}), 502
    response = jsonify({"status": "success", "data": data})
    response.headers["Cache-Control"] = "no-store"
    return response


def _inventory_list_params(args):
    shelf_text = str(args.get("shelf_id") or "").strip()
    return {
        "search": str(args.get("search") or "").strip(),
        "shelf_id": int(shelf_text) if shelf_text else None,
        "stock_status": str(args.get("stock_status") or "positive").strip(),
        "page": _parse_int_param(args, "page", 1, 1, 1_000_000),
        "page_size": _parse_int_param(args, "page_size", 50, 10, 200),
    }


@app.route('/api/inventory/stocks', methods=['GET'])
@login_required
def api_inventory_stocks():
    try:
        result = bit_db_api.list_inventory_stock(**_inventory_list_params(request.args))
        response = jsonify({"status": "success", "data": result})
        response.headers["Cache-Control"] = "no-store"
        return response
    except (TypeError, ValueError) as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("库存列表加载失败")
        return jsonify({"status": "error", "message": f"库存列表加载失败：{exc}"}), 502


@app.route('/api/inventory/shelves', methods=['GET', 'POST'])
@login_required
def api_inventory_shelves():
    try:
        if request.method == "GET":
            include_inactive = str(request.args.get("include_inactive") or "1").lower() not in (
                "0", "false", "no", "off",
            )
            result = bit_db_api.list_inventory_shelves(include_inactive=include_inactive)
        else:
            payload = request.get_json(silent=True)
            if not isinstance(payload, dict):
                return jsonify({"status": "error", "message": "货架内容必须是对象"}), 422
            result = bit_db_api.create_inventory_shelf(payload)
        response = jsonify({"status": "success", "data": result})
        response.headers["Cache-Control"] = "no-store"
        return response
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("货架操作失败")
        return jsonify({"status": "error", "message": f"货架操作失败：{exc}"}), 502


@app.route('/api/inventory/shelves/<int:shelf_id>', methods=['PATCH'])
@login_required
def api_update_inventory_shelf(shelf_id):
    payload = request.get_json(silent=True)
    if not isinstance(payload, dict):
        return jsonify({"status": "error", "message": "货架内容必须是对象"}), 422
    try:
        result = bit_db_api.update_inventory_shelf(shelf_id, payload)
        return jsonify({"status": "success", "data": result})
    except KeyError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 404
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("更新货架失败")
        return jsonify({"status": "error", "message": f"更新货架失败：{exc}"}), 502


@app.route('/api/inventory/matches', methods=['GET'])
@login_required
def api_inventory_matches():
    try:
        result = bit_db_api.list_inventory_matches(
            search=str(request.args.get("search") or "").strip(),
            limit=_parse_int_param(request.args, "limit", 30, 1, 100),
        )
        response = jsonify({"status": "success", "data": result})
        response.headers["Cache-Control"] = "no-store"
        return response
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("库存入库匹配加载失败")
        return jsonify({"status": "error", "message": f"订单和产品匹配失败：{exc}"}), 502


def _inventory_movement_params(args):
    shelf_text = str(args.get("shelf_id") or "").strip()
    return {
        "search": str(args.get("search") or "").strip(),
        "movement_type": str(args.get("movement_type") or "").strip(),
        "shelf_id": int(shelf_text) if shelf_text else None,
        "date_from": str(args.get("date_from") or "").strip(),
        "date_to": str(args.get("date_to") or "").strip(),
        "page": _parse_int_param(args, "page", 1, 1, 1_000_000),
        "page_size": _parse_int_param(args, "page_size", 50, 10, 200),
    }


@app.route('/api/inventory/movements', methods=['GET', 'POST'])
@login_required
def api_inventory_movements():
    try:
        if request.method == "GET":
            result = bit_db_api.list_inventory_movements(
                **_inventory_movement_params(request.args)
            )
        else:
            payload = request.get_json(silent=True)
            if not isinstance(payload, dict):
                return jsonify({"status": "error", "message": "出入库内容必须是对象"}), 422
            user = get_current_workbench_user() or {}
            payload["operator_id"] = user.get("id")
            payload["operator_name"] = (
                user.get("display_name") or user.get("username") or ""
            )
            result = bit_db_api.create_inventory_movement(payload)
        response = jsonify({"status": "success", "data": result})
        response.headers["Cache-Control"] = "no-store"
        return response
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("库存出入库操作失败")
        return jsonify({"status": "error", "message": f"库存出入库操作失败：{exc}"}), 502


@app.route('/api/order-sync/start', methods=['POST'])
@login_required
def api_start_order_sync():
    data = request.get_json(silent=True) or {}
    token_ids = data.get("token_ids") or []
    if not isinstance(token_ids, list):
        return jsonify({"status": "error", "message": "token_ids 必须是数组"}), 422
    try:
        result = bit_db_api.start_order_sync(
            start_date=str(data.get("start_date") or "").strip(),
            end_date=str(data.get("end_date") or "").strip(),
            token_ids=token_ids,
            mode="manual",
        )
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("启动订单同步失败")
        return jsonify({"status": "error", "message": f"启动订单同步失败：{exc}"}), 502
    started = bool(result.get("started"))
    return jsonify({
        "status": "success" if started else "running",
        "message": "订单拉取任务已启动" if started else "已有订单同步任务正在运行",
        "data": result.get("state") or {},
    }), 202 if started else 409


@app.route('/api/order-sync/status', methods=['GET'])
@login_required
def api_order_sync_status():
    try:
        data = bit_db_api.get_order_sync_status() or {}
        response = jsonify({"status": "success", "data": data})
        response.headers["Cache-Control"] = "no-store"
        return response
    except Exception as exc:
        logging.exception("读取订单同步状态失败")
        return jsonify({"status": "error", "message": f"读取订单同步状态失败：{exc}"}), 502


@app.route('/api/store-links', methods=['GET'])
@login_required
def api_store_links():
    try:
        token_text = str(request.args.get("token_id") or "").strip()
        data = bit_db_api.list_mercado_store_links(
            search=str(request.args.get("search") or "").strip(),
            token_id=int(token_text) if token_text else None,
            site_id=str(request.args.get("site_id") or "").strip(),
            group_name=str(request.args.get("group_name") or "").strip(),
            status=str(request.args.get("status") or "").strip(),
            management_category_id=str(
                request.args.get("management_category_id") or ""
            ).strip(),
            mercado_category=str(
                request.args.get("mercado_category") or ""
            ).strip(),
            sales_sort=str(request.args.get("sales_sort") or "desc").strip(),
            current_only=str(request.args.get("current_only") or "1").strip().lower()
            not in ("0", "false", "no", "off"),
            page=_parse_int_param(request.args, "page", 1, 1, 1000000),
            page_size=1000,
        )
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("店铺链接列表加载失败")
        return jsonify({"status": "error", "message": f"店铺链接列表加载失败：{exc}"}), 502
    response = jsonify({"status": "success", "data": data})
    response.headers["Cache-Control"] = "no-store"
    return response


@app.route('/api/store-links/bulk-update', methods=['POST'])
@login_required
def api_bulk_update_store_links():
    data = request.get_json(silent=True) or {}
    link_ids = data.get("link_ids") or []
    if not isinstance(link_ids, list):
        return jsonify({"status": "error", "message": "link_ids 必须是数组"}), 422
    allowed = (
        "price", "weight_g", "package_length_cm", "package_width_cm",
        "package_height_cm", "net_proceeds_usd",
    )
    changes = {field: data.get(field) for field in allowed if field in data}
    try:
        result = bit_db_api.bulk_update_mercado_store_links(link_ids, **changes)
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("批量更新店铺链接失败")
        return jsonify({"status": "error", "message": f"批量更新店铺链接失败：{exc}"}), 502
    started = bool(result.get("started"))
    return jsonify({
        "status": "success" if started else "running",
        "message": (
            "美客多后台修改任务已启动"
            if started else "已有美客多后台修改任务正在运行"
        ),
        "data": result.get("state") or {},
    }), 202 if started else 409


@app.route('/api/store-links/bulk-update/status', methods=['GET'])
@login_required
def api_store_link_remote_update_status():
    try:
        data = bit_db_api.get_mercado_store_link_remote_update_status() or {}
        response = jsonify({"status": "success", "data": data})
        response.headers["Cache-Control"] = "no-store"
        return response
    except Exception as exc:
        logging.exception("读取美客多后台修改任务状态失败")
        return jsonify({"status": "error", "message": f"读取后台修改状态失败：{exc}"}), 502


@app.route('/api/store-links/sync/start', methods=['POST'])
@login_required
def api_start_store_link_sync():
    data = request.get_json(silent=True) or {}
    sync_all = data.get("sync_all") is True
    token_ids = [] if sync_all else (data.get("token_ids") or [])
    if not isinstance(token_ids, list):
        return jsonify({"status": "error", "message": "token_ids 必须是数组"}), 422
    try:
        result = bit_db_api.start_store_link_sync(token_ids)
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("启动店铺链接同步失败")
        return jsonify({"status": "error", "message": f"启动店铺链接同步失败：{exc}"}), 502
    started = bool(result.get("started"))
    return jsonify({
        "status": "success" if started else "running",
        "message": "店铺链接同步已启动" if started else "已有店铺链接同步任务正在运行",
        "data": result.get("state") or {},
    }), 202 if started else 409


@app.route('/api/store-links/sync/status', methods=['GET'])
@login_required
def api_store_link_sync_status():
    try:
        data = bit_db_api.get_store_link_sync_status() or {}
        response = jsonify({"status": "success", "data": data})
        response.headers["Cache-Control"] = "no-store"
        return response
    except Exception as exc:
        logging.exception("读取店铺链接同步状态失败")
        return jsonify({"status": "error", "message": f"读取店铺链接同步状态失败：{exc}"}), 502


@app.route('/api/prohibited-listings', methods=['GET'])
@login_required
def api_prohibited_listings():
    try:
        token_text = str(request.args.get("token_id") or "").strip()
        data = bit_db_api.list_mercado_prohibited_listings(
            search=str(request.args.get("search") or "").strip(),
            token_id=int(token_text) if token_text else None,
            site_id=str(request.args.get("site_id") or "").strip(),
            salesperson=str(request.args.get("salesperson") or "").strip(),
            risk_type=str(request.args.get("risk_type") or "").strip(),
            page=_parse_int_param(request.args, "page", 1, 1, 1000000),
            page_size=_parse_int_param(request.args, "page_size", 100, 20, 500),
        )
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("禁限售列表加载失败")
        return jsonify({"status": "error", "message": f"禁限售列表加载失败：{exc}"}), 502
    response = jsonify({"status": "success", "data": data})
    response.headers["Cache-Control"] = "no-store"
    return response


@app.route('/api/prohibited-listings/sync/start', methods=['POST'])
@login_required
def api_start_prohibited_listing_sync():
    data = request.get_json(silent=True) or {}
    sync_all = data.get("sync_all") is True
    token_ids = [] if sync_all else (data.get("token_ids") or [])
    if not isinstance(token_ids, list):
        return jsonify({"status": "error", "message": "token_ids 必须是数组"}), 422
    try:
        salesperson = str(data.get("salesperson") or "").strip()
        if not sync_all and not token_ids and salesperson:
            token_rows = (bit_db_api.list_mercado_store_tokens() or {}).get("rows") or []
            token_ids = [
                int(row["id"])
                for row in token_rows
                if any(
                    str(setting.get("salesperson") or "").strip() == salesperson
                    for setting in row.get("site_settings") or []
                )
            ]
            if not token_ids:
                raise ValueError(f"业务员“{salesperson}”暂无已授权店铺")
        result = bit_db_api.start_prohibited_listing_sync(token_ids)
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("启动禁限售同步失败")
        return jsonify({"status": "error", "message": f"启动禁限售同步失败：{exc}"}), 502
    started = bool(result.get("started"))
    return jsonify({
        "status": "success" if started else "running",
        "message": "禁限售同步已启动" if started else "已有禁限售同步任务正在运行",
        "data": result.get("state") or {},
    }), 202 if started else 409


@app.route('/api/prohibited-listings/sync/status', methods=['GET'])
@login_required
def api_prohibited_listing_sync_status():
    try:
        data = bit_db_api.get_prohibited_listing_sync_status() or {}
        response = jsonify({"status": "success", "data": data})
        response.headers["Cache-Control"] = "no-store"
        return response
    except Exception as exc:
        logging.exception("读取禁限售同步状态失败")
        return jsonify({"status": "error", "message": f"读取禁限售同步状态失败：{exc}"}), 502


@app.route('/api/tasks/daily/start', methods=['POST'])
@login_required
def api_start_daily_task():
    data = request.get_json(silent=True) or {}
    try:
        params = build_daily_task_params(data)
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400

    if data.get("execution_target") == "agent":
        if params["execution_target"] != "agent":
            return jsonify({"status": "error", "message": "请在公网工作台选择 Agent"}), 400
        return enqueue_local_agent_daily_task(data.get("agent_id"), params)

    max_concurrent = _daily_task_max_concurrent()
    with _daily_task_lock:
        running_count = sum(
            1 for state in _daily_tasks.values() if state.get("running")
        )
        if running_count >= max_concurrent:
            limit_reached = True
            task_id = ""
            task_log_path = None
        else:
            limit_reached = False
            task_id = secrets.token_hex(8)
            while task_id in _daily_tasks:
                task_id = secrets.token_hex(8)
            task_log_path = _daily_task_log_file(task_id)
            started_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            task_state = {
                "task_id": task_id,
                "name": _daily_task_display_name(params),
                "execution_target": params["execution_target"],
                "running": True,
                "started_at": started_at,
                "finished_at": "",
                "status": "starting",
                "message": "daily_task 正在启动",
                "stop_requested": False,
                "can_stop": False,
                "log_path": str(task_log_path),
                "params": params,
            }
            # 在同一把锁内占住并发名额，避免两个并发 start 同时越过上限。
            _daily_tasks[task_id] = task_state
            _daily_task_state.update(task_state)

    if limit_reached:
        return jsonify({
            "status": "running",
            "data": _daily_tasks_snapshot(),
            "message": (
                f"同时运行的 daily_task 已达到上限 {max_concurrent}，"
                "请先停止或等待现有任务结束"
            ),
        }), 409

    task_lock = bit_daily_task.acquire_daily_task_lock(
        owner=f"bit_interface.py:{task_id}",
        mode=params["mode"],
        task_id=task_id,
    )
    if task_lock is None:
        owner = bit_daily_task.get_daily_task_lock_owner(task_id)
        _update_daily_task_state(
            task_id,
            running=False,
            finished_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            status="error",
            message="无法创建独立任务锁，请重试",
            can_stop=False,
        )
        _prune_daily_task_history()
        return jsonify({
            "status": "error",
            "data": {
                **_daily_task_snapshot(task_id),
                "lock_owner": owner,
            },
            "message": "无法创建独立任务锁，请重试",
        }), 409

    try:
        _reset_daily_task_log(task_log_path)
    except OSError as exc:
        task_lock.release()
        _update_daily_task_state(
            task_id,
            running=False,
            finished_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            status="error",
            message=f"无法创建 daily_task 运行日志：{exc}",
            can_stop=False,
        )
        _prune_daily_task_history()
        return jsonify({
            "status": "error",
            "data": _daily_task_snapshot(task_id),
            "message": f"无法创建 daily_task 运行日志：{exc}",
        }), 500

    stop_manager = None
    try:
        stop_manager = multiprocessing.Manager()
        stop_event = stop_manager.Event()
        owned_window_ids = stop_manager.dict()
    except Exception as exc:
        task_lock.release()
        if stop_manager is not None:
            try:
                stop_manager.shutdown()
            except Exception:
                pass
        _append_daily_task_log(
            f"daily_task 停止控制初始化失败：{exc}\n",
            task_log_path,
        )
        _update_daily_task_state(
            task_id,
            running=False,
            finished_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            status="error",
            message=f"daily_task 停止控制初始化失败：{exc}",
            can_stop=False,
        )
        _prune_daily_task_history()
        return jsonify({
            "status": "error",
            "data": _daily_task_snapshot(task_id),
            "message": f"daily_task 启动失败：{exc}",
        }), 500

    with _daily_task_lock:
        task_state = _daily_tasks[task_id]
        task_state.update({
            "status": "running",
            "message": "daily_task 已启动",
            "can_stop": True,
        })
        _daily_task_controls[task_id] = {
            "stop_event": stop_event,
            "stop_manager": stop_manager,
            "task_lock": task_lock,
            "thread": None,
            "owned_window_ids": owned_window_ids,
        }
        _daily_task_state.update(task_state)
    _prune_daily_task_history()

    try:
        task_thread = threading.Thread(
            target=run_daily_task_job,
            args=(
                params,
                task_lock,
                stop_event,
                task_id,
                task_log_path,
                owned_window_ids,
            ),
            name=f"daily-task-{task_id}",
            daemon=True,
        )
        task_thread.start()
        with _daily_task_lock:
            control = _daily_task_controls.get(task_id)
            if control is not None:
                control["thread"] = task_thread
    except Exception as exc:
        task_lock.release()
        with _daily_task_lock:
            _daily_task_controls.pop(task_id, None)
        _update_daily_task_state(
            task_id,
            running=False,
            finished_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            status="error",
            message=f"daily_task 启动失败：{exc}",
            can_stop=False,
        )
        try:
            stop_manager.shutdown()
        except Exception:
            pass
        return jsonify({
            "status": "error",
            "data": _daily_task_snapshot(task_id),
            "message": f"daily_task 启动失败：{exc}",
        }), 500
    return jsonify({
        "status": "success",
        "data": _daily_task_snapshot(task_id),
        "message": f"{task_state['name']} 已在后台启动",
    })


@app.route('/api/tasks/daily/stop', methods=['POST'])
@login_required
def api_stop_daily_task():
    global _daily_task_stop_event
    data = request.get_json(silent=True) or {}
    requested_task_id = str(data.get("task_id") or request.args.get("task_id") or "").strip()
    if request.args.get("execution_target") == "agent" or data.get("execution_target") == "agent":
        try:
            job = get_local_agent_store().get_job(requested_task_id)
        except ValueError as exc:
            return jsonify({"status": "error", "message": str(exc)}), 400
        if not job or job["job_type"] != "daily_task":
            return jsonify({"status": "error", "message": "没有找到指定 Agent 任务"}), 404
        get_local_agent_store().request_cancel(requested_task_id)
        job = get_local_agent_store().get_job(requested_task_id)
        return jsonify({"status": "success", "data": daily_agent_task_snapshot(job),
                        "message": "Agent 任务停止请求已提交"})
    task_id = _resolve_daily_task_id_for_stop(requested_task_id)
    if not task_id:
        with _daily_task_lock:
            running_count = sum(
                1 for state in _daily_tasks.values() if state.get("running")
            )
            legacy_running = bool(
                not _daily_tasks and _daily_task_state.get("running")
            )
            legacy_stop_event = _daily_task_stop_event
        if not requested_task_id and legacy_running:
            if legacy_stop_event is None:
                return jsonify({
                    "status": "error",
                    "data": _daily_tasks_snapshot(),
                    "message": "该任务无法从当前控制台停止",
                }), 409
            legacy_stop_event.set()
            with _daily_task_lock:
                _daily_task_state.update({
                    "status": "stopping",
                    "message": "已请求停止，正在终止任务进程并关闭浏览器窗口",
                    "stop_requested": True,
                })
            return jsonify({
                "status": "success",
                "data": _daily_tasks_snapshot(),
                "message": "daily_task 停止请求已提交",
            })
        message = (
            "没有找到指定任务"
            if requested_task_id
            else "daily_task 当前没有正在运行的任务"
            if running_count == 0
            else "有多个任务正在运行，请指定要停止的任务"
        )
        return jsonify({
            "status": "error",
            "data": _daily_tasks_snapshot(),
            "message": message,
        }), 404 if requested_task_id else 409

    with _daily_task_lock:
        task_state = _daily_tasks.get(task_id)
        if not task_state or not task_state.get("running"):
            return jsonify({
                "status": "idle",
                "data": _daily_task_snapshot(task_id),
                "message": "该 daily_task 当前未运行",
            }), 409
        control = _daily_task_controls.get(task_id) or {}
        stop_event = control.get("stop_event")
        if stop_event is None:
            return jsonify({
                "status": "error",
                "data": _daily_task_snapshot(task_id),
                "message": "该任务无法从当前控制台停止",
            }), 409
        stop_event.set()
        task_state.update({
            "status": "stopping",
            "message": "已请求停止该任务，正在终止其进程并关闭空闲窗口",
            "stop_requested": True,
        })
        if str(_daily_task_state.get("task_id") or "") == task_id:
            _daily_task_state.update(task_state)
        state = dict(task_state)
    state["log"] = _read_daily_task_log(log_path=state.get("log_path"))
    return jsonify({
        "status": "success",
        "data": state,
        "message": f"{state.get('name') or task_id} 停止请求已提交",
    })


@app.route('/api/tasks/daily/options', methods=['GET'])
@login_required
def api_daily_task_options():
    salespeople = []
    groups = []
    try:
        users = _workbench_backend("list_workbench_users") or []
        salespeople.extend(
            str(user.get("display_name") or user.get("username") or "").strip()
            for user in users
            if user.get("is_active") is not False
        )
    except Exception:
        logging.exception("读取任务模块业务员列表失败，回退到店铺授权配置")
    try:
        token_data = bit_db_api.list_mercado_store_tokens() or {}
        for token in token_data.get("rows") or ():
            for setting in token.get("site_settings") or ():
                salespeople.append(str(setting.get("salesperson") or "").strip())
                groups.append(str(setting.get("group_name") or "").strip())
    except Exception:
        logging.exception("从店铺授权读取任务模块业务员和店铺组失败")
    unique_salespeople = sorted(
        {name for name in salespeople if name},
        key=lambda value: value.casefold(),
    )
    unique_groups = sorted(
        {name for name in groups if name},
        key=lambda value: value.casefold(),
    )
    return jsonify({
        "status": "success",
        "data": {
            "salespeople": unique_salespeople,
            "groups": unique_groups,
        },
    })


@app.route('/api/tasks/daily/status', methods=['GET'])
@login_required
def api_daily_task_status():
    requested_task_id = str(request.args.get("task_id") or "").strip()
    if request.args.get("execution_target") == "agent":
        if requested_task_id:
            try:
                job = get_local_agent_store().get_job(requested_task_id)
            except ValueError as exc:
                return jsonify({"status": "error", "message": str(exc)}), 400
            if not job or job["job_type"] != "daily_task":
                return jsonify({"status": "error", "message": "没有找到指定 Agent 任务"}), 404
            data = daily_agent_task_snapshot(job, include_log=True)
        else:
            tasks = [daily_agent_task_snapshot(job) for job in
                     get_local_agent_store().list_jobs(job_type="daily_task", limit=500)]
            data = {"tasks": tasks, "running": any(task["running"] for task in tasks),
                    "running_count": sum(task["running"] for task in tasks), "total_count": len(tasks)}
        return jsonify({"status": "success", "data": data})
    if requested_task_id:
        task = _daily_task_snapshot(requested_task_id)
        if not task:
            return jsonify({
                "status": "error",
                "message": "没有找到指定任务",
            }), 404
        data = task
    else:
        data = _daily_tasks_snapshot()
        # 保留独立脚本/bit_main 的全局锁提示，但它不阻止页面新建独立任务。
        external_owner = bit_daily_task.get_daily_task_lock_owner()
        if external_owner:
            external_task = {
                "task_id": "external-daily-task",
                "name": "外部 daily_task",
                "execution_target": (
                    "local" if RUNTIME_SETTINGS.is_client else "server"
                ),
                "running": True,
                "status": "running",
                "message": "daily_task 正在其他进程中运行",
                "started_at": external_owner.get("acquired_at", ""),
                "finished_at": "",
                "stop_requested": False,
                "lock_owner": external_owner,
                "can_stop": False,
                "params": {},
            }
            local_tasks = list(data.get("tasks") or ())
            local_running = any(task.get("running") for task in local_tasks)
            data["external_task"] = external_task
            data["running"] = True
            data["running_count"] = int(data.get("running_count") or 0) + 1
            data["total_count"] = int(data.get("total_count") or 0) + 1
            if not local_running:
                # 旧客户端只读取顶层单任务字段；没有本地运行任务时应投影
                # 外部任务，同时保留新列表客户端需要的聚合字段。
                tasks = data.get("tasks", [])
                running_count = data["running_count"]
                total_count = data["total_count"]
                data.update(external_task)
                data.update({
                    "tasks": tasks,
                    "external_task": external_task,
                    "running": True,
                    "running_count": running_count,
                    "total_count": total_count,
                })
    return jsonify({
        "status": "success",
        "data": data,
    })


def _local_executor_database_preflight():
    """Check the existing data connection before accepting a browser job."""
    try:
        health = bit_db_api.get_database_api_health() or {}
        if health.get("role") != "server":
            raise RuntimeError("客户端配置的地址不是数据服务端")
    except Exception:
        return jsonify({
            "status": "error",
            "message": (
                "本机执行端已连接，但无法访问服务端数据接口。"
                "请检查客户端服务端地址；数据接口启用令牌认证时，"
                "两端需配置相同的 BIT_DB_API_TOKEN。"
            ),
        }), 503
    return None


def _agent_request_identity(data):
    agent_id = normalize_agent_id((data or {}).get("agent_id"))
    claims = getattr(g, "local_agent_claims", {}) or {}
    if claims.get("agent_id") not in ("*", agent_id):
        raise PermissionError("Agent 凭证与电脑编号不匹配")
    return agent_id


@app.route("/api/local-agents/enroll", methods=["POST"])
def api_local_agent_enroll():
    if USE_DB_API:
        return jsonify({"status": "error", "message": "请连接服务端工作台"}), 503
    authorization = str(request.headers.get("Authorization") or "")
    token = authorization[7:].strip() if authorization.lower().startswith("bearer ") else ""
    user_claims = _local_agent_enrollment_user(token)
    if not user_claims:
        return jsonify({
            "status": "error",
            "message": "Agent 注册链接无效或已过期，请从泽顺控制台重新下载",
        }), 401
    data = request.get_json(silent=True) or {}
    try:
        agent_id = normalize_agent_id(data.get("agent_id"))
        agent = get_local_agent_store().heartbeat(
            agent_id,
            name=data.get("name"),
            hostname=data.get("hostname"),
            platform=data.get("platform"),
            agent_version=data.get("agent_version"),
            business_version="",
            capabilities=data.get("capabilities") or ("appeal",),
        )
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    credential = create_local_agent_credential(agent_id, user_claims.get("id"))
    return jsonify({
        "status": "success",
        "data": {"agent": agent, "agent_token": credential},
    })


@app.route("/api/local-agents/heartbeat", methods=["POST"])
@local_agent_required
def api_local_agent_heartbeat():
    data = request.get_json(silent=True) or {}
    try:
        agent_id = _agent_request_identity(data)
        agent = get_local_agent_store().heartbeat(
            agent_id,
            name=data.get("name"),
            hostname=data.get("hostname"),
            platform=data.get("platform"),
            agent_version=data.get("agent_version"),
            business_version=data.get("business_version"),
            capabilities=data.get("capabilities") or ("appeal",),
        )
        bundle = current_local_agent_bundle()
        claims = getattr(g, "local_agent_claims", {}) or {}
        refreshed_credential = ""
        if time.time() - float(claims.get("issued_at") or 0) >= 30 * 24 * 60 * 60:
            refreshed_credential = create_local_agent_credential(
                agent_id, claims.get("user_id")
            )
        return jsonify({
            "status": "success",
            "data": {
                "agent": agent,
                "agent_token": refreshed_credential,
                "cancel_job_ids": get_local_agent_store().cancellation_job_ids(agent_id),
                "bundle": {
                    "version": bundle["version"],
                    "sha256": bundle["sha256"],
                    "size": bundle["size"],
                },
            },
        })
    except PermissionError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 403
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400


@app.route("/api/local-agents/jobs/claim", methods=["POST"])
@local_agent_required
def api_local_agent_claim_job():
    data = request.get_json(silent=True) or {}
    try:
        agent_id = _agent_request_identity(data)
        job = get_local_agent_store().claim_job(agent_id)
        return jsonify({"status": "success", "data": {"job": job}})
    except PermissionError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 403
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400


@app.route("/api/local-agents/jobs/<job_id>/events", methods=["POST"])
@local_agent_required
def api_local_agent_job_event(job_id):
    data = request.get_json(silent=True) or {}
    try:
        agent_id = _agent_request_identity(data)
        job = get_local_agent_store().append_event(
            job_id,
            agent_id,
            content=data.get("content"),
            event_type=data.get("event_type") or "log",
            status=data.get("status"),
            message=data.get("message"),
            result=data.get("result"),
        )
        return jsonify({"status": "success", "data": {"job": job}})
    except PermissionError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 403
    except KeyError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 404
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400


@app.route("/api/local-agents/business-bundle", methods=["GET"])
@local_agent_required
def api_local_agent_business_bundle():
    bundle = current_local_agent_bundle()
    response = send_file(
        BytesIO(bundle["content"]),
        mimetype="application/zip",
        as_attachment=True,
        download_name=f"mercado-business-{bundle['version']}.zip",
        max_age=0,
    )
    response.headers["X-Business-Version"] = bundle["version"]
    response.headers["X-Bundle-SHA256"] = bundle["sha256"]
    response.headers["Cache-Control"] = "no-store"
    return response


@app.route("/api/execution-agents", methods=["GET"])
@login_required
def api_execution_agents():
    agents = get_local_agent_store().list_agents(
        online_seconds=LOCAL_AGENT_ONLINE_SECONDS,
        capability=str(request.args.get("capability") or "appeal"),
    )
    response = jsonify({"status": "success", "data": {"agents": agents}})
    response.headers["Cache-Control"] = "no-store"
    return response


@app.route("/api/local-agents/download", methods=["GET"])
@login_required
def api_download_local_agent():
    user = get_current_workbench_user()
    if not any(workbench_user_has_permission(user, permission)
               for permission in ("appeal.execute", "tasks.execute")):
        return jsonify({"status": "error", "message": "当前账号没有 Agent 任务执行权限"}), 403
    enrollment_token = create_local_agent_enrollment_token(user)
    public_url = str(
        os.environ.get("BIT_PUBLIC_WORKBENCH_URL")
        or RUNTIME_SETTINGS.api_base_url
        or request.url_root
    ).strip().rstrip("/")
    package = build_agent_distribution(
        PROJECT_ROOT,
        server_url=public_url,
        enrollment_token=enrollment_token,
    )
    response = send_file(
        BytesIO(package["content"]),
        mimetype="application/zip",
        as_attachment=True,
        download_name="Zeshun-MercadoLocalAgent.zip",
        max_age=0,
    )
    response.headers["X-Agent-Package-Format"] = package["format"]
    response.headers["Cache-Control"] = "no-store"
    return response


def stream_local_agent_job(job_id):
    event_id = 0
    last_heartbeat = time.monotonic()
    while True:
        events = get_local_agent_store().events_after(job_id, event_id)
        for event in events:
            event_id = max(event_id, int(event["event_id"]))
            if event.get("content"):
                yield get_local_agent_store().render_event_content(event)
        job = get_local_agent_store().get_job(job_id)
        if not job:
            yield "Agent 任务记录不存在\n"
            return
        if job.get("status") in TERMINAL_JOB_STATUSES and not events:
            if job.get("status") == "error" and job.get("message"):
                yield f"任务失败：{job['message']}\n"
            return
        now = time.monotonic()
        if not events and now - last_heartbeat >= APPEAL_STREAM_HEARTBEAT_SECONDS:
            yield "\n"
            last_heartbeat = now
        time.sleep(0.5)


def daily_agent_task_snapshot(job, *, include_log=False):
    params = dict(job.get("payload") or {})
    result = job.get("result") or {}
    running = job["status"] not in TERMINAL_JOB_STATUSES
    def timestamp(value):
        return datetime.fromtimestamp(value).strftime("%Y-%m-%d %H:%M:%S") if value else ""
    state = {
        "task_id": job["job_id"], "agent_id": job["agent_id"],
        "agent_name": params.get("agent_name") or job["agent_id"],
        "name": _daily_task_display_name(params), "execution_target": "agent",
        "running": running, "can_stop": running,
        "status": "partial" if job["status"] == "success" and result.get("status") == "partial" else job["status"],
        "message": job["message"], "stop_requested": job["cancel_requested"],
        "started_at": timestamp(job.get("started_at")),
        "created_at": timestamp(job.get("created_at")),
        "finished_at": timestamp(job.get("finished_at")),
        "params": params, "execution_counts": result.get("execution_counts", {}),
    }
    if include_log:
        state["log"] = get_local_agent_store().recent_log(job["job_id"])
    return state


def enqueue_local_agent_daily_task(agent_id, params):
    store = get_local_agent_store()
    try:
        agent_id = normalize_agent_id(agent_id)
        agent = store.get_agent(agent_id, online_seconds=LOCAL_AGENT_ONLINE_SECONDS)
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    if not agent or not agent["online"]:
        return jsonify({"status": "error", "message": "所选 Agent 不在线，请启动 Agent 并刷新电脑"}), 409
    if "daily_task" not in agent.get("capabilities", ()):
        return jsonify({"status": "error", "message": "请将所选电脑的 Agent 更新至 1.1.0 或更高版本"}), 409
    params = {
        **params,
        "execution_target": "agent",
        "agent_id": agent_id,
        "agent_name": agent["name"],
    }
    user = get_current_workbench_user() or {}
    job = store.enqueue_job(
        secrets.token_hex(16), agent_id, "daily_task", params,
        required_version=current_local_agent_bundle()["version"],
        created_by_id=user.get("id"),
        created_by_name=user.get("display_name") or user.get("username") or "",
    )
    return jsonify({"status": "success", "data": daily_agent_task_snapshot(job),
                    "message": f"任务已提交到 {agent['name']}，等待 Agent 执行"})


def enqueue_local_agent_appeal(
    *, task_id, agent_id, name, sites, forms, message, mode, loop_count
):
    try:
        agent_id = normalize_agent_id(agent_id)
        agent = get_local_agent_store().get_agent(
            agent_id, online_seconds=LOCAL_AGENT_ONLINE_SECONDS
        )
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    if not agent or not agent.get("online"):
        return jsonify({
            "status": "error",
            "message": "所选本机 Agent 当前不在线，请刷新电脑列表或启动 Agent",
        }), 409
    if "appeal" not in agent.get("capabilities", ()):
        return jsonify({"status": "error", "message": "所选 Agent 不支持申诉任务"}), 409
    user = get_current_workbench_user() or {}
    bundle = current_local_agent_bundle()
    payload = {
        "name": name,
        "sites": list(sites),
        "forms": list(forms),
        "message": message,
        "mode": mode,
        "loop_count": "永久" if loop_count == PERMANENT_APPEAL_LOOP_COUNT else loop_count,
        "execution_target": "agent",
        "agent_id": agent_id,
        "agent_name": agent["name"],
    }
    try:
        get_local_agent_store().enqueue_job(
            task_id,
            agent_id,
            "appeal",
            payload,
            required_version=bundle["version"],
            created_by_id=user.get("id"),
            created_by_name=user.get("display_name") or user.get("username") or "",
        )
    except Exception as exc:
        logging.exception("创建本机 Agent 申诉任务失败")
        return jsonify({
            "status": "error",
            "message": f"创建本机 Agent 任务失败：{exc}",
        }), 409

    response = Response(stream_local_agent_job(task_id), mimetype="text/plain; charset=utf-8")
    response.headers["Cache-Control"] = "no-cache"
    response.headers["X-Accel-Buffering"] = "no"
    response.headers["X-Appeal-Task-ID"] = task_id
    response.headers["X-Execution-Target"] = "agent"
    response.headers["X-Execution-Agent-ID"] = agent_id
    return response


@app.route('/api/local-executor/health', methods=['GET', 'OPTIONS'])
@local_executor_required(
    "appeal.execute",
    "tasks.view",
    "tasks.execute",
)
def api_local_executor_health():
    return jsonify({
        "status": "success",
        "data": {
            "ready": True,
            "runtime_role": RUNTIME_SETTINGS.role,
            "execution_target": "local",
            "label": "本机比特浏览器",
        },
    })


@app.route('/api/local-executor/run_shensu', methods=['GET', 'OPTIONS'])
@local_executor_required("appeal.execute")
def api_local_executor_run_shensu():
    blocked = _local_executor_database_preflight()
    if blocked is not None:
        return blocked
    return api_run_shensu.__wrapped__()


@app.route('/api/local-executor/run_shensu/stop', methods=['POST', 'OPTIONS'])
@local_executor_required("appeal.execute")
def api_local_executor_stop_shensu():
    return api_stop_shensu.__wrapped__()


@app.route('/api/local-executor/tasks/daily/start', methods=['POST', 'OPTIONS'])
@local_executor_required("tasks.execute")
def api_local_executor_start_daily_task():
    blocked = _local_executor_database_preflight()
    if blocked is not None:
        return blocked
    return api_start_daily_task.__wrapped__()


@app.route('/api/local-executor/tasks/daily/stop', methods=['POST', 'OPTIONS'])
@local_executor_required("tasks.execute")
def api_local_executor_stop_daily_task():
    return api_stop_daily_task.__wrapped__()


@app.route('/api/local-executor/tasks/daily/status', methods=['GET', 'OPTIONS'])
@local_executor_required("tasks.view", "tasks.execute")
def api_local_executor_daily_task_status():
    return api_daily_task_status.__wrapped__()


@app.route('/api/risk-check/categories', methods=['GET'])
@login_required
def api_risk_check_categories():
    try:
        return jsonify({
            "status": "success",
            "data": db_list_zying_risk_categories(),
        })
    except Exception as exc:
        logging.error("读取智赢侵权检测分类失败：%s", exc)
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.route('/api/risk-check/results', methods=['GET'])
@login_required
def api_risk_check_results():
    try:
        params = _risk_result_query_params(request.args)
        return jsonify({
            "status": "success",
            "data": db_get_zying_risk_results(**params),
        })
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.error("读取智赢侵权检测结果失败：%s", exc)
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.route('/api/risk-check/results/export', methods=['GET'])
@login_required
def api_export_risk_check_results():
    try:
        params = _risk_result_query_params(request.args, export=True)
        data = db_get_zying_risk_results(**params) or {}
        rows = data.get("rows") or []

        wb = Workbook()
        ws = wb.active
        ws.title = "侵权检测结果"
        columns = [
            ("数据行", "row_id"),
            ("产品编号", "product_id"),
            ("标题", "title"),
            ("智赢分类编号", "zying_category_id"),
            ("智赢产品分类", "zying_category"),
            ("产品分类", "product_category"),
            ("风险级别", "risk_level"),
            ("侵权关键词/品牌", "keywords"),
            ("主图链接", "main_image_url"),
            ("采集时间", "collected_at"),
            ("提交时间", "submitted_at"),
        ]
        ws.append([label for label, _ in columns])
        risk_labels = {
            "0": "0 - 无可疑",
            "1": "1 - 疑似/需复核",
            "2": "2 - 侵权",
        }
        for row in rows:
            values = []
            for _, key in columns:
                value = row.get(key, "")
                if key == "risk_level":
                    value = risk_labels.get(str(value or ""), "未检测")
                values.append(value)
            ws.append(values)

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        header_font = Font(bold=True, color="1F2937")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 48)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"侵权检测结果_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        response = send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )
        response.headers["Content-Disposition"] = (
            f"attachment; filename*=UTF-8''{quote(filename)}"
        )
        return response
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.error("导出智赢侵权检测结果失败：%s", exc)
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.route('/api/zying-collection/categories', methods=['GET'])
@login_required
def api_zying_collection_categories():
    try:
        return jsonify({
            "status": "success",
            "data": db_list_zying_risk_categories(),
        })
    except Exception as exc:
        logging.error("读取智赢产品采集分类失败：%s", exc)
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.route('/api/zying-collection/auth/open', methods=['POST'])
@login_required
def api_open_zying_collection_login():
    try:
        params = build_zying_login_params(request.get_json(silent=True) or {})
        result = bit_zying_caiji.open_zying_login_window(**params)
        return jsonify({
            "status": "success",
            "message": result.get("message") or "智赢登录窗口已打开",
            "data": {
                **result,
                "auth": bit_zying_caiji.get_zying_auth_status(),
            },
        })
    except Exception as exc:
        logging.error("打开智赢登录窗口失败：%s", exc)
        return jsonify({"status": "error", "message": str(exc)}), 400


@app.route('/api/zying-collection/auth/capture', methods=['POST'])
@login_required
def api_capture_zying_collection_login():
    try:
        params = build_zying_login_params(request.get_json(silent=True) or {})
        auth = bit_zying_caiji.capture_zying_login_from_browser(**params)
        with _zying_collection_state_lock:
            _zying_collection_state["requires_login"] = False
            if not _zying_collection_state.get("running"):
                _zying_collection_state["message"] = "智赢登录状态已保存，可以启动后台采集"
        return jsonify({
            "status": "success",
            "message": "智赢登录状态有效，凭证已保存",
            "data": {"auth": auth},
        })
    except bit_zying_caiji.ZyingAuthenticationError as exc:
        return jsonify({
            "status": "error",
            "message": str(exc),
            "requires_login": True,
        }), 409
    except Exception as exc:
        logging.error("保存智赢登录状态失败：%s", exc)
        return jsonify({"status": "error", "message": str(exc)}), 400


@app.route('/api/zying-collection/start', methods=['POST'])
@login_required
def api_start_zying_collection():
    try:
        params = build_zying_collection_params(request.get_json(silent=True) or {})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    if not _zying_collection_lock.acquire(blocking=False):
        with _zying_collection_state_lock:
            data = {
                **dict(_zying_collection_state),
                "logs": list(_zying_collection_logs),
            }
        return jsonify({
            "status": "error",
            "message": "智赢产品采集任务正在运行",
            "data": data,
        }), 409

    with _zying_collection_state_lock:
        _zying_collection_stop_event.clear()
        _zying_collection_logs.clear()
        _zying_collection_state.update(
            {
                "running": True,
                "started_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "finished_at": "",
                "status": "running",
                "message": "正在启动智赢产品采集",
                "params": dict(params),
                "summary": {},
                "requires_login": False,
            }
        )
        _append_zying_collection_log(
            f"智赢采集任务已启动：第 {params['start_page']}-{params['number']} 页，"
            f"分类 {params.get('category_name') or params.get('category') or '全部'}，"
            "模式 后台 API；"
            "数据库已有产品将直接跳过"
        )
        data = {
            **dict(_zying_collection_state),
            "logs": list(_zying_collection_logs),
        }
    try:
        threading.Thread(
            target=run_zying_collection_job,
            args=(params, _zying_collection_lock),
            daemon=True,
            name="zying-product-collection",
        ).start()
    except Exception:
        _zying_collection_lock.release()
        with _zying_collection_state_lock:
            _zying_collection_state.update(
                {
                    "running": False,
                    "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "status": "error",
                    "message": "智赢产品采集后台线程启动失败",
                }
            )
        raise
    return jsonify({"status": "success", "data": data})


@app.route('/api/zying-collection/status', methods=['GET'])
@login_required
def api_zying_collection_status():
    with _zying_collection_state_lock:
        data = {
            **dict(_zying_collection_state),
            "params": dict(_zying_collection_state.get("params") or {}),
            "summary": dict(_zying_collection_state.get("summary") or {}),
            "logs": list(_zying_collection_logs),
            "auth": bit_zying_caiji.get_zying_auth_status(),
            "defaults": {
                "window_id": bit_zying_caiji.DEFAULT_ZYING_WINDOW_ID,
                "browser_type": bit_zying_caiji.normalize_zying_browser_type(
                    bit_zying_caiji.DEFAULT_ZYING_BROWSER_TYPE
                ),
                "window_name": "",
                "start_page": bit_zying_caiji.DEFAULT_ZYING_START_PAGE,
                "end_page": bit_zying_caiji.DEFAULT_ZYING_PAGE_COUNT,
            },
        }
    return jsonify({"status": "success", "data": data})


@app.route('/api/zying-collection/stop', methods=['POST'])
@login_required
def api_stop_zying_collection():
    with _zying_collection_state_lock:
        if not _zying_collection_state.get("running"):
            return jsonify({
                "status": "error",
                "message": "当前没有正在运行的智赢产品采集任务",
                "data": {
                    **dict(_zying_collection_state),
                    "logs": list(_zying_collection_logs),
                },
            }), 409
        _zying_collection_stop_event.set()
        _zying_collection_state.update(
            {
                "status": "stopping",
                "message": "正在安全结束智赢产品采集，请等待当前接口或入库节点完成",
            }
        )
        _append_zying_collection_log("已收到结束指令，正在安全停止采集")
        data = {
            **dict(_zying_collection_state),
            "logs": list(_zying_collection_logs),
        }
    return jsonify({
        "status": "success",
        "message": "已发送结束指令",
        "data": data,
    })


@app.route('/api/risk-check/start', methods=['POST'])
@login_required
def api_start_risk_check():
    params = build_risk_check_params(request.get_json(silent=True) or {})
    if not _risk_check_lock.acquire(blocking=False):
        with _risk_check_state_lock:
            data = dict(_risk_check_state)
        return jsonify({
            "status": "error",
            "message": "侵权检测任务正在运行",
            "data": data,
        }), 409

    with _risk_check_state_lock:
        _risk_check_logs.clear()
        _risk_check_state.update(
            {
                "running": True,
                "started_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "finished_at": "",
                "status": "running",
                "message": "正在读取商品并进行 AI 侵权检测",
                "params": dict(params),
                "summary": {},
            }
        )
        _append_risk_check_log("侵权检测任务已启动：当前仅识别商品标题，主图 Logo 暂不检测")
        data = {**dict(_risk_check_state), "logs": list(_risk_check_logs)}
    try:
        threading.Thread(
            target=run_risk_check_job,
            args=(params, _risk_check_lock),
            daemon=True,
            name="zying-risk-check",
        ).start()
    except Exception:
        _risk_check_lock.release()
        with _risk_check_state_lock:
            _risk_check_state.update(
                {
                    "running": False,
                    "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "status": "error",
                    "message": "侵权检测后台线程启动失败",
                }
            )
        raise
    return jsonify({"status": "success", "data": data})


@app.route('/api/risk-check/status', methods=['GET'])
@login_required
def api_risk_check_status():
    with _risk_check_state_lock:
        data = {
            **dict(_risk_check_state),
            "params": dict(_risk_check_state.get("params") or {}),
            "summary": dict(_risk_check_state.get("summary") or {}),
            "logs": list(_risk_check_logs),
        }
    return jsonify({"status": "success", "data": data})


def _mercado_collection_state_update(**changes):
    with _mercado_collection_lock:
        _mercado_collection_state.update(changes)


def _mercado_collection_finish_status(processed, completed, failed, requested=None):
    processed = max(0, int(processed or 0))
    completed = max(0, int(completed or 0))
    failed = max(0, int(failed or 0))
    requested = max(0, int(requested or 0))
    shortfall = max(0, requested - processed) if requested else 0
    if not processed or (failed >= processed and completed == 0):
        status = "error"
        prefix = "采集失败"
    elif failed or shortfall:
        status = "partial"
        prefix = "采集部分完成"
    else:
        status = "completed"
        prefix = "采集完成"
    message = (
        f"{prefix}：入库 {processed} 件，"
        f"重量尺寸完整 {completed} 件，待补充 {failed} 件"
    )
    if shortfall:
        message += f"，距离目标还差 {shortfall} 件"
    return status, message


def _mercado_collection_elapsed_seconds(state, task=None, now=None):
    """Return a stable live/final task duration for the status API."""
    state = state or {}
    task = task or {}
    started_value = state.get("started_at") or task.get("started_at")
    if not started_value:
        return max(0, int(state.get("elapsed_seconds") or task.get("elapsed_seconds") or 0))
    finished_value = state.get("finished_at") or task.get("finished_at")

    def parse(value):
        if isinstance(value, datetime):
            return value
        text = str(value or "").strip()
        if not text:
            return None
        try:
            return datetime.fromisoformat(text)
        except ValueError:
            return None

    started_at = parse(started_value)
    finished_at = parse(finished_value) or now or datetime.now()
    if started_at is None:
        return max(0, int(state.get("elapsed_seconds") or task.get("elapsed_seconds") or 0))
    return max(0, int((finished_at - started_at).total_seconds()))


def _format_mercado_elapsed(seconds):
    seconds = max(0, int(seconds or 0))
    hours, remainder = divmod(seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    if hours:
        return f"{hours}小时{minutes}分{seconds}秒"
    if minutes:
        return f"{minutes}分{seconds}秒"
    return f"{seconds}秒"


def _mercado_shipping_rate_refresh_snapshot():
    with _mercado_shipping_rate_refresh_lock:
        return {
            **_mercado_shipping_rate_refresh_state,
            "sites": [dict(row) for row in _mercado_shipping_rate_refresh_state.get("sites", [])],
            "errors": [dict(row) for row in _mercado_shipping_rate_refresh_state.get("errors", [])],
            "unavailable_sites": [
                dict(row)
                for row in _mercado_shipping_rate_refresh_state.get("unavailable_sites", [])
            ],
        }


def _run_mercado_shipping_rate_refresh():
    started_monotonic = time.monotonic()
    started_at = datetime.now().replace(microsecond=0).strftime("%Y-%m-%d %H:%M:%S")
    with _mercado_shipping_rate_refresh_lock:
        _mercado_shipping_rate_refresh_state.update({
            "running": True,
            "status": "running",
            "message": "正在读取 Global Selling 官方最新运费公告与汇率",
            "started_at": started_at,
            "finished_at": "",
            "elapsed_seconds": 0,
            "success_sites": 0,
            "failed_sites": 0,
            "unavailable_site_count": 0,
            "recalculation_count": 0,
            "sites": [],
            "errors": [],
            "unavailable_sites": [],
        })
    try:
        from erp.mercadolibre_collection_store import mark_all_profitability_stale
        from erp.mercadolibre_profitability import (
            MercadoProfitabilityClient,
            active_store_token,
        )
        from erp.mercadolibre_shipping_rate_cards import (
            refresh_official_shipping_rate_cards,
        )

        # The announcements are public. Mercado Libre currently requires an
        # authorized app token for its official currency-conversion endpoint.
        result = refresh_official_shipping_rate_cards(
            MercadoProfitabilityClient(active_store_token())
        )
        recalculation_count = mark_all_profitability_stale()
        success_sites = int(result.get("success_sites") or 0)
        failed_sites = int(result.get("failed_sites") or 0)
        unavailable_site_count = int(result.get("unavailable_site_count") or 0)
        status = "completed" if success_sites and not failed_sites else (
            "partial" if success_sites else "failed"
        )
        message = (
            f"官方标准更新完成：已更新 {success_sites} 个站点，"
            f"官方未公布 {unavailable_site_count} 个，失败 {failed_sites} 个；"
            f"已安排 {recalculation_count} 条商品成本重算"
        )
        with _mercado_shipping_rate_refresh_lock:
            _mercado_shipping_rate_refresh_state.update({
                "running": False,
                "status": status,
                "message": message,
                "finished_at": datetime.now().replace(microsecond=0).strftime(
                    "%Y-%m-%d %H:%M:%S"
                ),
                "elapsed_seconds": round(time.monotonic() - started_monotonic, 1),
                "success_sites": success_sites,
                "failed_sites": failed_sites,
                "unavailable_site_count": unavailable_site_count,
                "recalculation_count": recalculation_count,
                "sites": list(result.get("sites") or []),
                "errors": list(result.get("errors") or []),
                "unavailable_sites": list(result.get("unavailable_sites") or []),
            })
    except Exception as exc:
        logging.exception("从官方刷新 Mercado 运费表失败")
        with _mercado_shipping_rate_refresh_lock:
            _mercado_shipping_rate_refresh_state.update({
                "running": False,
                "status": "failed",
                "message": f"Global Selling 官方标准更新失败：{exc}",
                "finished_at": datetime.now().replace(microsecond=0).strftime(
                    "%Y-%m-%d %H:%M:%S"
                ),
                "elapsed_seconds": round(time.monotonic() - started_monotonic, 1),
                "failed_sites": 5,
                "errors": [{"site_id": "", "country_name": "全部站点", "error": str(exc)}],
            })


def _start_mercado_shipping_rate_refresh(*, automatic=False):
    with _mercado_shipping_rate_refresh_lock:
        if _mercado_shipping_rate_refresh_state.get("running"):
            return False
        _mercado_shipping_rate_refresh_state.update({
            "running": True,
            "status": "queued",
            "message": "官方标准已加入后台更新队列" if not automatic else "每日官方标准更新已加入队列",
        })
    threading.Thread(
        target=_run_mercado_shipping_rate_refresh,
        name="mercado-official-shipping-rate-refresh",
        daemon=True,
    ).start()
    return True


def _mercado_profit_refresh_loop():
    """Refresh stale official fee snapshots without blocking the workbench UI."""

    from erp.ecb_exchange_rates import refresh_usd_cny_daily_rates
    from erp.mercadolibre_collection_store import (
        backfill_item_exchange_prices,
        list_stale_profitability_items,
        update_item_profitability,
    )
    from erp.mercadolibre_profitability import (
        MercadoProfitabilityClient,
        SUPPORTED_SITE_CURRENCIES,
        active_store_token,
        enrich_profitability,
        refresh_supported_exchange_rates,
    )

    stale_hours = 24
    try:
        batch_size = max(1, min(int(os.environ.get("MERCADO_PROFIT_REFRESH_BATCH", "50")), 500))
    except ValueError:
        batch_size = 50
    try:
        interval = max(60, int(os.environ.get("MERCADO_PROFIT_REFRESH_SECONDS", "300")))
    except ValueError:
        interval = 300

    next_reference_check = 0.0
    while not _mercado_profit_refresh_stop_event.is_set():
        processed_rows = False
        try:
            stale_before = (datetime.now() - timedelta(hours=stale_hours)).strftime(
                "%Y-%m-%d %H:%M:%S"
            )
            rows = list_stale_profitability_items(
                stale_before=stale_before,
                limit=batch_size,
            )
            client = None
            if time.monotonic() >= next_reference_check:
                client = MercadoProfitabilityClient(active_store_token())
                site_rates = refresh_supported_exchange_rates(client)
                backfill_item_exchange_prices({
                    SUPPORTED_SITE_CURRENCIES[site_id]: snapshot
                    for site_id, snapshot in site_rates.items()
                })
                refresh_usd_cny_daily_rates()
                from erp.mercadolibre_shipping_rate_cards import (
                    OfficialShippingRateCardStore,
                )
                if OfficialShippingRateCardStore().needs_refresh(max_age_hours=24):
                    _start_mercado_shipping_rate_refresh(automatic=True)
                # Refresh reference data and backfill converted prices once a day.
                next_reference_check = time.monotonic() + 24 * 60 * 60
            if rows:
                processed_rows = True
                from concurrent.futures import ThreadPoolExecutor, as_completed

                token = active_store_token()
                worker_state = threading.local()

                def refresh_row(row):
                    row_client = getattr(worker_state, "client", None)
                    if row_client is None:
                        row_client = MercadoProfitabilityClient(token)
                        worker_state.client = row_client
                    enriched = enrich_profitability(row, client=row_client)
                    update_item_profitability(
                        str(row.get("source_item_id") or ""), enriched
                    )
                    return str(row.get("source_item_id") or "")

                worker_count = min(10, len(rows))
                with ThreadPoolExecutor(
                    max_workers=worker_count,
                    thread_name_prefix="mercado-profit",
                ) as executor:
                    futures = [executor.submit(refresh_row, row) for row in rows]
                    for future in as_completed(futures):
                        if _mercado_profit_refresh_stop_event.is_set():
                            break
                        try:
                            future.result()
                        except Exception:
                            logging.exception("Mercado 商品成本重算失败")
        except Exception:
            logging.exception("自动更新 Mercado 官网佣金、汇率和运费失败")
        # Drain a recalculation backlog continuously in 50-row/10-thread
        # batches. The long interval is only for the idle daily poll.
        _mercado_profit_refresh_stop_event.wait(0.5 if processed_rows else interval)


def ensure_mercado_profit_refresh_worker():
    global _mercado_profit_refresh_started
    if (
        USE_DB_API
        or app.testing
        or _truthy_env(os.environ.get("MERCADO_PROFIT_REFRESH_DISABLED"))
    ):
        return
    with _mercado_profit_refresh_lock:
        if _mercado_profit_refresh_started:
            return
        thread = threading.Thread(
            target=_mercado_profit_refresh_loop,
            name="mercado-profit-refresh",
            daemon=True,
        )
        thread.start()
        _mercado_profit_refresh_started = True


@app.before_request
def _start_mercado_profit_refresh_worker():
    ensure_mercado_profit_refresh_worker()


@app.before_request
def _start_order_sync_scheduler():
    if not USE_DB_API and not app.testing:
        bit_order_sync.ensure_order_sync_scheduler()
        bit_order_sync.ensure_order_financial_backfill_worker()
        bit_order_sync.ensure_order_image_backfill_worker()


def _mercado_collection_rows_needing_repair(rows):
    """Retry only rows whose current weight/dimensions are actually incomplete."""
    from erp.mercadolibre_collection_store import has_complete_weight_dimensions

    return [row for row in rows or [] if not has_complete_weight_dimensions(row)]


def _mercado_collection_db_call(operation, *args, attempts=6, **kwargs):
    """Retry transient database/network failures without killing browser workers."""
    attempts = max(1, min(int(attempts), 10))
    for attempt in range(attempts):
        try:
            return operation(*args, **kwargs)
        except Exception:
            if attempt + 1 >= attempts:
                raise
            delay = min(0.5 * (2 ** attempt), 8.0)
            logging.warning(
                "Mercado 采集数据库操作失败，%.1f 秒后重试（%s/%s）",
                delay,
                attempt + 1,
                attempts,
                exc_info=True,
            )
            time.sleep(delay)


def _run_mercado_collection_task(
    task_id, source_url, requested_count, worker_count, collection_scope
):
    from erp.mercadolibre_batch_collector import (
        CollectionStopped,
        DEFAULT_ZYING_WINDOW_ID,
        collect_marketplace_listing,
    )
    counters = {"candidate": 0, "processed": 0, "completed": 0, "failed": 0}
    counters_lock = threading.Lock()
    item_statuses = {}
    pending_rows = []
    pending_rows_lock = threading.Lock()
    collection_write_batch_size = 25
    started_monotonic = time.monotonic()
    last_task_status_write = 0.0

    def elapsed_seconds():
        return max(0, int(time.monotonic() - started_monotonic))

    def result_summary(message):
        return (
            f"{message} · 成功 {counters['completed']} 件 · "
            f"失败 {counters['failed']} 件 · 并发 {worker_count} · "
            f"耗时 {_format_mercado_elapsed(elapsed_seconds())}"
        )

    def on_page(info):
        counters["candidate"] = int(info.get("candidate_count") or 0)
        current_page = int(info.get("page") or 0)
        stage = str(info.get("stage") or "")
        message = str(info.get("message") or "") or (
            f"已扫描第 {current_page} 页，找到 {counters['candidate']} 个不重复商品"
        )
        state_status = "waiting_verification" if stage == "waiting_verification" else "running"
        _mercado_collection_state_update(
            status=state_status,
            candidate_count=counters["candidate"],
            current_page=current_page,
            message=message,
        )
        _mercado_collection_db_call(
            db_update_mercado_collection_task,
            task_id,
            status="running",
            collected_count=counters["candidate"],
            current_page=current_page,
            message=message,
        )

    def on_progress(info):
        _mercado_collection_state_update(
            current_item_id=str(info.get("item_id") or ""),
            message=str(info.get("message") or "正在采集商品详情"),
        )

    def buffer_collection_row(row, *, force=False):
        batch = []
        with pending_rows_lock:
            if row is not None:
                pending_rows.append(dict(row))
            if force or len(pending_rows) >= collection_write_batch_size:
                batch.extend(pending_rows)
                pending_rows.clear()
        if batch:
            _mercado_collection_db_call(
                db_upsert_mercado_collection_items, task_id, batch
            )

    def on_item(row):
        nonlocal last_task_status_write
        # Persist collection results immediately.  Official commission,
        # exchange-rate and shipping estimates are filled by the existing
        # background profitability worker and must not block browser slots.
        row = {
            **row,
            "profitability_updated_at": None,
            "profitability_source": "mercadolibre_official_api_pending",
            "profitability_error": "",
        }
        buffer_collection_row(row)
        with counters_lock:
            item_id = str(row.get("source_item_id") or "")
            incoming_status = "ok" if row.get("scrape_status") == "ok" else "partial"
            previous_status = item_statuses.get(item_id)
            effective_status = (
                "ok" if previous_status == "ok" or incoming_status == "ok" else "partial"
            )
            if previous_status is None:
                counters["processed"] += 1
                counters["completed" if effective_status == "ok" else "failed"] += 1
            elif previous_status != effective_status:
                counters["completed" if previous_status == "ok" else "failed"] -= 1
                counters["completed" if effective_status == "ok" else "failed"] += 1
            item_statuses[item_id] = effective_status
            message = (
                f"已完成 {counters['processed']}/{counters['candidate']}，"
                f"完整 {counters['completed']}，待补充 {counters['failed']}；"
                "佣金和运费在后台计算"
            )
            _mercado_collection_state_update(
                processed_count=counters["processed"],
                completed_count=counters["completed"],
                failed_count=counters["failed"],
                message=message,
            )
            now_monotonic = time.monotonic()
            persist_task_status = (
                now_monotonic - last_task_status_write >= 1.0
                or counters["processed"] >= max(counters["candidate"], requested_count)
            )
            if persist_task_status:
                last_task_status_write = now_monotonic
                task_snapshot = {
                    "collected_count": max(
                        counters["candidate"], counters["processed"]
                    ),
                    "completed_count": counters["completed"],
                    "failed_count": counters["failed"],
                    "message": message,
                }
            else:
                task_snapshot = None
        if task_snapshot is not None:
            _mercado_collection_db_call(
                db_update_mercado_collection_task,
                task_id,
                **task_snapshot,
            )

    try:
        _mercado_collection_db_call(
            db_update_mercado_collection_task,
            task_id,
            status="running",
            message="正在打开采集浏览器",
            worker_count=worker_count,
            started=True,
        )
        result = collect_marketplace_listing(
            source_url,
            requested_count,
            max_workers=worker_count,
            collection_scope=collection_scope,
            on_page=on_page,
            on_item=on_item,
            on_progress=on_progress,
            stop_event=_mercado_collection_stop_event,
        )
        buffer_collection_row(None, force=True)
        incomplete_rows = _mercado_collection_rows_needing_repair(result.get("rows"))
        if incomplete_rows and not _mercado_collection_stop_event.is_set():
            from erp.mercadolibre_playwright_collector import (
                repair_marketplace_items_playwright,
            )

            _mercado_collection_state_update(
                message=f"开始低并发补采 {len(incomplete_rows)} 件缺少重量尺寸的商品"
            )
            repair_marketplace_items_playwright(
                incomplete_rows,
                window_id=DEFAULT_ZYING_WINDOW_ID,
                plugin_timeout=15.0,
                attempts=1,
                on_item=on_item,
                on_progress=on_progress,
                stop_event=_mercado_collection_stop_event,
            )
            buffer_collection_row(None, force=True)
        status, message = _mercado_collection_finish_status(
            counters["processed"],
            counters["completed"],
            counters["failed"],
            requested_count,
        )
        message = result_summary(message)
        _mercado_collection_db_call(
            db_update_mercado_collection_task,
            task_id,
            status=status,
            message=message,
            collected_count=int(result.get("candidate_count") or counters["candidate"]),
            completed_count=counters["completed"],
            failed_count=counters["failed"],
            worker_count=worker_count,
            elapsed_seconds=elapsed_seconds(),
            finished=True,
        )
        _mercado_collection_state_update(
            status=status,
            message=message,
            elapsed_seconds=elapsed_seconds(),
        )
    except CollectionStopped:
        message = result_summary(
            f"采集已停止，已保留入库的 {counters['processed']} 件商品"
        )
        try:
            _mercado_collection_db_call(
                db_update_mercado_collection_task,
                task_id,
                status="stopped",
                message=message,
                completed_count=counters["completed"],
                failed_count=counters["failed"],
                worker_count=worker_count,
                elapsed_seconds=elapsed_seconds(),
                finished=True,
            )
        except Exception:
            logging.exception("更新 Mercado 采集停止状态失败")
        _mercado_collection_state_update(
            status="stopped", message=message, elapsed_seconds=elapsed_seconds()
        )
    except Exception as exc:
        message = result_summary(f"采集失败：{exc}")
        logging.exception("Mercado 列表采集失败")
        try:
            _mercado_collection_db_call(
                db_update_mercado_collection_task,
                task_id,
                status="error",
                message=message,
                completed_count=counters["completed"],
                failed_count=counters["failed"],
                worker_count=worker_count,
                elapsed_seconds=elapsed_seconds(),
                finished=True,
            )
        except Exception:
            logging.exception("更新 Mercado 采集失败状态失败")
        _mercado_collection_state_update(
            status="error", message=message, elapsed_seconds=elapsed_seconds()
        )
    finally:
        try:
            buffer_collection_row(None, force=True)
        except Exception:
            logging.exception("刷新 Mercado 采集批量写入缓冲区失败")
        _mercado_collection_state_update(
            running=False,
            current_item_id="",
            finished_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        )


@app.route('/api/mercado-collection/start', methods=['POST'])
@login_required
def api_start_mercado_collection():
    from erp.mercadolibre_batch_collector import (
        DEFAULT_COLLECTION_WORKERS,
        build_marketplace_search_url,
        normalize_collection_scope,
        normalize_collection_workers,
        validate_collection_request,
    )
    from erp.mercadolibre_translation import (
        marketplace_site_name,
        normalize_marketplace_site,
    )

    data = request.get_json(silent=True) or {}
    try:
        raw_source_url = str(data.get("source_url") or "").strip()
        keyword = str(data.get("keyword") or "").strip()
        source_site_id = normalize_marketplace_site(data.get("site_id") or "MLM")
        source_site_name = marketplace_site_name(source_site_id)
        collection_scope = normalize_collection_scope(
            data.get("collection_scope") or "all"
        )
        if not raw_source_url:
            raw_source_url = build_marketplace_search_url(
                keyword, source_site_id, collection_scope
            )
        source_url, requested_count = validate_collection_request(
            raw_source_url, data.get("requested_count", 20)
        )
        worker_count = normalize_collection_workers(
            data.get("worker_count", DEFAULT_COLLECTION_WORKERS)
        )
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400

    with _mercado_collection_lock:
        if _mercado_playwright_setup_state.get("running"):
            return jsonify({
                "status": "error",
                "message": "Playwright 登录窗口仍开着，请登录智赢后关闭该窗口，再开始采集",
            }), 409
        if _mercado_collection_state.get("running"):
            return jsonify({
                "status": "error",
                "message": "已有采集任务正在运行，请等待当前任务完成",
            }), 409
        user = session.get("workbench_user") or {}
        try:
            task_id = db_create_mercado_collection_task(
                source_url,
                requested_count,
                str(user.get("display_name") or user.get("username") or ""),
                worker_count=worker_count,
            )
        except Exception as exc:
            logging.exception("创建 Mercado 采集任务失败")
            return jsonify({"status": "error", "message": f"创建采集任务失败：{exc}"}), 500
        _mercado_collection_stop_event.clear()
        _mercado_collection_state.update(
            {
                "running": True,
                "task_id": int(task_id),
                "status": "starting",
                "message": (
                    f"正在启动{source_site_name}"
                    f"{'跨境卖家专区' if collection_scope == 'cross_border' else '全部商品'}"
                    f"采集（并发数 {worker_count}）"
                ),
                "source_url": source_url,
                "keyword": keyword,
                "source_site_id": source_site_id,
                "source_site_name": source_site_name,
                "collection_scope": collection_scope,
                "requested_count": requested_count,
                "worker_count": worker_count,
                "candidate_count": 0,
                "processed_count": 0,
                "completed_count": 0,
                "failed_count": 0,
                "elapsed_seconds": 0,
                "current_page": 0,
                "current_item_id": "",
                "started_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "finished_at": "",
            }
        )
        worker = threading.Thread(
            target=_run_mercado_collection_task,
            args=(
                int(task_id),
                source_url,
                requested_count,
                worker_count,
                collection_scope,
            ),
            name=f"mercado-collection-{task_id}",
            daemon=True,
        )
        try:
            worker.start()
        except Exception:
            _mercado_collection_state["running"] = False
            raise
    return jsonify({
        "status": "success",
        "data": {"task_id": int(task_id), **dict(_mercado_collection_state)},
    })


def _run_mercado_playwright_setup():
    try:
        from erp.mercadolibre_playwright_collector import open_playwright_login_setup

        open_playwright_login_setup()
        with _mercado_collection_lock:
            _mercado_playwright_setup_state.update({
                "running": False,
                "status": "completed",
                "message": "采集浏览器已关闭，可以开始采集",
            })
    except Exception as exc:
        logging.exception("打开 Playwright 智赢登录窗口失败")
        with _mercado_collection_lock:
            _mercado_playwright_setup_state.update({
                "running": False,
                "status": "error",
                "message": f"打开采集浏览器失败：{exc}",
            })


@app.route('/api/mercado-collection/playwright-setup', methods=['POST'])
@login_required
def api_open_mercado_playwright_setup():
    with _mercado_collection_lock:
        if _mercado_collection_state.get("running"):
            return jsonify({
                "status": "error",
                "message": "采集任务运行中，不能同时打开登录窗口",
            }), 409
        if _mercado_playwright_setup_state.get("running"):
            return jsonify({
                "status": "success",
                "data": dict(_mercado_playwright_setup_state),
            })
        _mercado_playwright_setup_state.update({
            "running": True,
            "status": "starting",
            "message": "正在打开 Playwright 采集浏览器",
        })
        worker = threading.Thread(
            target=_run_mercado_playwright_setup,
            name="mercado-playwright-login-setup",
            daemon=True,
        )
        worker.start()
        return jsonify({
            "status": "success",
            "data": dict(_mercado_playwright_setup_state),
        })


@app.route('/api/mercado-collection/stop', methods=['POST'])
@login_required
def api_stop_mercado_collection():
    with _mercado_collection_lock:
        if not _mercado_collection_state.get("running"):
            return jsonify({"status": "success", "data": dict(_mercado_collection_state)})
        _mercado_collection_stop_event.set()
        _mercado_collection_state["message"] = "正在停止，当前商品处理后将结束"
        data = dict(_mercado_collection_state)
    return jsonify({"status": "success", "data": data})


@app.route('/api/mercado-collection/status', methods=['GET'])
@login_required
def api_mercado_collection_status():
    with _mercado_collection_lock:
        data = dict(_mercado_collection_state)
    task_id = data.get("task_id")
    if task_id:
        try:
            data["task"] = db_get_mercado_collection_task(task_id)
        except Exception as exc:
            data["database_message"] = str(exc)
    task = data.get("task") or {}
    if not data.get("worker_count") and task.get("worker_count"):
        data["worker_count"] = int(task["worker_count"])
    data["elapsed_seconds"] = _mercado_collection_elapsed_seconds(data, task)
    response = jsonify({"status": "success", "data": data})
    response.headers["Cache-Control"] = "no-store"
    return response


@app.route('/api/mercado-collection/items', methods=['GET', 'DELETE'])
@login_required
def api_mercado_collection_items():
    try:
        if request.method == "DELETE":
            data = request.get_json(silent=True) or {}
            item_ids = data.get("collection_item_ids") or []
            if not isinstance(item_ids, list):
                return jsonify({"status": "error", "message": "collection_item_ids 必须是数组"}), 422
            result = db_delete_mercado_collection_items(item_ids)
            return jsonify({"status": "success", "data": result})
        task_id = request.args.get("task_id")
        result = db_list_mercado_collection_items(
            search=str(request.args.get("search") or "").strip(),
            limit=_parse_int_param(request.args, "limit", 500, 1, 1000),
            offset=_parse_int_param(request.args, "offset", 0, 0, 1000000),
            task_id=int(task_id) if str(task_id or "").strip() else None,
            weight_min=str(request.args.get("weight_min") or "").strip(),
            weight_max=str(request.args.get("weight_max") or "").strip(),
            price_min=str(request.args.get("price_min") or "").strip(),
            price_max=str(request.args.get("price_max") or "").strip(),
            net_proceeds_min=str(request.args.get("net_proceeds_min") or "").strip(),
            net_proceeds_max=str(request.args.get("net_proceeds_max") or "").strip(),
            date_from=str(request.args.get("date_from") or "").strip(),
            date_to=str(request.args.get("date_to") or "").strip(),
            management_category_id=str(
                request.args.get("management_category_id") or ""
            ).strip(),
            exclude_added=True,
        )
        return jsonify({"status": "success", "data": result})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("读取 Mercado 采集列表失败")
        return jsonify({"status": "error", "message": f"读取采集列表失败：{exc}"}), 500


@app.route('/api/mercado-products', methods=['GET', 'DELETE'])
@login_required
def api_mercado_products():
    try:
        if request.method == "DELETE":
            with _mercado_publish_lock:
                if _mercado_publish_state.get("running"):
                    return jsonify({
                        "status": "error",
                        "message": "批量上架正在运行，完成后再删除产品",
                    }), 409
            data = request.get_json(silent=True) or {}
            item_ids = data.get("product_item_ids") or []
            if not isinstance(item_ids, list):
                return jsonify({"status": "error", "message": "product_item_ids 必须是数组"}), 422
            result = db_delete_mercado_product_items(item_ids)
            return jsonify({"status": "success", "data": result})
        result = db_list_mercado_product_items(
            search=str(request.args.get("search") or "").strip(),
            limit=_parse_int_param(request.args, "limit", 500, 1, 1000),
            offset=_parse_int_param(request.args, "offset", 0, 0, 1000000),
            source_type=str(request.args.get("source_type") or "").strip(),
            review_status=str(request.args.get("review_status") or "").strip(),
            publish_status=str(request.args.get("publish_status") or "").strip(),
            weight_min=str(request.args.get("weight_min") or "").strip(),
            weight_max=str(request.args.get("weight_max") or "").strip(),
            price_min=str(request.args.get("price_min") or "").strip(),
            price_max=str(request.args.get("price_max") or "").strip(),
            net_proceeds_min=str(request.args.get("net_proceeds_min") or "").strip(),
            net_proceeds_max=str(request.args.get("net_proceeds_max") or "").strip(),
            date_from=str(request.args.get("date_from") or "").strip(),
            date_to=str(request.args.get("date_to") or "").strip(),
            management_category_id=str(
                request.args.get("management_category_id") or ""
            ).strip(),
            mercado_category=str(
                request.args.get("mercado_category") or ""
            ).strip(),
        )
        return jsonify({"status": "success", "data": result})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("读取 Mercado 产品列表失败")
        return jsonify({"status": "error", "message": f"读取产品列表失败：{exc}"}), 500


@app.route('/api/mercado-management-categories', methods=['GET', 'POST'])
@login_required
def api_mercado_management_categories():
    try:
        if request.method == "GET":
            result = db_list_mercado_management_categories()
        else:
            data = request.get_json(silent=True) or {}
            result = db_create_mercado_management_category(data.get("name", ""))
        return jsonify({"status": "success", "data": result})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("管理 Mercado 运营分类失败")
        return jsonify({"status": "error", "message": f"分类管理失败：{exc}"}), 500


@app.route('/api/mercado-management-categories/<int:category_id>', methods=['PATCH', 'DELETE'])
@login_required
def api_mercado_management_category(category_id):
    try:
        if request.method == "DELETE":
            result = db_delete_mercado_management_category(category_id)
        else:
            data = request.get_json(silent=True) or {}
            result = db_update_mercado_management_category(
                category_id, data.get("name", "")
            )
        return jsonify({"status": "success", "data": result})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except KeyError as exc:
        return jsonify({"status": "error", "message": exc.args[0]}), 404
    except Exception as exc:
        logging.exception("修改 Mercado 运营分类失败")
        return jsonify({"status": "error", "message": f"分类管理失败：{exc}"}), 500


@app.route('/api/mercado-management-categories/assign', methods=['POST'])
@login_required
def api_assign_mercado_management_category():
    data = request.get_json(silent=True) or {}
    item_ids = data.get("item_ids") or []
    if not isinstance(item_ids, list):
        return jsonify({"status": "error", "message": "item_ids 必须是数组"}), 422
    try:
        result = db_assign_mercado_management_category(
            data.get("item_type", ""), item_ids, data.get("category_id")
        )
        return jsonify({"status": "success", "data": result})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except KeyError as exc:
        return jsonify({"status": "error", "message": exc.args[0]}), 404
    except Exception as exc:
        logging.exception("设置 Mercado 商品运营分类失败")
        return jsonify({"status": "error", "message": f"设置分类失败：{exc}"}), 500


@app.route('/api/mercado-products/add', methods=['POST'])
@login_required
def api_add_mercado_products():
    data = request.get_json(silent=True) or {}
    item_ids = data.get("collection_item_ids") or []
    if not isinstance(item_ids, list):
        return jsonify({"status": "error", "message": "collection_item_ids 必须是数组"}), 422
    try:
        result = db_add_mercado_collection_items_to_products(item_ids)
        return jsonify({"status": "success", "data": result})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("加入 Mercado 产品列表失败")
        return jsonify({"status": "error", "message": f"加入产品列表失败：{exc}"}), 500


@app.route('/api/mercado-products/<int:product_item_id>', methods=['PATCH'])
@login_required
def api_update_mercado_product(product_item_id):
    with _mercado_publish_lock:
        if _mercado_publish_state.get("running"):
            return jsonify({
                "status": "error",
                "message": "批量上架正在运行，完成后再修改产品",
            }), 409
    data = request.get_json(silent=True)
    if not isinstance(data, dict):
        return jsonify({"status": "error", "message": "产品内容必须是对象"}), 422
    allowed = {
        "title", "description_text", "main_image_url", "category_id", "price",
        "weight_g", "package_length_cm", "package_width_cm", "package_height_cm",
    }
    try:
        result = db_update_mercado_product_item(
            product_item_id,
            {key: value for key, value in data.items() if key in allowed},
        )
        return jsonify({"status": "success", "data": result})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except KeyError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 404
    except Exception as exc:
        logging.exception("修改 Mercado 产品内容失败")
        return jsonify({"status": "error", "message": f"修改产品失败：{exc}"}), 500


@app.route('/api/mercado-products/bulk-edit', methods=['PATCH'])
@login_required
def api_bulk_update_mercado_products():
    with _mercado_publish_lock:
        if _mercado_publish_state.get("running"):
            return jsonify({
                "status": "error",
                "message": "批量上架正在运行，完成后再修改产品",
            }), 409
    data = request.get_json(silent=True)
    if not isinstance(data, dict):
        return jsonify({"status": "error", "message": "批量修改内容必须是对象"}), 422
    item_ids = data.get("product_item_ids") or []
    changes = data.get("changes")
    if not isinstance(item_ids, list):
        return jsonify({"status": "error", "message": "product_item_ids 必须是数组"}), 422
    if not isinstance(changes, dict):
        return jsonify({"status": "error", "message": "changes 必须是对象"}), 422
    allowed = {
        "title", "description_text", "main_image_url", "category_id", "price",
        "weight_g", "package_length_cm", "package_width_cm", "package_height_cm",
    }
    try:
        result = db_update_mercado_product_items(
            item_ids,
            {key: value for key, value in changes.items() if key in allowed},
        )
        return jsonify({"status": "success", "data": result})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("批量修改 Mercado 产品内容失败")
        return jsonify({"status": "error", "message": f"批量修改产品失败：{exc}"}), 500


@app.route('/api/mercado-products/review-status', methods=['POST'])
@login_required
def api_update_mercado_product_review_status():
    data = request.get_json(silent=True) or {}
    item_ids = data.get("product_item_ids") or []
    if not isinstance(item_ids, list):
        return jsonify({"status": "error", "message": "product_item_ids 必须是数组"}), 422
    try:
        result = db_update_mercado_product_review_status(
            item_ids, str(data.get("review_status") or "").strip()
        )
        return jsonify({"status": "success", "data": result})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("更新 Mercado 产品审核状态失败")
        return jsonify({"status": "error", "message": f"更新审核状态失败：{exc}"}), 500


def _mercado_publish_state_update(**changes):
    with _mercado_publish_lock:
        _mercado_publish_state.update(changes)


def _run_mercado_product_publish(
    product_rows, token_id, site_id, site_name, quantity, worker_count, store_name,
    batch_id, created_by, discount_rate, moved_to_collection_count=0,
):
    """Backward-compatible single-target entry point."""
    return _run_mercado_product_publish_targets(
        product_rows,
        targets=[{
            "token_id": int(token_id),
            "store_name": str(store_name),
            "site_id": str(site_id),
            "site_name": str(site_name),
            "discount_rate": float(discount_rate),
        }],
        quantity=quantity,
        worker_count=worker_count,
        batch_id=batch_id,
        created_by=created_by,
        moved_to_collection_count=moved_to_collection_count,
    )


def _run_mercado_product_publish_targets(
    product_rows, targets, quantity, worker_count, batch_id, created_by,
    moved_to_collection_count=0,
):
    from erp.mercadolibre_batch_publish import publish_product_batch

    rows = [dict(row) for row in product_rows or []]
    target_rows = [dict(target) for target in targets or []]
    target_count = len(target_rows)
    total_requested = sum(
        len(target.get("product_rows") or rows) for target in target_rows
    )
    processed = published = failed = skipped_published = 0
    all_results = []
    stage_duration_totals: dict[str, float] = {}
    stage_sample_count = 0
    started_monotonic = time.monotonic()

    def timing_metrics(done_count):
        elapsed = max(0.0, time.monotonic() - started_monotonic)
        done = max(0, int(done_count or 0))
        attempted_done = max(0, done - skipped_published)
        rate_per_second = (
            attempted_done / elapsed
        ) if attempted_done and elapsed else 0.0
        remaining = max(0, total_requested - done)
        return {
            "elapsed_seconds": round(elapsed, 1),
            "average_seconds_per_item": round(
                elapsed / attempted_done, 2
            ) if attempted_done else 0,
            "items_per_minute": round(rate_per_second * 60, 2),
            "estimated_remaining_seconds": round(
                remaining / rate_per_second, 1
            ) if rate_per_second else 0,
        }

    for target_index, target in enumerate(target_rows, start=1):
        token_id = int(target.get("token_id") or 0)
        site_id = str(target.get("site_id") or "")
        site_name = str(target.get("site_name") or site_id)
        store_name = str(target.get("store_name") or token_id)
        discount_rate = float(target.get("discount_rate") or 0)
        current_rows = [dict(row) for row in (target.get("product_rows") or rows)]
        current_quantity = int(target.get("quantity") or quantity)
        try:
            published_product_ids = set(db_get_published_mercado_product_item_ids(
                [int(row.get("id") or 0) for row in current_rows],
                token_id=token_id,
                site_id=site_id,
            ))
        except Exception:
            logging.exception(
                "读取历史成功上架记录失败，将继续当前组合: token_id=%s site_id=%s",
                token_id,
                site_id,
            )
            published_product_ids = set()
        target_rows_to_publish = [
            row for row in current_rows
            if int(row.get("id") or 0) not in published_product_ids
        ]
        target_skipped = len(current_rows) - len(target_rows_to_publish)
        skipped_published += target_skipped
        processed += target_skipped
        base_processed = processed
        base_published = published
        base_failed = failed

        def on_progress(info, *, _target_index=target_index):
            current = int(info.get("current") or 0)
            current_published = int(info.get("published_count") or 0)
            current_failed = int(info.get("failed_count") or 0)
            aggregate_processed = base_processed + current
            _mercado_publish_state_update(
                processed_count=aggregate_processed,
                published_count=base_published + current_published,
                failed_count=base_failed + current_failed,
                skipped_published_count=skipped_published,
                completed_target_count=_target_index - 1,
                worker_count=int(info.get("worker_count") or worker_count),
                current_item_id=str(info.get("source_item_id") or ""),
                average_stage_seconds=dict(
                    info.get("average_stage_seconds") or {}
                ),
                message=(
                    f"组合 {_target_index}/{target_count} · {store_name} · {site_name}"
                ),
                **timing_metrics(aggregate_processed),
            )

        try:
            target_batch_id = f"{batch_id}-{target_index}"
            if target_rows_to_publish:
                result = publish_product_batch(
                    target_rows_to_publish,
                    token_id=token_id,
                    site_id=site_id,
                    quantity=current_quantity,
                    workers=int(worker_count),
                    discount_rate=discount_rate,
                    update_state=db_update_mercado_product_publish_state,
                    on_progress=on_progress,
                    batch_id=target_batch_id,
                    created_by=str(created_by),
                    create_records=db_create_mercado_product_publish_records,
                    update_record=db_update_mercado_product_publish_record,
                )
                target_requested = int(
                    result.get("requested_count") or len(target_rows_to_publish)
                )
                target_published = int(result.get("published_count") or 0)
                target_failed = int(result.get("failed_count") or 0)
                target_results = list(result.get("results") or [])
                target_stage_averages = dict(
                    result.get("average_stage_seconds") or {}
                )
            else:
                target_requested = target_published = target_failed = 0
                target_results = []
                target_stage_averages = {}
        except Exception as exc:
            logging.exception(
                "Mercado 产品上架组合失败: token_id=%s site_id=%s",
                token_id,
                site_id,
            )
            target_requested = len(target_rows_to_publish)
            target_published = 0
            target_failed = len(target_rows_to_publish)
            target_results = [{
                "product_id": int(row.get("id") or 0),
                "source_item_id": str(row.get("source_item_id") or ""),
                "status": "failed",
                "published_item_id": "",
                "message": f"上架组合启动失败：{exc}"[:2000],
            } for row in target_rows_to_publish]
            target_stage_averages = {}

        processed += target_requested
        published += target_published
        failed += target_failed
        if target_results:
            sample_size = len(target_results)
            stage_sample_count += sample_size
            for stage, average in target_stage_averages.items():
                try:
                    stage_duration_totals[str(stage)] = (
                        stage_duration_totals.get(str(stage), 0.0)
                        + float(average or 0) * sample_size
                    )
                except (TypeError, ValueError):
                    continue
        aggregate_stage_averages = {
            stage: round(total / max(1, stage_sample_count), 4)
            for stage, total in stage_duration_totals.items()
        }
        for item in target_results:
            all_results.append({
                **dict(item),
                "target_index": target_index,
                "token_id": token_id,
                "store_name": store_name,
                "site_id": site_id,
                "site_name": site_name,
                "discount_rate": discount_rate,
            })
        _mercado_publish_state_update(
            processed_count=processed,
            published_count=published,
            failed_count=failed,
            skipped_published_count=skipped_published,
            completed_target_count=target_index,
            current_item_id="",
            average_stage_seconds=aggregate_stage_averages,
            results=list(all_results),
            message=f"已完成组合 {target_index}/{target_count} · {store_name} · {site_name}",
            **timing_metrics(processed),
        )

    if failed == 0:
        status = "completed"
    elif published:
        status = "partial"
    else:
        status = "error"
    message = f"批量上架完成：{target_count} 个账号-站点组合"
    if moved_to_collection_count:
        message += f"；已忽略并移回采集列表 {moved_to_collection_count} 件不可上架商品"
    _mercado_publish_state_update(
        running=False,
        status=status,
        message=message,
        requested_count=total_requested,
        processed_count=processed,
        published_count=published,
        failed_count=failed,
        moved_to_collection_count=int(moved_to_collection_count or 0),
        skipped_published_count=skipped_published,
        current_item_id="",
        completed_target_count=target_count,
        finished_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        results=list(all_results),
        average_stage_seconds={
            stage: round(total / max(1, stage_sample_count), 4)
            for stage, total in stage_duration_totals.items()
        },
        batch_id=str(batch_id),
        **timing_metrics(processed),
    )


@app.route('/api/mercado-products/publish', methods=['POST'])
@login_required
def api_publish_mercado_products():
    data = request.get_json(silent=True) or {}
    item_ids = data.get("product_item_ids") or []
    if not isinstance(item_ids, list):
        return jsonify({"status": "error", "message": "product_item_ids 必须是数组"}), 422
    with _mercado_publish_lock:
        if _mercado_publish_state.get("running"):
            return jsonify({
                "status": "error",
                "message": "已有批量上架任务正在运行",
            }), 409
    try:
        from erp.mercadolibre_translation import (
            marketplace_site_name,
            normalize_marketplace_site,
        )
        from erp.mercadolibre_batch_publish import (
            product_publish_issues,
            site_discount_rate,
            validate_publishable_products,
        )

        selection_mode = str(data.get("selection_mode") or "accounts").strip().lower()
        if selection_mode not in {"accounts", "groups"}:
            raise ValueError("上架选择方式必须是账号或分组")

        raw_token_ids = data.get("token_ids")
        if raw_token_ids is None:
            raw_token_ids = [data.get("token_id")]
        if not isinstance(raw_token_ids, list):
            raise ValueError("token_ids 必须是数组")
        token_ids = []
        for value in raw_token_ids:
            token_id_value = int(value or 0)
            if token_id_value > 0 and token_id_value not in token_ids:
                token_ids.append(token_id_value)

        raw_group_names = data.get("group_names") or []
        if not isinstance(raw_group_names, list):
            raise ValueError("group_names 必须是数组")
        group_names = []
        for value in raw_group_names:
            group_name = str(value or "").strip()
            if group_name and group_name not in group_names:
                group_names.append(group_name)

        raw_site_ids = data.get("site_ids")
        if raw_site_ids is None:
            raw_site_ids = [data.get("site_id") or "MLM"]
        if not isinstance(raw_site_ids, list):
            raise ValueError("site_ids 必须是数组")
        site_ids = []
        for value in raw_site_ids:
            site_id_value = normalize_marketplace_site(value)
            if site_id_value not in site_ids:
                site_ids.append(site_id_value)
        if not site_ids:
            raise ValueError("请选择至少一个目标站点")
        if selection_mode == "accounts" and not token_ids:
            raise ValueError("请选择至少一个要上架的授权账号")
        if selection_mode == "groups" and not group_names:
            raise ValueError("请选择至少一个账号分组")

        token_id = token_ids[0] if token_ids else None
        site_id = site_ids[0]
        site_name = marketplace_site_name(site_id)
        raw_quantity = data.get("quantity")
        raw_worker_count = data.get("worker_count")
        quantity = int(500 if raw_quantity in (None, "") else raw_quantity)
        worker_count = int(10 if raw_worker_count in (None, "") else raw_worker_count)
        if quantity < 1 or quantity > 9999:
            raise ValueError("上架库存必须在 1-9999 之间")
        if worker_count < 1:
            raise ValueError("上架并发必须是大于 0 的整数")
        rows = db_get_mercado_product_items_by_ids(item_ids)
        if len(rows) != len({int(value) for value in item_ids}):
            raise ValueError("部分勾选产品已不存在，请刷新列表后重试")
        blocked_rows = []
        publish_rows = []
        for row in rows:
            issues = product_publish_issues(row)
            if issues:
                blocked_rows.append((row, issues))
            else:
                publish_rows.append(row)
        moved_to_collection_count = 0
        if blocked_rows:
            reason_counts: dict[str, int] = {}
            for _row, issues in blocked_rows:
                for issue in issues:
                    reason_counts[issue] = reason_counts.get(issue, 0) + 1
            reason_summary = "；".join(
                f"{reason} {count} 件" for reason, count in reason_counts.items()
            )
            move_result = db_move_mercado_product_items_to_collection(
                [int(row.get("id") or 0) for row, _issues in blocked_rows],
                reason=reason_summary,
            )
            moved_to_collection_count = int(
                move_result.get("moved") or move_result.get("deleted") or 0
            )
        rows = publish_rows
        if not rows:
            message = (
                f"已忽略并移回采集列表 {moved_to_collection_count} 件不可上架商品，"
                "本次没有可上架产品"
            )
            with _mercado_publish_lock:
                _mercado_publish_state.update({
                    "running": False,
                    "batch_id": "",
                    "status": "completed",
                    "message": message,
                    "selection_mode": selection_mode,
                    "token_id": token_id,
                    "token_ids": token_ids,
                    "group_names": group_names,
                    "site_id": site_id,
                    "site_ids": site_ids,
                    "site_name": site_name,
                    "target_count": 0,
                    "completed_target_count": 0,
                    "skipped_target_count": 0,
                    "quantity": quantity,
                    "worker_count": 0,
                    "requested_count": 0,
                    "processed_count": 0,
                    "published_count": 0,
                    "failed_count": 0,
                    "moved_to_collection_count": moved_to_collection_count,
                    "skipped_published_count": 0,
                    "elapsed_seconds": 0,
                    "average_seconds_per_item": 0,
                    "items_per_minute": 0,
                    "estimated_remaining_seconds": 0,
                    "average_stage_seconds": {},
                    "current_item_id": "",
                    "started_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "results": [],
                })
                state = dict(_mercado_publish_state)
            return jsonify({"status": "success", "data": state})
        validate_publishable_products(rows)
        token_rows = list((bit_db_api.list_mercado_store_tokens() or {}).get("rows") or [])
        token_by_id = {
            int(row.get("id") or 0): row
            for row in token_rows
            if int(row.get("id") or 0) > 0
        }
        if selection_mode == "accounts":
            missing_token_ids = [value for value in token_ids if value not in token_by_id]
            if missing_token_ids:
                raise ValueError(
                    "部分授权账号不存在，请刷新后重试："
                    + "、".join(str(value) for value in missing_token_ids)
                )
            candidate_tokens = [token_by_id[value] for value in token_ids]
        else:
            candidate_tokens = [
                row for row in token_rows
                if row.get("status") != "expired" or row.get("has_refresh_token")
            ]

        targets = []
        skipped_target_count = 0
        for token in candidate_tokens:
            current_token_id = int(token.get("id") or 0)
            token_site_id = str(token.get("site_id") or "").strip().upper()
            store_name = str(
                token.get("display_name") or token.get("nickname") or current_token_id
            )
            settings_by_site = {
                str(setting.get("site_id") or "").strip().upper(): setting
                for setting in (token.get("site_settings") or [])
                if str(setting.get("site_id") or "").strip()
            }
            for current_site_id in site_ids:
                compatible = (
                    not token_site_id
                    or token_site_id == "CBT"
                    or token_site_id == current_site_id
                )
                if not compatible:
                    if selection_mode == "accounts":
                        skipped_target_count += 1
                    continue
                site_setting = settings_by_site.get(current_site_id) or {}
                configured_group = str(site_setting.get("group_name") or "").strip()
                normalized_group = configured_group or "__ungrouped__"
                if selection_mode == "groups" and normalized_group not in group_names:
                    continue
                targets.append({
                    "token_id": current_token_id,
                    "store_name": store_name,
                    "site_id": current_site_id,
                    "site_name": marketplace_site_name(current_site_id),
                    "group_name": configured_group,
                    "discount_rate": float(site_discount_rate(token, current_site_id)),
                })

        if not targets:
            if (
                selection_mode == "accounts"
                and len(token_ids) == 1
                and len(site_ids) == 1
                and token_ids[0] in token_by_id
            ):
                selected_token_site = str(
                    token_by_id[token_ids[0]].get("site_id") or ""
                ).strip().upper()
                if selected_token_site and selected_token_site != "CBT":
                    raise ValueError(
                        f"所选店铺属于 {selected_token_site} 站点，不能发布到 {site_ids[0]}；"
                        "请选择 Global Selling 店铺或对应本地站点"
                    )
            if selection_mode == "groups":
                raise ValueError("所选分组和站点没有可用的账号-站点组合")
            raise ValueError("所选账号和站点没有兼容的上架组合")

        target_token_ids = list(dict.fromkeys(
            int(target["token_id"]) for target in targets
        ))
        target_site_ids = list(dict.fromkeys(
            str(target["site_id"]) for target in targets
        ))
        target_store_names = list(dict.fromkeys(
            str(target["store_name"]) for target in targets
        ))
        target_site_names = list(dict.fromkeys(
            str(target["site_name"]) for target in targets
        ))
        store_name = (
            target_store_names[0]
            if len(target_store_names) == 1
            else f"{len(target_store_names)} 个账号"
        )
        site_name = (
            target_site_names[0]
            if len(target_site_names) == 1
            else f"{len(target_site_names)} 个站点"
        )
        discount_rate = (
            float(targets[0]["discount_rate"])
            if len(targets) == 1 else None
        )
        current_user = get_current_workbench_user() or {}
        created_by = str(
            current_user.get("display_name") or current_user.get("username") or ""
        )[:128]
        batch_id = f"{datetime.now().strftime('%Y%m%d%H%M%S')}-{secrets.token_hex(4)}"
    except (TypeError, ValueError) as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("准备 Mercado 批量上架失败")
        return jsonify({"status": "error", "message": f"准备批量上架失败：{exc}"}), 500

    with _mercado_publish_lock:
        if _mercado_publish_state.get("running"):
            return jsonify({
                "status": "error",
                "message": "已有批量上架任务正在运行",
            }), 409
        _mercado_publish_state.update({
            "running": True,
            "batch_id": batch_id,
            "status": "running",
            "message": (
                f"准备使用 {min(worker_count, len(rows))} 个线程上架 "
                f"{len(rows)} 件产品到 {len(targets)} 个账号-站点组合 · "
                "各组合按对应站点折扣计算净收益"
                + (
                    f"；已忽略并移回采集列表 {moved_to_collection_count} 件不可上架商品"
                    if moved_to_collection_count else ""
                )
            ),
            "selection_mode": selection_mode,
            "token_id": token_id,
            "token_ids": target_token_ids,
            "group_names": group_names,
            "store_name": store_name,
            "site_id": target_site_ids[0] if len(target_site_ids) == 1 else "",
            "site_ids": target_site_ids,
            "site_name": site_name,
            "target_count": len(targets),
            "completed_target_count": 0,
            "skipped_target_count": skipped_target_count,
            "quantity": quantity,
            "worker_count": min(worker_count, len(rows)),
            "discount_rate": discount_rate,
            "requested_count": len(rows) * len(targets),
            "processed_count": 0,
            "published_count": 0,
            "failed_count": 0,
            "moved_to_collection_count": moved_to_collection_count,
            "skipped_published_count": 0,
            "elapsed_seconds": 0,
            "average_seconds_per_item": 0,
            "items_per_minute": 0,
            "estimated_remaining_seconds": 0,
            "average_stage_seconds": {},
            "current_item_id": "",
            "started_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "finished_at": "",
            "results": [],
        })
        thread = threading.Thread(
            target=_run_mercado_product_publish_targets,
            args=(
                rows,
                targets,
                quantity,
                worker_count,
                batch_id,
                created_by,
                moved_to_collection_count,
            ),
            name=f"mercado-publish-{batch_id}",
            daemon=True,
        )
        try:
            thread.start()
        except Exception as exc:
            _mercado_publish_state.update(
                running=False,
                status="error",
                message=f"批量上架线程启动失败：{exc}",
            )
            raise
        state = dict(_mercado_publish_state)
    return jsonify({"status": "success", "data": state})


@app.route('/api/mercado-products/publish/status', methods=['GET'])
@login_required
def api_mercado_product_publish_status():
    with _mercado_publish_lock:
        state = {**_mercado_publish_state, "results": list(_mercado_publish_state.get("results") or [])}
    response = jsonify({"status": "success", "data": state})
    response.headers["Cache-Control"] = "no-store"
    return response


@app.route('/api/mercado-shipping-rates', methods=['GET'])
@login_required
def api_mercado_shipping_rates():
    try:
        from erp.mercadolibre_shipping_rate_cards import OfficialShippingRateCardStore

        site_id = str(request.args.get("site_id") or "").strip().upper()
        result = OfficialShippingRateCardStore().list_rates(site_id)
        result["refresh"] = _mercado_shipping_rate_refresh_snapshot()
        response = jsonify({"status": "success", "data": result})
        response.headers["Cache-Control"] = "no-store"
        return response
    except Exception as exc:
        logging.exception("读取 Mercado 官方运费表失败")
        return jsonify({"status": "error", "message": f"读取官方运费表失败：{exc}"}), 500


@app.route('/api/mercado-shipping-rates/refresh', methods=['POST'])
@login_required
def api_refresh_mercado_shipping_rates():
    started = _start_mercado_shipping_rate_refresh()
    data = _mercado_shipping_rate_refresh_snapshot()
    return jsonify({
        "status": "success",
        "message": "已开始更新 Global Selling 官方最新标准" if started else "官方标准更新正在运行",
        "data": data,
    }), (202 if started else 200)


@app.route('/api/mercado-publish-records', methods=['GET'])
@login_required
def api_mercado_product_publish_records():
    try:
        result = db_list_mercado_product_publish_records(
            search=str(request.args.get("search") or "").strip(),
            status=str(request.args.get("status") or "").strip(),
            store_name=str(request.args.get("store_name") or "").strip(),
            site_id=str(request.args.get("site_id") or "").strip(),
            limit=_parse_int_param(request.args, "limit", 500, 1, 1000),
            offset=_parse_int_param(request.args, "offset", 0, 0, 1000000),
        )
        with _mercado_publish_lock:
            result["publish_running"] = bool(_mercado_publish_state.get("running"))
        response = jsonify({"status": "success", "data": result})
        response.headers["Cache-Control"] = "no-store"
        return response
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("读取 Mercado 产品上架记录失败")
        return jsonify({"status": "error", "message": f"读取产品上架记录失败：{exc}"}), 500


@app.route('/api/mercado-publish-records/retry', methods=['POST'])
@login_required
def api_retry_mercado_product_publish_records():
    data = request.get_json(silent=True) or {}
    record_ids = data.get("record_ids") or []
    if not isinstance(record_ids, list):
        return jsonify({"status": "error", "message": "record_ids 必须是数组"}), 422
    with _mercado_publish_lock:
        if _mercado_publish_state.get("running"):
            return jsonify({
                "status": "error",
                "message": "已有批量上架任务正在运行，完成后再重新上架",
            }), 409
    try:
        from erp.mercadolibre_batch_publish import (
            site_discount_rate,
            validate_publishable_products,
        )
        from erp.mercadolibre_collection_store import (
            PRODUCT_PUBLISH_RETRYABLE_STATUSES,
        )
        from erp.mercadolibre_translation import (
            marketplace_site_name,
            normalize_marketplace_site,
        )

        normalized_record_ids = list(dict.fromkeys(
            int(value) for value in record_ids if int(value) > 0
        ))
        if not normalized_record_ids:
            raise ValueError("请至少勾选一条可重新上架的记录")
        records = db_get_mercado_product_publish_records_by_ids(normalized_record_ids)
        if len(records) != len(normalized_record_ids):
            raise ValueError("部分上架记录已不存在，请刷新后重试")
        blocked = [
            record for record in records
            if str(record.get("status") or "") not in PRODUCT_PUBLISH_RETRYABLE_STATUSES
        ]
        if blocked:
            raise ValueError("只有上架暂停或上架失败的记录可以重新上架")

        # A product may have several failed attempts. Keep only the newest selected
        # attempt for the same product/account/site to prevent duplicate listings.
        retry_records = []
        seen_product_targets = set()
        for record in records:
            key = (
                int(record.get("product_item_id") or 0),
                int(record.get("token_id") or 0),
                str(record.get("site_id") or "").strip().upper(),
            )
            if min(key[0], key[1]) <= 0 or not key[2]:
                raise ValueError("所选记录缺少产品、账号或站点信息")
            if key in seen_product_targets:
                continue
            seen_product_targets.add(key)
            retry_records.append(record)
        duplicate_selection_count = len(records) - len(retry_records)

        product_ids = list(dict.fromkeys(
            int(record.get("product_item_id") or 0) for record in retry_records
        ))
        products = db_get_mercado_product_items_by_ids(product_ids)
        product_by_id = {int(row.get("id") or 0): dict(row) for row in products}
        missing_product_ids = [value for value in product_ids if value not in product_by_id]
        if missing_product_ids:
            raise ValueError("部分记录对应的产品已从产品列表删除，无法重新上架")
        validate_publishable_products(products)

        token_rows = list((bit_db_api.list_mercado_store_tokens() or {}).get("rows") or [])
        token_by_id = {
            int(row.get("id") or 0): dict(row)
            for row in token_rows if int(row.get("id") or 0) > 0
        }
        grouped_targets: dict[tuple[int, str, int], dict] = {}
        for record in retry_records:
            token_id = int(record.get("token_id") or 0)
            token = token_by_id.get(token_id)
            if not token:
                raise ValueError(f"上架账号 {token_id} 已不存在，请重新授权后再试")
            site_id = normalize_marketplace_site(record.get("site_id") or "")
            token_site_id = str(token.get("site_id") or "").strip().upper()
            if token_site_id and token_site_id != "CBT" and token_site_id != site_id:
                raise ValueError(
                    f"账号 {token_id} 属于 {token_site_id} 站点，不能重新发布到 {site_id}"
                )
            quantity = int(record.get("quantity") or 1)
            if quantity < 1 or quantity > 9999:
                raise ValueError("历史记录中的上架库存无效")
            group_key = (token_id, site_id, quantity)
            target = grouped_targets.setdefault(group_key, {
                "token_id": token_id,
                "store_name": str(
                    token.get("display_name") or token.get("nickname")
                    or record.get("store_name") or token_id
                ),
                "site_id": site_id,
                "site_name": marketplace_site_name(site_id),
                "discount_rate": float(site_discount_rate(token, site_id)),
                "quantity": quantity,
                "product_rows": [],
            })
            product = product_by_id[int(record.get("product_item_id") or 0)]
            target["product_rows"].append(product)

        targets = list(grouped_targets.values())
        requested_count = sum(len(target["product_rows"]) for target in targets)
        worker_count = int(data.get("worker_count") or 10)
        if worker_count < 1:
            raise ValueError("重新上架并发必须是大于 0 的整数")
        current_user = get_current_workbench_user() or {}
        created_by = str(
            current_user.get("display_name") or current_user.get("username") or ""
        )[:128]
        batch_id = f"retry-{datetime.now().strftime('%Y%m%d%H%M%S')}-{secrets.token_hex(4)}"
        token_ids = list(dict.fromkeys(int(target["token_id"]) for target in targets))
        site_ids = list(dict.fromkeys(str(target["site_id"]) for target in targets))
        quantities = list(dict.fromkeys(int(target["quantity"]) for target in targets))
    except (TypeError, ValueError) as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("准备重新上架 Mercado 产品失败")
        return jsonify({"status": "error", "message": f"准备重新上架失败：{exc}"}), 500

    with _mercado_publish_lock:
        if _mercado_publish_state.get("running"):
            return jsonify({
                "status": "error",
                "message": "已有批量上架任务正在运行，完成后再重新上架",
            }), 409
        _mercado_publish_state.update({
            "running": True,
            "batch_id": batch_id,
            "status": "running",
            "message": f"正在重新上架 {requested_count} 件产品",
            "selection_mode": "retry",
            "token_id": token_ids[0] if len(token_ids) == 1 else None,
            "token_ids": token_ids,
            "group_names": [],
            "store_name": (
                targets[0]["store_name"] if len(targets) == 1
                else f"{len(token_ids)} 个账号"
            ),
            "site_id": site_ids[0] if len(site_ids) == 1 else "",
            "site_ids": site_ids,
            "site_name": (
                targets[0]["site_name"] if len(site_ids) == 1
                else f"{len(site_ids)} 个站点"
            ),
            "target_count": len(targets),
            "completed_target_count": 0,
            "skipped_target_count": 0,
            "quantity": quantities[0] if len(quantities) == 1 else 0,
            "worker_count": min(worker_count, requested_count),
            "discount_rate": None,
            "requested_count": requested_count,
            "processed_count": 0,
            "published_count": 0,
            "failed_count": 0,
            "moved_to_collection_count": 0,
            "skipped_published_count": 0,
            "duplicate_selection_count": duplicate_selection_count,
            "elapsed_seconds": 0,
            "average_seconds_per_item": 0,
            "items_per_minute": 0,
            "estimated_remaining_seconds": 0,
            "current_item_id": "",
            "started_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "finished_at": "",
            "results": [],
        })
        thread = threading.Thread(
            target=_run_mercado_product_publish_targets,
            args=([], targets, 1, worker_count, batch_id, created_by, 0),
            name=f"mercado-publish-retry-{batch_id}",
            daemon=True,
        )
        try:
            thread.start()
        except Exception as exc:
            _mercado_publish_state.update(
                running=False,
                status="error",
                message=f"重新上架线程启动失败：{exc}",
            )
            raise
        state = dict(_mercado_publish_state)
    return jsonify({"status": "success", "data": state})


@app.route('/api/db/mercado-collection/tasks', methods=['POST'])
@internal_api_required
def api_db_create_mercado_collection_task():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    try:
        task_id = db_create_mercado_collection_task(
            data.get("source_url", ""),
            data.get("requested_count", 20),
            data.get("created_by", ""),
        )
        return jsonify({"status": "success", "data": {"task_id": int(task_id)}})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400


@app.route('/api/db/mercado-collection/tasks/<int:task_id>', methods=['GET', 'PATCH'])
@internal_api_required
def api_db_mercado_collection_task(task_id):
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    if request.method == "GET":
        row = db_get_mercado_collection_task(task_id)
        if not row:
            return jsonify({"status": "error", "message": "采集任务不存在"}), 404
        return jsonify({"status": "success", "data": row})
    changes = request.get_json(silent=True) or {}
    allowed = {
        "status", "message", "collected_count", "completed_count",
        "failed_count", "current_page", "started", "finished",
    }
    db_update_mercado_collection_task(
        task_id, **{key: value for key, value in changes.items() if key in allowed}
    )
    return jsonify({"status": "success", "data": {"task_id": task_id}})


@app.route('/api/db/mercado-collection/items', methods=['GET', 'POST'])
@internal_api_required
def api_db_mercado_collection_items():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    if request.method == "GET":
        task_id = request.args.get("task_id")
        result = db_list_mercado_collection_items(
            search=str(request.args.get("search") or "").strip(),
            limit=_parse_int_param(request.args, "limit", 500, 1, 1000),
            offset=_parse_int_param(request.args, "offset", 0, 0, 1000000),
            task_id=int(task_id) if str(task_id or "").strip() else None,
            weight_min=str(request.args.get("weight_min") or "").strip(),
            weight_max=str(request.args.get("weight_max") or "").strip(),
            price_min=str(request.args.get("price_min") or "").strip(),
            price_max=str(request.args.get("price_max") or "").strip(),
            net_proceeds_min=str(request.args.get("net_proceeds_min") or "").strip(),
            net_proceeds_max=str(request.args.get("net_proceeds_max") or "").strip(),
            date_from=str(request.args.get("date_from") or "").strip(),
            date_to=str(request.args.get("date_to") or "").strip(),
            management_category_id=str(
                request.args.get("management_category_id") or ""
            ).strip(),
            exclude_added=str(request.args.get("exclude_added") or "").lower()
            in {"1", "true", "yes"},
        )
        return jsonify({"status": "success", "data": result})
    data = request.get_json(silent=True) or {}
    rows = data.get("rows") or []
    if not isinstance(rows, list):
        return jsonify({"status": "error", "message": "rows 必须是数组"}), 422
    count = db_upsert_mercado_collection_items(int(data.get("task_id")), rows)
    return jsonify({"status": "success", "data": {"count": count}})


@app.route('/api/db/mercado-products', methods=['GET'])
@internal_api_required
def api_db_mercado_products():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    result = db_list_mercado_product_items(
        search=str(request.args.get("search") or "").strip(),
        limit=_parse_int_param(request.args, "limit", 500, 1, 1000),
        offset=_parse_int_param(request.args, "offset", 0, 0, 1000000),
        source_type=str(request.args.get("source_type") or "").strip(),
        review_status=str(request.args.get("review_status") or "").strip(),
        publish_status=str(request.args.get("publish_status") or "").strip(),
        weight_min=str(request.args.get("weight_min") or "").strip(),
        weight_max=str(request.args.get("weight_max") or "").strip(),
        price_min=str(request.args.get("price_min") or "").strip(),
        price_max=str(request.args.get("price_max") or "").strip(),
        net_proceeds_min=str(request.args.get("net_proceeds_min") or "").strip(),
        net_proceeds_max=str(request.args.get("net_proceeds_max") or "").strip(),
        date_from=str(request.args.get("date_from") or "").strip(),
        date_to=str(request.args.get("date_to") or "").strip(),
        management_category_id=str(
            request.args.get("management_category_id") or ""
        ).strip(),
        mercado_category=str(
            request.args.get("mercado_category") or ""
        ).strip(),
    )
    return jsonify({"status": "success", "data": result})


@app.route('/api/db/mercado-products/<int:product_item_id>', methods=['PATCH'])
@internal_api_required
def api_db_update_mercado_product(product_item_id):
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True)
    if not isinstance(data, dict):
        return jsonify({"status": "error", "message": "产品内容必须是对象"}), 422
    try:
        result = db_update_mercado_product_item(product_item_id, data)
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except KeyError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 404
    return jsonify({"status": "success", "data": result})


@app.route('/api/db/mercado-publish-records', methods=['GET'])
@internal_api_required
def api_db_mercado_product_publish_records():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    result = db_list_mercado_product_publish_records(
        search=str(request.args.get("search") or "").strip(),
        status=str(request.args.get("status") or "").strip(),
        store_name=str(request.args.get("store_name") or "").strip(),
        site_id=str(request.args.get("site_id") or "").strip(),
        limit=_parse_int_param(request.args, "limit", 500, 1, 1000),
        offset=_parse_int_param(request.args, "offset", 0, 0, 1000000),
    )
    return jsonify({"status": "success", "data": result})


@app.route('/api/db/mercado-publish-records/published-product-ids', methods=['POST'])
@internal_api_required
def api_db_published_mercado_product_item_ids():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    item_ids = data.get("product_item_ids") or []
    if not isinstance(item_ids, list):
        return jsonify({"status": "error", "message": "product_item_ids 必须是数组"}), 422
    try:
        published_ids = db_get_published_mercado_product_item_ids(
            item_ids,
            token_id=int(data.get("token_id") or 0),
            site_id=str(data.get("site_id") or ""),
        )
    except (TypeError, ValueError) as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    return jsonify({
        "status": "success",
        "data": {"product_item_ids": published_ids},
    })


@app.route('/api/db/mercado-publish-records/by-ids', methods=['POST'])
@internal_api_required
def api_db_mercado_product_publish_records_by_ids():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    record_ids = data.get("record_ids") or []
    if not isinstance(record_ids, list):
        return jsonify({"status": "error", "message": "record_ids 必须是数组"}), 422
    try:
        rows = db_get_mercado_product_publish_records_by_ids(record_ids)
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    return jsonify({"status": "success", "data": {"rows": rows}})


@app.route('/api/db/mercado-publish-records/bulk', methods=['POST'])
@internal_api_required
def api_db_create_mercado_product_publish_records():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    rows = data.get("rows") or []
    if not isinstance(rows, list):
        return jsonify({"status": "error", "message": "rows 必须是数组"}), 422
    try:
        record_ids = db_create_mercado_product_publish_records(
            rows,
            batch_id=data.get("batch_id", ""),
            token_id=data.get("token_id", 0),
            store_name=data.get("store_name", ""),
            site_id=data.get("site_id", ""),
            site_name=data.get("site_name", ""),
            quantity=data.get("quantity", 1),
            created_by=data.get("created_by", ""),
        )
    except (TypeError, ValueError) as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    return jsonify({
        "status": "success",
        "data": {"record_ids": {str(key): value for key, value in record_ids.items()}},
    })


@app.route('/api/db/mercado-publish-records/<int:record_id>', methods=['PATCH'])
@internal_api_required
def api_db_update_mercado_product_publish_record(record_id):
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    allowed = {
        "status", "published_item_id", "failure_reason", "result", "started", "finished",
    }
    try:
        db_update_mercado_product_publish_record(
            record_id,
            **{key: value for key, value in data.items() if key in allowed},
        )
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except KeyError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 404
    return jsonify({"status": "success", "data": {"record_id": record_id}})


@app.route('/api/db/mercado-products/review-status', methods=['POST'])
@internal_api_required
def api_db_update_mercado_product_review_status():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    item_ids = data.get("product_item_ids") or []
    if not isinstance(item_ids, list):
        return jsonify({"status": "error", "message": "product_item_ids 必须是数组"}), 422
    try:
        result = db_update_mercado_product_review_status(
            item_ids, str(data.get("review_status") or "").strip()
        )
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    return jsonify({"status": "success", "data": result})


@app.route('/api/db/mercado-products/move-to-collection', methods=['POST'])
@internal_api_required
def api_db_move_mercado_products_to_collection():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    item_ids = data.get("product_item_ids") or []
    if not isinstance(item_ids, list):
        return jsonify({"status": "error", "message": "product_item_ids 必须是数组"}), 422
    try:
        result = db_move_mercado_product_items_to_collection(
            item_ids,
            reason=str(data.get("reason") or "不可上架"),
        )
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    return jsonify({"status": "success", "data": result})


@app.route('/api/db/mercado-products/add', methods=['POST'])
@internal_api_required
def api_db_add_mercado_products():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    item_ids = data.get("collection_item_ids") or []
    if not isinstance(item_ids, list):
        return jsonify({"status": "error", "message": "collection_item_ids 必须是数组"}), 422
    result = db_add_mercado_collection_items_to_products(item_ids)
    return jsonify({"status": "success", "data": result})


@app.route('/api/db/mercado-collection/items/delete', methods=['POST'])
@internal_api_required
def api_db_delete_mercado_collection_items():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    item_ids = data.get("collection_item_ids") or []
    if not isinstance(item_ids, list):
        return jsonify({"status": "error", "message": "collection_item_ids 必须是数组"}), 422
    return jsonify({"status": "success", "data": db_delete_mercado_collection_items(item_ids)})


@app.route('/api/db/mercado-products/delete', methods=['POST'])
@internal_api_required
def api_db_delete_mercado_product_items():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    item_ids = data.get("product_item_ids") or []
    if not isinstance(item_ids, list):
        return jsonify({"status": "error", "message": "product_item_ids 必须是数组"}), 422
    return jsonify({"status": "success", "data": db_delete_mercado_product_items(item_ids)})


@app.route('/api/db/mercado-products/by-ids', methods=['POST'])
@internal_api_required
def api_db_get_mercado_product_items_by_ids():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    item_ids = data.get("product_item_ids") or []
    if not isinstance(item_ids, list):
        return jsonify({"status": "error", "message": "product_item_ids 必须是数组"}), 422
    rows = db_get_mercado_product_items_by_ids(item_ids)
    return jsonify({"status": "success", "data": {"rows": rows}})


@app.route('/api/db/mercado-products/<int:product_item_id>/publish-state', methods=['PATCH'])
@internal_api_required
def api_db_update_mercado_product_publish_state(product_item_id):
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    allowed = {
        "status", "store_name", "token_id", "published_item_id",
        "error_message", "result", "finished",
    }
    db_update_mercado_product_publish_state(
        product_item_id,
        **{key: value for key, value in data.items() if key in allowed},
    )
    return jsonify({"status": "success", "data": {"product_item_id": product_item_id}})


@app.route('/api/db/task-records', methods=['POST'])
@internal_api_required
def api_db_insert_task_records():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    records = data.get("records") or []
    db_insert_task_record(records)
    return jsonify({"status": "success", "data": {"count": len(records)}})


@app.route('/api/db/task-records/order-print/latest', methods=['GET'])
@internal_api_required
def api_db_latest_order_print_records():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    return jsonify({
        "status": "success",
        "data": db_get_latest_order_print_records(),
    })


@app.route('/api/db/health', methods=['GET'])
@internal_api_required
def api_db_health():
    """Let clients verify that they reached the database-owning process."""
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    return jsonify({
        "status": "success",
        "data": {
            "role": "server",
            "database_host": mysql_config.get("host"),
        },
    })


@app.route('/api/db/official-infractions/dashboard', methods=['GET'])
@internal_api_required
def api_db_official_infraction_dashboard():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    try:
        data = list_infraction_dashboard(
            days=request.args.get("days", 30),
            view_mode=request.args.get("view_mode", "current"),
            group_name=request.args.get("group_name", ""),
            salesperson=request.args.get("salesperson", ""),
            source_type=request.args.get("source_type", ""),
            category=request.args.get("category", ""),
            search=request.args.get("search", ""),
            detail_token_id=request.args.get("detail_token_id", 0),
            page=request.args.get("page", 1),
            page_size=request.args.get("page_size", 100),
        )
        return jsonify({"status": "success", "data": data})
    except (TypeError, ValueError) as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400


@app.route('/api/db/official-infractions/current-counts', methods=['GET'])
@internal_api_required
def api_db_official_infraction_current_counts():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    snapshot = current_infraction_counts_by_token_site(
        request.args.get("days", 100)
    ) or {}
    count_rows = []
    for (token_id, site_id), values in (snapshot.get("counts") or {}).items():
        count_rows.append({
            "token_id": int(token_id),
            "site_id": str(site_id),
            **dict(values or {}),
        })
    data = {key: value for key, value in snapshot.items() if key != "counts"}
    data["count_rows"] = count_rows
    return jsonify({"status": "success", "data": data})


@app.route('/api/db/official-infractions/live', methods=['POST'])
@internal_api_required
def api_db_collect_live_official_infractions():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    payload = request.get_json(silent=True) or {}
    targets = payload.get("targets") or []
    if not isinstance(targets, list):
        return jsonify({"status": "error", "message": "targets 必须是数组"}), 400
    try:
        data = mercado_infraction_sync.collect_live_detection_infractions(
            targets,
            recent_days=payload.get("recent_days", 100),
            max_workers=payload.get("max_workers", 8),
        )
        return jsonify({"status": "success", "data": data})
    except (TypeError, ValueError) as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400


@app.route('/api/db/official-infractions/sync', methods=['POST'])
@internal_api_required
def api_db_start_official_infraction_sync():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    payload = request.get_json(silent=True) or {}
    token_ids = payload.get("token_ids") or []
    if not isinstance(token_ids, list):
        return jsonify({"status": "error", "message": "token_ids 必须是数组"}), 400
    try:
        started, state = mercado_infraction_sync.start_official_infraction_sync(token_ids)
        return jsonify({
            "status": "success",
            "data": {"started": bool(started), "state": state},
        })
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400


@app.route('/api/db/official-infractions/sync/status', methods=['GET'])
@internal_api_required
def api_db_official_infraction_sync_status():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    return jsonify({
        "status": "success",
        "data": mercado_infraction_sync.official_infraction_sync_status(),
    })


def _mercado_token_error_response(exc):
    if isinstance(exc, KeyError):
        return jsonify({"status": "error", "message": str(exc.args[0])}), 404
    if isinstance(
        exc,
        (
            ValueError,
            mercado_tokens.MercadoTokenError,
            mercado_reputation.MercadoReputationError,
        ),
    ):
        return jsonify({"status": "error", "message": str(exc)}), 400
    logging.exception("店铺授权操作失败")
    return jsonify({"status": "error", "message": f"店铺授权操作失败：{exc}"}), 500


def _mercado_communication_error_response(exc):
    if isinstance(exc, KeyError):
        return jsonify({"status": "error", "message": str(exc.args[0])}), 404
    if isinstance(exc, ValueError):
        return jsonify({"status": "error", "message": str(exc)}), 400
    if isinstance(exc, mercado_communications.MercadoCommunicationError):
        status_code = int(exc.status_code or 502)
        if status_code not in (400, 401, 403, 404, 409, 422, 429):
            status_code = 502
        return jsonify({
            "status": "error",
            "message": str(exc),
            "model_6_restricted": bool(exc.model_6_restricted),
        }), status_code
    logging.exception("美客多客户消息操作失败")
    return jsonify({"status": "error", "message": f"美客多客户消息操作失败：{exc}"}), 500


@app.route('/api/db/mercado-tokens/authorization', methods=['GET'])
@internal_api_required
def api_db_mercado_token_authorization():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    return jsonify({
        "status": "success",
        "data": bit_db_api.get_mercado_token_authorization_info(),
    })


@app.route('/api/db/mercado-tokens', methods=['GET'])
@internal_api_required
def api_db_list_mercado_tokens():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    return jsonify({"status": "success", "data": bit_db_api.list_mercado_store_tokens()})


@app.route('/api/db/mercado-tokens/<int:token_id>/site-settings', methods=['GET', 'PUT'])
@internal_api_required
def api_db_mercado_token_site_settings(token_id):
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    try:
        if request.method == "GET":
            result = bit_db_api.list_mercado_store_site_settings(token_id)
        else:
            data = request.get_json(silent=True) or {}
            result = bit_db_api.update_mercado_store_site_settings(
                token_id, data.get("settings", [])
            )
        return jsonify({"status": "success", "data": result})
    except Exception as exc:
        return _mercado_token_error_response(exc)


@app.route('/api/db/mercado-tokens/exchange', methods=['POST'])
@internal_api_required
def api_db_exchange_mercado_token():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    try:
        result = bit_db_api.exchange_mercado_store_token(
            data.get("display_name", ""), data.get("code", "")
        )
        return jsonify({"status": "success", "data": result})
    except Exception as exc:
        return _mercado_token_error_response(exc)


@app.route('/api/db/mercado-tokens/<int:token_id>/refresh', methods=['POST'])
@internal_api_required
def api_db_refresh_mercado_token(token_id):
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    try:
        result = bit_db_api.refresh_mercado_store_token(token_id)
        return jsonify({"status": "success", "data": result})
    except Exception as exc:
        return _mercado_token_error_response(exc)


@app.route('/api/db/mercado-tokens/<int:token_id>/reputation', methods=['GET'])
@internal_api_required
def api_db_mercado_token_reputation(token_id):
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    try:
        result = bit_db_api.get_mercado_store_reputation(token_id)
        return jsonify({"status": "success", "data": result})
    except Exception as exc:
        return _mercado_token_error_response(exc)


@app.route(
    '/api/db/mercado-communications/<int:token_id>/<action>',
    methods=['POST'],
)
@internal_api_required
def api_db_mercado_communication(token_id, action):
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    try:
        result = bit_db_api.execute_mercado_store_communication(
            token_id, action, request.get_json(silent=True) or {}
        )
        return jsonify({"status": "success", "data": result})
    except Exception as exc:
        return _mercado_communication_error_response(exc)


@app.route('/api/db/mercado-tokens/<int:token_id>', methods=['PATCH', 'DELETE'])
@internal_api_required
def api_db_update_mercado_token(token_id):
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    try:
        if request.method == "DELETE":
            affected = bit_db_api.delete_mercado_store_token(token_id)
            if not affected:
                raise KeyError("店铺授权不存在")
            return jsonify({"status": "success", "data": {"deleted": affected}})
        data = request.get_json(silent=True) or {}
        if "enabled" in data:
            if not isinstance(data.get("enabled"), bool):
                raise ValueError("店铺启用状态必须是布尔值")
            result = bit_db_api.set_mercado_store_token_enabled(
                token_id, data["enabled"]
            )
        else:
            result = bit_db_api.rename_mercado_store_token(
                token_id, data.get("display_name", "")
            )
        return jsonify({"status": "success", "data": result})
    except Exception as exc:
        return _mercado_token_error_response(exc)


@app.route('/api/db/reputation/bulk', methods=['POST'])
@internal_api_required
def api_db_insert_reputation():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    rows = data.get("rows") or []
    replace_targets = data.get("replace_targets") or []
    count = db_inset_reputation_info(
        rows,
        merge_latest=bool(data.get("merge_latest", False)),
        replace_targets=replace_targets,
    )
    return jsonify({"status": "success", "data": {"count": count}})


@app.route('/api/db/infractions/bulk', methods=['POST'])
@internal_api_required
def api_db_insert_infractions():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    rows = data.get("rows") or []
    replace_targets = data.get("replace_targets") or []
    count = db_inset_infraction_info(
        rows,
        merge_latest=bool(data.get("merge_latest", False)),
        replace_targets=replace_targets,
    )
    return jsonify({"status": "success", "data": {"count": count}})


@app.route('/api/db/delays/bulk', methods=['POST'])
@internal_api_required
def api_db_insert_delays():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    rows = data.get("rows") or []
    db_inset_delay_info(rows)
    return jsonify({"status": "success", "data": {"count": len(rows)}})


@app.route('/api/db/pago/bulk', methods=['POST'])
@internal_api_required
def api_db_insert_pago():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    rows = data.get("rows") or []
    db_inset_pago_info(rows)
    return jsonify({"status": "success", "data": {"count": len(rows)}})


@app.route('/api/db/zying-products/bulk', methods=['POST'])
@internal_api_required
def api_db_insert_zying_products():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    rows = data.get("rows") or []
    count = db_insert_zying_product_info(rows)
    return jsonify({"status": "success", "data": {"count": count}})


@app.route('/api/db/zying-products/existing', methods=['POST'])
@internal_api_required
def api_db_existing_zying_product_ids():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    product_ids = data.get("product_ids") or []
    if not isinstance(product_ids, list):
        return jsonify({
            "status": "error",
            "message": "product_ids 必须是数组",
        }), 400
    existing_ids = sorted(db_get_existing_zying_product_ids(product_ids))
    return jsonify({
        "status": "success",
        "data": {"product_ids": existing_ids},
    })


@app.route('/api/db/zying-products/product-list', methods=['POST'])
@internal_api_required
def api_db_upsert_zying_product_list():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    rows = data.get("rows") or []
    if not isinstance(rows, list):
        return jsonify({"status": "error", "message": "rows 必须是数组"}), 400
    result = db_upsert_zying_products_to_products(rows)
    return jsonify({"status": "success", "data": result})


@app.route('/api/db/zying-risk/candidates', methods=['GET'])
@internal_api_required
def api_db_zying_risk_candidates():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    rows = db_get_zying_risk_candidates(
        hours=_parse_int_param(request.args, "hours", 24, 0, 87600),
        limit=_parse_int_param(request.args, "limit", 0, 0, 50000),
        zying_category=str(request.args.get("category") or "").strip() or None,
        include_checked=_parse_bool_param(request.args, "include_checked", False),
    )
    return jsonify({"status": "success", "data": rows})


@app.route('/api/db/zying-risk/bulk', methods=['POST'])
@internal_api_required
def api_db_update_zying_risks():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    results = data.get("results") or []
    if not isinstance(results, list):
        return jsonify({"status": "error", "message": "results must be an array"}), 422
    count = db_update_zying_product_risks(results)
    return jsonify({"status": "success", "data": {"count": count}})


@app.route('/api/db/zying-risk/categories', methods=['GET'])
@internal_api_required
def api_db_zying_risk_categories():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    return jsonify({"status": "success", "data": db_list_zying_risk_categories()})


@app.route('/api/db/zying-risk/results', methods=['GET'])
@internal_api_required
def api_db_zying_risk_results():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    try:
        params = _risk_result_query_params(
            request.args,
            export=str(request.args.get("limit") or "").strip() == "0",
        )
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    return jsonify({
        "status": "success",
        "data": db_get_zying_risk_results(**params),
    })


@app.route('/api/db/inventory/stocks', methods=['GET'])
@internal_api_required
def api_db_inventory_stocks():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    try:
        result = bit_inventory.list_inventory_stock(
            **_inventory_list_params(request.args)
        )
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    return jsonify({"status": "success", "data": result})


@app.route('/api/db/mercado-management-categories', methods=['GET', 'POST'])
@internal_api_required
def api_db_mercado_management_categories():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    try:
        if request.method == "GET":
            result = db_list_mercado_management_categories()
        else:
            data = request.get_json(silent=True) or {}
            result = db_create_mercado_management_category(data.get("name", ""))
        return jsonify({"status": "success", "data": result})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400


@app.route('/api/db/mercado-management-categories/<int:category_id>', methods=['PATCH', 'DELETE'])
@internal_api_required
def api_db_mercado_management_category(category_id):
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    try:
        if request.method == "DELETE":
            result = db_delete_mercado_management_category(category_id)
        else:
            data = request.get_json(silent=True) or {}
            result = db_update_mercado_management_category(
                category_id, data.get("name", "")
            )
        return jsonify({"status": "success", "data": result})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except KeyError as exc:
        return jsonify({"status": "error", "message": exc.args[0]}), 404


@app.route('/api/db/mercado-management-categories/assign', methods=['POST'])
@internal_api_required
def api_db_assign_mercado_management_category():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    item_ids = data.get("item_ids") or []
    if not isinstance(item_ids, list):
        return jsonify({"status": "error", "message": "item_ids 必须是数组"}), 422
    try:
        result = db_assign_mercado_management_category(
            data.get("item_type", ""), item_ids, data.get("category_id")
        )
        return jsonify({"status": "success", "data": result})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except KeyError as exc:
        return jsonify({"status": "error", "message": exc.args[0]}), 404


@app.route('/api/db/inventory/shelves', methods=['GET', 'POST'])
@internal_api_required
def api_db_inventory_shelves():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    try:
        if request.method == "GET":
            include_inactive = str(request.args.get("include_inactive") or "1").lower() not in (
                "0", "false", "no", "off",
            )
            result = bit_inventory.list_inventory_shelves(
                include_inactive=include_inactive
            )
        else:
            payload = request.get_json(silent=True)
            if not isinstance(payload, dict):
                return jsonify({"status": "error", "message": "record must be an object"}), 422
            result = bit_inventory.create_inventory_shelf(payload)
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    return jsonify({"status": "success", "data": result})


@app.route('/api/db/inventory/shelves/<int:shelf_id>', methods=['PATCH'])
@internal_api_required
def api_db_update_inventory_shelf(shelf_id):
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    payload = request.get_json(silent=True)
    if not isinstance(payload, dict):
        return jsonify({"status": "error", "message": "record must be an object"}), 422
    try:
        result = bit_inventory.update_inventory_shelf(shelf_id, payload)
    except KeyError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 404
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    return jsonify({"status": "success", "data": result})


@app.route('/api/db/inventory/matches', methods=['GET'])
@internal_api_required
def api_db_inventory_matches():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    result = bit_inventory.list_inventory_matches(
        search=str(request.args.get("search") or "").strip(),
        limit=_parse_int_param(request.args, "limit", 30, 1, 100),
    )
    return jsonify({"status": "success", "data": result})


@app.route('/api/db/inventory/movements', methods=['GET', 'POST'])
@internal_api_required
def api_db_inventory_movements():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    try:
        if request.method == "GET":
            result = bit_inventory.list_inventory_movements(
                **_inventory_movement_params(request.args)
            )
        else:
            payload = request.get_json(silent=True)
            if not isinstance(payload, dict):
                return jsonify({"status": "error", "message": "record must be an object"}), 422
            result = bit_inventory.create_inventory_movement(payload)
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    return jsonify({"status": "success", "data": result})


@app.route('/api/db/orders', methods=['GET'])
@internal_api_required
def api_db_list_orders():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    try:
        data = db_list_orders(**_order_list_query_params(request.args))
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    return jsonify({"status": "success", "data": data})


@app.route('/api/db/orders/weight-quote', methods=['POST'])
@internal_api_required
def api_db_order_weight_quote():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    order_ids = data.get("order_ids") or []
    if not isinstance(order_ids, list):
        return jsonify({"status": "error", "message": "order_ids must be an array"}), 422
    try:
        quote = bit_order_sync.bit_mysql.get_mercado_order_weight_quote(order_ids)
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except KeyError as exc:
        return jsonify({"status": "error", "message": str(exc).strip("'")}), 404
    return jsonify({"status": "success", "data": quote})


@app.route('/api/db/order-sync/start', methods=['POST'])
@internal_api_required
def api_db_start_order_sync():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    token_ids = data.get("token_ids") or []
    if not isinstance(token_ids, list):
        return jsonify({"status": "error", "message": "token_ids must be an array"}), 422
    try:
        started, state = bit_order_sync.start_order_sync(
            start_date=str(data.get("start_date") or "").strip(),
            end_date=str(data.get("end_date") or "").strip(),
            token_ids=token_ids,
            mode="automatic" if str(data.get("mode")) == "automatic" else "manual",
        )
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    return jsonify({"status": "success", "data": {"started": started, "state": state}})


@app.route('/api/db/order-sync/status', methods=['GET'])
@internal_api_required
def api_db_order_sync_status():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    return jsonify({"status": "success", "data": bit_order_sync.order_sync_status()})


@app.route('/api/db/store-links', methods=['GET'])
@internal_api_required
def api_db_store_links():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    try:
        token_text = str(request.args.get("token_id") or "").strip()
        data = db_list_mercado_store_links(
            search=str(request.args.get("search") or "").strip(),
            token_id=int(token_text) if token_text else None,
            site_id=str(request.args.get("site_id") or "").strip(),
            group_name=str(request.args.get("group_name") or "").strip(),
            status=str(request.args.get("status") or "").strip(),
            management_category_id=str(
                request.args.get("management_category_id") or ""
            ).strip(),
            mercado_category=str(
                request.args.get("mercado_category") or ""
            ).strip(),
            sales_sort=str(request.args.get("sales_sort") or "desc").strip(),
            current_only=str(request.args.get("current_only") or "1").strip().lower()
            not in ("0", "false", "no", "off"),
            page=_parse_int_param(request.args, "page", 1, 1, 1000000),
            page_size=1000,
        )
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    return jsonify({"status": "success", "data": data})


@app.route('/api/db/store-links/bulk-update', methods=['POST'])
@internal_api_required
def api_db_bulk_update_store_links():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    link_ids = data.get("link_ids") or []
    changes = data.get("changes") or {}
    if not isinstance(link_ids, list) or not isinstance(changes, dict):
        return jsonify({"status": "error", "message": "link_ids must be an array and changes an object"}), 422
    try:
        started, state = bit_store_link_remote_update.start_store_link_remote_update(
            link_ids, changes
        )
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    return jsonify({"status": "success", "data": {"started": started, "state": state}})


@app.route('/api/db/store-links/bulk-update/status', methods=['GET'])
@internal_api_required
def api_db_store_link_remote_update_status():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    return jsonify({
        "status": "success",
        "data": bit_store_link_remote_update.store_link_remote_update_status(),
    })


@app.route('/api/db/store-links/sync/start', methods=['POST'])
@internal_api_required
def api_db_start_store_link_sync():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    sync_all = data.get("sync_all") is True
    token_ids = [] if sync_all else (data.get("token_ids") or [])
    if not isinstance(token_ids, list):
        return jsonify({"status": "error", "message": "token_ids must be an array"}), 422
    try:
        started, state = bit_store_link_sync.start_store_link_sync(token_ids)
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    return jsonify({"status": "success", "data": {"started": started, "state": state}})


@app.route('/api/db/store-links/sync/status', methods=['GET'])
@internal_api_required
def api_db_store_link_sync_status():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    return jsonify({"status": "success", "data": bit_store_link_sync.store_link_sync_status()})


@app.route('/api/db/prohibited-listings', methods=['GET'])
@internal_api_required
def api_db_prohibited_listings():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    from erp.mercadolibre_prohibited_store import list_prohibited_listings

    try:
        token_text = str(request.args.get("token_id") or "").strip()
        data = list_prohibited_listings(
            search=str(request.args.get("search") or "").strip(),
            token_id=int(token_text) if token_text else None,
            site_id=str(request.args.get("site_id") or "").strip(),
            salesperson=str(request.args.get("salesperson") or "").strip(),
            risk_type=str(request.args.get("risk_type") or "").strip(),
            page=_parse_int_param(request.args, "page", 1, 1, 1000000),
            page_size=_parse_int_param(request.args, "page_size", 100, 20, 500),
        )
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    return jsonify({"status": "success", "data": data})


@app.route('/api/db/prohibited-listings/sync/start', methods=['POST'])
@internal_api_required
def api_db_start_prohibited_listing_sync():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    sync_all = data.get("sync_all") is True
    token_ids = [] if sync_all else (data.get("token_ids") or [])
    if not isinstance(token_ids, list):
        return jsonify({"status": "error", "message": "token_ids must be an array"}), 422
    try:
        started, state = bit_prohibited_listing_sync.start_prohibited_listing_sync(token_ids)
        rights_started, rights_state = mercado_infraction_sync.start_official_infraction_sync(token_ids)
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    combined_state = dict(state or {})
    combined_state["prohibited_running"] = bool(combined_state.get("running"))
    combined_state["rights_holder_sync"] = dict(rights_state or {})
    combined_state["running"] = bool(
        combined_state.get("prohibited_running") or rights_state.get("running")
    )
    return jsonify({
        "status": "success",
        "data": {"started": bool(started or rights_started), "state": combined_state},
    })


@app.route('/api/db/prohibited-listings/sync/status', methods=['GET'])
@internal_api_required
def api_db_prohibited_listing_sync_status():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    state = dict(bit_prohibited_listing_sync.prohibited_listing_sync_status() or {})
    rights_state = dict(mercado_infraction_sync.official_infraction_sync_status() or {})
    state["prohibited_running"] = bool(state.get("running"))
    state["rights_holder_sync"] = rights_state
    state["running"] = bool(state.get("prohibited_running") or rights_state.get("running"))
    return jsonify({"status": "success", "data": state})


@app.route('/api/db/orders/bulk-update', methods=['POST'])
@internal_api_required
def api_db_bulk_update_orders():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    order_ids = data.get("order_ids") or []
    if not isinstance(order_ids, list):
        return jsonify({"status": "error", "message": "order_ids must be an array"}), 422
    changes = {}
    for field in (
        "workflow_status", "purchase_order", "purchase_tracking",
        "logistics_company", "purchase_cost", "purchase_remark",
    ):
        if field in data:
            changes[field] = data.get(field)
    try:
        result = bit_order_sync.bit_mysql.bulk_update_mercado_orders(
            order_ids,
            operator_id=data.get("operator_id"),
            operator_name=data.get("operator_name") or "",
            **changes,
        )
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    return jsonify({"status": "success", "data": result})


@app.route('/api/db/orders/labels', methods=['POST'])
@internal_api_required
def api_db_order_labels():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    order_ids = data.get("order_ids") or []
    if not isinstance(order_ids, list):
        return jsonify({"status": "error", "message": "order_ids must be an array"}), 422
    try:
        from bit.bit_order_labels import download_order_labels

        result = download_order_labels(order_ids)
    except (ValueError, RuntimeError) as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    response = send_file(
        BytesIO(result["content"]),
        mimetype="application/pdf",
        as_attachment=False,
        download_name=result.get("filename") or "mercado-labels.pdf",
        max_age=0,
    )
    response.headers["X-Mercado-Label-Filename"] = result.get("filename") or "mercado-labels.pdf"
    response.headers["X-Mercado-Shipment-Count"] = str(result.get("shipment_count") or 0)
    successful_order_ids = [str(value) for value in result.get("order_ids") or []]
    response.headers["X-Mercado-Printed-Order-Ids"] = ",".join(successful_order_ids)
    response.headers["X-Mercado-Printed-Order-Count"] = str(len(successful_order_ids))
    response.headers["X-Mercado-Skipped-Order-Count"] = str(
        len(result.get("skipped_order_ids") or [])
    )
    response.headers["X-Mercado-Failed-Order-Count"] = str(
        len(result.get("failed_order_ids") or [])
    )
    response.headers["Cache-Control"] = "no-store"
    return response


@app.route('/api/db/orders/print-logs', methods=['POST'])
@internal_api_required
def api_db_order_print_logs():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    order_ids = data.get("order_ids") or []
    if not isinstance(order_ids, list):
        return jsonify({"status": "error", "message": "order_ids must be an array"}), 422
    count = bit_order_sync.bit_mysql.record_mercado_order_print_logs(
        order_ids,
        operator_id=data.get("operator_id"),
        operator_name=data.get("operator_name") or "",
    )
    return jsonify({"status": "success", "data": {"count": count}})


@app.route('/api/db/orders/<order_id>/logs', methods=['GET'])
@internal_api_required
def api_db_order_operation_logs(order_id):
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    rows = bit_order_sync.bit_mysql.list_mercado_order_operation_logs(
        order_id,
        limit=_parse_int_param(request.args, "limit", 100, 1, 200),
    )
    return jsonify({"status": "success", "data": rows})


@app.route('/api/db/orders/<order_id>/tracking', methods=['GET'])
@internal_api_required
def api_db_order_tracking(order_id):
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    from bit.bit_logistics import query_order_tracking

    try:
        data = query_order_tracking(order_id)
    except (KeyError, ValueError) as exc:
        return jsonify({"status": "error", "message": str(exc)}), 404
    return jsonify({"status": "success", "data": data})


@app.route('/api/db/orders/bulk', methods=['POST'])
@internal_api_required
def api_db_insert_orders():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    rows = data.get("rows") or []
    db_insert_orders(rows)
    return jsonify({"status": "success", "data": {"count": len(rows)}})


@app.route('/api/db/orders/high-after-sales', methods=['GET'])
@internal_api_required
def api_db_high_after_sale_alerts():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    try:
        data = db_get_high_after_sale_alerts(
            **_high_after_sale_query_params(request.args)
        )
        return jsonify({"status": "success", "data": data})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400


@app.route('/api/db/orders/high-profits', methods=['GET'])
@internal_api_required
def api_db_high_profit_products():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    try:
        data = db_get_high_profit_products(
            **_high_profit_query_params(request.args)
        )
        return jsonify({"status": "success", "data": data})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400


@app.route('/api/db/chat', methods=['POST'])
@internal_api_required
def api_db_insert_chat():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    required_fields = ["name", "site", "message", "chat", "response", "time"]
    missing_fields = [field for field in required_fields if data.get(field) in (None, "")]
    if missing_fields:
        return jsonify({"status": "error", "message": "Missing required fields", "fields": missing_fields}), 422
    chat_id = db_insert_chat_info(
        data["name"],
        data["site"],
        data["message"],
        data["chat"],
        data["response"],
        data["time"],
    )
    return jsonify({"status": "success", "data": {"id": chat_id}})


@app.route('/api/db/appeal-chat-records', methods=['POST'])
@internal_api_required
def api_db_insert_appeal_chat_record():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    record = data.get("record") or {}
    if not isinstance(record, dict):
        return jsonify({"status": "error", "message": "record must be an object"}), 422
    record_id = db_insert_appeal_chat_record(record)
    return jsonify({"status": "success", "data": {"id": record_id}})


@app.route('/api/db/ai-appeal-records', methods=['POST'])
@internal_api_required
def api_db_insert_ai_appeal_record():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    record = data.get("record") or {}
    if not isinstance(record, dict):
        return jsonify({"status": "error", "message": "record must be an object"}), 422
    record_id = db_insert_ai_appeal_record(record)
    return jsonify({"status": "success", "data": {"id": record_id}})


@app.route('/api/db/ai-appeal-records', methods=['GET'])
@internal_api_required
def api_db_get_ai_appeal_records():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    limit = request.args.get("limit", 100)
    return jsonify({"status": "success", "data": db_get_ai_appeal_records(limit)})


@app.route('/api/db/appeal-phrases', methods=['GET', 'POST'])
@internal_api_required
def api_db_appeal_phrases():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    try:
        if request.method == "POST":
            result = db_create_appeal_phrase(request.get_json(silent=True) or {})
        else:
            result = db_list_appeal_phrases()
        return jsonify({"status": "success", "data": result})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400


@app.route('/api/db/appeal-phrases/random', methods=['GET'])
@internal_api_required
def api_db_random_appeal_phrase():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    try:
        result = db_get_random_appeal_phrase(request.args.get("appeal_type", ""))
        return jsonify({"status": "success", "data": result})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400


@app.route('/api/db/appeal-phrases/<int:phrase_id>', methods=['PUT', 'DELETE'])
@internal_api_required
def api_db_appeal_phrase_detail(phrase_id):
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    try:
        if request.method == "PUT":
            result = db_update_appeal_phrase(
                phrase_id,
                request.get_json(silent=True) or {},
            )
        else:
            result = db_delete_appeal_phrase(phrase_id)
        return jsonify({"status": "success", "data": result})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400


@app.route('/api/db/infringement-knowledge', methods=['GET', 'POST'])
@internal_api_required
def api_db_infringement_knowledge():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    try:
        if request.method == "POST":
            result = db_create_infringement_knowledge(
                request.get_json(silent=True) or {}
            )
        else:
            result = db_list_infringement_knowledge(
                list_type=request.args.get("list_type", ""),
                search=request.args.get("search", ""),
                limit=request.args.get("limit", 2000),
            )
        return jsonify({"status": "success", "data": result})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400


@app.route('/api/db/infringement-knowledge/bulk', methods=['POST'])
@internal_api_required
def api_db_bulk_infringement_knowledge():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    try:
        data = request.get_json(silent=True) or {}
        result = db_bulk_create_infringement_knowledge(data.get("records") or [])
        return jsonify({"status": "success", "data": result})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400


@app.route('/api/db/infringement-knowledge/analysis-sources', methods=['GET'])
@internal_api_required
def api_db_infringement_knowledge_analysis_sources():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    try:
        result = db_get_infringement_knowledge_analysis_sources(
            infraction_limit=request.args.get("infraction_limit", 10000),
            active_limit=request.args.get("active_limit", 5000),
        )
        return jsonify({"status": "success", "data": result})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400


@app.route('/api/db/infringement-knowledge/analyzed', methods=['POST'])
@internal_api_required
def api_db_upsert_analyzed_infringement_knowledge():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    try:
        data = request.get_json(silent=True) or {}
        result = db_upsert_analyzed_infringement_knowledge(data.get("records") or [])
        return jsonify({"status": "success", "data": result})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400


@app.route('/api/db/infringement-knowledge/<int:record_id>', methods=['PUT', 'DELETE'])
@internal_api_required
def api_db_infringement_knowledge_detail(record_id):
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    try:
        if request.method == "PUT":
            result = db_update_infringement_knowledge(
                record_id,
                request.get_json(silent=True) or {},
            )
        else:
            result = db_delete_infringement_knowledge(record_id)
        return jsonify({"status": "success", "data": result})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400


@app.route('/api/db/infractions/latest', methods=['GET'])
@internal_api_required
def api_db_latest_infractions():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    recent_days = request.args.get("days", 30)
    return jsonify({"status": "success", "data": db_get_latest_infraction_info(recent_days)})


@app.route('/api/db/reputation/latest', methods=['GET'])
@internal_api_required
def api_db_latest_reputation():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    return jsonify({"status": "success", "data": db_get_latest_reputation_info()})


@app.route('/api/db/pago/latest', methods=['GET'])
@internal_api_required
def api_db_latest_pago():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    salesperson = str(request.args.get("salesperson") or "").strip()
    return jsonify({
        "status": "success",
        "data": db_get_latest_pago_info(salesperson),
    })


@app.route('/api/db/window-anomalies', methods=['POST'])
@internal_api_required
def api_db_upsert_window_anomaly():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    if not str(data.get("window_id") or "").strip():
        return jsonify({"status": "error", "message": "Missing window_id"}), 422
    db_upsert_window_anomaly(
        data.get("window_id"),
        data.get("window_name", ""),
        data.get("site", ""),
        data.get("anomaly_type", "需要登录"),
        data.get("reason", ""),
        data.get("source", "bit_daily_task"),
    )
    return jsonify({"status": "success", "data": {"window_id": data.get("window_id")}})


@app.route('/api/db/window-anomalies', methods=['GET'])
@internal_api_required
def api_db_get_window_anomalies():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    active_only = str(request.args.get("active_only", "1")).strip().lower() not in ("0", "false", "no")
    limit = request.args.get("limit", 500)
    return jsonify({"status": "success", "data": db_get_window_anomalies(active_only, limit)})


@app.route('/api/db/window-anomalies/resolve', methods=['POST'])
@internal_api_required
def api_db_resolve_window_anomaly():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    affected = db_resolve_window_anomaly(data.get("window_id"))
    return jsonify({"status": "success", "data": {"affected": affected}})


@app.route('/api/ai-appeal-records', methods=['GET'])
@login_required
def api_ai_appeal_records():
    try:
        limit = request.args.get("limit", 100)
        return jsonify({"status": "success", "data": db_get_ai_appeal_records(limit)})
    except Exception as e:
        logging.error("AI appeal records query failed: %s", e)
        return jsonify({"status": "error", "message": f"Database error: {str(e)}"}), 500


@app.route('/api/window-anomalies', methods=['GET'])
@login_required
def api_window_anomalies():
    try:
        active_only = str(request.args.get("active_only", "1")).strip().lower() not in ("0", "false", "no")
        limit = request.args.get("limit", 500)
        anomaly_data = filter_shop_status_anomalies(
            db_get_window_anomalies(active_only, limit)
        )
        response = jsonify({
            "status": "success",
            "data": enrich_window_anomaly_salespersons(anomaly_data),
        })
        response.headers["Cache-Control"] = "no-store"
        return response
    except Exception as e:
        logging.error("Window anomalies query failed: %s", e)
        return jsonify({"status": "error", "message": f"Database error: {str(e)}"}), 500


def filter_shop_status_anomalies(anomaly_data):
    """店铺状态展示美客多退出登录和需要人工处理的人机验证。"""
    data = dict(anomaly_data or {})
    rows = [
        dict(row)
        for row in (data.get("rows") or [])
        if is_shop_status_anomaly(row)
    ]
    data["rows"] = rows
    data["total"] = len(rows)
    return data


def filter_human_verification_anomalies(anomaly_data):
    """兼容旧调用名；店铺状态现同时包含退出登录。"""
    return filter_shop_status_anomalies(anomaly_data)


def enrich_window_anomaly_salespersons(anomaly_data):
    """按窗口和店铺配置补充归属人与邮箱，不修改历史异常记录。"""
    data = dict(anomaly_data or {})
    rows = [dict(row) for row in (data.get("rows") or [])]
    data["rows"] = rows
    if not rows:
        return data

    try:
        configs = list_shop_configs(include_ignored=True) or []
    except Exception as exc:
        logging.warning("读取店铺归属人和邮箱失败：%s", exc)
        return data

    exact_owners = {}
    window_owners = {}
    exact_emails = {}
    window_emails = {}

    for config in configs:
        window_id = str(config.get("window_id") or "").strip()
        shop_name = str(config.get("shop_name") or "").strip()
        salesperson = str(
            config.get("salesperson") or config.get("业务员") or ""
        ).strip()
        email = str(config.get("email") or config.get("邮箱") or "").strip()
        if not window_id:
            continue
        if salesperson:
            exact_owners.setdefault((window_id, shop_name), salesperson)
            window_owners.setdefault(window_id, salesperson)
        if email:
            exact_emails.setdefault((window_id, shop_name), email)
            window_emails.setdefault(window_id, email)

    for row in rows:
        window_id = str(row.get("window_id") or "").strip()
        shop_name = str(row.get("window_name") or "").strip()
        row["salesperson"] = str(
            row.get("salesperson")
            or row.get("业务员")
            or exact_owners.get((window_id, shop_name))
            or window_owners.get(window_id)
            or ""
        ).strip()
        row["email"] = str(
            row.get("email")
            or row.get("邮箱")
            or exact_emails.get((window_id, shop_name))
            or window_emails.get(window_id)
            or ""
        ).strip()
    return data


@app.route('/api/window-anomalies/<window_id>/resolve', methods=['POST'])
@login_required
def api_resolve_window_anomaly(window_id):
    try:
        stopped_task_ids = request_mercado_login_window_tasks_stop(window_id)
        affected = db_resolve_window_anomaly(window_id)
        return jsonify(
            {
                "status": "success",
                "data": {
                    "affected": affected,
                    "stopped_count": len(stopped_task_ids),
                    "stopped_task_ids": stopped_task_ids,
                },
            }
        )
    except Exception as e:
        logging.error("Resolve window anomaly failed: %s", e)
        return jsonify({"status": "error", "message": f"Database error: {str(e)}"}), 500


@app.route('/api/window-anomalies/mercado-login/status', methods=['GET'])
@login_required
def api_mercado_login_console_status():
    response = jsonify(
        {"status": "success", "data": _mercado_login_task_snapshot()}
    )
    response.headers["Cache-Control"] = "no-store"
    return response


@app.route('/api/window-anomalies/mercado-login/stop', methods=['POST'])
@login_required
def api_stop_mercado_login_console():
    data = request.get_json(silent=True) or {}
    task_id = normalize_appeal_task_id(data.get("task_id"))
    if not task_id:
        return jsonify({"status": "error", "message": "缺少有效任务编号"}), 400
    stopped, task_state = request_mercado_login_task_stop(task_id)
    if not stopped:
        return jsonify(
            {
                "status": "error",
                "data": task_state,
                "message": "登录任务已结束或不存在",
            }
        ), 404
    return jsonify(
        {
            "status": "success",
            "data": task_state,
            "message": "已提交停止请求，正在关闭任务进程和浏览器窗口",
        }
    )


@app.route('/api/window-anomalies/mercado-login/start', methods=['POST'])
@login_required
def api_start_mercado_login_console():
    data = request.get_json(silent=True) or {}
    worker_count = _parse_int_param(
        data,
        "workers",
        DEFAULT_MERCADO_LOGIN_WORKERS,
        1,
        MAX_MERCADO_LOGIN_WORKERS,
    )
    window_id = str(data.get("window_id") or "").strip()
    shop_name = ""
    selected_shops = []
    if "window_ids" in data:
        raw_window_ids = data.get("window_ids")
        if not isinstance(raw_window_ids, list):
            return jsonify(
                {"status": "error", "message": "window_ids 必须是窗口 ID 数组"}
            ), 400
        window_ids = list(
            dict.fromkeys(
                str(item or "").strip()
                for item in raw_window_ids
                if str(item or "").strip()
            )
        )
        if not window_ids:
            return jsonify(
                {"status": "error", "message": "请至少选择一个待处理登录异常店铺"}
            ), 400
        if len(window_ids) > 500:
            return jsonify(
                {"status": "error", "message": "单次最多选择 500 个待处理登录异常店铺"}
            ), 400
        try:
            anomaly_data = filter_shop_status_anomalies(
                db_get_window_anomalies(active_only=True, limit=1000) or {}
            )
        except Exception as exc:
            return jsonify(
                {"status": "error", "message": f"读取店铺状态失败：{exc}"}
            ), 500
        anomaly_by_window_id = {
            str(row.get("window_id") or "").strip(): row
            for row in (anomaly_data.get("rows") or [])
            if str(row.get("window_id") or "").strip()
        }
        missing_window_ids = [
            item for item in window_ids if item not in anomaly_by_window_id
        ]
        if missing_window_ids:
            return jsonify(
                {
                    "status": "error",
                    "message": "部分店铺状态不存在或已恢复，请刷新后重试："
                    + "、".join(missing_window_ids),
                }
            ), 404
        selected_shops = [
            {
                "window_id": item,
                "window_name": str(
                    anomaly_by_window_id[item].get("window_name") or ""
                ).strip(),
            }
            for item in window_ids
        ]
        if any(not shop["window_name"] for shop in selected_shops):
            return jsonify(
                {"status": "error", "message": "部分店铺状态记录缺少店铺名"}
            ), 422
    elif window_id:
        try:
            anomaly_data = filter_shop_status_anomalies(
                db_get_window_anomalies(active_only=True, limit=1000) or {}
            )
            anomaly = next(
                (
                    row
                    for row in (anomaly_data.get("rows") or [])
                    if str(row.get("window_id") or "").strip() == window_id
                ),
                None,
            )
        except Exception as exc:
            return jsonify(
                {"status": "error", "message": f"读取店铺状态失败：{exc}"}
            ), 500
        if anomaly is None:
            return jsonify({"status": "error", "message": "店铺状态不存在或已恢复"}), 404
        shop_name = str(anomaly.get("window_name") or "").strip()
        if not shop_name:
            return jsonify({"status": "error", "message": "店铺状态记录缺少店铺名"}), 422

    if selected_shops:
        started, task_state = start_mercado_login_console_job(
            selected_shops=selected_shops,
            workers=worker_count,
        )
        if not started:
            return jsonify(
                {
                    "status": "running",
                    "data": task_state,
                    "message": task_state.get("message")
                    or "所选店铺已有自动登录任务正在运行",
                }
            ), 409
        started_task_id = str(task_state.get("started_task_id") or "").strip()
        actual_worker_count = min(worker_count, len(selected_shops))
        task_state.update(
            {
                "started_count": len(selected_shops),
                "started_task_ids": [started_task_id] if started_task_id else [],
                "failed_count": 0,
                "failed_shops": [],
                "workers": actual_worker_count,
            }
        )
        return jsonify(
            {
                "status": "success",
                "data": task_state,
                "message": (
                    f"已启动 {len(selected_shops)} 家店铺的重新检测，"
                    f"并发进程数：{actual_worker_count}"
                ),
            }
        )

    job_args = {
        "shop_name": shop_name,
        "window_id": window_id,
        "workers": worker_count,
    }
    started, task_state = start_mercado_login_console_job(**job_args)
    if not started:
        return jsonify(
            {
                "status": "running",
                "data": task_state,
                "message": task_state.get("message")
                or "所选店铺已有自动登录任务正在运行",
            }
        ), 409
    return jsonify(
        {
            "status": "success",
            "data": task_state,
            "message": f"{task_state['target']} 登录任务已启动",
        }
    )


@app.route('/api/db/workbench/session-user', methods=['GET'])
@internal_api_required
def api_db_workbench_session_user():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    user_id = request.args.get("user_id", type=int)
    if not user_id:
        return jsonify({"status": "error", "message": "缺少用户 ID"}), 400
    row = get_workbench_user(user_id=user_id)
    user = (
        build_workbench_session_user(row)
        if row and row.get("is_active")
        else None
    )
    return jsonify({"status": "success", "data": user})


@app.route('/api/db/workbench/roles', methods=['GET', 'POST'])
@internal_api_required
def api_db_workbench_roles():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    try:
        if request.method == "GET":
            data = list_workbench_roles_local()
        else:
            data = create_workbench_role_local(request.get_json(silent=True) or {})
        return jsonify({"status": "success", "data": data})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("工作台角色数据库接口失败")
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.route('/api/db/workbench/roles/<role_key>', methods=['PUT', 'DELETE'])
@internal_api_required
def api_db_workbench_role_detail(role_key):
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    try:
        if request.method == "PUT":
            update_workbench_role_local(
                role_key,
                request.get_json(silent=True) or {},
            )
        else:
            delete_workbench_role_local(role_key)
        return jsonify({"status": "success"})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("工作台角色数据库接口失败")
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.route('/api/db/workbench/users', methods=['GET', 'POST'])
@internal_api_required
def api_db_workbench_users():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    try:
        if request.method == "GET":
            data = list_workbench_users_local()
        else:
            data = create_workbench_user_local(request.get_json(silent=True) or {})
        return jsonify({"status": "success", "data": data})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("工作台账号数据库接口失败")
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.route('/api/db/workbench/users/<int:user_id>', methods=['PUT'])
@internal_api_required
def api_db_workbench_user_detail(user_id):
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    try:
        update_workbench_user_local(user_id, request.get_json(silent=True) or {})
        return jsonify({"status": "success"})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("工作台账号数据库接口失败")
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.route('/api/db/workbench/users/<int:user_id>/password', methods=['POST'])
@internal_api_required
def api_db_workbench_user_password(user_id):
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    try:
        data = request.get_json(silent=True) or {}
        reset_workbench_user_password_local(user_id, data.get("password"))
        return jsonify({"status": "success"})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("工作台密码数据库接口失败")
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.route('/api/access/catalog', methods=['GET'])
@login_required
def api_access_catalog():
    return jsonify(
        {
            "status": "success",
            "data": {"groups": workbench_permission_catalog()},
        }
    )


@app.route('/api/access/roles', methods=['GET', 'POST'])
@login_required
def api_access_roles():
    try:
        if request.method == "GET":
            data = _workbench_backend("list_workbench_roles")
        else:
            data = _workbench_backend(
                "create_workbench_role",
                request.get_json(silent=True) or {},
            )
        return jsonify({"status": "success", "data": data})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("人员与权限角色接口失败")
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.route('/api/access/roles/<role_key>', methods=['PUT', 'DELETE'])
@login_required
def api_access_role_detail(role_key):
    try:
        if request.method == "PUT":
            _workbench_backend(
                "update_workbench_role",
                role_key,
                request.get_json(silent=True) or {},
            )
        else:
            _workbench_backend("delete_workbench_role", role_key)
        return jsonify({"status": "success"})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("人员与权限角色接口失败")
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.route('/api/access/users', methods=['GET', 'POST'])
@login_required
def api_access_users():
    try:
        if request.method == "GET":
            data = _workbench_backend("list_workbench_users")
        else:
            data = _workbench_backend(
                "create_workbench_user",
                request.get_json(silent=True) or {},
            )
        return jsonify({"status": "success", "data": data})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("人员与权限账号接口失败")
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.route('/api/access/users/<int:user_id>', methods=['PUT'])
@login_required
def api_access_user_detail(user_id):
    data = request.get_json(silent=True) or {}
    current_user = get_current_workbench_user() or {}
    if int(current_user.get("id") or 0) == user_id and not data.get("is_active", True):
        return jsonify({"status": "error", "message": "不能停用当前登录账号"}), 400
    try:
        _workbench_backend("update_workbench_user", user_id, data)
        # 当前账号资料或角色变化后立即刷新本会话。
        if int(current_user.get("id") or 0) == user_id:
            g.pop("workbench_user", None)
            refreshed = get_current_workbench_user()
            if refreshed:
                session["workbench_user"] = refreshed
        return jsonify({"status": "success"})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("人员与权限账号接口失败")
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.route('/api/access/users/<int:user_id>/password', methods=['POST'])
@login_required
def api_access_user_password(user_id):
    try:
        data = request.get_json(silent=True) or {}
        _workbench_backend(
            "reset_workbench_user_password",
            user_id,
            data.get("password"),
        )
        return jsonify({"status": "success"})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("人员与权限密码接口失败")
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.route('/api/db/workbench/login', methods=['POST'])
@internal_api_required
def api_db_workbench_login():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked
    data = request.get_json(silent=True) or {}
    username = str(data.get("username", "")).strip()
    password = str(data.get("password", ""))
    if not username or not password:
        return jsonify({"status": "error", "message": "请输入账号和密码"}), 400

    user = authenticate_workbench_user(username, password)
    if not user:
        return jsonify({"status": "error", "message": "账号或密码错误"}), 401
    return jsonify({"status": "success", "data": user})


@app.route('/api/mercado-tokens/authorization', methods=['GET'])
@login_required
def api_mercado_token_authorization():
    try:
        data = bit_db_api.get_mercado_token_authorization_info()
        return jsonify({"status": "success", "data": data})
    except Exception as exc:
        return _mercado_token_error_response(exc)


@app.route('/api/mercado-tokens', methods=['GET'])
@login_required
def api_list_mercado_tokens():
    try:
        data = bit_db_api.list_mercado_store_tokens()
        return jsonify({"status": "success", "data": data})
    except Exception as exc:
        return _mercado_token_error_response(exc)


@app.route('/api/mercado-tokens/<int:token_id>/site-settings', methods=['GET', 'PUT'])
@login_required
def api_mercado_token_site_settings(token_id):
    try:
        if request.method == "GET":
            result = bit_db_api.list_mercado_store_site_settings(token_id)
            message = ""
        else:
            data = request.get_json(silent=True) or {}
            result = bit_db_api.update_mercado_store_site_settings(
                token_id, data.get("settings", [])
            )
            message = "店铺配置已保存"
        return jsonify({"status": "success", "data": result, "message": message})
    except Exception as exc:
        return _mercado_token_error_response(exc)


@app.route('/api/mercado-tokens/exchange', methods=['POST'])
@login_required
def api_exchange_mercado_token():
    data = request.get_json(silent=True) or {}
    try:
        result = bit_db_api.exchange_mercado_store_token(
            data.get("display_name", ""), data.get("code", "")
        )
        response_data = dict(result or {})
        token_id = int(response_data.get("id") or 0)
        auto_sync = {"started": False, "queued": False}
        if token_id and response_data.get("enabled", True):
            try:
                sync_result = bit_db_api.start_store_link_sync([token_id]) or {}
                auto_sync = {
                    "started": bool(sync_result.get("started")),
                    "queued": True,
                    "task_id": str((sync_result.get("state") or {}).get("task_id") or ""),
                }
            except Exception as sync_exc:
                logging.exception("店铺授权成功，但自动拉取店铺链接启动失败")
                auto_sync["error"] = str(sync_exc)
        response_data["auto_link_sync"] = auto_sync
        message = (
            "店铺授权成功，正在自动拉取全部链接"
            if auto_sync.get("started")
            else "店铺授权成功，全部链接已加入自动拉取队列"
        )
        if auto_sync.get("error"):
            message = "店铺授权成功；自动拉取暂未启动，系统会在下次周期检查时重试"
        return jsonify({
            "status": "success",
            "data": response_data,
            "message": message,
        })
    except Exception as exc:
        return _mercado_token_error_response(exc)


@app.route('/api/mercado-tokens/<int:token_id>/refresh', methods=['POST'])
@login_required
def api_refresh_mercado_token(token_id):
    try:
        result = bit_db_api.refresh_mercado_store_token(token_id)
        return jsonify({
            "status": "success",
            "data": result,
            "message": "Token 已刷新并保存",
        })
    except Exception as exc:
        return _mercado_token_error_response(exc)


def _append_api_reputation_log(message):
    line = f"[{datetime.now().strftime('%H:%M:%S')}] {str(message or '').strip()}"
    with _api_reputation_lock:
        _api_reputation_logs.append(line)


def _next_api_reputation_run(now=None):
    """Return the next local 00:00/12:00 reputation refresh boundary."""

    current = now or datetime.now()
    candidates = []
    for hour in API_REPUTATION_AUTO_REFRESH_HOURS:
        candidate = current.replace(hour=hour, minute=0, second=0, microsecond=0)
        if candidate <= current:
            candidate += timedelta(days=1)
        candidates.append(candidate)
    return min(candidates)


def _api_reputation_snapshot_unlocked():
    """Copy mutable state while the caller already owns the state lock."""

    data = dict(_api_reputation_state)
    data["rows"] = [dict(row) for row in _api_reputation_state.get("rows", [])]
    data["failures"] = [
        dict(row) for row in _api_reputation_state.get("failures", [])
    ]
    data["logs"] = list(_api_reputation_logs)
    return data


def _api_reputation_snapshot():
    with _api_reputation_lock:
        data = _api_reputation_snapshot_unlocked()
    if data.get("running"):
        data["elapsed_seconds"] = _mercado_collection_elapsed_seconds(data)
    data.update({
        "auto_refresh_enabled": not USE_DB_API,
        "auto_refresh_hours": list(API_REPUTATION_AUTO_REFRESH_HOURS),
        "max_workers": API_REPUTATION_MAX_WORKERS,
        "next_auto_refresh_at": _next_api_reputation_run().strftime(
            "%Y-%m-%d %H:%M:%S"
        ),
    })
    return data


def _legacy_reputation_percentage(value):
    match = re.search(r"-?\d+(?:\.\d+)?", str(value or "").replace(",", ""))
    return float(match.group(0)) if match else None


def _hydrate_api_reputation_from_database():
    global _api_reputation_database_hydration_attempted
    with _api_reputation_lock:
        if (
            _api_reputation_database_hydration_attempted
            or _api_reputation_state.get("running")
            or _api_reputation_state.get("rows")
        ):
            return False
        _api_reputation_database_hydration_attempted = True

    try:
        latest = db_get_latest_reputation_info() or {}
        _attach_reputation_token_ids(latest)
        _attach_latest_reputation_infraction_counts(latest)
    except Exception as exc:
        logging.warning("API 声誉读取上一次入库数据失败：%s", exc)
        return False

    database_rows = latest.get("rows") or []
    rows = []
    for source in database_rows:
        if not isinstance(source, dict):
            continue
        hydrated_row = {
            "store_name": source.get("店铺名") or "",
            "site_name": source.get("站点") or "",
            "level_name": source.get("声誉颜色") or "",
            "sales_completed": source.get("总单量"),
            "claims_rate_percent": _legacy_reputation_percentage(source.get("投诉率")),
            "delayed_handling_rate_percent": _legacy_reputation_percentage(source.get("延误率")),
            "cancellations_rate_percent": _legacy_reputation_percentage(source.get("取消率")),
            "infraction_count": int(_legacy_reputation_percentage(source.get("侵权数量")) or 0),
            "rights_holder_count": int(_legacy_reputation_percentage(source.get("权利人数量")) or 0),
        }
        if "站点状态" in source:
            hydrated_row.update({
                "direction": source.get("增加或减少") or "",
                "gradient_rate": source.get("近七天变化率") or "",
                "site_status_display": source.get("站点状态") or "未知",
                "token_id": int(source.get("token_id") or 0),
            })
        rows.append(hydrated_row)
    if not rows:
        return False

    store_count = len({str(row.get("store_name") or "") for row in rows})
    finished_at = str(latest.get("latest_submit_time") or "")
    with _api_reputation_lock:
        if _api_reputation_state.get("running") or _api_reputation_state.get("rows"):
            return False
        _api_reputation_state.update({
            "running": False,
            "status": "success",
            "message": "已展示上一次入库的 API 声誉数据",
            "started_at": "",
            "finished_at": finished_at,
            "elapsed_seconds": 0,
            "total_stores": store_count,
            "completed_stores": store_count,
            "success_stores": store_count,
            "failed_stores": 0,
            "total_sites": len(rows),
            "rows": rows,
            "failures": [],
        })
    return True


def _persist_api_reputation_snapshot(state_path=None):
    data = _api_reputation_snapshot()
    if data.get("status") not in {"success", "partial"} or not data.get("rows"):
        return False

    data["running"] = False
    path = Path(state_path or API_REPUTATION_STATE_PATH)
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = path.with_suffix(
        f"{path.suffix}.{os.getpid()}.{threading.get_ident()}.tmp"
    )
    try:
        temporary_path.write_text(
            json.dumps(data, ensure_ascii=False, indent=2, default=str),
            encoding="utf-8",
        )
        os.replace(temporary_path, path)
        return True
    except OSError as exc:
        logging.warning("无法保存 API 声誉上次结果：%s", exc)
        return False
    finally:
        try:
            temporary_path.unlink()
        except OSError:
            pass


def _run_all_api_reputation_refresh():
    started_monotonic = time.monotonic()

    def update_progress(progress):
        event = str((progress or {}).get("event") or "")
        with _api_reputation_lock:
            if event == "initialized":
                total_stores = int(progress.get("total_stores") or 0)
                _api_reputation_state["total_stores"] = total_stores
                _api_reputation_state["message"] = (
                    f"正在更新 {total_stores} 家已开启声誉更新的店铺"
                    if total_stores
                    else "没有开启声誉更新的店铺"
                )
            elif event == "store_success":
                rows = [dict(row) for row in (progress.get("rows") or [])]
                _api_reputation_state["completed_stores"] += 1
                _api_reputation_state["success_stores"] += 1
                _api_reputation_state["total_sites"] += len(rows)
                _api_reputation_state["rows"].extend(rows)
            elif event == "store_failure":
                _api_reputation_state["completed_stores"] += 1
                _api_reputation_state["failed_stores"] += 1
                _api_reputation_state["failures"].append({
                    "token_id": int(progress.get("token_id") or 0),
                    "store_name": str(progress.get("store_name") or ""),
                    "error": str(progress.get("error") or ""),
                })

    try:
        result = bit_reputation_info.main(
            max_workers=API_REPUTATION_MAX_WORKERS,
            retry_failed=True,
            send_email=False,
            export_excel=False,
            # 流量只能从后台页面读取；其余声誉、订单趋势、站点状态等
            # 均继续通过 Mercado Libre 官方 API 刷新。
            collect_browser_auxiliary=True,
            log_callback=_append_api_reputation_log,
            progress_callback=update_progress,
        ) or {}
        finished_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        rows = [dict(row) for row in (result.get("api_rows") or [])]
        failures = [dict(row) for row in (result.get("failures") or [])]
        total_stores = int(result.get("total_stores") or 0)
        completed_stores = int(result.get("completed_stores") or total_stores)
        success_count = int(result.get("success_stores") or 0)
        failed_count = int(result.get("failed_stores") or len(failures))
        total_sites = int(result.get("total_sites") or len(rows))
        with _api_reputation_lock:
            _api_reputation_state.update({
                "running": False,
                "status": (
                    "success" if failed_count == 0
                    else "partial" if success_count else "error"
                ),
                "message": (
                    f"全量更新完成：成功 {success_count} 家，失败 {failed_count} 家"
                    if total_stores
                    else "没有开启声誉更新的店铺"
                ),
                "finished_at": finished_at,
                "elapsed_seconds": max(0, int(time.monotonic() - started_monotonic)),
                "total_stores": total_stores,
                "completed_stores": completed_stores,
                "success_stores": success_count,
                "failed_stores": failed_count,
                "total_sites": total_sites,
                "rows": rows,
                "failures": failures,
            })
        _persist_api_reputation_snapshot()
    except Exception as exc:
        logging.exception("API 声誉全量更新失败")
        with _api_reputation_lock:
            _api_reputation_state.update({
                "running": False,
                "status": "error",
                "message": f"全量更新失败：{exc}",
                "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "elapsed_seconds": max(0, int(time.monotonic() - started_monotonic)),
            })
        _append_api_reputation_log(f"任务异常终止：{exc}")


def _start_api_reputation_refresh(*, automatic=False):
    with _api_reputation_lock:
        if _api_reputation_state.get("running"):
            return False
        _api_reputation_logs.clear()
        _api_reputation_state.update({
            "running": True,
            "status": "running",
            "message": (
                "定时任务正在读取授权店铺"
                if automatic
                else "正在读取授权店铺"
            ),
            "started_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "finished_at": "",
            "elapsed_seconds": 0,
            "total_stores": 0,
            "completed_stores": 0,
            "success_stores": 0,
            "failed_stores": 0,
            "total_sites": 0,
            "rows": [],
            "failures": [],
        })
    threading.Thread(
        target=_run_all_api_reputation_refresh,
        name=(
            "api-reputation-auto-refresh"
            if automatic
            else "api-reputation-refresh"
        ),
        daemon=True,
    ).start()
    if automatic:
        _append_api_reputation_log(
            f"定时刷新已启动：流量并发 {API_REPUTATION_MAX_WORKERS}，"
            "其余数据使用官方 API"
        )
    return True


def _api_reputation_auto_refresh_loop(stop_event=None):
    stop_event = stop_event or _api_reputation_scheduler_stop_event
    while not stop_event.is_set():
        now = datetime.now()
        next_run = _next_api_reputation_run(now)
        wait_seconds = max(0.0, (next_run - now).total_seconds())
        if stop_event.wait(wait_seconds):
            return
        if not _start_api_reputation_refresh(automatic=True):
            logging.warning("API 声誉定时刷新到点时已有任务运行，本轮跳过")


def start_api_reputation_scheduler_bootstrap():
    """Start the central 00:00/12:00 mixed reputation refresh scheduler."""

    global _api_reputation_scheduler_thread
    with _api_reputation_scheduler_guard:
        if (
            _api_reputation_scheduler_thread
            and _api_reputation_scheduler_thread.is_alive()
        ):
            return None
        _api_reputation_scheduler_stop_event.clear()
        _api_reputation_scheduler_thread = threading.Thread(
            target=_api_reputation_auto_refresh_loop,
            name="api-reputation-auto-scheduler",
            daemon=True,
        )
        _api_reputation_scheduler_thread.start()
        logging.info(
            "API 声誉自动刷新调度已启动：每天 00:00、12:00，流量并发 %s，"
            "其余数据使用官方 API",
            API_REPUTATION_MAX_WORKERS,
        )
        return _api_reputation_scheduler_thread


@app.route('/api/mercado-reputation/refresh', methods=['POST'])
@login_required
def api_refresh_all_mercado_reputation():
    if not _start_api_reputation_refresh():
        return jsonify({
            "status": "running",
            "data": _api_reputation_snapshot(),
            "message": "API 声誉全量更新任务正在运行",
        }), 409
    return jsonify({
        "status": "success",
        "data": _api_reputation_snapshot(),
        "message": "API 声誉全量更新已启动",
    })


@app.route('/api/mercado-reputation/status', methods=['GET'])
@login_required
def api_mercado_reputation_status():
    _hydrate_api_reputation_from_database()
    response = jsonify({"status": "success", "data": _api_reputation_snapshot()})
    response.headers["Cache-Control"] = "no-store"
    return response


@app.route('/api/mercado-tokens/<int:token_id>/reputation', methods=['GET'])
@login_required
def api_mercado_token_reputation(token_id):
    try:
        result = bit_db_api.get_mercado_store_reputation(token_id)
        response = jsonify({
            "status": "success",
            "data": result,
            "message": "已获取美客多官方声誉",
        })
        response.headers["Cache-Control"] = "no-store"
        return response
    except Exception as exc:
        return _mercado_token_error_response(exc)


def _query_pre_sale_store_pages(token_id, filters):
    rows = []
    official_total = 0
    offset = 0
    while offset <= 1000:
        page = bit_db_api.execute_mercado_store_communication(
            int(token_id),
            "pre-sale-list",
            {
                **filters,
                "sort_fields": "date_created",
                "sort_types": "DESC",
                "limit": 100,
                "offset": offset,
            },
        ) or {}
        questions = [
            dict(row) for row in (page.get("questions") or []) if isinstance(row, dict)
        ]
        rows.extend(questions)
        official_total = max(official_total, int(page.get("total") or 0))
        if not questions or len(rows) >= official_total:
            break
        offset += len(questions)
        if offset > 1000:
            break
    return {
        "token_id": int(token_id),
        "questions": rows,
        "official_total": official_total,
        "truncated": len(rows) < official_total,
    }


@app.route('/api/mercado-communications/pre-sale-aggregate', methods=['POST'])
@login_required
def api_mercado_pre_sale_aggregate():
    data = request.get_json(silent=True) or {}
    raw_token_ids = data.get("token_ids")
    if not isinstance(raw_token_ids, list):
        return jsonify({"status": "error", "message": "店铺 ID 必须是数组"}), 400
    try:
        token_ids = list(dict.fromkeys(
            int(value) for value in raw_token_ids if int(value) > 0
        ))
    except (TypeError, ValueError):
        return jsonify({"status": "error", "message": "店铺 ID 格式错误"}), 400
    if not token_ids:
        return jsonify({"status": "error", "message": "请选择要查询的店铺"}), 400
    if len(token_ids) > 100:
        return jsonify({"status": "error", "message": "单次最多并发查询 100 家店铺"}), 400

    filters = {
        key: data.get(key)
        for key in ("item_id", "user_id", "status")
        if str(data.get(key) or "").strip()
    }
    requested_workers = data.get("workers", 8)
    try:
        worker_count = max(1, min(12, int(requested_workers), len(token_ids)))
    except (TypeError, ValueError):
        return jsonify({"status": "error", "message": "并发线程数格式错误"}), 400

    started = time.perf_counter()
    rows = []
    failures = []
    with ThreadPoolExecutor(
        max_workers=worker_count,
        thread_name_prefix="pre-sale-query",
    ) as executor:
        future_tokens = {
            executor.submit(_query_pre_sale_store_pages, token_id, filters): token_id
            for token_id in token_ids
        }
        for future in as_completed(future_tokens):
            token_id = future_tokens[future]
            try:
                rows.append(future.result())
            except Exception as exc:
                failures.append({"token_id": token_id, "message": str(exc)})

    order = {token_id: index for index, token_id in enumerate(token_ids)}
    rows.sort(key=lambda row: order.get(int(row.get("token_id") or 0), len(order)))
    failures.sort(key=lambda row: order.get(int(row.get("token_id") or 0), len(order)))
    response = jsonify({
        "status": "success",
        "data": {
            "stores": rows,
            "failures": failures,
            "total_stores": len(token_ids),
            "success_stores": len(rows),
            "failed_stores": len(failures),
            "workers": worker_count,
            "elapsed_seconds": round(time.perf_counter() - started, 3),
        },
    })
    response.headers["Cache-Control"] = "no-store"
    return response


def _query_customer_service_claims_store(
    token_id, filters, required_rows, status_filter=""
):
    rows = []
    status_totals = {"opened": 0, "closed": 0}
    truncated = False
    marketplace_errors = []
    for claim_status in ("opened", "closed"):
        status_rows = []
        offset = 0
        while len(status_rows) < required_rows and offset <= 1000:
            limit = min(100, required_rows - len(status_rows))
            page = bit_db_api.execute_mercado_store_communication(
                int(token_id),
                "claims-list",
                {
                    **filters,
                    "status": claim_status,
                    "limit": limit,
                    "offset": offset,
                },
            ) or {}
            batch = [
                dict(row) for row in (page.get("data") or [])
                if isinstance(row, dict)
            ]
            for row in batch:
                row["token_id"] = int(token_id)
            status_rows.extend(batch)
            total = int((page.get("paging") or {}).get("total") or 0)
            status_totals[claim_status] = max(
                status_totals[claim_status], total, len(status_rows)
            )
            marketplace_errors.extend(
                dict(error) for error in (page.get("marketplace_errors") or [])
                if isinstance(error, dict)
            )
            if not batch or offset + len(batch) >= total:
                break
            offset += len(batch)
        if len(status_rows) < status_totals[claim_status] and offset > 1000:
            truncated = True
        if not status_filter or status_filter == claim_status:
            rows.extend(status_rows)

    deduplicated = list({
        str(row.get("id") or index): row for index, row in enumerate(rows)
    }.values())
    return {
        "token_id": int(token_id),
        "rows": deduplicated,
        "total": (
            status_totals.get(status_filter, 0)
            if status_filter
            else status_totals["opened"] + status_totals["closed"]
        ),
        "status_totals": status_totals,
        "marketplace_errors": marketplace_errors,
        "truncated": truncated,
    }


def _query_customer_service_post_sale_store(token_id, search=""):
    identifier = int(token_id)
    pack_search = str(search or "").strip()
    if pack_search:
        bundle = bit_db_api.execute_mercado_store_communication(
            identifier,
            "post-sale-messages",
            {"pack_id": pack_search, "limit": 100},
        ) or {}
        return {
            "token_id": identifier,
            "search_pack_id": pack_search,
            "search_bundle": bundle,
            "conversations": [],
        }

    unread = bit_db_api.execute_mercado_store_communication(
        identifier, "post-sale-unread", {}
    ) or {}
    conversations = []
    for row in unread.get("results") or []:
        if not isinstance(row, dict):
            continue
        pack_id = str(row.get("resource") or "").rstrip("/").split("/")[-1].strip()
        if not pack_id:
            continue
        conversation = {
            "token_id": identifier,
            "pack_id": pack_id,
            "unread_count": int(row.get("count") or 0),
            "resource": row.get("resource"),
            "order_context": dict(row.get("order_context") or {}),
        }
        try:
            conversation["bundle"] = (
                bit_db_api.execute_mercado_store_communication(
                    identifier,
                    "post-sale-messages",
                    {"pack_id": pack_id, "limit": 100},
                ) or {}
            )
        except Exception as exc:
            conversation["load_error"] = str(exc)
        conversations.append(conversation)
    return {
        "token_id": identifier,
        "unread_total": len(unread.get("results") or []),
        "conversations": conversations,
    }


@app.route(
    '/api/mercado-communications/customer-service-aggregate',
    methods=['POST'],
)
@login_required
def api_mercado_customer_service_aggregate():
    data = request.get_json(silent=True) or {}
    raw_token_ids = data.get("token_ids")
    if not isinstance(raw_token_ids, list):
        return jsonify({"status": "error", "message": "店铺 ID 必须是数组"}), 400
    try:
        token_ids = list(dict.fromkeys(
            int(value) for value in raw_token_ids if int(value) > 0
        ))
    except (TypeError, ValueError):
        return jsonify({"status": "error", "message": "店铺 ID 格式错误"}), 400
    if not token_ids:
        return jsonify({"status": "error", "message": "请选择要查询的店铺"}), 400
    if len(token_ids) > 100:
        return jsonify({"status": "error", "message": "单次最多并发查询 100 家店铺"}), 400

    mode = str(data.get("mode") or "").strip().lower()
    if mode not in ("claims", "post-sale"):
        return jsonify({"status": "error", "message": "查询类型格式错误"}), 400
    try:
        worker_count = max(
            1, min(12, int(data.get("workers", 8)), len(token_ids))
        )
    except (TypeError, ValueError):
        return jsonify({"status": "error", "message": "并发线程数格式错误"}), 400

    status_filter = str(data.get("status_filter") or "").strip().lower()
    if status_filter not in ("", "opened", "closed"):
        return jsonify({"status": "error", "message": "索赔状态格式错误"}), 400
    try:
        required_rows = max(1, min(1000, int(data.get("required_rows", 20))))
    except (TypeError, ValueError):
        return jsonify({"status": "error", "message": "查询数量格式错误"}), 400
    raw_filters = data.get("filters") or {}
    if not isinstance(raw_filters, dict):
        return jsonify({"status": "error", "message": "筛选条件格式错误"}), 400
    claim_filters = {
        key: raw_filters.get(key)
        for key in (
            "claim_type", "claim_id", "order_id", "pack_id",
            "date_from", "date_to",
        )
        if str(raw_filters.get(key) or "").strip()
    }
    search = str(data.get("search") or "").strip()

    def query_store(token_id):
        if mode == "claims":
            return _query_customer_service_claims_store(
                token_id, claim_filters, required_rows, status_filter
            )
        return _query_customer_service_post_sale_store(token_id, search)

    started = time.perf_counter()
    stores = []
    failures = []
    with ThreadPoolExecutor(
        max_workers=worker_count,
        thread_name_prefix="customer-service-query",
    ) as executor:
        future_tokens = {
            executor.submit(query_store, token_id): token_id
            for token_id in token_ids
        }
        for future in as_completed(future_tokens):
            token_id = future_tokens[future]
            try:
                stores.append(future.result())
            except Exception as exc:
                failures.append({"token_id": token_id, "message": str(exc)})

    order = {token_id: index for index, token_id in enumerate(token_ids)}
    stores.sort(key=lambda row: order.get(int(row.get("token_id") or 0), len(order)))
    failures.sort(key=lambda row: order.get(int(row.get("token_id") or 0), len(order)))
    response = jsonify({
        "status": "success",
        "data": {
            "mode": mode,
            "stores": stores,
            "failures": failures,
            "total_stores": len(token_ids),
            "success_stores": len(stores),
            "failed_stores": len(failures),
            "workers": worker_count,
            "elapsed_seconds": round(time.perf_counter() - started, 3),
        },
    })
    response.headers["Cache-Control"] = "no-store"
    return response


MERCADO_COMMUNICATION_READ_ACTIONS = frozenset((
    "pre-sale-list",
    "pre-sale-summary",
    "pre-sale-detail",
    "post-sale-unread",
    "post-sale-messages",
    "claims-list",
    "claims-detail",
))
MERCADO_COMMUNICATION_WRITE_ACTIONS = frozenset((
    "pre-sale-answer",
    "pre-sale-delete",
    "post-sale-send",
    "claims-send",
))
MERCADO_COMMUNICATION_VIEW_POST_ACTIONS = frozenset((
    "pre-sale-translate",
))


def _open_bitbrowser_page(open_result, target_url):
    target = str(target_url or "").strip()
    if not target:
        return None
    data = open_result.get("data") if isinstance(open_result, dict) else None
    debugger_address = str(
        data.get("http") if isinstance(data, dict) else ""
    ).strip()
    if not debugger_address:
        raise RuntimeError("比特浏览器未返回调试地址，无法打开指定页面")
    debugger_url = (
        debugger_address
        if debugger_address.startswith(("http://", "https://"))
        else f"http://{debugger_address}"
    )
    request_url = (
        f"{debugger_url.rstrip('/')}/json/new?{quote(target, safe='')}"
    )
    try:
        request = Request(request_url, method="PUT")
        with urlopen(request, timeout=10) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except Exception as exc:
        raise RuntimeError(f"比特浏览器已启动，但打开目标页面失败：{exc}") from exc
    if not isinstance(payload, dict) or not payload.get("webSocketDebuggerUrl"):
        raise RuntimeError("比特浏览器已启动，但目标页面未成功创建")
    return payload


def _open_mercado_claim_browser(
    token_id,
    claim_id="",
    shop_name_hint="",
    target_url="",
):
    """按 Mercado 授权店铺名称实时匹配并打开 BitBrowser 窗口。"""
    identifier = int(token_id)
    token_rows = list((bit_db_api.list_mercado_store_tokens() or {}).get("rows") or [])
    token = next(
        (row for row in token_rows if int(row.get("id") or 0) == identifier),
        None,
    )
    if not token:
        raise KeyError("店铺授权不存在")

    candidate_names = list(dict.fromkeys(
        str(value or "").strip()
        for value in (
            shop_name_hint,
            token.get("display_name"),
            token.get("nickname"),
        )
        if str(value or "").strip()
    ))
    if not candidate_names:
        raise ValueError("该店铺授权没有名称，无法匹配比特浏览器窗口")

    window_id = ""
    matched_name = ""
    errors = []
    configured_windows = [
        row for row in list_shop_configs(include_ignored=True)
        if isinstance(row, dict)
    ]
    for candidate in candidate_names:
        exact_matches = [
            row for row in configured_windows
            if str(row.get("shop_name") or "").strip() == candidate
        ]
        if not exact_matches:
            folded_candidate = candidate.casefold()
            exact_matches = [
                row for row in configured_windows
                if str(row.get("shop_name") or "").strip().casefold()
                == folded_candidate
            ]
        configured_ids = list(dict.fromkeys(
            str(row.get("window_id") or "").strip()
            for row in exact_matches
            if str(row.get("window_id") or "").strip()
        ))
        if len(configured_ids) == 1:
            window_id = configured_ids[0]
            matched_name = candidate
            break
        if len(configured_ids) > 1:
            errors.append(
                f"店铺“{candidate}”绑定了多个比特浏览器窗口，请先保留唯一绑定"
            )
            continue
        try:
            window_id = getBrowserIdByName(candidate)
            matched_name = candidate
            break
        except RuntimeError as exc:
            errors.append(str(exc))
    if not window_id:
        raise ValueError(
            f"店铺授权“{candidate_names[0]}”未绑定比特浏览器窗口；"
            "请让授权显示名称或 Mercado 昵称与比特浏览器窗口名称完全一致。"
            f"{errors[-1] if errors else ''}"
        )
    shop_name = str(token.get("display_name") or matched_name).strip()

    try:
        result = openBrowser(
            window_id,
            api_lock_timeout=5,
            request_timeout=20,
        )
        if not isinstance(result, dict):
            raise RuntimeError(f"比特浏览器启动接口返回格式异常：{result}")
        if result.get("success") is False:
            message = str(result.get("msg") or "启动失败").strip()
            owner = result.get("lockOwner")
            if owner:
                message = f"{message}（当前任务：{owner}）"
            raise RuntimeError(message)
        _open_bitbrowser_page(result, target_url)
    finally:
        # 人工处理只负责启动窗口和页面，不应长期持有自动任务锁。
        releaseBrowserLease(window_id)
    return {
        "token_id": identifier,
        "claim_id": str(claim_id or "").strip(),
        "shop_name": shop_name,
        "window_id": window_id,
        "target_url": str(target_url or "").strip(),
    }


@app.route(
    '/api/mercado-claims/<int:token_id>/open-browser',
    methods=['POST'],
)
@login_required
def api_open_mercado_claim_browser(token_id):
    data = request.get_json(silent=True) or {}
    try:
        result = _open_mercado_claim_browser(
            token_id,
            data.get("claim_id"),
            data.get("shop_name"),
        )
        response = jsonify({
            "status": "success",
            "data": result,
            "message": f"已启动 {result['shop_name']} 的比特浏览器窗口",
        })
        response.headers["Cache-Control"] = "no-store"
        return response
    except KeyError as exc:
        return jsonify({"status": "error", "message": str(exc.args[0])}), 404
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("启动订单索赔对应的比特浏览器窗口失败")
        return jsonify({
            "status": "error",
            "message": f"启动比特浏览器窗口失败：{exc}",
        }), 502


@app.route(
    '/api/reputation/<int:token_id>/open-browser',
    methods=['POST'],
)
@login_required
def api_open_reputation_browser(token_id):
    data = request.get_json(silent=True) or {}
    try:
        result = _open_mercado_claim_browser(
            token_id,
            shop_name_hint=data.get("shop_name"),
            target_url=bit_reputation_info.REPUTATION_URL,
        )
        response = jsonify({
            "status": "success",
            "data": result,
            "message": f"已打开 {result['shop_name']} 的声誉页面",
        })
        response.headers["Cache-Control"] = "no-store"
        return response
    except KeyError as exc:
        return jsonify({"status": "error", "message": str(exc.args[0])}), 404
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("打开声誉对应的比特浏览器页面失败")
        return jsonify({
            "status": "error",
            "message": f"打开比特浏览器声誉页面失败：{exc}",
        }), 502


@app.route(
    '/api/mercado-communications/<int:token_id>/<action>',
    methods=['GET', 'POST'],
)
@login_required
def api_mercado_communication(token_id, action):
    normalized_action = str(action or "").strip().lower()
    allowed = (
        normalized_action in MERCADO_COMMUNICATION_READ_ACTIONS
        if request.method == "GET"
        else normalized_action in (
            MERCADO_COMMUNICATION_WRITE_ACTIONS
            | MERCADO_COMMUNICATION_VIEW_POST_ACTIONS
        )
    )
    if not allowed:
        return jsonify({"status": "error", "message": "不支持的美客多消息操作"}), 404
    payload = request.args.to_dict() if request.method == "GET" else (request.get_json(silent=True) or {})
    try:
        result = bit_db_api.execute_mercado_store_communication(
            token_id, normalized_action, payload
        )
        response = jsonify({"status": "success", "data": result})
        response.headers["Cache-Control"] = "no-store"
        return response
    except Exception as exc:
        return _mercado_communication_error_response(exc)


@app.route('/api/mercado-tokens/<int:token_id>', methods=['PATCH', 'DELETE'])
@login_required
def api_update_mercado_token(token_id):
    try:
        if request.method == "DELETE":
            affected = bit_db_api.delete_mercado_store_token(token_id)
            if not affected:
                raise KeyError("店铺授权不存在")
            return jsonify({"status": "success", "data": {"deleted": affected}})
        data = request.get_json(silent=True) or {}
        if "enabled" in data:
            if not isinstance(data.get("enabled"), bool):
                raise ValueError("店铺启用状态必须是布尔值")
            result = bit_db_api.set_mercado_store_token_enabled(
                token_id, data["enabled"]
            )
            message = "店铺已启用" if data["enabled"] else "店铺已关闭，所有业务操作将跳过"
        else:
            result = bit_db_api.rename_mercado_store_token(
                token_id, data.get("display_name", "")
            )
            message = "店铺名称已更新"
        return jsonify({
            "status": "success",
            "data": result,
            "message": message,
        })
    except Exception as exc:
        return _mercado_token_error_response(exc)


@app.route("/api/yandex-console/status", methods=["GET"])
@login_required
def api_yandex_console_status():
    running = _yandex_console_health()
    public_urls = _yandex_console_public_urls()
    return jsonify(
        {
            "status": "success",
            "data": {
                "running": running,
                **public_urls,
                "port": YANDEX_CONSOLE_PORT,
                "pid": (
                    _yandex_console_process.pid
                    if _yandex_console_process is not None
                    and _yandex_console_process.poll() is None
                    else None
                ),
            },
        }
    )


@app.route("/api/yandex-console/start", methods=["POST"])
@login_required
def api_yandex_console_start():
    running, message = ensure_yandex_console()
    public_urls = _yandex_console_public_urls()
    return (
        jsonify(
            {
                "status": "success" if running else "error",
                "message": message,
                "data": {
                    "running": running,
                    **public_urls,
                    "port": YANDEX_CONSOLE_PORT,
                },
            }
        ),
        200 if running else 503,
    )


@app.route(
    f"{YANDEX_CONSOLE_PROXY_PATH}/",
    defaults={"proxy_path": ""},
    methods=["GET", "HEAD"],
)
@app.route(
    f"{YANDEX_CONSOLE_PROXY_PATH}/<path:proxy_path>",
    methods=["GET", "POST", "PUT", "PATCH", "DELETE", "HEAD"],
)
@login_required
def proxy_yandex_console(proxy_path):
    if not proxy_path:
        running, message = ensure_yandex_console()
        if not running:
            return jsonify({"detail": message}), 503
    return _proxy_yandex_console_request(proxy_path)


@app.route("/")
@login_required
def index():
    return render_template(
        'index.html',
        current_user=session.get("workbench_user") or {},
        mercado_authorization=mercado_tokens.authorization_info(),
        runtime_role=RUNTIME_SETTINGS.role,
    )


@app.route("/login")
def login_page():
    if session.get("workbench_user"):
        return redirect(url_for("index"))
    return render_template("login.html")


@app.route("/api/login", methods=["POST"])
def api_login():
    data = request.get_json(silent=True) or request.form
    username = str(data.get("username", "")).strip()
    password = str(data.get("password", ""))
    remember = str(data.get("remember", "")).strip().lower() in (
        "1",
        "true",
        "yes",
        "on",
    )
    if not username or not password:
        return jsonify({"status": "error", "message": "请输入账号和密码"}), 400

    try:
        user = authenticate_workbench_user(username, password)
    except Exception as e:
        logging.error("工作台登录接口调用失败: %s", e)
        return jsonify({"status": "error", "message": f"登录接口调用失败: {str(e)}"}), 500

    if not user:
        return jsonify({"status": "error", "message": "账号或密码错误"}), 401

    session.clear()
    session.permanent = remember
    session["workbench_user"] = user
    return jsonify({
        "status": "success",
        "data": session["workbench_user"],
        "remember": remember,
        "expires_in": WORKBENCH_REMEMBER_HOURS * 60 * 60 if remember else None,
    })


@app.route("/api/execution-targets/local-token", methods=["POST"])
@login_required
def api_local_executor_token():
    data = request.get_json(silent=True) or {}
    permission = str(data.get("permission") or "").strip()
    if permission not in LOCAL_EXECUTOR_PERMISSIONS:
        return jsonify({
            "status": "error",
            "message": "不支持的本机执行权限",
        }), 400
    user = get_current_workbench_user()
    if not workbench_user_has_permission(user, permission):
        return jsonify({
            "status": "error",
            "message": "当前账号没有执行该操作的权限",
        }), 403
    try:
        token = create_local_executor_token(user, permission)
    except RuntimeError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 503
    response = jsonify({
        "status": "success",
        "data": {
            "token": token,
            "base_url": LOCAL_EXECUTOR_BROWSER_URL,
            "target_address_space": local_executor_target_address_space(),
            "expires_in": LOCAL_EXECUTOR_TOKEN_MAX_AGE_SECONDS,
        },
    })
    response.headers["Cache-Control"] = "no-store"
    return response


@app.route("/api/execution-targets/local-token/verify", methods=["POST"])
def api_verify_local_executor_token():
    if USE_DB_API:
        return jsonify({"status": "error", "message": "请向服务端验证本机执行凭证"}), 503
    authorization = str(request.headers.get("Authorization") or "")
    token = authorization[7:].strip() if authorization.lower().startswith("bearer ") else ""
    try:
        user = _verify_local_executor_session_token(token)
    except Exception:
        logging.exception("验证本机执行凭证失败")
        return jsonify({"status": "error", "message": "服务端暂时无法校验账号权限，请重试"}), 503
    if not user:
        return jsonify({"status": "error", "message": "本机执行凭证无效、已过期或账号权限已变更"}), 401
    response = jsonify({"status": "success", "data": user})
    response.headers["Cache-Control"] = "no-store"
    return response


@app.route("/api/browser-extension/login", methods=["POST"])
def api_browser_extension_login():
    data = request.get_json(silent=True) or {}
    username = str(data.get("username") or "").strip()
    password = str(data.get("password") or "")
    if not username or not password:
        return jsonify({"status": "error", "message": "请输入泽顺控制台账号和密码"}), 400
    try:
        user = authenticate_workbench_user(username, password)
    except Exception as exc:
        logging.exception("泽顺商品采集助手登录失败")
        return jsonify({"status": "error", "message": f"登录接口调用失败：{exc}"}), 500
    if not user:
        return jsonify({"status": "error", "message": "账号或密码错误"}), 401
    return jsonify({
        "status": "success",
        "data": {
            "token": create_browser_extension_token(user),
            "user": user,
            "expires_in": WORKBENCH_REMEMBER_HOURS * 60 * 60,
        },
    })


@app.route("/api/browser-extension/session", methods=["GET"])
@browser_extension_login_required
def api_browser_extension_session():
    return jsonify({"status": "success", "data": {"user": g.browser_extension_user}})


@app.route("/api/browser-extension/collect", methods=["POST"])
@browser_extension_login_required
def api_browser_extension_collect():
    from erp.mercadolibre_batch_collector import validate_collection_request

    data = request.get_json(silent=True) or {}
    product = data.get("product") if isinstance(data.get("product"), dict) else data
    item_id = str(product.get("source_item_id") or "").strip().upper()
    title = str(product.get("title") or "").strip()
    source_url = str(product.get("source_url") or product.get("final_url") or "").strip()
    if not re.fullmatch(r"(?:ML[A-Z]|CBT)\d{5,}", item_id):
        return jsonify({"status": "error", "message": "未识别到有效的 Mercado Libre 商品编号"}), 400
    if not title:
        return jsonify({"status": "error", "message": "商品标题不能为空"}), 400
    try:
        source_url, _ = validate_collection_request(source_url, 1)
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    product = dict(product)
    product["source_item_id"] = item_id
    product["source_url"] = source_url
    created_by = str(
        g.browser_extension_user.get("display_name")
        or g.browser_extension_user.get("username")
        or "浏览器插件"
    )
    task_id = None
    try:
        task_id = db_create_mercado_collection_task(
            source_url, 1, f"浏览器插件：{created_by}"
        )
        db_upsert_mercado_collection_items(task_id, [product])
        complete = str(product.get("scrape_status") or "partial") == "ok"
        status = "completed" if complete else "partial"
        message = (
            "浏览器插件采集完成"
            if complete
            else "浏览器插件快速采集完成，重量尺寸或详情待补充"
        )
        db_update_mercado_collection_task(
            task_id,
            status=status,
            message=message,
            collected_count=1,
            completed_count=1 if complete else 0,
            failed_count=0 if complete else 1,
            current_page=1,
            started=True,
            finished=True,
        )
    except Exception as exc:
        if task_id:
            try:
                db_update_mercado_collection_task(
                    task_id,
                    status="error",
                    message=f"浏览器插件采集失败：{exc}",
                    failed_count=1,
                    started=True,
                    finished=True,
                )
            except Exception:
                logging.exception("更新浏览器插件采集失败任务状态失败")
        logging.exception("浏览器插件采集商品入库失败")
        return jsonify({"status": "error", "message": f"商品入库失败：{exc}"}), 500
    return jsonify({
        "status": "success",
        "data": {
            "task_id": int(task_id),
            "source_item_id": item_id,
            "scrape_status": product.get("scrape_status") or "partial",
        },
    }), 201


@app.route("/logout", methods=["GET", "POST"])
def logout():
    session.clear()
    if request.method == "POST" or request.path.startswith("/api/"):
        return jsonify({"status": "success"})
    return redirect(url_for("login_page"))


# 定义路由和返回内容
@app.route("/zs")
def hello_whzs():
    callback_code = request.args.get("code", "")
    if callback_code:
        try:
            code = mercado_tokens.extract_authorization_code(callback_code)
        except ValueError:
            return Response("无效的 Mercado Libre 授权码", status=400, content_type="text/plain; charset=utf-8")
        escaped_code = html.escape(code)
        response = Response(
            f"""<!doctype html>
<html lang="zh-CN"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width">
<meta name="referrer" content="no-referrer"><title>授权成功</title>
<style>body{{margin:0;background:#f4f7fb;color:#172033;font-family:Arial,'Microsoft YaHei',sans-serif}}main{{max-width:640px;margin:12vh auto;padding:28px;background:#fff;border:1px solid #d9e2ef;border-radius:12px;box-shadow:0 16px 40px #10284014}}h1{{font-size:22px}}code{{display:block;margin:18px 0;padding:14px;background:#f1f5f9;border-radius:8px;word-break:break-all}}button{{border:0;border-radius:8px;background:#2563eb;color:#fff;padding:10px 16px;font-weight:700;cursor:pointer}}p{{color:#667085}}</style></head>
<body><main><h1>Mercado Libre 授权成功</h1><p>复制下面的 TG Code，回到“店铺授权”模块完成添加。TG Code 只能使用一次。</p>
<code id="tg-code">{escaped_code}</code><button type="button" onclick="navigator.clipboard.writeText(document.getElementById('tg-code').textContent).then(()=>this.textContent='已复制')">复制 TG Code</button></main></body></html>""",
            content_type="text/html; charset=utf-8",
        )
        response.headers["Cache-Control"] = "no-store"
        response.headers["Pragma"] = "no-cache"
        return response
    if not session.get("workbench_user"):
        return redirect(url_for("login_page"))
    return "武汉泽顺"


# --- 新增：1688大模型找货数据插入接口 ---
# @app.route('/api/v1/chat', methods=['POST'])
def api_insert_chat_info():
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"status": "error", "message": "Missing JSON payload"}), 400

    required_fields = ["name", "site", "message", "chat", "response", "time"]
    missing_fields = [field for field in required_fields if data.get(field) in (None, "")]
    if missing_fields:
        return jsonify({
            "status": "error",
            "message": "Missing required fields",
            "fields": missing_fields
        }), 422

    try:
        chat_id = db_insert_chat_info(
            data["name"],
            data["site"],
            data["message"],
            data["chat"],
            data["response"],
            data["time"]
        )
        return jsonify({
            "status": "success",
            "message": "Chat info inserted successfully",
            "id": chat_id
        }), 201
    except Exception as e:
        logging.error(f"Chat info insert failed: {str(e)}")
        return jsonify({"status": "error", "message": f"Database error: {str(e)}"}), 500


# @app.route('/api/v1/records', methods=['POST'])
def insert_record():
    blocked = reject_db_api_client_mode()
    if blocked:
        return blocked

    # 获取客户端发送的 JSON 数据
    data = request.get_json()
    if not data:
        return jsonify({"status": "error", "message": "Missing JSON payload"}), 400

    # 安全提取 product_id（必填项项校验）
    product_id = data.get('product_id')
    if not product_id:
        return jsonify({"status": "error", "message": "Field 'product_id' is required"}), 422

    # 提取其余字段，并设置默认值（与你的数据表结构严格对应）
    zhiying_category = data.get('zhiying_category', None)
    original_img_url = data.get('original_img_url', None)
    is_same_style = int(data.get('is_same_style', 0))
    title = data.get('title', None)
    identified_weight = int(data.get('identified_weight', 0))
    pre_modified_weight = int(data.get('pre_modified_weight', 0))
    post_modified_weight = int(data.get('post_modified_weight', 0))

    # 金额与置信度转换为 Decimal 类型，防止精度丢失
    pre_modified_cost_usd = Decimal(str(data.get('pre_modified_cost_usd', '0.0000')))
    post_modified_cost_usd = Decimal(str(data.get('post_modified_cost_usd', '0.0000')))
    max_sku_price_cny = Decimal(str(data.get('max_sku_price_cny', '0.00')))
    model_confidence = Decimal(str(data.get('model_confidence', '0.00')))

    max_sku_spec = data.get('max_sku_spec', None)
    max_sku_id = data.get('max_sku_id', None)
    weight_issue = data.get('weight_issue', None)
    matched_1688_url = data.get('matched_1688_url', None)
    reason = data.get('reason', None)

    sql = """
        INSERT INTO product_mapping_records (
            crawl_time, zhiying_category, original_img_url, is_same_style, 
            product_id, title, identified_weight, pre_modified_weight, 
            post_modified_weight, pre_modified_cost_usd, post_modified_cost_usd, 
            max_sku_price_cny, max_sku_spec, max_sku_id, model_confidence, 
            weight_issue, matched_1688_url, reason
        ) VALUES (
            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
        )
    """

    current_time = datetime.now()
    params = (
        current_time, zhiying_category, original_img_url, is_same_style,
        product_id, title, identified_weight, pre_modified_weight,
        post_modified_weight, pre_modified_cost_usd, post_modified_cost_usd,
        max_sku_price_cny, max_sku_spec, max_sku_id, model_confidence,
        weight_issue, matched_1688_url, reason
    )

    conn = None
    cursor = None
    try:
        # 从连接池中取得一条连接
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(sql, params)
        conn.commit()

        logging.info(f"成功录入产品数据，Product ID: {product_id}")
        return jsonify({
            "status": "success",
            "message": "Record inserted successfully",
            "id": cursor.lastrowid
        }), 201

    except Exception as e:
        if conn:
            conn.rollback()
        logging.error(f"数据库写入失败: {str(e)}")
        return jsonify({"status": "error", "message": f"Database error: {str(e)}"}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()  # 将连接放回连接池


def recover_interrupted_mercado_collection_tasks():
    """Mark tasks owned by an earlier workbench process as interrupted."""
    from erp.mercadolibre_collection_store import recover_interrupted_collection_tasks

    cutoff = datetime.now().replace(microsecond=0).strftime("%Y-%m-%d %H:%M:%S")
    recovered = recover_interrupted_collection_tasks(cutoff=cutoff)
    if recovered:
        logging.warning("已恢复 %s 个异常中断的 Mercado 商品采集任务", recovered)
    return recovered


def start_interrupted_collection_recovery():
    """Run startup recovery without delaying the HTTP service from listening."""

    def recover_safely():
        try:
            recover_interrupted_mercado_collection_tasks()
        except Exception:
            logging.exception("恢复异常中断的 Mercado 商品采集任务失败")

    recovery_thread = threading.Thread(
        target=recover_safely,
        name="mercado-collection-startup-recovery",
        daemon=True,
    )
    recovery_thread.start()
    return recovery_thread


def start_store_link_scheduler_bootstrap():
    """Start the optional store-link scheduler without blocking Flask startup."""

    if bit_db_api.DB_MODE != "mysql":
        return None

    def start_safely():
        try:
            bit_store_link_sync.start_store_link_auto_scheduler()
            logging.info(
                "店铺链接自动同步调度已启动：每 %s 天同步一次",
                bit_store_link_sync.STORE_LINK_AUTO_SYNC_DAYS,
            )
        except Exception:
            logging.exception("启动店铺链接自动同步调度失败")

    scheduler_thread = threading.Thread(
        target=start_safely,
        name="mercado-store-link-scheduler-bootstrap",
        daemon=True,
    )
    scheduler_thread.start()
    return scheduler_thread


def start_token_refresh_scheduler_bootstrap():
    """Start proactive Mercado token renewal without delaying HTTP startup."""

    def start_safely():
        try:
            mercado_tokens.start_token_auto_refresh_scheduler()
            logging.info(
                "店铺 Token 自动刷新已启动：到期前 %s 分钟刷新，每 %s 秒检查",
                mercado_tokens.TOKEN_AUTO_REFRESH_BEFORE_MINUTES,
                mercado_tokens.TOKEN_AUTO_REFRESH_CHECK_SECONDS,
            )
        except Exception:
            logging.exception("启动店铺 Token 自动刷新失败")

    scheduler_thread = threading.Thread(
        target=start_safely,
        name="mercado-token-refresh-scheduler-bootstrap",
        daemon=True,
    )
    scheduler_thread.start()
    return scheduler_thread


def start_store_email_sync_scheduler_bootstrap():
    """Backfill seller emails for existing local store authorizations."""

    if bit_db_api.DB_MODE != "mysql":
        return None

    def start_safely():
        try:
            mercado_tokens.start_store_email_sync_scheduler()
            logging.info(
                "店铺邮箱自动同步已启动：每 %s 秒补读缺失邮箱",
                mercado_tokens.STORE_EMAIL_SYNC_INTERVAL_SECONDS,
            )
        except Exception:
            logging.exception("启动店铺邮箱自动同步失败")

    scheduler_thread = threading.Thread(
        target=start_safely,
        name="mercado-store-email-sync-bootstrap",
        daemon=True,
    )
    scheduler_thread.start()
    return scheduler_thread


def start_prohibited_listing_scheduler_bootstrap():
    """Start the daily official-API prohibited-listing scheduler."""

    if bit_db_api.DB_MODE != "mysql":
        return None

    def start_safely():
        try:
            bit_prohibited_listing_sync.start_prohibited_listing_auto_scheduler()
            logging.info(
                "禁限售列表自动同步调度已启动：每 %s 小时同步一次",
                bit_prohibited_listing_sync.PROHIBITED_AUTO_SYNC_HOURS,
            )
        except Exception:
            logging.exception("启动禁限售列表自动同步调度失败")

    scheduler_thread = threading.Thread(
        target=start_safely,
        name="mercado-prohibited-scheduler-bootstrap",
        daemon=True,
    )
    scheduler_thread.start()
    return scheduler_thread


def start_official_infraction_scheduler_bootstrap():
    """Start the 12-hour official infringement scheduler."""

    if bit_db_api.DB_MODE != "mysql":
        return None

    def start_safely():
        try:
            mercado_infraction_sync.start_official_infraction_auto_scheduler()
            logging.info(
                "官方侵权数据自动同步已启动：每 %s 小时同步一次",
                mercado_infraction_sync.INFRACTION_AUTO_SYNC_HOURS,
            )
        except Exception:
            logging.exception("启动官方侵权数据自动同步失败")

    scheduler_thread = threading.Thread(
        target=start_safely,
        name="mercado-official-infraction-scheduler-bootstrap",
        daemon=True,
    )
    scheduler_thread.start()
    return scheduler_thread


def interface_hot_reload_enabled(value=None):
    """Enable source reloading for source runs, never for frozen executables."""

    # Werkzeug's reloader passes a listening socket to its child process via a
    # file descriptor.  That descriptor hand-off is not reliable for a frozen
    # Windows executable (and is unnecessary because bundled sources cannot be
    # hot-reloaded), where it can fail with WinError 10038 in socket.fromfd().
    if getattr(sys, "frozen", False):
        return False

    configured = (
        os.environ.get("BIT_INTERFACE_HOT_RELOAD", "1")
        if value is None
        else value
    )
    return _truthy_env(configured)


def is_werkzeug_reloader_child(environ=None):
    environment = os.environ if environ is None else environ
    return str(environment.get("WERKZEUG_RUN_MAIN", "")).strip().lower() == "true"


def start_interface_background_services():
    # 客户端只承载本机界面与浏览器自动化。所有会读取数据库、刷新 Token
    # 或维护中心数据的后台线程统一由服务端进程运行。
    if USE_DB_API:
        logging.info("客户端模式不启动数据库维护与中心调度线程")
        return

    start_interrupted_collection_recovery()
    start_store_link_scheduler_bootstrap()
    start_prohibited_listing_scheduler_bootstrap()
    start_official_infraction_scheduler_bootstrap()
    start_api_reputation_scheduler_bootstrap()
    start_token_refresh_scheduler_bootstrap()
    start_store_email_sync_scheduler_bootstrap()
    bit_order_sync.ensure_order_sync_scheduler()
    bit_order_sync.ensure_order_financial_backfill_worker()
    bit_order_sync.ensure_order_image_backfill_worker()


def run_interface_server():
    hot_reload = interface_hot_reload_enabled()
    reloader_child = hot_reload and is_werkzeug_reloader_child()
    interface_lock = None

    # The reloader parent owns the singleton lock for its whole lifetime. Its
    # child process serves requests and is replaced automatically on edits.
    if not reloader_child:
        interface_lock = InterProcessLock(
            "bit_interface_singleton",
            owner="bit_interface.py",
            metadata={
                "port": 5000,
                "project": str(PROJECT_ROOT),
                "hot_reload": hot_reload,
                "runtime_role": RUNTIME_SETTINGS.role,
            },
        )
        if not interface_lock.acquire(timeout=0):
            owner = interface_lock.read_owner()
            logging.error(
                "泽顺工作台服务已经运行，本次重复进程退出：pid=%s",
                owner.get("pid") or "unknown",
            )
            return False

    try:
        if not hot_reload or reloader_child:
            start_interface_background_services()
        logging.info(
            "工作台以 %s 角色启动；代码热更新%s",
            RUNTIME_SETTINGS.role,
            "已启用" if hot_reload else "已关闭",
        )
        app.run(
            host="0.0.0.0",
            port=5000,
            threaded=True,
            debug=False,
            use_debugger=False,
            use_reloader=hot_reload,
        )
        return True
    finally:
        if interface_lock is not None:
            interface_lock.release()


def run_interface_main():
    """Prepare frozen multiprocessing support before starting the server."""

    # PyInstaller replaces this helper so ProcessPoolExecutor worker command
    # lines are dispatched to multiprocessing instead of re-running the Flask
    # entry point.  Calling it is harmless for normal source runs.
    multiprocessing.freeze_support()
    return run_interface_server()


if __name__ == '__main__':
    run_interface_main()
