import time
from sys import prefix
from concurrent.futures import ThreadPoolExecutor, as_completed

from oss2 import is_valid_endpoint
from selenium import webdriver
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.wait import WebDriverWait

from bit.bit_utils import get_now_time
from bit.bit_api import *
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import pyautogui
from bit.bit_switch_country import *
from openpyxl import load_workbook
from bit_send_mail import *
import pandas as pd

from datetime import datetime
from pathlib import Path
from bit.bit_clash import *
from bit.bit_mysql import *


SITE_PREFIX_MAP = {
    "墨西哥": "MLM",
    "巴西": "MLB",
    "哥伦比亚": "MCO",
    "智利": "MLC",
    "阿根廷": "MLA",
    "乌拉圭": "MLU",
}

SITE_SWITCH_SELECTOR_MAP = {
    "墨西哥": 'div[data-value="MLM-remote"]',
    "巴西": 'div[data-value="MLB-remote"]',
    "哥伦比亚": 'div[data-value="MCO-remote"]',
    "智利": 'div[data-value="MLC-remote"]',
    "阿根廷": 'div[data-value="MLA-remote"]',
    "乌拉圭": 'div[data-value="MLU-remote"]',
}


def _get_text_list(driver, class_name):
    elements = WebDriverWait(driver, 30).until(
        EC.presence_of_all_elements_located((By.CLASS_NAME, class_name))
    )
    return [el.get_attribute("textContent").strip() for el in elements]


def _get_page_signature(driver):
    try:
        ids = _get_text_list(driver, "infraction-item__id")
        return tuple(ids)
    except Exception:
        return tuple()


def _is_next_button_disabled(button):
    disabled = button.get_attribute("disabled")
    aria_disabled = button.get_attribute("aria-disabled")
    class_name = button.get_attribute("class") or ""
    parent_class = ""
    try:
        parent_class = button.find_element(By.XPATH, "./ancestor::li[1]").get_attribute("class") or ""
    except Exception:
        pass
    return (
        disabled is not None
        or aria_disabled == "true"
        or "disabled" in class_name.lower()
        or "disabled" in parent_class.lower()
    )


def _find_next_button(driver):
    selectors = [
        (By.XPATH, "//a[.//span[contains(@class, 'andes-pagination__arrow-title') and normalize-space()='Next']]"),
        (By.XPATH, "//button[.//span[contains(@class, 'andes-pagination__arrow-title') and normalize-space()='Next']]"),
        (By.XPATH, "//*[contains(@class, 'andes-pagination__button')][.//*[normalize-space()='Next']]"),
        (By.XPATH, "//span[contains(@class, 'andes-pagination__arrow-title') and normalize-space()='Next']/ancestor::*[self::a or self::button][1]"),
    ]
    for by, selector in selectors:
        elements = driver.find_elements(by, selector)
        for element in elements:
            if element.is_displayed():
                return element
    return None


def _click_next_page(driver, previous_signature, page_no):
    for attempt in range(3):
        try:
            next_button = _find_next_button(driver)
            if next_button is None or _is_next_button_disabled(next_button):
                print("当前为最后一页，循环结束")
                return False

            driver.execute_script(
                "arguments[0].scrollIntoView({block: 'center', inline: 'center'});",
                next_button,
            )
            time.sleep(0.5)
            try:
                WebDriverWait(driver, 10).until(lambda d: next_button.is_enabled())
                next_button.click()
            except Exception:
                driver.execute_script("arguments[0].click();", next_button)

            WebDriverWait(driver, 30).until(
                lambda d: _get_page_signature(d) != previous_signature
            )
            print(f"成功点击下一页，当前第{page_no + 1}页")
            time.sleep(1)
            return True
        except StaleElementReferenceException:
            time.sleep(1)
            continue
        except TimeoutException:
            print(f"点击下一页后页面未变化，重试第{attempt + 1}次")
            time.sleep(2)
            continue
        except Exception as e:
            print(f"翻页失败，重试第{attempt + 1}次:", e)
            time.sleep(2)
            continue

    print("翻页多次失败，结束当前站点抓取")
    return False


def _read_current_infractions_page(driver, name, site):
    ids = _get_text_list(driver, "infraction-item__id")
    titles = _get_text_list(driver, "infraction-item__title")
    dates = _get_text_list(driver, "infraction-denounce__date")

    prefix = SITE_PREFIX_MAP.get(site, "")
    rows = []
    for id_text, title, date in zip(ids, titles, dates):
        rows.append(
            [
                name,
                site,
                id_text.replace("#", prefix),
                title,
                date,
                get_now_time(),
            ]
        )
    return rows


def get_infractions_info(window_id, name, site):
    res = openBrowser(window_id)  # 窗口ID从窗口配置界面中复制，或者api创建后返回

    print(res)
    driverPath = res["data"]["driver"]
    debuggerAddress = res["data"]["http"]

    # selenium 连接代码
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_experimental_option("debuggerAddress", debuggerAddress)

    chrome_service = Service(driverPath)
    driver = webdriver.Chrome(service=chrome_service, options=chrome_options)

    driver.implicitly_wait(10)
    # 设置最长等待时间为 10 秒
    wait = WebDriverWait(driver, 10)

    driver.get(
        "https://global-selling.mercadolibre.com/noindex/pppi/infractions?tab=detections&offset=0"
    )
    time.sleep(10)
    i = 0
    while i < 3:
        i = i + 1
        try:
            # 打开站点选择器
            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable(
                    (By.CLASS_NAME, "nav-header-cbt__site-switcher")
                )
            ).click()

            print(name + "打开站点选择器")
            time.sleep(5)
            path = SITE_SWITCH_SELECTOR_MAP.get(site, 'div[data-value="MLM-remote"]')

            WebDriverWait(driver, 30).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, path))
            ).click()

            driver.refresh()
            time.sleep(3)
            print(get_now_time() + name + site + "选择站点：", site)
            break
        except Exception as e:
            print(get_now_time() + name + site + "重新执行选择站点")
            switch_random_hongkong_node()
            get_public_ip()
            continue
    infractions_list = []
    seen_ids = set()
    page_no = 1
    while True:
        page_rows = _read_current_infractions_page(driver, name, site)
        new_count = 0
        for row in page_rows:
            row_id = row[2]
            if row_id in seen_ids:
                continue
            seen_ids.add(row_id)
            infractions_list.append(row)
            new_count += 1

        print(f"{get_now_time()}{name}{site}第{page_no}页抓取{len(page_rows)}条，新增{new_count}条")
        previous_signature = _get_page_signature(driver)
        if not previous_signature:
            print("当前页面没有侵权数据，结束当前站点抓取")
            break

        if not _click_next_page(driver, previous_signature, page_no):
            break
        page_no += 1

    return infractions_list


def _run_infractions_for_browser(row):
    id = row[0]
    name = row[1]
    remark = row[2]
    if remark == "忽略":
        return [], []

    print(get_now_time() + "开始打开窗口:" + name)
    if not row[3]:
        return [], [("获取侵权信息", name, "", "失败：未配置站点", get_now_time())]

    site_list = row[3].split("，")
    infraction_info_sum = []
    result = []

    for site in site_list:
        site = str(site).strip()
        if not site:
            continue

        for i in range(1, 4):
            try:
                infraction_info = get_infractions_info(id, name, site)
                infraction_info_sum.extend(infraction_info)
                print(get_now_time() + name + site + "成功")
                result.append(("获取侵权信息", name, site, "成功", get_now_time()))
                break
            except Exception as e:
                print(get_now_time() + name + site + "执行失败", e)
                if i == 3:
                    result.append(("获取侵权信息", name, site, "失败", get_now_time()))
                else:
                    time.sleep(5)

    print(get_now_time() + "结束，正在关闭窗口")

    try:
        closeBrowser(id)
    except Exception as e:
        print(get_now_time() + name + "关闭窗口失败", e)
    print(get_now_time() + "已经关闭窗口")
    return infraction_info_sum, result


def get_infractions_info_all(max_workers=10):
    start = int(time.time())
    print(start)
    root_path = Path(__file__).resolve().parent
    # file_path = root_path / "比特配置文件.xlsx"
    file_path = root_path / "比特配置文件.xlsx"

    wb = load_workbook(file_path)
    sheet = wb.active
    infraction_info_sum = []
    result = []
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    rows = [row for row in rows if row and row[0] and row[2] != "忽略"]

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_map = {
            executor.submit(_run_infractions_for_browser, row): row for row in rows
        }
        for future in as_completed(future_map):
            row = future_map[future]
            name = row[1]
            try:
                browser_infractions, browser_result = future.result()
                infraction_info_sum.extend(browser_infractions)
                result.extend(browser_result)
                print(get_now_time() + name + "窗口任务完成")
            except Exception as e:
                print(get_now_time() + name + "窗口任务异常", e)
                result.append(("获取侵权信息", name, "", "失败", get_now_time()))

    infraction_info_sum_str = "\n".join(map(str, infraction_info_sum))
    print(infraction_info_sum_str)

    end = int(time.time())
    print(get_now_time() + "总花费", end - start)
    df = pd.DataFrame(
        infraction_info_sum,
        columns=["店铺名", "站点", "编号", "标题", "侵权时间", "执行时间"],
    )

    now = datetime.now()
    date_str = datetime.now().strftime("%Y-%m-%d-%H")

    df.to_excel(
        root_path / ("美客多侵权/武汉泽顺店铺侵权信息汇总" + date_str + ".xlsx"),
        index=False,
    )

    send_info(
        "美客多所有店铺侵权汇总",
        infraction_info_sum_str,
        root_path / ("美客多侵权/武汉泽顺店铺侵权信息汇总" + date_str + ".xlsx"),
        r"武汉泽顺店铺侵权信息汇总" + date_str + ".xlsx",
    )
    print(get_now_time() + "发送邮件成功")

    insert_task_record(result)
    inset_infraction_info(infraction_info_sum)



if __name__ == '__main__':

    # inf=get_infractions_info('1495e31cb630406bb690ba187f264fe7','vngbjkk','墨西哥')
    # print(inf)
    get_infractions_info_all()
