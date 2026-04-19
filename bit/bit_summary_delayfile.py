import time
from ctypes.wintypes import DOUBLE

import pandas as pd
from bit_api import *
from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime
import pandas
from send_mail import *

from utils import *
import sys
from bit_mysql import *


def summary_delayFile():
    start = int(time.time())
    print(start)
    file_path = Path(__file__).resolve().parent / "比特配置文件.xlsx"
    wb = load_workbook(file_path)
    sheet = wb.active
    reputation_info_sum = []
    save_fold=""

    if sys.platform == "win32":
        print("当前环境是 Windows")
        save_fold = r"C:/BitDownload/"
    elif sys.platform == "darwin":
        print("当前环境是 macOS")
        save_fold = "/Users/active11/Downloads/"

    # 使用 min_row=2 跳过第一行
    file_dict = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        id = row[0]
        name = row[1]
        remark = row[2]
        seq = str(row[4])

        if (remark == '忽略'):
            continue
        fold = Path(save_fold + seq)
        print(fold)
        for file in fold.glob('*.csv'):
            print(f"文件名: {file.name}, 绝对路径: {file.absolute()}")
            part = file.name.split("_")
            country = part[10]
            file_time = float((part[12].replace('.csv', '')).split(" (")[0])
            file_datetime = datetime.fromtimestamp(file_time / 1000.0)
            filename = file_dict.get(name + "-" + country)
            if (filename == None):
                file_dict[name + "-" + country] = str(file.absolute())
            else:
                part = filename.split("_")
                filename_time = float((part[12].replace('.csv', '')).split(" (")[0])
                if (file_time > filename_time):
                    file_dict[name + "-" + country] = str(file.absolute())
    print(file_dict)

    for key, value in file_dict.items():
        print(key + "|||" + value)

    line = []

    rp_file = Path(__file__).resolve().parent / "美客多声誉"
    filename = get_latest_modified_file(rp_file)
    filepath = rp_file / filename
    print(filepath)
    dict_delay = {}
    df = pd.read_excel(filepath, engine='openpyxl')
    for index, row in df.iterrows():
        print(row[0])
        if (len(row) > 6):
            delayrate = row['延误率']
            name = row['店铺名']
            site = row['站点']
            key = name + site
            dict_delay[key] = delayrate
    print(dict_delay)

    for key, filepath in file_dict.items():

        name = key.split("-")[0]
        site = key.split("-")[1]
        if (site == 'AR'):
            site = "阿根廷"
        if (site == 'MX'):
            site = '墨西哥'
        if (site == 'BR'):
            site = '巴西'
        if (site == 'CL'):
            site = '智利'
        if (site == 'CO'):
            site = '哥伦比亚'
        if (site == 'UY'):
            site = '乌拉圭'
        delayrate=""
        try:
            delayrate = dict_delay[name + site]
        except Exception as e:
            continue

        part = filepath.split("_")
        filename_time = float((part[12].replace('.csv', '')).split(" (")[0]) / 1000
        dt_filename = datetime.fromtimestamp(filename_time)

        # 将datetime对象转换为指定格式的字符串
        date_string = dt_filename.strftime('%Y-%m-%d %H:%M:%S')

        try:
            df = pd.read_csv(filepath, header=None, skiprows=1)
            for index, row in df.iterrows():
                # 假设第一列是订单号，第三列是金额（转化为浮点数）

                try:

                    print((name, site, delayrate, row[0], row[1], row[2], row[5], row[6], date_string, filepath))
                    line.append((name, site, delayrate, row[0], row[1], row[2], row[5], row[6], date_string, filepath))

                except Exception as e:
                    print(e)
        except Exception as e:
            print("错误文件为", filepath)

    print(line)
    df = pd.DataFrame(line, columns=['店铺', '站点', '延误率', '下单时间', '销售单号', '订单标题', '截止延误时间',
                                     '实际揽收时间', '更新时间', '文件路径'])
    now = datetime.now()
    date_str = datetime.now().strftime("%Y-%m-%d-%H")

    df.to_excel(Path(__file__).resolve().parent/("美客多延误\武汉泽顺店铺延误信息汇总" + date_str + ".xlsx"), index=False)

    send_info('美客多所有店铺延误信息汇总', "美客多所有店铺延误信息汇总",Path(__file__).resolve().parent/("美客多延误\武汉泽顺店铺延误信息汇总" + date_str + ".xlsx"),
              r"武汉泽顺店铺延误信息汇总" + date_str + ".xlsx")
    inset_delay_info(line)



if __name__ == '__main__':
   summary_delayFile()

