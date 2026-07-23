"""比特浏览器店铺配置的数据访问层。

运行时只从 MySQL 或数据库 HTTP 接口读取。Excel 仅用于显式执行一次性迁移。
"""

import argparse
import re
from pathlib import Path

from bit import bit_db_api


CONFIG_FIELDS = (
    "window_id",
    "shop_name",
    "status",
    "sites",
    "sequence_no",
    "salesperson",
    "email",
)
HEADER_ALIASES = {
    "window_id": ("窗口ID", "窗口 id", "浏览器窗口ID"),
    "shop_name": ("账号名", "店铺名", "窗口名称"),
    "status": ("状态（若为忽略则跳过）", "状态", "备注"),
    "sites": ("站点", "国家", "国家/站点"),
    "sequence_no": ("序号", "下载序号", "文件夹序号"),
    "salesperson": ("业务员", "负责人"),
    "email": ("邮箱", "电子邮箱", "美客多邮箱", "登录邮箱", "email", "e-mail"),
}


def _text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalized_header(value):
    return "".join(_text(value).lower().split())


def is_ignored_config(value):
    if isinstance(value, dict):
        value = value.get("status")
    return "忽略" in _text(value)


def split_config_sites(value):
    text = _text(value)
    if not text or text.casefold() == "nan":
        return []
    sites = []
    for site in re.split(r"[，,、/;；|\s]+", text):
        site = site.strip()
        if site and site not in sites:
            sites.append(site)
    return sites


def normalize_config_record(record):
    record = dict(record or {})
    return {field: _text(record.get(field)) for field in CONFIG_FIELDS}


def _browser_config_api_route_missing(exc):
    message = str(exc or "")
    return (
        bit_db_api.DB_MODE == "api"
        and "404" in message
        and "/api/db/browser-configs" in message
    )


def list_shop_configs(include_ignored=True):
    try:
        rows = bit_db_api.list_bit_browser_configs(include_ignored=include_ignored) or []
    except RuntimeError as exc:
        if not _browser_config_api_route_missing(exc):
            raise
        # 兼容云端尚未部署 browser-configs 路由的过渡期；仍然只读取数据库，
        # 不回退到本地 Excel。直连不可用时保留原接口错误，便于定位部署问题。
        try:
            from bit.bit_mysql import list_bit_browser_configs

            rows = list_bit_browser_configs(include_ignored=include_ignored) or []
        except Exception:
            raise exc
    normalized = [normalize_config_record(row) for row in rows]
    if include_ignored:
        return normalized
    return [row for row in normalized if not is_ignored_config(row)]


def list_config_rows(include_ignored=True):
    """返回兼容旧脚本的 7 列元组，但数据源已经是数据库。"""
    return [
        tuple(record[field] for field in CONFIG_FIELDS)
        for record in list_shop_configs(include_ignored=include_ignored)
    ]


def get_shop_config(shop_name="", window_id="", include_ignored=True):
    shop_name = _text(shop_name)
    window_id = _text(window_id)
    lookup_error = None
    try:
        record = bit_db_api.get_bit_browser_config(
            shop_name=shop_name,
            window_id=window_id,
            include_ignored=include_ignored,
        )
        if not record and shop_name and window_id:
            record = bit_db_api.get_bit_browser_config(
                shop_name=shop_name,
                include_ignored=include_ignored,
            )
        if not record and window_id:
            record = bit_db_api.get_bit_browser_config(
                window_id=window_id,
                include_ignored=include_ignored,
            )
        if record:
            return normalize_config_record(record)
    except RuntimeError as exc:
        # 云端服务可能尚未部署单条 lookup 路由；列表路由包含同一份数据库数据，
        # 因而可以安全地在本地做精确匹配，避免单店任务被旧版服务阻断。
        lookup_error = exc

    try:
        records = list_shop_configs(include_ignored=include_ignored)
    except RuntimeError:
        if lookup_error is not None:
            raise lookup_error
        raise

    candidates = records
    if shop_name and window_id:
        exact = [
            row
            for row in candidates
            if row["shop_name"] == shop_name and row["window_id"] == window_id
        ]
        if exact:
            return exact[0]
    if shop_name:
        exact = [row for row in candidates if row["shop_name"] == shop_name]
        if exact:
            return exact[0]
    if window_id:
        exact = [row for row in candidates if row["window_id"] == window_id]
        if exact:
            return exact[0]
    return None


def require_shop_config(shop_name="", window_id="", include_ignored=True):
    record = get_shop_config(shop_name, window_id, include_ignored)
    if record:
        return record
    identifier = _text(window_id) or _text(shop_name)
    raise RuntimeError(f"未在数据库比特浏览器配置中找到店铺：{identifier}")


def get_window_id_by_shop_name(shop_name):
    return require_shop_config(shop_name=shop_name)["window_id"]


def _header_column_map(headers):
    normalized = {
        _normalized_header(value): index
        for index, value in enumerate(headers or [])
        if _normalized_header(value)
    }
    result = {}
    for field, aliases in HEADER_ALIASES.items():
        result[field] = next(
            (
                normalized[_normalized_header(alias)]
                for alias in aliases
                if _normalized_header(alias) in normalized
            ),
            None,
        )
    return result


def read_config_excel_for_migration(config_path=None):
    """只供显式迁移命令调用，不参与任何业务运行时读取。"""
    from openpyxl import load_workbook

    path = Path(config_path or (Path(__file__).resolve().parent / "比特配置文件.xlsx"))
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, ())
        columns = _header_column_map(headers)
        if columns["window_id"] is None or columns["shop_name"] is None:
            raise RuntimeError("比特配置文件缺少“窗口ID”或“账号名”表头")

        records = []
        for row in rows:
            record = {}
            for field in CONFIG_FIELDS:
                index = columns[field]
                record[field] = row[index] if index is not None and index < len(row) else ""
            record = {field: _text(value) for field, value in record.items()}
            if not record["window_id"] and not record["shop_name"]:
                continue
            if not record["window_id"] or not record["shop_name"]:
                raise RuntimeError(
                    f"比特配置文件存在缺少窗口ID或账号名的行：{record}"
                )
            records.append(record)
        return records
    finally:
        workbook.close()


def import_config_excel(config_path=None, replace=True):
    records = read_config_excel_for_migration(config_path)
    result = bit_db_api.upsert_bit_browser_configs(records, replace=replace) or {}
    return {
        "source": str(
            Path(config_path or (Path(__file__).resolve().parent / "比特配置文件.xlsx")).resolve()
        ),
        "count": len(records),
        "replace": bool(replace),
        "database_result": result,
    }


def build_command_line_parser():
    parser = argparse.ArgumentParser(description="把比特配置文件一次性导入数据库")
    parser.add_argument(
        "--import-excel",
        dest="config_path",
        default=str(Path(__file__).resolve().parent / "比特配置文件.xlsx"),
        help="待导入的比特配置 Excel 路径",
    )
    parser.add_argument(
        "--merge",
        action="store_true",
        help="合并写入；默认清空配置表后按 Excel 完整替换",
    )
    return parser


def main(argv=None):
    args = build_command_line_parser().parse_args(argv)
    result = import_config_excel(args.config_path, replace=not args.merge)
    print(
        f"比特浏览器配置导入完成：{result['count']} 条，"
        f"模式：{'完整替换' if result['replace'] else '合并更新'}"
    )
    return result


if __name__ == "__main__":
    main()
