import json
import os
import re
from datetime import datetime, timedelta
from pathlib import Path

import pymysql
from openpyxl import load_workbook

# 1. 配置数据库连接信息
config = {
    'host': os.environ.get('MYSQL_HOST', os.environ.get('DB_HOST', '192.168.1.11')),
    'user': os.environ.get('MYSQL_USER', os.environ.get('DB_USER', 'mercado')),
    'password': os.environ.get('MYSQL_PASSWORD', os.environ.get('DB_PASSWORD', 'mercado')),
    'database': os.environ.get('MYSQL_DATABASE', os.environ.get('DB_NAME', 'mercado')),
    'charset': os.environ.get('MYSQL_CHARSET', 'utf8mb4'),
    'port': int(os.environ.get('MYSQL_PORT', os.environ.get('DB_PORT', '3306'))),
    'cursorclass': pymysql.cursors.DictCursor  # 让查询结果以字典形式返回
}
# config = {
#     'host': 'c766667e.natappfree.cc',
#     'user': 'root',
#     'password': 'zzw@951208',
#     'database': 'mercado',
#     'charset': 'utf8mb4',
#     'port': 39181,
#     'cursorclass': pymysql.cursors.DictCursor  # 让查询结果以字典形式返回
# }
# yuming=c766667e.natappfree.cc:39181


def _ensure_column(cursor, table_name, column_name, column_definition):
    cursor.execute(f"SHOW COLUMNS FROM `{table_name}` LIKE %s", (column_name,))
    if cursor.fetchone():
        return
    cursor.execute(f"ALTER TABLE `{table_name}` ADD COLUMN `{column_name}` {column_definition}")


def _parse_number(value):
    text = str(value or "").replace(",", "").strip()
    number_text = "".join(ch for ch in text if ch.isdigit() or ch in ".-")
    try:
        return float(number_text) if number_text else 0
    except ValueError:
        return 0


def mysql_demo():
    # 建立连接
    connection = pymysql.connect(**config)

    try:
        with connection.cursor() as cursor:
            # --- 增 (Create) ---
            sql_insert = "INSERT INTO `users` (`username`, `email`) VALUES (%s, %s)"
            cursor.execute(sql_insert, ("Gemini", "gemini@example.com"))
            print(f"新增记录ID: {cursor.lastrowid}")

            # --- 查 (Read) ---
            sql_select = "SELECT * FROM `users` WHERE `username` = %s"
            cursor.execute(sql_select, ("Gemini",))
            result = cursor.fetchone()
            print(f"查询结果: {result}")

            # --- 改 (Update) ---
            sql_update = "UPDATE `users` SET `email` = %s WHERE `username` = %s"
            cursor.execute(sql_update, ("new_gemini@example.com", "Gemini"))
            print(f"修改行数: {cursor.rowcount}")

            # --- 删 (Delete) ---
            sql_delete = "DELETE FROM `users` WHERE `username` = %s"
            cursor.execute(sql_delete, ("Gemini",))
            print(f"删除行数: {cursor.rowcount}")

        # 核心：涉及写操作（增删改）必须提交事务
        connection.commit()
        print("事务已提交")

    except Exception as e:
        # 发生错误则回滚
        connection.rollback()
        print(f"操作失败，已回滚: {e}")
    finally:
        # 关闭连接
        connection.close()


def insert_task_record(record_list):
    # 建立连接
    connection = pymysql.connect(**config)

    try:
        with connection.cursor() as cursor:
            # --- 增 (Create) ---
            sql_insert = "insert into record (type,name,site,isSuccess,datetime) VALUES (%s,%s,%s,%s,%s)"
            cursor.executemany(sql_insert, record_list)
            print("执行sql成功", sql_insert)

        # 核心：涉及写操作（增删改）必须提交事务
        connection.commit()

    except Exception as e:
        # 发生错误则回滚
        connection.rollback()
        print(f"操作失败，已回滚: {e}")
    finally:
        # 关闭连接
        connection.close()


def inset_reputation_info(reputation_list):
    connection = pymysql.connect(**config)

    try:
        with connection.cursor() as cursor:
            _ensure_column(cursor, "reputation", "取消率", "VARCHAR(255) NULL")
            _ensure_column(cursor, "reputation", "一周流量趋势", "TEXT NULL")
            submit_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            normalized_list = []
            for row in reputation_list:
                row = list(row)
                if len(row) == 6:
                    color, orders, complain, delay, name, site = row
                    row = [
                        name,
                        site,
                        color,
                        orders,
                        complain,
                        delay,
                        "",
                        "",
                        "",
                        "",
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "",
                    ]
                elif len(row) == 10:
                    row = row[:6] + [""] + row[6:] + [""]
                elif len(row) == 11:
                    row = row[:6] + [""] + row[6:]
                elif len(row) > 13:
                    row = row[:13]
                if len(row) < 12:
                    row.extend([""] * (12 - len(row)))
                if len(row) == 12:
                    row.append(submit_time)
                elif len(row) == 13:
                    row[12] = submit_time
                normalized_list.append(row)
            print(f"准备插入声誉记录 {len(normalized_list)} 条，提交时间 {submit_time}")

            # --- 增 (Create) ---
            sql_insert = """
    INSERT INTO reputation (
         店铺名, 站点, 声誉颜色, 总单量, 
        投诉率, 延误率, 取消率, 增加或减少, 近七天变化率,
        系统告警, 更新时间, 一周流量趋势, 提交时间
    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """
            cursor.executemany(sql_insert, normalized_list)
            print("执行sql成功", sql_insert)

        # 核心：涉及写操作（增删改）必须提交事务
        connection.commit()

    except Exception as e:
        # 发生错误则回滚
        connection.rollback()
        print(f"操作失败，已回滚: {e}")
    finally:
        # 关闭连接
        connection.close()


def inset_infraction_info(infraction_list):
    connection = pymysql.connect(**config)

    try:
        with connection.cursor() as cursor:
            normalized_list = []
            submit_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            for row in infraction_list:
                row = list(row)
                if len(row) == 6:
                    row.insert(5, "")
                    row.append("侵权")
                elif len(row) == 7:
                    row.append("侵权")
                if len(row) >= 8:
                    row[5] = submit_time
                normalized_list.append(row)
            submit_time_count = sum(1 for row in normalized_list if len(row) > 5 and row[5])
            print(f"准备插入侵权记录 {len(normalized_list)} 条，提交时间 {submit_time}，非空 {submit_time_count} 条")

            # --- 增 (Create) ---
            sql_insert = """
    INSERT INTO infraction (
         店铺名,站点,编号,标题,侵权时间,提交时间,执行时间,类型

    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
    """
            cursor.executemany(sql_insert, normalized_list)
            print("执行sql成功", sql_insert)

        # 核心：涉及写操作（增删改）必须提交事务
        connection.commit()

    except Exception as e:
        # 发生错误则回滚
        connection.rollback()
        print(f"操作失败，已回滚: {e}")
    finally:
        # 关闭连接
        connection.close()


def _parse_infraction_date(value):
    text = str(value or "").strip()
    if not text:
        return None

    formats = [
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%m/%d/%y",
        "%m/%d/%Y",
        "%d/%m/%y",
        "%d/%m/%Y",
        "%b %d, %Y",
        "%B %d, %Y",
        "%d %b %Y",
        "%d %B %Y",
    ]
    candidates = []
    for fmt in formats:
        try:
            candidates.append(datetime.strptime(text, fmt))
        except ValueError:
            pass

    if not candidates:
        return None

    now = datetime.now()
    not_future = [item for item in candidates if item <= now]
    return max(not_future or candidates)


def _split_config_sites(value):
    text = str(value or "").strip()
    if not text or text.lower() == "nan":
        return []
    parts = re.split(r"[，,、/;\s]+", text)
    return [part.strip() for part in parts if part and part.strip()]


def _is_ignored_config_value(value):
    return "忽略" in str(value or "").strip()


def _load_configured_shop_sites():
    config_path = Path(__file__).resolve().parent / "比特配置文件.xlsx"
    if not config_path.exists():
        return []

    wb = load_workbook(config_path, data_only=True)
    sheet = wb.active
    configured = []
    seen = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        name = str(row[1] or "").strip() if len(row) > 1 else ""
        status = str(row[2] or "").strip() if len(row) > 2 else ""
        sites_text = row[3] if len(row) > 3 else ""
        if not name or _is_ignored_config_value(status):
            continue
        for site in _split_config_sites(sites_text):
            key = (name, site)
            if key in seen:
                continue
            seen.add(key)
            configured.append({"店铺名": name, "站点": site})
    return configured


def _get_latest_infraction_task_status(cursor):
    cursor.execute(
        """
        SELECT `name`, `site`, `isSuccess`, `datetime`
        FROM record
        WHERE `type` = '获取侵权信息'
          AND `name` IS NOT NULL AND `name` <> ''
          AND `site` IS NOT NULL AND `site` <> ''
        ORDER BY `datetime` DESC, `id` DESC
        """
    )
    status_map = {}
    for row in cursor.fetchall():
        key = (str(row.get("name") or "").strip(), str(row.get("site") or "").strip())
        if not key[0] or not key[1] or key in status_map:
            continue
        status_map[key] = {
            "状态": str(row.get("isSuccess") or "").strip() or "未知",
            "状态时间": str(row.get("datetime") or ""),
        }
    return status_map


def get_latest_infraction_info(recent_days=30):
    try:
        recent_days = int(recent_days)
    except (TypeError, ValueError):
        recent_days = 30
    if recent_days not in (7, 30, 100, 365):
        recent_days = 30

    connection = pymysql.connect(**config)

    try:
        with connection.cursor() as cursor:
            _ensure_column(cursor, "reputation", "一周流量趋势", "TEXT NULL")
            cursor.execute(
                """
                SELECT MAX(`提交时间`) AS latest_submit_time
                FROM infraction
                WHERE `提交时间` IS NOT NULL AND `提交时间` <> ''
                """
            )
            latest = cursor.fetchone() or {}
            latest_submit_time = latest.get("latest_submit_time")
            if not latest_submit_time:
                task_status_map = _get_latest_infraction_task_status(cursor)
                summary = []
                for configured_site in _load_configured_shop_sites():
                    key = (configured_site["店铺名"], configured_site["站点"])
                    task_status = task_status_map.get(key) or {}
                    summary.append({
                        "店铺名": key[0],
                        "站点": key[1],
                        "总数": 0,
                        "侵权": 0,
                        "权利人": 0,
                        "状态": task_status.get("状态") or "无数据",
                        "状态时间": task_status.get("状态时间") or "",
                    })
                return {"latest_submit_time": "", "total": 0, "summary": summary, "rows": []}

            cursor.execute(
                """
                SELECT
                    `店铺名`, `站点`, `编号`, `标题`, `侵权时间`,
                    `提交时间`, `执行时间`, `类型`
                FROM infraction
                WHERE `提交时间` = %s
                """,
                (latest_submit_time,),
            )
            rows = cursor.fetchall()
            cutoff = datetime.now() - timedelta(days=recent_days)
            recent_rows = []
            for row in rows:
                infraction_date = _parse_infraction_date(row.get("侵权时间"))
                if infraction_date and infraction_date >= cutoff:
                    row["_侵权时间排序"] = infraction_date
                    recent_rows.append(row)
            recent_rows.sort(
                key=lambda row: row.get("_侵权时间排序") or datetime.min,
                reverse=True,
            )
            for row in recent_rows:
                row.pop("_侵权时间排序", None)

            summary_map = {}
            for row in recent_rows:
                key = (row.get("店铺名") or "", row.get("站点") or "")
                item = summary_map.setdefault(
                    key,
                    {
                        "店铺名": key[0],
                        "站点": key[1],
                        "总数": 0,
                        "侵权": 0,
                        "权利人": 0,
                        "状态": "成功",
                        "状态时间": "",
                    },
                )
                item["总数"] += 1
                infraction_type = row.get("类型") or "侵权"
                if infraction_type == "权利人":
                    item["权利人"] += 1
                else:
                    item["侵权"] += 1

            task_status_map = _get_latest_infraction_task_status(cursor)
            for item in summary_map.values():
                task_status = task_status_map.get((item["店铺名"], item["站点"])) or {}
                item["状态"] = task_status.get("状态") or "成功"
                item["状态时间"] = task_status.get("状态时间") or ""

            for configured_site in _load_configured_shop_sites():
                key = (configured_site["店铺名"], configured_site["站点"])
                if key in summary_map:
                    continue
                task_status = task_status_map.get(key) or {}
                status = task_status.get("状态") or "无数据"
                summary_map[key] = {
                    "店铺名": key[0],
                    "站点": key[1],
                    "总数": 0,
                    "侵权": 0,
                    "权利人": 0,
                    "状态": status,
                    "状态时间": task_status.get("状态时间") or "",
                }

            summary = sorted(
                summary_map.values(),
                key=lambda item: (
                    item["状态"] != "失败",
                    item["总数"],
                    item["侵权"],
                    item["权利人"],
                    item["店铺名"],
                    item["站点"],
                ),
                reverse=True,
            )

            return {
                "latest_submit_time": str(latest_submit_time),
                "latest_total": len(rows),
                "recent_days": recent_days,
                "total": len(recent_rows),
                "summary": summary,
                "rows": recent_rows,
            }
    except Exception as e:
        print(f"查询最新侵权数据失败: {e}")
        raise
    finally:
        connection.close()


def get_latest_reputation_info():
    connection = pymysql.connect(**config)

    try:
        with connection.cursor() as cursor:
            _ensure_column(cursor, "reputation", "取消率", "VARCHAR(255) NULL")
            _ensure_column(cursor, "reputation", "一周流量趋势", "TEXT NULL")
            cursor.execute(
                """
                SELECT MAX(`提交时间`) AS latest_submit_time
                FROM reputation
                WHERE `提交时间` IS NOT NULL
                """
            )
            latest = cursor.fetchone() or {}
            latest_submit_time = latest.get("latest_submit_time")
            if not latest_submit_time:
                return {"latest_submit_time": "", "total": 0, "rows": []}

            cursor.execute(
                """
                SELECT
                    `店铺名`, `站点`, `声誉颜色`, `总单量`, `投诉率`,
                    `延误率`, `取消率`, `增加或减少`, `近七天变化率`,
                    `系统告警`, `更新时间`, `一周流量趋势`, `提交时间`
                FROM reputation
                WHERE `提交时间` = %s
                """,
                (latest_submit_time,),
            )
            rows = cursor.fetchall()
            for row in rows:
                for key in ("更新时间", "提交时间"):
                    if row.get(key) is not None:
                        row[key] = str(row[key])
            rows.sort(
                key=lambda row: (
                    -_parse_number(row.get("总单量")),
                    str(row.get("店铺名") or ""),
                    str(row.get("站点") or ""),
                )
            )
            return {
                "latest_submit_time": str(latest_submit_time),
                "total": len(rows),
                "rows": rows,
            }
    except Exception as e:
        print(f"查询最新声誉数据失败: {e}")
        raise
    finally:
        connection.close()


def insert_orders(line):
    connection = pymysql.connect(**config)

    try:
        with connection.cursor() as cursor:
            # --- 增 (Create) ---
            sql_insert = """
            INSERT INTO orders (
                `id`, `编号`, `时间`, `业务员`, `来源`, `状态`, 
                `金额`, `费用`, `退款`, `人民币收入`, `采购成本`, `采购单号`, 
                `采购追踪`, `利润`, `产品id`, `产品分类`, `标题`, 
                `图片`, `数量`, `订单运费`,`订单备注`, `地区`, `买家姓名`
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            cursor.executemany(sql_insert, line)
            print("执行sql成功", sql_insert)

        # 核心：涉及写操作（增删改）必须提交事务
        connection.commit()

    except Exception as e:
        # 发生错误则回滚
        connection.rollback()
        print(f"操作失败，已回滚: {e}")
    finally:
        # 关闭连接
        connection.close()


def inset_delay_info(delay_list):
    connection = pymysql.connect(**config)

    try:
        with connection.cursor() as cursor:
            # --- 增 (Create) ---
            sql_insert = """
            INSERT INTO `delay` (
    `店铺`, 
    `站点`, 
    `延误率`, 
    `下单时间`, 
    `销售单号`, 
    `订单标题`, 
    `截止延误时间`, 
    `实际揽收时间`, 
    `更新时间`, 
    `文件路径`
) VALUES (
    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
);    
    """
            cursor.executemany(sql_insert, delay_list)
            print("执行sql成功", sql_insert)

        # 核心：涉及写操作（增删改）必须提交事务
        connection.commit()

    except Exception as e:
        # 发生错误则回滚
        connection.rollback()
        print(f"操作失败，已回滚: {e}")
    finally:
        # 关闭连接
        connection.close()


def _ensure_pago_table(cursor):
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS `pago` (
            `id` BIGINT NOT NULL AUTO_INCREMENT,
            `店铺名` VARCHAR(128) NULL,
            `站点` VARCHAR(64) NULL,
            `已释放美元` VARCHAR(64) NULL,
            `未释放美元` VARCHAR(64) NULL,
            `状态` VARCHAR(64) NULL,
            `更新时间` DATETIME NULL,
            `页面原始信息` LONGTEXT NULL,
            `提交时间` DATETIME NULL,
            PRIMARY KEY (`id`),
            KEY `idx_pago_submit_time` (`提交时间`),
            KEY `idx_pago_shop_site` (`店铺名`, `站点`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """
    )


def inset_pago_info(pago_list):
    connection = pymysql.connect(**config)

    try:
        with connection.cursor() as cursor:
            _ensure_pago_table(cursor)
            submit_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            normalized_list = []
            for row in pago_list:
                row = list(row)
                if len(row) < 7:
                    row.extend([""] * (7 - len(row)))
                elif len(row) > 7:
                    row = row[:7]
                row.append(submit_time)
                normalized_list.append(row)

            sql_insert = """
                INSERT INTO `pago` (
                    `店铺名`,
                    `站点`,
                    `已释放美元`,
                    `未释放美元`,
                    `状态`,
                    `更新时间`,
                    `页面原始信息`,
                    `提交时间`
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """
            cursor.executemany(sql_insert, normalized_list)
            print(f"执行sql成功 {sql_insert}，准备插入款项记录 {len(normalized_list)} 条，提交时间 {submit_time}")

        connection.commit()

    except Exception as e:
        connection.rollback()
        print(f"操作失败，已回滚: {e}")
        raise
    finally:
        connection.close()


def _ensure_zying_product_table(cursor):
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS `zying_product` (
            `id` BIGINT NOT NULL AUTO_INCREMENT,
            `产品编号` VARCHAR(128) NULL,
            `分类编号` VARCHAR(64) NULL,
            `产品分类` VARCHAR(2048) NULL,
            `主图链接` TEXT NULL,
            `标题` VARCHAR(1024) NULL,
            `售价` VARCHAR(128) NULL,
            `净收益` VARCHAR(128) NULL,
            `包装毛重` VARCHAR(128) NULL,
            `包装尺寸` VARCHAR(255) NULL,
            `审核状态` VARCHAR(128) NULL,
            `疑似侵权` VARCHAR(8) NULL,
            `采集页码` INT NULL,
            `采集时间` DATETIME NOT NULL,
            `页面原始信息` LONGTEXT NULL,
            `提交时间` DATETIME NOT NULL,
            PRIMARY KEY (`id`),
            KEY `idx_zying_product_collect_time` (`采集时间`),
            KEY `idx_zying_product_id_time` (`产品编号`, `采集时间`),
            KEY `idx_zying_product_review_status` (`审核状态`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """
    )
    _ensure_column(cursor, "zying_product", "分类编号", "VARCHAR(64) NULL")
    _ensure_column(cursor, "zying_product", "产品分类", "VARCHAR(2048) NULL")
    _ensure_column(cursor, "zying_product", "疑似侵权", "VARCHAR(8) NULL")


def insert_zying_product_info(product_list):
    """创建智赢产品表，并将本次采集结果作为快照批量写入。"""
    connection = pymysql.connect(**config)
    try:
        with connection.cursor() as cursor:
            _ensure_zying_product_table(cursor)
            submit_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            normalized_list = []
            for record in product_list or []:
                if isinstance(record, dict):
                    normalized_list.append(
                        (
                            record.get("product_id", record.get("产品编号", "")),
                            record.get(
                                "product_category_id",
                                record.get("分类编号", ""),
                            ),
                            record.get(
                                "product_category",
                                record.get("产品分类", ""),
                            ),
                            record.get("main_image_url", record.get("主图链接", "")),
                            record.get("title", record.get("标题", "")),
                            record.get("sale_price", record.get("售价", "")),
                            record.get("net_income", record.get("净收益", "")),
                            record.get("package_gross_weight", record.get("包装毛重", "")),
                            record.get("package_dimensions", record.get("包装尺寸", "")),
                            record.get("review_status", record.get("审核状态", "")),
                            record.get("page_number", record.get("采集页码")),
                            record.get("collected_at", record.get("采集时间")) or submit_time,
                            record.get("raw_text", record.get("页面原始信息", "")),
                            submit_time,
                        )
                    )
                    continue

                row = list(record)
                if len(row) >= 13:
                    normalized_list.append(tuple(row[:13] + [submit_time]))
                    continue
                if len(row) < 11:
                    row.extend([""] * (11 - len(row)))
                normalized_list.append(
                    tuple([row[0], "", ""] + row[1:11] + [submit_time])
                )

            if normalized_list:
                cursor.executemany(
                    """
                    INSERT INTO `zying_product` (
                        `产品编号`, `分类编号`, `产品分类`, `主图链接`, `标题`, `售价`, `净收益`,
                        `包装毛重`, `包装尺寸`, `审核状态`, `采集页码`,
                        `采集时间`, `页面原始信息`, `提交时间`
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    normalized_list,
                )
        connection.commit()
        return len(normalized_list)
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()


def get_zying_risk_candidates(hours=24, limit=0):
    """读取最近入库的智赢商品，供标题和主图侵权风险检查。"""
    hours = max(1, int(hours))
    limit = max(0, int(limit or 0))
    since = datetime.now() - timedelta(hours=hours)
    connection = pymysql.connect(**config)
    try:
        with connection.cursor() as cursor:
            _ensure_zying_product_table(cursor)
            sql = """
                SELECT
                    `id` AS `row_id`,
                    `产品编号` AS `product_id`,
                    `主图链接` AS `main_image_url`,
                    `标题` AS `title`,
                    `产品分类` AS `product_category`,
                    `提交时间` AS `submitted_at`,
                    `疑似侵权` AS `suspected_infringement`
                FROM `zying_product`
                WHERE `提交时间` >= %s
                  AND COALESCE(`疑似侵权`, '') <> '是'
                ORDER BY `id` ASC
            """
            params = [since.strftime("%Y-%m-%d %H:%M:%S")]
            if limit:
                sql += " LIMIT %s"
                params.append(limit)
            cursor.execute(sql, tuple(params))
            return cursor.fetchall()
    finally:
        connection.close()


def mark_zying_products_suspected(row_ids):
    """把指定 zying_product 数据行的“疑似侵权”字段标记为“是”。"""
    normalized_ids = sorted(
        {
            int(row_id)
            for row_id in (row_ids or [])
            if str(row_id or "").strip().isdigit()
        }
    )
    if not normalized_ids:
        return 0

    connection = pymysql.connect(**config)
    try:
        with connection.cursor() as cursor:
            _ensure_zying_product_table(cursor)
            cursor.executemany(
                "UPDATE `zying_product` SET `疑似侵权` = '是' WHERE `id` = %s",
                [(row_id,) for row_id in normalized_ids],
            )
            updated_count = cursor.rowcount
        connection.commit()
        return updated_count
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()


def insert_chat_info(name, site, message, chat, response, time):
    connection = pymysql.connect(**config)

    try:
        with connection.cursor() as cursor:
            # --- 增 (Create) ---
            sql_insert = """
                INSERT INTO `record_chat` (
        `店铺`, 
        `站点`, 
        `话术`, 
        `客服消息`, 
        `回复`, 
        `时间`
    ) VALUES (
        %s, %s, %s, %s, %s, %s
    );    
        """
            cursor.execute(sql_insert, (name, site, message, chat, response, time))
            chat_id = cursor.lastrowid
            print("执行sql成功", sql_insert)
        connection.commit()
        return chat_id

    except Exception as e:
        # 发生错误则回滚
        connection.rollback()
        print(f"操作失败，已回滚: {e}")
    finally:
        # 关闭连接
        connection.close()


def _ensure_appeal_chat_records_table(cursor):
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS `appeal_chat_records` (
            `id` BIGINT NOT NULL AUTO_INCREMENT,
            `record_time` DATETIME NULL,
            `window_name` VARCHAR(128) NULL,
            `site` VARCHAR(64) NULL,
            `event` VARCHAR(128) NULL,
            `message` LONGTEXT NULL,
            `response` LONGTEXT NULL,
            `chat_json` LONGTEXT NULL,
            `extra_json` LONGTEXT NULL,
            `raw_json` LONGTEXT NULL,
            `created_at` DATETIME NOT NULL,
            PRIMARY KEY (`id`),
            KEY `idx_appeal_chat_time` (`record_time`),
            KEY `idx_appeal_chat_shop_site` (`window_name`, `site`),
            KEY `idx_appeal_chat_event` (`event`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """
    )


def insert_appeal_chat_record(record):
    connection = pymysql.connect(**config)
    try:
        with connection.cursor() as cursor:
            _ensure_appeal_chat_records_table(cursor)
            record = dict(record or {})
            record_time = record.get("time") or None
            sql_insert = """
                INSERT INTO `appeal_chat_records` (
                    `record_time`,
                    `window_name`,
                    `site`,
                    `event`,
                    `message`,
                    `response`,
                    `chat_json`,
                    `extra_json`,
                    `raw_json`,
                    `created_at`
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            cursor.execute(
                sql_insert,
                (
                    record_time,
                    record.get("window", ""),
                    record.get("site", ""),
                    record.get("event", ""),
                    record.get("message", ""),
                    record.get("response", ""),
                    json.dumps(record.get("chat", []), ensure_ascii=False),
                    json.dumps(record.get("extra", {}), ensure_ascii=False),
                    json.dumps(record, ensure_ascii=False),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                ),
            )
            record_id = cursor.lastrowid
        connection.commit()
        return record_id
    except Exception as e:
        connection.rollback()
        print(f"AI申诉聊天记录写入失败，已回滚: {e}")
        raise
    finally:
        connection.close()


def _ensure_ai_appeal_records_table(cursor):
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS `ai_appeal_records` (
            `id` BIGINT NOT NULL AUTO_INCREMENT,
            `appeal_time` DATETIME NULL,
            `appeal_type` VARCHAR(64) NULL,
            `shop_name` VARCHAR(128) NULL,
            `site` VARCHAR(64) NULL,
            `status` VARCHAR(64) NULL,
            `appeal_content` LONGTEXT NULL,
            `identifiers_json` LONGTEXT NULL,
            `success_ids_json` LONGTEXT NULL,
            `failed_ids_json` LONGTEXT NULL,
            `ai_replies_json` LONGTEXT NULL,
            `ai_summary` LONGTEXT NULL,
            `error` LONGTEXT NULL,
            `raw_json` LONGTEXT NULL,
            `created_at` DATETIME NOT NULL,
            PRIMARY KEY (`id`),
            KEY `idx_ai_appeal_time` (`appeal_time`),
            KEY `idx_ai_appeal_shop_site` (`shop_name`, `site`),
            KEY `idx_ai_appeal_type_status` (`appeal_type`, `status`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """
    )


def insert_ai_appeal_record(record):
    connection = pymysql.connect(**config)
    try:
        with connection.cursor() as cursor:
            _ensure_ai_appeal_records_table(cursor)
            record = dict(record or {})
            sql_insert = """
                INSERT INTO `ai_appeal_records` (
                    `appeal_time`,
                    `appeal_type`,
                    `shop_name`,
                    `site`,
                    `status`,
                    `appeal_content`,
                    `identifiers_json`,
                    `success_ids_json`,
                    `failed_ids_json`,
                    `ai_replies_json`,
                    `ai_summary`,
                    `error`,
                    `raw_json`,
                    `created_at`
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            cursor.execute(
                sql_insert,
                (
                    record.get("appeal_time") or None,
                    record.get("appeal_type", ""),
                    record.get("shop_name", ""),
                    record.get("site", ""),
                    record.get("status", ""),
                    record.get("appeal_content", ""),
                    json.dumps(record.get("identifiers", []), ensure_ascii=False),
                    json.dumps(record.get("success_ids", []), ensure_ascii=False),
                    json.dumps(record.get("failed_ids", []), ensure_ascii=False),
                    json.dumps(record.get("ai_replies", []), ensure_ascii=False),
                    record.get("ai_summary", ""),
                    record.get("error", ""),
                    json.dumps(record, ensure_ascii=False),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                ),
            )
            record_id = cursor.lastrowid
        connection.commit()
        return record_id
    except Exception as e:
        connection.rollback()
        print(f"AI申诉汇总记录写入失败，已回滚: {e}")
        raise
    finally:
        connection.close()


def get_ai_appeal_records(limit=100):
    try:
        limit = max(1, min(int(limit), 500))
    except (TypeError, ValueError):
        limit = 100

    connection = pymysql.connect(**config)
    try:
        with connection.cursor() as cursor:
            _ensure_ai_appeal_records_table(cursor)
            cursor.execute(
                """
                SELECT
                    `id`,
                    `appeal_time`,
                    `appeal_type`,
                    `shop_name`,
                    `site`,
                    `status`,
                    `appeal_content`,
                    `identifiers_json`,
                    `success_ids_json`,
                    `failed_ids_json`,
                    `ai_replies_json`,
                    `ai_summary`,
                    `error`,
                    `created_at`
                FROM `ai_appeal_records`
                ORDER BY `appeal_time` DESC, `id` DESC
                LIMIT %s
                """,
                (limit,),
            )
            rows = cursor.fetchall()
            for row in rows:
                for time_key in ("appeal_time", "created_at"):
                    if row.get(time_key) is not None:
                        row[time_key] = str(row[time_key])
                for json_key in ("identifiers_json", "success_ids_json", "failed_ids_json", "ai_replies_json"):
                    value = row.get(json_key)
                    try:
                        row[json_key] = json.loads(value) if value else []
                    except Exception:
                        row[json_key] = []
            return {"total": len(rows), "rows": rows}
    finally:
        connection.close()


def _ensure_window_anomalies_table(cursor):
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS `window_anomalies` (
            `id` BIGINT NOT NULL AUTO_INCREMENT,
            `window_id` VARCHAR(64) NOT NULL,
            `window_name` VARCHAR(128) NULL,
            `site` VARCHAR(64) NULL,
            `anomaly_type` VARCHAR(64) NOT NULL DEFAULT '需要登录',
            `reason` LONGTEXT NULL,
            `source` VARCHAR(64) NULL,
            `active` TINYINT(1) NOT NULL DEFAULT 1,
            `occurrence_count` INT NOT NULL DEFAULT 1,
            `first_detected_at` DATETIME NOT NULL,
            `last_detected_at` DATETIME NOT NULL,
            `resolved_at` DATETIME NULL,
            PRIMARY KEY (`id`),
            UNIQUE KEY `uniq_window_anomaly_window` (`window_id`),
            KEY `idx_window_anomaly_active_time` (`active`, `last_detected_at`),
            KEY `idx_window_anomaly_name` (`window_name`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """
    )


def upsert_window_anomaly(
    window_id,
    window_name,
    site="",
    anomaly_type="需要登录",
    reason="",
    source="bit_daily_task",
):
    window_id = str(window_id or "").strip()
    if not window_id:
        raise ValueError("窗口异常缺少 window_id")
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    connection = pymysql.connect(**config)
    try:
        with connection.cursor() as cursor:
            _ensure_window_anomalies_table(cursor)
            cursor.execute(
                """
                INSERT INTO `window_anomalies` (
                    `window_id`, `window_name`, `site`, `anomaly_type`, `reason`,
                    `source`, `active`, `occurrence_count`, `first_detected_at`,
                    `last_detected_at`, `resolved_at`
                ) VALUES (%s, %s, %s, %s, %s, %s, 1, 1, %s, %s, NULL)
                ON DUPLICATE KEY UPDATE
                    `window_name` = VALUES(`window_name`),
                    `site` = VALUES(`site`),
                    `anomaly_type` = VALUES(`anomaly_type`),
                    `reason` = VALUES(`reason`),
                    `source` = VALUES(`source`),
                    `active` = 1,
                    `occurrence_count` = `occurrence_count` + 1,
                    `last_detected_at` = VALUES(`last_detected_at`),
                    `resolved_at` = NULL
                """,
                (
                    window_id,
                    str(window_name or ""),
                    str(site or ""),
                    str(anomaly_type or "需要登录"),
                    str(reason or ""),
                    str(source or "bit_daily_task"),
                    now,
                    now,
                ),
            )
        connection.commit()
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()


def resolve_window_anomaly(window_id):
    window_id = str(window_id or "").strip()
    if not window_id:
        return 0
    connection = pymysql.connect(**config)
    try:
        with connection.cursor() as cursor:
            _ensure_window_anomalies_table(cursor)
            cursor.execute(
                """
                UPDATE `window_anomalies`
                SET `active` = 0, `resolved_at` = %s
                WHERE `window_id` = %s AND `active` = 1
                """,
                (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), window_id),
            )
            affected = cursor.rowcount
        connection.commit()
        return affected
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()


def get_window_anomalies(active_only=True, limit=500):
    try:
        limit = max(1, min(int(limit), 1000))
    except (TypeError, ValueError):
        limit = 500
    connection = pymysql.connect(**config)
    try:
        with connection.cursor() as cursor:
            _ensure_window_anomalies_table(cursor)
            where = "WHERE `active` = 1" if active_only else ""
            cursor.execute(
                f"""
                SELECT `id`, `window_id`, `window_name`, `site`, `anomaly_type`,
                       `reason`, `source`, `active`, `occurrence_count`,
                       `first_detected_at`, `last_detected_at`, `resolved_at`
                FROM `window_anomalies`
                {where}
                ORDER BY `active` DESC, `last_detected_at` DESC, `id` DESC
                LIMIT %s
                """,
                (limit,),
            )
            rows = cursor.fetchall()
            for row in rows:
                for key in ("first_detected_at", "last_detected_at", "resolved_at"):
                    if row.get(key) is not None:
                        row[key] = str(row[key])
                row["active"] = bool(row.get("active"))
            return {"total": len(rows), "rows": rows}
    finally:
        connection.close()


if __name__ == "__main__":
    mysql_demo()
