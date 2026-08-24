from collections import Counter
from pathlib import Path
import re

from openpyxl import load_workbook


input_dir = Path("outputs/zying_boutique_17074_20260812_run3")
counts = Counter()
samples = []
total = 0
for path in sorted(input_dir.glob("*.xlsx")):
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows = workbook.active.iter_rows(values_only=True)
    headers = [str(value or "") for value in next(rows)]
    for row in rows:
        total += 1
        for index, value in enumerate(row):
            text = str(value or "")
            urls = re.findall(r"https?://[^\s\"'<>]+", text)
            images = [
                url for url in urls
                if re.search(r"(?i)\.(?:jpg|jpeg|png|webp)(?:\?|$)", url)
            ]
            if images:
                counts[headers[index]] += 1
                if len(samples) < 20:
                    samples.append(
                        (path.name, row[0], headers[index], images[:3], text[:500])
                    )
    workbook.close()
print("ROWS", total, "COUNTS", counts)
for sample in samples:
    print(sample)
