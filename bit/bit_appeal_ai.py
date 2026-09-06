"""
美客多 AI 客服申诉自动化脚本。

主要职责：
1. 连接比特浏览器中指定店铺窗口。
2. 切换到指定美客多站点。
3. 打开 Mercado Libre Help 页面里的 AI 客服悬浮窗。
4. 针对延误、侵权、投诉等场景生成申诉话术并发送。
5. 对侵权编号进行分组发送，并根据 AI 回复自动补充站点/误判说明。

适用环境：Python 3 + Selenium + 比特浏览器本地 API。
"""

import time
import traceback
import json
import os
import socket
import sys
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
import math
import uuid
from contextvars import ContextVar
from contextlib import contextmanager
from bit.chat_log import write_local_record


# 允许脚本被直接运行时也能正常导入项目内的 bit、AI_Agent 等包。
# 项目根目录必须排在 bit 目录之前，否则 ``bit_playwright`` 会误解析成
# ``bit/bit_playwright.py``，而不是同名 package。
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
project_root_text = str(PROJECT_ROOT)
script_dir_text = str(SCRIPT_DIR)
for import_path in (project_root_text, script_dir_text):
    while import_path in sys.path:
        sys.path.remove(import_path)
sys.path.insert(0, script_dir_text)
sys.path.insert(0, project_root_text)

import requests
from pydantic.v1.datetime_parse import parse_date
from selenium.webdriver.chrome.service import Service

from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import random

from bit.bit_utils import get_latest_modified_file, get_bit_path, parser_delay_date, get_now_time, getWindowidByName
from bit.bit_api import *
from bit.bit_collection_control import env_int
from bit.bit_runtime_lock import create_window_lease, current_thread_window_lease
from bit.bit_appeal_phrases import (
    get_current_appeal_phrase,
    render_appeal_phrase,
    select_appeal_phrase,
    use_appeal_phrase,
)
from bit.bit_config import (
    get_shop_config,
    get_window_id_by_shop_name as get_config_window_id_by_shop_name,
    list_shop_configs,
)
from bit.bit_mercado_login import (
    is_mercado_login_page as _is_mercado_login_page,
    open_mercado_backend_page,
)
from bit.bit_mercado_limit import MERCADO_RATE_LIMIT_TEXT
from bit.bit_download import download_relay_mail
from bit import mercado_infraction_sync
from bit.bit_db_api import (
    insert_ai_appeal_record,
    list_mercado_prohibited_listings,
    list_mercado_store_tokens,
)
from bit.bit_reputation_info import get_cancellation_orders, get_complaint_orders
import pandas as pd
from datetime import datetime, timedelta
from datetime import datetime
from bit.bit_appeal_state import AppealExecutionError, execution_result, result_from_logs, STATUS_LABELS
from bit.bit_ai_chat_protocol import ChatMessages, read_snapshot, new_messages, normalized_text
import re
from openpyxl import load_workbook
import traceback

try:
    from bit.chat_log import (
        append_chat_log,
        get_appeal_log_records,
        start_appeal_log_collection,
        stop_appeal_log_collection,
    )
except Exception:
    from chat_log import (
        append_chat_log,
        get_appeal_log_records,
        start_appeal_log_collection,
        stop_appeal_log_collection,
    )

# 聊天记录入库接口；AI 与人工客服回复都会通过这个接口记录。
CHAT_INFO_API_URL = "https://zeshun.nat100.top/api/v1/chat"

# 美客多帮助中心入口，AI 客服悬浮窗通常挂在这些页面中。
HELP_URL = "https://global-selling.mercadolibre.com/help"

# AI 悬浮窗 iframe 的特征。不同账号/页面版本的 src/title 可能不同，所以这里保留多个标记。
AI_FRAME_URL_MARKERS = ("meli-ai-chat", "maxwell/new-chat")
AI_FRAME_MARKERS = ("meli-ai-chat", "maxwell", "new-chat", "ai chat", "assistant", "chat", "meli")
AI_CHAT_MODE_INLINE = "inline_dom"
AI_CHAT_MODE_IFRAME = "legacy_iframe"
AI_BACKEND_SETTLE_SECONDS = env_int(
    "BIT_DAILY_BACKEND_SETTLE_SECONDS",
    12,
    minimum=5,
)
AI_CHAT_READY_TIMEOUT_SECONDS = env_int(
    "BIT_DAILY_AI_CHAT_READY_TIMEOUT_SECONDS",
    45,
    minimum=15,
)
AI_CHAT_ENTRY_TIMEOUT_SECONDS = env_int(
    "BIT_DAILY_AI_CHAT_ENTRY_TIMEOUT_SECONDS",
    30,
    minimum=12,
)
AI_CHAT_INPUT_TIMEOUT_SECONDS = env_int(
    "BIT_DAILY_AI_CHAT_INPUT_TIMEOUT_SECONDS",
    45,
    minimum=15,
)
AI_AGENT_REPLY_TIMEOUT_SECONDS = env_int(
    "BIT_DAILY_AI_REPLY_TIMEOUT_SECONDS",
    300,
    minimum=180,
)
AI_AGENT_REPLY_POLL_SECONDS = 2
AI_SEND_CONFIRM_TIMEOUT_SECONDS = env_int("BIT_AI_SEND_CONFIRM_TIMEOUT_SECONDS", 30, minimum=5)
AI_SITE_BUDGET_SECONDS = env_int("BIT_AI_SITE_BUDGET_SECONDS", 7200, minimum=60)
AI_GROUPS_PER_CONVERSATION = env_int("BIT_AI_GROUPS_PER_CONVERSATION", 3, minimum=1)
_APPEAL_STOP_EVENT = ContextVar("appeal_stop_event", default=None)


@contextmanager
def appeal_controls(stop_event):
    token = _APPEAL_STOP_EVENT.set(stop_event)
    try:
        yield
    finally:
        _APPEAL_STOP_EVENT.reset(token)

AI_HELP_URLS = (
    HELP_URL,
    "https://global-selling.mercadolibre.com/help/v2",
)
SITE_OPTION_MENU_OPTIONS = (
    "Mexico (Direct to consumer)",
    "Mexico (Fulfillment)",
    "Brazil",
    "Chile",
    "Colombia",
    "Argentina",
    "Uruguay",
)

CANCELLATION_DEFAULT_APPEAL_TEMPLATE = """尊敬的平台审核专员：

1. 订单编号：{order_ids}

2. 订单取消原因：页面显示【Mercado Libre取消的包裹，我们已取消此交易】，本次订单为平台系统主动取消交易，并非我方卖家主动发起订单取消。

3. 订单节点说明：该订单产生时，我方商品链接正常在售、库存充足、已经备好货物、完全按平台时效要求准备安排发货，不存在缺货、超时、虚假发货等任何卖家违规行为。

4. 诉求：本次交易取消责任完全不在我方，本次不良记录严重影响我方店铺信誉与店铺评分，现正式申诉，恳请平台核实系统后台记录，撤销本次订单的负面处罚。"""

COMPLAINT_DEFAULT_APPEAL_MESSAGE = (
    "亲爱的客服，我叫 Jack，这个产品没有任何证据证明产品有质量问题，"
    "这是买家想白嫖，能帮我消除对我声誉的影响吗？"
)
COMPLAINT_GROUP_SIZE = 2

# 将中文站点名、平台代码统一转成内部站点码，便于回复 AI 的站点确认问题。
SITE_CODE_MAP = {
    "MX": "MX",
    "MLM": "MX",
    "墨西哥": "MX",
    "BR": "BR",
    "MLB": "BR",
    "巴西": "BR",
    "AR": "AR",
    "MLA": "AR",
    "阿根廷": "AR",
    "CL": "CL",
    "MLC": "CL",
    "智利": "CL",
    "CO": "CO",
    "MCO": "CO",
    "哥伦比亚": "CO",
    "UY": "UY",
    "MLU": "UY",
    "Uruguay": "UY",
    "乌拉圭": "UY",
}

SITE_CODE_TO_NAME = {
    "MX": "墨西哥",
    "BR": "巴西",
    "CO": "哥伦比亚",
    "CL": "智利",
    "AR": "阿根廷",
    "UY": "乌拉圭",
}

# 美客多顶部站点切换器中各站点对应的 data-value。
SITE_SWITCH_SELECTOR_MAP = {
    "墨西哥": 'div[data-value="MLM-remote"]',
    "巴西": 'div[data-value="MLB-remote"]',
    "哥伦比亚": 'div[data-value="MCO-remote"]',
    "智利": 'div[data-value="MLC-remote"]',
    "阿根廷": 'div[data-value="MLA-remote"]',
    "乌拉圭": 'div[data-value="MLU-remote"]',
}

SITE_REMOTE_VALUE_MAP = {
    "墨西哥": "MLM-remote",
    "巴西": "MLB-remote",
    "哥伦比亚": "MCO-remote",
    "智利": "MLC-remote",
    "阿根廷": "MLA-remote",
    "乌拉圭": "MLU-remote",
}

SITE_ID_MAP = {
    "墨西哥": "MLM",
    "巴西": "MLB",
    "哥伦比亚": "MCO",
    "智利": "MLC",
    "阿根廷": "MLA",
    "乌拉圭": "MLU",
}

SITE_SHORT_CODE_MAP = {
    "墨西哥": "MX",
    "巴西": "BR",
    "哥伦比亚": "CO",
    "智利": "CL",
    "阿根廷": "AR",
    "乌拉圭": "UY",
}

SITE_LABEL_MAP = {
    "墨西哥": ("Mexico", "México", "墨西哥", "MLM"),
    "巴西": ("Brazil", "Brasil", "巴西", "MLB"),
    "哥伦比亚": ("Colombia", "哥伦比亚", "MCO"),
    "智利": ("Chile", "智利", "MLC"),
    "阿根廷": ("Argentina", "阿根廷", "MLA"),
    "乌拉圭": ("Uruguay", "乌拉圭", "MLU"),
}


def insert_chat_info_by_api(name, site, message, chat, response, time):
    """把客服对话记录通过公网接口写入数据库。"""
    payload = {
        "name": name,
        "site": site,
        "message": message,
        "chat": chat,
        "response": response,
        "time": time
    }
    res = requests.post(CHAT_INFO_API_URL, json=payload, timeout=10)
    res.raise_for_status()
    return res.json()


def connect_bit_browser(window_id):
    """打开指定比特浏览器窗口，并通过 Selenium 连接到该窗口的调试端口。"""
    last_res = None
    max_attempts = 5
    for attempt in range(1, max_attempts + 1):
        try:
            res = openBrowser(window_id)
        except Exception as e:
            last_res = {"success": False, "msg": str(e)}
            print(
                f"{get_now_time()} 打开比特浏览器失败，"
                f"第 {attempt}/{max_attempts} 次: {e}<br>"
            )
            if attempt < max_attempts:
                time.sleep(8 if "频率" in str(e) or "正在打开" in str(e) else 3)
            continue

        print(res)
        last_res = res
        data = res.get("data") if isinstance(res, dict) else None
        if (
            isinstance(res, dict)
            and res.get("success") is not False
            and isinstance(data, dict)
            and data.get("driver")
            and data.get("http")
        ):
            driver_path = data["driver"]
            debugger_address = data["http"]

            chrome_options = webdriver.ChromeOptions()
            chrome_options.add_experimental_option("debuggerAddress", debugger_address)

            chrome_service = Service(driver_path)
            driver = webdriver.Chrome(service=chrome_service, options=chrome_options)
            driver.implicitly_wait(0)
            driver.set_page_load_timeout(60)
            driver.set_script_timeout(30)
            return driver, res

        print(
            f"{get_now_time()} 比特浏览器返回异常，"
            f"第 {attempt}/{max_attempts} 次: {res}<br>"
        )
        if attempt < max_attempts:
            response_text = str(res or "")
            retry_delay = (
                8
                if "降低接口请求频率" in response_text or "正在打开" in response_text
                else 3
            )
            time.sleep(retry_delay)

    raise RuntimeError(f"打开比特浏览器失败，窗口ID={window_id}，返回={last_res}")


def is_mercado_login_required_page(driver):
    """兼容旧调用点，统一使用共享的 Mercado 登录页识别。"""
    return _is_mercado_login_page(driver)


def _abort_ai_appeal_after_backend_recovery(
    result,
    name="",
    site="",
    abort_after_rate_limit_recovery=True,
):
    """校验业务页恢复结果；自动登录成功时继续当前申诉。

    ``open_mercado_backend_page`` 会在未登录时使用店铺授权邮箱和浏览器保存的
    默认密码自动登录，并在成功后重新打开原业务页。因此 ``login_retry_count``
    大于零且状态为 ``ready`` 是可继续的成功状态，不能再被任务模块当成失败。
    限频恢复仍可按调用方策略结束本次浏览器任务。
    """
    if not isinstance(result, dict):
        raise RuntimeError(f"{name} {site} 美客多后台返回无效结果：{result}")

    status = str(result.get("status") or "").strip()
    message = str(result.get("message") or status or "美客多后台不可用").strip()
    rate_limit_retry_count = int(result.get("rate_limit_retry_count") or 0)
    login_retry_count = int(result.get("login_retry_count") or 0)

    if status == "rate_limited" or (
        abort_after_rate_limit_recovery and rate_limit_retry_count
    ):
        raise RuntimeError(
            f"{name} {site} 检测到访问限频（{MERCADO_RATE_LIMIT_TEXT}），"
            "已终止自动找客服并准备关闭浏览器"
        )
    if status == "logged_out":
        raise RuntimeError(
            f"{name} {site} 检测到登录态失效，自动登录未成功："
            f"{message}"
        )
    if not result.get("ok"):
        raise RuntimeError(f"{name} {site} 窗口页面打开验证失败：{message}")
    if login_retry_count:
        print(
            f"{get_now_time()} {name} {site} 自动登录成功，"
            "已重新打开原业务页，继续当前申诉<br>"
        )
    return result


def open_help_page_with_daily_validation(
    driver,
    name="",
    site="",
    max_hongkong_switches=3,
    switch_wait_seconds=8,
    window_id="",
    abort_after_rate_limit_recovery=None,
):
    """打开帮助页，统一处理限频、退出登录和页面有效性。"""
    result = open_mercado_backend_page(
        driver,
        HELP_URL,
        name,
        window_id,
        settle_seconds=AI_BACKEND_SETTLE_SECONDS,
        max_rate_limit_retries=max_hongkong_switches,
        rate_limit_retry_wait_seconds=switch_wait_seconds,
        anomaly_site=site,
        anomaly_source="AI申诉",
    )
    if abort_after_rate_limit_recovery is None:
        abort_after_rate_limit_recovery = bool(
            getattr(driver, "_bit_abort_ai_after_rate_limit_recovery", False)
        )
    _abort_ai_appeal_after_backend_recovery(
        result,
        name,
        site,
        abort_after_rate_limit_recovery=abort_after_rate_limit_recovery,
    )

    state = result.get("state") or {}
    current_url = str(state.get("current_url") or "").strip()
    has_page_content = bool(
        str(state.get("page_text") or "").strip()
        or str(state.get("title") or "").strip()
        or str(state.get("page_source") or "").strip()
    )
    if not current_url or current_url == "about:blank" or not has_page_content:
        raise RuntimeError(
            f"{name} {site} 窗口页面未正常打开：url={current_url or 'empty'}"
        )
    print(f"{get_now_time()} {name} {site} 窗口打开验证通过：{current_url}<br>")
    return True


def close_current_tab_keep_browser(driver, name="", site=""):
    """关闭当前标签页，但保留 BitBrowser 窗口本身。"""
    if not driver:
        return False
    try:
        driver.switch_to.default_content()
    except Exception:
        pass

    try:
        handles = list(driver.window_handles)
        current = driver.current_window_handle
        if len(handles) <= 1:
            driver.execute_script("window.open('about:blank', '_blank');")
            time.sleep(0.5)
            handles = list(driver.window_handles)
            driver.switch_to.window(current)

        driver.close()
        remaining = [handle for handle in driver.window_handles if handle != current]
        if remaining:
            driver.switch_to.window(remaining[-1])
        print(f"{get_now_time()} {name}{site} 已关闭当前标签页<br>")
        return True
    except Exception as e:
        print(f"{get_now_time()} {name}{site} 关闭当前标签页失败：{e}<br>")
        return False


def get_window_id_by_shop_name(name):
    """从已开启申诉的店铺授权匹配 BitBrowser 窗口 ID。"""
    return get_config_window_id_by_shop_name(
        name,
        authorization_flag="appeal_enabled",
    )


def select_site(driver, name, site):
    """在美客多全球销售后台顶部站点切换器中切换到指定站点。"""
    site_name = normalize_site_name(site)
    if select_mercado_site_fast(driver, name, site_name):
        return True
    if select_mercado_site_by_cookie(driver, name, site_name):
        return True

    try:
        WebDriverWait(driver, AI_CHAT_ENTRY_TIMEOUT_SECONDS).until(
            EC.element_to_be_clickable((By.CLASS_NAME, "nav-header-cbt__site-switcher"))
        ).click()
        path = SITE_SWITCH_SELECTOR_MAP.get(site_name, 'div[data-value="MLM-remote"]')
        WebDriverWait(driver, AI_CHAT_ENTRY_TIMEOUT_SECONDS).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, path))
        ).click()
        driver.refresh()
        matched = False
        for _ in range(AI_CHAT_READY_TIMEOUT_SECONDS):
            time.sleep(1)
            if verify_selected_site(driver, site_name):
                matched = True
                break
        if not matched:
            raise RuntimeError(f"选择后页面严格校验站点不匹配，state={get_selected_site_state(driver)}")
        print(f"{get_now_time()} {name} {site_name} '选择站点成功'<br>")
        return True
    except Exception as e:
        print(f"{get_now_time()} {name} {site_name} '选择站点失败': {e}<br>")
        raise RuntimeError(f"{name} 切换站点失败：目标={site_name}, 原始参数={site}") from e


def get_selected_site_state(driver):
    """读取顶部站点切换器的真实当前站点，不扫描整页普通文本。"""
    try:
        return driver.execute_script(
            """
            function allElements(root) {
                const out = [];
                const walk = (node) => {
                    const elements = node.querySelectorAll ? Array.from(node.querySelectorAll('*')) : [];
                    for (const el of elements) {
                        out.push(el);
                        if (el.shadowRoot) walk(el.shadowRoot);
                    }
                };
                walk(root || document);
                return out;
            }

            function textOf(selector) {
                const node = document.querySelector(selector);
                return node ? (node.innerText || node.textContent || '').trim() : '';
            }

            const currentNode = document.querySelector('.nav-header-cbt__current, .nav-header-cbt__site-switcher, [class*="site-switcher"]');
            const currentText = currentNode ? (currentNode.innerText || currentNode.textContent || '').replace(/\\s+/g, ' ').trim() : '';
            const currentFlag = document.querySelector('.nav-header-cbt__current-flag img, .nav-header-cbt__current img, [class*="current"] img');
            const currentFlagAlt = currentFlag ? (currentFlag.getAttribute('alt') || currentFlag.getAttribute('title') || '') : '';
            const selected = document.querySelector('#nav-header-cbt__switcher [class*="option-selected"], [data-value$="-remote"][class*="selected"]');
            const selectedRemote = selected ? (selected.getAttribute('data-value') || selected.querySelector('[data-value]')?.getAttribute('data-value') || '') : '';
            const selectedText = selected ? (selected.innerText || selected.textContent || selected.getAttribute('aria-label') || '').replace(/\\s+/g, ' ').trim() : '';
            const available = [...document.querySelectorAll('#nav-header-cbt__switcher [data-value$="-remote"], [class*="site-switcher"] [data-value$="-remote"]')]
                .map((node) => ({
                    value: node.getAttribute('data-value') || '',
                    text: (node.innerText || node.textContent || node.getAttribute('alt') || '').replace(/\\s+/g, ' ').trim(),
                    selected: String(node.className || '').includes('selected')
                }))
                .filter((item, index, arr) => item.value && arr.findIndex((other) => other.value === item.value) === index);
            const scriptsText = [...document.scripts].map((script) => script.textContent || '').join('\\n');
            const operatingMatch = scriptsText.match(/operating_site_id["']?\\s*:\\s*["']([A-Z]{3})["']/);
            const siteIdMatch = scriptsText.match(/"siteId"\\s*:\\s*"([A-Z]{3})"/);
            const cookieMatch = document.cookie.match(/(?:^|;\\s*)cbtSiteId=([^;]+)/);
            return {
                currentShort: textOf('.nav-header-cbt__current-site'),
                currentText,
                currentFlagAlt,
                selectedRemote,
                selectedText,
                operatingSiteId: operatingMatch ? operatingMatch[1] : '',
                siteId: siteIdMatch ? siteIdMatch[1] : '',
                cookieRemote: cookieMatch ? decodeURIComponent(cookieMatch[1]) : '',
                available,
                url: location.href,
                title: document.title
            };
            """
        ) or {}
    except Exception as e:
        return {"error": str(e)}


def _site_state_matches(state, site):
    state = state or {}
    site_name = normalize_site_name(site)
    remote_value = SITE_REMOTE_VALUE_MAP.get(site_name, "")
    site_id = SITE_ID_MAP.get(site_name, "")
    short_code = SITE_SHORT_CODE_MAP.get(site_name, "")
    labels = tuple(label.lower() for label in SITE_LABEL_MAP.get(site_name, ()))

    selected_remote = str(state.get("selectedRemote") or "").strip()
    operating_site_id = str(state.get("operatingSiteId") or "").strip().upper()
    current_short = str(state.get("currentShort") or "").strip().upper()
    page_site_id = str(state.get("siteId") or "").strip().upper()
    cookie_remote = str(state.get("cookieRemote") or "").strip()
    selected_text = str(state.get("selectedText") or "").lower()
    current_text = str(state.get("currentText") or "").lower()
    current_flag_alt = str(state.get("currentFlagAlt") or "").lower()
    # Reject contradictory explicit signals even if one of them matches.
    for actual, expected in ((selected_remote, remote_value),
                             (operating_site_id, site_id),
                             (current_short, short_code),
                             (page_site_id if page_site_id != "CBT" else "", site_id)):
        if actual and expected and actual != expected:
            return False
    explicit_match = any(
        [
            remote_value and selected_remote == remote_value,
            site_id and operating_site_id == site_id,
            short_code and current_short == short_code,
            site_id and page_site_id == site_id,
            labels and any(label and label in current_flag_alt for label in labels),
            labels and any(label and label in selected_text for label in labels),
            short_code and current_text == short_code.lower(),
        ]
    )
    if explicit_match:
        return True

    # Help 页面有时不渲染站点切换器，siteId 也只返回通用的 CBT。
    # 此时 cbtSiteId 是页面上唯一可用的当前站点证据；若存在任何明确页面状态，
    # 仍然优先相信页面，避免用陈旧 cookie 覆盖真实站点。
    has_explicit_state = any(
        [
            selected_remote,
            operating_site_id,
            current_short,
            current_text,
            current_flag_alt,
            selected_text,
            page_site_id not in ("", "CBT"),
        ]
    )
    return bool(
        not has_explicit_state
        and remote_value
        and cookie_remote == remote_value
    )


def verify_selected_site(driver, site):
    site_name = normalize_site_name(site)
    state = get_selected_site_state(driver)
    matched = _site_state_matches(state, site_name)
    if not matched:
        print(f"{get_now_time()} 站点严格校验失败：目标={site_name}，当前状态={state}<br>")
    return matched


def select_mercado_site_by_cookie(driver, name, site):
    """直接写入美客多站点 cookie 后刷新，作为点击切换失败时的兜底。"""
    site_name = normalize_site_name(site)
    remote_value = SITE_REMOTE_VALUE_MAP.get(site_name, "MLM-remote")
    state = get_selected_site_state(driver)
    available = state.get("available") or []
    if available and not any(item.get("value") == remote_value for item in available):
        print(f"{get_now_time()} {name} {site_name} 当前店铺站点列表不包含目标：{remote_value}，available={available}<br>")
        return False

    try:
        print(f"{get_now_time()} {name} {site_name} 尝试通过 cbtSiteId cookie 切换站点：{remote_value}<br>")
        try:
            driver.delete_cookie("cbtSiteId")
        except Exception:
            pass
        try:
            driver.add_cookie({"name": "cbtSiteId", "value": remote_value, "path": "/"})
        except Exception:
            pass
        driver.execute_script(
            """
            const value = arguments[0];
            document.cookie = `cbtSiteId=${value}; path=/`;
            """,
            remote_value,
        )
        driver.refresh()
        for _ in range(AI_CHAT_READY_TIMEOUT_SECONDS):
            time.sleep(1)
            if verify_selected_site(driver, site_name):
                print(f"{get_now_time()} {name} {site_name} cookie 切换站点成功<br>")
                return True
        print(f"{get_now_time()} {name} {site_name} cookie 切换后仍未通过严格校验，state={get_selected_site_state(driver)}<br>")
        return False
    except Exception as e:
        print(f"{get_now_time()} {name} {site_name} cookie 切换站点异常：{e}<br>")
        return False


def select_mercado_site_fast(driver, name, site):
    """快速切换美客多站点，优先用 JS 深度查找，减少重复重试。"""
    site_name = normalize_site_name(site)
    if verify_selected_site(driver, site_name):
        print(f"{get_now_time()} {name} {site_name} 当前站点严格校验已匹配，无需切换<br>")
        return True

    remote_value = SITE_REMOTE_VALUE_MAP.get(site_name, "MLM-remote")
    labels = SITE_LABEL_MAP.get(site_name, SITE_LABEL_MAP["墨西哥"])
    result = driver.execute_script(
        """
        const remoteValue = arguments[0];
        const labels = arguments[1].map(item => String(item).toLowerCase());

        function allElements(root) {
            const out = [];
            const walk = (node) => {
                const elements = node.querySelectorAll ? Array.from(node.querySelectorAll('*')) : [];
                for (const el of elements) {
                    out.push(el);
                    if (el.shadowRoot) walk(el.shadowRoot);
                }
            };
            walk(root || document);
            return out;
        }

        function visible(el) {
            const rect = el.getBoundingClientRect();
            return rect.width > 0 && rect.height > 0;
        }

        function textOf(el) {
            return [
                el.innerText || '',
                el.textContent || '',
                el.getAttribute('aria-label') || '',
                el.getAttribute('title') || '',
                el.getAttribute('data-value') || ''
            ].join(' ').toLowerCase();
        }

        const elements = allElements(document);
        const targetByValue = elements.find(el =>
            visible(el) && (
                el.getAttribute('data-value') === remoteValue ||
                el.querySelector?.(`[data-value="${remoteValue}"]`)
            )
        );
        if (targetByValue) {
            const target = targetByValue.getAttribute('data-value') === remoteValue
                ? targetByValue
                : targetByValue.querySelector(`[data-value="${remoteValue}"]`);
            const clickable = target.closest('li, button, a, [role="option"], [role="menuitem"], [class*="option-switcher"]') || target;
            clickable.scrollIntoView({block: 'center', inline: 'center'});
            clickable.click();
            for (const type of ['pointerdown', 'mousedown', 'pointerup', 'mouseup', 'click']) {
                clickable.dispatchEvent(new MouseEvent(type, {bubbles: true, cancelable: true, view: window}));
            }
            return `clicked_data_value:${remoteValue}`;
        }

        const switcher = elements.find(el =>
            visible(el) && (
                String(el.className || '').includes('nav-header-cbt__site-switcher') ||
                String(el.className || '').includes('nav-header-cbt__trigger') ||
                el.id === 'nav-header-cbt__logged' ||
                /select\\s+(country|site)|country|site/i.test(textOf(el))
            )
        );
        if (switcher) {
            switcher.click();
        }

        const switcherRoot = document.querySelector('#nav-header-cbt__switcher, #nav-header-cbt__logged-options, [class*="site-switcher"]') || document;
        const target = allElements(switcherRoot).filter(visible).find(el =>
            visible(el) && labels.some(label => textOf(el).includes(label))
        );
        if (!target) {
            return switcher ? 'opened_no_target' : 'no_switcher';
        }
        const clickable = target.closest('li, button, a, [role="option"], [role="menuitem"], [class*="option-switcher"]') || target;
        clickable.scrollIntoView({block: 'center'});
        clickable.click();
        for (const type of ['pointerdown', 'mousedown', 'pointerup', 'mouseup', 'click']) {
            clickable.dispatchEvent(new MouseEvent(type, {bubbles: true, cancelable: true, view: window}));
        }
        return 'clicked_text_in_switcher';
        """,
        remote_value,
        list(labels),
    )
    print(f"{get_now_time()} {name} {site_name} 快速选择站点结果：{result}<br>")
    if str(result or "").startswith("clicked"):
        time.sleep(3)
        driver.refresh()
        for _ in range(AI_CHAT_READY_TIMEOUT_SECONDS):
            time.sleep(1)
            if verify_selected_site(driver, site_name):
                return True
        print(f"{get_now_time()} {name} {site_name} 快速选择后严格校验失败<br>")
    return False


def build_appeal_message(window_id, name, site, form, message, nickname):
    """根据申诉类型构造首条发送给 AI 客服的申诉话术。"""
    if message:
        return message

    words = []
    if form == "延误":
        orders_random = get_delay_orders_download_random(window_id, name, site, 5)
        if orders_random == "":
            return ""
        selected_phrase = get_current_appeal_phrase()
        if selected_phrase:
            return render_appeal_phrase(
                selected_phrase,
                nickname=nickname,
                order_ids=orders_random,
                appeal_type=form,
            )
        words = [
            f"亲爱的客服，我叫{nickname}！这些订单因合作物流车辆临时出现故障，导致未能及时揽收，并非我这边发货延误，麻烦您帮忙处理一下，消除对店铺声誉的影响，非常感谢！",
            f"亲爱的客服，我叫{nickname}！这些订单因为菜鸟物流原因，并非我这边发货延误，麻烦您帮忙处理一下，消除对店铺声誉的影响，非常感谢！",
        ]
        return orders_random + random.choice(words)

    if form == "侵权":
        infraction_random = get_infraction_orders_random(window_id, name, site, 10)
        selected_phrase = get_current_appeal_phrase()
        if selected_phrase:
            return render_appeal_phrase(
                selected_phrase,
                nickname=nickname,
                order_ids=infraction_random,
                appeal_type=form,
            )
        words = [
            f"亲爱的客服，我叫{nickname}！这些产品是通用品牌产品，被系统误检测为侵权产品，你能帮我核查并消除记录吗？",
            f"亲爱的客服，我叫{nickname}！这些产品是通用产品，并没有侵犯品牌权益，麻烦你帮我重新审核并恢复产品，谢谢！",
        ]
        return infraction_random + random.choice(words)

    if form == "投诉":
        selected_phrase = get_current_appeal_phrase()
        if selected_phrase:
            return render_appeal_phrase(
                selected_phrase,
                nickname=nickname,
                appeal_type=form,
            )
        return COMPLAINT_DEFAULT_APPEAL_MESSAGE

    return message


def get_frame_info(driver, frame):
    """读取 iframe 的关键属性与位置，用来判断它是不是 AI 客服悬浮窗。"""
    return driver.execute_script(
        """
        const frame = arguments[0];
        const rect = frame.getBoundingClientRect();
        return {
            src: frame.getAttribute('src') || '',
            title: frame.getAttribute('title') || '',
            name: frame.getAttribute('name') || '',
            id: frame.getAttribute('id') || '',
            cls: String(frame.className || ''),
            top: rect.top,
            left: rect.left,
            right: rect.right,
            bottom: rect.bottom,
            width: rect.width,
            height: rect.height,
            visible: !!(rect.width || rect.height || frame.getClientRects().length)
        };
        """,
        frame,
    )


def is_ai_frame_info(info):
    """根据 iframe 的 src/title/name/id/class 等文本特征识别 AI 客服 iframe。"""
    text = " ".join(str(info.get(key, "")) for key in ("src", "title", "name", "id", "cls")).lower()
    return any(marker in text for marker in AI_FRAME_MARKERS)


def find_frames_including_shadow_dom(driver):
    """返回当前文档及所有开放 Shadow Root 中的 iframe。

    2026-08 版 seller-assistant loader 会把 ``#sa-assistant-chat`` 和真正的
    Maxwell iframe 渲染进 ``#sof-seller-assistant-frm-host`` 的 Shadow Root。
    Selenium 的普通 ``find_elements`` 和 ``document.querySelectorAll`` 都不会穿透
    Shadow DOM，因此必须显式递归开放的 Shadow Root。
    """
    try:
        return driver.execute_script(
            """
            function deepElements(root = document) {
                const out = [];
                const walk = (node) => {
                    const elements = node.querySelectorAll ? Array.from(node.querySelectorAll('*')) : [];
                    for (const el of elements) {
                        out.push(el);
                        if (el.shadowRoot) walk(el.shadowRoot);
                    }
                };
                walk(root);
                return out;
            }
            return deepElements(document).filter((el) =>
                String(el.tagName || '').toUpperCase() === 'IFRAME'
            );
            """
        ) or []
    except Exception:
        # 旧浏览器或测试替身不支持 execute_script 时保留原定位方式。
        try:
            return driver.find_elements(By.TAG_NAME, "iframe")
        except Exception:
            return []


def classify_ai_chat_variant(state):
    """根据页面探测结果区分新版内嵌助手和旧版 iframe 助手。"""
    state = state or {}
    if state.get("legacy_frame_count", 0):
        return AI_CHAT_MODE_IFRAME
    if state.get("inline_shell"):
        return AI_CHAT_MODE_INLINE
    return ""


def get_ai_chat_dom_state(driver):
    """探测 AI 助手页面结构，不点击页面，也不依赖界面语言。"""
    try:
        driver.switch_to.default_content()
    except Exception:
        pass

    try:
        return driver.execute_script(
            """
            const frameMarkers = arguments[0] || [];

            function deepElements(root = document) {
                const out = [];
                const walk = (node) => {
                    const elements = node.querySelectorAll ? Array.from(node.querySelectorAll('*')) : [];
                    for (const el of elements) {
                        out.push(el);
                        if (el.shadowRoot) walk(el.shadowRoot);
                    }
                };
                walk(root);
                return out;
            }

            function visible(el) {
                if (!el) return false;
                const rect = el.getBoundingClientRect();
                const style = window.getComputedStyle(el);
                return !!(rect.width || rect.height || el.getClientRects().length)
                    && style.visibility !== 'hidden'
                    && style.display !== 'none';
            }

            function isTextbox(el) {
                if (!el || !visible(el) || el.disabled) return false;
                const tag = String(el.tagName || '').toUpperCase();
                const role = (el.getAttribute('role') || '').toLowerCase();
                return tag === 'TEXTAREA' || tag === 'INPUT' || role === 'textbox'
                    || el.isContentEditable || el.getAttribute('contenteditable') === 'true';
            }

            const all = deepElements(document);
            const root = all.find((el) => el.id === 'sa-assistant-chat') || null;
            const opener = all.find((el) =>
                el.id === 'sa-icon-button-wrapper'
                || el.getAttribute('aria-controls') === 'sa-assistant-chat'
            ) || null;
            const inlineInput = root
                ? deepElements(root).find((el) => isTextbox(el)) || null
                : null;
            const rootClass = root ? String(root.className || '') : '';
            const ariaExpanded = opener ? (opener.getAttribute('aria-expanded') || '') : '';
            const rootClasses = rootClass.split(/\\s+/).filter(Boolean);
            const inlineOpen = !!inlineInput || ariaExpanded === 'true'
                || (!!root && rootClasses.includes('show') && !rootClasses.includes('minimized'));

            // 新版 loader 把 Maxwell iframe 放在开放的 Shadow Root 中，必须从 all
            // 里取 iframe；document.querySelectorAll('iframe') 只能看到 Hotjar 等顶层 frame。
            const aiFrames = all.filter((frame) => {
                if (String(frame.tagName || '').toUpperCase() !== 'IFRAME') return false;
                const text = [
                    frame.getAttribute('src') || '',
                    frame.getAttribute('title') || '',
                    frame.getAttribute('name') || '',
                    frame.getAttribute('id') || '',
                    String(frame.className || '')
                ].join(' ').toLowerCase();
                return frameMarkers.some((marker) => text.includes(String(marker).toLowerCase()));
            });
            const visibleAiFrames = aiFrames.filter((frame) => visible(frame));

            return {
                inline_shell: !!root || !!opener || all.some((el) =>
                    el.matches && el.matches('button.action-button[aria-label*="助手"]')
                ),
                inline_open: inlineOpen,
                inline_has_input: !!inlineInput,
                inline_root_class: rootClass,
                inline_aria_expanded: ariaExpanded,
                ai_frame_count: aiFrames.length,
                visible_ai_frame_count: visibleAiFrames.length,
                shadow_ai_frame_count: aiFrames.filter((frame) => {
                    const frameRoot = frame.getRootNode ? frame.getRootNode() : null;
                    return !!(frameRoot && frameRoot.host);
                }).length,
                // 保留原字段供分类与旧调用方使用；现在包含 Shadow DOM 中可见的 iframe。
                legacy_frame_count: visibleAiFrames.length
            };
            """,
            list(AI_FRAME_MARKERS),
        ) or {}
    except Exception:
        return {}


def detect_ai_chat_variant(driver):
    """返回 inline_dom、legacy_iframe 或空字符串。"""
    return classify_ai_chat_variant(get_ai_chat_dom_state(driver))


def find_inline_chat_input(driver, timeout=AI_CHAT_INPUT_TIMEOUT_SECONDS):
    """只在新版 #sa-assistant-chat 内查找输入框，避免误命中帮助页顶部搜索框。"""
    try:
        driver.switch_to.default_content()
    except Exception:
        pass

    end_time = time.time() + max(0, timeout)
    while time.time() < end_time:
        try:
            element = driver.execute_script(
                """
                function deepElements(root = document) {
                    const out = [];
                    const walk = (node) => {
                        const elements = node.querySelectorAll ? Array.from(node.querySelectorAll('*')) : [];
                        for (const el of elements) {
                            out.push(el);
                            if (el.shadowRoot) walk(el.shadowRoot);
                        }
                    };
                    walk(root);
                    return out;
                }

                function visible(el) {
                    const rect = el.getBoundingClientRect();
                    const style = window.getComputedStyle(el);
                    return !!(rect.width || rect.height || el.getClientRects().length)
                        && style.visibility !== 'hidden'
                        && style.display !== 'none'
                        && !el.disabled;
                }

                const all = deepElements(document);
                const root = all.find((el) => el.id === 'sa-assistant-chat');
                if (!root) return null;
                const candidates = deepElements(root).filter((el) => {
                    if (!visible(el)) return false;
                    const tag = String(el.tagName || '').toUpperCase();
                    const role = (el.getAttribute('role') || '').toLowerCase();
                    const editable = el.isContentEditable || el.getAttribute('contenteditable') === 'true';
                    return tag === 'TEXTAREA' || tag === 'INPUT' || role === 'textbox' || editable;
                }).map((el) => {
                    const rect = el.getBoundingClientRect();
                    const placeholder = (el.getAttribute('placeholder') || '').toLowerCase();
                    const aria = (el.getAttribute('aria-label') || '').toLowerCase();
                    const id = (el.id || '').toLowerCase();
                    let score = rect.bottom + rect.right;
                    if (id === 'chat-input' || id.includes('chat-input')) score += 12000;
                    if (/message|mensaje|mensagem|消息|输入|提问/.test(aria)) score += 8000;
                    if (/ask|escribe|digite|消息|输入|提问/.test(placeholder)) score += 6000;
                    if (String(el.tagName || '').toUpperCase() === 'TEXTAREA') score += 3000;
                    return {el, score};
                });
                candidates.sort((a, b) => b.score - a.score);
                return candidates.length ? candidates[0].el : null;
                """
            )
        except Exception:
            element = None

        if element:
            try:
                driver.execute_script("arguments[0].focus();", element)
            except Exception:
                pass
            return element
        time.sleep(0.25)
    return None


def click_inline_ai_assistant_entry(
    driver,
    name="",
    site="",
    timeout=AI_CHAT_ENTRY_TIMEOUT_SECONDS,
):
    """点击新版助手入口，并用内嵌输入框或 Shadow DOM iframe 验证结果。"""
    try:
        driver.switch_to.default_content()
    except Exception:
        pass

    if switch_to_ai_chat_frame(driver, require_input=True):
        setattr(driver, "_mercado_ai_chat_mode", AI_CHAT_MODE_IFRAME)
        return AI_CHAT_MODE_IFRAME
    if find_inline_chat_input(driver, timeout=0.2):
        setattr(driver, "_mercado_ai_chat_mode", AI_CHAT_MODE_INLINE)
        return AI_CHAT_MODE_INLINE

    try:
        clicked = driver.execute_script(
            """
            function deepElements(root = document) {
                const out = [];
                const walk = (node) => {
                    const elements = node.querySelectorAll ? Array.from(node.querySelectorAll('*')) : [];
                    for (const el of elements) {
                        out.push(el);
                        if (el.shadowRoot) walk(el.shadowRoot);
                    }
                };
                walk(root);
                return out;
            }

            function visible(el) {
                if (!el) return false;
                const rect = el.getBoundingClientRect();
                const style = window.getComputedStyle(el);
                return !!(rect.width || rect.height || el.getClientRects().length)
                    && style.visibility !== 'hidden'
                    && style.display !== 'none'
                    && !el.disabled;
            }

            const all = deepElements(document);
            const selectors = [
                'button.action-button[aria-label*="向助手提问"]',
                'button.action-button[data-component="WIDGET"]',
                '#sa-icon-button-wrapper[aria-expanded="false"]',
                '[aria-controls="sa-assistant-chat"][aria-expanded="false"]',
                'button[aria-label*="Ask the assistant"]',
                'button[aria-label*="Preguntar al asistente"]',
                'button[aria-label*="Perguntar ao assistente"]'
            ];
            let target = null;
            for (const selector of selectors) {
                target = all.find((el) => el.matches && el.matches(selector) && visible(el));
                if (target) break;
            }
            if (!target) return null;
            target.scrollIntoView({block: 'center', inline: 'center'});
            target.click();
            return {
                id: target.id || '',
                aria: target.getAttribute('aria-label') || '',
                component: target.getAttribute('data-component') || ''
            };
            """
        )
    except Exception:
        clicked = None

    if not clicked:
        return False

    print(f"{get_now_time()} {name} {site} 点击新版内嵌 AI 助手入口：{clicked}<br>")
    end_time = time.time() + max(0, timeout)
    while time.time() < end_time:
        if switch_to_ai_chat_frame(driver, require_input=True):
            setattr(driver, "_mercado_ai_chat_mode", AI_CHAT_MODE_IFRAME)
            return AI_CHAT_MODE_IFRAME
        if find_inline_chat_input(driver, timeout=0.5):
            setattr(driver, "_mercado_ai_chat_mode", AI_CHAT_MODE_INLINE)
            return AI_CHAT_MODE_INLINE
        state = get_ai_chat_dom_state(driver)
        if state.get("inline_open"):
            # 容器已经展开时继续等待异步加载输入框，不重复点击造成开关反转。
            time.sleep(0.5)
            continue
        time.sleep(0.5)
    return False


def activate_ai_chat_context(driver, require_input=False):
    """恢复已识别的聊天上下文，并在必要时自动重新探测模式。"""
    preferred = getattr(driver, "_mercado_ai_chat_mode", "")
    if preferred == AI_CHAT_MODE_INLINE:
        state = get_ai_chat_dom_state(driver)
        if state.get("inline_shell") and (not require_input or find_inline_chat_input(driver, timeout=1)):
            driver.switch_to.default_content()
            return AI_CHAT_MODE_INLINE
    elif preferred == AI_CHAT_MODE_IFRAME:
        if switch_to_ai_chat_frame(driver, require_input=require_input):
            return AI_CHAT_MODE_IFRAME

    variant = detect_ai_chat_variant(driver)
    if variant == AI_CHAT_MODE_INLINE:
        state = get_ai_chat_dom_state(driver)
        if state.get("inline_shell") and (not require_input or find_inline_chat_input(driver, timeout=1)):
            driver.switch_to.default_content()
            setattr(driver, "_mercado_ai_chat_mode", AI_CHAT_MODE_INLINE)
            return AI_CHAT_MODE_INLINE

    if switch_to_ai_chat_frame(driver, require_input=require_input):
        setattr(driver, "_mercado_ai_chat_mode", AI_CHAT_MODE_IFRAME)
        return AI_CHAT_MODE_IFRAME

    driver.switch_to.default_content()
    return ""


def reset_expired_ai_iframe(driver, name="", site=""):
    """父页面 iframe 地址已经过期时，直接重置为新的 AI 会话地址。"""
    try:
        driver.switch_to.default_content()
        reset = driver.execute_script(
            """
            function deepElements(root = document) {
                const out = [];
                const walk = (node) => {
                    const elements = node.querySelectorAll ? Array.from(node.querySelectorAll('*')) : [];
                    for (const el of elements) {
                        out.push(el);
                        if (el.shadowRoot) walk(el.shadowRoot);
                    }
                };
                walk(root);
                return out;
            }

            function cleanNewChatUrl(rawUrl) {
                const origin = window.location.origin && window.location.origin !== 'null'
                    ? window.location.origin
                    : 'https://global-selling.mercadolibre.com';
                const url = new URL(rawUrl || '/maxwell/new-chat', origin);
                url.pathname = '/maxwell/new-chat';
                url.searchParams.set('flavor', 'seller-assistant');
                url.searchParams.set('hideHeader', 'true');
                url.searchParams.set('origin', 'sa');
                url.searchParams.set('andes_ui', 'legacy');
                url.searchParams.set('customLocale', 'en-US');
                url.searchParams.delete('lifecycle');
                url.searchParams.delete('conversation_id');
                url.hash = '';
                return url.toString();
            }

            const frames = deepElements(document).filter((el) =>
                String(el.tagName || '').toUpperCase() === 'IFRAME'
            );
            const target = frames.find((frame) => {
                const src = frame.getAttribute('src') || '';
                return /maxwell\\/new-chat|meli-ai-chat/i.test(src) &&
                    /inactivity_expiration|conversation_id=/i.test(src);
            });
            if (!target) return false;
            const nextUrl = cleanNewChatUrl(target.getAttribute('src') || '');
            target.setAttribute('src', nextUrl);
            return nextUrl;
            """
        )
        if reset:
            print(f"{get_now_time()} {name} {site} AI会话 iframe 已过期，已重置为新会话：{reset}<br>")
            time.sleep(4)
            return True
    except Exception as e:
        print(f"{get_now_time()} {name} {site} 重置过期 AI iframe 失败：{e}<br>")
    finally:
        try:
            driver.switch_to.default_content()
        except Exception:
            pass
    return False


def switch_to_ai_chat_frame(driver, require_input=False, max_depth=2):
    """递归切换到 AI 客服 iframe。

    require_input=True 时，会进一步确认 iframe 内存在聊天输入框，避免误进帮助页顶部搜索框。
    """
    reset_expired_ai_iframe(driver)
    driver.switch_to.default_content()

    def search_frames(depth):
        """按页面位置和 AI 特征给 iframe 打分，优先尝试右下方的悬浮窗。"""
        frames = find_frames_including_shadow_dom(driver)
        frame_infos = []
        for frame in frames:
            try:
                info = get_frame_info(driver, frame)
                if not info.get("visible"):
                    continue
                score = info.get("bottom", 0) + info.get("right", 0)
                if is_ai_frame_info(info):
                    score += 20000
                if info.get("top", 0) > 100:
                    score += 2000
                frame_infos.append((score, frame, info))
            except Exception:
                continue

        for _, frame, info in sorted(frame_infos, key=lambda item: item[0], reverse=True):
            driver.switch_to.parent_frame() if depth > 0 else driver.switch_to.default_content()
            try:
                driver.switch_to.frame(frame)
            except Exception:
                continue

            if is_ai_frame_info(info):
                if not require_input or find_chat_input(driver, timeout=2, allow_default_content=False):
                    return True

            if depth < max_depth and search_frames(depth + 1):
                return True

        if depth > 0:
            try:
                driver.switch_to.parent_frame()
            except Exception:
                driver.switch_to.default_content()
        return False

    try:
        if search_frames(0):
            return True
    except Exception:
        driver.switch_to.default_content()

    driver.switch_to.default_content()
    return False


def switch_to_ai_chat_frame_old(driver):
    """旧版 iframe 定位逻辑，仅按 src 判断，保留用于对照排查。"""
    driver.switch_to.default_content()
    frames = driver.find_elements(By.TAG_NAME, "iframe")
    for frame in frames:
        try:
            if not frame.is_displayed():
                continue
            src = (frame.get_attribute("src") or "").lower()
            if not any(marker in src for marker in AI_FRAME_URL_MARKERS):
                continue
            driver.switch_to.frame(frame)
            return True
        except Exception:
            driver.switch_to.default_content()
            continue

    driver.switch_to.default_content()
    return False


def dump_iframe_debug_info(driver):
    """打印当前页面所有 iframe 的调试信息，方便定位 AI 悬浮窗加载失败原因。"""
    driver.switch_to.default_content()
    try:
        frames = driver.execute_script(
            """
            function deepElements(root = document) {
                const out = [];
                const walk = (node) => {
                    const elements = node.querySelectorAll ? Array.from(node.querySelectorAll('*')) : [];
                    for (const el of elements) {
                        out.push(el);
                        if (el.shadowRoot) walk(el.shadowRoot);
                    }
                };
                walk(root);
                return out;
            }
            return deepElements(document)
                .filter((el) => String(el.tagName || '').toUpperCase() === 'IFRAME')
                .map((frame, index) => {
                const rect = frame.getBoundingClientRect();
                const root = frame.getRootNode ? frame.getRootNode() : null;
                const host = root && root.host ? root.host : null;
                return {
                    index,
                    title: frame.getAttribute('title') || '',
                    name: frame.getAttribute('name') || '',
                    src: frame.getAttribute('src') || '',
                    top: Math.round(rect.top),
                    bottom: Math.round(rect.bottom),
                    width: Math.round(rect.width),
                    height: Math.round(rect.height),
                    visible: !!(rect.width || rect.height || frame.getClientRects().length),
                    shadowHost: host ? (host.id || host.className || host.tagName || '') : ''
                };
            });
            """
        )
        print(f"{get_now_time()} 当前页面 iframe 信息：{frames}<br>")
    except Exception as e:
        print(f"{get_now_time()} 获取 iframe 调试信息失败：{e}<br>")


def save_ai_open_debug_artifacts(driver, name, site):
    """AI 悬浮窗打开失败时保存截图、HTML 和 Shadow DOM 结构化快照。"""
    driver.switch_to.default_content()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    prefix = Path(__file__).resolve().parent / f"ai_chat_open_failed_{name}_{site}_{timestamp}"
    try:
        driver.save_screenshot(str(prefix.with_suffix(".png")))
        print(f"{get_now_time()} 已保存AI悬浮窗失败截图：{prefix.with_suffix('.png')}<br>")
    except Exception as e:
        print(f"{get_now_time()} 保存AI悬浮窗失败截图失败：{e}<br>")
    try:
        prefix.with_suffix(".html").write_text(driver.page_source, encoding="utf-8")
        print(f"{get_now_time()} 已保存AI悬浮窗失败HTML：{prefix.with_suffix('.html')}<br>")
    except Exception as e:
        print(f"{get_now_time()} 保存AI悬浮窗失败HTML失败：{e}<br>")
    try:
        snapshot = driver.execute_script(
            """
            function deepElements(root = document) {
                const out = [];
                const walk = (node) => {
                    const elements = node.querySelectorAll ? Array.from(node.querySelectorAll('*')) : [];
                    for (const el of elements) {
                        out.push(el);
                        if (el.shadowRoot) walk(el.shadowRoot);
                    }
                };
                walk(root);
                return out;
            }
            function visible(el) {
                const rect = el.getBoundingClientRect();
                const style = window.getComputedStyle(el);
                return !!(rect.width || rect.height || el.getClientRects().length)
                    && style.visibility !== 'hidden'
                    && style.display !== 'none';
            }
            const all = deepElements(document);
            return {
                title: document.title || '',
                readyState: document.readyState || '',
                shadowHosts: all
                    .filter((el) => !!el.shadowRoot)
                    .map((el) => ({
                        tag: el.tagName || '',
                        id: el.id || '',
                        className: String(el.className || '')
                    })),
                aiShells: all
                    .filter((el) =>
                        el.id === 'sa-assistant-chat'
                        || el.id === 'sa-icon-button-wrapper'
                        || el.getAttribute('aria-controls') === 'sa-assistant-chat'
                    )
                    .map((el) => ({
                        tag: el.tagName || '',
                        id: el.id || '',
                        className: String(el.className || ''),
                        ariaExpanded: el.getAttribute('aria-expanded') || '',
                        visible: visible(el)
                    })),
                frames: all
                    .filter((el) => String(el.tagName || '').toUpperCase() === 'IFRAME')
                    .map((frame) => {
                        const rect = frame.getBoundingClientRect();
                        const root = frame.getRootNode ? frame.getRootNode() : null;
                        const host = root && root.host ? root.host : null;
                        return {
                            src: frame.getAttribute('src') || '',
                            title: frame.getAttribute('title') || '',
                            id: frame.id || '',
                            className: String(frame.className || ''),
                            width: Math.round(rect.width),
                            height: Math.round(rect.height),
                            visible: visible(frame),
                            shadowHost: host
                                ? (host.id || String(host.className || '') || host.tagName || '')
                                : ''
                        };
                    })
            };
            """
        ) or {}
        debug_info = {
            "captured_at": datetime.now().isoformat(timespec="seconds"),
            "current_url": driver.current_url,
            "chat_state": get_ai_chat_dom_state(driver),
            "dom_snapshot": snapshot,
        }
        try:
            debug_info["browser_logs"] = driver.get_log("browser")[-100:]
        except Exception:
            debug_info["browser_logs"] = []
        debug_path = prefix.with_suffix(".debug.json")
        debug_path.write_text(
            json.dumps(debug_info, ensure_ascii=False, indent=2, default=str),
            encoding="utf-8",
        )
        print(f"{get_now_time()} 已保存AI悬浮窗失败结构化快照：{debug_path}<br>")
    except Exception as e:
        print(f"{get_now_time()} 保存AI悬浮窗失败结构化快照失败：{e}<br>")


def dump_ai_entry_debug_info(driver):
    """打印页面上疑似 AI/Help/Contact 入口的元素信息。"""
    driver.switch_to.default_content()
    try:
        entries = driver.execute_script(
            """
            function deepElements(root = document) {
                const out = [];
                const walk = (node) => {
                    const elements = node.querySelectorAll ? Array.from(node.querySelectorAll('*')) : [];
                    for (const el of elements) {
                        out.push(el);
                        if (el.shadowRoot) walk(el.shadowRoot);
                    }
                };
                walk(root);
                return out;
            }
            return deepElements(document)
                .filter((node) => ['BUTTON', 'A', 'DIV', 'SPAN'].includes(node.tagName))
                .map((node, index) => {
                    const rect = node.getBoundingClientRect();
                    const label = [
                        node.innerText || '',
                        node.getAttribute('aria-label') || '',
                        node.getAttribute('title') || '',
                        node.getAttribute('data-testid') || '',
                        String(node.className || '')
                    ].join(' ').trim();
                    return {
                        index,
                        tag: node.tagName,
                        label: label.slice(0, 180),
                        top: Math.round(rect.top),
                        left: Math.round(rect.left),
                        width: Math.round(rect.width),
                        height: Math.round(rect.height),
                        visible: rect.width > 0 && rect.height > 0
                    };
                })
                .filter((item) =>
                    item.visible &&
                    /assistant|chat|help|maxwell|contact|助手|助理|提问|咨询|输入/i.test(item.label)
                )
                .slice(0, 30);
            """
        )
        print(f"{get_now_time()} 当前页面AI入口候选：{entries}<br>")
    except Exception as e:
        print(f"{get_now_time()} 获取AI入口候选失败：{e}<br>")


def dump_ai_chat_mode_debug_info(driver):
    """同时输出新版内嵌助手和旧版 iframe 的探测状态。"""
    try:
        state = get_ai_chat_dom_state(driver)
        print(
            f"{get_now_time()} AI聊天模式探测："
            f"variant={classify_ai_chat_variant(state) or 'unknown'}，state={state}<br>"
        )
    except Exception as e:
        print(f"{get_now_time()} 获取AI聊天模式探测信息失败：{e}<br>")


def find_chat_input(
    driver,
    timeout=AI_CHAT_INPUT_TIMEOUT_SECONDS,
    allow_default_content=False,
):
    """在当前 iframe 内查找 AI 客服真正的 textarea 输入框。

    这里优先匹配 id=chat-input、aria-label、placeholder，并给页面下半部分元素更高分，
    目的是避开页面顶部的帮助中心搜索框。
    """
    end_time = time.time() + timeout
    while time.time() < end_time:
        element = driver.execute_script(
            """
            function deepElements(root = document) {
                const out = [];
                const walk = (node) => {
                    const elements = node.querySelectorAll ? Array.from(node.querySelectorAll('*')) : [];
                    for (const el of elements) {
                        out.push(el);
                        if (el.shadowRoot) walk(el.shadowRoot);
                    }
                };
                walk(root);
                return out;
            }

            function visible(el) {
                const rect = el.getBoundingClientRect();
                const style = window.getComputedStyle(el);
                return !!(rect.width || rect.height || el.getClientRects().length)
                    && style.visibility !== 'hidden'
                    && style.display !== 'none'
                    && !el.disabled;
            }

            const viewportHeight = window.innerHeight || document.documentElement.clientHeight;
            const viewportWidth = window.innerWidth || document.documentElement.clientWidth;
            const candidates = deepElements().filter((el) => {
                const placeholder = el.getAttribute('placeholder') || '';
                const aria = el.getAttribute('aria-label') || '';
                const role = el.getAttribute('role') || '';
                const tag = el.tagName || '';
                const editable = el.isContentEditable || el.getAttribute('contenteditable') === 'true';
                const textboxLike = tag === 'TEXTAREA' || tag === 'INPUT' || role === 'textbox' || editable;
                return visible(el) && textboxLike && (
                    el.id === 'chat-input' ||
                    aria.includes('Chat message input') ||
                    placeholder === 'Ask me' ||
                    placeholder.includes('Ask the assistant') ||
                    placeholder.includes('Escribe') ||
                    placeholder.includes('Digite') ||
                    placeholder.includes('输入') ||
                    (editable && (aria.toLowerCase().includes('message') || role === 'textbox'))
                );
            }).map((el) => {
                const rect = el.getBoundingClientRect();
                const placeholder = el.getAttribute('placeholder') || '';
                const aria = el.getAttribute('aria-label') || '';
                let score = rect.bottom + rect.right;
                if (el.id === 'chat-input') score += 10000;
                if (aria.includes('Chat message input')) score += 5000;
                if (el.isContentEditable || el.getAttribute('contenteditable') === 'true') score += 3500;
                if (placeholder === 'Ask me') score += 3000;
                if (rect.top > viewportHeight * 0.30) score += 8000;
                if (rect.left > viewportWidth * 0.35) score += 1000;
                return { el, rect, score };
            });

            const bottomCandidates = candidates.filter((item) =>
                item.rect.top > viewportHeight * 0.25 || item.rect.bottom > viewportHeight * 0.55
            );
            const pool = bottomCandidates.length ? bottomCandidates : candidates;
            pool.sort((a, b) => b.score - a.score);
            return pool.length ? pool[0].el : null;
            """
        )
        if element:
            try:
                driver.execute_script("arguments[0].focus();", element)
                element.click()
                return element
            except Exception:
                try:
                    driver.execute_script("arguments[0].focus();", element)
                    return element
                except Exception:
                    pass
        time.sleep(0.2)
    return None


def recover_expired_ai_conversation(
    driver,
    name="",
    site="",
    timeout=AI_CHAT_READY_TIMEOUT_SECONDS,
    force=False,
):
    """AI 会话过期时优先使用页面提供的 New conversation 链接进入新对话。"""
    try:
        result = driver.execute_script(
            """
            const force = arguments[0];
            function fallbackNewChatUrl(rawUrl) {
                const origin = window.location.origin && window.location.origin !== 'null'
                    ? window.location.origin
                    : 'https://global-selling.mercadolibre.com';
                const url = new URL(rawUrl || '/maxwell/new-chat', origin);
                url.pathname = '/maxwell/new-chat';
                url.searchParams.set('flavor', 'seller-assistant');
                url.searchParams.set('hideHeader', 'true');
                url.searchParams.set('origin', 'sa');
                url.searchParams.set('andes_ui', 'legacy');
                url.searchParams.set('customLocale', 'en-US');
                url.searchParams.delete('lifecycle');
                url.searchParams.delete('conversation_id');
                url.hash = '';
                return url.toString();
            }

            function deepElements(root = document) {
                const out = [];
                const walk = (node) => {
                    const elements = node.querySelectorAll ? Array.from(node.querySelectorAll('*')) : [];
                    for (const el of elements) {
                        out.push(el);
                        if (el.shadowRoot) walk(el.shadowRoot);
                    }
                };
                walk(root);
                return out;
            }

            function visible(el) {
                const rect = el.getBoundingClientRect();
                const style = window.getComputedStyle(el);
                return !!(rect.width || rect.height || el.getClientRects().length)
                    && style.visibility !== 'hidden'
                    && style.display !== 'none'
                    && !el.disabled;
            }

            const nodes = deepElements();
            const pageText = nodes.map(el => el.innerText || el.textContent || '').join('\\n');
            const currentUrl = window.location.href || '';
            const hasInput = nodes.some((el) => {
                const tag = el.tagName || '';
                const role = el.getAttribute('role') || '';
                const placeholder = el.getAttribute('placeholder') || '';
                const aria = el.getAttribute('aria-label') || '';
                return visible(el) && (
                    tag === 'TEXTAREA' ||
                    tag === 'INPUT' ||
                    role === 'textbox' ||
                    el.isContentEditable ||
                    el.getAttribute('contenteditable') === 'true'
                ) && (
                    el.id === 'chat-input' ||
                    aria.includes('Chat message input') ||
                    placeholder === 'Ask me' ||
                    placeholder.includes('Ask the assistant') ||
                    placeholder.includes('Escribe') ||
                    placeholder.includes('Digite') ||
                    placeholder.includes('输入')
                );
            });
            const expired = force ||
                /this conversation has ended|conversation has ended|inactivity_expiration/i.test(pageText + '\\n' + currentUrl) ||
                (!hasInput && /new conversation|iniciar otra consulta|发起新咨询|新的对话/i.test(pageText));
            if (!expired) return false;

            // 先精确查找页面真正显示的 New conversation 链接。SPA 版本的入口
            // 可能使用 # 或 click handler，不能要求 href 必须包含 /maxwell/new-chat。
            const links = nodes.filter((el) => {
                if (String(el.tagName || '').toUpperCase() !== 'A') return false;
                const href = el.getAttribute('href') || '';
                const label = [
                    el.innerText || '',
                    el.getAttribute('aria-label') || '',
                    el.getAttribute('title') || '',
                ].join(' ');
                return /new conversation|iniciar otra consulta|发起新咨询|新的对话/i.test(label);
            });
            links.sort((a, b) => Number(visible(b)) - Number(visible(a)));
            const link = links[0];
            if (link) {
                const href = link.getAttribute('href') || '';
                if (href && href !== '#' && !/^javascript:/i.test(href)) {
                    // 服务端生成的 conversation_id 必须原样保留。
                    const nextUrl = new URL(href, window.location.href).toString();
                    window.location.assign(nextUrl);
                    return { action: 'navigate_new_conversation_href', url: nextUrl };
                }
                link.scrollIntoView({block: 'center', inline: 'center'});
                link.click();
                return { action: 'click_new_conversation_link', href: href };
            }

            // 某些版本使用按钮而不是链接；只匹配按钮本身，避免误点包含文案的祖先节点。
            const button = nodes.find((el) => {
                const tag = String(el.tagName || '').toUpperCase();
                const role = el.getAttribute('role') || '';
                if (!visible(el) || (tag !== 'BUTTON' && role !== 'button')) return false;
                const label = [
                    el.innerText || '',
                    el.getAttribute('aria-label') || '',
                    el.getAttribute('title') || '',
                ].join(' ');
                return /new conversation|iniciar otra consulta|发起新咨询|新的对话/i.test(label);
            });
            if (button) {
                button.scrollIntoView({block: 'center', inline: 'center'});
                button.click();
                return { action: 'click_new_conversation_button' };
            }

            // 页面确实没有提供入口时，才使用不带旧 conversation_id 的标准地址兜底。
            const fallbackUrl = new URL(fallbackNewChatUrl(currentUrl || '/maxwell/new-chat'));
            // 避免浏览器把与当前地址相同的兜底导航优化掉。
            fallbackUrl.searchParams.set('_new_conversation_ts', String(Date.now()));
            const nextUrl = fallbackUrl.toString();
            window.location.assign(nextUrl);
            return { action: 'navigate_fallback', url: nextUrl };
            """,
            bool(force),
        )
        if result in ("expired_no_button", "expired_no_clickable_button"):
            print(f"{get_now_time()} {name} {site} AI会话已结束，但没有找到可用的 New conversation 入口<br>")
            return False
        if result:
            print(f"{get_now_time()} {name} {site} AI会话已结束，已进入 New conversation：{result}<br>")
            end_time = time.time() + timeout
            while time.time() < end_time:
                time.sleep(1)
                try:
                    if find_chat_input(driver, timeout=1, allow_default_content=False):
                        print(f"{get_now_time()} {name} {site} 已进入新的 AI 会话<br>")
                        return True
                except Exception:
                    # iframe 跳转期间旧 document 可能短暂失效，继续等待新页面加载。
                    continue
            print(f"{get_now_time()} {name} {site} New conversation 已跳转，但未出现输入框<br>")
            return False
    except Exception as e:
        print(f"{get_now_time()} {name} {site} 检查 AI 会话过期失败：{e}<br>")
    return False


def click_send_button(driver, mode=AI_CHAT_MODE_IFRAME):
    """查找并点击 AI 客服输入框旁的发送按钮，失败时交给回车发送兜底。"""
    button = driver.execute_script(
        """
        function deepElements(root = document) {
            const out = [];
            const walk = (node) => {
                const elements = node.querySelectorAll ? Array.from(node.querySelectorAll('*')) : [];
                for (const el of elements) {
                    out.push(el);
                    if (el.shadowRoot) walk(el.shadowRoot);
                }
            };
            walk(root);
            return out;
        }

        function visible(el) {
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            return !!(rect.width || rect.height || el.getClientRects().length)
                && style.visibility !== 'hidden'
                && style.display !== 'none'
                && !el.disabled;
        }

        const inlineMode = arguments[0] === 'inline_dom';
        const all = deepElements(document);
        const inlineRoot = inlineMode ? all.find((el) => el.id === 'sa-assistant-chat') : null;
        const searchRoot = inlineMode ? inlineRoot : document;
        if (!searchRoot) return null;
        const viewportHeight = window.innerHeight || document.documentElement.clientHeight;
        const candidates = deepElements(searchRoot).filter((el) => {
            const aria = el.getAttribute('aria-label') || '';
            const title = el.getAttribute('title') || '';
            const cls = String(el.className || '');
            const buttonLike = el.tagName === 'BUTTON' || el.getAttribute('role') === 'button';
            return visible(el) && buttonLike && (
                /send|enviar|发送|提交/i.test(aria) ||
                /send|enviar|发送|提交/i.test(title) ||
                cls.includes('new-chat-input__right-button') ||
                el.type === 'submit'
            );
        }).map((el) => {
            const rect = el.getBoundingClientRect();
            let score = rect.bottom + rect.right;
            const aria = el.getAttribute('aria-label') || '';
            const cls = String(el.className || '');
            if (aria.includes('Send message')) score += 10000;
            if (cls.includes('new-chat-input__right-button')) score += 8000;
            if (rect.top > viewportHeight * 0.30) score += 5000;
            return { el, score };
        });
        candidates.sort((a, b) => b.score - a.score);
        return candidates.length ? candidates[0].el : null;
        """,
        mode,
    )
    if not button:
        return False
    # The browser may have accepted the click even when its response is lost.
    driver.execute_script("arguments[0].click();", button)
    return True


def send_ai_chat_message(driver, message):
    """向 AI 客服窗口发送一条消息。

    Mercado 的 textarea 有时不接受普通 click/send_keys，所以优先用 JS 原生 setter 写值并触发 input/change 事件，
    如果页面仍未接收到内容，再使用 ActionChains 兜底。
    """
    mode = activate_ai_chat_context(driver, require_input=False)
    if not mode:
        raise RuntimeError("没有找到 AI 客服聊天窗口")

    if mode == AI_CHAT_MODE_INLINE:
        input_box = find_inline_chat_input(
            driver,
            timeout=AI_CHAT_INPUT_TIMEOUT_SECONDS,
        )
    else:
        recover_expired_ai_conversation(driver)
        input_box = find_chat_input(
            driver,
            timeout=AI_CHAT_INPUT_TIMEOUT_SECONDS,
            allow_default_content=False,
        )
        if input_box is None and recover_expired_ai_conversation(driver):
            input_box = find_chat_input(
                driver,
                timeout=AI_CHAT_INPUT_TIMEOUT_SECONDS,
                allow_default_content=False,
            )
        if input_box is None and recover_expired_ai_conversation(driver, force=True):
            switch_to_ai_chat_frame(driver, require_input=False)
            input_box = find_chat_input(
                driver,
                timeout=AI_CHAT_INPUT_TIMEOUT_SECONDS,
                allow_default_content=False,
            )
    if input_box is None:
        raise RuntimeError("没有找到 AI 客服输入框")

    if (input_box.tag_name or "").lower() == "textarea":
        driver.execute_script(
            """
            const input = arguments[0];
            const value = arguments[1];
            input.scrollIntoView({block: 'center', inline: 'center'});
            input.focus();
            const setter = Object.getOwnPropertyDescriptor(HTMLTextAreaElement.prototype, 'value')?.set
                || Object.getOwnPropertyDescriptor(input.constructor.prototype, 'value')?.set;
            if (setter) {
                setter.call(input, value);
            } else {
                input.value = value;
            }
            input.dispatchEvent(new InputEvent('input', { bubbles: true, inputType: 'insertText', data: value }));
            input.dispatchEvent(new Event('change', { bubbles: true }));
            input.dispatchEvent(new KeyboardEvent('keydown', { bubbles: true, key: 'a' }));
            input.dispatchEvent(new KeyboardEvent('keyup', { bubbles: true, key: 'a' }));
            """,
            input_box,
            message,
        )
        current_value = driver.execute_script("return arguments[0].value || '';", input_box)
        if current_value != message:
            input_box.clear()
            input_box.send_keys(message)
    else:
        try:
            input_box.click()
            if (input_box.get_attribute("contenteditable") or "").lower() == "true":
                driver.execute_script(
                    """
                    const input = arguments[0];
                    const value = arguments[1];
                    input.scrollIntoView({block: 'center', inline: 'center'});
                    input.focus();
                    input.innerText = value;
                    input.textContent = value;
                    input.dispatchEvent(new InputEvent('input', { bubbles: true, inputType: 'insertText', data: value }));
                    input.dispatchEvent(new Event('change', { bubbles: true }));
                    """,
                    input_box,
                    message,
                )
            else:
                input_box.clear()
                input_box.send_keys(message)
        except Exception:
            driver.execute_script(
                """
                const input = arguments[0];
                const value = arguments[1];
                input.scrollIntoView({block: 'center', inline: 'center'});
                input.focus();
                if (input.isContentEditable || input.getAttribute('contenteditable') === 'true') {
                    input.innerText = value;
                    input.textContent = value;
                } else {
                    input.value = value;
                }
                input.dispatchEvent(new InputEvent('input', { bubbles: true, inputType: 'insertText', data: value }));
                input.dispatchEvent(new Event('change', { bubbles: true }));
                """,
                input_box,
                message,
            )
    before_send = read_snapshot(driver)
    _check_appeal_control(driver)
    try:
        if not click_send_button(driver, mode=mode):
            input_box.send_keys(Keys.ENTER)
        deadline = time.monotonic() + AI_SEND_CONFIRM_TIMEOUT_SECONDS
        while time.monotonic() < deadline:
            _check_appeal_control(driver)
            try:
                after_send = read_snapshot(driver)
                echoed = [m for m in new_messages(before_send, after_send, "user")
                          if normalized_text(m["text"]) == normalized_text(message)]
                current_value = driver.execute_script(
                    "return arguments[0].value || arguments[0].innerText || '';", input_box,
                )
                if echoed and not normalized_text(current_value):
                    return {"acknowledged": True, "reply_baseline": ChatMessages(before_send),
                            "message_id": echoed[-1]["id"],
                            "conversation_id": after_send.get("conversation_id", ""),
                            "chat_snapshot": after_send}
            except AppealExecutionError:
                raise
            except Exception:
                pass
            _appeal_pause(driver, 0.5)
        raise AppealExecutionError("已执行发送，但未确认消息气泡和输入框状态", "sent_unknown", sent=True)
    except AppealExecutionError as exc:
        # Stop/deadline after the click must still preserve possible submission.
        exc.sent = True
        raise
    except Exception as exc:
        raise AppealExecutionError(f"发送结果不确定：{exc}", "sent_unknown", sent=True) from exc


def safe_get_agent_messages(driver):
    """读取失败不能伪装成空列表，否则旧回复可能成为下一组的新回复。"""
    if not activate_ai_chat_context(driver, require_input=False):
        raise RuntimeError("没有找到 AI 客服聊天窗口")
    return ChatMessages(read_snapshot(driver))


def wait_for_ai_agent_reply(
    driver, previous_messages, timeout=AI_AGENT_REPLY_TIMEOUT_SECONDS,
    poll_interval=AI_AGENT_REPLY_POLL_SECONDS,
):
    before = getattr(previous_messages, "snapshot", None)
    previous_messages = previous_messages or []
    deadline = time.monotonic() + timeout
    latest = previous_messages
    candidate, stable_since = "", None
    while time.monotonic() < deadline:
        _check_appeal_control(driver)
        try:
            latest = safe_get_agent_messages(driver)
            after = getattr(latest, "snapshot", None)
            if before is not None and after is not None:
                changes = new_messages(before, after, "assistant")
                response = changes[-1]["text"] if changes else ""
                busy = after.get("busy", False)
            else:
                response = latest[-1] if latest and (
                    len(latest) > len(previous_messages)
                    or latest[-1] != (previous_messages[-1] if previous_messages else "")
                ) else ""
                busy = False
            if response and not busy:
                if response != candidate:
                    candidate, stable_since = response, time.monotonic()
                elif stable_since is not None and time.monotonic() - stable_since >= 2:
                    return response, latest
            else:
                candidate, stable_since = "", None
        except AppealExecutionError:
            raise
        except Exception as exc:
            candidate, stable_since = "", None
            print(f"{get_now_time()} 读取客服回复暂时失败：{exc}<br>")
        _appeal_pause(driver, min(poll_interval, max(0, deadline - time.monotonic())))
    return "", latest


def should_intervene_ai_response(response_text):
    """判断 AI 回复是否需要脚本继续补充说明或坚持申诉。"""
    if not response_text:
        return False
    if is_site_option_question(response_text):
        return True
    keywords = [
        "无法", "不能", "不可以", "拒绝", "不支持", "没有权限", "无法删除",
        "无法撤销", "无法回滚", "rollback", "not possible", "can't", "cannot",
        "processing", "正在处理", "稍后", "等待", "try again",
    ]
    lower_text = response_text.lower()
    return any(keyword.lower() in lower_text for keyword in keywords)


def build_site_option_reply(site):
    """根据当前站点生成 AI 要求选择站点时的标准回复。"""
    site_code = normalize_site_code(site)
    option_map = {
        "MX": "Mexico (Direct to consumer)",
        "BR": "Brazil",
        "CL": "Chile",
        "CO": "Colombia",
        "AR": "Argentina",
        "UY": "Uruguay",
    }
    if site_code not in option_map:
        raise AppealExecutionError(f"无法确认客服所需站点：{site}", "needs_human")
    return option_map[site_code]


def contains_site_option_menu(response_text):
    """识别 AI 发出的完整站点选项列表。

    AI 有时只返回“Mexico ... Uruguay”这类纯选项列表，不一定带“可选”或问句；
    只要看到多个站点选项同时出现，就按站点确认问题处理。
    """
    lower_text = re.sub(r"\s+", " ", response_text or "").lower()
    lower_text = lower_text.replace("（", "(").replace("）", ")")
    compact_text = re.sub(r"[\s。．.、,，:：;；]+", "", lower_text)
    compact_menu = "".join(SITE_OPTION_MENU_OPTIONS).lower()
    compact_menu = re.sub(r"[\s。．.、,，:：;；]+", "", compact_menu)

    # 硬规则：只要文本包含完整“Mexico ... Uruguay”菜单，就必须自动回复站点。
    if compact_menu in compact_text:
        return True

    compact_options = [
        re.sub(r"[\s。．.、,，:：;；]+", "", option.lower())
        for option in SITE_OPTION_MENU_OPTIONS
    ]
    matched_count = sum(1 for option in compact_options if option in compact_text)
    if matched_count >= 5:
        return True

    return ("可选" in lower_text or "option" in lower_text or "站点" in lower_text) and matched_count >= 3


def find_site_option_message(messages):
    """从最近聊天消息中找出包含站点选项菜单的 AI 回复。"""
    for message in reversed(messages or []):
        if is_site_option_question(message):
            return message
    joined_text = "\n".join(messages or [])
    if is_site_option_question(joined_text):
        return joined_text
    return ""


def is_site_option_question(response_text):
    """识别 AI 是否在询问“针对哪个站点/选项提出咨询”。"""
    if contains_site_option_menu(response_text):
        return True

    lower_text = (response_text or "").lower()
    option_markers = [
        "mexico (direct to consumer)",
        "mexico (fulfillment)",
        "brazil",
        "chile",
        "colombia",
        "argentina",
        "uruguay",
        "可选",
    ]
    question_markers = [
        "which country",
        "which site",
        "which option",
        "select a country",
        "select your country",
        "choose a country",
        "confirm your country",
        "confirm the country",
        "select a site",
        "choose a site",
        "qual país",
        "qual pais",
        "qué país",
        "que país",
        "qué sitio",
        "哪个国家",
        "哪个站点",
        "针对哪个",
        "哪个选项",
        "请选择站点",
        "请选择国家",
        "请确认站点",
        "请确认国家",
    ]
    return (
        any(marker in lower_text for marker in option_markers)
        and any(marker in lower_text for marker in question_markers)
    )


def build_infraction_followup_message(infraction_ids, site):
    """生成侵权申诉被 AI 拒绝或要求确认后继续坚持核查的话术。"""
    return (
        f"{infraction_ids} 请继续帮我重新核查。我的店铺对应的是 {build_site_option_reply(site)} 站点，"
        f"这些商品是通用品牌/通用款产品，并非侵权产品，也没有使用他人品牌商标，"
        f"这是系统误判。请直接帮我复查并删除侵权记录，谢谢。"
    )


def reply_site_option_menu_if_present(
    driver,
    name,
    site,
    timeout=AI_CHAT_READY_TIMEOUT_SECONDS,
):
    """只要当前 AI 窗口文本包含完整站点选项菜单，就按当前站点自动回复。"""
    end_time = time.time() + timeout
    while time.time() < end_time:
        messages = safe_get_agent_messages(driver)
        site_option_message = find_site_option_message(messages)
        if site_option_message:
            reply = build_site_option_reply(site)
            print(f"{get_now_time()} {name} {site} 检测到完整站点选项菜单，自动回复：{reply}<br>")
            send_ai_chat_message(driver, reply)
            append_chat_log(
                name,
                site,
                "auto_reply_site_option_menu",
                message=reply,
                response=site_option_message,
                chat=messages,
            )
            return True, site_option_message, messages
        time.sleep(1)
    return False, "", []


def _check_appeal_control(driver):
    stop_event = getattr(driver, "_bit_appeal_stop_event", None)
    if stop_event is not None and stop_event.is_set():
        raise AppealExecutionError("已收到停止请求", "stopped")
    deadline = getattr(driver, "_bit_appeal_deadline", None)
    if deadline is not None and time.monotonic() >= deadline:
        raise AppealExecutionError("本站点超过执行时间预算", "deadline_exceeded")


def _appeal_pause(driver, seconds):
    remaining = max(0, seconds)
    while remaining > 0:
        _check_appeal_control(driver)
        interval = min(0.5, remaining)
        time.sleep(interval)
        remaining -= interval


def _prepare_group_conversation(driver, name, site, group_index):
    _check_appeal_control(driver)
    if getattr(driver, "_bit_target_site", ""):
        driver.switch_to.default_content()
        if not verify_selected_site(driver, site):
            raise AppealExecutionError("发送前校验发现站点已经改变", "pre_send_failed")
    if getattr(driver, "_bit_ai_reset_before_group", False) or (
        group_index > 1 and (group_index - 1) % AI_GROUPS_PER_CONVERSATION == 0
    ):
        if not restart_ai_conversation(driver, name, site):
            raise AppealExecutionError("未能建立新的客服会话", "pre_send_failed")
        setattr(driver, "_bit_ai_reset_before_group", False)
        print(f"{get_now_time()} {name} {site} 已为第 {group_index} 组建立新会话<br>")


def restart_ai_conversation(driver, name, site):
    """Confirm a new session; never navigate the top-level inline page to human chat."""
    mode = activate_ai_chat_context(driver, require_input=False)
    if not mode:
        return False
    before = read_snapshot(driver)
    if mode == AI_CHAT_MODE_IFRAME:
        if not recover_expired_ai_conversation(driver, name, site, force=True):
            return False
    else:
        clicked = driver.execute_script(r"""
            function all(root) {
                const out = [];
                for (const el of root.querySelectorAll('*')) {
                    out.push(el);
                    if (el.shadowRoot) out.push(...all(el.shadowRoot));
                }
                return out;
            }
            const root = all(document).find(el => el.id === 'sa-assistant-chat');
            if (!root) return false;
            const entry = all(root).find(el => {
                const r = el.getBoundingClientRect();
                return r.width && r.height && el.matches('button, a, [role="button"]') &&
                    /new conversation|new chat|iniciar otra consulta|nova conversa|发起新咨询|新的对话/i.test(
                        el.innerText || el.getAttribute('aria-label') || '');
            });
            if (!entry) return false;
            entry.click();
            return true;
        """)
        if not clicked:
            return False
    deadline = time.monotonic() + AI_CHAT_READY_TIMEOUT_SECONDS
    while time.monotonic() < deadline:
        _check_appeal_control(driver)
        try:
            if activate_ai_chat_context(driver, require_input=True):
                after = read_snapshot(driver)
                if (after.get("epoch") != before.get("epoch")
                        or after.get("conversation_id") != before.get("conversation_id")
                        or (before["messages"] and not after["messages"])):
                    return True
        except Exception:
            pass
        _appeal_pause(driver, 0.5)
    return False


def send_infraction_message_with_retry(
    driver, huashu, infraction_ids, name, site, group_index, total_groups,
    appeal_kind="侵权",
):
    """话术回显即视为申诉成功；客服回复只作为附加记录。

    只重试发送前失败。发送成功后仍等待并记录回复，但回复超时、
    站点追问或后续读取失败不再把已发送的申诉改判为失败。
    """
    identifier_key, event_name = {
        "取消率": ("cancellation_ids", "cancellation"),
        "投诉": ("complaint_order_ids", "complaint"),
        "禁限售": ("prohibited_ids", "prohibited"),
        "延误": ("delay_ids", "delay"),
    }.get(appeal_kind, ("infraction_ids", "infraction"))
    base_extra = {"group_index": group_index, "total_groups": total_groups,
                  identifier_key: infraction_ids, "appeal_kind": appeal_kind}
    result = execution_result("pre_send_failed")
    latest_chat = []
    try:
        _prepare_group_conversation(driver, name, site, group_index)
        message, site_answers = huashu, 0
        while True:
            _check_appeal_control(driver)
            before = safe_get_agent_messages(driver)
            latest_chat = before
            for attempt in range(1, 3):
                try:
                    ack = send_ai_chat_message(driver, message)
                    break
                except AppealExecutionError:
                    raise
                except Exception as exc:
                    append_chat_log(name, site, f"send_{event_name}_error", message=message,
                                    extra={**base_extra, "attempt": attempt, "error": str(exc)})
                    if attempt == 2:
                        raise AppealExecutionError(str(exc), "pre_send_failed") from exc
                    _appeal_pause(driver, 5 * attempt + random.uniform(0, 2))
            # The initial appeal phrase is the success boundary. Site-option
            # follow-ups enrich the conversation but cannot revoke this result.
            if site_answers == 0:
                result.update(
                    status="sent",
                    execution_status="sent",
                    message=STATUS_LABELS["sent"],
                    sent=True,
                    acknowledged=True,
                    retryable=False,
                )
            if isinstance(ack, dict):
                before = ack.get("reply_baseline", before)
                latest_chat = ack.get("chat_snapshot") or latest_chat
                if site_answers == 0:
                    result.update(conversation_id=ack.get("conversation_id", ""),
                                  message_id=ack.get("message_id", ""))
            append_chat_log(
                name, site,
                ("send_delay_group" if appeal_kind == "延误" else f"send_{event_name}")
                if site_answers == 0 else "send_followup",
                message=message,
                chat=latest_chat,
                extra={
                    **base_extra,
                    "generated_by": "local_rule",
                    "acknowledged": True,
                    "conversation_id": ack.get("conversation_id", "") if isinstance(ack, dict) else "",
                },
            )
            response, latest = wait_for_ai_agent_reply(
                driver, before, timeout=AI_AGENT_REPLY_TIMEOUT_SECONDS,
                poll_interval=AI_AGENT_REPLY_POLL_SECONDS,
            )
            if getattr(latest, "snapshot", None) is not None or latest:
                latest_chat = latest
            append_chat_log(
                name, site,
                ("delay_agent_reply" if appeal_kind == "延误" else "agent_reply")
                if site_answers == 0 else "agent_reply_after_site_option",
                message=huashu, response=response, chat=latest_chat, extra=base_extra,
            )
            if not response:
                result.update(reply_status="reply_timeout", reply_timed_out=True)
                append_chat_log(
                    name,
                    site,
                    f"{event_name}_reply_timeout",
                    message=huashu,
                    chat=latest_chat,
                    extra={**base_extra, "timeout_seconds": AI_AGENT_REPLY_TIMEOUT_SECONDS},
                )
                break
            result.update(response=response, reply_received=True, reply_status="replied")
            if not is_site_option_question(response):
                break
            if site_answers >= 2:
                result.update(reply_status="needs_human", needs_human=True)
                break
            message = build_site_option_reply(site)
            site_answers += 1
    except AppealExecutionError as exc:
        if result.get("acknowledged"):
            result.update(
                status="sent",
                execution_status="sent",
                message=STATUS_LABELS["sent"],
                post_send_status=exc.status,
                post_send_error=str(exc),
                retryable=False,
            )
        else:
            result.update(status=exc.status, execution_status=exc.status,
                          message=STATUS_LABELS.get(exc.status, exc.status),
                          error=str(exc), sent=result["sent"] or exc.sent, retryable=exc.retryable)
            if exc.status in {"stopped", "deadline_exceeded"}:
                raise
    except Exception as exc:
        if result.get("acknowledged"):
            result.update(
                status="sent",
                execution_status="sent",
                message=STATUS_LABELS["sent"],
                post_send_status="failed",
                post_send_error=str(exc),
                retryable=False,
            )
        else:
            result.update(status="failed", execution_status="failed", error=str(exc),
                          message=STATUS_LABELS["failed"])
    finally:
        append_chat_log(name, site, "group_result", message=huashu,
                        response=result.get("response", ""), chat=latest_chat,
                        extra={**base_extra, "result": result})
    print(f"{get_now_time()} {name} {site} 第 {group_index}/{total_groups} 组：{result['message']}<br>")
    if result.get("reply_status") in {"reply_timeout", "needs_human"} or result.get("post_send_status"):
        setattr(driver, "_bit_ai_reset_before_group", True)
    return result


def click_contact_us(driver, name, site):
    """点击帮助页底部或页面中的 Contact us 入口。"""
    driver.switch_to.default_content()
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(2)

    contact_selectors = [
        (By.XPATH, "//*[self::a or self::button][contains(normalize-space(), 'Contact us')]"),
        (By.XPATH, "//*[contains(normalize-space(), 'Contact us')]"),
    ]
    for by, selector in contact_selectors:
        elements = driver.find_elements(by, selector)
        for element in elements:
            try:
                if not element.is_displayed():
                    continue
                driver.execute_script(
                    "arguments[0].scrollIntoView({block: 'center', inline: 'center'});",
                    element,
                )
                time.sleep(1)
                element.click()
                print(f"{get_now_time()} {name} {site} 点击 Contact us<br>")
                return True
            except Exception:
                continue

    clicked = driver.execute_script(
        """
        const candidates = [...document.querySelectorAll('a, button, span, div')]
            .filter((node) => node.innerText && node.innerText.trim().includes('Contact us'));
        const node = candidates.find((item) => {
            const rect = item.getBoundingClientRect();
            return rect.width > 0 && rect.height > 0;
        });
        if (!node) return false;
        node.scrollIntoView({block: 'center', inline: 'center'});
        node.click();
        return true;
        """
    )
    if clicked:
        print(f"{get_now_time()} {name} {site} JS 点击 Contact us<br>")
        return True
    return False


def click_ai_assistant_entry(driver, name, site):
    """优先按常见文案点击 AI Assistant 入口。

    页面可能使用 Shadow DOM，所以通过 JS 深度遍历元素，而不是只用普通 CSS 查找。
    """
    driver.switch_to.default_content()
    entry_texts = ("Ask the assistant", "AI Assistant", "Assistant", "助手", "助理", "个人助手")
    for text in entry_texts:
        try:
            clicked = driver.execute_script(
                """
                const text = arguments[0];
                function deepElements(root = document) {
                    const out = [];
                    const walk = (node) => {
                        const elements = node.querySelectorAll ? Array.from(node.querySelectorAll('*')) : [];
                        for (const el of elements) {
                            out.push(el);
                            if (el.shadowRoot) walk(el.shadowRoot);
                        }
                    };
                    walk(root);
                    return out;
                }
                const needle = text.toLowerCase();
                const candidates = deepElements()
                    .filter((node) =>
                        ['BUTTON', 'A'].includes(node.tagName) || node.getAttribute('role') === 'button'
                    )
                    .map((node) => {
                        const rect = node.getBoundingClientRect();
                        const label = [
                            node.innerText || '',
                            node.getAttribute('aria-label') || '',
                            node.getAttribute('title') || '',
                            node.getAttribute('data-testid') || '',
                            String(node.className || '')
                        ].join(' ').toLowerCase();
                        let score = rect.bottom + rect.right;
                        if (label.includes(needle)) score += 10000;
                        if (label.includes('assistant') || label.includes('maxwell')) score += 6000;
                        if (label.includes('助手') || label.includes('助理') || label.includes('个人助手')) score += 6000;
                        if (label.includes('chat')) score += 2500;
                        if (rect.top > window.innerHeight * 0.35) score += 3000;
                        if (rect.left > window.innerWidth * 0.45) score += 1500;
                        return {node, rect, label, score};
                    })
                    .filter((item) => item.rect.width > 0 && item.rect.height > 0)
                    .filter((item) => item.label.includes(needle))
                    .sort((a, b) => b.score - a.score);
                const node = candidates.length ? candidates[0].node : null;
                if (!node) return false;
                node.scrollIntoView({block: 'center', inline: 'center'});
                node.click();
                return true;
                """,
                text,
            )
            if clicked:
                print(f"{get_now_time()} {name} {site} 点击 {text}<br>")
                return True
        except Exception:
            continue
    return False


def click_ai_entry_fallback(driver, name, site):
    """AI 入口兜底点击：只按 Assistant/Maxwell 关键词寻找悬浮窗入口。"""
    driver.switch_to.default_content()
    try:
        clicked = driver.execute_script(
            """
            function deepElements(root = document) {
                const out = [];
                const walk = (node) => {
                    const elements = node.querySelectorAll ? Array.from(node.querySelectorAll('*')) : [];
                    for (const el of elements) {
                        out.push(el);
                        if (el.shadowRoot) walk(el.shadowRoot);
                    }
                };
                walk(root);
                return out;
            }
            const candidates = deepElements()
                .filter((node) =>
                    ['BUTTON', 'A'].includes(node.tagName) || node.getAttribute('role') === 'button'
                )
                .map((node) => {
                    const rect = node.getBoundingClientRect();
                    const label = [
                        node.innerText || '',
                        node.getAttribute('aria-label') || '',
                        node.getAttribute('title') || '',
                        node.getAttribute('data-testid') || '',
                        String(node.className || '')
                    ].join(' ').toLowerCase();
                    let score = rect.bottom + rect.right;
                    if (label.includes('maxwell')) score += 10000;
                    if (label.includes('assistant')) score += 8000;
                    if (label.includes('助手') || label.includes('助理') || label.includes('个人助手')) score += 8000;
                    if (label.includes('help')) score += 1500;
                    if (rect.top > window.innerHeight * 0.45) score += 3500;
                    if (rect.left > window.innerWidth * 0.55) score += 2500;
                    return {node, rect, label, score};
                })
                .filter((item) => item.rect.width > 0 && item.rect.height > 0)
                .filter((item) =>
                    item.label.includes('maxwell') ||
                    item.label.includes('assistant') ||
                    item.label.includes('助手') ||
                    item.label.includes('助理') ||
                    item.label.includes('个人助手')
                )
                .sort((a, b) => b.score - a.score);
            const node = candidates.length ? candidates[0].node : null;
            if (!node) return false;
            node.scrollIntoView({block: 'center', inline: 'center'});
            node.click();
            return true;
            """
        )
        if clicked:
            print(f"{get_now_time()} {name} {site} 点击AI入口兜底候选<br>")
            return True
    except Exception:
        pass
    return False


def is_top_level_human_customer_service_page(driver):
    """AI 客服只认悬浮窗；顶层 chat/maxwell 页面属于人工客服页面。"""
    try:
        driver.switch_to.default_content()
    except Exception:
        pass
    try:
        current_url = (driver.current_url or "").lower()
    except Exception:
        current_url = ""
    return "/help/chat" in current_url or "/maxwell/new-chat" in current_url


def wait_for_ai_chat_frame(driver, timeout=AI_CHAT_READY_TIMEOUT_SECONDS):
    """在限定时间内等待 AI 客服 iframe 出现并可切换。"""
    end_time = time.time() + timeout
    while time.time() < end_time:
        _check_appeal_control(driver)
        if is_top_level_human_customer_service_page(driver):
            return False
        if switch_to_ai_chat_frame(driver):
            return True
        time.sleep(0.5)
    dump_iframe_debug_info(driver)
    return False


def wait_for_ai_chat_ready(
    driver,
    timeout=AI_CHAT_READY_TIMEOUT_SECONDS,
    require_input=False,
):
    """等待任一受支持的聊天结构就绪，并返回对应模式。"""
    end_time = time.time() + timeout
    while time.time() < end_time:
        _check_appeal_control(driver)
        variant = detect_ai_chat_variant(driver)
        if variant == AI_CHAT_MODE_INLINE:
            state = get_ai_chat_dom_state(driver)
            if require_input:
                ready = bool(find_inline_chat_input(driver, timeout=0.5))
            else:
                ready = bool(state.get("inline_open") or state.get("inline_has_input"))
            if ready:
                driver.switch_to.default_content()
                setattr(driver, "_mercado_ai_chat_mode", AI_CHAT_MODE_INLINE)
                return AI_CHAT_MODE_INLINE

        # 旧版 iframe 可能嵌套在没有明显标记的外层 frame 中，始终保留递归探测兜底。
        if switch_to_ai_chat_frame(driver, require_input=require_input):
            setattr(driver, "_mercado_ai_chat_mode", AI_CHAT_MODE_IFRAME)
            return AI_CHAT_MODE_IFRAME
        time.sleep(0.5)

    driver.switch_to.default_content()
    return ""


def open_ai_contact_window(driver, name, site, window_id=""):
    """打开 Help 页面并自动进入新版内嵌助手或旧版 iframe 助手。

    不同账号的入口页和按钮文案可能不同，所以依次尝试多个 URL、多种入口点击方式；
    若失败会保存页面截图、HTML 和候选元素信息。
    """
    opened_mode = ""
    entered_human_page = False
    last_variant = ""
    for url in AI_HELP_URLS:
        _check_appeal_control(driver)
        driver.switch_to.default_content()
        backend_result = open_mercado_backend_page(
            driver,
            url,
            name,
            window_id,
            settle_seconds=AI_BACKEND_SETTLE_SECONDS,
            anomaly_site=site,
            anomaly_source="AI申诉",
        )
        _abort_ai_appeal_after_backend_recovery(
            backend_result,
            name,
            site,
            abort_after_rate_limit_recovery=bool(
                getattr(driver, "_bit_abort_ai_after_rate_limit_recovery", False)
            ),
        )
        print(f"{get_now_time()} {name} {site} 打开AI客服入口页面：{url}<br>")

        for attempt in range(1, 5):
            _check_appeal_control(driver)
            variant = detect_ai_chat_variant(driver)
            state = get_ai_chat_dom_state(driver)
            if variant:
                last_variant = variant
            print(
                f"{get_now_time()} {name} {site} 尝试打开AI客服，第 {attempt} 次，"
                f"探测模式={variant or 'unknown'}<br>"
            )

            opened_mode = wait_for_ai_chat_ready(
                driver,
                timeout=min(10, AI_CHAT_READY_TIMEOUT_SECONDS),
                require_input=True,
            )
            if opened_mode:
                break

            if state.get("visible_ai_frame_count", 0):
                # 面板已经展开时只等待 iframe 内部就绪。再次点击入口会把面板关掉，
                # 这正是旧逻辑在新 Shadow DOM 结构下反复开关、最终误报的原因。
                print(
                    f"{get_now_time()} {name} {site} AI客服面板已展开，等待 iframe 输入框加载<br>"
                )
                opened_mode = wait_for_ai_chat_ready(
                    driver,
                    timeout=AI_CHAT_READY_TIMEOUT_SECONDS,
                    require_input=True,
                )
                if opened_mode:
                    break
                driver.switch_to.default_content()
                time.sleep(1)
                continue

            if variant == AI_CHAT_MODE_INLINE:
                opened_mode = click_inline_ai_assistant_entry(
                    driver,
                    name,
                    site,
                    timeout=AI_CHAT_ENTRY_TIMEOUT_SECONDS,
                )
                if opened_mode:
                    break
                driver.switch_to.default_content()
                time.sleep(2)
                continue

            if click_ai_assistant_entry(driver, name, site):
                if is_top_level_human_customer_service_page(driver):
                    entered_human_page = True
                    print(f"{get_now_time()} {name} {site} 点击 Assistant 后进入人工客服页面，不按 AI 悬浮窗处理<br>")
                    break
                opened_mode = wait_for_ai_chat_ready(
                    driver,
                    timeout=AI_CHAT_READY_TIMEOUT_SECONDS,
                    require_input=False,
                )
                if opened_mode:
                    break

            if click_ai_entry_fallback(driver, name, site):
                if is_top_level_human_customer_service_page(driver):
                    entered_human_page = True
                    print(f"{get_now_time()} {name} {site} 兜底点击后进入人工客服页面，不按 AI 悬浮窗处理<br>")
                    break
                opened_mode = wait_for_ai_chat_ready(
                    driver,
                    timeout=AI_CHAT_READY_TIMEOUT_SECONDS,
                    require_input=False,
                )
                if opened_mode:
                    break

            driver.switch_to.default_content()
            time.sleep(2)

        if opened_mode or entered_human_page:
            break

    if not opened_mode:
        dump_iframe_debug_info(driver)
        dump_ai_entry_debug_info(driver)
        dump_ai_chat_mode_debug_info(driver)
        save_ai_open_debug_artifacts(driver, name, site)
        if entered_human_page:
            raise RuntimeError("进入了人工客服页面，不是 AI 客服悬浮窗")
        if last_variant == AI_CHAT_MODE_INLINE:
            raise RuntimeError("检测到新版内嵌 AI 助手，但打开后没有找到聊天输入框")
        raise RuntimeError("没有找到 AI 客服悬浮窗 iframe")

    setattr(driver, "_mercado_ai_chat_mode", opened_mode)
    if opened_mode == AI_CHAT_MODE_INLINE:
        driver.switch_to.default_content()
        input_box = find_inline_chat_input(
            driver,
            timeout=AI_CHAT_INPUT_TIMEOUT_SECONDS,
        )
    else:
        if not switch_to_ai_chat_frame(driver, require_input=False):
            dump_iframe_debug_info(driver)
            dump_ai_entry_debug_info(driver)
            dump_ai_chat_mode_debug_info(driver)
            save_ai_open_debug_artifacts(driver, name, site)
            raise RuntimeError("没有切换到旧版 AI 客服 iframe")
        recover_expired_ai_conversation(driver, name, site)
        input_box = find_chat_input(
            driver,
            timeout=AI_CHAT_INPUT_TIMEOUT_SECONDS,
            allow_default_content=False,
        )
        if not input_box and recover_expired_ai_conversation(driver, name, site):
            input_box = find_chat_input(
                driver,
                timeout=AI_CHAT_ENTRY_TIMEOUT_SECONDS,
                allow_default_content=False,
            )
        if not input_box and recover_expired_ai_conversation(driver, name, site, force=True):
            switch_to_ai_chat_frame(driver, require_input=False)
            input_box = find_chat_input(
                driver,
                timeout=AI_CHAT_ENTRY_TIMEOUT_SECONDS,
                allow_default_content=False,
            )
    if not input_box:
        dump_iframe_debug_info(driver)
        dump_ai_entry_debug_info(driver)
        dump_ai_chat_mode_debug_info(driver)
        save_ai_open_debug_artifacts(driver, name, site)
        raise RuntimeError(f"{opened_mode} AI 客服已打开，但没有找到输入框")
    print(f"{get_now_time()} {name} {site} 进入 AI 客服，模式={opened_mode}<br>")
    return opened_mode


def use_one_browser_run_task(info):
    """循环使用单个比特浏览器窗口执行申诉任务。"""
    # /browser/open 接口会返回 selenium使用的http地址，以及webdriver的path，直接使用即可
    name = info[0]
    site = info[1]
    form = info[2]
    message = info[3]

    try:
        ip_usable = True
        if ip_usable:
            while True:
                print("ip检测通过，打开店铺平台主页")

                try:
                    shensu(name, site, form, message)
                except Exception as e:
                    traceback.print_exc()
                    print("申诉执行异常", e)
                finally:
                    print(f"{get_now_time()} {name}{site} 本轮结束，标签页由申诉流程清理<br>")
                # 5分钟执行一次这个方法
                time.sleep(300)
        else:
            print("ip检测不通过，请检查")
    except:
        print("脚本运行异常:" + traceback.format_exc())


def normalize_site_code(site):
    """把用户传入的中文站点、平台代码或英文缩写统一为内部站点码。"""
    key = str(site or "").strip().upper()
    if key in SITE_CODE_MAP:
        return SITE_CODE_MAP[key]
    return SITE_CODE_MAP.get(str(site or "").strip(), key or "MX")


def normalize_site_name(site):
    """把用户传入的中文站点、平台代码或英文缩写统一为中文站点名，供 Selenium 切站点使用。"""
    site_code = normalize_site_code(site)
    return SITE_CODE_TO_NAME.get(site_code, "墨西哥")


def _split_config_sites(value):
    """把配置表中的中文站点列表拆成独立站点。"""
    text = str(value or "").strip()
    if not text:
        return []
    for sep in ("，", "、", ";", "；", "|"):
        text = text.replace(sep, ",")
    return [item.strip() for item in text.split(",") if item.strip()]


def load_active_shop_site_config():
    """从店铺授权读取已开启申诉的店铺站点。"""
    active = {}
    for row in list_shop_configs(
        include_ignored=False,
        authorization_flag="appeal_enabled",
    ):
        name = row["shop_name"]
        sites = _split_config_sites(row["sites"])
        active[str(name)] = {normalize_site_code(site) for site in sites}
    return active


def build_top_infraction_shop_plan(
    top_shops=5,
    recent_days=30,
    max_workers=8,
):
    """通过官方 API 重新读取侵权，并生成带完整编号的授权站点计划。"""

    # 延迟导入避免 bit_daily_task -> bit_appeal_ai 的模块循环；调用发生时
    # 两个模块均已初始化。任务模块与自动 AI 因而共用同一套 API 筛选规则。
    from bit import bit_daily_task

    return bit_daily_task.build_latest_infraction_appeal_plan(
        top_n=top_shops,
        recent_days=recent_days,
        min_infraction_count=0,
        max_workers=max_workers,
    )


def run_top_infraction_shop_once(shop_plan, site_pause=30):
    """按单个店铺内各站点侵权数量降序执行一轮 AI 侵权申诉。"""
    name = shop_plan["name"]
    results = []
    for site in shop_plan["sites"]:
        site_code = site["site_code"]
        count = site["count"]
        appeal_type = str(site.get("appeal_type") or "侵权")
        try:
            print(
                f"{get_now_time()} {name} {site_code} 开始处理{appeal_type}，"
                f"当前站点数量 {count}<br>"
            )
            appeal_kwargs = {}
            if appeal_type == "禁限售":
                appeal_kwargs["prohibited_ids"] = site.get("prohibited_ids") or ()
            else:
                appeal_kwargs["infraction_ids"] = site.get("infraction_ids") or ()
            result = shensu(
                name,
                site_code,
                appeal_type,
                "",
                **appeal_kwargs,
            )
            results.append({
                "site": site_code,
                "appeal_type": appeal_type,
                "count": count,
                "result": result,
            })
            print(
                f"{get_now_time()} {name} {site_code} {appeal_type}处理完成："
                f"{result}<br>"
            )
        except Exception as e:
            results.append({
                "site": site_code,
                "appeal_type": appeal_type,
                "count": count,
                "error": str(e),
            })
            print(f"{get_now_time()} {name} {site_code} {appeal_type}处理失败：{e}<br>")
            traceback.print_exc()
        if site_pause > 0:
            time.sleep(site_pause)
    return {"name": name, "total": shop_plan["total"], "results": results}


def run_top_infraction_appeal_round(max_windows=5, top_shops=5, recent_days=30, site_pause=30):
    """执行一轮：选择侵权最多的店铺，最多同时打开 max_windows 个窗口处理。"""
    print(f"{get_now_time()} 本轮开始，重新通过官方 API 读取全部授权站点侵权<br>")
    plan = build_top_infraction_shop_plan(
        top_shops=top_shops,
        recent_days=recent_days,
        max_workers=max_windows,
    )
    print(f"{get_now_time()} 本轮 API 侵权店铺计划：{plan}<br>")
    if not plan:
        print(f"{get_now_time()} 没有找到可处理的侵权店铺<br>")
        return []

    max_workers = max(1, min(int(max_windows), len(plan)))
    results = []
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = [executor.submit(run_top_infraction_shop_once, shop, site_pause) for shop in plan]
        for future in futures:
            try:
                results.append(future.result())
            except Exception as e:
                results.append({"error": str(e)})
                traceback.print_exc()
    print(f"{get_now_time()} 本轮所有店铺、全部授权站点侵权编号已处理完毕<br>")
    return results


def auto_appeal_top_infractions_loop(
    max_windows=5,
    top_shops=5,
    recent_days=30,
    round_interval=300,
    site_pause=30,
):
    """无限循环处理侵权最多的店铺；默认同时 5 个窗口、选择 5 个店铺。"""
    round_no = 1
    while True:
        started = time.time()
        try:
            print(
                f"{get_now_time()} 开始第 {round_no} 轮 Top 侵权店铺 AI 申诉，"
                "先重新读取 API 侵权列表<br>"
            )
            results = run_top_infraction_appeal_round(
                max_windows=max_windows,
                top_shops=top_shops,
                recent_days=recent_days,
                site_pause=site_pause,
            )
            print(f"{get_now_time()} 第 {round_no} 轮 Top 侵权店铺 AI 申诉完成：{results}<br>")
        except Exception as e:
            print(f"{get_now_time()} 第 {round_no} 轮 Top 侵权店铺 AI 申诉异常：{e}<br>")
            traceback.print_exc()

        sleep_seconds = max(0, int(round_interval) - (time.time() - started))
        print(
            f"{get_now_time()} 第 {round_no} 轮全部站点结束，等待 {sleep_seconds:.1f} 秒后"
            "重新读取 API 侵权列表<br>"
        )
        time.sleep(sleep_seconds)
        round_no += 1


def _unique_text_list(values):
    result = []
    seen = set()
    for value in values or []:
        if isinstance(value, (list, tuple, set)):
            items = value
        else:
            items = [value]
        for item in items:
            text = str(item or "").strip()
            if not text or text in seen:
                continue
            seen.add(text)
            result.append(text)
    return result


def _extract_identifiers_from_text(text):
    text = str(text or "")
    values = re.findall(r"\bML[A-Z]{1,3}\d{6,}\b|\b\d{8,}\b", text)
    return _unique_text_list(values)


def _chat_snapshot(value):
    snapshot = getattr(value, "snapshot", None)
    if isinstance(snapshot, dict):
        return snapshot
    if isinstance(value, dict) and isinstance(value.get("messages"), list):
        return value
    if isinstance(value, list):
        messages = []
        for index, item in enumerate(value):
            if isinstance(item, dict):
                messages.append(dict(item))
            elif str(item or "").strip():
                messages.append({
                    "id": f"legacy-{index}",
                    "role": "assistant",
                    "text": str(item).strip(),
                })
        return {"conversation_id": "", "epoch": "", "messages": messages}
    return None


def _collect_full_chat_history(log_records, final_agent_messages=None):
    """Merge overlapping DOM snapshots into complete, role-aware conversations."""
    sessions = {}
    values = [record.get("chat") for record in (log_records or [])]
    if final_agent_messages is not None:
        values.append(final_agent_messages)

    for snapshot_index, value in enumerate(values):
        snapshot = _chat_snapshot(value)
        if not snapshot:
            continue
        conversation_id = str(snapshot.get("conversation_id") or "")
        epoch = str(snapshot.get("epoch") or "")
        session_key = (conversation_id, epoch)
        if not any(session_key):
            session_key = ("legacy", "")
        session = sessions.setdefault(session_key, {
            "conversation_id": conversation_id,
            "epoch": epoch,
            "messages": [],
            "_message_indexes": {},
            "_fallback_keys": set(),
        })
        for position, original in enumerate(snapshot.get("messages") or []):
            if not isinstance(original, dict):
                continue
            message = {
                "id": str(original.get("id") or ""),
                "role": str(original.get("role") or ""),
                "text": str(original.get("text") or ""),
            }
            if not message["text"].strip():
                continue
            if message["id"]:
                identity = (message["role"], message["id"])
                existing_index = session["_message_indexes"].get(identity)
                if existing_index is None:
                    session["_message_indexes"][identity] = len(session["messages"])
                    session["messages"].append(message)
                else:
                    # Streaming responses keep one DOM id while the text grows.
                    session["messages"][existing_index] = message
            else:
                fallback = (snapshot_index, position, message["role"], message["text"])
                if fallback not in session["_fallback_keys"]:
                    session["_fallback_keys"].add(fallback)
                    session["messages"].append(message)

    result = []
    for session in sessions.values():
        session.pop("_message_indexes", None)
        session.pop("_fallback_keys", None)
        if session["messages"]:
            result.append(session)
    return result


def _collect_appeal_record_fields(log_records, final_agent_messages=None):
    appeal_messages = []
    identifiers = []
    ai_replies = []

    for record in log_records or []:
        message = record.get("message")
        if message:
            appeal_messages.append(message)
            identifiers.extend(_extract_identifiers_from_text(message))

        response = record.get("response")
        if response:
            ai_replies.append(response)


        extra = record.get("extra") or {}
        if isinstance(extra, dict):
            for key in (
                "infraction_ids",
                "prohibited_ids",
                "cancellation_ids",
                "complaint_order_ids",
                "delay_ids",
                "order_ids",
                "product_ids",
                "ids",
            ):
                value = extra.get(key)
                if isinstance(value, str):
                    identifiers.extend(_extract_identifiers_from_text(value))
                elif isinstance(value, (list, tuple, set)):
                    identifiers.extend(value)

    # Whole-window snapshots contain previous groups; only explicit response events are authoritative.
    ai_replies.extend(final_agent_messages or [])
    return {
        "appeal_content": "\n".join(_unique_text_list(appeal_messages)),
        "identifiers": _unique_text_list(identifiers),
        "ai_replies": _unique_text_list(ai_replies),
        "chat_history": _collect_full_chat_history(log_records, final_agent_messages),
    }


def summarize_ai_appeal_result(appeal_type, identifiers, appeal_content, ai_replies, force=False):
    """整理申诉记录，不再调用外部模型判断或总结申诉结果。

    客服原始回复仍会完整保存在 ``ai_replies`` 字段。这里仅把最后一条回复复制到
    列表页原有的摘要字段，方便查看；结果和成功/失败编号留待人工确认，避免本地
    规则误判客服语义。
    """
    identifiers = _unique_text_list(identifiers)
    ai_replies = _unique_text_list(ai_replies)
    latest_reply = ai_replies[-1] if ai_replies else ""
    return {
        "status": "待确认",
        "summary": (
            f"客服最后回复：{latest_reply}"
            if latest_reply
            else "未读取到 AI 客服回复，无法判断申诉结果。"
        ),
        "success_ids": [],
        "failed_ids": [],
        "error": "",
    }


def appeal_executor_metadata():
    """Identify the machine that performed an appeal without exposing addresses."""
    runtime_role = str(os.environ.get("BIT_RUNTIME_ROLE") or "server").strip().lower()
    execution_target = "local" if runtime_role == "client" else "server"
    return {
        "runtime_role": runtime_role,
        "execution_target": execution_target,
        "hostname": socket.gethostname(),
    }


def save_ai_appeal_record(
    appeal_time,
    appeal_type,
    shop_name,
    site,
    log_records,
    final_agent_messages=None,
    error="",
    execution=None,
):
    fields = _collect_appeal_record_fields(log_records, final_agent_messages)
    if str(error or "").strip() == "未登录":
        summary = {
            "status": "未登录",
            "summary": "店铺窗口显示登录邮箱输入页，本次未执行申诉操作。",
            "success_ids": [],
            "failed_ids": [],
            "error": "",
        }
    else:
        summary = summarize_ai_appeal_result(
            appeal_type,
            fields["identifiers"],
            fields["appeal_content"],
            fields["ai_replies"],
        )
    record = {
        "event_id": uuid.uuid4().hex,
        "appeal_time": appeal_time,
        "appeal_type": appeal_type,
        "shop_name": shop_name,
        "site": site,
        "status": summary["status"],
        "appeal_content": fields["appeal_content"],
        "identifiers": fields["identifiers"],
        "success_ids": summary["success_ids"],
        "failed_ids": summary["failed_ids"],
        "ai_replies": fields["ai_replies"],
        "chat_history": fields["chat_history"],
        "ai_summary": summary["summary"],
        "error": "\n".join(_unique_text_list([error, summary.get("error", "")])),
        "executor": appeal_executor_metadata(),
    }
    if execution:
        record["status"] = execution["message"]
        record["execution"] = execution
    write_local_record({"event": "appeal_record", "event_id": record["event_id"], "record": record})
    try:
        insert_ai_appeal_record(record)
        print(f"{get_now_time()} {shop_name} {site} AI申诉记录已入库<br>")
    except Exception as e:
        print(f"{get_now_time()} {shop_name} {site} AI申诉记录入库失败：{e}<br>")
    return record


def _collect_group_ai_replies(group_records):
    replies = []
    response_events = {
        "group_result",
        "delay_agent_reply",
        "agent_reply",
        "agent_reply_after_auto_reply",
        "agent_reply_after_site_option",
        "send_followup",
        "send_insist_after_site_option",
        "infraction_auto_reply_limit_reached",
        "send_followup_error",
        "send_infraction_error",
        "send_cancellation_error",
        "cancellation_auto_reply_limit_reached",
    }
    for record in group_records or []:
        event = record.get("event")
        response = record.get("response")
        if event in response_events and response:
            replies.append(response)
    return _unique_text_list(replies)


def _filter_group_log_records(
    group_records,
    shop_name,
    site,
    group_index,
    total_groups,
    infraction_ids,
    identifier_key="infraction_ids",
):
    filtered = []
    for record in group_records or []:
        extra = record.get("extra") if isinstance(record.get("extra"), dict) else {}
        if str(record.get("window") or "") != str(shop_name or ""):
            continue
        if str(record.get("site") or "") != str(site or ""):
            continue
        if str(extra.get("group_index") or "") != str(group_index):
            continue
        if str(extra.get("total_groups") or "") != str(total_groups):
            continue
        if str(extra.get(identifier_key) or "") != str(infraction_ids or ""):
            continue
        filtered.append(record)
    return filtered


def save_ai_appeal_group_record(
    appeal_time,
    shop_name,
    site,
    group_index,
    total_groups,
    infraction_ids,
    appeal_content,
    group_records,
    error="",
    appeal_kind="侵权",
):
    """每组侵权/取消率申诉结束后写入原始结果，不调用 DeepSeek 总结。"""
    identifiers = _extract_identifiers_from_text(infraction_ids)
    ai_replies = _collect_group_ai_replies(group_records)
    chat_history = _collect_full_chat_history(group_records)
    appeal_type = f"{appeal_kind}-第{group_index}/{total_groups}组"
    summary = summarize_ai_appeal_result(
        appeal_type,
        identifiers,
        appeal_content,
        ai_replies,
        force=True,
    )
    record = {
        "event_id": uuid.uuid4().hex,
        "appeal_time": appeal_time,
        "appeal_type": appeal_type,
        "shop_name": shop_name,
        "site": site,
        "status": summary["status"],
        "appeal_content": appeal_content,
        "identifiers": identifiers,
        "success_ids": summary["success_ids"],
        "failed_ids": summary["failed_ids"],
        "ai_replies": ai_replies,
        "chat_history": chat_history,
        "ai_summary": f"第{group_index}/{total_groups}组：{summary['summary']}",
        "error": "\n".join(_unique_text_list([error, summary.get("error", "")])),
        "record_scope": "group",
        "group_index": group_index,
        "total_groups": total_groups,
        "executor": appeal_executor_metadata(),
    }
    execution = result_from_logs(group_records, error=error)
    record["execution"] = execution
    record["status"] = execution["message"]
    record["error"] = error or "\n".join(g.get("error", "") for g in execution["groups"] if g.get("error"))
    write_local_record({"event": "appeal_record", "event_id": record["event_id"], "record": record})
    try:
        insert_ai_appeal_record(record)
        print(f"{get_now_time()} {shop_name} {site} 第{group_index}/{total_groups}组AI申诉记录已入库<br>")
    except Exception as e:
        print(f"{get_now_time()} {shop_name} {site} 第{group_index}/{total_groups}组AI申诉记录入库失败：{e}<br>")
    return record


# 申诉
def shensu(
    name, site, form, message, validate_open=False, infraction_ids=None,
    prohibited_ids=None, stop_event=None,
):
    """返回执行状态；收到回复不等于平台已批准申诉。"""
    print(f"{name} {site} 开始进行{form}申诉，自定义话术为{message}<br>")
    appeal_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    start_appeal_log_collection()
    site_name = normalize_site_name(site)
    driver, owned_window_lease = None, None
    window_id, appeal_error, failure_status = "", "", ""
    skip_close_tab = False
    outcome = execution_result("no_data")
    started = time.monotonic()
    try:
        nickname = random.choice(["Bruce", "Jack", "Lucy", "James"])
        selected_phrase = select_appeal_phrase(form) if not str(message or "").strip() else ""
        window_id = get_window_id_by_shop_name(name)
        if current_thread_window_lease(window_id) is None:
            owned_window_lease = create_window_lease(
                window_id, owner=f"ai_appeal:{name}", shop_name=name, task_type="ai_appeal",
            )
            if not owned_window_lease.acquire(timeout=0):
                raise AppealExecutionError("窗口正在被其他任务占用", "window_busy")
        driver, res = connect_bit_browser(window_id)
        name = res.get("data", {}).get("name") or name
        driver._bit_appeal_stop_event = stop_event if stop_event is not None else _APPEAL_STOP_EVENT.get()
        driver._bit_appeal_deadline = started + AI_SITE_BUDGET_SECONDS
        driver._bit_abort_ai_after_rate_limit_recovery = bool(validate_open)
        _check_appeal_control(driver)
        try:
            open_help_page_with_daily_validation(driver, name, site_name, window_id=window_id)
        except Exception:
            skip_close_tab = True
            raise
        select_site(driver, name, site_name)
        driver._bit_target_site = site_name
        with use_appeal_phrase(selected_phrase):
            if form == "侵权":
                handle_infraction(window_id, driver, name, site_name, message, nickname,
                                  infraction_ids=infraction_ids)
            elif form == "禁限售":
                handle_prohibited(window_id, driver, name, site_name, message, nickname,
                                  prohibited_ids=prohibited_ids)
            elif form == "延误":
                handle_delay(window_id, driver, name, site_name, message, nickname)
            elif form == "取消率":
                handle_cancellation(window_id, driver, name, site_name, message, nickname)
            elif form == "投诉":
                handle_complaint(window_id, driver, name, site_name, message, nickname)
            else:
                raise ValueError(f"不支持的申诉类型：{form}")
    except Exception as exc:
        appeal_error = str(exc)
        if isinstance(exc, AppealExecutionError):
            failure_status = exc.status
        elif "登录" in appeal_error:
            failure_status = "login_required"
        elif MERCADO_RATE_LIMIT_TEXT in appeal_error or "限频" in appeal_error:
            failure_status = "rate_limited"
        else:
            failure_status = "failed"
        print(f"{get_now_time()} {name} {site} 申诉执行异常：{appeal_error}<br>")
    finally:
        try:
            records = get_appeal_log_records()
            outcome.update(result_from_logs(records, error=appeal_error, status=failure_status))
            outcome["elapsed_seconds"] = round(time.monotonic() - started, 2)
            # Retry only before any message might have been submitted.
            outcome["retryable"] = bool(appeal_error and not outcome["sent"]
                                        and failure_status in {"failed", "rate_limited"})
            save_ai_appeal_record(appeal_time, form, name, site_name, records,
                                 error=appeal_error, execution=outcome)
            print(f"{get_now_time()} {name} {site} {outcome['message']}：{outcome['metrics']}<br>")
        except Exception as exc:
            print(f"{get_now_time()} 保存申诉执行结果失败：{exc}<br>")
        finally:
            stop_appeal_log_collection()
            try:
                if driver is not None:
                    if not skip_close_tab:
                        close_current_tab_keep_browser(driver, name, site)
                    # Stop the driver service without quitting the user's BitBrowser.
                    service = getattr(driver, "service", None)
                    if service is not None:
                        try:
                            service.stop()
                        except Exception as exc:
                            print(f"{get_now_time()} {name} {site} 驱动服务清理失败：{exc}<br>")
            finally:
                if owned_window_lease is not None:
                    owned_window_lease.release()
    return outcome


def handle_infraction(
    window_id,
    driver,
    name,
    site,
    message,
    nickname,
    infraction_ids=None,
):
    """处理侵权申诉：直接使用 API 编号，按最多 10 个逐组发送。"""
    group = 10
    inf_list = []
    seen_ids = set()
    for raw_id in infraction_ids or ():
        item_id = str(raw_id or "").strip().upper()
        if item_id and item_id not in seen_ids:
            seen_ids.add(item_id)
            inf_list.append(item_id)
    if infraction_ids is None:
        inf_list = get_infraction_orders(window_id, name, site)
    if not inf_list:
        print(f"{get_now_time()} {name} {site} 没有可以申诉的侵权编号<br>")
        return

    groups = [inf_list[i:i + group] for i in range(0, len(inf_list), group)]
    print(
        f"{get_now_time()} {name} {site} API侵权编号共 {len(inf_list)} 个，"
        f"每次对话最多 {group} 个，共 {len(groups)} 组<br>"
    )

    appeal_suffix = (
        f"这几个产品是通用品牌产品，并非侵权产品，这是系统误判，"
        f"麻烦帮我重新核查并删除侵权记录，谢谢"
    )
    selected_phrase = get_current_appeal_phrase()

    open_ai_contact_window(driver, name, site, window_id)
    for index, current_group in enumerate(groups, start=1):
        infraction_ids = "、".join(str(item) for item in current_group)
        huashu = (
            f"{infraction_ids}{message}"
            if message
            else render_appeal_phrase(
                selected_phrase,
                nickname=nickname,
                order_ids=infraction_ids,
                appeal_type="侵权",
            )
            if selected_phrase
            else f"{infraction_ids}{appeal_suffix}"
        )
        print(f"{get_now_time()} {name} {site} 开始发送第 {index}/{len(groups)} 组侵权申诉：{huashu}<br>")
        group_appeal_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        group_log_start = len(get_appeal_log_records())
        group_error = ""
        try:
            send_infraction_message_with_retry(driver, huashu, infraction_ids, name, site, index, len(groups))
        except Exception as e:
            group_error = str(e)
            raise
        finally:
            group_records = _filter_group_log_records(
                get_appeal_log_records()[group_log_start:],
                name,
                site,
                index,
                len(groups),
                infraction_ids,
            )
            save_ai_appeal_group_record(
                group_appeal_time,
                name,
                site,
                index,
                len(groups),
                infraction_ids,
                huashu,
                group_records,
                error=group_error,
            )
        print(f"{get_now_time()} {name} {site} 第 {index}/{len(groups)} 组侵权申诉处理完成<br>")
        if index < len(groups):
            _appeal_pause(driver, 20)


def handle_prohibited(
    window_id,
    driver,
    name,
    site,
    message,
    nickname,
    prohibited_ids=None,
):
    """处理禁限售申诉：读取禁限售列表，按最多 10 个编号独立发送。"""
    group_size = 10
    item_ids = []
    seen_ids = set()
    for raw_id in prohibited_ids or ():
        item_id = str(raw_id or "").strip().upper()
        if item_id and item_id not in seen_ids:
            seen_ids.add(item_id)
            item_ids.append(item_id)
    if prohibited_ids is None:
        item_ids = get_prohibited_listing_ids(name, site)
    if not item_ids:
        print(f"{get_now_time()} {name} {site} 禁限售列表没有可以申诉的产品编号<br>")
        return

    groups = [
        item_ids[index:index + group_size]
        for index in range(0, len(item_ids), group_size)
    ]
    print(
        f"{get_now_time()} {name} {site} 禁限售列表编号共 {len(item_ids)} 个，"
        f"每次对话最多 {group_size} 个，共 {len(groups)} 组；"
        "与普通侵权申诉分开处理<br>"
    )
    default_message = "亲爱客服，这个产品不是禁限售产品，他被系统误判了，麻烦你帮我恢复"
    selected_phrase = get_current_appeal_phrase()

    open_ai_contact_window(driver, name, site, window_id)
    for index, current_group in enumerate(groups, start=1):
        group_ids = "、".join(current_group)
        huashu = (
            f"{group_ids}{message}"
            if message
            else render_appeal_phrase(
                selected_phrase,
                nickname=nickname,
                order_ids=group_ids,
                appeal_type="禁限售",
            )
            if selected_phrase
            else f"{group_ids}{default_message}"
        )
        print(
            f"{get_now_time()} {name} {site} 开始发送第 {index}/{len(groups)} 组"
            f"禁限售申诉：{huashu}<br>"
        )
        group_appeal_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        group_log_start = len(get_appeal_log_records())
        group_error = ""
        try:
            send_infraction_message_with_retry(
                driver,
                huashu,
                group_ids,
                name,
                site,
                index,
                len(groups),
                appeal_kind="禁限售",
            )
        except Exception as exc:
            group_error = str(exc)
            raise
        finally:
            group_records = _filter_group_log_records(
                get_appeal_log_records()[group_log_start:],
                name,
                site,
                index,
                len(groups),
                group_ids,
                identifier_key="prohibited_ids",
            )
            save_ai_appeal_group_record(
                group_appeal_time,
                name,
                site,
                index,
                len(groups),
                group_ids,
                huashu,
                group_records,
                error=group_error,
                appeal_kind="禁限售",
            )
        print(
            f"{get_now_time()} {name} {site} 第 {index}/{len(groups)} 组"
            "禁限售申诉处理完成<br>"
        )
        if index < len(groups):
            _appeal_pause(driver, 20)



def handle_delay(window_id, driver, name, site, message, nickname):
    """延误与其他申诉共用发送确认、站点问答和结果记录。"""
    orders = get_delay_orders_download_list(window_id, name, site)
    if not orders:
        return
    groups = [orders[i:i + 5] for i in range(0, len(orders), 5)]
    selected_phrase = get_current_appeal_phrase()
    open_ai_contact_window(driver, name, site, window_id)
    for index, group in enumerate(groups, 1):
        ids = "、".join(str(item) for item in group)
        text = f"{ids}{message}" if message else render_appeal_phrase(
            selected_phrase or "请核查这些订单的延误责任，并复核对店铺声誉的影响。",
            nickname=nickname, order_ids=ids, appeal_type="延误",
        )
        group_started = len(get_appeal_log_records())
        error = ""
        try:
            send_infraction_message_with_retry(
                driver, text, ids, name, site, index, len(groups), appeal_kind="延误",
            )
        except Exception as exc:
            error = str(exc)
            raise
        finally:
            records = _filter_group_log_records(
                get_appeal_log_records()[group_started:], name, site, index, len(groups),
                ids, identifier_key="delay_ids",
            )
            save_ai_appeal_group_record(
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"), name, site,
                index, len(groups), ids, text, records, error=error, appeal_kind="延误",
            )
        if index < len(groups):
            _appeal_pause(driver, 20)


def handle_cancellation(window_id, driver, name, site, message, nickname):
    """处理取消率申诉：从声誉 Metrics 读取全部取消订单，按侵权规则分组处理。"""
    group_size = 10
    cancellation_orders = get_cancellation_orders(driver, name, site)
    if not cancellation_orders:
        print(f"{get_now_time()} {name} {site} 没有可以申诉的取消订单<br>")
        return

    groups = [
        cancellation_orders[index:index + group_size]
        for index in range(0, len(cancellation_orders), group_size)
    ]
    print(
        f"{get_now_time()} {name} {site} 取消订单共 {len(cancellation_orders)} 个，"
        f"按每组 {group_size} 个分为 {len(groups)} 组<br>"
    )

    # 取消订单读取结束时位于 Metrics 页面，返回帮助页后再打开 AI 客服。
    open_help_page_with_daily_validation(
        driver,
        name,
        site,
        window_id=window_id,
    )
    select_site(driver, name, site)
    open_ai_contact_window(driver, name, site, window_id)
    selected_phrase = get_current_appeal_phrase()

    for index, current_group in enumerate(groups, start=1):
        cancellation_ids = "、".join(str(item) for item in current_group)
        custom_message = str(message or "").strip()
        huashu = (
            f"{cancellation_ids}{custom_message}"
            if custom_message
            else render_appeal_phrase(
                selected_phrase,
                nickname=nickname,
                order_ids=cancellation_ids,
                appeal_type="取消率",
            )
            if selected_phrase
            else CANCELLATION_DEFAULT_APPEAL_TEMPLATE.format(
                order_ids=cancellation_ids,
            )
        )
        print(
            f"{get_now_time()} {name} {site} 开始发送第 {index}/{len(groups)} 组"
            f"取消率申诉：{huashu}<br>"
        )
        group_appeal_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        group_log_start = len(get_appeal_log_records())
        group_error = ""
        try:
            send_infraction_message_with_retry(
                driver,
                huashu,
                cancellation_ids,
                name,
                site,
                index,
                len(groups),
                appeal_kind="取消率",
            )
        except Exception as exc:
            group_error = str(exc)
            raise
        finally:
            group_records = _filter_group_log_records(
                get_appeal_log_records()[group_log_start:],
                name,
                site,
                index,
                len(groups),
                cancellation_ids,
                identifier_key="cancellation_ids",
            )
            save_ai_appeal_group_record(
                group_appeal_time,
                name,
                site,
                index,
                len(groups),
                cancellation_ids,
                huashu,
                group_records,
                error=group_error,
                appeal_kind="取消率",
            )
        print(
            f"{get_now_time()} {name} {site} 第 {index}/{len(groups)} 组"
            "取消率申诉处理完成<br>"
        )
        if index < len(groups):
            _appeal_pause(driver, 20)


def handle_complaint(window_id, driver, name, site, message, nickname):
    """处理投诉申诉：读取全部销售单号，每两个一组提交给 AI 客服。"""
    group_size = COMPLAINT_GROUP_SIZE
    complaint_orders = get_complaint_orders(driver, name, site)
    if not complaint_orders:
        print(f"{get_now_time()} {name} {site} 没有影响声誉的投诉销售单<br>")
        return

    groups = [
        complaint_orders[index:index + group_size]
        for index in range(0, len(complaint_orders), group_size)
    ]
    print(
        f"{get_now_time()} {name} {site} 投诉销售单共 {len(complaint_orders)} 个，"
        f"按每组 {group_size} 个分为 {len(groups)} 组<br>"
    )

    # 销售单读取结束时位于 Metrics 页面，返回帮助页后再打开 AI 客服。
    open_help_page_with_daily_validation(
        driver,
        name,
        site,
        window_id=window_id,
    )
    select_site(driver, name, site)
    open_ai_contact_window(driver, name, site, window_id)
    selected_phrase = get_current_appeal_phrase()

    for index, current_group in enumerate(groups, start=1):
        complaint_ids = "、".join(str(item) for item in current_group)
        custom_message = str(message or "").strip()
        huashu = (
            f"{complaint_ids}{custom_message}"
            if custom_message
            else render_appeal_phrase(
                selected_phrase,
                nickname=nickname,
                order_ids=complaint_ids,
                appeal_type="投诉",
            )
            if selected_phrase
            else f"销售单号：{complaint_ids}\n{COMPLAINT_DEFAULT_APPEAL_MESSAGE}"
        )
        print(
            f"{get_now_time()} {name} {site} 开始发送第 {index}/{len(groups)} 组"
            f"投诉申诉：{huashu}<br>"
        )
        group_appeal_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        group_log_start = len(get_appeal_log_records())
        group_error = ""
        try:
            send_infraction_message_with_retry(
                driver,
                huashu,
                complaint_ids,
                name,
                site,
                index,
                len(groups),
                appeal_kind="投诉",
            )
        except Exception as exc:
            group_error = str(exc)
            raise
        finally:
            group_records = _filter_group_log_records(
                get_appeal_log_records()[group_log_start:],
                name,
                site,
                index,
                len(groups),
                complaint_ids,
                identifier_key="complaint_order_ids",
            )
            save_ai_appeal_group_record(
                group_appeal_time,
                name,
                site,
                index,
                len(groups),
                complaint_ids,
                huashu,
                group_records,
                error=group_error,
                appeal_kind="投诉",
            )
        print(
            f"{get_now_time()} {name} {site} 第 {index}/{len(groups)} 组"
            "投诉申诉处理完成<br>"
        )
        if index < len(groups):
            _appeal_pause(driver, 20)


def handle_complain(driver, name, site, message, nickname):
    """保留旧函数名；新流程请调用需要 window_id 的 ``handle_complaint``。"""
    window_id = get_window_id_by_shop_name(name)
    return handle_complaint(window_id, driver, name, site, message, nickname)


def get_delay_orders_random(name, site, nums):
    """从最新延误表中随机抽取最近 15 天内的延误订单号。"""
    order_random = ""
    try:
        delay_folder_path = get_bit_path() / "美客多延误"
        delay_file = get_latest_modified_file(delay_folder_path)
        delay_file_path = delay_folder_path / delay_file
        fifteen_days_ago = datetime.now() - timedelta(days=15)
        order_list = []
        df = pd.read_excel(delay_file_path, engine='openpyxl')
        for index, row in df.iterrows():
            # print(row)
            line_name = row['店铺']
            line_site = row['站点']
            order_date = row['下单时间']
            order_num = row['销售单号']
            dispatch_date = row['实际揽收时间']
            if (line_name == name and line_site == site and dispatch_date != "Not yet dispatched"):
                order_date = parser_delay_date(order_date)
                if (order_date > fifteen_days_ago):
                    order_list.append(order_num)
        print(get_now_time() + name + site + "最近15天的延误个数:", len(order_list))

        if len(order_list) >= nums:
            order_random = str(random.sample(order_list, nums))
        else:
            order_random = str(order_list)
        order_random = re.sub(r'[^\d,]', '', order_random)

        print(get_now_time() + name + site + "随机得到的延误销售单号为", order_random)
    except Exception as e:
        print("获取延误表格信息失败", e)
    return order_random

def get_delay_orders_download_list(window_id, name, site):
    """下载最新延误报表，返回最近 15 天内已实际揽收的全部延误订单号。"""
    order_list = []
    try:
        download_start_time = time.time()
        message = download_relay_mail(window_id, site)
        print(get_now_time() + name + site + "延误报表下载结果:", message)

        if message != "下载文件成功":
            raise RuntimeError(f"延误报表下载失败：{message}")

        delay_file_path = save_latest_delay_report_to_excel(
            window_id, name, site, download_start_time
        )
        if delay_file_path is None:
            raise RuntimeError("没有找到刚下载的延误报表")

        fifteen_days_ago = datetime.now() - timedelta(days=15)
        df = pd.read_excel(delay_file_path, engine='openpyxl')
        for index, row in df.iterrows():
            # print(row)
            line_name = row['店铺']
            line_site = row['站点']
            order_date = row['下单时间']
            order_num = row['销售单号']
            dispatch_date = row['实际揽收时间']
            if (line_name == name and line_site == site and dispatch_date != "Not yet dispatched"):
                order_date = parser_delay_date(order_date)
                if (order_date > fifteen_days_ago):
                    order_list.append(str(order_num).strip().lstrip("'"))
        print(get_now_time() + name + site + "最近15天的延误个数:", len(order_list))
    except Exception as e:
        raise RuntimeError(f"获取延误表格信息失败：{e}") from e
    return order_list


def get_delay_orders_download_random(window_id,name, site, nums):
    """下载最新延误报表，写入汇总 Excel，再随机抽取最近 15 天内的延误订单号。"""
    order_random = ""
    try:
        order_list = get_delay_orders_download_list(window_id, name, site)
        if len(order_list) >= nums:
            order_random = str(random.sample(order_list, nums))
        else:
            order_random = str(order_list)
        order_random = re.sub(r'[^\d,]', '', order_random)
        print(get_now_time() + name + site + "随机得到的延误销售单号为", order_random)
    except Exception as e:
        print("获取延误表格信息失败", e)
    return order_random


def save_latest_delay_report_to_excel(window_id, name, site, download_start_time):
    """把当前店铺刚下载的延误 CSV/XLSX 转成统一格式的延误汇总 Excel。"""
    report_file = find_latest_delay_report(window_id, name, site, download_start_time)
    if report_file is None:
        return None

    delayrate = get_latest_delay_rate(name, site)
    updated_at = datetime.fromtimestamp(report_file.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")
    rows = read_delay_report_rows(report_file, name, site, delayrate, updated_at)
    if not rows:
        print(get_now_time() + name + site + "下载的延误报表为空:", report_file)
        return None

    df = pd.DataFrame(
        rows,
        columns=[
            "店铺",
            "站点",
            "延误率",
            "下单时间",
            "销售单号",
            "订单标题",
            "截止延误时间",
            "实际揽收时间",
            "更新时间",
            "文件路径",
        ],
    )
    delay_folder_path = get_bit_path() / "美客多延误"
    delay_folder_path.mkdir(parents=True, exist_ok=True)
    date_str = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
    safe_name = re.sub(r'[\\/:*?"<>|]', "_", str(name))
    safe_site = re.sub(r'[\\/:*?"<>|]', "_", str(site))
    output_path = delay_folder_path / f"{safe_name}{safe_site}延误信息汇总{date_str}.xlsx"
    df.to_excel(output_path, index=False)
    print(get_now_time() + name + site + "延误数据已写入:", output_path)
    return output_path


def find_latest_delay_report(window_id, name, site, download_start_time):
    """在当前 Bit 窗口下载目录中找到刚生成的延误报表。"""
    download_folder = get_browser_download_folder(window_id, name)
    if download_folder is None:
        return None

    site_code = get_site_code(site)
    patterns = ["Report_of_orders_that_you_shipped_with_delayed_handling_time_*"]
    deadline = time.time() + 60
    latest_files = []
    while time.time() < deadline:
        files = []
        for pattern in patterns:
            files.extend(download_folder.glob(pattern + ".csv"))
            files.extend(download_folder.glob(pattern + ".xlsx"))

        if site_code:
            files = [file for file in files if f"_{site_code}_" in file.name]
        latest_files = files

        fresh_files = [
            file for file in files
            if file.is_file() and file.stat().st_mtime >= download_start_time - 300
        ]
        if fresh_files:
            return max(fresh_files, key=lambda file: file.stat().st_mtime)
        time.sleep(3)

    if latest_files:
        return max(latest_files, key=lambda file: file.stat().st_mtime)
    return None


def get_browser_download_folder(window_id, name):
    """根据数据库配置中的下载序号定位浏览器下载目录。"""
    seq = ""
    try:
        record = get_shop_config(shop_name=name, window_id=window_id)
        seq = record["sequence_no"] if record else ""
    except Exception as e:
        print("读取比特下载目录配置失败", e)

    candidates = []
    if sys.platform == "win32" and seq:
        candidates.append(Path("C:/BitDownload") / seq)
    if seq:
        candidates.extend([
            Path.home() / "Downloads" / seq,
            Path("/Users/active11/Downloads") / seq,
        ])
    candidates.append(Path.home() / "Downloads")

    for folder in candidates:
        if folder.exists():
            return folder
    print("没有找到 Bit 下载目录，候选目录:", candidates)
    return None


def read_delay_report_rows(report_file, name, site, delayrate, updated_at):
    """读取 Mercado 延误报表，转换为项目统一的延误列。"""
    rows = []
    if report_file.suffix.lower() == ".csv":
        df = pd.read_csv(report_file, header=None, skiprows=1)
    else:
        df = pd.read_excel(report_file, header=None, skiprows=1, engine="openpyxl")

    for _, row in df.iterrows():
        try:
            rows.append(
                (
                    name,
                    site,
                    delayrate,
                    row.iloc[0],
                    row.iloc[1],
                    row.iloc[2],
                    row.iloc[5],
                    row.iloc[6],
                    updated_at,
                    str(report_file),
                )
            )
        except Exception as e:
            print("解析延误报表行失败", e)
    return rows


def get_latest_delay_rate(name, site):
    """从最新声誉表中取当前店铺站点的延误率。"""
    try:
        reputation_folder = get_bit_path() / "美客多声誉"
        reputation_file = get_latest_modified_file(reputation_folder)
        if reputation_file is None:
            return ""
        df = pd.read_excel(reputation_folder / reputation_file, engine="openpyxl")
        for _, row in df.iterrows():
            if row.get("店铺名") == name and row.get("站点") == site:
                delayrate = row.get("延误率")
                if pd.isna(delayrate):
                    return ""
                return delayrate
    except Exception as e:
        print("读取最新声誉延误率失败", e)
    return ""


def get_site_code(site):
    site_map = {
        "墨西哥": "MX",
        "巴西": "BR",
        "哥伦比亚": "CO",
        "智利": "CL",
        "阿根廷": "AR",
        "乌拉圭": "UY",
    }
    return site_map.get(site, "")


def _is_ai_infringement_record(record):
    """防御性过滤：AI 侵权申诉不接受 reports/权利人举报记录。"""
    if isinstance(record, dict):
        record_type = record.get("类型", record.get("type", ""))
    else:
        try:
            record_type = record[7] if len(record) > 7 else ""
        except (TypeError, IndexError):
            return False
    normalized_type = str(record_type or "").strip().lower()
    if not normalized_type:
        # 兼容旧采集器的三列返回值；新采集器会明确返回类型。
        return True
    return normalized_type in {
        "侵权",
        "infringement",
        "infringements",
        "detection",
        "detections",
    }


def get_infraction_orders_random(window_id, name, site, nums):
    """仅从官方 API 的侵权列表中随机抽取指定数量的编号。"""
    try:
        inf_list = get_infraction_orders(window_id, name, site)
        selected = (
            random.sample(inf_list, nums)
            if len(inf_list) > nums
            else inf_list
        )
        inf_list = "、".join(str(item) for item in selected)
        print(
            get_now_time() + name + site
            + f"从官方 API 随机得到 {len(selected)} 个侵权编号：",
            inf_list,
        )
    except Exception as e:
        print("获取侵权订单信息失败", e)
    return inf_list


def _appeal_setting_enabled(value):
    if isinstance(value, str):
        return value.strip().casefold() not in ("", "0", "false", "no", "off")
    return bool(value)


def _find_infraction_api_target(name, site):
    """Resolve the authorized token/site used by the direct moderation API."""

    requested_name = str(name or "").strip().casefold()
    site_name = normalize_site_name(site)
    site_id = SITE_ID_MAP.get(site_name, "")
    for token in (list_mercado_store_tokens() or {}).get("rows") or ():
        if not bool(token.get("enabled", True)):
            continue
        aliases = [
            str(value or "").strip()
            for value in (token.get("display_name"), token.get("nickname"))
            if str(value or "").strip()
        ]
        if not requested_name or not any(
            alias.casefold() == requested_name for alias in aliases
        ):
            continue
        enabled_sites = {
            str(setting.get("site_id") or "").strip().upper()
            for setting in (token.get("site_settings") or ())
            if _appeal_setting_enabled(setting.get("appeal_enabled"))
        }
        if site_id not in enabled_sites:
            return None
        try:
            token_id = int(token.get("id") or 0)
        except (TypeError, ValueError):
            return None
        if token_id <= 0:
            return None
        return {
            "token_id": token_id,
            "name": aliases[0] if aliases else str(name or "").strip(),
            "aliases": aliases,
            "site_ids": [site_id],
        }
    return None


def get_prohibited_listing_ids(name, site):
    """从当前禁限售列表读取指定店铺、站点的全部产品编号。"""
    target = _find_infraction_api_target(name, site)
    if not target:
        print(
            f"{get_now_time()} {name} {site} 未找到已开启申诉的店铺授权站点，"
            "不读取禁限售列表<br>"
        )
        return []

    site_code = normalize_site_code(site)
    item_ids = []
    seen_ids = set()
    page = 1
    while True:
        data = list_mercado_prohibited_listings(
            token_id=target["token_id"],
            risk_type="prohibited",
            page=page,
            page_size=500,
        ) or {}
        for row in data.get("rows") or ():
            if normalize_site_code(row.get("site_id")) != site_code:
                continue
            item_id = str(row.get("item_id") or "").strip().upper()
            if item_id and item_id not in seen_ids:
                seen_ids.add(item_id)
                item_ids.append(item_id)
        if page >= max(1, int(data.get("pages") or 1)):
            break
        page += 1
    print(
        f"{get_now_time()} {name} {site} 当前禁限售列表得到产品编号 "
        f"{len(item_ids)} 个<br>"
    )
    return item_ids


def get_infraction_orders(window_id, name, site):
    """直接从官方 Moderations API 读取侵权编号，绝不读取侵权网页。"""

    del window_id
    inf_list = []
    try:
        target = _find_infraction_api_target(name, site)
        if not target:
            print(
                f"{get_now_time()} {name} {site} 未找到已开启申诉的店铺授权站点，"
                "不读取侵权网页<br>"
            )
            return []
        site_id = target["site_ids"][0]
        result = mercado_infraction_sync.collect_live_detection_infractions(
            [target],
            recent_days=100,
            max_workers=1,
        )
        if result.get("failed_stores"):
            raise RuntimeError(f"侵权 API 采集不完整：{result['failed_stores']}")
        seen_ids = set()
        for row in result.get("data") or ():
            if not _is_ai_infringement_record(row):
                continue
            if not mercado_infraction_sync.is_auto_appeal_eligible_detection(row):
                continue
            if mercado_infraction_sync.is_prohibited_detection(row):
                continue
            if str(row.get("站点") or "").strip().upper() != site_id:
                continue
            item_id = str(row.get("编号") or "").strip().upper()
            if item_id and item_id not in seen_ids:
                seen_ids.add(item_id)
                inf_list.append(item_id)
        print(
            f"{get_now_time()} {name} {site} API得到侵权编号 {len(inf_list)} 个；"
            "已排除权利人案件及禁限售产品；禁限售将独立申诉<br>"
        )
    except Exception as e:
        raise RuntimeError(f"{name} {site} API获取侵权订单信息失败：{e}") from e
    return inf_list


# 检查聊天是否结束
def checkChatEnd(driver, name, site):
    """检查 AI 客服会话是否已经结束。"""
    try:
        if not activate_ai_chat_context(driver, require_input=False):
            return False
        WebDriverWait(driver, 5).until(
            EC.visibility_of_element_located(
                (
                    By.XPATH,
                    "//*[contains(text(), 'This chat has ended') or contains(text(), 'chat has ended') "
                    "or contains(text(), '聊天已结束') or contains(text(), '对话已结束')]",
                )
            )
        )
        print(f"{get_now_time()} {name}{site}聊天已经结束,结束AI找客服<br>")
        return True
    except Exception as e:
        return False
    return False


def get_inline_agent_messages(driver):
    """从新版顶层 #sa-assistant-chat 容器读取客服侧消息。"""
    driver.switch_to.default_content()
    try:
        return driver.execute_script(
            """
            function deepElements(root = document) {
                const out = [];
                const walk = (node) => {
                    const elements = node.querySelectorAll ? Array.from(node.querySelectorAll('*')) : [];
                    for (const el of elements) {
                        out.push(el);
                        if (el.shadowRoot) walk(el.shadowRoot);
                    }
                };
                walk(root);
                return out;
            }

            function visible(el) {
                const rect = el.getBoundingClientRect();
                const style = window.getComputedStyle(el);
                return !!(rect.width || rect.height || el.getClientRects().length)
                    && style.visibility !== 'hidden'
                    && style.display !== 'none';
            }

            function marker(el) {
                return [
                    String(el.className || ''),
                    el.getAttribute('data-testid') || '',
                    el.getAttribute('data-role') || '',
                    el.getAttribute('data-author') || '',
                    el.getAttribute('data-sender') || '',
                    el.getAttribute('data-message-author') || '',
                    el.getAttribute('aria-label') || ''
                ].join(' ').toLowerCase();
            }

            function isAgentMessage(el) {
                const value = marker(el);
                const agent = /from-agent|agent-message|assistant-message|message[^ ]*[-_ ](agent|assistant)|sender[^ ]*[-_ ](agent|assistant)|\\b(agent|assistant)\\b/.test(value);
                const user = /from-user|user-message|seller-message|message[^ ]*[-_ ](user|seller|client)/.test(value);
                return agent && !user;
            }

            const all = deepElements(document);
            const root = all.find((el) => el.id === 'sa-assistant-chat');
            if (!root) return [];
            return deepElements(root)
                .filter((el) => visible(el) && isAgentMessage(el))
                .filter((el) => !deepElements(el).some((child) => visible(child) && isAgentMessage(child)))
                .map((el) => (el.innerText || '').trim())
                .filter(Boolean);
            """
        ) or []
    except Exception:
        return []


def get_agent_messages(driver):
    """读取 AI 客服窗口中客服侧的消息文本。"""
    mode = activate_ai_chat_context(driver, require_input=False)
    if not mode:
        return []
    if mode == AI_CHAT_MODE_INLINE:
        messages = get_inline_agent_messages(driver)
        if messages:
            print("AI客服明文回复：<br>" + "<br>".join(messages) + "<br>")
        else:
            print("AI客服明文回复：暂无<br>")
        return messages

    message_selectors = [
        # 2026-07 新版 Maxwell：完成的回复放在 message-item--assistant 下。
        # 限定到 message-item 可排除仍在生成中的 thinking-indicator。
        (By.CSS_SELECTOR, ".message-item--assistant .chat-message__content"),
        (By.CSS_SELECTOR, ".chat-ui-message-bubble.chat-ui-message-bubble--from-agent"),
        (By.CSS_SELECTOR, "[class*='message-bubble--from-agent']"),
        (By.CSS_SELECTOR, "[class*='message'][class*='from-agent']"),
        (By.CSS_SELECTOR, "[class*='message'][class*='agent']"),
    ]
    messages = []
    for by, selector in message_selectors:
        elements = driver.find_elements(by, selector)
        for element in elements:
            try:
                # Selenium 的 text 只返回浏览器中可见的 innerText；不读取隐藏脚本、
                # textContent、aria-label 或整页源码。
                if not element.is_displayed():
                    continue
                text = element.text.strip()
                # 不按文本去重：AI 对不同分组可能连续回复完全相同的站点选项菜单，
                # 去重会导致后续相同回复无法被 wait_for_ai_agent_reply 识别为新消息。
                if text:
                    messages.append(text)
            except Exception:
                continue
        if messages:
            break
    if messages:
        print("AI客服明文回复：<br>" + "<br>".join(messages) + "<br>")
    else:
        print("AI客服明文回复：暂无<br>")
    return messages


def chat_script(driver):
    """旧版聊天脚本预留函数。"""
    return None


def use_all_browser_run_task(browser_list):
    """
    循环打开所有店铺运行脚本
    :param browser_list: 店铺列表
    """
    for browser in browser_list:
        use_one_browser_run_task(browser)


def use_all_browser_run_task_with_thread_pool(browser_list, max_threads=10):
    """
    使用线程池控制最大并发线程数
    :param browser_list: 店铺列表
    :param max_threads: 最大并发线程数
    """
    with ThreadPoolExecutor(max_workers=max_threads) as executor:
        executor.map(use_one_browser_run_task, browser_list)


def auto_appeal_delay():
    """根据最新延误报表筛选延误率较高的店铺，并批量发起延误申诉。"""
    fold_path = get_bit_path() / "美客多延误"
    file_path = fold_path / get_latest_modified_file(fold_path)
    wb = load_workbook(file_path)
    sheet = wb.active
    # 使用 min_row=2 跳过第一行

    name_site = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        delayrate = row[2]
        if delayrate != None and delayrate != "":
            delay_value = 0.0
            if type(delayrate) == str:
                delay_value = float(delayrate.strip("%")) / 100
            else:
                delay_value = float(delayrate)
            if delay_value >= 0.07:
                name_site.add((row[0], row[1], delay_value))

    print(len(name_site))

    list_appeal = []
    for i in name_site:
        list_appeal.append((i[0], i[1], "延误", ""))

    print(list_appeal)

    use_all_browser_run_task_with_thread_pool(list_appeal, 5)


if __name__ == "__main__":
    auto_appeal_top_infractions_loop()
