"""
# 适用环境python3
"""

import json
import time
import traceback
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

import requests
from pydantic.v1.datetime_parse import parse_date
from selenium.webdriver.chrome.service import Service

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import random

from bit.bit_utils import get_latest_modified_file, get_bit_path, parser_delay_date, get_now_time, getWindowidByName
from bit.bit_api import *
from bit.bit_config import get_window_id_by_shop_name
from bit.bit_mercado_limit import is_mercado_rate_limited_page
from bit.bit_mercado_login import open_mercado_backend_page
from bit.bit_appeal_phrases import render_appeal_phrase, select_appeal_phrase
from bit.bit_reputation_info import get_cancellation_orders
from AI_Agent.qianwen import *
import pandas as pd
from datetime import datetime, timedelta
from datetime import datetime
from AI_Agent.deepseek import *
import re
from openpyxl import load_workbook
import traceback


CHAT_INFO_API_URL = "https://zeshun.nat100.top/api/v1/chat"
HUMAN_SERVICE_HUB_URL = "https://global-selling.mercadolibre.com/help/hub/30928?source"
HUMAN_SERVICE_CHAT_V2_URL = "https://global-selling.mercadolibre.com/help/chat/v2"
SITE_REMOTE_VALUE_MAP = {
    "墨西哥": "MLM-remote",
    "巴西": "MLB-remote",
    "哥伦比亚": "MCO-remote",
    "智利": "MLC-remote",
    "阿根廷": "MLA-remote",
    "乌拉圭": "MLU-remote",
}

SITE_LABEL_MAP = {
    "墨西哥": ("Mexico", "México", "墨西哥", "MLM"),
    "巴西": ("Brazil", "Brasil", "巴西", "MLB"),
    "哥伦比亚": ("Colombia", "哥伦比亚", "MCO"),
    "智利": ("Chile", "智利", "MLC"),
    "阿根廷": ("Argentina", "阿根廷", "MLA"),
    "乌拉圭": ("Uruguay", "乌拉圭", "MLU"),
}


def select_mercado_site_fast(driver, name, site):
    """快速切换美客多站点，优先用 JS 深度查找，减少重复重试。"""
    remote_value = SITE_REMOTE_VALUE_MAP.get(site, "MLM-remote")
    labels = SITE_LABEL_MAP.get(site, SITE_LABEL_MAP["墨西哥"])
    result = driver.execute_script(
        """
        const remoteValue = arguments[0];
        const labels = arguments[1].map(item => String(item).toLowerCase());

        function allElements(root) {
            const out = [];
            const walk = (node) => {
                const elements = node.querySelectorAll ? Array.from(node.querySelectorAll('*')) : [];
                for (const el of elements) {
                    out.push(el);
                    if (el.shadowRoot) walk(el.shadowRoot);
                }
            };
            walk(root || document);
            return out;
        }

        function visible(el) {
            const rect = el.getBoundingClientRect();
            return rect.width > 0 && rect.height > 0;
        }

        function textOf(el) {
            return [
                el.innerText || '',
                el.textContent || '',
                el.getAttribute('aria-label') || '',
                el.getAttribute('title') || '',
                el.getAttribute('data-value') || ''
            ].join(' ').toLowerCase();
        }

        const elements = allElements(document);
        const switcher = elements.find(el =>
            visible(el) && (
                String(el.className || '').includes('nav-header-cbt__site-switcher') ||
                /select\\s+(country|site)|country|site/i.test(textOf(el))
            )
        );

        if (switcher && labels.some(label => textOf(switcher).includes(label))) {
            return 'already';
        }

        if (switcher) {
            switcher.click();
        }

        const target = allElements(document).find(el =>
            visible(el) && (
                el.getAttribute('data-value') === remoteValue ||
                labels.some(label => textOf(el).includes(label))
            )
        );
        if (!target) {
            return switcher ? 'opened_no_target' : 'no_switcher';
        }
        target.scrollIntoView({block: 'center'});
        target.click();
        return 'clicked';
        """,
        remote_value,
        list(labels),
    )
    print(f"{get_now_time()} {name} {site} 快速选择站点结果：{result}<br>")
    if result in ("already", "clicked"):
        time.sleep(2)
        if result == "clicked":
            driver.refresh()
            time.sleep(2)
        return True
    return False


def insert_chat_info_by_api(name, site, message, chat, response, time):
    payload = {
        "name": name,
        "site": site,
        "message": message,
        "chat": chat,
        "response": response,
        "time": time
    }
    res = requests.post(CHAT_INFO_API_URL, json=payload, timeout=10)
    res.raise_for_status()
    return res.json()


def fast_navigate(driver, url, stop_after=4):
    """快速跳转页面：不等待所有资源加载，避免 Mercado help 页面长时间卡住。"""
    driver.switch_to.default_content()
    try:
        driver.execute_script("window.location.href = arguments[0];", url)
    except Exception:
        try:
            driver.set_page_load_timeout(stop_after)
            driver.get(url)
        except Exception:
            pass
    time.sleep(stop_after)
    try:
        driver.execute_script("window.stop();")
    except Exception:
        pass


def open_human_service_hub_with_ip_retry(
    driver,
    name,
    site,
    max_hongkong_switches=3,
    window_id="",
):
    """打开人工客服 Hub，同时处理指定限频页和退出登录。"""
    result = open_mercado_backend_page(
        driver,
        HUMAN_SERVICE_HUB_URL,
        name,
        window_id,
        settle_seconds=3,
        max_rate_limit_retries=max_hongkong_switches,
        rate_limit_retry_wait_seconds=0,
        navigate=lambda url: fast_navigate(driver, url, stop_after=3),
        anomaly_site=site,
        anomaly_source="人工申诉",
    )
    if not result.get("ok"):
        raise RuntimeError(f"{name} {site} {result.get('message') or result.get('status')}")
    return True


def fast_open_new_tab(driver, url, stop_after=4):
    """用新标签页打开页面，并停止长时间加载的资源。"""
    driver.switch_to.default_content()
    try:
        before_handles = set(driver.window_handles)
    except Exception:
        before_handles = set()
    try:
        driver.execute_script("window.open(arguments[0], '_blank');", url)
        handles = driver.window_handles
        new_handles = [handle for handle in handles if handle not in before_handles]
        target_handle = new_handles[-1] if new_handles else handles[-1]
        driver.switch_to.window(target_handle)
    except Exception:
        try:
            driver.switch_to.new_window("tab")
        except Exception:
            pass
        fast_navigate(driver, url, stop_after=stop_after)
        try:
            return driver.current_window_handle
        except Exception:
            return None
    time.sleep(stop_after)
    try:
        driver.execute_script("window.stop();")
    except Exception:
        pass
    try:
        return driver.current_window_handle
    except Exception:
        return None


def page_contains_all_texts(driver, texts, timeout=6):
    end_time = time.time() + timeout
    while time.time() < end_time:
        try:
            page_text = driver.execute_script("return document.body ? document.body.innerText : ''") or ""
            if all(re.search(rf"\b{re.escape(text)}\b", page_text, re.I) for text in texts):
                return True
        except Exception:
            pass
        time.sleep(0.5)
    return False


def switch_to_latest_valid_hub_tab(driver, expected_texts, timeout=2):
    """从右往左找到正常 hub 标签并切换过去。"""
    try:
        handles = driver.window_handles[:]
    except Exception:
        return False

    for handle in reversed(handles):
        try:
            driver.switch_to.window(handle)
            current_url = driver.current_url or ""
            if "global-selling.mercadolibre.com/help/hub/30928" not in current_url:
                continue
            if page_contains_all_texts(driver, expected_texts, timeout=timeout):
                return True
        except Exception:
            continue
    return False


def open_hub_new_tab_with_retry(driver, name, site, retries=3, window_id=""):
    hub_url = "https://global-selling.mercadolibre.com/help/hub/30928?source"
    expected_texts = ("Chat", "Email")
    for attempt in range(1, retries + 1):
        try:
            before_handles = set(driver.window_handles)
            previous_handle = driver.current_window_handle
        except Exception:
            before_handles = set()
            previous_handle = None

        opened_handle = fast_open_new_tab(driver, "about:blank", stop_after=0)
        try:
            if opened_handle in driver.window_handles:
                driver.switch_to.window(opened_handle)
        except Exception:
            pass

        open_result = open_mercado_backend_page(
            driver,
            hub_url,
            name,
            window_id,
            settle_seconds=0,
            rate_limit_retry_wait_seconds=0,
            navigate=lambda url: fast_navigate(driver, url, stop_after=3),
            anomaly_site=site,
            anomaly_source="人工申诉",
        )

        if open_result.get("ok") and (
            page_contains_all_texts(driver, expected_texts, timeout=6)
            or switch_to_latest_valid_hub_tab(driver, expected_texts)
        ):
            print(f"{get_now_time()} {name} {site} hub 页面打开正常，第{attempt}次成功<br>")
            return True

        reason = open_result.get("message") or "未看到 Chat 和 Email"
        print(f"{get_now_time()} {name} {site} hub 页面异常：{reason}，第{attempt}次重试<br>")
        try:
            current_handle = opened_handle or driver.current_window_handle
            if current_handle not in before_handles and current_handle in driver.window_handles and len(driver.window_handles) > 1:
                driver.switch_to.window(current_handle)
                driver.close()
            remaining_handles = driver.window_handles
            if previous_handle in remaining_handles:
                driver.switch_to.window(previous_handle)
            elif remaining_handles:
                driver.switch_to.window(remaining_handles[-1])
        except Exception:
            pass

    return False


def close_appeal_tabs(driver, base_handles, name, site):
    """关闭本次申诉中新开的标签页，避免长时间循环导致标签堆积。"""
    try:
        current_handles = driver.window_handles
    except Exception as e:
        print(f"{get_now_time()} {name}{site} 获取标签页失败：{e}<br>")
        return

    base_handles = set(base_handles or [])
    new_handles = [handle for handle in current_handles if handle not in base_handles]
    for handle in new_handles:
        try:
            driver.switch_to.window(handle)
            driver.close()
            print(f"{get_now_time()} {name}{site} 已关闭本次申诉标签页<br>")
        except Exception as e:
            print(f"{get_now_time()} {name}{site} 关闭申诉标签页失败：{e}<br>")

    try:
        remaining_handles = driver.window_handles
        for handle in current_handles:
            if handle in base_handles and handle in remaining_handles:
                driver.switch_to.window(handle)
                return
        if remaining_handles:
            driver.switch_to.window(remaining_handles[-1])
    except Exception:
        pass


def open_human_service_chat(driver, name, site, window_id=""):
    """当前标签优先打开指定 hub；进入对话窗失败时再回退 chat/v2。"""
    entry_xpath = "//*[self::button or self::a][contains(., 'We’ll send you a message in less than 5 min') or contains(., \"We'll send you a message in less than 5 min\")]"

    try:
        current_url = driver.current_url or ""
    except Exception:
        current_url = ""
    if "/help/hub/30928" not in current_url or is_mercado_rate_limited_page(driver):
        open_human_service_hub_with_ip_retry(
            driver,
            name,
            site,
            window_id=window_id,
        )

    if wait_for_human_chat_input(driver, timeout=2) or switch_to_latest_chat_input_tab(driver, timeout=1):
        print(f"{get_now_time()} {name} {site} hub 页面人工客服输入框已就绪<br>")
        return True

    try:
        element = WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.XPATH, entry_xpath)))
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
        driver.execute_script("arguments[0].click();", element)
        print(f"{get_now_time()} {name} {site} hub 页面点击人工客服 chat 入口成功<br>")
        return wait_after_human_entry_click(driver, name, site)
    except Exception as e:
        print(f"{get_now_time()} {name} {site} hub 进入对话窗失败，回退 chat/v2：{e}<br>")

    chat_result = open_mercado_backend_page(
        driver,
        HUMAN_SERVICE_CHAT_V2_URL,
        name,
        window_id,
        settle_seconds=0,
        rate_limit_retry_wait_seconds=0,
        navigate=lambda url: fast_navigate(driver, url, stop_after=3),
        anomaly_site=site,
        anomaly_source="人工申诉",
    )
    if not chat_result.get("ok"):
        raise RuntimeError(chat_result.get("message") or "chat/v2 页面打开失败")
    if wait_for_human_chat_input(driver, timeout=2) or switch_to_latest_chat_input_tab(driver, timeout=1):
        print(f"{get_now_time()} {name} {site} chat/v2 聊天页已就绪<br>")
        return True
    save_human_service_debug_artifacts(driver, name, site)
    raise RuntimeError("进入人工客服失败：hub 对话窗和 chat/v2 均没有找到输入框")


def wait_after_human_entry_click(driver, name, site):
    """点击人工客服入口后快速确认聊天输入框，避免继续慢速遍历其他候选入口。"""
    time.sleep(1)
    try:
        driver.switch_to.window(driver.window_handles[-1])
    except Exception:
        pass

    if wait_for_human_chat_input(driver, timeout=2) or switch_to_latest_chat_input_tab(driver, timeout=1):
        print(f"{get_now_time()} {name} {site} 人工客服输入框已出现<br>")
        return True

    raise RuntimeError("点击 hub 人工客服入口后没有出现聊天输入框")


def get_human_chat_input(driver, timeout=30):
    """查找人工客服聊天输入框。"""
    input_box = find_human_chat_input_in_frames(driver, timeout=timeout)
    if input_box is not None:
        return input_box
    raise RuntimeError("没有找到人工客服输入框")


def wait_for_human_chat_input(driver, timeout=20):
    return find_human_chat_input_in_frames(driver, timeout=timeout) is not None


def switch_to_latest_chat_input_tab(driver, timeout=1):
    """从右往左查找已经打开的聊天页输入框，找到后停留在该标签/iframe。"""
    try:
        handles = driver.window_handles[:]
    except Exception:
        return False

    for handle in reversed(handles):
        try:
            driver.switch_to.window(handle)
            if find_human_chat_input_in_frames(driver, timeout=timeout) is not None:
                return True
        except Exception:
            continue
    return False


def find_human_chat_input_in_frames(driver, timeout=30):
    """在默认页面和 iframe 中查找人工客服输入框，并停留在找到输入框的 frame。"""
    selectors = [
        (By.XPATH, "//div[@aria-placeholder='Write your question or problem']"),
        (By.XPATH, "//div[@contenteditable='true' and contains(@aria-placeholder, 'Write')]"),
        (By.XPATH, "//textarea[contains(@placeholder, 'question') or contains(@aria-label, 'question') or contains(@placeholder, 'message') or contains(@aria-label, 'message')]"),
        (By.XPATH, "//input[contains(@placeholder, 'question') or contains(@aria-label, 'question') or contains(@placeholder, 'message') or contains(@aria-label, 'message')]"),
        (By.XPATH, "//*[@contenteditable='true' and (@role='textbox' or @aria-multiline='true')]"),
        (By.CSS_SELECTOR, "div[contenteditable='true']"),
        (By.CSS_SELECTOR, "textarea"),
        (By.CSS_SELECTOR, "input[type='text']"),
        (By.CSS_SELECTOR, "[role='textbox']"),
    ]
    end_time = time.time() + timeout
    while time.time() < end_time:
        driver.switch_to.default_content()
        found = _find_human_input_current_context(driver, selectors)
        if found is not None:
            return found

        frames = driver.find_elements(By.TAG_NAME, "iframe")
        for frame in frames:
            try:
                driver.switch_to.default_content()
                driver.switch_to.frame(frame)
                found = _find_human_input_current_context(driver, selectors)
                if found is not None:
                    return found
            except Exception:
                continue
        time.sleep(0.5)
    driver.switch_to.default_content()
    return None


def _find_human_input_current_context(driver, selectors):
    try:
        element = driver.execute_script(
            """
            function visible(el) {
                const rect = el.getBoundingClientRect();
                const style = window.getComputedStyle(el);
                return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
            }
            function textOf(el) {
                return [
                    el.getAttribute('placeholder') || '',
                    el.getAttribute('aria-placeholder') || '',
                    el.getAttribute('aria-label') || '',
                    el.getAttribute('role') || '',
                    el.className || ''
                ].join(' ').toLowerCase();
            }
            const candidates = [...document.querySelectorAll('textarea,input,[contenteditable="true"],[role="textbox"]')]
                .filter(el => visible(el) && !el.disabled && !el.readOnly)
                .filter(el => {
                    const text = textOf(el);
                    return el.isContentEditable || /textbox|question|message|write|problem|chat/.test(text);
                });
            candidates.sort((a, b) => {
                const score = el => {
                    const text = textOf(el);
                    let value = 0;
                    if (el.isContentEditable) value += 30;
                    if (el.tagName === 'TEXTAREA') value += 20;
                    if (/write|question|message|problem/.test(text)) value += 50;
                    return value;
                };
                return score(b) - score(a);
            });
            return candidates[0] || null;
            """
        )
        if element is not None and element.is_displayed():
            return element
    except Exception:
        pass

    for by, selector in selectors:
        try:
            elements = driver.find_elements(by, selector)
            for element in elements:
                if element.is_displayed() and element.is_enabled():
                    return element
        except Exception:
            continue
    return None


def save_human_service_debug_artifacts(driver, name, site):
    """保存进入人工客服失败时的页面信息，便于定位实际入口。"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = re.sub(r'[\\/:*?"<>|]', "_", str(name))
    safe_site = re.sub(r'[\\/:*?"<>|]', "_", str(site))
    prefix = get_bit_path() / f"human_service_failed_{safe_name}_{safe_site}_{timestamp}"
    try:
        driver.save_screenshot(str(prefix) + ".png")
    except Exception:
        pass
    try:
        Path(str(prefix) + ".html").write_text(driver.page_source, encoding="utf-8", errors="ignore")
    except Exception:
        pass
    try:
        buttons = driver.execute_script(
            """
            return [...document.querySelectorAll('button,a,[role="button"]')]
                .map((el, index) => ({
                    index,
                    tag: el.tagName,
                    text: (el.innerText || el.textContent || '').trim().slice(0, 160),
                    aria: el.getAttribute('aria-label') || '',
                    title: el.getAttribute('title') || '',
                    href: el.getAttribute('href') || ''
                }))
                .filter(item => item.text || item.aria || item.title || item.href)
                .slice(0, 120);
            """
        )
        Path(str(prefix) + "_elements.json").write_text(
            json.dumps(buttons, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except Exception:
        pass
    print(f"{get_now_time()} {name} {site} 已保存人工客服入口失败调试文件：{prefix}.*<br>")


def click_human_send_button(driver, timeout=30):
    """点击人工客服发送按钮，兼容 title/aria-label/文本变化。"""
    selectors = [
        (By.CSS_SELECTOR, 'button[title="Send"]'),
        (By.CSS_SELECTOR, 'button[aria-label="Send"]'),
        (By.XPATH, "//*[self::button or @role='button'][contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'send')]"),
        (By.XPATH, "//*[self::button or @role='button'][contains(translate(@aria-label, 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'send')]"),
    ]
    end_time = time.time() + timeout
    while time.time() < end_time:
        for by, selector in selectors:
            try:
                elements = driver.find_elements(by, selector)
                for element in elements:
                    if element.is_displayed() and element.is_enabled():
                        try:
                            element.click()
                        except Exception:
                            driver.execute_script("arguments[0].click();", element)
                        return True
            except Exception:
                continue
        try:
            clicked = driver.execute_script(
                """
                const candidates = [...document.querySelectorAll('button,[role="button"]')].filter(el => {
                    const text = [
                        el.innerText || '',
                        el.textContent || '',
                        el.getAttribute('aria-label') || '',
                        el.getAttribute('title') || ''
                    ].join(' ');
                    const rect = el.getBoundingClientRect();
                    return /send/i.test(text) && rect.width > 0 && rect.height > 0;
                });
                if (!candidates.length) return false;
                candidates[0].click();
                return true;
                """
            )
            if clicked:
                return True
        except Exception:
            pass
        time.sleep(0.5)
    raise RuntimeError("没有找到人工客服发送按钮")


def use_one_browser_run_task(info):
    # /browser/open 接口会返回 selenium使用的http地址，以及webdriver的path，直接使用即可
    name = info[0]
    site = info[1]
    form = info[2]
    message = info[3]

    try:
        ip_usable = True
        if ip_usable:
            while True:
                print("ip检测通过，打开店铺平台主页")

                try:
                    shensu(name, site, form, message,"人工客服")
                except Exception as e:
                    traceback.print_exc()
                    print("申诉执行异常", e)
                finally:
                    window_id = getWindowidByName(name)
                    try:
                        closeBrowser(window_id)
                    except Exception as e:
                        continue
                    time.sleep(1800)

        else:
            print("ip检测不通过，请检查")
    except:
        print("脚本运行异常:" + traceback.format_exc())


def should_load_infraction_orders(form, message):
    """仅无自定义话术的侵权申诉才需要遍历侵权列表。"""
    return form == "侵权" and message == ""


# 申诉
def shensu(name, site, form, message, mode="人工客服"):
    print(f"{name} {site} 开始进行{form}申诉，话术为{message}<br>")
    selected_phrase = select_appeal_phrase(form) if message == "" else ""
    if selected_phrase:
        print(f"{get_now_time()} {name} {site} 从{form}话术库随机选取：{selected_phrase}<br>")
    window_id = get_window_id_by_shop_name(
        name,
        authorization_flag="appeal_enabled",
    )

    res = openBrowser(window_id)  # 窗口ID从窗口配置界面中复制，或者api创建后返回

    print(res)
    name = res["data"]["name"]

    driverPath = res["data"]["driver"]
    debuggerAddress = res["data"]["http"]

    # selenium 连接代码
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_experimental_option("debuggerAddress", debuggerAddress)

    chrome_service = Service(driverPath)
    driver = webdriver.Chrome(service=chrome_service, options=chrome_options)

    driver.implicitly_wait(10)
    try:
        driver.set_page_load_timeout(12)
    except Exception:
        pass
    # 设置最长等待时间为 10 秒
    wait = WebDriverWait(driver, 15)
    appeal_base_handles = []
    try:
        appeal_base_handles = driver.window_handles[:]
    except Exception:
        pass

    # driver.switch_to.new_window('tab') 决定是否打开新窗口
    try:
        open_human_service_hub_with_ip_retry(
            driver,
            name,
            site,
            window_id=window_id,
        )
    except Exception as exc:
        print(f"{get_now_time()} {name} {site} {exc}<br>")
        return str(exc)

    words = []
    nickname_list = ["Bruce", "Jack", "Lucy", "James"]
    nickname = random.choice(nickname_list)
    if form == "延误":
        words = [
            f"亲爱的客服，我叫{nickname}！这些订单因合作物流车辆临时出现故障，导致未能及时揽收，并非我这边发货延误，麻烦您帮忙处理一下，消除对店铺声誉的影响，非常感谢！",
            f"亲爱的客服，我叫{nickname}！这些订单因为菜鸟，并非我这边发货延误，麻烦您帮忙处理一下，消除对店铺声誉的影响，非常感谢！",
        ]

    if form == "侵权":
        words = [
            f"亲爱的客服，我叫{nickname}！这些产品是通用品牌产品，他们被系统误检测为侵权产品，你能帮我消除记录吗？",
            f"亲爱的客服，我叫{nickname}！这些产品是通用品牌产品，他们被系统误检测为侵权产品，你能帮我消除记录吗？",
        ]

    if form == "取消率":
        words = [
            f"亲爱的客服，我叫{nickname}！这些订单并非因卖家责任取消，麻烦您重新核查订单记录，并移除这些订单对店铺取消率和声誉的影响，非常感谢！",
            f"亲爱的客服，我叫{nickname}！这些订单的取消不应计入卖家责任，麻烦您帮我复核并消除对店铺取消率的影响，谢谢！",
        ]

    if form == "投诉":
        words = [
            f"亲爱的客服，我叫{nickname}！我的产品没有任何质量问题，客户没有给出确凿的证据证明他出了问题，我认为客户是想免费购物，你能消除对我声誉的影响吗"
        ]

    words_random = random.choice(words)

    if not select_mercado_site_fast(driver, name, site):
        try:
            WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.CLASS_NAME, "nav-header-cbt__site-switcher"))
            ).click()
            path = f'div[data-value="{SITE_REMOTE_VALUE_MAP.get(site, "MLM-remote")}"]'
            WebDriverWait(driver, 8).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, path))
            ).click()
            driver.refresh()
            time.sleep(2)
            print(f"{get_now_time()} {name} {site} '选择站点成功'<br>")
        except Exception as e:
            print(f"{get_now_time()} {name} {site} '选择站点失败，继续使用当前页面': {e}<br>")
    orders_random=""
    infraction_random=""
    cancellation_random=""
    initial_huashu = ""

    if(form=="延误" and message == ""):
        orders_random = get_delay_orders_for_human_service(window_id, name, site, 10)
        if (orders_random == "" and message == ""):
            close_appeal_tabs(driver, appeal_base_handles, name, site)
            return "没有可以申诉的订单"
    if should_load_infraction_orders(form, message):
        infraction_random = get_infraction_orders_random(window_id,name, site, 10)
    if form == "取消率" and message == "":
        cancellation_orders = get_cancellation_orders(driver, name, site)
        if not cancellation_orders:
            close_appeal_tabs(driver, appeal_base_handles, name, site)
            return "没有可以申诉的取消订单"
        selected_orders = (
            random.sample(cancellation_orders, 10)
            if len(cancellation_orders) > 10
            else cancellation_orders
        )
        cancellation_random = "、".join(str(order_id) for order_id in selected_orders)
        print(
            f"{get_now_time()} {name} {site} 共获取 {len(cancellation_orders)} 个取消订单，"
            f"本轮人工客服按侵权规则发送 {len(selected_orders)} 个：{cancellation_random}<br>"
        )
        # 获取订单后当前位于 Metrics 页面，重新回到人工客服入口。
        open_human_service_hub_with_ip_retry(
            driver,
            name,
            site,
            window_id=window_id,
        )
        select_mercado_site_fast(driver, name, site)
    try:
        open_human_service_chat(driver, name, site, window_id=window_id)

        # 发消息
        print(f"{get_now_time()} {name} {site} '进入人工客服'<br>")

        if message == "":

            if form == "延误":
                initial_huashu = (
                    render_appeal_phrase(
                        selected_phrase,
                        nickname=nickname,
                        order_ids=orders_random,
                        appeal_type=form,
                    )
                    if selected_phrase
                    else orders_random + words_random
                )
                get_human_chat_input(driver, 30).send_keys(initial_huashu)
                time.sleep(3)
                click_human_send_button(driver, 30)
                print(
                    f"{get_now_time()} {name}  {site} 发送延误订单：{initial_huashu}<br>"
                )
                chat_ai(
                    driver, name, site, form, initial_huashu, nickname
                )
            if form == "侵权":
                initial_huashu = (
                    render_appeal_phrase(
                        selected_phrase,
                        nickname=nickname,
                        order_ids=infraction_random,
                        appeal_type=form,
                    )
                    if selected_phrase
                    else infraction_random + words_random
                )
                get_human_chat_input(driver, 30).send_keys(initial_huashu)
                time.sleep(3)
                click_human_send_button(driver, 30)
                print(
                    f"{get_now_time()} {name} {site} '发送侵权的 id：{initial_huashu}<br>"
                )
                chat_ai(
                    driver, name, site, form, initial_huashu, nickname
                )
            if form == "取消率":
                initial_huashu = (
                    render_appeal_phrase(
                        selected_phrase,
                        nickname=nickname,
                        order_ids=cancellation_random,
                        appeal_type=form,
                    )
                    if selected_phrase
                    else cancellation_random + words_random
                )
                get_human_chat_input(driver, 30).send_keys(initial_huashu)
                time.sleep(3)
                click_human_send_button(driver, 30)
                print(
                    f"{get_now_time()} {name} {site} 发送取消订单：{initial_huashu}<br>"
                )
                chat_ai(
                    driver, name, site, form, initial_huashu, nickname
                )
            if form == "投诉":
                initial_huashu = (
                    render_appeal_phrase(
                        selected_phrase or words_random,
                        nickname=nickname,
                        appeal_type=form,
                    )
                )
                get_human_chat_input(driver, 30).send_keys(initial_huashu)
                time.sleep(3)
                click_human_send_button(driver, 30)
                print(
                    f"{get_now_time()} {name} {site} 发送投诉申诉：{initial_huashu}<br>"
                )
                chat_ai(
                    driver, name, site, form, initial_huashu, nickname
                )
        else:
            initial_huashu = message
            get_human_chat_input(driver, 30).send_keys(message)
            time.sleep(3)
            click_human_send_button(driver, 30)
            print(f"{get_now_time()} {name} {site} 自动发送自定义话术：{message}<br>")
            chat_ai(
                driver, name, site, form, initial_huashu, nickname
            )

    except Exception as e:
        chat_result = open_mercado_backend_page(
            driver,
            HUMAN_SERVICE_CHAT_V2_URL,
            name,
            window_id,
            settle_seconds=0,
            rate_limit_retry_wait_seconds=0,
            navigate=lambda url: fast_navigate(driver, url, stop_after=3),
            anomaly_site=site,
            anomaly_source="人工申诉",
        )
        if not chat_result.get("ok"):
            raise RuntimeError(chat_result.get("message") or str(e)) from e
        print(get_now_time() + name + site + "继续与客服对话")
        # 全部聊天记录
        chat_ai(driver, name, site, form, initial_huashu, nickname)
    finally:
        print(f"{get_now_time()} {name}{site}找客服执行完毕<br>")
        close_appeal_tabs(driver, appeal_base_handles, name, site)
        print(f"{get_now_time()} {name}{site} 关闭浏览器<br>")


def get_delay_orders_random(name, site, nums):
    order_random = ""
    try:
        delay_folder_path = get_bit_path() / "美客多延误"
        delay_file = get_latest_modified_file(delay_folder_path)
        delay_file_path = delay_folder_path / delay_file
        fifteen_days_ago = datetime.now() - timedelta(days=15)
        order_list = []
        df = pd.read_excel(delay_file_path, engine='openpyxl')
        for index, row in df.iterrows():
            # print(row)
            line_name = row['店铺']
            line_site = row['站点']
            order_date = row['下单时间']
            order_num = row['销售单号']
            dispatch_date = row['实际揽收时间']
            if (line_name == name and line_site == site and dispatch_date != "Not yet dispatched"):
                order_date = parser_delay_date(order_date)
                if (order_date > fifteen_days_ago):
                    order_list.append(order_num)
        print(get_now_time() + name + site + "最近15天的延误个数:", len(order_list))

        if len(order_list) >= nums:
            order_random = str(random.sample(order_list, nums))
        else:
            order_random = str(order_list)
        order_random = re.sub(r'[^\d,]', '', order_random)

        print(get_now_time() + name + site + "随机得到的延误销售单号为", order_random)
    except Exception as e:
        print("获取延误表格信息失败",e)
    return order_random


def get_delay_orders_for_human_service(window_id, name, site, nums=10):
    """人工客服延误申诉：自动下载最新延误表，并一次最多返回 nums 个订单号。"""
    try:
        from bit.bit_appeal_ai import get_delay_orders_download_list

        order_list = get_delay_orders_download_list(window_id, name, site)
        if not order_list:
            return ""
        selected_orders = order_list[:nums]
        order_text = "、".join(str(order).strip().lstrip("'") for order in selected_orders if str(order).strip())
        print(get_now_time() + name + site + f"人工客服延误本次发送 {len(selected_orders)} 个订单:", order_text)
        return order_text
    except Exception as e:
        print("人工客服自动下载并读取延误订单失败", e)
        traceback.print_exc()
        return ""


def get_infraction_orders_random(window_id,name, site, nums):
    """人工客服旧入口也统一从官方 API 读取，网页采集器不再参与。"""

    try:
        from bit.bit_appeal_ai import get_infraction_orders as get_api_infraction_orders

        inf_list = get_api_infraction_orders(window_id, name, site)
        selected = (
            random.sample(inf_list, nums)
            if len(inf_list) > nums
            else inf_list
        )
        result = "、".join(str(item) for item in selected)
        print(
            get_now_time() + name + site
            + f"从官方 API 随机得到 {len(selected)} 个侵权编号：",
            result,
        )
        return result
    except Exception as e:
        print("通过官方 API 获取侵权订单信息失败",e)
        return ""


# 检查聊天是否结束
def checkChatEnd(driver, name, site):
    try:
        elements = driver.find_elements(By.XPATH, "//p[contains(text(), 'This chat has ended')]")
        if any(element.is_displayed() for element in elements):
            print(f"{get_now_time()} {name}{site}聊天已经结束,结束AI找客服<br>")
            return True
    except Exception as e:
        return False
    return False


def get_human_chat_context(driver):
    """读取当前聊天上下文，尽量包含客服和卖家的完整消息。"""
    try:
        messages = driver.execute_script(
            """
            const nodes = [...document.querySelectorAll('[class*="chat-ui-message-bubble"], [role="log"] [role="listitem"], [data-testid*="message"]')];
            const lines = [];
            const seen = new Set();
            for (const node of nodes) {
                const text = (node.innerText || node.textContent || '').trim();
                if (!text || seen.has(text)) continue;
                seen.add(text);
                const cls = String(node.className || '');
                let role = '对话';
                if (/from-agent/i.test(cls)) role = '客服';
                if (/from-user|from-client|from-customer/i.test(cls)) role = '我';
                lines.push(`${role}: ${text}`);
            }
            return lines;
            """
        )
        if messages:
            return "\n".join(messages)
    except Exception:
        pass

    lines = []
    try:
        elements = driver.find_elements(By.CSS_SELECTOR, '[class*="chat-ui-message-bubble"]')
        for element in elements:
            text = element.text.strip()
            if text and text not in lines:
                lines.append(text)
    except Exception:
        pass
    return "\n".join(lines)


def get_deepseek_human_service_reply(context, form, nickname):
    task_map = {
        "延误": "我正在申诉延误订单，希望客服消除对店铺声誉的影响。",
        "侵权": "我正在申诉侵权记录，希望客服认可这是通用品牌产品并消除记录。",
        "取消率": "我正在申诉取消订单，希望客服复核并消除这些订单对店铺取消率和声誉的影响。",
        "投诉": "我正在申诉投诉订单，希望客服消除对店铺声誉的影响。",
    }
    messages = [
        {
            "role": "system",
            "content": (
                "你正在帮助卖家和 Mercado Libre 人工客服对话。"
                "每次必须根据完整上下文判断下一句回复。"
                "如果客服已经明确拒绝继续处理、明确表示无法移除/无法申诉/最终决定不变，rejected 返回 true。"
                "如果只是询问信息、要求订单号、解释流程或还没明确拒绝，rejected 返回 false。"
                "只返回 JSON，不要返回 Markdown。格式：{\"reply\":\"不超过40个中文字的自然回复\",\"rejected\":false,\"reason\":\"简短判断依据\"}"
            ),
        },
        {
            "role": "user",
            "content": (
                f"我的英文名/称呼是 {nickname}。\n"
                f"当前业务：{task_map.get(form, form)}\n"
                "如果 rejected=true，reply 必须是在礼貌接受结果后，请客服帮忙关闭聊天窗口。\n"
                "如果 rejected=false，reply 需要继续推进申诉，不要重复已经发过的订单号。\n\n"
                f"完整聊天上下文：\n{context}"
            ),
        },
    ]
    raw = chat_deepseek(messages, temperature=0.2, max_tokens=300)
    try:
        data = json.loads(raw.strip())
    except Exception:
        match = re.search(r"\{.*\}", raw, re.S)
        data = json.loads(match.group(0)) if match else {"reply": raw, "rejected": False, "reason": "非 JSON 返回"}

    reply = str(data.get("reply", "")).strip()
    rejected = bool(data.get("rejected", False))
    reason = str(data.get("reason", "")).strip()
    if rejected:
        reply = "好的，我明白了。麻烦您帮我关闭聊天窗口，谢谢。"
    if not reply:
        reply = "麻烦您再帮我确认一下，谢谢。"
    return reply, rejected, reason, raw


def chat_ai(driver, name, site, form, huashu, nickname):
    last_context = ""
    idle_times = 0
    chat_logs = []
    i = 0
    while True:
        i += 1
        context = ""
        response = ""
        should_stop = False
        isEnd = checkChatEnd(driver, name, site)
        if isEnd:
            break

        try:
            print(f"{get_now_time()} {name}{site} 进入人工客服处理流程，第{i}轮读取上下文<br>")
            context = get_human_chat_context(driver)
            print(f"{get_now_time()} {name}{site} 聊天上下文：<br>{context}<br>")
            if not context:
                print(f"{get_now_time()} {name}{site} 暂未读取到聊天内容，继续等待<br>")
                idle_times += 1
                continue

            if context == last_context:
                idle_times += 1
                print(f"{get_now_time()} {name}{site} 客服暂无新回复，第{idle_times}次等待<br>")
                continue

            idle_times = 0
            last_context = context
            response, rejected, reason, raw_response = get_deepseek_human_service_reply(context, form, nickname)
            chat_logs.append({
                "round": i,
                "context": context,
                "response": response,
                "rejected": rejected,
                "reason": reason,
                "raw_response": raw_response,
            })
            print(f"{get_now_time()} {name}{site} DeepSeek判断：rejected={rejected}, reason={reason}<br>")
            print(f"{get_now_time()} {name}{site} DeepSeek回复：{response}<br>")
            try:
                # 发消息
                get_human_chat_input(driver, 30).send_keys(response)
                time.sleep(3)
                click_human_send_button(driver, 30)
                print(f"{get_now_time()} {name}{site}自动发送消息:{response}<br>")
                # 聊天记录插入数据库
                result = insert_chat_info_by_api(name, site, huashu, context, response, get_now_time())
                print(f"{get_now_time()} {name}{site}聊天记录接口入库成功:{result}<br>")

            except Exception as e:
                print(f"{get_now_time()} {name}{site}发送消息失败<br>")
                print(e)
                traceback.print_exc()

            if rejected:
                print(f"{get_now_time()} {name}{site}客服已明确拒绝，已请求客服关闭聊天窗口<br>")
                should_stop = True
                break
        except Exception as e:
            print(e)
            traceback.print_exc()
        finally:
            if not should_stop:
                print(f"{get_now_time()} {name}{site}等待客服回复，120秒后继续轮询<br>")
                time.sleep(120)

    print(f"{get_now_time()} {name}{site}聊天日志：{json.dumps(chat_logs, ensure_ascii=False)}<br>")
    print(f"{get_now_time()} {name}{site}结束AI客服回复<br>")


def chat_script(driver):
    return None


def use_all_browser_run_task(browser_list):
    """
    循环打开所有店铺运行脚本
    :param browser_list: 店铺列表
    """
    for browser in browser_list:
        use_one_browser_run_task(browser)


def use_all_browser_run_task_with_thread_pool(browser_list, max_threads=10):
    """
    使用线程池控制最大并发线程数
    :param browser_list: 店铺列表
    :param max_threads: 最大并发线程数
    """
    with ThreadPoolExecutor(max_workers=max_threads) as executor:
        executor.map(use_one_browser_run_task, browser_list)


def auto_appeal_delay():
    fold_path = get_bit_path() / "美客多延误"
    file_path = fold_path / get_latest_modified_file(fold_path)
    wb = load_workbook(file_path)
    sheet = wb.active
    # 使用 min_row=2 跳过第一行

    name_site = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        delayrate = row[2]
        if delayrate != None and delayrate != "":
            delay_value = 0.0
            if type(delayrate) == str:
                delay_value = float(delayrate.strip("%")) / 100
            else:
                delay_value = float(delayrate)
            if delay_value >= 0.07:
                name_site.add((row[0], row[1], delay_value))

    print(len(name_site))

    list_appeal = []
    for i in name_site:
        list_appeal.append((i[0], i[1], "延误", ""))

    print(list_appeal)

    use_all_browser_run_task_with_thread_pool(list_appeal, 5)


def main():
    # 在这里指定要运行的店铺参数，不使用命令行传参。
    shop_name = "跃马扬鞭"
    site = "墨西哥"
    appeal_type = "侵权"
    message = ""

    use_one_browser_run_task((shop_name, site, appeal_type, message))



if __name__ == "__main__":
    main()
