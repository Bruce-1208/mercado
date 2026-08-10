"""从订单 Excel 文件批量更新数据库。"""

from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook

from bit.bit_mysql import insert_orders


EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
ORDER_FIELDS = (
    ("id", ("id", "订单id", "订单ID")),
    ("编号", ("编号", "订单编号", "销售单号")),
    ("时间", ("时间", "订单时间")),
    ("业务员", ("业务员",)),
    ("来源", ("来源",)),
    ("状态", ("状态",)),
    ("金额", ("金额",)),
    ("费用", ("费用",)),
    ("退款", ("退款",)),
    ("人民币收入", ("人民币收入",)),
    ("采购成本", ("采购成本",)),
    ("采购单号", ("采购单号",)),
    ("采购追踪", ("采购追踪",)),
    ("利润", ("利润",)),
    ("产品id", ("产品id", "产品ID")),
    ("产品分类", ("产品分类",)),
    ("标题", ("标题",)),
    ("图片", ("图片",)),
    ("数量", ("数量",)),
    ("订单运费", ("订单运费",)),
    ("订单备注", ("订单备注",)),
    ("地区", ("地区",)),
    ("买家姓名", ("买家名称", "买家姓名")),
)


def _normalize_header(value):
    return str(value or "").strip()


def _normalize_order_id(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def is_order_excel_file(name):
    path = Path(str(name or ""))
    return (
        path.suffix.casefold() in EXCEL_SUFFIXES
        and not path.name.startswith(("~$", ".~", "."))
    )


def discover_order_excel_files(folder=None):
    """递归查找文件夹内的订单工作簿，并按修改时间从旧到新排序。"""

    root = Path(folder or (Path(__file__).resolve().parent / "美客多订单")).expanduser()
    if not root.is_dir():
        raise ValueError(f"订单文件夹不存在：{root}")
    files = [
        path
        for path in root.rglob("*")
        if path.is_file() and is_order_excel_file(path.name)
    ]
    return sorted(
        files,
        key=lambda path: (path.stat().st_mtime_ns, str(path).casefold()),
    )


def _source_parts(source):
    if isinstance(source, (str, Path)):
        path = Path(source)
        return str(path), path
    if isinstance(source, tuple) and len(source) == 2:
        name, stream = source
        return str(name or getattr(stream, "name", "订单工作簿")), stream
    return str(getattr(source, "name", "订单工作簿")), source


def read_order_workbook(source):
    """读取单个工作簿的所有有效工作表，返回标准数据库行及统计。"""

    source_name, workbook_source = _source_parts(source)
    if not is_order_excel_file(source_name):
        raise ValueError(f"不支持的 Excel 文件：{source_name}")
    if hasattr(workbook_source, "seek"):
        workbook_source.seek(0)

    workbook = load_workbook(
        workbook_source,
        read_only=True,
        data_only=True,
        keep_vba=Path(source_name).suffix.casefold() == ".xlsm",
    )
    rows = []
    blank_order_ids = 0
    parsed_sheets = 0
    try:
        for sheet in workbook.worksheets:
            iterator = sheet.iter_rows(values_only=True)
            headers = next(iterator, None)
            if not headers:
                continue
            header_index = {
                _normalize_header(header): index
                for index, header in enumerate(headers)
                if _normalize_header(header)
            }
            id_aliases = ORDER_FIELDS[0][1]
            if not any(alias in header_index for alias in id_aliases):
                continue
            parsed_sheets += 1

            resolved_indexes = []
            for _database_name, aliases in ORDER_FIELDS:
                resolved_indexes.append(
                    next(
                        (header_index[alias] for alias in aliases if alias in header_index),
                        None,
                    )
                )

            for values in iterator:
                if not values or not any(value not in (None, "") for value in values):
                    continue
                order_id_index = resolved_indexes[0]
                order_id = _normalize_order_id(
                    values[order_id_index]
                    if order_id_index is not None and order_id_index < len(values)
                    else None
                )
                if not order_id:
                    blank_order_ids += 1
                    continue
                row = []
                for index in resolved_indexes:
                    row.append(values[index] if index is not None and index < len(values) else None)
                row[0] = order_id
                rows.append(row)
    finally:
        workbook.close()

    if parsed_sheets == 0:
        raise ValueError(f"没有找到包含 id 列的订单工作表：{source_name}")
    return rows, {
        "file": source_name,
        "sheets": parsed_sheets,
        "rows": len(rows),
        "blank_order_ids": blank_order_ids,
    }


def collect_latest_order_rows(sources):
    """合并多个工作簿；同一订单后读取到的记录覆盖先读取到的记录。"""

    latest_rows = OrderedDict()
    reports = []
    errors = []
    raw_rows = 0
    blank_order_ids = 0
    duplicate_rows = 0

    for source in sources:
        source_name, _stream = _source_parts(source)
        try:
            rows, report = read_order_workbook(source)
        except Exception as exc:
            errors.append({"file": source_name, "error": str(exc)})
            continue
        reports.append(report)
        raw_rows += len(rows)
        blank_order_ids += report["blank_order_ids"]
        for row in rows:
            order_id = row[0]
            if order_id in latest_rows:
                duplicate_rows += 1
            latest_rows[order_id] = row

    return list(latest_rows.values()), {
        "files_processed": len(reports),
        "raw_rows": raw_rows,
        "unique_orders": len(latest_rows),
        "duplicate_rows": duplicate_rows,
        "blank_order_ids": blank_order_ids,
        "files": reports,
        "errors": errors,
    }


def update_order_sources(sources, insert_func=insert_orders):
    """读取并 UPSERT 一组有顺序的 Excel 来源，返回导入统计。"""

    sources = list(sources or ())
    if not sources:
        raise ValueError("没有找到可导入的 Excel 文件")
    rows, result = collect_latest_order_rows(sources)
    result["files_discovered"] = len(sources)
    if not rows:
        detail = result["errors"][0]["error"] if result["errors"] else "没有有效订单"
        raise ValueError(detail)
    insert_func(rows)
    result["imported_orders"] = len(rows)
    return result


def update_order_mysql(folder=None, insert_func=insert_orders):
    """兼容旧入口：导入指定文件夹；未指定时仍使用 ``美客多订单``。"""

    files = discover_order_excel_files(folder)
    return update_order_sources(files, insert_func=insert_func)


if __name__ == "__main__":
    print(update_order_mysql())
