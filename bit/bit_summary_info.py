import time

from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.wait import WebDriverWait

from bit.utils import get_now_time
from bit_api import *
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import pyautogui
from switch_country import *
from openpyxl import load_workbook
from send_mail import *
import pandas as pd

from datetime import datetime
from pathlib import Path


def get_reputation_info(window_id, site):
    res = openBrowser(window_id)  # 窗口ID从窗口配置界面中复制，或者api创建后返回

    print(res)

    driverPath = res['data']['driver']
    debuggerAddress = res['data']['http']

    # selenium 连接代码
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_experimental_option("debuggerAddress", debuggerAddress)

    chrome_service = Service(driverPath)
    driver = webdriver.Chrome(service=chrome_service, options=chrome_options)

    driver.implicitly_wait(10)
    # 设置最长等待时间为 10 秒
    wait = WebDriverWait(driver, 10)

    driver.get("https://global-selling.mercadolibre.com/sales-summary")
    driver.refresh()
    time.sleep(10)
    i = 0
    while (i < 3):
        i = i + 1
        try:

            # 打开站点选择器
            oepn_country_switch(driver)
            # 选择站点
            name = ''
            if site == '墨西哥':
                name = 'Mexico'
            if site == '巴西':
                name = 'Brazil'
            if site == '哥伦比亚':
                name = 'Colombia'
            if site == '智利':
                name = 'Chile'
            if site == '阿根廷':
                name = 'Argentina'
            if site == '乌拉圭':
                name = 'Uruguay'
            #
            force_select_country(driver, name)
            print(get_now_time() + name + '成功选择站点:', site)
            break
        except Exception as e:
            print(get_now_time() + name + '选择站点失败:', site)

    data_delay=""
    data_complain=""
    data_cancel=""
    try:

        xpath_selector = "//p[text()='Non-compliant shipments']/ancestor::div[@class='metric']//div[contains(@class, 'metric__title')]"
        data_delay = target_element = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, xpath_selector))
        ).text
        xpath_selector = "//p[text()='Complaints']/ancestor::div[@class='metric']//div[contains(@class, 'metric__title')]"
        data_complain = target_element = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, xpath_selector))
        ).text
        xpath_selector = "//p[text()='Canceled by you']/ancestor::div[@class='metric']//div[contains(@class, 'metric__title')]"
        data_cancel = target_element = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, xpath_selector))
        ).text
    except Exception as e:
        print(get_now_time()+name+site+"获取声誉数据失败")

    print("提取到的延误率为:", data_delay)
    data_color = driver.find_element(By.CLASS_NAME, 'panel-segment__focus-item-title-container').text
    print("账号的声誉为:", data_color)

    data_orders = driver.find_element(By.CLASS_NAME, 'value__sales').text
    print("总单数为：", data_orders)

    list = []
    if (data_color.__contains__("Green")):
        data_color = '绿色'
    if (data_color.__contains__("Yellow")):
        data_color = '黄色'
    if (data_color.__contains__("Orange")):
        data_color = '橘色'
    if (data_color.__contains__("Red")):
        data_color = '红色'
    if (data_color.__contains__("You still have no color")):
        data_color = '无色'

    list.append(data_color)
    list.append(data_orders)
    list.append(data_complain)
    list.append(data_delay)

    driver.get("")

    return list


def get_reputation_info_all():
    start = int(time.time())
    print(start)
    root_path = Path(__file__).resolve().parent
    file_path = root_path / "比特配置文件.xlsx"

    wb = load_workbook(file_path)
    sheet = wb.active
    reputation_info_sum = []
    reuslt = []
    # 使用 min_row=2 跳过第一行
    for row in sheet.iter_rows(min_row=2, values_only=True):
        print(row)  # row 是一个元组，包含该行所有数据
        id = row[0]
        name = row[1]
        remark = row[2]
        if remark == '忽略':
            continue
        print(get_now_time() + "开始打开窗口:" + name)
        site_list = row[3].split("，")
        for site in site_list:
            i = 0
            while (i < 3):
                i = i + 1
                try:
                    reputation_info = get_reputation_info(id, site)
                    reputation_info.append(name)
                    reputation_info.append(site)
                    reputation_info_sum.append(reputation_info)
                    print(get_now_time() + name + site + "成功")
                    reuslt.append(name + site + "获取声誉信息执行成功")
                    break
                except Exception as e:
                    print(get_now_time() + name + site + "执行失败", e)
                    reuslt.append(name + site + "获取声誉信息执行失败")
                    time.sleep(180)

            time.sleep(10)
        print(get_now_time() + "结束，正在关闭窗口")

        try:
            closeBrowser(id)
        except Exception as e:
            continue
        print(get_now_time() + "已经关闭窗口")
        time.sleep(5)

    result = "\n".join(map(str, reputation_info_sum))
    print(result)

    end = int(time.time())
    print(get_now_time() + "总花费", end - start)
    df = pd.DataFrame(reputation_info_sum, columns=['声誉颜色', '总单量', '投诉率', '延误率', '店铺名', '站点'])

    now = datetime.now()
    date_str = datetime.now().strftime("%Y-%m-%d-%H")

    df.to_excel(root_path / ("美客多声誉/武汉泽顺店铺声誉信息汇总" + date_str + ".xlsx"), index=False)

    send_info('美客多所有店铺声誉汇总', result,
              root_path / ("美客多声誉/武汉泽顺店铺声誉信息汇总" + date_str + ".xlsx"),
              r"武汉泽顺店铺声誉信息汇总" + date_str + ".xlsx")
    print(get_now_time() + "发送邮件成功")


if __name__ == '__main__':
    get_reputation_info_all()
