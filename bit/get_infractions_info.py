import time
from sys import prefix

from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.wait import WebDriverWait

from bit.bit_utils import get_now_time
from bit.bit_api import *
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import pyautogui
from bit.switch_country import *
from openpyxl import load_workbook
from send_mail import *
import pandas as pd

from datetime import datetime
from pathlib import Path
from bit.bit_clash import *
from bit.bit_mysql import *


def get_infractions_info(window_id, name, site):
    res = openBrowser(window_id)  # 窗口ID从窗口配置界面中复制，或者api创建后返回

    print(res)
    driverPath = res['data']['driver']
    debuggerAddress = res['data']['http']

    # selenium 连接代码
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_experimental_option("debuggerAddress", debuggerAddress)

    chrome_service = Service(driverPath)
    driver = webdriver.Chrome(service=chrome_service, options=chrome_options)

    driver.implicitly_wait(10)
    # 设置最长等待时间为 10 秒
    wait = WebDriverWait(driver, 10)

    driver.get("https://global-selling.mercadolibre.com/noindex/pppi/infractions?tab=detections&offset=0")
    time.sleep(10)
    i = 0
    while (i < 3):
        i = i + 1
        try:
            # 打开站点选择器
            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.CLASS_NAME, "nav-header-cbt__site-switcher"))).click()

            print(name + "打开站点选择器")
            time.sleep(5)
            path = 'div[data-value="MLM-remote"]'
            if site == "墨西哥":
                path = 'div[data-value="MLM-remote"]'
            if site == "巴西":
                path = 'div[data-value="MLB-remote"]'
            if site == "哥伦比亚":
                path = 'div[data-value="MCO-remote"]'
            if site == "智利":
                path = 'div[data-value="MLC-remote"]'
            if site == "阿根廷":
                path = 'div[data-value="MLA-remote"]'
            if site == "乌拉圭":
                path = 'div[data-value="MLU-remote"]'

            WebDriverWait(driver, 30).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, path))).click()

            driver.refresh()
            time.sleep(3)
            print(get_now_time() + name + site + "选择站点：", site)
            break
        except Exception as e:
            print(get_now_time() + name + site + "重新执行选择站点")
            switch_random_hongkong_node()
            get_public_ip()
            continue

    # 使用 presence_of_all_elements_located 等待所有匹配的元素出现在 DOM 中
    # 注意：这里传入的是一个元组 (By.CLASS_NAME, "value")
    ids = wait.until(
        EC.presence_of_all_elements_located((By.CLASS_NAME, "infraction-item__id"))
    )
    ids = [el.get_attribute("textContent") for el in ids]
    # infraction-item__title
    titles = wait.until(
        EC.presence_of_all_elements_located((By.CLASS_NAME, "infraction-item__title"))
    )
    titles = [el.get_attribute("textContent") for el in titles]
    # infraction-denounce__date
    dates = wait.until(
        EC.presence_of_all_elements_located((By.CLASS_NAME, "infraction-denounce__date"))
    )
    dates = [el.get_attribute("textContent") for el in dates]
    results = list(zip(ids, titles, dates))
    infractions_list = []
    for row in results:
        new_row = []
        id = row[0]
        title = row[1]
        date = row[2]
        prefix = ""
        if (site == "墨西哥"):
            prefix = "MLM"
        if (site == "巴西"):
            prefix = "MLB"
        if (site == "哥伦比亚"):
            prefix = "MCO"
        if (site == "智利"):
            prefix = "MLC"
        if (site == "阿根廷"):
            prefix = "MLA"
        if (site == "乌拉圭"):
            prefix = "MLU"
        id = id.replace("#",  prefix)
        new_row.append(name)
        new_row.append(site)
        new_row.append(id)
        new_row.append(title)
        new_row.append(date)
        new_row.append(get_now_time())
        infractions_list.append(new_row)
    return infractions_list


def get_infractions_info_all():
    start = int(time.time())
    print(start)
    root_path = Path(__file__).resolve().parent
    # file_path = root_path / "比特配置文件.xlsx"
    file_path = root_path / "比特配置文件.xlsx"


    wb = load_workbook(file_path)
    sheet = wb.active
    infraction_info_sum = []
    result = []
    # 使用 min_row=2 跳过第一行
    for row in sheet.iter_rows(min_row=2, values_only=True):
        print(row)  # row 是一个元组，包含该行所有数据
        id = row[0]
        name = row[1]
        remark = row[2]
        if remark == '忽略':
            continue
        print(get_now_time() + "开始打开窗口:" + name)
        site_list = row[3].split("，")
        for site in site_list:
            i = 0
            while (i < 3):
                i = i + 1
                try:
                    infraction_info = get_infractions_info(id, name, site)
                    infraction_info_sum = infraction_info_sum + infraction_info
                    print(get_now_time() + name + site + "成功")
                    result.append(('获取侵权信息', name, site, "成功", get_now_time()))
                    break
                except Exception as e:
                    print(get_now_time() + name + site + "执行失败", e)
                    if(i==3):
                        result.append(('获取声誉信息', name, site, "失败", get_now_time()))

        print(get_now_time() + "结束，正在关闭窗口")

        try:
            closeBrowser(id)
        except Exception as e:
            continue
        print(get_now_time() + "已经关闭窗口")

    infraction_info_sum_str = "\n".join(map(str, infraction_info_sum))
    print(infraction_info_sum_str)

    end = int(time.time())
    print(get_now_time() + "总花费", end - start)
    df = pd.DataFrame(infraction_info_sum,
                      columns=['店铺名','站点','编号','标题','侵权时间','执行时间'])

    now = datetime.now()
    date_str = datetime.now().strftime("%Y-%m-%d-%H")

    df.to_excel(root_path / ("美客多侵权/武汉泽顺店铺侵权信息汇总" + date_str + ".xlsx"), index=False)

    send_info('美客多所有店铺侵权汇总', infraction_info_sum_str,
              root_path / ("美客多侵权/武汉泽顺店铺侵权信息汇总" + date_str + ".xlsx"),
              r"武汉泽顺店铺侵权信息汇总" + date_str + ".xlsx")
    print(get_now_time() + "发送邮件成功")

    insert_task_record(result)
    inset_infraction_info(infraction_info_sum)



if __name__ == '__main__':
    get_infractions_info_all()
