from openpyxl import load_workbook
from datetime import datetime, timedelta

from pydantic.v1.datetime_parse import parse_date
from datetime import datetime
import pytz
import random
import time

def parser_date(time_str):
    date = time_str.split("-")[0]
    date_format = "%A %d %b %Y "

    return datetime.strptime(date, date_format)








def get_order_number_excel(sheet_name,file_path):
    list_order_number = []
    try:
        # 加载工作簿（Excel文件）
        wb = load_workbook(file_path)

        # 获取活动的工作表（或通过名称获取特定的工作表）
        sheet = wb[sheet_name]  # 或者 wb['Sheet1'] 如果你知道工作表的名称

        fifteen_days_ago = datetime.now() - timedelta(days=30)


        # 读取数据
        for row in sheet.iter_rows(values_only=True):
            order_date = row[0]
            order_number = row[1]
            dispatched_date = row[3]

            try:
                date = parser_date(order_date)
                print(date)
                if dispatched_date != 'Not yet dispatched' and date > fifteen_days_ago:
                    list_order_number.append(order_number)
            except Exception as e:
                print(e)
    except Exception as e :
        print(e,"无法读取"+sheet_name)
    print(sheet_name+"延误订单有", list_order_number)
    return  list_order_number


if __name__ == '__main__':
    print(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
    my_list=get_order_number_excel('龙（张泽文）',r'C:\Users\Admin\PycharmProjects\MercadoApp\订单延误.xlsx')
    if len(my_list) >= 10:
        random_elements = random.sample(my_list, 10)
        print(str(random_elements))
    else:
        print(str(my_list))