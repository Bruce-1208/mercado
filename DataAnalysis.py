from contextlib import nullcontext
import requests
import pandas as pd
import time
from webbrowser import Error
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor


start_time=time.time()
id="8820539028080485"
pwd="7jkMtY3ghAgIkuq3XdgbmrOwaXOlMRnA"
refresh_token="TG-672355e16b6ec300019fb696-1742669993"
#获取最新的token
def getToken(refresh):
     url="https://api.mercadolibre.com/oauth/token?grant_type=refresh_token&client_id="+id+"&client_secret="+pwd+"&refresh_token="+refresh
     json=requests.post(url).json()
     print(json)
     access_token=json['access_token']
     global refresh_token
     refresh_token=json['refresh_token']
     return access_token


headers = {
    'Authorization': 'Bearer '+getToken(refresh_token)  # 假设API需要身份验证令牌
}





# 获取指定时间流量
def getVisitByDate(dict):
        number=dict['number']
        start_date='2024-10-01'
        end_date='2024-10-31'
        url="https://api.mercadolibre.com/items/visits?ids="+number+"&date_from="+start_date+"&date_to="+end_date
        rep = requests.get(url, headers=headers)
        visits=0
        try:
            visits=rep.json()[0]["total_visits"]
        except Exception as e:
            print("************",e)
            visits=-1
        dict.update({'visits': visits})
        print(dict)
        return dict

def getAllVisit(number):
    try:
        url = "https://api.mercadolibre.com/visits/items?ids=" + number
        rep = requests.get(url, headers=headers)
        print(rep.json())
        total_visits = rep.json()['item_id']
        return total_visits
    except Exception as e:
        print(e)

wb = load_workbook("Listings-2024_10_31-08_42.xlsx")
# 获取叫"Sheet"的工作表
sheet = wb["Listings"]
# 获取 2到4行，1到3列 范围的所有单元格


result=[]
list_number=[]

#去重，获取关键属性
for row in sheet.iter_rows(min_row=6):
   number = row[0].value
   a = row[1].value
   b = row[2].value
   Category = row[12].value
   title = row[13].value
   if list_number.__contains__(number):
       continue
   else:
       list_number.append(number)
       dict = {"number": str(number), "category": str(Category), "title": str(title)}
       result.append(dict)
print(len(result))

list_dict=[]
with ThreadPoolExecutor(max_workers=10) as executor:
    results = executor.map(getVisitByDate, result)
list_dict=list(results)
end_time=time.time()

print(list_dict)
print(end_time-start_time)


df = pd.DataFrame(result)
df.to_excel('龙-墨西哥-10月流量分析.xlsx', index=False)